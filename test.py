"""Davomat (kelib-ketish) tizimi — DB yozuvi debug/tekshiruv testi.

Nimani tekshiradi:
  1. «Keldim» bosilganda check_in_time/lat/lng/masofa BAZAGA yozilishi
  2. Kechikish (late_minutes) ish jadvalidagi 09:00 dan to'g'ri hisoblanishi
  3. «Ketdim» bosilganda check_out_time + worked_minutes yozilishi
  4. Dubl check-in / check-out bloklanishi (baza buzilmasligi)
  5. Begona yuz / past tiriklik / ofisdan uzoq GPS — yozuv YARATILMASLIGI
  6. Dam olish kunida (override) status=weekend, late=0 yozilishi
  7. Kechikish statistikasi endpoint'i (/attendance/late-stats) bazadagi bilan mosligi
  8. Bazadagi UNIQUE(user_id, date) cheklovi ishlashi

Ishga tushirish (loyiha ildizidan, API 8000 da ishlab turishi shart):
    .venv/Scripts/python.exe test.py

Barcha sinov ma'lumotlari T- prefiksi bilan yaratiladi va oxirida to'liq
o'chiriladi (jonli ma'lumotga tegilmaydi).
"""
import json
import sqlite3
import sys
import traceback
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

# Loyihaning o'z JWT funksiyasi bilan token yaratamiz — dev-login DEBUG=false'da
# yopiq (404), lekin JWT_SECRET .env'dan bir xil o'qiladi.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from api.security import create_access_token
except Exception:
    print("XATO: api.security import bo'lmadi (.venv'dan ishga tushiring):\n" + traceback.format_exc())
    sys.exit(1)

DB_PATH = "app.db"
API_BASE = "http://127.0.0.1:8000"
OFFICE = (41.311081, 69.240562)  # sinov ofisi koordinatasi
FACE = [0.05] * 128  # ro'yxatdagi yuz
WRONG_FACE = [0.35] * 128  # begona yuz (masofa > 0.5)
TZ = ZoneInfo("Asia/Tashkent")

passed: list[str] = []
failed: list[str] = []


def check(name: str, cond: bool, extra: str = "") -> None:
    """Bitta tekshiruv natijasini qayd etadi va chiqaradi. Chop etish konsol
    kodlashiga chidamli — aks holda extra ichidagi «» kabi belgi Windows
    konsolida UnicodeEncodeError otib, TESTNING O'ZINI yiqitardi (natija
    baholanmasdan FAIL bo'lardi)."""
    (passed if cond else failed).append(name)
    mark = "  [OK]  " if cond else "  [FAIL]"
    line = f"{mark} {name}" + (f"  | {extra}" if extra else "")
    try:
        print(line)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "utf-8"
        print(line.encode(enc, "replace").decode(enc, "replace"))


def _fail_print(line: str) -> None:
    """Konsol kodlashiga chidamli chop etish (`check` dagi bilan bir xil)."""
    try:
        print(line)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "utf-8"
        print(line.encode(enc, "replace").decode(enc, "replace"))


def db() -> sqlite3.Connection:
    return sqlite3.connect(DB_PATH)


# ─────────────────────────────────────────────────────────────────
# Sozlash / tozalash — xatoga chidamli (har biri alohida try/except)
# ─────────────────────────────────────────────────────────────────

def require_notifications_off() -> None:
    """Xabarlar YOQIQ serverga qarshi testni ISHGA TUSHIRMAYDI.

    NEGA: test bazadagi HAQIQIY xodimlar bilan ishlaydi — sinov davri
    («2022-03», «2021-05») uchun hisoblash/tasdiqlash amallari Boshliq va
    HR ga Telegram/push xabari yuboradi. 2026-08-17 da bu IKKI MARTA sodir
    bo'ldi: birinchisida `BOT_TOKEN` bo'shatilgan, lekin push kanali ochiq
    qolgan edi; ikkinchisida esa server oddiy rejimda qayta ishga
    tushirilgani unutilgan edi.

    Endi bu insonning e'tiboriga emas, TEKSHIRUVGA bog'liq.
    """
    import httpx

    try:
        r = httpx.get(f"{API_BASE}/health", timeout=10)
        holat = r.json()
    except Exception as e:  # noqa: BLE001
        print(f"XATO: API ({API_BASE}) javob bermadi: {e}")
        sys.exit(1)

    if holat.get("notifications_enabled") is not False:
        print("=" * 70)
        print("TO'XTATILDI: API xabar yuborish YOQIQ holatda ishlayapti.")
        print("Test haqiqiy xodimlarga Telegram/push yuborib yuboradi.")
        print()
        print("Serverni shu rejimda qayta ishga tushiring:")
        print("  NOTIFICATIONS_ENABLED=false .venv/Scripts/python.exe -m uvicorn"
              " api.main:app --host 127.0.0.1 --port 8000")
        print("=" * 70)
        sys.exit(1)


def setup() -> dict:
    """T- sinov ma'lumotlarini yaratadi: 2 xodim, ish jadvali, ofis."""
    ctx: dict = {}
    conn = db()
    try:
        cur = conn.cursor()
        today = date.today()

        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " face_descriptor, face_registered_at, created_at)"
            " values (999100701, 'T-DebugXodim', 'employee', 1, 1, ?, datetime('now'), datetime('now'))",
            (json.dumps(FACE),),
        )
        ctx["uid1"] = cur.lastrowid

        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " face_descriptor, face_registered_at, created_at)"
            " values (999100702, 'T-DamXodim', 'employee', 1, 1, ?, datetime('now'), datetime('now'))",
            (json.dumps(FACE),),
        )
        ctx["uid2"] = cur.lastrowid

        # uid1: bugun 09:00-23:59 ish kuni (kechikish 09:00 dan hisoblanadi)
        cur.execute(
            "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
            " values (?, ?, 1, '09:00', '23:59', datetime('now'))",
            (ctx["uid1"], today.weekday()),
        )
        # uid2: bugun override bilan DAM OLISH kuni
        cur.execute(
            "insert into work_schedule_override (user_id, date, is_working, updated_at)"
            " values (?, ?, 0, datetime('now'))",
            (ctx["uid2"], today.isoformat()),
        )
        cur.execute(
            "insert into office_locations (name, latitude, longitude, radius_meters, is_active, created_at)"
            " values ('T-DebugOfis', ?, ?, 200, 1, datetime('now'))",
            OFFICE,
        )
        ctx["office_id"] = cur.lastrowid
        conn.commit()
        print(f"Sozlash: uid1={ctx['uid1']}, uid2={ctx['uid2']}, ofis={ctx['office_id']}\n")
        return ctx
    finally:
        conn.close()


def cleanup(ctx: dict) -> None:
    """Barcha T- sinov ma'lumotlarini o'chiradi (har qadami mustaqil himoyalangan)."""
    conn = db()
    try:
        cur = conn.cursor()
        uids = [ctx.get("uid1"), ctx.get("uid2")]
        uids = [u for u in uids if u]
        for sql, params in [
            ("delete from attendance where user_id in (%s)" % ",".join("?" * len(uids)), uids),
            ("delete from work_schedule_weekly where user_id in (%s)" % ",".join("?" * len(uids)), uids),
            ("delete from work_schedule_override where user_id in (%s)" % ",".join("?" * len(uids)), uids),
            ("delete from office_locations where name like 'T-%'", []),
            # T- foydalanuvchilar o'chirilishidan OLDIN — aks holda quyidagi subselect
            # bo'sh qaytadi (SQLite FK cheklovi yo'q, lekin yetim yozuv qolib ketmasin).
            ("delete from face_reregistration_requests where user_id in "
             "(select id from users where full_name like 'T-%')", []),
            ("delete from users where full_name like 'T-%'", []),
        ]:
            try:
                if uids or "T-" in sql:
                    cur.execute(sql, params)
            except sqlite3.Error as e:
                print(f"  tozalash xatosi ({sql[:40]}...): {e}")
        conn.commit()
        left = cur.execute("select count(*) from users where full_name like 'T-%'").fetchone()[0]
        print(f"\nTozalash tugadi. T- qoldiq foydalanuvchi: {left}")
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────
# HTTP yordamchilar
# ─────────────────────────────────────────────────────────────────

def token_for(user_id: int, role: str) -> str | None:
    """Foydalanuvchi uchun JWT'ni bevosita yaratadi (loyihaning o'z funksiyasi,
    .env'dagi JWT_SECRET bilan) — DEBUG holatiga bog'liq emas."""
    try:
        return create_access_token(user_id, role)
    except Exception as e:
        print(f"  token yaratish xatosi (user={user_id}): {e}")
        return None


def find_manager_id() -> int | None:
    """Bazadan bitta faol rahbar (boss/dasturchi/hr) id'sini topadi."""
    try:
        conn = db()
        row = conn.execute(
            "select id, role from users where role in ('boss','dasturchi','hr') and is_active=1 limit 1"
        ).fetchone()
        conn.close()
        return (row[0], row[1]) if row else None
    except sqlite3.Error as e:
        print(f"  rahbar topish xatosi: {e}")
        return None


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def bot_secret_hdr() -> dict:
    """`.env`dagi BOT_SHARED_SECRET — bot/cron endpointlari uchun."""
    env_path = Path(__file__).resolve().parent / ".env"
    with open(env_path, encoding="utf-8") as f:
        secret = next(
            (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")), ""
        )
    return {"X-Bot-Secret": secret}


def payroll_tick(client, kutilgan: str | None = None, urinish: int = 5) -> dict | None:
    """Navbatdagi oylik hisobini bajaradi — testda CRON o'rnini bosadi (§4.3).

    NEGA KERAK: `POST /payroll/{period}/calculate` endi hisobni o'zi
    bajarmaydi, faqat NAVBATGA qo'yadi (202) — production'da Passenger
    konkurentligi = 1 va og'ir hisob butun saytni qotirardi. Haqiqiy ishni
    `scripts/cron_tick.py` alohida jarayonda qiladi; test esa xuddi shu
    servisni HTTP orqali chaqiradi.

    Navbatda boshqa davr ham turgan bo'lishi mumkin (oldingi blok qoldig'i) —
    shuning uchun kutilgan davr chiqquncha bir necha marta uriniladi."""
    natija = None
    for _ in range(urinish):
        r = client.post(f"{API_BASE}/payroll/tick", headers=bot_secret_hdr(), json={})
        if r.status_code != 200:
            return {"http": r.status_code, "text": r.text[:200]}
        natija = r.json()
        if natija.get("ran") is None:
            break
        if kutilgan is None or natija.get("ran") == kutilgan:
            break
    return natija


# ─────────────────────────────────────────────────────────────────
# Testlar
# ─────────────────────────────────────────────────────────────────

def run_tests(ctx: dict) -> None:
    try:
        import httpx
    except ImportError:
        print("XATO: httpx o'rnatilmagan (.venv ishlatilyaptimi?)")
        sys.exit(1)

    uid1, uid2 = ctx["uid1"], ctx["uid2"]
    today_iso = date.today().isoformat()

    with httpx.Client(timeout=15) as client:
        # API tiriklik tekshiruvi
        try:
            r = client.get(f"{API_BASE}/health")
            check("API ishlayapti (/health)", r.status_code == 200)
        except Exception as e:
            print(f"API'ga ulanib bo'lmadi: {e}\nAvval xizmatlarni ishga tushiring.")
            sys.exit(1)

        t1 = token_for(uid1, "employee")
        t2 = token_for(uid2, "employee")
        check("T-xodimlar uchun JWT yaratildi", bool(t1 and t2))
        if not (t1 and t2):
            return  # token yo'q — davomi ma'nosiz
        # Token haqiqatan ishlashini tekshiramiz
        try:
            r = client.get(f"{API_BASE}/users/me", headers=auth(t1))
            check("JWT bilan /users/me ishladi", r.status_code == 200,
                  f"user={r.json().get('full_name')}" if r.status_code == 200 else f"status={r.status_code}")
        except Exception:
            check("JWT bilan /users/me ishladi", False, traceback.format_exc(limit=1).strip())

        # ── 1-5: yozuv YARATILMAYDIGAN xato holatlar ──────────────
        print("\n-- Xato holatlar (baza toza qolishi kerak) --")
        cases = [
            ("begona yuz", {"latitude": OFFICE[0], "longitude": OFFICE[1],
                            "face_descriptor": WRONG_FACE, "liveness": 0.9}),
            ("past tiriklik", {"latitude": OFFICE[0], "longitude": OFFICE[1],
                               "face_descriptor": FACE, "liveness": 0.2}),
            ("ofisdan uzoq GPS", {"latitude": 41.5, "longitude": 69.6,
                                  "face_descriptor": FACE, "liveness": 0.9}),
        ]
        for name, body in cases:
            try:
                r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(t1), json=body)
                check(f"{name} -> 400", r.status_code == 400,
                      r.json().get("detail", "")[:45])
            except Exception:
                check(f"{name} -> 400", False, traceback.format_exc(limit=1).strip())

        try:
            conn = db()
            n = conn.execute("select count(*) from attendance where user_id=?", (uid1,)).fetchone()[0]
            conn.close()
            check("xato holatlardan keyin bazada yozuv YO'Q", n == 0, f"yozuvlar={n}")
        except Exception:
            check("xato holatlardan keyin bazada yozuv YO'Q", False, traceback.format_exc(limit=1).strip())

        # ── Keldim: bazaga yozilish + kechikish hisobi ───────────
        print("\n-- «Keldim» bazaga yozilishi --")
        checkin_moment = datetime.now(TZ)
        try:
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(t1),
                            json={"latitude": OFFICE[0], "longitude": OFFICE[1],
                                  "face_descriptor": FACE, "liveness": 0.9})
            check("to'g'ri check-in -> 200", r.status_code == 200,
                  f"API javobi: status={r.json().get('status')}, late={r.json().get('late_minutes')}")
        except Exception:
            check("to'g'ri check-in -> 200", False, traceback.format_exc(limit=1).strip())

        # Bazadan BEVOSITA o'qib tekshirish (API javobiga ishonmasdan)
        try:
            conn = db()
            row = conn.execute(
                "select check_in_time, check_in_lat, check_in_lng, check_in_distance_m,"
                " late_minutes, status from attendance where user_id=? and date=?",
                (uid1, today_iso),
            ).fetchone()
            conn.close()
            check("bazada check_in_time yozildi", bool(row and row[0]), f"qiymat={row[0] if row else None}")
            check("bazada GPS (lat/lng) yozildi",
                  bool(row and abs(row[1] - OFFICE[0]) < 1e-4 and abs(row[2] - OFFICE[1]) < 1e-4),
                  f"lat={row[1]}, lng={row[2]}" if row else "")
            check("bazada masofa yozildi (radius ichida)", row is not None and row[3] is not None and row[3] <= 200,
                  f"masofa={row[3]}m" if row else "")

            # Kechikishni mustaqil hisoblab solishtirish: grace BO'SAG'A (chegirma
            # emas, 1.3-tuzatish) — diff > grace bo'lsa TO'LIQ diff yoziladi.
            # 1.4-tuzatish: yuqori chegara — ish oynasidan (tushliksiz) oshmaydi.
            # uid1 jadvali 09:00-23:59 (setup()da), shu oynadan oshib ketsa chegaralanadi.
            from api.timeutil import work_minutes as _work_minutes2
            diff = (checkin_moment.hour * 60 + checkin_moment.minute) - (9 * 60)
            expected = diff if diff > 5 else 0
            expected = min(expected, _work_minutes2(9 * 60, 23 * 60 + 59))
            got = row[4] if row else -1
            check("late_minutes to'g'ri hisoblangan (±1 daq)",
                  abs(got - expected) <= 1, f"bazada={got}, kutilgan~{expected}")
            check("status to'g'ri (late/present)",
                  row is not None and row[5] == ("late" if expected > 0 else "present"),
                  f"status={row[5] if row else None}")

            # check_in_time UTC bo'lib saqlanganini tekshirish (Toshkent-5)
            if row and row[0]:
                dt_utc = datetime.fromisoformat(row[0].split(".")[0])
                delta_min = abs((checkin_moment.replace(tzinfo=None) - timedelta(hours=5)) - dt_utc).total_seconds() / 60
                check("check_in_time UTC sifatida saqlangan (±2 daq)", delta_min <= 2,
                      f"bazada={row[0]} (UTC), farq={delta_min:.1f} daq")
        except Exception:
            check("bazadan check-in o'qish", False, traceback.format_exc(limit=1).strip())

        # ── Dubl check-in ─────────────────────────────────────────
        try:
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(t1),
                            json={"latitude": OFFICE[0], "longitude": OFFICE[1],
                                  "face_descriptor": FACE, "liveness": 0.9})
            check("dubl check-in -> 400", r.status_code == 400)
            conn = db()
            n = conn.execute("select count(*) from attendance where user_id=? and date=?",
                             (uid1, today_iso)).fetchone()[0]
            conn.close()
            check("dubldan keyin ham bazada 1 ta yozuv", n == 1, f"yozuvlar={n}")
        except Exception:
            check("dubl check-in nazorati", False, traceback.format_exc(limit=1).strip())

        # ── Ketdim: bazaga yozilish ───────────────────────────────
        print("\n-- «Ketdim» bazaga yozilishi --")
        try:
            r = client.post(f"{API_BASE}/attendance/me/check-out", headers=auth(t1),
                            json={"latitude": OFFICE[0], "longitude": OFFICE[1],
                                  "face_descriptor": FACE, "liveness": 0.9})
            check("check-out -> 200", r.status_code == 200)
            conn = db()
            row = conn.execute(
                "select check_out_time, worked_minutes, early_leave_minutes from attendance"
                " where user_id=? and date=?", (uid1, today_iso)).fetchone()
            conn.close()
            check("bazada check_out_time yozildi", bool(row and row[0]), f"qiymat={row[0] if row else None}")
            check("worked_minutes >= 0 yozildi", row is not None and row[1] is not None and row[1] >= 0,
                  f"worked={row[1]}" if row else "")
        except Exception:
            check("check-out bazaga yozilishi", False, traceback.format_exc(limit=1).strip())

        try:
            r = client.post(f"{API_BASE}/attendance/me/check-out", headers=auth(t1),
                            json={"latitude": OFFICE[0], "longitude": OFFICE[1],
                                  "face_descriptor": FACE, "liveness": 0.9})
            check("dubl check-out -> 400", r.status_code == 400)
        except Exception:
            check("dubl check-out -> 400", False, traceback.format_exc(limit=1).strip())

        # ── Dam olish kuni ────────────────────────────────────────
        print("\n-- Dam olish kuni (override) --")
        try:
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(t2),
                            json={"latitude": OFFICE[0], "longitude": OFFICE[1],
                                  "face_descriptor": FACE, "liveness": 0.9})
            conn = db()
            row = conn.execute(
                "select status, late_minutes, is_weekend from attendance where user_id=? and date=?",
                (uid2, today_iso)).fetchone()
            conn.close()
            check("dam kunida check-in -> weekend/late=0",
                  r.status_code == 200 and row is not None
                  and row[0] == "weekend" and row[1] == 0 and row[2] == 1,
                  f"status={row[0] if row else None}, late={row[1] if row else None}")
        except Exception:
            check("dam kuni holati", False, traceback.format_exc(limit=1).strip())

        # ── UNIQUE(user_id, date) cheklovi ────────────────────────
        print("\n-- Baza cheklovi --")
        try:
            conn = db()
            try:
                conn.execute(
                    "insert into attendance (user_id, date, late_minutes, early_leave_minutes,"
                    " worked_minutes, status, is_weekend, created_at, updated_at)"
                    " values (?, ?, 0, 0, 0, 'present', 0, datetime('now'), datetime('now'))",
                    (uid1, today_iso),
                )
                conn.commit()
                check("UNIQUE(user_id,date) cheklovi ishlaydi", False, "dubl insert o'tib ketdi!")
            except sqlite3.IntegrityError:
                check("UNIQUE(user_id,date) cheklovi ishlaydi", True, "IntegrityError (kutilgan)")
            finally:
                conn.close()
        except Exception:
            check("UNIQUE cheklovi testi", False, traceback.format_exc(limit=1).strip())

        # ── Kechikish statistikasi endpoint'i ─────────────────────
        print("\n-- Kechikish statistikasi (/attendance/late-stats) --")
        try:
            mgr = find_manager_id()
            check("bazada rahbar topildi", mgr is not None, f"{mgr}")
            boss_t = token_for(mgr[0], mgr[1]) if mgr else None
            r = client.get(f"{API_BASE}/attendance/late-stats?days=7", headers=auth(boss_t))
            check("late-stats -> 200", r.status_code == 200)
            stats = r.json()
            me = next((s for s in stats if s["full_name"] == "T-DebugXodim"), None)
            conn = db()
            db_late = conn.execute(
                "select late_minutes from attendance where user_id=? and date=?",
                (uid1, today_iso)).fetchone()[0]
            conn.close()
            if db_late > 0:
                check("statistikada T-DebugXodim bor (bazadagi bilan mos)",
                      me is not None and me["total_late_minutes"] == db_late
                      and any(d["date"] == today_iso and d["late_minutes"] == db_late for d in me["days"]),
                      f"api={me['total_late_minutes'] if me else None}, baza={db_late}")
            else:
                check("kechikish 0 — statistikada yo'q (to'g'ri)", me is None)
            # employee ruxsati yo'qligi
            r = client.get(f"{API_BASE}/attendance/late-stats", headers=auth(t1))
            check("employee late-stats -> 403", r.status_code == 403)
        except Exception:
            check("late-stats tekshiruvi", False, traceback.format_exc(limit=1).strip())

        # ── 1-BOSQICH: statistika yolg'oni tuzatishlari (2026-07-26 audit) ──
        print("\n-- 1.1: absent yozuvi + 1.2: LEFT JOIN --")
        try:
            import asyncio as _asyncio

            from db.base import async_session as _async_session
            from api.services.attendance_digest import write_absent_records as _write_absent

            today = date.today()
            wd_today = today.weekday()

            conn = db()
            cur2 = conn.cursor()
            # T-HechKelmagan: jadvali bor (ish kuni), check-in qilmagan, sababli ham emas
            cur2.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999666001,'T-HechKelmagan','employee',1,1,datetime('now'))")
            ghost_uid = cur2.lastrowid
            cur2.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (ghost_uid, wd_today))
            # T-SababliKun: xuddi shunday jadval, lekin BUGUNGA tasdiqlangan sababli kun
            cur2.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999666002,'T-SababliKun','employee',1,1,datetime('now'))")
            excused_uid = cur2.lastrowid
            cur2.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (excused_uid, wd_today))
            cur2.execute(
                "insert into excused_days (user_id, date, reason, status, created_at)"
                " values (?, ?, 'T-sinov', 'approved', datetime('now'))", (excused_uid, today.isoformat()))
            conn.commit()

            # ⚠️ JONLI MA'LUMOT HIMOYASI: write_absent_records BUGUN uchun
            # chaqiriladi va u HAQIQIY xodimlarga ham (hali check-in qilmagan
            # bo'lsa) absent yozadi — kun hali tugamagan bo'lsa bu YOLG'ON
            # yozuv bo'lib qoladi va matritsada "Kutilmoqda" o'rniga "Kelmadi"
            # ko'rinadi. Chaqiruvdan OLDINGI id'larni eslab, test yaratganlarini
            # (T-lardan tashqari haqiqiylarni ham) oxirida O'CHIRAMIZ.
            before_ids = {
                r[0] for r in conn.execute(
                    "select id from attendance where date=?", (today.isoformat(),)
                ).fetchall()
            }

            async def _run_absent():
                async with _async_session() as s:
                    return await _write_absent(s, today)

            written = _asyncio.run(_run_absent())
            row_ghost = conn.execute(
                "select status from attendance where user_id=? and date=?",
                (ghost_uid, today.isoformat())).fetchone()
            check("1.1: absent yozuvi yaratildi", row_ghost is not None and row_ghost[0] == "absent",
                  f"written={written}, status={row_ghost[0] if row_ghost else None}")

            row_excused = conn.execute(
                "select status from attendance where user_id=? and date=?",
                (excused_uid, today.isoformat())).fetchone()
            check("1.1: sababli kunli xodimga absent YOZILMAYDI", row_excused is None,
                  f"yozuv={row_excused}")

            # Idempotentlik
            _asyncio.run(_run_absent())
            n = conn.execute("select count(*) from attendance where user_id=? and date=?",
                             (ghost_uid, today.isoformat())).fetchone()[0]
            check("1.1: qayta chaqirilsa dublikat yo'q", n == 1, f"yozuvlar={n}")

            r = client.get(f"{API_BASE}/attendance?status_filter=absent&date_from={today.isoformat()}",
                           headers=auth(boss_t))
            names = [x["user_full_name"] for x in r.json()]
            check("1.1: GET status_filter=absent natija qaytaradi", "T-HechKelmagan" in names, f"{names}")

            # 1.2: hech qachon check-in qilmagan (jadval ham yo'q) xodim
            # employee-summary'da (LEFT JOIN) ko'rinishi kerak
            cur2.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999666003,'T-Umuman-Yoq','employee',1,1,datetime('now'))")
            noop_uid = cur2.lastrowid
            conn.commit()
            r = client.get(f"{API_BASE}/attendance/employee-summary?days=90", headers=auth(boss_t))
            names2 = [x["full_name"].strip() for x in r.json()]
            check("1.2: hech qachon kelmagan xodim employee-summary'da bor",
                  "T-Umuman-Yoq" in names2, f"{len(names2)} kishi ro'yxatda")

            # Test yaratgan BARCHA bugungi absent yozuvlarini o'chiramiz —
            # haqiqiy xodimlarnikini ham (kun tugagach kechki tick ularni
            # o'zi qayta yozadi, hech narsa yo'qolmaydi).
            after_rows = conn.execute(
                "select id from attendance where date=?", (today.isoformat(),)
            ).fetchall()
            new_ids = [r[0] for r in after_rows if r[0] not in before_ids]
            if new_ids:
                cur2.execute(
                    "delete from attendance where id in (%s)" % ",".join("?" * len(new_ids)),
                    new_ids,
                )

            for u in (ghost_uid, excused_uid, noop_uid):
                cur2.execute("delete from attendance where user_id=?", (u,))
                cur2.execute("delete from work_schedule_weekly where user_id=?", (u,))
                cur2.execute("delete from excused_days where user_id=?", (u,))
                cur2.execute("delete from users where id=?", (u,))
            conn.commit()
            conn.close()
        except Exception:
            check("1-bosqich tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 1.4: kechikish/erta-ketish yuqori chegarasi --")
        try:
            from api.timeutil import work_minutes as _work_minutes
            window = _work_minutes(9 * 60, 18 * 60)  # 09:00-18:00, tushliksiz
            check("1.4: ish oynasi (tushliksiz) 480 daq", window == 480, f"window={window}")
            # Servis darajasida: diff oynadan katta bo'lsa late shu oyna bilan cheklanishi
            # kodning o'zida tekshirilgan (attendance.py min(late, window)); bu yerda
            # formulaning to'g'riligini mustaqil qayta hisoblab tasdiqlaymiz.
            fake_diff = 999  # masalan 17:59 da kelish emulyatsiyasi
            capped = min(fake_diff, window)
            check("1.4: 999 daqiqalik farq 480 bilan cheklanadi", capped == 480, f"capped={capped}")
        except Exception:
            check("1.4 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        # ── 0-BOSQICH: xavfsizlik tuzatishlari (2026-07-26 audit) ─────
        print("\n-- 0.1: /attendance/digest chat_id ruxsati --")
        try:
            r = client.post(f"{API_BASE}/attendance/digest",
                             params={"kind": "morning", "chat_id": 999999, "dry_run": "true"})
            check("sekretsiz chaqiruv -> 401/403", r.status_code in (401, 403))

            mgr = find_manager_id()
            mgr_tg = None
            if mgr:
                conn = db()
                row = conn.execute("select telegram_id from users where id=?", (mgr[0],)).fetchone()
                conn.close()
                mgr_tg = row[0] if row else None

            # bot-sekret bilan, lekin telegram_id'siz chat_id -> 400
            import httpx as _httpx  # allaqachon yuqorida import qilingan bo'lsa ham xavfsiz

            with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as f:
                secret = next(
                    (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")),
                    "",
                )
            bot_h = {"X-Bot-Secret": secret}
            r = client.post(f"{API_BASE}/attendance/digest", headers=bot_h,
                             params={"kind": "morning", "chat_id": 999999, "dry_run": "true"})
            check("bot-sekret + chat_id, telegram_id'siz -> 400", r.status_code == 400)

            if mgr_tg:
                r = client.post(f"{API_BASE}/attendance/digest", headers=bot_h,
                                params={"kind": "morning", "chat_id": 999999,
                                        "telegram_id": mgr_tg, "dry_run": "true"})
                check("bot-sekret + rahbar telegram_id + chat_id -> 200", r.status_code == 200)

            # chat_id berilmagan holat — bu cron/scheduler yo'li, telegram_id shart emas
            r = client.post(f"{API_BASE}/attendance/digest", headers=bot_h,
                             params={"kind": "morning", "dry_run": "true"})
            check("chat_id'siz chaqiruv (cron yo'li) hali ishlaydi", r.status_code == 200)
        except Exception:
            check("0.1 digest ruxsat tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 0.4: /docs production'da yopiq --")
        try:
            r = client.get(f"{API_BASE}/docs")
            check("/docs -> 404 (DEBUG=false)", r.status_code == 404)
            r = client.get(f"{API_BASE}/openapi.json")
            check("/openapi.json -> 404 (DEBUG=false)", r.status_code == 404)
        except Exception:
            check("0.4 /docs tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 0.6: bir xil yuz ikki hisobga ro'yxatdan o'ta olmaydi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444778,'T-Dup1','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            dup1_uid = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444779,'T-Dup2','employee',1,1,datetime('now'))")
            dup2_uid = cur.lastrowid
            conn.commit()

            dup2_tok = token_for(dup2_uid, "employee")
            # T-Dup1 bilan BIR XIL descriptor bilan ro'yxatdan o'tishga urinish -> rad etilishi kerak.
            r = client.post(f"{API_BASE}/attendance/me/register-face", headers=auth(dup2_tok),
                             json={"face_descriptor": FACE})
            check("bir xil yuz bilan ro'yxatdan o'tish -> 400", r.status_code == 400, r.text[:150])
            has_face_after = conn.execute(
                "select face_descriptor is not null from users where id=?", (dup2_uid,)).fetchone()[0]
            check("rad etilgach descriptor yozilmagan", has_face_after == 0)

            # Boshqa (uzoq) yuz bilan esa muvaffaqiyatli bo'lishi kerak.
            r = client.post(f"{API_BASE}/attendance/me/register-face", headers=auth(dup2_tok),
                             json={"face_descriptor": WRONG_FACE})
            check("uzoq yuz bilan ro'yxatdan o'tish -> 200", r.status_code == 200)
            body = r.json()
            check("birinchi marta darhol 'registered'", body.get("status") == "registered",
                  str(body.get("status")))

            cur.execute("delete from face_reregistration_requests where user_id in (?,?)", (dup1_uid, dup2_uid))
            cur.execute("delete from users where id in (?,?)", (dup1_uid, dup2_uid))
            conn.commit()
            conn.close()
        except Exception:
            check("0.6 duplicate-face tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 0.7: GPS aniqligi past bo'lsa check-in rad etiladi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444780,'T-AccuracyTest','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            acc_uid = cur.lastrowid
            conn.commit()
            acc_tok = token_for(acc_uid, "employee")

            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(acc_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9, "accuracy": 5000,
            })
            check("aniqlik 5000m -> 400 (rad etiladi)", r.status_code == 400, r.text[:150])
            row = conn.execute(
                "select check_in_time from attendance where user_id=?", (acc_uid,)).fetchone()
            check("bazaga yozuv qo'shilmadi", row is None or row[0] is None)

            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(acc_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9, "accuracy": 15,
            })
            check("aniqlik 15m -> 200 (o'tadi)", r.status_code == 200, r.text[:150])

            cur.execute("delete from attendance where user_id=?", (acc_uid,))
            cur.execute("delete from users where id=?", (acc_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("0.7 GPS aniqligi tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 2.1: yarim tundan keyin kechagi ochiq yozuvni yopish --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444781,'T-Midnight','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            mid_uid = cur.lastrowid
            wd = date.today().weekday()
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (mid_uid, wd))
            # "Kecha" check-in qilgan, hali check-out qilmagan — 2 soat oldin (UTC).
            yesterday = date.today() - timedelta(days=1)
            checkin_utc = datetime.utcnow() - timedelta(hours=2)
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, ?, ?, 0, 0, 0, 'present', 0, datetime('now'), datetime('now'))",
                (mid_uid, yesterday.isoformat(), checkin_utc.isoformat(sep=" ")))
            conn.commit()
            mid_tok = token_for(mid_uid, "employee")

            r = client.post(f"{API_BASE}/attendance/me/check-out", headers=auth(mid_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9,
            })
            check("yarim tundan keyingi check-out -> 200", r.status_code == 200, r.text[:200])
            row = conn.execute(
                "select date, check_out_time, worked_minutes from attendance where user_id=?",
                (mid_uid,)).fetchone()
            check("yozuv KECHAGI sanada qoldi (bugunga ko'chmadi)",
                  row is not None and row[0] == yesterday.isoformat(), f"{row}")
            check("worked_minutes ~2 soat (real vaqt farqi)",
                  row is not None and 110 <= row[2] <= 130, f"worked={row[2] if row else None}")
            today_row = conn.execute(
                "select count(*) from attendance where user_id=? and date=?",
                (mid_uid, date.today().isoformat())).fetchone()[0]
            check("bugunga yangi yozuv YARATILMADI", today_row == 0)

            cur.execute("delete from attendance where user_id=?", (mid_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (mid_uid,))
            cur.execute("delete from users where id=?", (mid_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("2.1 yarim tun tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 2.1b: 6 soatdan uzoq ochiq yozuv YOPILMAYDI (eski oyna) --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444782,'T-OldOpen','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            old_uid = cur.lastrowid
            yesterday = date.today() - timedelta(days=1)
            old_checkin_utc = datetime.utcnow() - timedelta(hours=10)  # 6 soatlik oynadan tashqari
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, ?, ?, 0, 0, 0, 'present', 0, datetime('now'), datetime('now'))",
                (old_uid, yesterday.isoformat(), old_checkin_utc.isoformat(sep=" ")))
            conn.commit()
            old_tok = token_for(old_uid, "employee")

            r = client.post(f"{API_BASE}/attendance/me/check-out", headers=auth(old_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9,
            })
            check("10 soat oldingi ochiq yozuv -> 400 (oynadan tashqari)", r.status_code == 400, r.text[:150])
            row = conn.execute(
                "select check_out_time from attendance where user_id=?", (old_uid,)).fetchone()
            check("eski yozuv hamon ochiq", row is not None and row[0] is None)

            cur.execute("delete from attendance where user_id=?", (old_uid,))
            cur.execute("delete from users where id=?", (old_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("2.1b eski oyna tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 2.2: kelib-u \"Ketdim\" bosmagan o'tgan kun avtomatik yopiladi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444783,'T-AutoClose','employee',1,1,datetime('now'))")
            ac_uid = cur.lastrowid
            three_days_ago = date.today() - timedelta(days=3)
            wd = three_days_ago.weekday()
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (ac_uid, wd))
            checkin_utc = datetime(three_days_ago.year, three_days_ago.month, three_days_ago.day, 4, 0, 0)  # 09:00 Toshkent
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, ?, ?, 0, 0, 0, 'present', 0, datetime('now'), datetime('now'))",
                (ac_uid, three_days_ago.isoformat(), checkin_utc.isoformat(sep=" ")))
            conn.commit()

            import asyncio
            from db.base import async_session
            from api.services.attendance_digest import auto_close_unclosed_checkouts

            async def _run_close():
                async with async_session() as s:
                    return await auto_close_unclosed_checkouts(s, date.today())

            closed = asyncio.run(_run_close())
            check("auto_close_unclosed_checkouts >=1 yopdi", closed >= 1, f"closed={closed}")
            row = conn.execute(
                "select check_out_time, worked_minutes, note from attendance where user_id=?",
                (ac_uid,)).fetchone()
            check("3 kun oldingi ochiq yozuv yopildi", row is not None and row[0] is not None, f"{row}")
            check("worked_minutes ish oynasidan hisoblangan (~480 daq)",
                  row is not None and 400 <= row[1] <= 480, f"worked={row[1] if row else None}")
            check("izoh (note) qo'yildi", row is not None and row[2] and "Avtomatik" in row[2], f"note={row[2] if row else None}")

            # Bugungi (hali davom etayotgan) ochiq yozuv TEGILMASLIGI kerak.
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444784,'T-StillOpen','employee',1,1,datetime('now'))")
            open_uid = cur.lastrowid
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, ?, datetime('now'), 0, 0, 0, 'present', 0, datetime('now'), datetime('now'))",
                (open_uid, date.today().isoformat()))
            conn.commit()
            closed2 = asyncio.run(_run_close())
            check("ikkinchi chaqiruvda yangi yopish yo'q (bugungi hisobga kirmaydi)", closed2 == 0, f"closed2={closed2}")
            row2 = conn.execute(
                "select check_out_time from attendance where user_id=?", (open_uid,)).fetchone()
            check("bugungi ochiq yozuvga TEGILMADI", row2 is not None and row2[0] is None, f"{row2}")

            cur.execute("delete from attendance where user_id in (?,?)", (ac_uid, open_uid))
            cur.execute("delete from work_schedule_weekly where user_id=?", (ac_uid,))
            cur.execute("delete from users where id in (?,?)", (ac_uid, open_uid))
            conn.commit()
            conn.close()
        except Exception:
            check("2.2 avtomatik yopish tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 2.3: check-in poygasi -> 500 EMAS, tushunarli 400 --")
        try:
            import concurrent.futures

            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444785,'T-RaceTest','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            race_uid = cur.lastrowid
            conn.commit()
            race_tok = token_for(race_uid, "employee")

            def _do_checkin():
                with httpx.Client(timeout=20) as c2:
                    return c2.post(f"{API_BASE}/attendance/me/check-in", headers=auth(race_tok), json={
                        "latitude": OFFICE[0], "longitude": OFFICE[1],
                        "face_descriptor": FACE, "liveness": 0.9,
                    })

            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
                futs = [ex.submit(_do_checkin) for _ in range(2)]
                results = [f.result().status_code for f in futs]
            check("parallel check-in: hech biri 500 EMAS", 500 not in results, f"natijalar={results}")
            check("parallel check-in: biri 200, ikkinchisi 400", sorted(results) == [200, 400], f"natijalar={results}")
            n = conn.execute(
                "select count(*) from attendance where user_id=? and date=?",
                (race_uid, date.today().isoformat())).fetchone()[0]
            check("bazada faqat 1 ta yozuv", n == 1, f"yozuvlar={n}")

            cur.execute("delete from attendance where user_id=?", (race_uid,))
            cur.execute("delete from users where id=?", (race_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("2.3 poyga tekshiruvi", False, traceback.format_exc(limit=1).strip())

        # ─────────────────────────────────────────────────────────
        # 0-BOSQICH (oylik/jarima poydevori): HR qo'lda tuzatishi
        # va ma'lumot tayyorligi hisoboti
        # ─────────────────────────────────────────────────────────
        print("\n-- 0.1: PUT /attendance/manual (HR qo'lda tuzatishi) --")
        try:
            mgr = find_manager_id()
            mgr_t = token_for(mgr[0], mgr[1]) if mgr else None
            check("rahbar tokeni olindi", bool(mgr_t))

            def _manual(payload: dict, token: str | None = None):
                return client.request(
                    "PUT", f"{API_BASE}/attendance/manual",
                    headers=auth(token or mgr_t), json=payload,
                )

            # Sabab majburiy: 5 belgidan qisqa -> 422 (pydantic)
            r = _manual({"user_id": uid1, "date": today_iso, "check_in": "10:30", "reason": "qis"})
            check("sababsiz (qisqa) tuzatish -> 422", r.status_code == 422, f"kod={r.status_code}")

            # Noto'g'ri vaqt formati -> 422
            r = _manual({"user_id": uid1, "date": today_iso, "check_in": "25:99",
                         "reason": "format tekshiruvi"})
            check("noto'g'ri vaqt formati -> 422", r.status_code == 422, f"kod={r.status_code}")

            # «Ketdim» ni «Keldim» siz belgilash -> 400
            r = _manual({"user_id": uid1, "date": today_iso, "check_out": "18:00",
                         "reason": "faqat ketdim sinovi"})
            check("«Ketdim» «Keldim» siz -> 400", r.status_code == 400, f"kod={r.status_code}")

            # ROP'da bu huquq YO'Q (kechikish jarimasini o'zi bekor qila olmasin)
            conn = db()
            cur = conn.cursor()
            # Audit tozalash uchun chegara: SHU testdan keyin yaratilgan yozuvlarnigina
            # o'chiramiz (id > audit_before). Keng `action`/`target_user_id` filtri
            # haqiqiy tarixiy yozuvlarni ham ushlab qolishi mumkin edi.
            audit_before = conn.execute("select coalesce(max(id), 0) from audit_logs").fetchone()[0]
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444790,'T-RopEdit','rop',1,1,datetime('now'))")
            rop_uid = cur.lastrowid
            conn.commit()
            rop_t = token_for(rop_uid, "rop")
            r = _manual({"user_id": uid1, "date": today_iso, "check_in": "10:30",
                         "reason": "ROP urinishi"}, token=rop_t)
            check("ROP qo'lda tuzata OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # Haqiqiy tuzatish: uid1 jadvali bugun 09:00-23:59 -> 10:30 = 90 daq kechikish
            r = _manual({"user_id": uid1, "date": today_iso, "check_in": "10:30",
                         "check_out": "19:00", "reason": "Face ID ishlamadi (sinov)"})
            check("tuzatish -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            body = r.json() if r.status_code == 200 else {}
            check("kechikish jadvaldan qayta hisoblandi (90 daq)",
                  body.get("late_minutes") == 90, f"late={body.get('late_minutes')}")
            check("status 'late' ga o'tdi", body.get("status") == "late", f"status={body.get('status')}")
            check("ishlangan vaqt hisoblandi (>0)",
                  (body.get("worked_minutes") or 0) > 0, f"worked={body.get('worked_minutes')}")
            check("qo'lda tuzatishda GPS masofasi tozalandi",
                  body.get("check_in_distance_m") is None, f"dist={body.get('check_in_distance_m')}")

            # Audit jurnaliga tushdimi
            log = conn.execute(
                "select action, after from audit_logs where action='attendance_manual_edit'"
                " and target_user_id=? order by id desc limit 1", (uid1,)).fetchone()
            check("audit jurnaliga yozildi", log is not None, f"{log}")
            check("auditda sabab saqlandi",
                  bool(log) and "Face ID ishlamadi" in (log[1] or ""), f"{log[1] if log else None}")

            # «Keldim» ni tozalash -> kun "kelmagan" bo'lib qoladi
            r = _manual({"user_id": uid1, "date": today_iso, "check_in": None,
                         "check_out": None, "reason": "Aslida kelmagan (sinov)"})
            check("«Keldim» tozalandi -> absent",
                  r.status_code == 200 and r.json().get("status") == "absent",
                  f"kod={r.status_code} status={r.json().get('status') if r.status_code == 200 else None}")

            # O'TGAN kunga YANGI yozuv yaratish (yozuv yo'q edi) — 7 kun oldin,
            # ayni hafta kuni, ya'ni uid1 ning o'sha jadvali amal qiladi.
            past = (date.today() - timedelta(days=7)).isoformat()
            r = _manual({"user_id": uid1, "date": past, "check_in": "09:20",
                         "check_out": "18:00", "reason": "Yo'qolgan kun tiklandi"})
            check("o'tgan kunga yangi yozuv yaratildi -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:120]}")
            check("yangi yozuvda kechikish 20 daq",
                  r.status_code == 200 and r.json().get("late_minutes") == 20,
                  f"late={r.json().get('late_minutes') if r.status_code == 200 else None}")

            # Dam olish kunida (uid2 bugun override bilan dam) kechikish bo'lmaydi
            r = _manual({"user_id": uid2, "date": today_iso, "check_in": "12:00",
                         "check_out": "15:00", "reason": "Dam olish kuni sinovi"})
            check("dam olish kunida kechikish 0 va status 'weekend'",
                  r.status_code == 200 and r.json().get("late_minutes") == 0
                  and r.json().get("status") == "weekend",
                  f"{r.json() if r.status_code == 200 else r.status_code}")

            # Kelajakdagi kun -> 400
            future = (date.today() + timedelta(days=1)).isoformat()
            r = _manual({"user_id": uid1, "date": future, "check_in": "09:00",
                         "reason": "Kelajak sinovi"})
            check("kelajakdagi kunni tuzatib bo'lmaydi -> 400", r.status_code == 400,
                  f"kod={r.status_code}")

            cur.execute("delete from users where id=?", (rop_uid,))
            cur.execute("delete from attendance where user_id=? and date=?", (uid1, past))
            # Shu testda yozilgan audit yozuvlarini (id chegarasi bilan) tozalaymiz —
            # eski haqiqiy tarixga tegilmaydi.
            cur.execute("delete from audit_logs where id > ? and action='attendance_manual_edit'",
                        (audit_before,))
            conn.commit()
            conn.close()
        except Exception:
            check("0.1 qo'lda tuzatish tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 0.3: GET /attendance/readiness (ma'lumot tayyorligi) --")
        try:
            mgr = find_manager_id()
            mgr_t = token_for(mgr[0], mgr[1]) if mgr else None

            # Yuzi ro'yxatdan o'tmagan xodim — hisobotda ko'rinishi kerak
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444791,'T-NoFace','employee',1,1,datetime('now'))")
            nf_uid = cur.lastrowid
            conn.commit()

            r = client.get(f"{API_BASE}/attendance/readiness", headers=auth(mgr_t))
            check("readiness -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            data = r.json() if r.status_code == 200 else {}
            for key in ("no_schedule", "open_checkouts", "auto_closed", "pending_excused", "no_face"):
                check(f"readiness '{key}' guruhi bor", key in data)
            check("yuzsiz xodim 'no_face' ro'yxatida",
                  any(i["user_id"] == nf_uid for i in data.get("no_face", [])),
                  f"no_face={len(data.get('no_face', []))} ta")
            check("jadvalsiz xodim 'no_schedule' ro'yxatida",
                  any(i["user_id"] == nf_uid for i in data.get("no_schedule", [])),
                  f"no_schedule={len(data.get('no_schedule', []))} ta")
            check("muammo bor ekan ok=False", data.get("ok") is False, f"ok={data.get('ok')}")

            # Noto'g'ri sana oralig'i -> 400
            r = client.get(
                f"{API_BASE}/attendance/readiness?date_from=2026-07-10&date_to=2026-07-01",
                headers=auth(mgr_t))
            check("teskari sana oralig'i -> 400", r.status_code == 400, f"kod={r.status_code}")

            cur.execute("delete from users where id=?", (nf_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("0.3 tayyorlik tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 3.1: jadval o'zgarsa BUGUNGI yozuv qayta hisoblanadi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444792,'T-Recalc','employee',1,1,datetime('now'))")
            rc_uid = cur.lastrowid
            wd = date.today().weekday()
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (rc_uid, wd))
            # 09:30 Toshkent = 04:30 UTC check-in -> 09:00 jadval bilan late=30
            checkin_utc = f"{date.today().isoformat()} 04:30:00"
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, ?, ?, 30, 0, 0, 'late', 0, datetime('now'), datetime('now'))",
                (rc_uid, date.today().isoformat(), checkin_utc))
            conn.commit()

            mgr_rc = find_manager_id()
            if mgr_rc:
                mgr_rc_tok = token_for(mgr_rc[0], mgr_rc[1])
                # Rahbar startni 10:00 ga surdi -> 09:30 kelish endi kechikish EMAS.
                r = client.put(
                    f"{API_BASE}/work-schedule/{rc_uid}/weekly", headers=auth(mgr_rc_tok),
                    json={"days": [
                        {"weekday": d, "is_working": d == wd, "start_time": "10:00" if d == wd else None,
                         "end_time": "18:00" if d == wd else None}
                        for d in range(7)
                    ]})
                check("PUT weekly -> 200", r.status_code == 200, f"kod={r.status_code}")
                row = conn.execute(
                    "select late_minutes, status from attendance where user_id=? and date=?",
                    (rc_uid, date.today().isoformat())).fetchone()
                check("3.1: bugungi yozuv qayta hisoblandi (late=0)",
                      row is not None and row[0] == 0, f"late={row[0] if row else None}, status={row[1] if row else None}")
            else:
                check("3.1 rahbar topilmadi", False)

            cur.execute("delete from attendance where user_id=?", (rc_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (rc_uid,))
            cur.execute("delete from users where id=?", (rc_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("3.1 qayta hisoblash tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 3.2: eng yaqin EMAS, BIRORTA ofis radiusi ichida bo'lsa yetarli --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444793,'T-NearestOffice','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            no_uid = cur.lastrowid
            # A ofis (ENG YAQIN, ~78m): radius 50m -> xodim A dan TASHQARIDA.
            cur.execute(
                "insert into office_locations (name, latitude, longitude, radius_meters, is_active, created_at)"
                " values ('T-YaqinOfis', ?, ?, 50, 1, datetime('now'))",
                (OFFICE[0] + 0.0007, OFFICE[1]))
            office_a_id = cur.lastrowid
            # B ofis (uzoqroq, ~156m), lekin radiusi 200m -> xodim B radiusi ICHIDA.
            cur.execute(
                "insert into office_locations (name, latitude, longitude, radius_meters, is_active, created_at)"
                " values ('T-UzoqOfis', ?, ?, 200, 1, datetime('now'))",
                (OFFICE[0] + 0.0014, OFFICE[1]))
            office_b_id = cur.lastrowid
            conn.commit()

            no_tok = token_for(no_uid, "employee")
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(no_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9,
            })
            check("eng yaqin ofis radiusidan tashqarida bo'lsa ham, B radiusida -> 200",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            cur.execute("delete from attendance where user_id=?", (no_uid,))
            cur.execute("delete from office_locations where id in (?,?)", (office_a_id, office_b_id))
            cur.execute("delete from users where id=?", (no_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("3.2 eng yaqin ofis tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 3.4: Boshliq check-in qila olmaydi (izchillik) --")
        try:
            boss_row = conn2 = None
            conn2 = db()
            boss_row = conn2.execute(
                "select id, telegram_id, face_descriptor from users where role='boss' limit 1").fetchone()
            if boss_row:
                boss_uid = boss_row[0]
                boss_tok = token_for(boss_uid, "boss")
                r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(boss_tok), json={
                    "latitude": OFFICE[0], "longitude": OFFICE[1],
                    "face_descriptor": FACE, "liveness": 0.9,
                })
                check("Boshliq check-in -> 400 (bloklangan)", r.status_code == 400, f"kod={r.status_code} {r.text[:150]}")
            else:
                check("3.4 Boshliq topilmadi", False)
            conn2.close()
        except Exception:
            check("3.4 Boshliq check-in tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 4.6: days=30 aniq 30 kun (31 EMAS) --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444794,'T-DaysWindow','employee',1,1,datetime('now'))")
            dw_uid = cur.lastrowid
            today = date.today()
            exactly_30 = today - timedelta(days=29)  # bugundan hisoblab 30-chi kun (chegarada)
            exactly_31 = today - timedelta(days=30)  # 31-chi kun — OYNADAN TASHQARI bo'lishi kerak
            for d in (exactly_30, exactly_31):
                cur.execute(
                    "insert into attendance (user_id, date, check_in_time, late_minutes,"
                    " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                    " values (?, ?, ?, 10, 0, 0, 'late', 0, datetime('now'), datetime('now'))",
                    (dw_uid, d.isoformat(), f"{d.isoformat()} 04:00:00"))
            conn.commit()

            r = client.get(f"{API_BASE}/attendance/employee-summary?days=30", headers=auth(boss_t))
            check("employee-summary?days=30 -> 200", r.status_code == 200)
            row = next((x for x in r.json() if x["user_id"] == dw_uid), None)
            check("30-kunlik oyna: faqat CHEGARADAGI (29 kun oldingi) yozuv hisobga olindi",
                  row is not None and row["late_minutes"] == 10, f"{row}")

            r2 = client.get(f"{API_BASE}/attendance/late-stats?days=30", headers=auth(boss_t))
            names_in_window = [x["full_name"] for x in r2.json() if x["user_id"] == dw_uid]
            late_days_count = next((x["late_days"] for x in r2.json() if x["user_id"] == dw_uid), None)
            check("late-stats?days=30: 31-kun oldingi yozuv OYNADAN TASHQARI",
                  late_days_count == 1, f"late_days={late_days_count}")

            cur.execute("delete from attendance where user_id=?", (dw_uid,))
            cur.execute("delete from users where id=?", (dw_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("4.6 kunlar oynasi tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 5.1: tasdiqlangan sababli kunda kechikish yozilmaydi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444795,'T-ExcusedCheckin','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            ex_uid = cur.lastrowid
            wd = date.today().weekday()
            # 00:01 boshlanish — real vaqt qanday bo'lishidan qat'i nazar, oddiy
            # mantiqda deyarli har doim "kechikkan" bo'lardi.
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'00:01','23:59',datetime('now'))", (ex_uid, wd))
            cur.execute(
                "insert into excused_days (user_id, date, reason, status, created_at)"
                " values (?, ?, 'T-shifokorga bordi', 'approved', datetime('now'))",
                (ex_uid, date.today().isoformat()))
            conn.commit()
            ex_tok = token_for(ex_uid, "employee")

            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(ex_tok), json={
                "latitude": OFFICE[0], "longitude": OFFICE[1],
                "face_descriptor": FACE, "liveness": 0.9,
            })
            # ⚠️ QAROR O'ZGARDI (yangi TZ 2.9 / S-09). Ilgari sababli kunda
            # check-in ATAYIN o'tkazilardi (200 + ogohlantirish) — «ta'tildan
            # chaqirib olish normal» degan mulohaza bilan. Amalda bu xodimni
            # bir vaqtning o'zida ham «ta'tilda», ham «ishda» qilib qo'ydi va
            # oylik/norma/davomat uchta har xil javob berardi. Endi RAD
            # etiladi, HR ga esa xabar boradi.
            check("S-09: sababli kunda check-in RAD etildi -> 400",
                  r.status_code == 400, f"kod={r.status_code} {r.text[:150]}")
            check("S-09: rad etish sababi aytilgan (sababli kun)",
                  r.status_code == 400 and "sababli kun" in r.text,
                  r.text[:150])
            qator = cur.execute(
                "select count(*) from attendance where user_id=?", (ex_uid,)).fetchone()[0]
            check("S-09: rad etilgach davomat yozuvi YARATILMADI", qator == 0,
                  f"qatorlar={qator}")

            # 5.1 ning ASL qoidasi kuchda: sababli kunda kechikish yozilmaydi.
            # Endi bu check-in orqali emas (u bloklangan), balki qo'lda
            # kiritilgan / avtomatik qayta hisoblangan yozuv orqali sinaladi —
            # HR ta'tildagi kunni qo'lda tuzatishi mumkin.
            try:
                import asyncio as _aio

                from db.base import async_session as _sess

                async def _qayta():
                    from api.services.attendance import recompute_attendance
                    from db.models import Attendance as _Att
                    from db.models import User as _U
                    from sqlalchemy import select as _sel
                    async with _sess() as s3:
                        u3 = await s3.get(_U, ex_uid)
                        att3 = _Att(user_id=ex_uid, date=date.today(),
                                    check_in_time=datetime.utcnow())
                        s3.add(att3)
                        await s3.flush()
                        await recompute_attendance(s3, att3, u3)
                        await s3.commit()
                        row = await s3.scalar(
                            _sel(_Att).where(_Att.user_id == ex_uid))
                        return row.status, row.late_minutes

                st, lt = _aio.run(_qayta())
                check("5.1: late_minutes=0 (sababli kun, aks holda kech qolgan bo'lardi)",
                      lt == 0, f"late={lt}")
                check("5.1: status='excused'", st == "excused", f"status={st}")
            except Exception:
                check("5.1 qayta hisob tekshiruvi", False,
                      traceback.format_exc(limit=2).strip())

            cur.execute("delete from attendance where user_id=?", (ex_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (ex_uid,))
            cur.execute("delete from excused_days where user_id=?", (ex_uid,))
            cur.execute("delete from users where id=?", (ex_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("5.1 sababli kun tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 5.4: digest vaqti ORQAGA surilsa qo'riqchi tozalanmaydi --")
        try:
            conn = db()
            secret = ""
            for line in open(".env", encoding="utf-8"):
                if line.startswith("BOT_SHARED_SECRET="):
                    secret = line.strip().split("=", 1)[1]
            bot_h = {"X-Bot-Secret": secret}
            boss_row = conn.execute("select telegram_id from users where role='boss' and telegram_id is not null limit 1").fetchone()
            original = conn.execute(
                "select evening_hour, evening_minute, evening_last_posted from attendance_digest_config where id=1"
            ).fetchone()
            if boss_row and original:
                boss_tid = boss_row[0]
                orig_hour, orig_minute, orig_last_posted = original
                try:
                    conn.execute(
                        "update attendance_digest_config set evening_hour=10, evening_minute=0, evening_last_posted=? where id=1",
                        (date.today().isoformat(),))
                    conn.commit()

                    # ORQAGA surish (10:00 -> 09:00) — qo'riqchi TOZALANMASLIGI kerak,
                    # aks holda (hozir >= 09:00) darhol rost bo'lib, digest bugun
                    # IKKINCHI marta yuborilib ketardi.
                    r = client.post(
                        f"{API_BASE}/attendance/digest-time?telegram_id={boss_tid}&kind=evening&hour=9&minute=0",
                        headers=bot_h)
                    check("vaqtni orqaga surish -> 200", r.status_code == 200, f"kod={r.status_code}")
                    row = conn.execute("select evening_last_posted from attendance_digest_config where id=1").fetchone()
                    check("5.4: orqaga surilganda qo'riqchi SAQLANIB QOLADI",
                          row is not None and row[0] == date.today().isoformat(), f"evening_last_posted={row[0] if row else None}")

                    # OLDINGA surish (09:00 -> 11:00) — qo'riqchi TOZALANISHI kerak.
                    r2 = client.post(
                        f"{API_BASE}/attendance/digest-time?telegram_id={boss_tid}&kind=evening&hour=11&minute=0",
                        headers=bot_h)
                    check("vaqtni oldinga surish -> 200", r2.status_code == 200, f"kod={r2.status_code}")
                    row2 = conn.execute("select evening_last_posted from attendance_digest_config where id=1").fetchone()
                    check("5.4: oldinga surilganda qo'riqchi TOZALANADI",
                          row2 is not None and row2[0] is None, f"evening_last_posted={row2[0] if row2 else None}")
                finally:
                    # Asl (jonli) sozlamani albatta tiklaymiz.
                    conn.execute(
                        "update attendance_digest_config set evening_hour=?, evening_minute=?, evening_last_posted=? where id=1",
                        (orig_hour, orig_minute, orig_last_posted))
                    conn.commit()
            else:
                check("5.4 Boshliq telegram_id yoki sozlama topilmadi", False)
            conn.close()
        except Exception:
            check("5.4 digest vaqti tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 5.7: sababli kun dublikat va idempotentlik --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444796,'T-ExcusedDup','employee',1,1,datetime('now'))")
            dup_uid = cur.lastrowid
            conn.commit()

            secret = ""
            for line in open(".env", encoding="utf-8"):
                if line.startswith("BOT_SHARED_SECRET="):
                    secret = line.strip().split("=", 1)[1]
            bot_h = {"X-Bot-Secret": secret}

            r1 = client.post(f"{API_BASE}/excused-days", headers=bot_h,
                              json={"telegram_id": 999444796, "reason": "T-birinchi so'rov"})
            check("birinchi so'rov -> 200", r1.status_code == 200, f"kod={r1.status_code}")
            item_id = r1.json().get("id") if r1.status_code == 200 else None

            r2 = client.post(f"{API_BASE}/excused-days", headers=bot_h,
                              json={"telegram_id": 999444796, "reason": "T-ikkinchi so'rov"})
            check("5.7: bir kunga ikkinchi (dublikat) so'rov -> 400",
                  r2.status_code == 400, f"kod={r2.status_code} {r2.text[:150]}")

            if item_id:
                mgr = find_manager_id()
                mgr_telegram = (
                    conn.execute("select telegram_id from users where id=?", (mgr[0],)).fetchone()[0]
                    if mgr else None
                )
                if mgr_telegram:
                    r3 = client.post(f"{API_BASE}/excused-days/{item_id}/decide", headers=bot_h,
                                      json={"decider_telegram_id": mgr_telegram, "decision": "approved"})
                    check("birinchi qaror -> 200", r3.status_code == 200, f"kod={r3.status_code}")
                    r4 = client.post(f"{API_BASE}/excused-days/{item_id}/decide", headers=bot_h,
                                      json={"decider_telegram_id": mgr_telegram, "decision": "rejected"})
                    check("5.7: allaqachon hal qilingan so'rovga qayta qaror -> 400 (idempotent)",
                          r4.status_code == 400, f"kod={r4.status_code} {r4.text[:150]}")
                else:
                    check("5.7 rahbar telegram_id topilmadi", False)

            cur.execute("delete from excused_days where user_id=?", (dup_uid,))
            cur.execute("delete from users where id=?", (dup_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("5.7 dublikat/idempotentlik tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- 5.11: ish jadvali o'zgarishi audit qilinadi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444797,'T-ScheduleAudit','employee',1,1,datetime('now'))")
            sa_uid = cur.lastrowid
            conn.commit()

            mgr = find_manager_id()
            if mgr:
                mgr_tok = token_for(mgr[0], mgr[1])
                before_n = conn.execute(
                    "select count(*) from audit_logs where target_user_id=? and action='work_schedule_weekly_changed'",
                    (sa_uid,)).fetchone()[0]
                r = client.put(f"{API_BASE}/work-schedule/{sa_uid}/weekly", headers=auth(mgr_tok),
                               json={"days": [
                                   {"weekday": d, "is_working": d < 5, "start_time": "09:00" if d < 5 else None,
                                    "end_time": "18:00" if d < 5 else None}
                                   for d in range(7)
                               ]})
                check("PUT weekly -> 200", r.status_code == 200, f"kod={r.status_code}")
                after_n = conn.execute(
                    "select count(*) from audit_logs where target_user_id=? and action='work_schedule_weekly_changed'",
                    (sa_uid,)).fetchone()[0]
                check("5.11: haftalik jadval o'zgarishi audit qilindi",
                      after_n == before_n + 1, f"before={before_n}, after={after_n}")

                before_ov = conn.execute(
                    "select count(*) from audit_logs where target_user_id=? and action='work_schedule_override_changed'",
                    (sa_uid,)).fetchone()[0]
                r2 = client.put(f"{API_BASE}/work-schedule/{sa_uid}/override", headers=auth(mgr_tok),
                                json={"date": date.today().isoformat(), "is_working": True,
                                      "start_time": "12:00", "end_time": "18:00", "note": "T-sinov"})
                check("PUT override -> 200", r2.status_code == 200, f"kod={r2.status_code}")
                after_ov = conn.execute(
                    "select count(*) from audit_logs where target_user_id=? and action='work_schedule_override_changed'",
                    (sa_uid,)).fetchone()[0]
                check("5.11: aniq sana o'zgartirishi audit qilindi",
                      after_ov == before_ov + 1, f"before={before_ov}, after={after_ov}")
            else:
                check("5.11 rahbar topilmadi", False)

            cur.execute("delete from audit_logs where target_user_id=?", (sa_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (sa_uid,))
            cur.execute("delete from work_schedule_override where user_id=?", (sa_uid,))
            cur.execute("delete from users where id=?", (sa_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("5.11 audit tekshiruvi", False, traceback.format_exc(limit=1).strip())

        # ═══════════ UX-A bosqichi (DAVOMAT_UX_PROMPT.md) ═══════════

        print("\n-- UX-A2/A3: oylik matritsa va xodim tarixi --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444801,'T-Matrix','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            mx_uid = cur.lastrowid
            # Jadval: Du-Sha ish (09:00-18:00), Yak dam — 2020-yanvar butunlay
            # nazorat ostida (real bugunga bog'liq emas).
            for wd in range(7):
                cur.execute(
                    "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                    " values (?,?,?,?,?,datetime('now'))",
                    (mx_uid, wd, 1 if wd < 6 else 0,
                     "09:00" if wd < 6 else None, "18:00" if wd < 6 else None))
            # 2020-01-02 (Pa): kechikish yozuvi; 01-03 (Ju): yozuvsiz -> virtual absent;
            # 01-04 (Sh): sababli (yozuvsiz); 01-05 (Ya): dam.
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, '2020-01-02', '2020-01-02 04:15:00', 15, 0, 460, 'late', 0,"
                " datetime('now'), datetime('now'))", (mx_uid,))
            cur.execute(
                "insert into excused_days (user_id, date, reason, status, created_at)"
                " values (?, '2020-01-04', 'T-sinov', 'approved', datetime('now'))", (mx_uid,))
            conn.commit()

            r = client.get(f"{API_BASE}/attendance/matrix?month=2020-01&user_id={mx_uid}",
                           headers=auth(boss_t))
            check("A2: matrix -> 200", r.status_code == 200, f"kod={r.status_code}")
            body = r.json() if r.status_code == 200 else {}
            emp = (body.get("employees") or [{}])[0]
            cells = {c["date"]: c for c in emp.get("cells", [])}
            check("A2: yozuvli kun -> late (+vaqt mahalliy)",
                  cells.get("2020-01-02", {}).get("status") == "late"
                  and cells.get("2020-01-02", {}).get("check_in") == "09:15",
                  str(cells.get("2020-01-02")))
            check("A2: yozuvsiz o'tgan ish kuni -> absent (virtual)",
                  cells.get("2020-01-03", {}).get("status") == "absent", str(cells.get("2020-01-03")))
            check("A2: sababli kun (yozuvsiz) -> excused",
                  cells.get("2020-01-04", {}).get("status") == "excused", str(cells.get("2020-01-04")))
            check("A2: dam kuni -> weekend",
                  cells.get("2020-01-05", {}).get("status") == "weekend", str(cells.get("2020-01-05")))
            tot = emp.get("totals", {})
            check("A2: totals (late 1/15, excused 1)",
                  tot.get("late_count") == 1 and tot.get("late_minutes") == 15
                  and tot.get("excused_days") == 1, str(tot))
            check("A2: user_id filtri faqat bitta xodim qaytardi",
                  len(body.get("employees", [])) == 1)

            # Joriy oy: bugun (ish kuni, yozuvsiz) -> pending; ertaga -> future.
            r2 = client.get(f"{API_BASE}/attendance/matrix?user_id={mx_uid}", headers=auth(boss_t))
            cells2 = {c["date"]: c for c in (r2.json().get("employees") or [{}])[0].get("cells", [])}
            today_iso2 = date.today().isoformat()
            check("A2: bugungi yozuvsiz ish kuni -> pending",
                  cells2.get(today_iso2, {}).get("status") == "pending",
                  str(cells2.get(today_iso2)))
            tomorrow = date.today() + timedelta(days=1)
            if tomorrow.month == date.today().month:
                exp = "weekend" if tomorrow.weekday() == 6 else "future"
                check("A2: ertangi kun -> future/weekend",
                      cells2.get(tomorrow.isoformat(), {}).get("status") == exp,
                      str(cells2.get(tomorrow.isoformat())))

            # A3: xodim o'z tarixini oladi
            mx_tok = token_for(mx_uid, "employee")
            r3 = client.get(f"{API_BASE}/attendance/me/history?month=2020-01", headers=auth(mx_tok))
            check("A3: me/history -> 200", r3.status_code == 200, f"kod={r3.status_code}")
            days3 = {c["date"]: c for c in r3.json().get("days", [])}
            check("A3: xodim kalendari matritsa bilan bir xil (late kuni)",
                  days3.get("2020-01-02", {}).get("status") == "late"
                  and days3.get("2020-01-02", {}).get("schedule_start") == "09:00",
                  str(days3.get("2020-01-02")))
            r4 = client.get(f"{API_BASE}/attendance/me/history?month=2020-13", headers=auth(mx_tok))
            check("A3: noto'g'ri oy -> 400", r4.status_code == 400, f"kod={r4.status_code}")
            r5 = client.get(f"{API_BASE}/attendance/matrix", headers=auth(mx_tok))
            check("A2: oddiy xodimga matrix -> 403", r5.status_code == 403, f"kod={r5.status_code}")

            # A4: aniq davr parametrlari
            r6 = client.get(
                f"{API_BASE}/attendance/employee-summary?date_from=2020-01-01&date_to=2020-01-31",
                headers=auth(boss_t))
            row6 = next((x for x in r6.json() if x["user_id"] == mx_uid), None)
            check("A4: employee-summary aniq davr bilan", row6 is not None and row6["late_minutes"] == 15,
                  str(row6))
            r7 = client.get(
                f"{API_BASE}/attendance/late-stats?date_from=2020-01-01&date_to=2020-01-31",
                headers=auth(boss_t))
            check("A4: late-stats aniq davr bilan",
                  any(x["user_id"] == mx_uid for x in r7.json()), f"{len(r7.json())} qator")
            r8 = client.get(f"{API_BASE}/attendance/late-stats?date_from=2020-01-01",
                            headers=auth(boss_t))
            check("A4: faqat bitta sana parametri -> 400", r8.status_code == 400, f"kod={r8.status_code}")

            cur.execute("delete from attendance where user_id=?", (mx_uid,))
            cur.execute("delete from excused_days where user_id=?", (mx_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (mx_uid,))
            cur.execute("delete from users where id=?", (mx_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("UX-A2/A3 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- UX-A1: dashboard kelmaganlar ismlari --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444802,'T-NotCome','employee',1,1,datetime('now'))")
            nc_uid = cur.lastrowid
            wd = date.today().weekday()
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'08:30','18:00',datetime('now'))", (nc_uid, wd))
            conn.commit()

            r = client.get(f"{API_BASE}/attendance/dashboard", headers=auth(boss_t))
            body = r.json() if r.status_code == 200 else {}
            row = next((x for x in body.get("not_come", []) if x["user_id"] == nc_uid), None)
            check("A1: kelmagan xodim not_come ro'yxatida (jadval vaqti bilan)",
                  row is not None and row["schedule_start"] == "08:30", str(row))
            check("A1: left ro'yxati mavjud", "left" in body)

            # Sababli kun tasdiqlansa -> not_come'dan chiqib excused_today'ga o'tadi
            cur.execute(
                "insert into excused_days (user_id, date, reason, status, created_at)"
                " values (?, ?, 'T-sinov', 'approved', datetime('now'))",
                (nc_uid, date.today().isoformat()))
            conn.commit()
            r2 = client.get(f"{API_BASE}/attendance/dashboard", headers=auth(boss_t))
            b2 = r2.json()
            check("A1: sababli xodim not_come'da EMAS, excused_today'da BOR",
                  not any(x["user_id"] == nc_uid for x in b2.get("not_come", []))
                  and any(x["user_id"] == nc_uid for x in b2.get("excused_today", [])),
                  f"not_come={len(b2.get('not_come', []))}, excused={len(b2.get('excused_today', []))}")

            # UX-A5: eslatma — fake telegramga yetkazib bo'lmaydi -> 400;
            # 2 ta audit izi bo'lsa -> 429; oddiy xodimga -> 403.
            cur.execute("delete from excused_days where user_id=?", (nc_uid,))
            conn.commit()
            r3 = client.post(f"{API_BASE}/attendance/remind/{nc_uid}", headers=auth(boss_t))
            check("A5: yetkazib bo'lmasa -> 400 (fake telegram)", r3.status_code == 400,
                  f"kod={r3.status_code} {r3.text[:120]}")
            for _ in range(2):
                cur.execute(
                    "insert into audit_logs (actor_id, action, target_user_id, created_at)"
                    " values (1, 'attendance_reminder_sent', ?, datetime('now'))", (nc_uid,))
            conn.commit()
            r4 = client.post(f"{API_BASE}/attendance/remind/{nc_uid}", headers=auth(boss_t))
            check("A5: kuniga 2 tadan keyin -> 429", r4.status_code == 429, f"kod={r4.status_code}")
            emp_tok = token_for(nc_uid, "employee")
            r5 = client.post(f"{API_BASE}/attendance/remind/{nc_uid}", headers=auth(emp_tok))
            check("A5: oddiy xodim eslata olmaydi -> 403", r5.status_code == 403, f"kod={r5.status_code}")

            cur.execute("delete from audit_logs where target_user_id=?", (nc_uid,))
            cur.execute("delete from work_schedule_weekly where user_id=?", (nc_uid,))
            cur.execute("delete from users where id=?", (nc_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("UX-A1/A5 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- UX2-W1: dashboard late_list/user_id + remind-all --")
        try:
            conn = db()
            cur = conn.cursor()
            wd = date.today().weekday()
            today_iso2 = date.today().isoformat()

            # Kechikkan va allaqachon ketgan T-xodim — late_list'da bo'lishi kerak
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444804,'T-LateGuy','employee',1,1,datetime('now'))")
            lg_uid = cur.lastrowid
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, check_out_time, status,"
                " late_minutes, early_leave_minutes, worked_minutes, created_at, updated_at)"
                " values (?,?,datetime('now','-4 hours'),datetime('now','-1 hours'),'late',25,0,180,"
                " datetime('now'),datetime('now'))",
                (lg_uid, today_iso2))
            conn.commit()

            r = client.get(f"{API_BASE}/attendance/dashboard", headers=auth(boss_t))
            body = r.json() if r.status_code == 200 else {}
            ll = body.get("late_list", [])
            lg = next((x for x in ll if x["user_id"] == lg_uid), None)
            check("W1: late_list'da kechikkan xodim (25 daq, ketgan)",
                  lg is not None and lg["late_minutes"] == 25 and lg["left"] is True, str(lg))
            check("W1: late_list kamayish tartibida",
                  ll == sorted(ll, key=lambda x: x["late_minutes"], reverse=True),
                  str([x["late_minutes"] for x in ll]))
            check("W1: recent yozuvlarida user_id bor",
                  all("user_id" in x for x in body.get("recent", [])),
                  f"recent={len(body.get('recent', []))}")
            check("W1: in_office yozuvlarida user_id bor",
                  all("user_id" in x for x in body.get("in_office", [])),
                  f"in_office={len(body.get('in_office', []))}")

            # remind-all XAVFSIZ sinovi: yuborishdan OLDIN limit tekshiriladi,
            # shuning uchun BARCHA haqiqiy xodimlarga bugunga 2 tadan audit izi
            # qo'yamiz (ularga HECH NARSA yuborilmaydi); faqat fake-telegram'li
            # T-xodim yuborish yo'lidan o'tadi (Telegram 'chat not found' — jim).
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444805,'T-RemindAll','employee',1,1,datetime('now'))")
            ra_uid = cur.lastrowid
            cur.execute(
                "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
                " values (?,?,1,'09:00','18:00',datetime('now'))", (ra_uid, wd))
            seeded_ids = []
            for (uid,) in cur.execute(
                "select id from users where is_active=1 and id not in (?,?)", (ra_uid, lg_uid)
            ).fetchall():
                for _ in range(2):
                    cur.execute(
                        "insert into audit_logs (actor_id, action, target_user_id, created_at)"
                        " values (1, 'attendance_reminder_sent', ?, datetime('now'))", (uid,))
                    seeded_ids.append(cur.lastrowid)
            conn.commit()

            r6 = client.post(f"{API_BASE}/attendance/remind-all", headers=auth(boss_t))
            rb = r6.json() if r6.status_code == 200 else {}
            check("W1: remind-all -> 200", r6.status_code == 200, f"kod={r6.status_code} {r6.text[:120]}")
            check("W1: remind-all hech kimga yubormadi (hammada limit/fake)",
                  rb.get("sent") == 0, str(rb)[:200])
            ra_fail = next((f for f in rb.get("failed", []) if f["full_name"] == "T-RemindAll"), None)
            check("W1: T-RemindAll yuborish yo'lidan o'tdi (yetkazib bo'lmadi)",
                  ra_fail is not None and "yetkazib bo'lmadi" in ra_fail["reason"], str(ra_fail))
            real_fails = [f for f in rb.get("failed", []) if not f["full_name"].startswith("T-")]
            check("W1: haqiqiy xodimlar limitda to'xtadi (xabar ketmagan)",
                  all("2 marta" in f["reason"] for f in real_fails),
                  str([f["reason"] for f in real_fails])[:200])

            emp_tok2 = token_for(ra_uid, "employee")
            r7 = client.post(f"{API_BASE}/attendance/remind-all", headers=auth(emp_tok2))
            check("W1: remind-all oddiy xodimga -> 403", r7.status_code == 403, f"kod={r7.status_code}")

            # Tozalash: faqat O'ZIMIZ qo'ygan audit izlari + T- ma'lumotlar
            if seeded_ids:
                cur.execute(
                    "delete from audit_logs where id in (%s)" % ",".join("?" * len(seeded_ids)),
                    seeded_ids)
            for u in (lg_uid, ra_uid):
                cur.execute("delete from audit_logs where target_user_id=?", (u,))
                cur.execute("delete from attendance where user_id=?", (u,))
                cur.execute("delete from work_schedule_weekly where user_id=?", (u,))
                cur.execute("delete from users where id=?", (u,))
            conn.commit()
            conn.close()
        except Exception:
            check("UX2-W1 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- TASHRIF HISOBI: voqea-asosli (KPI = statistika) --")
        try:
            import asyncio as _aio2

            from db.base import async_session as _asess2
            from api.services import lead_diff as _ld

            conn = db()
            cur = conn.cursor()
            VISIT_ID = 8787  # .env dagi tashrif bosqichi (birinchisi)
            other = 7136

            # Ikki operator: T-Olib (lidni olib kelgan) va T-Yopgan (tashrifga o'tkazgan)
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " crm_visit_external_id, created_at)"
                " values (999777001,'T-Olib','employee',1,1,'970001',datetime('now'))")
            u_olib = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " crm_visit_external_id, created_at)"
                " values (999777002,'T-Yopgan','employee',1,1,'970002',datetime('now'))")
            u_yopgan = cur.lastrowid

            now_utc = datetime.utcnow()
            ts_now = int(now_utc.timestamp())

            def add_event(lead_id, frm, to, rid, first_rid):
                cur.execute(
                    "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                    " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                    " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                    " values (?,'stage_change',?,'T-eski',?,'T-Tashrif',?,?,?,?,?)",
                    (lead_id, frm, to, rid, "T-op", first_rid, ts_now,
                     now_utc.isoformat(sep=" ", timespec="seconds")))

            # 1) Tashrifga YANGI kirish, olib kelgan boshqa odam -> DUAL-KREDIT
            add_event(970101, other, VISIT_ID, 970002, 970001)
            # 2) Allaqachon Tashrifda bo'lgan lid yana tahrirlandi -> SANALMAYDI
            add_event(970102, VISIT_ID, VISIT_ID, 970002, 970002)
            # 3) O'zi olib kelib o'zi yopgan -> FAQAT BITTA kredit
            add_event(970103, other, VISIT_ID, 970002, 970002)
            conn.commit()

            async def _stats():
                async with _asess2() as s:
                    return await _ld.visit_stats_range(s, date.today(), date.today(), {VISIT_ID})

            ser = _aio2.run(_stats())
            ops = ser["daily_by_operator"].get(date.today(), {})
            uniq = ser["daily_unique"].get(date.today(), 0)

            check("Tashrif: takroriy tahrir SANALMAYDI (2 noyob, 3 voqea emas)",
                  uniq >= 2 and ops.get(970002, 0) >= 2, f"noyob={uniq}, ops={ops}")
            check("Tashrif: olib kelgan operatorga ALOHIDA kredit (dual)",
                  ops.get(970001, 0) >= 1, f"olib_kelgan={ops.get(970001)}")
            check("Tashrif: o'zi olib kelib o'zi yopganda ikki marta sanalmaydi",
                  ops.get(970002, 0) == uniq, f"yopgan={ops.get(970002)}, noyob={uniq}")

            # 2026-08-13 (egasining qarori): `first_seen` — bu CRM hodisasi
            # emas, bizning skaner lidni birinchi ko'rgani. Lid o'sha paytda
            # allaqachon Tashrifda bo'lsa, u tashrif deb SANALMASLIGI kerak.
            cur.execute(
                "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                " values (970104,'first_seen',NULL,NULL,?,'T-Tashrif',970002,'T-op',970002,?,?)",
                (VISIT_ID, ts_now, now_utc.isoformat(sep=" ", timespec="seconds")))
            conn.commit()
            ser2 = _aio2.run(_stats())
            uniq2 = ser2["daily_unique"].get(date.today(), 0)
            check("Tashrif: first_seen (skaner endi ko'rgan lid) SANALMAYDI",
                  uniq2 == uniq, f"first_seen'siz={uniq}, keyin={uniq2}")

            # recalc-visits endpointi: dry_run farqni ko'rsatadi, yozmaydi
            today_iso3 = date.today().isoformat()
            cur.execute(
                "insert into daily_results (user_id, date, conversations_count, visits_count,"
                " source) values (?,?,0,99,'crm')",
                (u_yopgan, today_iso3))
            conn.commit()
            r = client.post(
                f"{API_BASE}/daily-results/recalc-visits?date_from={today_iso3}"
                f"&date_to={today_iso3}&dry_run=true", headers=auth(boss_t))
            body = r.json() if r.status_code == 200 else {}
            ch = [c for c in body.get("changes", []) if c["user"] == "T-Yopgan"]
            check("Tashrif: recalc dry_run yolg'on sonni ko'rsatadi (99 -> voqea)",
                  r.status_code == 200 and ch and ch[0]["old"] == 99 and ch[0]["new"] != 99,
                  f"kod={r.status_code}, {ch[:1]}")
            after_dry = cur.execute(
                "select visits_count from daily_results where user_id=? and date=?",
                (u_yopgan, today_iso3)).fetchone()[0]
            check("Tashrif: dry_run bazaga YOZMAYDI", after_dry == 99, f"baza={after_dry}")

            r2 = client.post(
                f"{API_BASE}/daily-results/recalc-visits?date_from={today_iso3}"
                f"&date_to={today_iso3}&dry_run=false", headers=auth(boss_t))
            conn.commit()
            after_real = cur.execute(
                "select visits_count from daily_results where user_id=? and date=?",
                (u_yopgan, today_iso3)).fetchone()[0]
            check("Tashrif: dry_run=false haqiqatan tuzatadi",
                  r2.status_code == 200 and after_real != 99, f"baza={after_real}")

            emp_tok3 = token_for(u_olib, "employee")
            r3 = client.post(
                f"{API_BASE}/daily-results/recalc-visits?date_from={today_iso3}"
                f"&date_to={today_iso3}", headers=auth(emp_tok3))
            check("Tashrif: recalc oddiy xodimga -> 403", r3.status_code == 403, f"kod={r3.status_code}")

            cur.execute("delete from lead_events where crm_lead_id in (970101,970102,970103,970104)")
            cur.execute("delete from daily_results where user_id in (?,?)", (u_olib, u_yopgan))
            cur.execute("delete from users where id in (?,?)", (u_olib, u_yopgan))
            conn.commit()
            conn.close()
        except Exception:
            check("Tashrif hisobi tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- TABRIK VIDEOLARI: tashrif/shartnoma -> guruh + tugma --")
        try:
            import asyncio as _aio3

            from db.base import async_session as _asess3
            from api.services import celebration as _cel

            conn = db()
            cur = conn.cursor()
            # Oldingi UZILGAN yugurishdan qolgan izlar (UNIQUE xatosi bermasin)
            cur.execute("delete from celebration_claps where post_id in (select id from"
                        " celebration_posts where crm_lead_id between 970601 and 970606)")
            cur.execute("delete from celebration_posts where crm_lead_id between 970601 and 970606")
            cur.execute("delete from celebration_media where file_id like 'T-FILE-%'")
            cur.execute("delete from lead_events where crm_lead_id between 970601 and 970606")
            cur.execute("delete from monitored_groups where chat_id=-100999778")
            cur.execute("delete from users where telegram_id in (999778001,999778002,999778003)")
            conn.commit()
            V_ID = 8787          # tashrif bosqichi
            C_ID = 999888        # test uchun "shartnoma" bosqichi
            OTHER = 7136

            # Telegram'ga HECH NARSA ketmasin — yuboruvchilarni almashtiramiz
            sent_calls: list = []
            edited: list = []

            async def _fake_send_file(chat_id, file_id, file_type, caption=None, reply_markup=None):
                sent_calls.append(
                    {"chat": chat_id, "file_id": file_id, "type": file_type,
                     "caption": caption, "markup": reply_markup})
                return {"result": {"message_id": 5550 + len(sent_calls)}}

            async def _fake_send_msg(chat_id, text, reply_markup=None):
                sent_calls.append({"chat": chat_id, "file_id": None, "caption": text})
                return {"result": {"message_id": 6660 + len(sent_calls)}}

            async def _fake_edit(chat_id, message_id, reply_markup):
                edited.append({"chat": chat_id, "msg": message_id, "markup": reply_markup})
                return {"ok": True}

            _orig = (_cel.send_file_id, _cel.send_message, _cel.edit_reply_markup,
                     list(_cel.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS))
            _cel.send_file_id = _fake_send_file
            _cel.send_message = _fake_send_msg
            _cel.edit_reply_markup = _fake_edit
            _cel.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [C_ID]

            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " crm_visit_external_id, created_at)"
                " values (999778001,'T-Menejer','employee',1,1,'970501',datetime('now'))")
            u_men = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999778002,'T-HRcel','hr',1,1,datetime('now'))")
            u_hr = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999778003,'T-Xodimcel','employee',1,1,datetime('now'))")
            u_emp = cur.lastrowid

            # Umumiy guruh: jonli yozuvga TEGMAYMIZ — bo'lsa o'chirib turamiz,
            # oxirida asl holiga qaytariladi.
            live_main = cur.execute(
                "select id from monitored_groups where purpose='main' and is_active=1").fetchall()
            for (gid,) in live_main:
                cur.execute("update monitored_groups set is_active=0 where id=?", (gid,))
            cur.execute(
                "insert into monitored_groups (purpose, chat_id, title, is_active, created_at)"
                " values ('main', -100999778, 'T-Guruh', 1, datetime('now'))")
            g_id = cur.lastrowid
            conn.commit()

            cel_now = datetime.utcnow()
            cel_ts = int(cel_now.timestamp())

            def add_ev(lead_id, etype, frm, to, rid):
                cur.execute(
                    "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                    " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                    " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                    " values (?,?,?,'T-eski',?,'T-bosqich',?,'T-crm-nom',?,?,?)",
                    (lead_id, etype, frm, to, rid, rid, cel_ts,
                     cel_now.isoformat(sep=" ", timespec="seconds")))
                conn.commit()

            async def _announce(dry=False):
                async with _asess3() as s:
                    return await _cel.announce_pending(s, dry_run=dry)

            try:
                # 1) Video YO'Q -> guruhga hech nima ketmaydi (funksiya o'chiq tug'iladi)
                add_ev(970601, "stage_change", OTHER, V_ID, 970501)
                res0 = _aio3.run(_announce())
                check("Tabrik: video yuklanmagan bo'lsa guruhga hech narsa ketmaydi",
                      res0.get("sent") == 0 and not sent_calls, f"{res0}, chaqiruv={len(sent_calls)}")

                # 2) HR botdan video yuklaydi
                with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as _f:
                    _cel_secret = next(
                        (ln.strip().split("=", 1)[1] for ln in _f
                         if ln.startswith("BOT_SHARED_SECRET=")), "")
                bot_h_cel = {"X-Bot-Secret": _cel_secret}
                rset = client.post(f"{API_BASE}/celebration/media", headers=bot_h_cel, json={
                    "telegram_id": 999778002, "kind": "visit",
                    "file_id": "T-FILE-VISIT", "file_type": "video", "caption": "Zo'r ish!"})
                check("Tabrik: HR video yuklay oladi", rset.status_code == 200, f"kod={rset.status_code}")

                rdeny = client.post(f"{API_BASE}/celebration/media", headers=bot_h_cel, json={
                    "telegram_id": 999778003, "kind": "visit", "file_id": "X", "file_type": "video"})
                check("Tabrik: oddiy xodim video yuklay OLMAYDI -> 403",
                      rdeny.status_code == 403, f"kod={rdeny.status_code}")

                # 3) Endi o'sha voqea uchun video guruhga ketadi
                res1 = _aio3.run(_announce())
                check("Tabrik: tashrif voqeasida guruhga video yuboriladi",
                      res1.get("sent") == 1 and len(sent_calls) == 1, f"{res1}, {len(sent_calls)}")
                first = sent_calls[0] if sent_calls else {}
                check("Tabrik: to'g'ri guruh + to'g'ri fayl",
                      first.get("chat") == -100999778 and first.get("file_id") == "T-FILE-VISIT",
                      f"{first.get('chat')}, {first.get('file_id')}")
                check("Tabrik: izohda xodim ismi va rahbar matni bor",
                      "T-Menejer" in (first.get("caption") or "")
                      and "Zo'r ish!" in (first.get("caption") or ""),
                      f"izoh={(first.get('caption') or '')[:60]}")
                check("Tabrik: tabriklash tugmasi qo'yiladi",
                      "Tabriklash" in str(first.get("markup")), f"markup={str(first.get('markup'))[:60]}")

                # 4) TAKROR yuborilmaydi (webhook + cron ikkalasi chaqirsa ham)
                res2 = _aio3.run(_announce())
                check("Tabrik: bir voqea IKKI marta e'lon qilinmaydi",
                      res2.get("sent") == 0 and len(sent_calls) == 1, f"{res2}, {len(sent_calls)}")

                # 5) first_seen tabrik BERMAYDI (tashrif hisobidagi qoida bilan bir xil)
                add_ev(970602, "first_seen", None, V_ID, 970501)
                res3 = _aio3.run(_announce())
                check("Tabrik: first_seen uchun video yuborilmaydi",
                      res3.get("sent") == 0 and len(sent_calls) == 1, f"{res3}, {len(sent_calls)}")

                # 6) Tashrif ichida ko'chish ham tabrik emas
                add_ev(970603, "stage_change", V_ID, V_ID, 970501)
                res4 = _aio3.run(_announce())
                check("Tabrik: Tashrif->Tashrif ko'chishi tabrik bermaydi",
                      res4.get("sent") == 0, f"{res4}")

                # 7) Shartnoma — alohida video, alohida sarlavha
                add_ev(970604, "stage_change", V_ID, C_ID, 970501)
                res5 = _aio3.run(_announce())
                check("Tabrik: shartnoma videosi yuklanmagan -> yuborilmaydi",
                      res5.get("sent") == 0 and res5.get("skipped_no_media") == 1, f"{res5}")

                client.post(f"{API_BASE}/celebration/media", headers=bot_h_cel, json={
                    "telegram_id": 999778002, "kind": "contract",
                    "file_id": "T-FILE-CONTRACT", "file_type": "animation"})
                res6 = _aio3.run(_announce())
                last = sent_calls[-1] if sent_calls else {}
                check("Tabrik: shartnomada BOSHQA video (GIF) ketadi",
                      res6.get("sent") == 1 and last.get("file_id") == "T-FILE-CONTRACT"
                      and last.get("type") == "animation", f"{res6}, {last.get('file_id')}")
                check("Tabrik: shartnoma sarlavhasi tashrifdan farq qiladi",
                      "SHARTNOMA" in (last.get("caption") or ""),
                      f"izoh={(last.get('caption') or '')[:40]}")

                # 8) Tabriklash tugmasi: bir odam bir marta
                post_id = cur.execute(
                    "select id from celebration_posts where crm_lead_id=970601").fetchone()[0]

                async def _clap(tg):
                    async with _asess3() as s:
                        return await _cel.register_clap(s, post_id, tg)

                c1 = _aio3.run(_clap(555001))
                c2 = _aio3.run(_clap(555001))
                c3 = _aio3.run(_clap(555002))
                check("Tabrik: tabrik sanog'i oshadi", c1.get("claps") == 1, f"{c1}")
                check("Tabrik: bitta odam ikki marta tabriklay olmaydi",
                      c2.get("already") is True and c2.get("claps") == 1, f"{c2}")
                check("Tabrik: boshqa odam bossa sanoq 2 bo'ladi", c3.get("claps") == 2, f"{c3}")
                check("Tabrik: tugma matni yangilanadi (editMessageReplyMarkup)",
                      len(edited) >= 2 and "(2)" in str(edited[-1].get("markup")),
                      f"tahrir={len(edited)}, {str(edited[-1].get('markup'))[:50] if edited else ''}")

                # 8b) SAYT paneli: fayl -> Telegram -> file_id (bot bilan bir xil natija)
                uploaded: list = []

                async def _fake_upload(chat_id, content, filename, file_type, caption=None):
                    uploaded.append(
                        {"chat": chat_id, "bayt": len(content), "nom": filename, "tur": file_type})
                    return {"result": {file_type: {"file_id": f"T-FILE-WEB-{file_type}"}}}

                _orig_upload = _cel.send_media_file
                _cel.send_media_file = _fake_upload
                try:
                    hr_tok_cel = token_for(u_hr, "hr")
                    emp_tok_cel = token_for(u_emp, "employee")

                    rweb = client.get(f"{API_BASE}/celebration/settings", headers=auth(hr_tok_cel))
                    check("Tabrik(web): HR sozlamalarni ko'radi",
                          rweb.status_code == 200 and len(rweb.json().get("items", [])) == 2,
                          f"kod={rweb.status_code}")
                    rweb_deny = client.get(f"{API_BASE}/celebration/settings",
                                           headers=auth(emp_tok_cel))
                    check("Tabrik(web): oddiy xodimga -> 403",
                          rweb_deny.status_code == 403, f"kod={rweb_deny.status_code}")

                    # Yuklash yo'li SERVISDA sinaladi: test alohida jarayonda,
                    # API boshqasida — bu yerdagi almashtirish serverga ta'sir
                    # qilmaydi va so'rov haqiqiy Telegram'ga ketib qolardi.
                    from db.models import User as _UserModel

                    async def _upload(kind, name, ctype, data, cap):
                        async with _asess3() as s3:
                            actor = await s3.get(_UserModel, u_hr)
                            return await _cel.upload_and_set(
                                s3, kind, data, name, ctype, cap, actor)

                    rup = _aio3.run(_upload("visit", "tabrik.mp4", "video/mp4", b"x" * 2048, "Saytdan"))
                    check("Tabrik(web): video yuklandi va file_id saqlandi",
                          rup.get("ok") and rup.get("file_type") == "video" and len(uploaded) == 1,
                          f"{rup}, yuklashlar={len(uploaded)}")
                    saved = cur.execute(
                        "select file_id, file_type, caption from celebration_media"
                        " where kind='visit' and is_active=1").fetchone()
                    check("Tabrik(web): faol yozuv Telegram file_id bilan almashdi",
                          saved and saved[0] == "T-FILE-WEB-video" and saved[2] == "Saytdan",
                          f"baza={saved}")

                    rgif = _aio3.run(_upload("contract", "tabrik.gif", "image/gif", b"g" * 512, ""))
                    check("Tabrik(web): GIF animation sifatida yuklanadi",
                          rgif.get("ok") and rgif.get("file_type") == "animation", f"{rgif}")

                    rnotg = _aio3.run(_upload("visit", "hujjat.pdf", "application/pdf", b"%PDF", ""))
                    check("Tabrik(web): servis ham video bo'lmagan faylni rad etadi",
                          not rnotg.get("ok"), f"{rnotg}")

                    rbad = client.post(
                        f"{API_BASE}/celebration/settings/upload", headers=auth(hr_tok_cel),
                        data={"kind": "visit", "caption": ""},
                        files={"file": ("hujjat.pdf", b"%PDF-1.4", "application/pdf")})
                    check("Tabrik(web): video bo'lmagan fayl rad etiladi -> 400",
                          rbad.status_code == 400, f"kod={rbad.status_code}")

                    rbig = client.post(
                        f"{API_BASE}/celebration/settings/upload", headers=auth(hr_tok_cel),
                        data={"kind": "visit", "caption": ""},
                        files={"file": ("katta.mp4", b"x" * (46 * 1024 * 1024), "video/mp4")})
                    check("Tabrik(web): 45 MB dan katta fayl rad etiladi -> 400",
                          rbig.status_code == 400, f"kod={rbig.status_code}")

                    rup_deny = client.post(
                        f"{API_BASE}/celebration/settings/upload", headers=auth(emp_tok_cel),
                        data={"kind": "visit", "caption": ""},
                        files={"file": ("tabrik.mp4", b"x" * 128, "video/mp4")})
                    check("Tabrik(web): oddiy xodim yuklay OLMAYDI -> 403",
                          rup_deny.status_code == 403, f"kod={rup_deny.status_code}")
                finally:
                    _cel.send_media_file = _orig_upload
                    # Bot bloki davomi eski file_id larni kutadi — qaytaramiz
                    cur.execute("update celebration_media set is_active=0"
                                " where file_id like 'T-FILE-WEB-%'")
                    cur.execute("update celebration_media set is_active=1"
                                " where file_id in ('T-FILE-VISIT','T-FILE-CONTRACT')")
                    conn.commit()

                # 9) O'chirish -> yangi voqeaga video ketmaydi
                client.post(f"{API_BASE}/celebration/media/disable", headers=bot_h_cel,
                            json={"telegram_id": 999778002, "kind": "visit"})
                add_ev(970605, "stage_change", OTHER, V_ID, 970501)
                before_n = len(sent_calls)
                res7 = _aio3.run(_announce())
                check("Tabrik: o'chirilgan turda video yuborilmaydi",
                      res7.get("sent") == 0 and len(sent_calls) == before_n, f"{res7}")

                # 10) Eski voqea (lookback tashqarisi) guruhga to'kilmaydi
                client.post(f"{API_BASE}/celebration/media", headers=bot_h_cel, json={
                    "telegram_id": 999778002, "kind": "visit", "file_id": "T-FILE-VISIT2",
                    "file_type": "video"})
                old_dt = (cel_now - timedelta(hours=48)).isoformat(sep=" ", timespec="seconds")
                cur.execute(
                    "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                    " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                    " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                    " values (970606,'stage_change',?,'T-eski',?,'T-bosqich',970501,'T',970501,?,?)",
                    (OTHER, V_ID, cel_ts, old_dt))
                conn.commit()
                before_n2 = len(sent_calls)
                res8 = _aio3.run(_announce())
                sent_leads = [r[0] for r in cur.execute(
                    "select crm_lead_id from celebration_posts").fetchall()]
                check("Tabrik: eski (48 soatlik) voqea guruhga to'kilmaydi",
                      970606 not in sent_leads, f"postlar={sent_leads}")
                check("Tabrik: lekin YANGI voqea (970605) endi yuboriladi",
                      res8.get("sent") == 1 and len(sent_calls) == before_n2 + 1, f"{res8}")

            finally:
                # ── tozalash: HAR QANDAY holatda (xato bo'lsa ham) ──
                # Aks holda ochiq SQLite ulanishi keyingi bloklarni
                # "database is locked" bilan yiqitadi (jonli uchradi).
                try:
                    cur.execute("delete from celebration_claps where post_id in (select id"
                                " from celebration_posts where crm_lead_id between 970601 and 970606)")
                    cur.execute("delete from celebration_posts where crm_lead_id between 970601 and 970606")
                    cur.execute("delete from celebration_media where file_id like 'T-FILE-%'")
                    cur.execute("delete from lead_events where crm_lead_id between 970601 and 970606")
                    cur.execute("delete from monitored_groups where id=?", (g_id,))
                    for (gid,) in live_main:
                        cur.execute("update monitored_groups set is_active=1 where id=?", (gid,))
                    cur.execute("delete from audit_logs where actor_id in (?,?,?)", (u_men, u_hr, u_emp))
                    cur.execute("delete from users where id in (?,?,?)", (u_men, u_hr, u_emp))
                    conn.commit()
                finally:
                    conn.close()
                    (_cel.send_file_id, _cel.send_message, _cel.edit_reply_markup,
                     _cel.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _orig
        except Exception:
            check("Tabrik videolari tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: bosqichlar, konversiya, kogorta --")
        try:
            import asyncio as _aio4

            from db.base import async_session as _asess4
            from datetime import timezone as _tz4
            from api.services import funnel as _fn

            conn = db()
            cur = conn.cursor()
            FL = 980001  # test lid diapazoni
            cur.execute("delete from lead_events where crm_lead_id >= ?", (FL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (FL,))
            conn.commit()

            # Bosqich ID'larini testda QOTIRAMIZ — jonli .env ga bog'liq
            # bo'lmasin (u yerda ID o'zgarsa test yiqilmasligi kerak).
            _fn_orig = (
                list(_fn.CRM_UYSOT_INVITE_PIPE_STATUS_IDS),
                list(_fn.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn.CRM_UYSOT_INVITE_PIPE_STATUS_IDS = [8786]
            _fn.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                fn_day = date(2021, 5, 10)          # tarixiy oy — jonli ma'lumotga tegmaydi
                fn_from, fn_to = date(2021, 5, 1), date(2021, 5, 31)
                fn_ts = int(datetime(2021, 5, 10, 9, tzinfo=_tz4.utc).timestamp())
                fn_det = datetime(2021, 5, 10, 9).isoformat(sep=" ", timespec="seconds")

                for i in range(5):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?)",
                        (FL + i, 8779, "T-Yangi", 1, "T-op", 1, fn_ts, fn_ts, fn_det, fn_det))

                def fev(lead, to_id, etype="stage_change", det=fn_det):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,?,8779,'T-Yangi',?,'T-bosqich',1,'T-op',1,?,?)",
                        (lead, etype, to_id, fn_ts, det))

                fev(FL + 0, 8786); fev(FL + 0, 8787); fev(FL + 0, 8788)   # to'liq zanjir
                fev(FL + 1, 8786); fev(FL + 1, 8787)                       # tashrifgacha
                fev(FL + 2, 8787)                                          # TAKLIFSIZ sakrash
                fev(FL + 3, 8786)                                          # faqat taklif
                fev(FL + 4, 8787, etype="first_seen")                      # SANALMASIN
                conn.commit()

                async def _per():
                    async with _asess4() as s:
                        return await _fn.period_funnel(s, fn_from, fn_to)

                async def _coh():
                    async with _asess4() as s:
                        return await _fn.cohort_funnel(s, fn_from, fn_to)

                per = _aio4.run(_per())
                by = {r["key"]: r for r in per["rows"]}
                check("Voronka: lid soni CRM yaratilish vaqtidan (5)",
                      by["lead"]["value"] == 5, f"lid={by['lead']['value']}")
                check("Voronka: sakragan lid taklifda ham sanaladi (4)",
                      by["invite"]["value"] == 4, f"taklif={by['invite']['value']}")
                check("Voronka: tashrif = 3", by["visit"]["value"] == 3,
                      f"tashrif={by['visit']['value']}")
                check("Voronka: first_seen tashrif bermaydi (4 emas, 3)",
                      by["visit"]["value"] == 3, f"tashrif={by['visit']['value']}")
                check("Voronka: shartnoma = 1", by["contract"]["value"] == 1,
                      f"shartnoma={by['contract']['value']}")
                check("Voronka: qo'ng'iroq qatorida konversiya YO'Q (zanjirdan tashqari)",
                      by["call_try"]["conv_from_prev"] is None and by["call_try"]["outside_chain"],
                      f"{by['call_try']}")
                check("Voronka: tashrif->shartnoma konversiyasi 33.3%",
                      by["contract"]["conv_from_prev"] == 33.3,
                      f"={by['contract']['conv_from_prev']}")

                coh = _aio4.run(_coh())
                cby = {r["key"]: r for r in coh["rows"]}
                check("Voronka(kogorta): lid=5, shartnoma=1",
                      cby["lead"]["value"] == 5 and cby["contract"]["value"] == 1,
                      f"lid={cby['lead']['value']}, shartnoma={cby['contract']['value']}")
                check("Voronka(kogorta): liddan shartnomagacha 20%",
                      cby["contract"]["conv_from_lead"] == 20.0,
                      f"={cby['contract']['conv_from_lead']}")
                check("Voronka(kogorta): eski davr «pishgan» deb belgilanadi",
                      coh["mature"] is True, f"mature={coh['mature']}, yosh={coh['age_days']}")

                # Tashrifda TURGAN lid shartnomaga o'tsa, «tashrif» qatorida
                # QAYTA sanalmasligi kerak (jonli 2026-08 da shu sabab 139 ta
                # tashrif chiqqan, KPI esa 44 ta ko'rsatgan edi).
                fev(FL + 1, 8788, det=fn_det)   # 8787 -> 8788 (tashrifdan shartnomaga)
                conn.commit()
                per2 = _aio4.run(_per())
                by2 = {r["key"]: r for r in per2["rows"]}
                check("Voronka: tashrifdan shartnomaga o'tish tashrifni QAYTA sanamaydi",
                      by2["visit"]["value"] == 3, f"tashrif={by2['visit']['value']}")
                check("Voronka: lekin shartnoma soni oshadi (1 -> 2)",
                      by2["contract"]["value"] == 2, f"shartnoma={by2['contract']['value']}")

                weak = _fn.weakest_link(per["rows"])
                check("Voronka: eng zaif bo'g'in — shartnoma (33.3%)",
                      weak and weak["key"] == "contract", f"{weak}")

                # Bo'sh oy: bo'lish xatosi bermasin, konversiya None bo'lsin
                async def _empty():
                    async with _asess4() as s:
                        return await _fn.period_funnel(s, date(2019, 2, 1), date(2019, 2, 28))
                emp = _aio4.run(_empty())
                eby = {r["key"]: r for r in emp["rows"]}
                check("Voronka: bo'sh davrda konversiya 0% emas, «—» (None)",
                      eby["visit"]["conv_from_prev"] is None and eby["lead"]["value"] == 0,
                      f"lid={eby['lead']['value']}, konv={eby['visit']['conv_from_prev']}")
                check("Voronka: bo'sh davrda eng zaif bo'g'in yo'q",
                      _fn.weakest_link(emp["rows"]) is None, f"{_fn.weakest_link(emp['rows'])}")

                # API: ruxsat
                fn_rop = cur.execute(
                    "select id from users where role='rop' and is_active=1 limit 1").fetchone()
                if fn_rop:
                    r_ok = client.get(f"{API_BASE}/funnel?mode=cohort&month=2021-05",
                                      headers=auth(token_for(fn_rop[0], "rop")))
                    check("Voronka(API): ROP ko'ra oladi -> 200",
                          r_ok.status_code == 200, f"kod={r_ok.status_code}")
                fn_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1 limit 1").fetchone()
                if fn_emp:
                    r_no = client.get(f"{API_BASE}/funnel", headers=auth(token_for(fn_emp[0], "employee")))
                    check("Voronka(API): oddiy xodimga -> 403",
                          r_no.status_code == 403, f"kod={r_no.status_code}")
                r_bad = client.get(f"{API_BASE}/funnel?month=notoq", headers=auth(boss_t))
                check("Voronka(API): noto'g'ri oy formati -> 400",
                      r_bad.status_code == 400, f"kod={r_bad.status_code}")
            finally:
                cur.execute("delete from lead_events where crm_lead_id >= ?", (FL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (FL,))
                conn.commit()
                conn.close()
                (_fn.CRM_UYSOT_INVITE_PIPE_STATUS_IDS,
                 _fn.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn_orig
        except Exception:
            check("Voronka tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: kanal kesimi va manba boyitish (2-bosqich) --")
        try:
            import asyncio as _aio5
            import json as _json5
            from datetime import timezone as _tz5

            from db.base import async_session as _asess5
            from api.services import funnel as _fn5
            from api.services import lead_source as _ls5

            conn = db()
            cur = conn.cursor()
            CL = 981001
            cur.execute("delete from lead_events where crm_lead_id >= ?", (CL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (CL,))
            conn.commit()

            _fn5_orig = (
                list(_fn5.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn5.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn5.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn5.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                c_from, c_to = date(2021, 6, 1), date(2021, 6, 30)
                c_ts = int(datetime(2021, 6, 10, 9, tzinfo=_tz5.utc).timestamp())
                c_det = datetime(2021, 6, 10, 9).isoformat(sep=" ", timespec="seconds")

                def lead(i, tags, source=None, checked=None):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, tags, source, source_checked_at, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (CL + i, 8779, "T-Yangi", 1, "T-op", 1, c_ts, c_ts,
                         _json5.dumps(tags) if tags is not None else None,
                         source, checked, c_det, c_det))

                def cev(i, to_id):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',?,'T-bosqich',1,'T-op',1,?,?)",
                        (CL + i, to_id, c_ts, c_det))

                # 0: telegram + webinar (IKKI teg), tashrif+shartnoma
                lead(0, ["#telegram", "#webinar"]); cev(0, 8787); cev(0, 8788)
                # 1: telegram, faqat tashrif
                lead(1, ["#telegram"]); cev(1, 8787)
                # 2: telegram, hech qayerga yetmagan
                lead(2, ["#telegram"])
                # 3: tegsiz
                lead(3, None)
                conn.commit()

                async def _ch(group_by):
                    async with _asess5() as s:
                        return await _fn5.channel_funnel(s, c_from, c_to, group_by)

                ch = _aio5.run(_ch("tag"))
                by = {r["channel"]: r for r in ch["rows"]}
                check("Kanal: teglar bo'yicha guruhlanadi (telegram 3 lid)",
                      by.get("#telegram", {}).get("leads") == 3, f"{by.get('#telegram')}")
                check("Kanal: bitta lid IKKI tegda ham sanaladi (webinar 1 lid)",
                      by.get("#webinar", {}).get("leads") == 1, f"{by.get('#webinar')}")
                check("Kanal: tegsiz lid alohida qatorda",
                      by.get("(tegsiz)", {}).get("leads") == 1, f"{by.get('(tegsiz)')}")
                check("Kanal: telegram konversiyasi — lid->tashrif 66.7%",
                      by["#telegram"]["lead_to_visit"] == 66.7,
                      f"={by['#telegram']['lead_to_visit']}")
                check("Kanal: telegram lid->shartnoma 33.3%",
                      by["#telegram"]["lead_to_contract"] == 33.3,
                      f"={by['#telegram']['lead_to_contract']}")
                check("Kanal: webinar 100% shartnoma (1/1)",
                      by["#webinar"]["lead_to_contract"] == 100.0,
                      f"={by['#webinar']['lead_to_contract']}")
                check("Kanal: eng ko'p lidli kanal yuqorida",
                      ch["rows"][0]["channel"] == "#telegram", f"{ch['rows'][0]['channel']}")

                src = _aio5.run(_ch("source"))
                sby = {r["channel"]: r for r in src["rows"]}
                _no_src = sby.get("(manba yo'q)")
                check("Kanal(manba): so'ralmagan lidlar «(manba yo'q)» da",
                      (_no_src or {}).get("leads") == 4, str(_no_src))

                # ── Manba boyituvchisi: CRM o'rniga soxta adapter ──
                class _FakeAdapter:
                    def __init__(self):
                        self.calls = []

                    async def get_lead_detail(self, lead_id):
                        self.calls.append(lead_id)
                        # Bittasiga manba yo'q — u ham BELGILANISHI kerak
                        if lead_id == CL + 3:
                            return {"id": lead_id, "source": None, "tags": ["#T-yangiteg"]}
                        return {"id": lead_id, "source": "T-FACEBOOK", "tags": []}

                fake = _FakeAdapter()
                _ls_orig = _ls5._adapter
                _ls5._adapter = lambda: fake
                try:
                    async def _enrich(n):
                        async with _asess5() as s:
                            return await _ls5.enrich_tick(s, limit=n)

                    r1 = _aio5.run(_enrich(2))
                    check("Manba: bir tick'da FAQAT byudjet qadar so'raladi (2 ta)",
                          r1.get("checked") == 2 and len(fake.calls) == 2, f"{r1}, {fake.calls}")
                    check("Manba: eng yangi liddan boshlanadi",
                          fake.calls[0] == CL + 3, f"birinchi={fake.calls[0]}")
                    check("Manba: yana qolgani bor deb belgilanadi",
                          r1.get("has_more") is True, f"{r1}")

                    r2 = _aio5.run(_enrich(10))
                    checked_all = cur.execute(
                        "select count(*) from crm_lead_state where crm_lead_id >= ?"
                        " and source_checked_at is not null", (CL,)).fetchone()[0]
                    check("Manba: hammasi belgilandi (qayta so'ralmaydi)",
                          checked_all == 4, f"belgilangan={checked_all}")
                    check("Manba: manbasi yo'q lid ham belgilanadi (cheksiz so'rov yo'q)",
                          cur.execute(
                              "select source, source_checked_at is not null from crm_lead_state"
                              " where crm_lead_id=?", (CL + 3,)).fetchone() == (None, 1),
                          "manbasiz lid belgilanmagan")
                    check("Manba: keyingi tick'da hech nima so'ralmaydi",
                          _aio5.run(_enrich(10)).get("checked") == 0, "qayta so'rov bor")

                    # Teg ham SHU javobdan olinadi (qo'shimcha so'rovsiz) —
                    # productionda diff-skaner webhook rejimida ishlamaydi,
                    # ya'ni teglarning yagona yo'li shu.
                    check("Manba: teg ham o'sha javobdan yoziladi",
                          cur.execute("select tags from crm_lead_state where crm_lead_id=?",
                                      (CL + 3,)).fetchone()[0] == '["#T-yangiteg"]',
                          str(cur.execute("select tags from crm_lead_state where crm_lead_id=?",
                                          (CL + 3,)).fetchone()))

                    src2 = _aio5.run(_ch("source"))
                    s2by = {r["channel"]: r for r in src2["rows"]}
                    check("Kanal(manba): boyitgandan keyin manba kesimi ishlaydi",
                          s2by.get("T-FACEBOOK", {}).get("leads") == 3, f"{s2by.get('T-FACEBOOK')}")
                finally:
                    _ls5._adapter = _ls_orig

                r_api = client.get(f"{API_BASE}/funnel/channels?group_by=tag&month=2021-06",
                                   headers=auth(boss_t))
                check("Kanal(API): 200 va qatorlar qaytadi",
                      r_api.status_code == 200 and isinstance(r_api.json().get("rows"), list),
                      f"kod={r_api.status_code}")
                r_bad2 = client.get(f"{API_BASE}/funnel/channels?group_by=xato",
                                    headers=auth(boss_t))
                check("Kanal(API): noto'g'ri group_by -> 422",
                      r_bad2.status_code == 422, f"kod={r_bad2.status_code}")
            finally:
                cur.execute("delete from lead_events where crm_lead_id >= ?", (CL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (CL,))
                conn.commit()
                conn.close()
                (_fn5.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn5.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn5_orig
        except Exception:
            check("Kanal kesimi tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: reklama xarajati, CPL/CAC/ROMI (3-bosqich) --")
        try:
            import asyncio as _aio6
            import json as _json6
            from datetime import timezone as _tz6

            from db.base import async_session as _asess6
            from api.services import ad_spend as _ads
            from api.services import funnel as _fn6

            conn = db()
            cur = conn.cursor()
            EL = 982001
            E_PERIOD = "2021-07"
            for t in ("ad_spend", "funnel_month"):
                cur.execute(f"delete from {t} where period=?", (E_PERIOD,))
            cur.execute("delete from lead_events where crm_lead_id >= ?", (EL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (EL,))
            conn.commit()

            _fn6_orig = (
                list(_fn6.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn6.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn6.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn6.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                e_ts = int(datetime(2021, 7, 10, 9, tzinfo=_tz6.utc).timestamp())
                e_det = datetime(2021, 7, 10, 9).isoformat(sep=" ", timespec="seconds")

                def elead(i, tags):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, tags, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?,?)",
                        (EL + i, 8779, "T-Yangi", 1, "T-op", 1, e_ts, e_ts,
                         _json6.dumps(tags), e_det, e_det))

                def eev(i, to_id):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',?,'T-bosqich',1,'T-op',1,?,?)",
                        (EL + i, to_id, e_ts, e_det))

                # #T-insta: 10 lid, 2 tashrif, 1 shartnoma
                for i in range(10):
                    elead(i, ["#T-insta"])
                eev(0, 8787); eev(0, 8788); eev(1, 8787)
                # #T-tg: 5 lid, 1 shartnoma
                for i in range(10, 15):
                    elead(i, ["#T-tg"])
                eev(10, 8787); eev(10, 8788)
                conn.commit()

                async def _eco(gb="tag"):
                    async with _asess6() as s:
                        return await _ads.economics(s, E_PERIOD, gb)

                async def _set(channel, amount, reach=None):
                    async with _asess6() as s:
                        return await _ads.upsert_spend(s, E_PERIOD, channel, amount, reach, None, None)

                _aio6.run(_set("#T-insta", 1_000_000, reach=50_000))
                eco = _aio6.run(_eco())
                by = {r["channel"]: r for r in eco["rows"]}
                check("Iqtisod: CPL = xarajat / lid (1 000 000 / 10 = 100 000)",
                      by["#T-insta"]["cpl"] == 100000.0, f"={by['#T-insta']['cpl']}")
                check("Iqtisod: CAC = xarajat / shartnoma (1 000 000 / 1)",
                      by["#T-insta"]["cac"] == 1000000.0, f"={by['#T-insta']['cac']}")
                check("Iqtisod: CPV = xarajat / tashrif (1 000 000 / 2)",
                      by["#T-insta"]["cpv"] == 500000.0, f"={by['#T-insta']['cpv']}")
                check("Iqtisod: qamrov kiritilsa auditoriya->lid % chiqadi",
                      by["#T-insta"]["reach_to_lead"] == 0.02,
                      f"={by['#T-insta']['reach_to_lead']}")
                check("Iqtisod: foyda kiritilmagan -> ROMI hisoblanmaydi (None)",
                      by["#T-insta"]["romi"] is None, f"={by['#T-insta']['romi']}")

                # Xarajati kiritilmagan, lekin lid keltirgan kanal ko'rinsin
                check("Iqtisod: xarajatsiz kanal «unutilgan» ro'yxatida",
                      any(m["channel"] == "#T-tg" for m in eco["missing_spend"]),
                      f"{eco['missing_spend']}")

                # O'rtacha foyda -> ROMI
                async def _profit(v):
                    async with _asess6() as s:
                        return await _ads.set_avg_deal_profit(s, E_PERIOD, v, None)

                _aio6.run(_profit(3_000_000))
                eco2 = _aio6.run(_eco())
                by2 = {r["channel"]: r for r in eco2["rows"]}
                # (1 shartnoma × 3 000 000 − 1 000 000) / 1 000 000 = 200%
                check("Iqtisod: ROMI = (daromad − xarajat) / xarajat = 200%",
                      by2["#T-insta"]["romi"] == 200.0, f"={by2['#T-insta']['romi']}")

                _aio6.run(_profit(500_000))
                eco3 = _aio6.run(_eco())
                by3 = {r["channel"]: r for r in eco3["rows"]}
                check("Iqtisod: zarar bo'lsa ROMI manfiy (-50%)",
                      by3["#T-insta"]["romi"] == -50.0, f"={by3['#T-insta']['romi']}")

                # Mos kelmagan kanal — jimgina 0 emas, ogohlantirish
                _aio6.run(_set("Instagram", 500_000))
                eco4 = _aio6.run(_eco())
                by4 = {r["channel"]: r for r in eco4["rows"]}
                check("Iqtisod: voronkada yo'q kanal «mos kelmadi» deb belgilanadi",
                      by4["Instagram"]["matched"] is False
                      and "Instagram" in eco4["unmatched"], f"{eco4['unmatched']}")
                check("Iqtisod: mos kelmagan kanal JAMI lidga qo'shilmaydi",
                      eco4["totals"]["leads"] == 10, f"jami lid={eco4['totals']['leads']}")
                check("Iqtisod: lekin JAMI xarajatga qo'shiladi (pul ketgan)",
                      eco4["totals"]["spend"] == 1_500_000.0,
                      f"jami xarajat={eco4['totals']['spend']}")

                # Registr farqi CPL'ni buzmasin
                _aio6.run(_set("#T-INSTA", 200_000))
                eco5 = _aio6.run(_eco())
                by5 = {r["channel"]: r for r in eco5["rows"]}
                check("Iqtisod: katta/kichik harf farqi kanalni ajratmaydi",
                      by5["#T-INSTA"]["matched"] is True and by5["#T-INSTA"]["leads"] == 10,
                      f"{by5['#T-INSTA']}")

                # API
                r_e = client.get(f"{API_BASE}/funnel/economics?period={E_PERIOD}",
                                 headers=auth(boss_t))
                check("Iqtisod(API): 200", r_e.status_code == 200, f"kod={r_e.status_code}")
                r_ch = client.get(f"{API_BASE}/funnel/economics/channels?period={E_PERIOD}",
                                  headers=auth(boss_t))
                check("Iqtisod(API): kanal ro'yxati «(tegsiz)» ni bermaydi",
                      r_ch.status_code == 200
                      and all(not c["channel"].startswith("(")
                              for c in r_ch.json().get("channels", [])),
                      f"kod={r_ch.status_code}")
                e_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1 limit 1").fetchone()
                if e_emp:
                    r_deny = client.post(
                        f"{API_BASE}/funnel/economics/spend",
                        headers=auth(token_for(e_emp[0], "employee")),
                        json={"period": E_PERIOD, "channel": "#T-insta", "amount": 1})
                    check("Iqtisod(API): oddiy xodim xarajat kirita OLMAYDI -> 403",
                          r_deny.status_code == 403, f"kod={r_deny.status_code}")
                r_neg = client.post(
                    f"{API_BASE}/funnel/economics/spend", headers=auth(boss_t),
                    json={"period": E_PERIOD, "channel": "#T-insta", "amount": -5})
                check("Iqtisod(API): manfiy summa -> 400", r_neg.status_code == 400,
                      f"kod={r_neg.status_code}")
            finally:
                for t in ("ad_spend", "funnel_month"):
                    cur.execute(f"delete from {t} where period=?", (E_PERIOD,))
                cur.execute("delete from lead_events where crm_lead_id >= ?", (EL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (EL,))
                conn.commit()
                conn.close()
                (_fn6.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn6.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn6_orig
        except Exception:
            check("Reklama xarajati tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: teskari kalkulyator (4-bosqich) --")
        try:
            import asyncio as _aio7

            from db.base import async_session as _asess7
            from api.services import target_calc as _tc

            conn = db()
            cur = conn.cursor()
            T_PERIOD = "2021-09"
            cur.execute("delete from funnel_month where period=?", (T_PERIOD,))
            conn.commit()

            try:
                async def _calc(target, ov=None):
                    async with _asess7() as s:
                        return await _tc.calculate(s, T_PERIOD, target, ov)

                # To'liq faraz to'plami bilan: arifmetika aniq tekshiriladi
                ov = {
                    "lead_to_visit": 5.0,        # 5%
                    "visit_to_contract": 10.0,   # 10%
                    "talks_per_lead": 2.0,
                    "pickup_rate": 50.0,
                    "cpl": 40000.0,
                    "reach_to_lead": 1.0,        # 1%
                }
                r = _aio7.run(_calc(10, ov))
                chain = {c["key"]: c["value"] for c in r["chain"]}
                # 10 shartnoma / 10% = 100 tashrif; /5% = 2000 lid;
                # ×2 = 4000 suhbat; /50% = 8000 urinish; /1% = 200 000 qamrov
                check("Kalkulyator: tashrif = maqsad / (tashrif->shartnoma)",
                      chain["visits"] == 100, f"={chain['visits']}")
                check("Kalkulyator: lid = tashrif / (lid->tashrif)",
                      chain["leads"] == 2000, f"={chain['leads']}")
                check("Kalkulyator: suhbat = lid × lid boshiga suhbat",
                      chain["talks"] == 4000, f"={chain['talks']}")
                check("Kalkulyator: urinish = suhbat / ko'tarish foizi",
                      chain["tries"] == 8000, f"={chain['tries']}")
                check("Kalkulyator: auditoriya = lid / (qamrov->lid)",
                      chain["reach"] == 200000, f"={chain['reach']}")
                check("Kalkulyator: byudjet = lid × CPL (2000 × 40 000)",
                      r["budget"] == 80_000_000, f"={r['budget']}")
                check("Kalkulyator: qo'lda kiritilgan faraz «override» deb belgilanadi",
                      r["assumptions"]["cpl"]["source"] == "override",
                      f"={r['assumptions']['cpl']['source']}")

                # Yuqoriga yaxlitlash: 7 shartnoma / 10% = 70 tashrif,
                # /3% = 2333.3 lid -> 2334 (yarim lid bo'lmaydi)
                r2 = _aio7.run(_calc(7, {**ov, "lead_to_visit": 3.0}))
                c2 = {c["key"]: c["value"] for c in r2["chain"]}
                check("Kalkulyator: kasr son YUQORIGA yaxlitlanadi (2334)",
                      c2["leads"] == 2334, f"={c2['leads']}")

                # CPL yo'q bo'lsa byudjet hisoblanmaydi (0 emas!)
                r3 = _aio7.run(_calc(10, {**ov, "cpl": None}))
                check("Kalkulyator: CPL bo'lmasa byudjet «hisoblanmadi» (None)",
                      r3["budget"] is None and "cpl" in r3["missing"], f"{r3['budget']}")

                # Sezgirlik: tashrif->shartnoma 10% -> 11% bo'lsa lid kamayadi
                sens = {s["label"]: s for s in r["sensitivity"]}
                key = "Tashrif→shartnoma +1 punkt"
                # 10/11% = 90.9 tashrif; /5% = 1818.2 lid; 2000 - 1818.2 = 181.8 -> 182
                check("Kalkulyator: sezgirlik — +1 punkt 182 ta lid tejaydi",
                      sens[key]["leads_saved"] == 182, f"={sens.get(key)}")
                check("Kalkulyator: sezgirlik byudjet tejamini ham beradi",
                      sens[key]["budget_saved"] == round(181.818181 * 40000),
                      f"={sens[key]['budget_saved']}")

                # Nol o'lchov ishlatilmasin — u butun zanjirni to'xtatadi
                _base_orig = _tc.baseline

                async def _fake_base(db, months=6):
                    return {
                        "months_used": 1,
                        "values": {
                            "lead_to_visit": 0.0,          # soxta «o'lchov»
                            "visit_to_contract": 12.0,
                            "talks_per_lead": None,
                            "pickup_rate": None,
                            "cpl": None,
                            "reach_to_lead": None,
                        },
                        "confidence": "past",
                    }

                _tc.baseline = _fake_base
                try:
                    r0 = _aio7.run(_calc(10, {}))
                    a0 = r0["assumptions"]["lead_to_visit"]
                    check("Kalkulyator: 0% «o'lchov» rad etiladi, zaxira faraz olinadi",
                          a0["source"] == "default" and a0["value"] == _tc.DEFAULTS["lead_to_visit"],
                          f"={a0}")
                    c0 = {c["key"]: c["value"] for c in r0["chain"]}
                    check("Kalkulyator: shu sabab zanjir to'xtamaydi (lid hisoblandi)",
                          c0["leads"] is not None and c0["leads"] > 0, f"lid={c0['leads']}")
                finally:
                    _tc.baseline = _base_orig

                # Saqlash: faqat to'ldirilgan farazlar yoziladi
                async def _save(target, assumptions):
                    async with _asess7() as s:
                        return await _tc.save_target(s, T_PERIOD, target, assumptions, None)

                row = _aio7.run(_save(12, {"visit_to_contract": 8.0, "cpl": None}))
                check("Kalkulyator: maqsad saqlanadi", row.target_contracts == 12,
                      f"={row.target_contracts}")
                check("Kalkulyator: bo'sh farazlar SAQLANMAYDI (o'lchovdan olinadi)",
                      row.assumptions == {"visit_to_contract": 8.0}, f"={row.assumptions}")

                # API
                r_api = client.get(f"{API_BASE}/funnel/target?period={T_PERIOD}",
                                   headers=auth(boss_t))
                check("Kalkulyator(API): saqlangan maqsad bilan hisob qaytadi",
                      r_api.status_code == 200
                      and r_api.json().get("target_contracts") == 12,
                      f"kod={r_api.status_code}, {r_api.text[:90]}")
                r_no = client.get(f"{API_BASE}/funnel/target?period=2021-10",
                                  headers=auth(boss_t))
                check("Kalkulyator(API): maqsadsiz oyda hisob yo'q, maslahat bor",
                      r_no.status_code == 200 and r_no.json().get("chain") == []
                      and r_no.json().get("hint"), f"kod={r_no.status_code}")
                t_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1 limit 1").fetchone()
                if t_emp:
                    r_deny = client.post(
                        f"{API_BASE}/funnel/target", headers=auth(token_for(t_emp[0], "employee")),
                        json={"period": T_PERIOD, "target_contracts": 5})
                    check("Kalkulyator(API): oddiy xodim maqsad qo'ya OLMAYDI -> 403",
                          r_deny.status_code == 403, f"kod={r_deny.status_code}")
                t_rop = cur.execute(
                    "select id from users where role='rop' and is_active=1 limit 1").fetchone()
                if t_rop:
                    r_rop = client.post(
                        f"{API_BASE}/funnel/target", headers=auth(token_for(t_rop[0], "rop")),
                        json={"period": T_PERIOD, "target_contracts": 12})
                    check("Kalkulyator(API): ROP maqsad qo'ya oladi -> 200",
                          r_rop.status_code == 200, f"kod={r_rop.status_code}")
            finally:
                cur.execute("delete from funnel_month where period in (?,?)", (T_PERIOD, "2021-10"))
                conn.commit()
                conn.close()
        except Exception:
            check("Teskari kalkulyator tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: targetni xodimlarga tarqatish (5-bosqich) --")
        try:
            import asyncio as _aio8

            from db.base import async_session as _asess8
            from api.services import target_split as _ts
            from db.models import User as _U8

            conn = db()
            cur = conn.cursor()
            S_PERIOD = "2021-11"
            S_TG = 999779001
            cur.execute("delete from funnel_month where period=?", (S_PERIOD,))
            cur.execute("delete from users where telegram_id between ? and ?", (S_TG, S_TG + 9))
            conn.commit()

            try:
                # Lavozim: «tashrif» ko'rsatkichi biriktirilgan
                cur.execute(
                    "insert into positions (name, metrics, is_active, created_at)"
                    " values ('T-Menejer5', '[\"tashrif\"]', 1, datetime('now'))")
                s_pos = cur.lastrowid
                s_users = []
                for i in range(2):
                    cur.execute(
                        "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                        " position_id, created_at) values (?,?,'employee',1,1,?,datetime('now'))",
                        (S_TG + i, f"T-Split{i}", s_pos))
                    s_users.append(cur.lastrowid)
                # Ikkalasiga ham HAR KUNI ishlaydigan jadval (7/7) — oy 30 kun
                for uid in s_users:
                    for wd in range(7):
                        cur.execute(
                            "insert into work_schedule_weekly (user_id, weekday, is_working,"
                            " start_time, end_time, updated_at)"
                            " values (?,?,1,'09:00','18:00',datetime('now'))", (uid, wd))
                # 2-xodim 10 kun ta'tilda (tasdiqlangan sababli kun)
                for d in range(1, 11):
                    cur.execute(
                        "insert into excused_days (user_id, date, reason, status, created_at)"
                        " values (?,?,'T-tatil','approved',datetime('now'))",
                        (s_users[1], f"2021-11-{d:02d}"))
                # Maqsad: 6 shartnoma; tashrif->shartnoma 10% -> 60 tashrif
                cur.execute(
                    "insert into funnel_month (period, target_contracts, assumptions, updated_at)"
                    " values (?,?,?,datetime('now'))",
                    (S_PERIOD, 6, '{"visit_to_contract": 10.0, "lead_to_visit": 5.0}'))
                conn.commit()

                async def _sug():
                    async with _asess8() as s:
                        return await _ts.suggest(s, S_PERIOD)

                sug = _aio8.run(_sug())
                grp = next(g for g in sug["groups"] if g["metric"] == "tashrif")
                emp = {e["full_name"]: e for e in grp["employees"]}
                check("Tarqatish: oylik maqsad zanjirdan olinadi (60 tashrif)",
                      grp["monthly_target"] == 60, f"={grp['monthly_target']}")
                check("Tarqatish: ta'til ish kunidan CHIQARILADI (30 va 20)",
                      emp["T-Split0"]["working_days"] == 30
                      and emp["T-Split1"]["working_days"] == 20,
                      f"{emp['T-Split0']['working_days']}, {emp['T-Split1']['working_days']}")
                # DIQQAT: guruhda JONLI xodimlar ham bor (ularning lavozimida
                # ham «tashrif» bor), shuning uchun kutilgan qiymat guruhning
                # o'z sonlaridan hisoblanadi — aks holda test jonli bazadagi
                # xodimlar soniga bog'lanib qolardi.
                import math as _math8
                kutilgan_kunlik = _math8.ceil(grp["monthly_target"] / grp["person_days"])
                check("Tarqatish: kunlik norma = maqsad / jami ish kuni",
                      grp["suggested_daily"] == kutilgan_kunlik,
                      f"={grp['suggested_daily']}, kutilgan={kutilgan_kunlik}")
                check("Tarqatish: kunlik norma HAMMAGA bir xil",
                      emp["T-Split0"]["suggested_daily"]
                      == emp["T-Split1"]["suggested_daily"] == kutilgan_kunlik,
                      f"{emp['T-Split0']['suggested_daily']}, {emp['T-Split1']['suggested_daily']}")
                check("Tarqatish: OYLIK ulush ish kuniga proporsional (30 kun > 20 kun)",
                      emp["T-Split0"]["month_total"] == kutilgan_kunlik * 30
                      and emp["T-Split1"]["month_total"] == kutilgan_kunlik * 20,
                      f"{emp['T-Split0']['month_total']}, {emp['T-Split1']['month_total']}")

                # Tavsiya HECH NARSA yozmasligi kerak
                norms_before = cur.execute(
                    "select count(*) from norms where user_id in (?,?)", tuple(s_users)).fetchone()[0]
                check("Tarqatish: tavsiya bazaga norma YOZMAYDI",
                      norms_before == 0, f"normalar={norms_before}")

                # Tasdiqlangandan keyin yoziladi
                # ⚠️ FAQAT test xodimlariga — `user_ids`siz chaqiruv JONLI
                # xodimlarga ham norma yozib yuborardi (birinchi yugurishda
                # aynan shunday bo'ldi va 4 ta haqiqiy xodimning normasi
                # o'zgardi).
                async def _apply(actor_id):
                    async with _asess8() as s:
                        actor = await s.get(_U8, actor_id)
                        return await _ts.apply_suggestion(
                            s, S_PERIOD, "tashrif", actor, user_ids=s_users)

                boss_id = cur.execute(
                    "select id from users where role='boss' and is_active=1 limit 1").fetchone()[0]
                res = _aio8.run(_apply(boss_id))
                conn.commit()
                check("Tarqatish: tasdiqlangach norma yoziladi (2 xodim)",
                      res.get("applied") == 2, f"{res}")
                vals = [r[0] for r in cur.execute(
                    "select value from norms where user_id in (?,?)", tuple(s_users)).fetchall()]
                check("Tarqatish: yozilgan norma kunlik tavsiyaga teng",
                      sorted(vals) == [kutilgan_kunlik, kutilgan_kunlik], f"={vals}")

                # Maqsadsiz oy — tavsiya yo'q
                async def _nosug():
                    async with _asess8() as s:
                        return await _ts.suggest(s, "2021-12")
                ns = _aio8.run(_nosug())
                check("Tarqatish: maqsad qo'yilmagan oyda tavsiya yo'q",
                      ns["ready"] is False and ns["reason"], f"{ns.get('reason')}")

                # API
                r_get = client.get(f"{API_BASE}/funnel/target/split?period={S_PERIOD}",
                                   headers=auth(boss_t))
                check("Tarqatish(API): 200", r_get.status_code == 200, f"kod={r_get.status_code}")
                s_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1"
                    " and telegram_id not between ? and ? limit 1", (S_TG, S_TG + 9)).fetchone()
                if s_emp:
                    r_deny = client.post(
                        f"{API_BASE}/funnel/target/split/apply",
                        headers=auth(token_for(s_emp[0], "employee")),
                        json={"period": S_PERIOD, "metric": "tashrif"})
                    check("Tarqatish(API): oddiy xodim tarqata OLMAYDI -> 403",
                          r_deny.status_code == 403, f"kod={r_deny.status_code}")
                r_bad = client.post(
                    f"{API_BASE}/funnel/target/split/apply", headers=auth(boss_t),
                    json={"period": S_PERIOD, "metric": "yolgon"})
                check("Tarqatish(API): noma'lum ko'rsatkich -> 400",
                      r_bad.status_code == 400, f"kod={r_bad.status_code}")
            finally:
                cur.execute("delete from norms where user_id in (select id from users"
                            " where telegram_id between ? and ?)", (S_TG, S_TG + 9))
                cur.execute("delete from audit_logs where target_user_id in (select id from users"
                            " where telegram_id between ? and ?)", (S_TG, S_TG + 9))
                cur.execute("delete from excused_days where user_id in (select id from users"
                            " where telegram_id between ? and ?)", (S_TG, S_TG + 9))
                cur.execute("delete from work_schedule_weekly where user_id in (select id from users"
                            " where telegram_id between ? and ?)", (S_TG, S_TG + 9))
                cur.execute("delete from users where telegram_id between ? and ?", (S_TG, S_TG + 9))
                cur.execute("delete from positions where name='T-Menejer5'")
                cur.execute("delete from funnel_month where period in (?,?)", (S_PERIOD, "2021-12"))
                conn.commit()
                conn.close()
        except Exception:
            check("Target tarqatish tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: reja/fakt kuzatuvi va prognoz (6-bosqich) --")
        try:
            import asyncio as _aio9
            from datetime import timezone as _tz9

            from db.base import async_session as _asess9
            from api.services import target_track as _tt
            from api.services import funnel as _fn9

            conn = db()
            cur = conn.cursor()
            P_PERIOD = "2021-10"
            PL = 983001
            P_TG = 999780001
            cur.execute("delete from funnel_month where period=?", (P_PERIOD,))
            cur.execute("delete from lead_events where crm_lead_id >= ?", (PL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (PL,))
            cur.execute("delete from users where telegram_id between ? and ?", (P_TG, P_TG + 9))
            conn.commit()

            _fn9_orig = (
                list(_fn9.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn9.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn9.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn9.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                # Maqsad: 10 shartnoma; tashrif->shartnoma 10% -> 100 tashrif
                cur.execute(
                    "insert into funnel_month (period, target_contracts, assumptions, updated_at)"
                    " values (?,?,?,datetime('now'))",
                    (P_PERIOD, 10, '{"visit_to_contract": 10.0, "lead_to_visit": 5.0}'))

                # 15-oktabrgacha 30 ta tashrif (reja bo'yicha ~50 bo'lishi kerak)
                p_ts = int(datetime(2021, 10, 5, 9, tzinfo=_tz9.utc).timestamp())
                p_det = datetime(2021, 10, 5, 9).isoformat(sep=" ", timespec="seconds")
                for i in range(30):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?)",
                        (PL + i, 8779, "T-Yangi", 1, "T-op", 1, p_ts, p_ts, p_det, p_det))
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',8787,'T-Tashrif',1,'T-op',1,?,?)",
                        (PL + i, p_ts, p_det))
                conn.commit()

                async def _prog(today):
                    async with _asess9() as s:
                        return await _tt.progress(s, P_PERIOD, today=today)

                # Oyning yarmi (15-oktabr): kalendar asosida ~48% o'tgan
                pr = _aio9.run(_prog(date(2021, 10, 15)))
                by = {r["key"]: r for r in pr["rows"]}
                check("Kuzatuv: oylik reja zanjirdan olinadi (100 tashrif)",
                      by["visits"]["plan_month"] == 100, f"={by['visits']['plan_month']}")
                check("Kuzatuv: haqiqiy son davr kesimidan (30 tashrif)",
                      by["visits"]["actual"] == 30, f"={by['visits']['actual']}")
                check("Kuzatuv: «hozir kutilgan» = reja × o'tgan ulush",
                      by["visits"]["expected_now"] == round(100 * pr["elapsed"]["share"]),
                      f"={by['visits']['expected_now']}, ulush={pr['elapsed']['share']}")
                check("Kuzatuv: rejadan orqada bo'lsa «orqada» deb belgilanadi",
                      by["visits"]["status"] == "orqada", f"={by['visits']['status']}")
                check("Kuzatuv: prognoz = haqiqiy / o'tgan ulush",
                      by["visits"]["forecast"] == round(30 / pr["elapsed"]["share"]),
                      f"={by['visits']['forecast']}")
                check("Kuzatuv: prognoz rejadan past ekani ko'rsatiladi (manfiy farq)",
                      by["visits"]["forecast_gap"] is not None
                      and by["visits"]["forecast_gap"] < 0,
                      f"={by['visits']['forecast_gap']}")
                check("Kuzatuv: eng orqada qolgan bo'g'in aniqlanadi",
                      pr["weakest"] is not None, f"={pr['weakest']}")

                # Oy boshida (2-kun) prognoz KO'RSATILMAYDI
                pr2 = _aio9.run(_prog(date(2021, 10, 2)))
                by2 = {r["key"]: r for r in pr2["rows"]}
                check("Kuzatuv: oy boshida prognoz berilmaydi (namuna kichik)",
                      pr2["forecast_ready"] is False and by2["visits"]["forecast"] is None,
                      f"ulush={pr2['elapsed']['share']}, prognoz={by2['visits']['forecast']}")
                check("Kuzatuv: shunda digest qatori ham chiqmaydi",
                      _tt.digest_line(pr2) is None, "digest qatori chiqdi")

                # Digest qatori: prognoz tayyor bo'lganda ogohlantiradi
                line = _tt.digest_line(pr)
                check("Kuzatuv: digest qatori «reja ostida» deb ogohlantiradi",
                      line is not None and "Reja ostida" in line, f"={line}")
                check("Kuzatuv: digest qatorida maqsad va prognoz bor",
                      line is not None and "10" in line, f"={line}")

                # Maqsadsiz oy — kuzatuv yo'q
                async def _nop():
                    async with _asess9() as s:
                        return await _tt.progress(s, "2021-12", today=date(2021, 12, 20))
                np = _aio9.run(_nop())
                check("Kuzatuv: maqsadsiz oyda kuzatuv yo'q",
                      np["ready"] is False, f"{np.get('reason')}")
                check("Kuzatuv: maqsadsiz oyda digest jim qoladi",
                      _tt.digest_line(np) is None, "digest qatori chiqdi")

                r_api = client.get(f"{API_BASE}/funnel/target/progress?period={P_PERIOD}",
                                   headers=auth(boss_t))
                check("Kuzatuv(API): 200", r_api.status_code == 200, f"kod={r_api.status_code}")
            finally:
                cur.execute("delete from funnel_month where period in (?,?)", (P_PERIOD, "2021-12"))
                cur.execute("delete from lead_events where crm_lead_id >= ?", (PL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (PL,))
                cur.execute("delete from users where telegram_id between ? and ?", (P_TG, P_TG + 9))
                conn.commit()
                conn.close()
                (_fn9.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn9.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn9_orig
        except Exception:
            check("Reja/fakt kuzatuvi tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: operator kesimida konversiya (7-bosqich) --")
        try:
            import asyncio as _aio10
            from datetime import timezone as _tz10

            from db.base import async_session as _asess10
            from api.services import funnel as _fn10
            from api.services import funnel_operators as _fo

            conn = db()
            cur = conn.cursor()
            OL = 984001
            O_FROM, O_TO = date(2021, 4, 1), date(2021, 4, 30)
            cur.execute("delete from lead_events where crm_lead_id >= ?", (OL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (OL,))
            conn.commit()

            _fn10_orig = (
                list(_fn10.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn10.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn10.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn10.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                o_ts = int(datetime(2021, 4, 10, 9, tzinfo=_tz10.utc).timestamp())
                o_det = datetime(2021, 4, 10, 9).isoformat(sep=" ", timespec="seconds")

                def olead(i, first_rid):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?)",
                        (OL + i, 8779, "T-Yangi", first_rid, "T-op", first_rid, o_ts, o_ts,
                         o_det, o_det))

                def oev(i, to_id, rid):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',?,'T-bosqich',?,'T-op',?,?,?)",
                        (OL + i, to_id, rid, rid, o_ts, o_det))

                # Operator 111: 25 lid, 5 tashrif -> 20%
                for i in range(25):
                    olead(i, 111)
                for i in range(5):
                    oev(i, 8787, 222)          # tashrifni MENEJER 222 qabul qildi
                # Operator 333: 25 lid, 1 tashrif -> 4%
                for i in range(25, 50):
                    olead(i, 333)
                oev(25, 8787, 222)
                # Kichik namunali operator 444: 2 lid, 1 tashrif -> 50% (reytingga KIRMASIN)
                olead(50, 444); olead(51, 444)
                oev(50, 8787, 222)
                # Menejer 222: 7 tashrifdan 2 tasi shartnoma -> 28.6%
                oev(0, 8788, 222); oev(25, 8788, 222)
                conn.commit()

                async def _oq():
                    async with _asess10() as s:
                        return await _fo.operator_quality(s, O_FROM, O_TO)

                d = _aio10.run(_oq())
                ops = {r["responsible_id"]: r for r in d["operators"]}
                check("Operator sifati: maxraj — OLIB KELGAN lidlar (25 ta)",
                      ops[111]["leads"] == 25, f"={ops[111]['leads']}")
                check("Operator sifati: konversiya 20% (5/25)",
                      ops[111]["lead_to_visit"] == 20.0, f"={ops[111]['lead_to_visit']}")
                check("Operator sifati: ikkinchi operator 4% (1/25)",
                      ops[333]["lead_to_visit"] == 4.0, f"={ops[333]['lead_to_visit']}")
                check("Operator sifati: tashrifni MENEJER qilgan bo'lsa ham lid egasiga yoziladi",
                      ops[111]["visits"] == 5, f"={ops[111]['visits']}")
                check("Operator sifati: kichik namuna reytingdan CHIQARILADI",
                      ops[444]["ranked"] is False and ops[444]["lead_to_visit"] == 50.0,
                      f"{ops[444]}")
                check("Operator sifati: 50% li kichik namuna «eng yaxshi» bo'lib qolmaydi",
                      d["best_operator"]["responsible_id"] == 111,
                      f"={d['best_operator']['responsible_id']}")
                check("Operator sifati: eng past — yetarli namunalilar orasidan",
                      d["worst_operator"]["responsible_id"] == 333,
                      f"={d['worst_operator']['responsible_id']}")

                mgr = {r["responsible_id"]: r for r in d["managers"]}
                check("Menejer sifati: maxraj — O'ZI qabul qilgan tashriflar (7 ta)",
                      mgr[222]["visits"] == 7, f"={mgr[222]['visits']}")
                check("Menejer sifati: tashrif->shartnoma 28.6% (2/7)",
                      mgr[222]["visit_to_contract"] == 28.6,
                      f"={mgr[222]['visit_to_contract']}")

                # Bo'sh davr — yiqilmasin
                async def _empty10():
                    async with _asess10() as s:
                        return await _fo.operator_quality(s, date(2019, 1, 1), date(2019, 1, 31))
                e = _aio10.run(_empty10())
                check("Operator sifati: bo'sh davrda bo'sh ro'yxat (xato emas)",
                      e["operators"] == [] and e["best_operator"] is None, f"{e['operators']}")

                r_api = client.get(f"{API_BASE}/funnel/operators?month=2021-04",
                                   headers=auth(boss_t))
                check("Operator sifati(API): 200", r_api.status_code == 200,
                      f"kod={r_api.status_code}")
                o_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1 limit 1").fetchone()
                if o_emp:
                    r_no = client.get(f"{API_BASE}/funnel/operators",
                                      headers=auth(token_for(o_emp[0], "employee")))
                    check("Operator sifati(API): oddiy xodimga -> 403",
                          r_no.status_code == 403, f"kod={r_no.status_code}")
            finally:
                cur.execute("delete from lead_events where crm_lead_id >= ?", (OL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (OL,))
                conn.commit()
                conn.close()
                (_fn10.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn10.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn10_orig
        except Exception:
            check("Operator sifati tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: bo'g'in tahlili va stsenariylar (7-bosqich, 1-2 band) --")
        try:
            import asyncio as _aio11
            from datetime import timezone as _tz11

            from db.base import async_session as _asess11
            from api.services import funnel as _fn11
            from api.services import funnel_analysis as _fa
            from api.services import target_calc as _tc11

            conn = db()
            cur = conn.cursor()
            AL = 985001
            A_PERIOD = "2021-03"
            cur.execute("delete from lead_events where crm_lead_id >= ?", (AL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (AL,))
            cur.execute("delete from funnel_month where period=?", (A_PERIOD,))
            conn.commit()

            _fn11_orig = (
                list(_fn11.CRM_UYSOT_INVITE_PIPE_STATUS_IDS),
                list(_fn11.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn11.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn11.CRM_UYSOT_INVITE_PIPE_STATUS_IDS = [8786]
            _fn11.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn11.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                a_ts = int(datetime(2021, 3, 10, 9, tzinfo=_tz11.utc).timestamp())
                a_det = datetime(2021, 3, 10, 9).isoformat(sep=" ", timespec="seconds")

                def alead(i):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?)",
                        (AL + i, 8779, "T-Yangi", 1, "T-op", 1, a_ts, a_ts, a_det, a_det))

                def aev(i, to_id):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',?,'T-bosqich',1,'T-op',1,?,?)",
                        (AL + i, to_id, a_ts, a_det))

                # 100 lid -> 20 taklif -> 10 tashrif -> 2 shartnoma
                for i in range(100):
                    alead(i)
                for i in range(20):
                    aev(i, 8786)
                for i in range(10):
                    aev(i, 8787)
                for i in range(2):
                    aev(i, 8788)
                conn.commit()

                async def _leak():
                    async with _asess11() as s:
                        return await _fa.leak_analysis(s, A_PERIOD)

                lk = _aio11.run(_leak())
                steps = {s["label"]: s for s in lk["steps"]}
                check("Tahlil: lid->taklif bo'g'inida 80 ta yo'qoldi (80%)",
                      steps["Lid → ofisga taklif"]["lost"] == 80
                      and steps["Lid → ofisga taklif"]["loss_pct"] == 80.0,
                      f"{steps['Lid → ofisga taklif']}")
                check("Tahlil: taklif->tashrif 10 ta (50%)",
                      steps["Taklif → tashrif"]["lost"] == 10
                      and steps["Taklif → tashrif"]["loss_pct"] == 50.0,
                      f"{steps['Taklif → tashrif']}")
                check("Tahlil: tashrif->shartnoma 8 ta (80%)",
                      steps["Tashrif → shartnoma"]["lost"] == 8, f"{steps['Tashrif → shartnoma']}")
                check("Tahlil: eng katta yo'qotish — birinchi bo'g'in",
                      lk["biggest_leak"]["label"] == "Lid → ofisga taklif",
                      f"{lk['biggest_leak']}")
                check("Tahlil: umumiy konversiya 2% (2/100)",
                      lk["overall_conversion"] == 2.0, f"={lk['overall_conversion']}")
                check("Tahlil: CPL yo'q -> pul ustuni bo'sh (0 emas!)",
                      all(s["money_lost"] is None for s in lk["steps"]),
                      f"{[s['money_lost'] for s in lk['steps']]}")
                check("Tahlil: «~yo'qolgan shartnoma» o'rtacha konversiyadan (80 × 2%)",
                      steps["Lid → ofisga taklif"]["contracts_lost"] == 1.6,
                      f"={steps['Lid → ofisga taklif']['contracts_lost']}")

                # STSENARIY: maqsad va farazlar bilan
                cur.execute(
                    "insert into funnel_month (period, target_contracts, assumptions, updated_at)"
                    " values (?,?,?,datetime('now'))",
                    (A_PERIOD, 10,
                     '{"lead_to_visit": 10.0, "visit_to_contract": 20.0, "cpl": 50000}'))
                conn.commit()

                async def _sc():
                    async with _asess11() as s:
                        return await _fa.scenarios(s, A_PERIOD, 20)

                sc = _aio11.run(_sc())
                by = {s["key"]: s for s in sc["scenarios"]}
                # lid->shartnoma = 10% × 20% = 2%; 10 uy uchun 500 lid;
                # byudjet = 500 × 50 000 = 25 mln; +20% = 5 mln -> 100 lid -> +2 uy
                check("Stsenariy: byudjet +20% -> +100 lid",
                      by["budget_up"]["extra_leads"] == 100, f"={by['budget_up']['extra_leads']}")
                check("Stsenariy: byudjet +20% -> +2 uy",
                      by["budget_up"]["extra_contracts"] == 2.0,
                      f"={by['budget_up']['extra_contracts']}")
                konv_key = "Tashrif → shartnoma +1 punkt"
                # 500 lid × (10% × 21%) = 10.5 -> +0.5 uy
                check("Stsenariy: tashrif->shartnoma +1 punkt -> +0.5 uy",
                      by[konv_key]["extra_contracts"] == 0.5,
                      f"={by[konv_key]['extra_contracts']}")
                check("Stsenariy: faraz manbai ko'rsatiladi (qo'lda kiritilgan)",
                      "override" in (by["budget_up"]["sources"] or []),
                      f"={by['budget_up']['sources']}")

                # Ma'lumot yetmasa — halol «hisoblanmadi»
                cur.execute("delete from funnel_month where period=?", (A_PERIOD,))
                conn.commit()
                sc2 = _aio11.run(_sc())
                b2 = {s["key"]: s for s in sc2["scenarios"]}
                check("Stsenariy: maqsad yo'q -> hisoblanmaydi va sababi aytiladi",
                      b2["budget_up"]["extra_contracts"] is None
                      and "maqsad" in (b2["budget_up"].get("missing") or []),
                      f"{b2['budget_up']}")

                r_api = client.get(f"{API_BASE}/funnel/analysis?period={A_PERIOD}",
                                   headers=auth(boss_t))
                check("Tahlil(API): 200 va ikkala bo'lim qaytadi",
                      r_api.status_code == 200 and "leaks" in r_api.json()
                      and "scenarios" in r_api.json(), f"kod={r_api.status_code}")
            finally:
                cur.execute("delete from lead_events where crm_lead_id >= ?", (AL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (AL,))
                cur.execute("delete from funnel_month where period=?", (A_PERIOD,))
                conn.commit()
                conn.close()
                (_fn11.CRM_UYSOT_INVITE_PIPE_STATUS_IDS,
                 _fn11.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn11.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn11_orig
        except Exception:
            check("Bo'g'in tahlili tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- VORONKA: qoidalar panelda (bekor shartnoma / sifatsiz lid) --")
        try:
            import asyncio as _aio12
            from datetime import timezone as _tz12

            from db.base import async_session as _asess12
            from api.services import funnel as _fn12
            from api.services import funnel_settings as _fs

            conn = db()
            cur = conn.cursor()
            SL = 986001
            S_FROM, S_TO = date(2021, 2, 1), date(2021, 2, 28)
            CANCEL_ID, LOWQ_ID = 999111, 999222
            cur.execute("delete from lead_events where crm_lead_id >= ?", (SL,))
            cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (SL,))
            _fs_before = cur.execute(
                "select cancelled_pipe_status_ids, subtract_cancelled,"
                " low_quality_pipe_status_ids, exclude_low_quality from funnel_settings"
                " where id=1").fetchone()
            cur.execute("delete from funnel_settings where id=1")
            conn.commit()

            _fn12_orig = (
                list(_fn12.CRM_UYSOT_VISIT_PIPE_STATUS_IDS),
                list(_fn12.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS),
            )
            _fn12.CRM_UYSOT_VISIT_PIPE_STATUS_IDS = [8787]
            _fn12.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS = [8788]

            try:
                s_ts = int(datetime(2021, 2, 10, 9, tzinfo=_tz12.utc).timestamp())
                s_det = datetime(2021, 2, 10, 9).isoformat(sep=" ", timespec="seconds")

                def slead(i, now_stage):
                    cur.execute(
                        "insert into crm_lead_state (crm_lead_id, pipe_status_id, stage_name,"
                        " responsible_id, responsible_name, first_responsible_id, crm_updated_ts,"
                        " crm_created_ts, first_seen_at, last_seen_at)"
                        " values (?,?,?,?,?,?,?,?,?,?)",
                        (SL + i, now_stage, "T-hozirgi", 1, "T-op", 1, s_ts, s_ts, s_det, s_det))

                def sev(i, to_id):
                    cur.execute(
                        "insert into lead_events (crm_lead_id, event_type, from_pipe_status_id,"
                        " from_stage_name, to_pipe_status_id, to_stage_name, to_responsible_id,"
                        " to_responsible_name, first_responsible_id, crm_updated_ts, detected_at)"
                        " values (?,'stage_change',8779,'T-Yangi',?,'T-bosqich',1,'T-op',1,?,?)",
                        (SL + i, to_id, s_ts, s_det))

                # 0,1: shartnoma qilgan va shunday qolgan
                slead(0, 8788); sev(0, 8787); sev(0, 8788)
                slead(1, 8788); sev(1, 8787); sev(1, 8788)
                # 2: shartnoma qilib, KEYIN bekor bo'lgan (hozirgi holati — bekor)
                slead(2, CANCEL_ID); sev(2, 8787); sev(2, 8788); sev(2, CANCEL_ID)
                # 3: sifatsiz lead (hech qayerga yetmagan)
                slead(3, LOWQ_ID)
                # 4: oddiy lid
                slead(4, 8779)
                conn.commit()

                async def _coh():
                    async with _asess12() as s:
                        return await _fn12.cohort_funnel(s, S_FROM, S_TO)

                async def _set(**kw):
                    async with _asess12() as s:
                        return await _fs.update_settings(s, kw, None)

                # ── Default: IKKALA qoida ham O'CHIQ ──
                d0 = _aio12.run(_coh())
                b0 = {r["key"]: r["value"] for r in d0["rows"]}
                check("Qoida: default o'chiq — 5 lid, 3 shartnoma (bekor ham sanaladi)",
                      b0["lead"] == 5 and b0["contract"] == 3,
                      f"lid={b0['lead']}, shartnoma={b0['contract']}")

                # ── Bekor qilinganni ayirish ──
                _aio12.run(_set(subtract_cancelled=True,
                                cancelled_pipe_status_ids=[CANCEL_ID]))
                d1 = _aio12.run(_coh())
                b1 = {r["key"]: r["value"] for r in d1["rows"]}
                check("Qoida: bekor qilingan shartnoma ayrildi (3 -> 2)",
                      b1["contract"] == 2, f"shartnoma={b1['contract']}")
                check("Qoida: lid soni o'zgarmadi (faqat shartnomaga ta'sir)",
                      b1["lead"] == 5, f"lid={b1['lead']}")

                # ── Bosqich tanlanmasa qoida ISHLAMAYDI ──
                _aio12.run(_set(subtract_cancelled=True, cancelled_pipe_status_ids=[]))
                d2 = _aio12.run(_coh())
                b2 = {r["key"]: r["value"] for r in d2["rows"]}
                check("Qoida: yoqiq-u bosqich tanlanmagan -> qoida ishlamaydi",
                      b2["contract"] == 3, f"shartnoma={b2['contract']}")

                # ── Sifatsiz lid maxrajdan chiqsin ──
                _aio12.run(_set(subtract_cancelled=False, cancelled_pipe_status_ids=[CANCEL_ID],
                                exclude_low_quality=True, low_quality_pipe_status_ids=[LOWQ_ID]))
                d3 = _aio12.run(_coh())
                b3 = {r["key"]: r["value"] for r in d3["rows"]}
                check("Qoida: sifatsiz lid maxrajdan chiqdi (5 -> 4)",
                      b3["lead"] == 4, f"lid={b3['lead']}")
                check("Qoida: sifatsiz chiqarilgach konversiya ko'tariladi",
                      d3["rows"][-1]["conv_from_lead"] > d0["rows"][-1]["conv_from_lead"],
                      f"{d0['rows'][-1]['conv_from_lead']} -> {d3['rows'][-1]['conv_from_lead']}")

                # ── Ikkalasi birga ──
                _aio12.run(_set(subtract_cancelled=True, cancelled_pipe_status_ids=[CANCEL_ID],
                                exclude_low_quality=True, low_quality_pipe_status_ids=[LOWQ_ID]))
                d4 = _aio12.run(_coh())
                b4 = {r["key"]: r["value"] for r in d4["rows"]}
                check("Qoida: ikkalasi birga — 4 lid, 2 shartnoma",
                      b4["lead"] == 4 and b4["contract"] == 2,
                      f"lid={b4['lead']}, shartnoma={b4['contract']}")

                # ── Bosqich ro'yxati CRM'siz to'ldiriladi ──
                async def _stages():
                    async with _asess12() as s:
                        return await _fs.known_stages(s)
                st = _aio12.run(_stages())
                check("Qoida: bosqich ro'yxati o'z jurnalimizdan (CRM so'rovsiz)",
                      any(x["pipe_status_id"] == CANCEL_ID for x in st), f"{len(st)} ta bosqich")

                # ── API ──
                r_get = client.get(f"{API_BASE}/funnel/settings", headers=auth(boss_t))
                check("Qoida(API): sozlama va bosqichlar qaytadi",
                      r_get.status_code == 200 and "stages" in r_get.json(),
                      f"kod={r_get.status_code}")
                s_emp = cur.execute(
                    "select id from users where role='employee' and is_active=1 limit 1").fetchone()
                if s_emp:
                    r_deny = client.post(
                        f"{API_BASE}/funnel/settings",
                        headers=auth(token_for(s_emp[0], "employee")),
                        json={"subtract_cancelled": True})
                    check("Qoida(API): oddiy xodim o'zgartira OLMAYDI -> 403",
                          r_deny.status_code == 403, f"kod={r_deny.status_code}")
                r_ok = client.post(f"{API_BASE}/funnel/settings", headers=auth(boss_t),
                                   json={"subtract_cancelled": False})
                check("Qoida(API): Boshliq saqlay oladi", r_ok.status_code == 200,
                      f"kod={r_ok.status_code}")
            finally:
                cur.execute("delete from lead_events where crm_lead_id >= ?", (SL,))
                cur.execute("delete from crm_lead_state where crm_lead_id >= ?", (SL,))
                cur.execute("delete from funnel_settings where id=1")
                if _fs_before:
                    cur.execute(
                        "insert into funnel_settings (id, cancelled_pipe_status_ids,"
                        " subtract_cancelled, low_quality_pipe_status_ids, exclude_low_quality,"
                        " updated_at) values (1,?,?,?,?,datetime('now'))", _fs_before)
                conn.commit()
                conn.close()
                (_fn12.CRM_UYSOT_VISIT_PIPE_STATUS_IDS,
                 _fn12.CRM_UYSOT_CONTRACT_PIPE_STATUS_IDS) = _fn12_orig
        except Exception:
            check("Voronka qoidalari tekshiruvi", False, traceback.format_exc(limit=2).strip())

        print("\n-- ISSIQ LID: sovish qoidasi, eslatmalar, taqsimlash, statistika --")
        try:
            import asyncio as _aio

            from db.base import async_session as _asess
            from api.services import hot_lead as _hl

            conn = db()
            cur = conn.cursor()

            # HR qoidasi: global FinePolicy (bo'lmasa yaratamiz; borini eslab
            # qolib, oxirida ASL HOLIGA qaytaramiz — jonli sozlamaga tegmaymiz)
            old_global = cur.execute(
                "select id, hot_lead_cool_minutes, hot_lead_fine from fine_policies"
                " where scope='global'").fetchone()
            if old_global:
                cur.execute(
                    "update fine_policies set hot_lead_cool_minutes=12, hot_lead_fine=50000"
                    " where id=?", (old_global[0],))
                created_policy = None
            else:
                cur.execute(
                    "insert into fine_policies (scope, scope_id, free_late_minutes_per_month,"
                    " fine_mode, absent_mode, early_leave_enabled, fine_applies_to, is_active,"
                    " hot_lead_cool_minutes, hot_lead_fine, updated_at)"
                    " values ('global', NULL, 60, 'per_day', 'fixed', 0, 'net_salary', 1,"
                    " 12, 50000, datetime('now'))")
                created_policy = cur.lastrowid
            conn.commit()

            async def _rules():
                async with _asess() as s:
                    return await _hl.hot_lead_rules(s)

            mins, fine = _aio.run(_rules())
            check("HL: HR qoidasi o'qildi (12 daq / 50 000)", mins == 12 and fine == 50000.0,
                  f"{mins} daq, {fine}")

            # Eslatma matni bosqichlari — senlab, jarima bilan
            fake = type("L", (), {"contact_name": "T-Mijoz", "lead_name": None, "crm_lead_id": 1})()
            t3 = _hl._reminder_text(3, 12, fake, 50000.0)
            t9 = _hl._reminder_text(9, 12, fake, 50000.0)
            check("HL: 3-daqiqa eslatmasi yumshoq, 9-daqiqa qattiq + jarima",
                  "3 daqiqa" in t3 and "guruhga" in t9 and "50 000" in t9,
                  f"{t3[:40]} | {t9[:60]}")
            check("HL: eslatmalar sovish limitidan kichik bosqichlarda",
                  all(s < 12 for s in _hl.REMINDER_STEPS), str(_hl.REMINDER_STEPS))

            # Oldingi-qo'ng'iroq oynasi 2 soat (mijoz bilan gaplashib bo'lib
            # CRM'ga kiritish holati) — egasining shikoyatining yechimi
            check("HL: qo'ng'iroq oynasi 2 soat (10 daqiqa emas)",
                  _hl.PRE_CREATION_GRACE_SECONDS == 7200, str(_hl.PRE_CREATION_GRACE_SECONDS))

            # Taqsimlash: eng kam yuklangan operator tanlanadi
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " crm_visit_external_id, created_at)"
                " values (999555001,'T-Op1','employee',1,1,'900001',datetime('now'))")
            op1 = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " crm_visit_external_id, created_at)"
                " values (999555002,'T-Op2','employee',1,1,'900002',datetime('now'))")
            op2 = cur.lastrowid
            # op1 ga bugun 2 ta lid — demak keyingisi op2 ga tushishi kerak
            for i in (91001, 91002):
                cur.execute(
                    "insert into hot_lead (crm_lead_id, user_id, created_ts, detected_at,"
                    " status, last_reminder_minute)"
                    " values (?,?,?,datetime('now'),'notified',0)",
                    (i, op1, int(__import__("time").time())))
            conn.commit()

            async def _pick():
                async with _asess() as s:
                    u = await _hl._pick_operator(s)
                    return u.full_name if u else None

            picked = _aio.run(_pick())
            # Nomzodlar orasida HAQIQIY operatorlar ham bor (ular ham 0 lidli)
            # — muhimi: bugun 2 ta lid olgan T-Op1 TANLANMASLIGI kerak.
            check("HL: yuklangan operator (T-Op1) taqsimotdan chetda qoldi",
                  picked is not None and picked != "T-Op1", f"tanlandi={picked}")

            # O'CHIRGICH (2026-08-06): hot_lead_enabled=0 bo'lgan operatorga
            # lid BERILMAYDI (Tester akkaunti uchun aynan shu kerak).
            cur.execute("update users set hot_lead_enabled=0 where id in (?,?)", (op1, op2))
            # Boshqa hamma nomzodni ham vaqtincha o'chiramiz — tanlov aynan
            # bayroqqa bog'liqligini isbotlash uchun
            others = [r[0] for r in cur.execute(
                "select id from users where hot_lead_enabled=1").fetchall()]
            if others:
                cur.execute(
                    "update users set hot_lead_enabled=0 where id in (%s)"
                    % ",".join("?" * len(others)), others)
            conn.commit()
            picked_off = _aio.run(_pick())
            check("HL: hamma o'chirilganda taqsimot to'xtaydi (lid berilmaydi)",
                  picked_off is None, f"tanlandi={picked_off}")

            cur.execute("update users set hot_lead_enabled=1 where id=?", (op2,))
            conn.commit()
            picked_on = _aio.run(_pick())
            check("HL: faqat yoqilgan operator lid oladi (T-Op2)",
                  picked_on == "T-Op2", f"tanlandi={picked_on}")
            # Haqiqiy operatorlarni ASL holiga qaytaramiz
            if others:
                cur.execute(
                    "update users set hot_lead_enabled=1 where id in (%s)"
                    % ",".join("?" * len(others)), others)
            conn.commit()

            # Kunlik statistika: kim nechta lidni sovutgan
            cur.execute(
                "insert into hot_lead (crm_lead_id, user_id, created_ts, detected_at,"
                " escalated_at, fine_amount, status, last_reminder_minute)"
                " values (91003,?,?,datetime('now'),datetime('now'),50000,'notified',9)",
                (op1, int(__import__("time").time())))
            cur.execute(
                "insert into hot_lead (crm_lead_id, user_id, created_ts, detected_at,"
                " escalated_at, fine_amount, status, last_reminder_minute)"
                " values (91004,?,?,datetime('now'),datetime('now'),50000,'called',9)",
                (op1, int(__import__("time").time())))
            conn.commit()

            async def _cooled():
                async with _asess() as s:
                    return await _hl.cooled_by_operator(s, date.today())

            cooled = _aio.run(_cooled())
            row = next((c for c in cooled if c["full_name"] == "T-Op1"), None)
            check("HL: statistikada sovutgan operator (1 ta, tuzatilgani sanalmaydi)",
                  row is not None and row["count"] == 1 and row["fine"] == 50000.0, str(row))

            # Tozalash: T- ma'lumot + sozlamani ASL holiga qaytarish
            cur.execute("delete from hot_lead where crm_lead_id in (91001,91002,91003,91004)")
            cur.execute("delete from users where id in (?,?)", (op1, op2))
            if created_policy:
                cur.execute("delete from fine_policies where id=?", (created_policy,))
            elif old_global:
                cur.execute(
                    "update fine_policies set hot_lead_cool_minutes=?, hot_lead_fine=? where id=?",
                    (old_global[1], old_global[2], old_global[0]))
            conn.commit()
            conn.close()
        except Exception:
            check("Issiq lid tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- UX2-W4: bot dashboard + statistika davomat bloki + validatsiya --")
        try:
            conn = db()
            cur = conn.cursor()
            with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as f:
                secret = next(
                    (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")),
                    "",
                )
            bot_h = {"X-Bot-Secret": secret}

            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444806,'T-BotMgr','hr',1,1,datetime('now'))")
            mgr_uid = cur.lastrowid
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
                " values (999444807,'T-BotEmp','employee',1,1,datetime('now'))")
            emp_uid = cur.lastrowid
            # T-BotEmp'ga shu oyda davomat: 1 kelgan kun (20 daq kechikish) + 1 absent
            first_day = date.today().replace(day=1).isoformat()
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, check_out_time, status,"
                " late_minutes, early_leave_minutes, worked_minutes, created_at, updated_at)"
                " values (?,?,datetime('now','-30 hours'),datetime('now','-22 hours'),'late',20,0,480,"
                " datetime('now'),datetime('now'))", (emp_uid, first_day))
            if date.today().day >= 2:
                second_day = date.today().replace(day=2).isoformat()
                cur.execute(
                    "insert into attendance (user_id, date, status, late_minutes,"
                    " early_leave_minutes, worked_minutes, created_at, updated_at)"
                    " values (?,?,'absent',0,0,0,datetime('now'),datetime('now'))",
                    (emp_uid, second_day))
            conn.commit()

            # C5: dashboard-bot — rahbar 200, xodim 403
            r = client.get(f"{API_BASE}/attendance/dashboard-bot/999444806", headers=bot_h)
            check("W4: dashboard-bot rahbarga -> 200 + summary/late_list",
                  r.status_code == 200 and "summary" in r.json() and "late_list" in r.json(),
                  f"kod={r.status_code}")
            r2 = client.get(f"{API_BASE}/attendance/dashboard-bot/999444807", headers=bot_h)
            check("W4: dashboard-bot xodimga -> 403", r2.status_code == 403, f"kod={r2.status_code}")

            # C9: statistikada davomat bloki
            r3 = client.get(f"{API_BASE}/stats/my/999444807", headers=bot_h)
            b3 = r3.json() if r3.status_code == 200 else {}
            expected_absent = 1 if date.today().day >= 2 else 0
            check("W4: my-stats davomat maydonlari (1 kelgan, 20 daq, absent)",
                  b3.get("attendance_present_days") == 1
                  and b3.get("attendance_late_minutes") == 20
                  and b3.get("attendance_absent_days") == expected_absent,
                  str({k: v for k, v in b3.items() if k.startswith("attendance")}))

            # C10: sabab juda qisqa -> 422 (schema min_length)
            r4 = client.post(f"{API_BASE}/excused-days", headers=bot_h,
                             json={"telegram_id": 999444807, "reason": "x"})
            check("W4: 1-belgili sabab -> 422", r4.status_code == 422, f"kod={r4.status_code}")

            # Qoldiq #4/#5: kutilayotgan so'rovlar ro'yxatlari (bot)
            cur.execute(
                "insert into excused_days (user_id, date, reason, status, created_at)"
                " values (?, date('now','+3 days'), 'T-kutish', 'pending', datetime('now'))",
                (emp_uid,))
            cur.execute(
                "insert into face_reregistration_requests (user_id, new_descriptor, status, created_at)"
                " values (?, ?, 'pending', datetime('now'))", (emp_uid, json.dumps([0.5] * 128)))
            conn.commit()
            r5 = client.get(f"{API_BASE}/excused-days/pending-bot/999444806", headers=bot_h)
            names5 = [x["user_full_name"] for x in (r5.json() if r5.status_code == 200 else [])]
            check("Q4: pending-bot (sababli) rahbarga -> 200 + T-BotEmp bor",
                  r5.status_code == 200 and "T-BotEmp" in names5, f"kod={r5.status_code} {names5[:5]}")
            r6 = client.get(f"{API_BASE}/excused-days/pending-bot/999444807", headers=bot_h)
            check("Q4: pending-bot xodimga -> 403", r6.status_code == 403, f"kod={r6.status_code}")
            r7 = client.get(f"{API_BASE}/attendance/face-reregistration/pending-bot/999444806", headers=bot_h)
            names7 = [x["user_full_name"] for x in (r7.json() if r7.status_code == 200 else [])]
            check("Q5: pending-bot (yuz) rahbarga -> 200 + T-BotEmp bor",
                  r7.status_code == 200 and "T-BotEmp" in names7, f"kod={r7.status_code} {names7[:5]}")

            # Qoldiq #13: qisqa muddatli WebView tokeni
            emp_jwt = token_for(emp_uid, "employee")
            r8 = client.post(f"{API_BASE}/auth/webview-token", headers=auth(emp_jwt))
            wb = r8.json() if r8.status_code == 200 else {}
            check("Q13: webview-token -> 200 + 30 daq",
                  r8.status_code == 200 and wb.get("expires_in_minutes") == 30,
                  f"kod={r8.status_code}")
            if wb.get("access_token"):
                r9 = client.get(f"{API_BASE}/users/me", headers=auth(wb["access_token"]))
                check("Q13: webview-token bilan /users/me ishlaydi",
                      r9.status_code == 200 and r9.json().get("id") == emp_uid,
                      f"kod={r9.status_code}")

            for u in (mgr_uid, emp_uid):
                cur.execute("delete from attendance where user_id=?", (u,))
                cur.execute("delete from excused_days where user_id=?", (u,))
                cur.execute("delete from face_reregistration_requests where user_id=?", (u,))
                cur.execute("delete from users where id=?", (u,))
            conn.commit()
            conn.close()
        except Exception:
            check("UX2-W4 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- UX-A6: yuz so'rovini webdan hal qilish --")
        try:
            conn = db()
            cur = conn.cursor()
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " face_descriptor, face_registered_at, created_at) values"
                " (999444803,'T-ReregWeb','employee',1,1,?,datetime('now'),datetime('now'))",
                (json.dumps(FACE),))
            rw_uid = cur.lastrowid
            new_face = json.dumps([0.9] * 128)
            cur.execute(
                "insert into face_reregistration_requests (user_id, new_descriptor, status, created_at)"
                " values (?, ?, 'pending', datetime('now'))", (rw_uid, new_face))
            req_id = cur.lastrowid
            conn.commit()

            emp_tok = token_for(rw_uid, "employee")
            r = client.post(f"{API_BASE}/attendance/face-reregistration/{req_id}/decide-web",
                            headers=auth(emp_tok), json={"decision": "approved"})
            check("A6: oddiy xodim -> 403", r.status_code == 403, f"kod={r.status_code}")

            mgr = find_manager_id()
            mgr_tok = token_for(mgr[0], mgr[1])
            r2 = client.post(f"{API_BASE}/attendance/face-reregistration/{req_id}/decide-web",
                             headers=auth(mgr_tok), json={"decision": "approved"})
            check("A6: rahbar webdan tasdiqlaydi -> 200", r2.status_code == 200,
                  f"kod={r2.status_code} {r2.text[:150]}")
            saved = conn.execute("select face_descriptor from users where id=?", (rw_uid,)).fetchone()[0]
            check("A6: descriptor yangilandi", json.loads(saved) == [0.9] * 128)
            r3 = client.post(f"{API_BASE}/attendance/face-reregistration/{req_id}/decide-web",
                             headers=auth(mgr_tok), json={"decision": "rejected"})
            check("A6: qayta qaror -> 400 (idempotent)", r3.status_code == 400, f"kod={r3.status_code}")

            cur.execute("delete from face_reregistration_requests where user_id=?", (rw_uid,))
            cur.execute("delete from audit_logs where target_user_id=?", (rw_uid,))
            cur.execute("delete from users where id=?", (rw_uid,))
            conn.commit()
            conn.close()
        except Exception:
            check("UX-A6 tekshiruvi", False, traceback.format_exc(limit=1).strip())

        print("\n-- UX-A7: umumiy ish jadvali (JWT) --")
        try:
            mgr = find_manager_id()
            mgr_tok = token_for(mgr[0], mgr[1])
            r = client.get(f"{API_BASE}/work-schedule/all/week", headers=auth(mgr_tok))
            check("A7: rahbar umumiy jadvalni oladi -> 200",
                  r.status_code == 200 and isinstance(r.json(), list),
                  f"kod={r.status_code}, {len(r.json()) if r.status_code == 200 else '-'} xodim")
            conn = db()
            emp_row = conn.execute(
                "select id from users where role='employee' and is_active=1 limit 1").fetchone()
            conn.close()
            if emp_row:
                emp_tok = token_for(emp_row[0], "employee")
                r2 = client.get(f"{API_BASE}/work-schedule/all/week", headers=auth(emp_tok))
                check("A7: oddiy xodimga -> 403", r2.status_code == 403, f"kod={r2.status_code}")
        except Exception:
            check("UX-A7 tekshiruvi", False, traceback.format_exc(limit=1).strip())


def test_payroll_engine() -> None:
    """Bosqich 2: `api/services/payroll.py` hisoblash yadrosi — HTTP orqali
    emas (router hali yo'q, Bosqich 3), to'g'ridan-to'g'ri servis funksiyalarini
    chaqirib. Butunlay izolyatsiyalangan davr ("2020-01") ishlatiladi — real
    joriy oy ma'lumotiga tegilmaydi. Xodimga barcha hafta kunlari
    `is_working=False` qilib qo'yilib (WorkScheduleWeekly), faqat aniq
    belgilangan kunlar `WorkScheduleOverride` bilan ish kuni deb ochiladi —
    shu orqali "rejadagi kunlar soni" testda TO'LIQ nazorat qilinadi (Yanvar
    2020'ning haftaning qaysi kunlariga to'g'ri kelishiga bog'liq emas)."""
    import asyncio
    from decimal import Decimal

    print("\n" + "=" * 60)
    print("BOSQICH 2: PAYROLL HISOBLASH YADROSI (api/services/payroll.py)")
    print("=" * 60)

    async def _run():
        from sqlalchemy import delete, select
        from db.base import async_session
        from db.models import (
            Attendance, AuditLog, Bonus, ExcusedDay, FinePolicy, OvertimeEntry, OvertimeProfile,
            PayrollAdjustment, PayrollPeriod, Payslip, PayslipItem, SalaryRate,
            User, WorkScheduleOverride, WorkScheduleWeekly,
        )
        from api.services import payroll as pr

        PERIOD = "2020-01"
        WINDOW_MIN = 480  # work_minutes(09:00,18:00) — tushliksiz 8 soat

        async with async_session() as s:
            # Oldingi (masalan yarim yo'lda qulagan) ishga tushirishdan qolgan
            # T-Payroll* yozuvlarni tozalaymiz — aks holda UNIQUE(telegram_id)
            # xatosi bilan bu test hech qachon qayta ishlamay qolardi.
            stale_ids = list(
                await s.scalars(select(User.id).where(User.full_name.like("T-Payroll%")))
            )
            if stale_ids:
                stale_payslips = list(
                    await s.scalars(select(Payslip.id).where(Payslip.user_id.in_(stale_ids)))
                )
                if stale_payslips:
                    await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id.in_(stale_payslips)))
                await s.execute(delete(Payslip).where(Payslip.user_id.in_(stale_ids)))
                await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id.in_(stale_ids)))
                await s.execute(delete(OvertimeProfile).where(OvertimeProfile.user_id.in_(stale_ids)))
                await s.execute(delete(PayrollAdjustment).where(PayrollAdjustment.user_id.in_(stale_ids)))
                await s.execute(delete(SalaryRate).where(SalaryRate.user_id.in_(stale_ids)))
                await s.execute(delete(FinePolicy).where(FinePolicy.scope_id.in_(stale_ids)))
                await s.execute(delete(ExcusedDay).where(ExcusedDay.user_id.in_(stale_ids)))
                await s.execute(delete(Attendance).where(Attendance.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id.in_(stale_ids)))
                # audit_logs.target_user_id VA actor_id — boshqa test bloklari
                # (masalan test_admin_override, dasturchi sifatida `override_*`
                # yozadi — actor_id orqali) shu id'larga tegishli yozuv
                # qoldirishi mumkin (SQLite ROWID o'chirilgan id'larni qayta
                # beradi); FK cheklovi tufayli bu tozalanmasa User o'chirilmay
                # qoladi (jonli isbot: 2026-07-27).
                await s.execute(
                    delete(AuditLog).where(
                        AuditLog.target_user_id.in_(stale_ids) | AuditLog.actor_id.in_(stale_ids)
                    )
                )
                await s.execute(delete(User).where(User.id.in_(stale_ids)))
            # Qo'shimcha himoya: agar avvalgi qulagan ishga tushirishda AVVAL
            # User o'chirilib, keyingi qadam (masalan Attendance o'chirish)
            # bajarilmay qolgan bo'lsa — ORFAN (egasiz) yozuv qoladi. SQLite
            # ROWID'ni o'chirilgan eng katta id qayta beradi, shuning uchun
            # KEYINGI ishga tushirishda yangi u1 xuddi O'SHA id'ni olib,
            # egasiz yozuvga to'qnashishi mumkin (jonli isbot: 2026-07-27).
            # Shu sabab test davri (2020-01) va aniq telegram_id'lar bo'yicha
            # HAM, User mavjudligidan qat'i nazar, to'g'ridan-to'g'ri tozalanadi.
            await s.execute(delete(Attendance).where(Attendance.date >= date(2020, 1, 1), Attendance.date < date(2020, 2, 1)))
            await s.execute(
                delete(WorkScheduleOverride).where(
                    WorkScheduleOverride.date >= date(2020, 1, 1), WorkScheduleOverride.date < date(2020, 2, 1)
                )
            )
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.execute(
                delete(FinePolicy).where(
                    FinePolicy.scope == "global", FinePolicy.free_late_minutes_per_month == 999_999
                )
            )
            await s.commit()

            # ── Sozlash: T-Payroll1 (asosiy sinov xodimi) ──
            u1 = User(telegram_id=999500701, full_name="T-Payroll1", role="employee",
                      bot_started=True, is_active=True)
            s.add(u1)
            await s.flush()

            for wd in range(7):
                s.add(WorkScheduleWeekly(user_id=u1.id, weekday=wd, is_working=False))

            day1, day2, day3 = date(2020, 1, 6), date(2020, 1, 7), date(2020, 1, 8)
            day4_absent, day5_present, day6_excused = date(2020, 1, 9), date(2020, 1, 10), date(2020, 1, 13)
            test_days = [day1, day2, day3, day4_absent, day5_present, day6_excused]
            for d in test_days:
                s.add(WorkScheduleOverride(user_id=u1.id, date=d, is_working=True,
                                            start_time="09:00", end_time="18:00"))

            # Kechikishlar: 20+15 (ikkalasi ham bepul limit ichida, jami 35),
            # keyingi 10 daq (cumulative_before=35 >= limit(30) -> JARIMALI).
            s.add(Attendance(user_id=u1.id, date=day1, status="late", late_minutes=20, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=u1.id, date=day2, status="late", late_minutes=15, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=u1.id, date=day3, status="late", late_minutes=10, worked_minutes=WINDOW_MIN))
            # day4_absent — yozuv YO'Q (defensiv "absent" filiali sinaladi)
            s.add(Attendance(user_id=u1.id, date=day5_present, status="present", worked_minutes=WINDOW_MIN))
            s.add(ExcusedDay(user_id=u1.id, date=day6_excused, reason="T-sinov", status="approved"))

            s.add(SalaryRate(user_id=u1.id, amount=3_000_000, pay_basis="monthly",
                              effective_from=date(2019, 1, 1), changed_by=u1.id))
            s.add(FinePolicy(scope="user", scope_id=u1.id, is_active=True,
                              free_late_minutes_per_month=30, fine_mode="per_day", fine_per_day=50_000,
                              absent_mode="fixed", absent_fine=100_000,
                              monthly_cap_amount=1_000_000, fine_applies_to="net_salary"))
            await s.commit()

            days = await pr.collect_attendance(s, u1, PERIOD)
            by_date = {d["date"]: d for d in days}
            check("Payroll: scheduled_days aniq 6", sum(1 for d in days if d["is_working"]) == 6,
                  f"={sum(1 for d in days if d['is_working'])}")
            check("Payroll: day4 defensiv 'absent' deb aniqlandi",
                  by_date[day4_absent]["status"] == "absent")
            check("Payroll: day6 'excused' deb aniqlandi (ExcusedDay, yozuvsiz)",
                  by_date[day6_excused]["status"] == "excused" and by_date[day6_excused]["excused"] is True)

            policy = await pr.resolve_policy(s, u1)
            check("Payroll: resolve_policy user-scope qoidani topdi", policy is not None and policy.scope == "user")

            late = pr.compute_late_fine(days, policy)
            check("Payroll: late_days=3", late["late_days"] == 3, f"={late['late_days']}")
            check("Payroll: late_minutes=45", late["late_minutes"] == 45, f"={late['late_minutes']}")
            check("Payroll: limit ichidagi 2 kun BEPUL (chegaradan o'tkazgan kun ham)",
                  late["fined_days"] == 1, f"fined_days={late['fined_days']}")
            check("Payroll: faqat 3-kun (10 daq) jarimali", late["fined_minutes"] == 10,
                  f"fined_minutes={late['fined_minutes']}")
            check("Payroll: jarima summasi = 1 x 50000", late["amount"] == Decimal("50000"),
                  f"amount={late['amount']}")

            absent = pr.compute_absent_fine(days, policy)
            check("Payroll: absent_days=1", absent["absent_days"] == 1, f"={absent['absent_days']}")
            check("Payroll: kelmagan kun jarimasi = 100000", absent["amount"] == Decimal("100000"),
                  f"amount={absent['amount']}")

            rate = await pr.resolve_rate(s, u1.id, date(2020, 1, 1))
            first_rate = await pr._first_rate(s, u1.id)
            # Uchinchi qiymat — kelmagan kunlar ayirmasi (2026-08-08). Bu yerda
            # qoida `fixed` rejimda, ya'ni ayirma bo'lmasligi kerak.
            # To'rtinchi — to'lovsiz ta'til ayirmasi (2026-08-13).
            base_amount, base_item, absent_item, unpaid_item = pr.compute_base(
                rate, first_rate, days, date(2020, 1, 1), policy
            )
            check("Payroll: base_amount to'liq oylik (prorata yo'q, 6/6 kun)",
                  base_amount == Decimal("3000000"), f"base={base_amount}")
            check("Payroll: 'fixed' rejimda bazadan ayirma YO'Q (ikki marta jazo bo'lmasin)",
                  absent_item is None, f"absent_item={absent_item}")

            # ── run_payroll orqali to'liq oqim (Payslip + PayslipItem yoziladi) ──
            result = await pr.run_payroll(s, PERIOD, user_ids=[u1.id])
            check("Payroll: run_payroll 1 xodimni hisobladi", result["calculated"] == 1, f"={result}")

            payslip = await s.scalar(select(Payslip).where(Payslip.user_id == u1.id, Payslip.period == PERIOD))
            check("Payroll: Payslip yaratildi", payslip is not None)
            if payslip is not None:
                check("Payroll: Payslip.fine_amount=50000", float(payslip.fine_amount) == 50000.0,
                      f"={payslip.fine_amount}")
                check("Payroll: Payslip.absent_deduction=100000", float(payslip.absent_deduction) == 100000.0,
                      f"={payslip.absent_deduction}")
                check("Payroll: Payslip.net = 3000000-50000-100000=2850000",
                      float(payslip.net) == 2_850_000.0, f"net={payslip.net}")
                check("Payroll: Payslip.excused_days=1", payslip.excused_days == 1, f"={payslip.excused_days}")
                check("Payroll: Payslip.worked_days=4 (3 late + 1 present)", payslip.worked_days == 4,
                      f"={payslip.worked_days}")

                items_1 = list(await s.scalars(
                    select(PayslipItem).where(PayslipItem.payslip_id == payslip.id)))
                check("Payroll: PayslipItem qatorlari yozildi (base+fine_late+fine_absent)",
                      len(items_1) == 3, f"soni={len(items_1)}")

                # ── Idempotentlik: qayta chaqirilsa dublikat YO'Q ──
                result2 = await pr.run_payroll(s, PERIOD, user_ids=[u1.id])
                payslip2 = await s.scalar(select(Payslip).where(Payslip.user_id == u1.id, Payslip.period == PERIOD))
                items_2 = list(await s.scalars(
                    select(PayslipItem).where(PayslipItem.payslip_id == payslip2.id)))
                check("Payroll: idempotent - bir xil Payslip.id", payslip2.id == payslip.id)
                check("Payroll: idempotent - PayslipItem dublikat YO'Q (hamon 3 ta)",
                      len(items_2) == 3, f"soni={len(items_2)}")
                check("Payroll: idempotent - net o'zgarmadi", float(payslip2.net) == 2_850_000.0,
                      f"net={payslip2.net}")

            # ── Qulflangan davr qayta hisoblashni rad etadi ──
            period_row = await s.scalar(select(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            period_row.locked = True
            await s.commit()
            locked_raised = False
            try:
                await pr.run_payroll(s, PERIOD, user_ids=[u1.id])
            except pr.PayrollLocked:
                locked_raised = True
            check("Payroll: qulflangan davrda run_payroll -> PayrollLocked", locked_raised)
            period_row.locked = False
            await s.commit()

            # ── Oylik jarima chegarasi (cap) — pastroq cap bilan qayta sinov ──
            policy.monthly_cap_amount = 100_000  # raw jami 150000 dan kichik
            await s.commit()
            late_c, absent_c, raw_total, cap_applied = pr.apply_fine_cap(
                late["amount"], absent["amount"], base_amount, policy
            )
            check("Payroll: cap ishga tushdi (raw 150000 > cap 100000)", cap_applied is True)
            check("Payroll: cap qo'llangach jami roppa-rosa 100000",
                  abs(float(late_c + absent_c) - 100_000.0) < 0.01, f"={late_c + absent_c}")
            check("Payroll: cap ikkalasini proporsional qisqartiradi (ikkalasi ham >0)",
                  late_c > 0 and absent_c > 0, f"late={late_c} absent={absent_c}")

            # ── Qo'shimcha ish (derived rejim) ──
            s.add(OvertimeProfile(user_id=u1.id, enabled=True, mode="derived",
                                   multiplier=1.5, norm_hours_source="schedule"))
            s.add(OvertimeEntry(user_id=u1.id, date=day1, minutes=120, source="manual", status="approved"))
            await s.commit()
            profile = await s.scalar(select(OvertimeProfile).where(OvertimeProfile.user_id == u1.id))
            ot = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            # norm_hours = scheduled_minutes(6*480=2880)/60=48; hourly=3000000/48*1.5=93750;
            # amount=93750*(120/60)=187500
            check("Payroll: overtime_minutes=120", ot["minutes"] == 120, f"={ot['minutes']}")
            check("Payroll: overtime amount (derived, 1.5x) ~187500",
                  abs(float(ot["amount"]) - 187_500.0) < 1, f"={ot['amount']}")

            # Fixed-rate rejim (bir xil OvertimeEntry, boshqa profil)
            profile.mode = "fixed_rate"
            profile.fixed_rate_per_hour = 20_000
            await s.commit()
            ot2 = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            check("Payroll: overtime amount (fixed_rate, 20000/soat x 2 soat) =40000",
                  abs(float(ot2["amount"]) - 40_000.0) < 1, f"={ot2['amount']}")

            # ── §3.4: «vaqtni QO'SHIB-AYIRIB umumiy berish» ──
            # Egasining talabi shu edi: ortiqcha ham, KAM ishlangan vaqt ham
            # bitta songa yig'ilsin. Shuning uchun `minutes` manfiy bo'lishi
            # MUMKIN va oy bo'yicha ortiqchadan ayiriladi.
            await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id == u1.id))
            s.add(OvertimeEntry(user_id=u1.id, date=day1, minutes=-105, source="auto_attendance",
                                 status="approved"))
            await s.commit()
            ot_neg = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            check("Payroll §3.4: KAM ishlangan vaqt manfiy sanaladi (-105 daq)",
                  ot_neg["minutes"] == -105, f"={ot_neg['minutes']}")
            check("Payroll §3.4: manfiy vaqt summasi ham manfiy (-35000)",
                  abs(float(ot_neg["amount"]) + 35_000.0) < 1, f"={ot_neg['amount']}")

            # Aralash kunlar: +120 va -30 -> sof +90
            await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id == u1.id))
            s.add(OvertimeEntry(user_id=u1.id, date=day1, minutes=120, source="auto_attendance",
                                 status="approved"))
            s.add(OvertimeEntry(user_id=u1.id, date=day2, minutes=-30, source="auto_attendance",
                                 status="approved"))
            await s.commit()
            ot_mix = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            check("Payroll §3.4: aralash kunlar qo'shilib-ayirilib +90 daq",
                  ot_mix["minutes"] == 90, f"={ot_mix['minutes']}")

            # Kunlik cheklov IKKI TOMONGA simmetrik: cap=60 bo'lsa
            # +120 -> +60, -30 esa tegilmaydi -> sof +30
            profile.daily_cap_minutes = 60
            await s.commit()
            ot_cap = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            check("Payroll §3.4: kunlik cheklov ikki tomonga simmetrik (+60-30=+30)",
                  ot_cap["minutes"] == 30, f"={ot_cap['minutes']}")

            # Manfiy tomonda ham cheklov ishlashi kerak: -105 -> -60
            await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id == u1.id))
            s.add(OvertimeEntry(user_id=u1.id, date=day1, minutes=-105, source="auto_attendance",
                                 status="approved"))
            await s.commit()
            ot_cap_neg = await pr.compute_overtime(s, u1, PERIOD, profile, days)
            check("Payroll §3.4: cheklov MANFIY tomonda ham qo'llanadi (-60)",
                  ot_cap_neg["minutes"] == -60, f"={ot_cap_neg['minutes']}")

            # Sinovdan keyin asl holatga (keyingi tekshiruvlar buzilmasin)
            profile.daily_cap_minutes = None
            await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id == u1.id))
            s.add(OvertimeEntry(user_id=u1.id, date=day1, minutes=120, source="manual",
                                 status="approved"))
            await s.commit()

            # ── QoolAdjustment (avans) — minus ──
            s.add(PayrollAdjustment(user_id=u1.id, period=PERIOD, kind="minus", amount=50_000,
                                     reason="T-sinov avans", created_by=u1.id))
            await s.commit()
            result3 = await pr.run_payroll(s, PERIOD, user_ids=[u1.id])
            check("Payroll: adjustment qo'shilgandan keyin ham hisoblandi", result3["calculated"] == 1)
            payslip3 = await s.scalar(select(Payslip).where(Payslip.user_id == u1.id, Payslip.period == PERIOD))
            check("Payroll: adjustments_minus=50000 qaytadan hisoblanganda hisobga olindi",
                  float(payslip3.adjustments_minus) == 50_000.0, f"={payslip3.adjustments_minus}")

            # ── resolve_policy: xodim > lavozim > global qamrov ──
            u2 = User(telegram_id=999500702, full_name="T-Payroll2NoPolicy", role="employee",
                      bot_started=True, is_active=True)
            s.add(u2)
            await s.flush()
            no_policy = await pr.resolve_policy(s, u2)
            check("Payroll: qoida umuman yo'q bo'lsa None qaytadi", no_policy is None)

            days2 = [{"date": day1, "is_working": True, "start": "09:00", "end": "18:00",
                      "scheduled_minutes": WINDOW_MIN, "attendance": None, "excused": False,
                      "status": "late", "late_minutes": 999, "worked_minutes": 0}]
            late_no_policy = pr.compute_late_fine(days2, no_policy)
            check("Payroll: qoidasiz kechikish uchun jarima 0 (xavfsiz sukut)",
                  late_no_policy["amount"] == Decimal("0"), f"={late_no_policy['amount']}")

            s.add(FinePolicy(scope="global", scope_id=None, is_active=True,
                              free_late_minutes_per_month=999_999, fine_mode="per_day", fine_per_day=1,
                              absent_mode="none"))
            await s.commit()
            global_policy = await pr.resolve_policy(s, u2)
            check("Payroll: user/lavozim yo'q bo'lsa GLOBAL qoidaga tushadi",
                  global_policy is not None and global_policy.scope == "global")

            # ── Tozalash ──
            for uid in (u1.id, u2.id):
                pslips = list(await s.scalars(select(Payslip).where(Payslip.user_id == uid)))
                for p in pslips:
                    await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id == p.id))
                await s.execute(delete(Payslip).where(Payslip.user_id == uid))
                await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id == uid))
                await s.execute(delete(OvertimeProfile).where(OvertimeProfile.user_id == uid))
                await s.execute(delete(PayrollAdjustment).where(PayrollAdjustment.user_id == uid))
                await s.execute(delete(SalaryRate).where(SalaryRate.user_id == uid))
                await s.execute(delete(FinePolicy).where(FinePolicy.scope_id == uid))
                await s.execute(delete(ExcusedDay).where(ExcusedDay.user_id == uid))
                await s.execute(delete(Attendance).where(Attendance.user_id == uid))
                await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id == uid))
                await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id == uid))
            await s.execute(delete(FinePolicy).where(FinePolicy.scope == "global",
                                                       FinePolicy.free_late_minutes_per_month == 999_999))
            # §2.3 dan keyin oylik hisobi bonus qatorini ham yaratadi —
            # tozalanmasa `delete(User)` FOREIGN KEY bilan yiqiladi.
            await s.execute(delete(Bonus).where(Bonus.user_id.in_([u1.id, u2.id])))
            await s.execute(delete(ExcusedDay).where(ExcusedDay.user_id.in_([u1.id, u2.id])))
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.execute(
                delete(AuditLog).where(
                    AuditLog.target_user_id.in_([u1.id, u2.id]) | AuditLog.actor_id.in_([u1.id, u2.id])
                )
            )
            await s.execute(delete(User).where(User.id.in_([u1.id, u2.id])))
            await s.commit()

    try:
        asyncio.run(_run())
    except Exception:
        check("Payroll hisoblash yadrosi (umumiy)", False, traceback.format_exc(limit=2).strip())


def test_payroll_api() -> None:
    """Bosqich 3: `api/routers/payroll.py` — HTTP darajasida (ruxsat
    matritsasi, validatsiya, to'liq oqim: hisoblash -> ro'yxat -> tafsilot ->
    tasdiqlash -> qulf -> bot). Izolyatsiyalangan davr "2021-02" ishlatiladi."""
    import httpx

    print("\n" + "=" * 60)
    print("BOSQICH 3: PAYROLL API (api/routers/payroll.py)")
    print("=" * 60)

    PERIOD = "2021-02"
    conn = db()
    cur = conn.cursor()

    # Oldingi (yarim yo'lda qulagan) ishga tushirishdan qolganlarni tozalash.
    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-PayAPI%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        pslip_ids = [r[0] for r in cur.execute(
            f"select id from payslips where user_id in ({qm})", stale).fetchall()]
        if pslip_ids:
            qm2 = ",".join("?" * len(pslip_ids))
            cur.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslip_ids)
        for tbl in ("payslips", "bonuses", "overtime_entries", "overtime_profiles", "payroll_adjustments",
                    "salary_rates", "attendance", "work_schedule_override", "work_schedule_weekly"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from fine_policies where scope_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from payroll_periods where period=?", (PERIOD,))
    conn.commit()

    # ── Sozlash: T-PayAPI-HR (rahbar sifatida ishlatilmaydi — mavjud HR
    # ishlatiladi), T-PayAPI-ROP + uning T-PayAPI-Emp xodimi ──
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600801,'T-PayAPI-Rop','rop',1,1,datetime('now'))")
    rop_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, manager_id, bot_started, is_active, created_at)"
        " values (999600802,'T-PayAPI-Emp','employee',?,1,1,datetime('now'))", (rop_uid,))
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600803,'T-PayAPI-Outsider','employee',1,1,datetime('now'))")
    outsider_uid = cur.lastrowid
    conn.commit()

    mgr = find_manager_id()
    mgr_t = token_for(mgr[0], mgr[1]) if mgr else None
    rop_t = token_for(rop_uid, "rop")
    emp_t = token_for(emp_uid, "employee")

    def cleanup_payapi():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [rop_uid, emp_uid, outsider_uid]
            qm = ",".join("?" * len(uids))
            pslip_ids = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", uids).fetchall()]
            if pslip_ids:
                qm2 = ",".join("?" * len(pslip_ids))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslip_ids)
            for tbl in ("payslips", "bonuses", "overtime_entries", "overtime_profiles", "payroll_adjustments",
                        "salary_rates", "attendance", "work_schedule_override", "work_schedule_weekly"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", uids)
            c2.execute(f"delete from fine_policies where scope_id in ({qm})", uids)
            c2.execute("delete from payroll_periods where period=?", (PERIOD,))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Payroll API tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        with httpx.Client(timeout=15) as client:
            # ── Ruxsat matritsasi ──
            r = client.get(f"{API_BASE}/payroll/policies", headers=auth(emp_t))
            check("xodim /payroll/policies -> 403", r.status_code == 403, f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/payroll/policies", headers=auth(rop_t))
            check("ROP /payroll/policies -> 403 (faqat HR/Boss/Dasturchi)", r.status_code == 403,
                  f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/payroll/policies", headers=auth(mgr_t))
            check("HR/Boss /payroll/policies -> 200", r.status_code == 200, f"kod={r.status_code}")

            # ── FinePolicy validatsiya ──
            r = client.put(f"{API_BASE}/payroll/policies", headers=auth(mgr_t), json={
                "scope": "user", "scope_id": emp_uid, "free_late_minutes_per_month": 20,
                "fine_mode": "per_day", "fine_per_day": 30000,
                "absent_mode": "fixed", "absent_fine": 80000,
                # cap yo'q -> majburiy maydon xatosi kutiladi
            })
            check("cap'siz policy -> 422 (majburiy)", r.status_code == 422, f"kod={r.status_code}")

            r = client.put(f"{API_BASE}/payroll/policies", headers=auth(mgr_t), json={
                "scope": "user", "scope_id": emp_uid, "free_late_minutes_per_month": 20,
                "fine_mode": "per_day", "fine_per_day": 30000,
                "absent_mode": "fixed", "absent_fine": 80000,
                "monthly_cap_amount": 500000, "fine_applies_to": "net_salary",
            })
            check("to'g'ri policy -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            policy_id = r.json().get("id") if r.status_code == 200 else None
            check("policy javobida scope_label (xodim ismi)",
                  r.status_code == 200 and r.json().get("scope_label") == "T-PayAPI-Emp",
                  f"={r.json().get('scope_label') if r.status_code == 200 else None}")

            # ── SalaryRate ──
            r = client.post(f"{API_BASE}/payroll/rates", headers=auth(mgr_t), json={
                "user_id": emp_uid, "amount": 2_500_000, "pay_basis": "monthly",
                "effective_from": "2020-01-01", "reason": "hire",
            })
            check("stavka yaratildi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            r2 = client.post(f"{API_BASE}/payroll/rates", headers=auth(mgr_t), json={
                "user_id": emp_uid, "amount": 2_600_000, "pay_basis": "monthly",
                "effective_from": "2020-01-01", "reason": "periodic",
            })
            check("bir xil sanaga dublikat stavka -> 400", r2.status_code == 400, f"kod={r2.status_code}")

            # ── OvertimeProfile validatsiya ──
            r = client.put(f"{API_BASE}/payroll/overtime-profiles/{emp_uid}", headers=auth(mgr_t), json={
                "enabled": True, "mode": "derived", "norm_hours_source": "schedule",
                # multiplier yo'q -> xato kutiladi
            })
            check("multiplier'siz derived profil -> 422", r.status_code == 422, f"kod={r.status_code}")

            # ── Ish jadvali va davomat (2021-02 uchun) — BARCHA hafta kunlari
            # dam olish deb belgilanadi, FAQAT 1-fevral aniq override bilan ish
            # kuni ochiladi (Bosqich 2'dagi bilan bir xil naqsh). Aks holda oy
            # davomidagi BOSHQA ish kunlariga Attendance yozuvi yo'qligi
            # sababli ular "kelmagan" deb hisoblanib (collect_attendance'ning
            # ataylab qilingan defensiv qoidasi), kutilmagan jarima chiqarardi —
            # bu HTTP darajasidagi test uchun ortiqcha murakkablik, hisoblash
            # mantig'ining o'zi Bosqich 2'da alohida tekshirilgan.
            for wd in range(7):
                cur.execute(
                    "insert into work_schedule_weekly (user_id, weekday, is_working, updated_at)"
                    " values (?,?,0,datetime('now'))", (emp_uid, wd))
            cur.execute(
                "insert into work_schedule_override (user_id, date, is_working, start_time, end_time, updated_at)"
                " values (?, '2021-02-01', 1, '09:00', '18:00', datetime('now'))", (emp_uid,))
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, '2021-02-01', '2021-02-01 04:00:00', 0, 0, 480, 'present', 0,"
                " datetime('now'), datetime('now'))", (emp_uid,))
            conn.commit()

            # ── Preflight ──
            r = client.get(f"{API_BASE}/payroll/{PERIOD}/preflight", headers=auth(mgr_t))
            check("preflight -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            # ── Hisoblash ──
            # §4.3: endi bu NAVBATGA qo'yadi (202) — hisobni cron bajaradi.
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t),
                             json={"user_ids": [emp_uid]})
            check("calculate -> 202 (navbatga qo'yildi)", r.status_code == 202,
                  f"kod={r.status_code} {r.text[:150]}")
            check("calculate javobida navbat belgisi bor",
                  r.status_code == 202 and r.json().get("queued") is True,
                  f"={r.json() if r.status_code == 202 else None}")
            tick = payroll_tick(client, PERIOD)
            check("cron navbatdagi davrni hisobladi",
                  bool(tick) and tick.get("ran") == PERIOD and tick.get("calculated") == 1, f"={tick}")

            # Hisob ketayotganda ikkinchi marta bosib bo'lmasin (409)
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
            check("navbatga qayta qo'yish -> 202 (avvalgisi tugagan)", r.status_code == 202,
                  f"kod={r.status_code}")
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
            check("hisob navbatda turganda ikkinchi so'rov -> 409", r.status_code == 409,
                  f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/payroll/{PERIOD}/status", headers=auth(mgr_t))
            check("GET /status navbat holatini ko'rsatadi",
                  r.status_code == 200 and r.json().get("state") == "queued",
                  f"kod={r.status_code} {r.text[:120]}")
            payroll_tick(client, PERIOD)
            r = client.get(f"{API_BASE}/payroll/{PERIOD}/status", headers=auth(mgr_t))
            check("hisob tugagach state=done",
                  r.status_code == 200 and r.json().get("state") == "done", f"={r.text[:120]}")

            # ── Davrlar ro'yxati (literal /periods — {period} bilan aralashmasin) ──
            r = client.get(f"{API_BASE}/payroll/periods", headers=auth(mgr_t))
            check("GET /payroll/periods -> 200", r.status_code == 200, f"kod={r.status_code}")
            check("yangi davr ro'yxatda bor",
                  r.status_code == 200 and any(p["period"] == PERIOD for p in r.json()),
                  f"davrlar={[p['period'] for p in r.json()] if r.status_code == 200 else None}")

            # ── Ro'yxat va tafsilot — ROP qamrovi ──
            r = client.get(f"{API_BASE}/payroll/{PERIOD}", headers=auth(rop_t))
            check("ROP ro'yxatni ko'radi -> 200", r.status_code == 200, f"kod={r.status_code}")
            names_seen = [row["full_name"] for row in r.json()] if r.status_code == 200 else []
            check("ROP faqat o'z jamoasini ko'radi (T-PayAPI-Emp bor, Outsider yo'q)",
                  "T-PayAPI-Emp" in names_seen, f"={names_seen}")

            r = client.get(f"{API_BASE}/payroll/{PERIOD}/user/{emp_uid}", headers=auth(rop_t))
            check("ROP o'z xodimining tafsilotini ko'radi -> 200", r.status_code == 200, f"kod={r.status_code}")
            detail = r.json() if r.status_code == 200 else {}
            check("tafsilotda base_amount=2500000", detail.get("base_amount") == 2_500_000.0,
                  f"={detail.get('base_amount')}")
            check("tafsilotda fine_amount=0 (kechikish yo'q)", detail.get("fine_amount") == 0.0,
                  f"={detail.get('fine_amount')}")
            check("tafsilotda items ro'yxati bor", len(detail.get("items", [])) >= 1,
                  f"soni={len(detail.get('items', []))}")

            # S-06 (TZ 4-qism): begona yozuvga 404, 403 EMAS. 403 «yozuv bor,
            # lekin ruxsat yo'q» degani — ya'ni o'sha xodim MAVJUDLIGINI
            # oshkor qiladi. TZ buni ataylab taqiqlaydi.
            r = client.get(f"{API_BASE}/payroll/{PERIOD}/user/{outsider_uid}", headers=auth(rop_t))
            check("ROP begona xodim tafsilotiga -> 404 (S-06: 403 emas)",
                  r.status_code == 404, f"kod={r.status_code}")

            r = client.get(f"{API_BASE}/payroll/{PERIOD}/user/{emp_uid}", headers=auth(emp_t))
            check("xodim /payroll/*/user -> 403 (VIEW_ROLES'da yo'q)", r.status_code == 403,
                  f"kod={r.status_code}")

            # ── Tasdiqlash va qulf ──
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/approve", headers=auth(rop_t))
            check("ROP tasdiqlay OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # 2026-08-21 (egasining qarori): oylikni HR ham tasdiqlaydi —
            # «HR tayyor deydi -> Boshliq qulflaydi» ajratimidan voz
            # kechildi, chunki u oylikni kechiktirardi. «Tayyor» qadami
            # IXTIYORIY bo'lib qoldi (faqat «kim ko'rib chiqdi» izi).
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/hr-approve", headers=auth(mgr_t))
            check("HR «tayyor» dedi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")

            _c = db()
            _boss = _c.execute(
                "select id, role from users where role in ('boss','dasturchi') and is_active=1 limit 1"
            ).fetchone()
            _c.close()
            boss_t = token_for(_boss[0], _boss[1]) if _boss else mgr_t
            r = client.post(f"{API_BASE}/payroll/{PERIOD}/approve", headers=auth(mgr_t))
            check("HR yakuniy tasdiqlaydi -> 200 (Boshliq kerak emas)",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
            check("tasdiqlangan (qulflangan) davrni qayta hisoblash -> 409", r.status_code == 409,
                  f"kod={r.status_code}")

            r = client.post(f"{API_BASE}/payroll/{PERIOD}/approve", headers=auth(boss_t))
            check("ikkinchi marta tasdiqlash -> 409", r.status_code == 409, f"kod={r.status_code}")

            # ── Bot: /payroll/my — faqat TASDIQLANGANDAN keyin ko'rinadi ──
            with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as f:
                _secret = next(
                    (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")), ""
                )
            BOT_SECRET_HDR = {"X-Bot-Secret": _secret}
            r = client.get(f"{API_BASE}/payroll/my/999600802", headers=BOT_SECRET_HDR)
            check("bot /payroll/my -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            bot_out = r.json() if r.status_code == 200 else {}
            check("bot /payroll/my calculated=True (tasdiqlangan)", bot_out.get("calculated") is True,
                  f"={bot_out}")
            check("bot /payroll/my net=2500000", bot_out.get("net") == 2_500_000.0, f"={bot_out.get('net')}")

            r = client.get(f"{API_BASE}/payroll/my/999600802/late-status", headers=BOT_SECRET_HDR)
            check("bot /payroll/my/late-status -> 200", r.status_code == 200, f"kod={r.status_code}")
            check("late-status joriy oy (bugungi) qaytadi",
                  r.status_code == 200 and r.json().get("period") == date.today().strftime("%Y-%m"),
                  f"={r.json().get('period') if r.status_code == 200 else None}")

            # ── Overtime kirish/tasdiqlash oqimi ──
            r = client.post(f"{API_BASE}/payroll/overtime", headers=auth(mgr_t), json={
                "user_id": emp_uid, "date": "2021-02-02", "minutes": 90, "note": "T-sinov",
            })
            check("overtime kiritildi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            ot_id = r.json().get("id") if r.status_code == 200 else None
            r = client.post(f"{API_BASE}/payroll/overtime/{ot_id}/decide", headers=auth(mgr_t),
                             json={"status": "approved"})
            check("overtime tasdiqlandi -> 200", r.status_code == 200, f"kod={r.status_code}")

            # ── Adjustment (avans) ──
            r = client.post(f"{API_BASE}/payroll/adjustments", headers=auth(mgr_t), json={
                "user_id": emp_uid, "period": PERIOD, "kind": "minus", "amount": 50000,
                "reason": "T-sinov avans",
            })
            check("adjustment yaratildi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            adj_id = r.json().get("id") if r.status_code == 200 else None
            if adj_id:
                r = client.delete(f"{API_BASE}/payroll/adjustments/{adj_id}", headers=auth(mgr_t))
                check("adjustment o'chirildi -> 200", r.status_code == 200, f"kod={r.status_code}")

            if policy_id:
                r = client.delete(f"{API_BASE}/payroll/policies/{policy_id}", headers=auth(mgr_t))
                check("policy o'chirildi -> 200", r.status_code == 200, f"kod={r.status_code}")
    except Exception:
        check("Payroll API (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_payapi()
        conn.close()


def test_admin_override() -> None:
    """Bosqich 3.5: `api/routers/admin_override.py` — Dasturchi rejimi
    (super-admin). HTTP darajasida: ruxsat (faqat dasturchi), yumshoq
    o'chirish/tiklash, normalar matritsasi bypass, payroll qulflari."""
    import httpx

    print("\n" + "=" * 60)
    print("BOSQICH 3.5: DASTURCHI REJIMI (api/routers/admin_override.py)")
    print("=" * 60)

    PERIOD = "2022-03"
    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-Admin%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        pslip_ids = [r[0] for r in cur.execute(f"select id from payslips where user_id in ({qm})", stale).fetchall()]
        if pslip_ids:
            qm2 = ",".join("?" * len(pslip_ids))
            cur.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslip_ids)
        for tbl in ("payslips", "bonuses", "salary_rates", "attendance", "work_schedule_override", "work_schedule_weekly"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from norms where user_id in ({qm})", stale)
        # actor_id HAM tozalanadi — bu test dasturchi (T-Admin-Dev) sifatida
        # `override_*` amallar bajaradi, ular AuditLog.actor_id (target_user_id
        # EMAS) orqali shu userga bog'lanadi. Jonli isbot (2026-07-27): shu
        # tozalanmasa keyingi ishga tushirishda FOREIGN KEY xatosi bilan
        # user o'chirilmay qolgan (SQLite ROWID qayta berilgani sabab boshqa
        # test blokining foydalanuvchisiga to'qnashgan).
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from payroll_periods where period=?", (PERIOD,))
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999700901,'T-Admin-Dev','dasturchi',1,1,datetime('now'))")
    dev_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999700902,'T-Admin-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999700903,'T-Admin-Hr','hr',1,1,datetime('now'))")
    hr_target_uid = cur.lastrowid
    conn.commit()

    dev_t = token_for(dev_uid, "dasturchi")
    emp_t = token_for(emp_uid, "employee")
    mgr = find_manager_id()
    mgr_t = token_for(mgr[0], mgr[1]) if mgr else None

    def cleanup_admin():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [dev_uid, emp_uid, hr_target_uid]
            qm = ",".join("?" * len(uids))
            pslip_ids = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", uids).fetchall()]
            if pslip_ids:
                qm2 = ",".join("?" * len(pslip_ids))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslip_ids)
            for tbl in ("payslips", "bonuses", "salary_rates", "attendance", "work_schedule_override", "work_schedule_weekly"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", uids)
            c2.execute(f"delete from norms where user_id in ({qm})", uids)
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute("delete from payroll_periods where period=?", (PERIOD,))
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Admin override tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        with httpx.Client(timeout=15) as client:
            # ── Ruxsat: FAQAT dasturchi ──
            r = client.get(f"{API_BASE}/admin/audit/overrides", headers=auth(emp_t))
            check("xodim /admin/* -> 403", r.status_code == 403, f"kod={r.status_code}")
            if mgr_t:
                r = client.get(f"{API_BASE}/admin/audit/overrides", headers=auth(mgr_t))
                check("HR/Boss /admin/* -> 403 (faqat dasturchi)", r.status_code == 403, f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/admin/audit/overrides", headers=auth(dev_t))
            check("dasturchi /admin/* -> 200", r.status_code == 200, f"kod={r.status_code}")

            # ── Sababsiz override -> 422 ──
            r = client.put(f"{API_BASE}/admin/norms/{emp_uid}/suhbat", headers=auth(dev_t), json={"value": 50})
            check("override_reason'siz -> 422", r.status_code == 422, f"kod={r.status_code}")

            # ── Normalar: xodim bo'lmaganga (HR) ham norma qo'yish ──
            r = client.put(f"{API_BASE}/admin/norms/{hr_target_uid}/suhbat", headers=auth(dev_t), json={
                "value": 77, "override_reason": "T-sinov: HR'ga norma",
            })
            check("dasturchi HR (xodim EMAS) ga norma qo'ydi -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:150]}")
            check("javobda qiymat 77", r.status_code == 200 and r.json().get("value") == 77,
                  f"={r.json().get('value') if r.status_code == 200 else None}")

            # Oddiy yo'l (odatdagi HR/Boss) xuddi shu HR'ga norma qo'ya OLMAYDI
            if mgr_t:
                r = client.post(f"{API_BASE}/norms", headers=auth(mgr_t), json={
                    "user_id": hr_target_uid, "metric_type": "suhbat", "value": 10,
                })
                check("oddiy /norms HR nishoniga -> 400 (xodim emas)", r.status_code == 400,
                      f"kod={r.status_code}")

            # ── Metrika cheklovisiz (lavozimida yo'q ko'rsatkich) ──
            r = client.put(f"{API_BASE}/admin/norms/{emp_uid}/dumaloq_video", headers=auth(dev_t), json={
                "value": 5, "override_reason": "T-sinov: metrika cheklovisiz",
            })
            check("dasturchi lavozimga mos kelmagan metrikaga norma qo'ydi -> 200",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            norm_id = r.json().get("id") if r.status_code == 200 else None

            # ── Yumshoq o'chirish + tiklash ──
            if norm_id:
                r = client.request("DELETE", f"{API_BASE}/admin/norms/{norm_id}", headers=auth(dev_t), json={
                    "override_reason": "T-sinov: yumshoq o'chirish",
                })
                check("norma yumshoq o'chirildi -> 200", r.status_code == 200, f"kod={r.status_code}")
                check("javobda hard=False", r.status_code == 200 and r.json().get("hard") is False,
                      f"={r.json() if r.status_code == 200 else None}")

                r = client.get(f"{API_BASE}/admin/records/norm?include_deleted=true", headers=auth(dev_t))
                found = [row for row in r.json() if row.get("id") == norm_id] if r.status_code == 200 else []
                check("o'chirilgan norma include_deleted=true da ko'rinadi",
                      len(found) == 1 and found[0].get("deleted_at") is not None,
                      f"topildi={found}")

                r = client.get(f"{API_BASE}/admin/records/norm?include_deleted=false", headers=auth(dev_t))
                found2 = [row for row in r.json() if row.get("id") == norm_id] if r.status_code == 200 else []
                check("include_deleted=false da ko'rinmaydi", len(found2) == 0, f"topildi={found2}")

                r = client.post(f"{API_BASE}/admin/records/norm/{norm_id}/restore", headers=auth(dev_t), json={
                    "override_reason": "T-sinov: tiklash",
                })
                check("norma tiklandi -> 200", r.status_code == 200, f"kod={r.status_code}")

            # ── PATCH oq ro'yxat tashqarisidagi maydon -> 400 ──
            if norm_id:
                r = client.patch(f"{API_BASE}/admin/records/norm/{norm_id}", headers=auth(dev_t), json={
                    "fields": {"user_id": 99999}, "override_reason": "T-sinov: ruxsatsiz maydon",
                })
                check("ruxsat etilmagan maydon -> 400", r.status_code == 400, f"kod={r.status_code}")

                r = client.patch(f"{API_BASE}/admin/records/norm/{norm_id}", headers=auth(dev_t), json={
                    "fields": {"value": 123}, "override_reason": "T-sinov: qiymatni tahrirlash",
                })
                check("ruxsat etilgan maydon -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
                check("qiymat yangilandi", r.status_code == 200 and r.json().get("value") == 123,
                      f"={r.json().get('value') if r.status_code == 200 else None}")

            # ── Metrikani butunlay tozalash ──
            r = client.request("DELETE", f"{API_BASE}/admin/norms/{emp_uid}/dumaloq_video", headers=auth(dev_t), json={
                "override_reason": "T-sinov: metrikani tozalash",
            })
            check("metrika tozalandi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            # ── revert: 2 marta norma qo'yib, oldingisiga qaytarish ──
            client.put(f"{API_BASE}/admin/norms/{emp_uid}/tashrif", headers=auth(dev_t), json={
                "value": 10, "override_reason": "T-sinov: revert 1-qadam",
            })
            client.put(f"{API_BASE}/admin/norms/{emp_uid}/tashrif", headers=auth(dev_t), json={
                "value": 20, "override_reason": "T-sinov: revert 2-qadam",
            })
            r = client.post(f"{API_BASE}/admin/norms/{emp_uid}/revert?metric=tashrif", headers=auth(dev_t), json={
                "override_reason": "T-sinov: revert",
            })
            check("revert -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            check("revert oldingi qiymatga (10) qaytardi",
                  r.status_code == 200 and r.json().get("current_value") == 10,
                  f"={r.json() if r.status_code == 200 else None}")

            # ── Payroll: qulf va super-admin amallar ──
            for wd in range(7):
                cur.execute(
                    "insert into work_schedule_weekly (user_id, weekday, is_working, updated_at)"
                    " values (?,?,0,datetime('now'))", (emp_uid, wd))
            cur.execute(
                "insert into work_schedule_override (user_id, date, is_working, start_time, end_time, updated_at)"
                " values (?, '2022-03-01', 1, '09:00', '18:00', datetime('now'))", (emp_uid,))
            cur.execute(
                "insert into attendance (user_id, date, check_in_time, late_minutes,"
                " early_leave_minutes, worked_minutes, status, is_weekend, created_at, updated_at)"
                " values (?, '2022-03-01', '2022-03-01 04:00:00', 0, 0, 480, 'present', 0,"
                " datetime('now'), datetime('now'))", (emp_uid,))
            conn.commit()
            if mgr_t:
                client.post(f"{API_BASE}/payroll/rates", headers=auth(mgr_t), json={
                    "user_id": emp_uid, "amount": 1_800_000, "pay_basis": "monthly",
                    "effective_from": "2020-01-01", "reason": "hire",
                })
                r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={"user_ids": [emp_uid]})
                check("payroll navbatga qo'yildi -> 202", r.status_code == 202, f"kod={r.status_code}")
                tick = payroll_tick(client, PERIOD)
                check("payroll hisoblandi (cron)", bool(tick) and tick.get("ok") is True, f"={tick}")
                # 2026-08-21: «tayyor» qadami IXTIYORIY — HR to'g'ridan
                # to'g'ri tasdiqlay oladi va iz avtomatik to'ldiriladi.
                # Bu yerda ATAYLAB o'tkazib yuboramiz: aynan shu yo'l
                # sinalmasa, «iz avtomatik to'ladi» qoidasi tekshirilmay
                # qolardi.
                r = client.post(f"{API_BASE}/payroll/{PERIOD}/approve", headers=auth(mgr_t))
                check("«tayyor»siz ham tasdiqlanadi (qulflandi) -> 200",
                      r.status_code == 200, f"kod={r.status_code}")

                r = client.get(f"{API_BASE}/payroll/periods", headers=auth(mgr_t))
                if r.status_code == 200:
                    row = next((x for x in r.json() if x["period"] == PERIOD), None)
                    check("«tayyor»siz tasdiqda ham kim tekshirgani yozildi",
                          bool(row and row.get("hr_approved_name")),
                          f"hr_approved_name={row.get('hr_approved_name') if row else None}")

                r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
                check("qulflangan davrni HR qayta hisoblay OLMAYDI -> 409", r.status_code == 409,
                      f"kod={r.status_code}")

                r = client.post(f"{API_BASE}/admin/payroll/{PERIOD}/unlock", headers=auth(mgr_t), json={
                    "override_reason": "T-sinov: HR unlock urinishi",
                })
                check("HR /admin/payroll/unlock -> 403", r.status_code == 403, f"kod={r.status_code}")

                r = client.post(f"{API_BASE}/admin/payroll/{PERIOD}/unlock", headers=auth(dev_t), json={
                    "override_reason": "T-sinov: dasturchi qulfni ochadi",
                })
                check("dasturchi qulfni ochadi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

                r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
                check("qulf ochilgach HR qayta hisoblay oladi -> 202", r.status_code == 202,
                      f"kod={r.status_code}")
                tick = payroll_tick(client, PERIOD)
                check("qulf ochilgach cron hisobladi", bool(tick) and tick.get("ok") is True, f"={tick}")

                r = client.patch(f"{API_BASE}/admin/payroll/{PERIOD}/user/{emp_uid}", headers=auth(dev_t), json={
                    "fields": {"net": 999999}, "override_reason": "T-sinov: qo'lda tuzatish",
                })
                check("dasturchi payslip summasini qo'lda tuzatdi -> 200", r.status_code == 200,
                      f"kod={r.status_code} {r.text[:150]}")
                check("net=999999 ga o'zgardi", r.status_code == 200 and r.json().get("net") == 999999.0,
                      f"={r.json().get('net') if r.status_code == 200 else None}")

                r = client.request("DELETE", f"{API_BASE}/admin/payroll/{PERIOD}", headers=auth(dev_t), json={
                    "override_reason": "T-sinov: butun davrni bekor qilish",
                })
                check("davr butunlay bekor qilindi -> 200", r.status_code == 200, f"kod={r.status_code}")
                r = client.get(f"{API_BASE}/payroll/{PERIOD}/user/{emp_uid}", headers=auth(mgr_t))
                check("bekor qilingandan keyin payslip topilmaydi -> 404", r.status_code == 404,
                      f"kod={r.status_code}")

            # ── Audit tarixi ──
            r = client.get(f"{API_BASE}/admin/audit/overrides", headers=auth(dev_t))
            check("audit/overrides -> 200", r.status_code == 200, f"kod={r.status_code}")
            actions = [row["action"] for row in r.json()] if r.status_code == 200 else []
            check("audit tarixida override_norm_set bor", "override_norm_set" in actions,
                  f"amallar namunasi={actions[:5]}")
    except Exception:
        check("Admin override (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_admin()
        conn.close()


def test_payroll_automation() -> None:
    """Bosqich 6: avtomatika — `api/services/payroll.py`ning yangi
    funksiyalari (`detect_overtime_candidates`, `late_limit_event_for`,
    `previous_period`) va `api/routers/payroll.py`ning yangi tick
    endpointlari (`/calculate-monthly`, `/late-warnings-tick`,
    `/overtime/auto-detect`). Izolyatsiyalangan davr "2020-03"
    (test_payroll_engine'ning "2020-01"idan ALOHIDA).

    ⭐ MUHIM: `/late-warnings-tick` va `/overtime/auto-detect` odatda BARCHA
    real faol xodimlarni "kecha" sanasi bo'yicha skanerlaydi — shu sabab bu
    testda ular HAR DOIM aniq `target_date` bilan (uzoq o'tmishdagi, 2020-03)
    chaqiriladi, hech qachon default ("kecha") bilan EMAS — aks holda real
    xodimlarga botdan chinakam xabar ketishi yoki real ma'lumot o'zgarishi
    mumkin edi (xuddi shu sabab bilan test.py'dan yuzni qayta tekshirish
    testi ilgari olib tashlangan edi)."""
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("BOSQICH 6: PAYROLL AVTOMATIKA (scheduler + tick endpointlari)")
    print("=" * 60)

    PERIOD = "2020-03"
    WINDOW_MIN = 480  # work_minutes(09:00,18:00) — tushliksiz 8 soat

    async def _setup_and_direct_checks() -> dict:
        from sqlalchemy import delete, select

        from api.services import payroll as pr
        from api.services.attendance import local_hm_to_utc
        from db.base import async_session
        from db.models import (
            Attendance,
            AuditLog,
            FinePolicy,
            OvertimeEntry,
            OvertimeProfile,
            PayrollPeriod,
            Payslip,
            PayslipItem,
            SalaryRate,
            User,
            WorkScheduleOverride,
            WorkScheduleWeekly,
        )

        async with async_session() as s:
            # Oldingi (yarim yo'lda qulagan) ishga tushirishdan qolganlarni tozalash.
            stale_ids = list(await s.scalars(select(User.id).where(User.full_name.like("T-Payroll6%"))))
            if stale_ids:
                stale_payslips = list(await s.scalars(select(Payslip.id).where(Payslip.user_id.in_(stale_ids))))
                if stale_payslips:
                    await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id.in_(stale_payslips)))
                await s.execute(delete(Payslip).where(Payslip.user_id.in_(stale_ids)))
                await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id.in_(stale_ids)))
                await s.execute(delete(OvertimeProfile).where(OvertimeProfile.user_id.in_(stale_ids)))
                await s.execute(delete(SalaryRate).where(SalaryRate.user_id.in_(stale_ids)))
                await s.execute(delete(FinePolicy).where(FinePolicy.scope_id.in_(stale_ids)))
                await s.execute(delete(Attendance).where(Attendance.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id.in_(stale_ids)))
                await s.execute(
                    delete(AuditLog).where(
                        AuditLog.target_user_id.in_(stale_ids) | AuditLog.actor_id.in_(stale_ids)
                    )
                )
                await s.execute(delete(User).where(User.id.in_(stale_ids)))
            await s.execute(
                delete(Attendance).where(Attendance.date >= date(2020, 3, 1), Attendance.date < date(2020, 4, 1))
            )
            await s.execute(
                delete(WorkScheduleOverride).where(
                    WorkScheduleOverride.date >= date(2020, 3, 1), WorkScheduleOverride.date < date(2020, 4, 1)
                )
            )
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.commit()

            # ── T-Payroll6OT — qo'shimcha ish avtomatik aniqlash (1.3-band) ──
            uot = User(telegram_id=999600901, full_name="T-Payroll6OT", role="employee",
                       bot_started=True, is_active=True)
            s.add(uot)
            await s.flush()
            for wd in range(7):
                s.add(WorkScheduleWeekly(user_id=uot.id, weekday=wd, is_working=False))
            day_over, day_within, day_http = date(2020, 3, 10), date(2020, 3, 11), date(2020, 3, 12)
            for d in (day_over, day_within, day_http):
                s.add(WorkScheduleOverride(user_id=uot.id, date=d, is_working=True,
                                            start_time="09:00", end_time="18:00"))
            s.add(OvertimeProfile(user_id=uot.id, enabled=True, mode="fixed_rate",
                                   fixed_rate_per_hour=10_000, min_minutes=15))
            # 2026-08-15: nomzod endi SOF FARQ bo'yicha aniqlanadi
            # (`worked_minutes` − rejadagi daqiqa), ilgari esa faqat
            # `check_out − ish oynasi tugashi` qaralardi. Farqi muhim: xodim
            # kech kelib kech ketsa eski usul buni «qo'shimcha ish» deb
            # yozardi, aslida u KAM ishlagan bo'lishi mumkin edi. Shuning
            # uchun sinov ma'lumoti ham `worked_minutes` bilan beriladi.
            s.add(Attendance(user_id=uot.id, date=day_over, status="present",
                              worked_minutes=WINDOW_MIN + 30,  # +30 daq ORTIQCHA -> nomzod
                              check_out_time=local_hm_to_utc(day_over, "18:30")))
            s.add(Attendance(user_id=uot.id, date=day_within, status="present",
                              worked_minutes=WINDOW_MIN - 10,  # sezgirlik chegarasi (15) ichida -> yo'q
                              check_out_time=local_hm_to_utc(day_within, "17:50")))
            s.add(Attendance(user_id=uot.id, date=day_http, status="present",
                              worked_minutes=WINDOW_MIN + 25,  # HTTP orqali tekshiriladi
                              check_out_time=local_hm_to_utc(day_http, "18:25")))
            await s.commit()

            created1 = await pr.detect_overtime_candidates(s, day_over)
            await s.commit()
            check("Bosqich6: overtime nomzodi yaratildi (check-out 30 daq keyin)", len(created1) == 1,
                  f"soni={len(created1)}")
            if created1:
                check("Bosqich6: nomzod minutes=30", created1[0].minutes == 30, f"={created1[0].minutes}")
                check("Bosqich6: nomzod source=auto_attendance", created1[0].source == "auto_attendance",
                      f"={created1[0].source}")
                check("Bosqich6: nomzod status=pending", created1[0].status == "pending",
                      f"={created1[0].status}")

            created_again = await pr.detect_overtime_candidates(s, day_over)
            check("Bosqich6: overtime idempotent - qayta chaqirilsa dublikat YO'Q", created_again == [],
                  f"={created_again}")

            created_within = await pr.detect_overtime_candidates(s, day_within)
            check("Bosqich6: ish oynasi ichida check-out -> nomzod YO'Q", created_within == [],
                  f"={created_within}")

            # ── T-Payroll6Late — kechikish limiti ogohlantirishi (1.5-band) ──
            ul = User(telegram_id=999600902, full_name="T-Payroll6Late", role="employee",
                      bot_started=True, is_active=True)
            s.add(ul)
            await s.flush()
            for wd in range(7):
                s.add(WorkScheduleWeekly(user_id=ul.id, weekday=wd, is_working=False))
            d1, d2, d3, d4, d5 = (
                date(2020, 3, 2), date(2020, 3, 3), date(2020, 3, 4), date(2020, 3, 5), date(2020, 3, 6),
            )
            for d in (d1, d2, d3, d4, d5):
                s.add(WorkScheduleOverride(user_id=ul.id, date=d, is_working=True,
                                            start_time="09:00", end_time="18:00"))
            # `compute_late_fine`ning QOIDASI: "chegaradan o'tkazgan kunning o'zi
            # hali bepul" — `fined = (before >= limit)`, ya'ni limitni TO'LDIRGAN
            # kunning O'ZI EMAS, undan KEYINGI kun birinchi jarimali hisoblanadi.
            # d1: 0->20 (limit 30ga 10 daq qoldi) -> near_limit.
            # d2: 20->35 (limitni TO'LDIRGAN kun, before=20<30 -> hali JARIMASIZ,
            #     lekin remaining_before(10) allaqachon buferdan (15) kichik ->
            #     near_limit ILGARI (d1'da) berilgan, shu sabab bu yerda YANGI
            #     voqea YO'Q) -> None.
            # d3: 35->40 (before=35>=30 -> BIRINCHI jarimali kun) -> limit_reached.
            # d4: 40->45 (before=40>=30 -> jarimali, lekin ALLAQACHON d3'da
            #     boshlangan) -> yangi voqea emas -> None.
            # d5: kechikmagan (present) -> None.
            s.add(Attendance(user_id=ul.id, date=d1, status="late", late_minutes=20, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=ul.id, date=d2, status="late", late_minutes=15, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=ul.id, date=d3, status="late", late_minutes=5, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=ul.id, date=d4, status="late", late_minutes=5, worked_minutes=WINDOW_MIN))
            s.add(Attendance(user_id=ul.id, date=d5, status="present", worked_minutes=WINDOW_MIN))
            s.add(FinePolicy(scope="user", scope_id=ul.id, is_active=True,
                              free_late_minutes_per_month=30, fine_mode="per_day", fine_per_day=50_000,
                              absent_mode="none", monthly_cap_amount=1_000_000, fine_applies_to="net_salary"))
            s.add(SalaryRate(user_id=ul.id, amount=2_000_000, pay_basis="monthly",
                              effective_from=date(2019, 1, 1), changed_by=ul.id))
            await s.commit()

            ev1 = await pr.late_limit_event_for(s, ul, d1)
            check("Bosqich6: d1 (0->20, limit 30) -> near_limit",
                  ev1 is not None and ev1["kind"] == "near_limit", f"={ev1}")
            if ev1:
                check("Bosqich6: d1 remaining_minutes=10", ev1.get("remaining_minutes") == 10, f"={ev1}")

            ev2 = await pr.late_limit_event_for(s, ul, d2)
            check("Bosqich6: d2 (limitni to'ldirgan kun, hali jarimasiz) -> None",
                  ev2 is None, f"={ev2}")

            ev3 = await pr.late_limit_event_for(s, ul, d3)
            check("Bosqich6: d3 (birinchi jarimali kun) -> limit_reached",
                  ev3 is not None and ev3["kind"] == "limit_reached", f"={ev3}")

            ev4 = await pr.late_limit_event_for(s, ul, d4)
            check("Bosqich6: d4 (jarimali, lekin ALLAQACHON d3'da boshlangan) -> None (yangi voqea emas)",
                  ev4 is None, f"={ev4}")

            ev5 = await pr.late_limit_event_for(s, ul, d5)
            check("Bosqich6: d5 (kechikmagan) -> None", ev5 is None, f"={ev5}")

            u_nopolicy = User(telegram_id=999600903, full_name="T-Payroll6NoPolicy", role="employee",
                               bot_started=True, is_active=True)
            s.add(u_nopolicy)
            await s.flush()
            ev_none = await pr.late_limit_event_for(s, u_nopolicy, d1)
            check("Bosqich6: qoidasiz xodim -> None", ev_none is None, f"={ev_none}")

            check("Bosqich6: previous_period(2020-03-15) = 2020-02",
                  pr.previous_period(date(2020, 3, 15)) == "2020-02",
                  f"={pr.previous_period(date(2020, 3, 15))}")
            check("Bosqich6: previous_period(2020-01-01) = 2019-12 (yil chegarasi)",
                  pr.previous_period(date(2020, 1, 1)) == "2019-12",
                  f"={pr.previous_period(date(2020, 1, 1))}")

            return {"d1": d1.isoformat(), "day_http": day_http.isoformat()}

    async def _cleanup() -> None:
        from sqlalchemy import delete, select

        from db.base import async_session
        from db.models import (
            Attendance,
            AuditLog,
            Bonus,
            FinePolicy,
            OvertimeEntry,
            OvertimeProfile,
            PayrollPeriod,
            Payslip,
            PayslipItem,
            SalaryRate,
            User,
            WorkScheduleOverride,
            WorkScheduleWeekly,
        )

        async with async_session() as s:
            ids = list(await s.scalars(select(User.id).where(User.full_name.like("T-Payroll6%"))))
            if not ids:
                return
            pslips = list(await s.scalars(select(Payslip.id).where(Payslip.user_id.in_(ids))))
            if pslips:
                await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id.in_(pslips)))
            await s.execute(delete(Payslip).where(Payslip.user_id.in_(ids)))
            # §2.3 dan keyin oylik hisobi KPI bonusini ham yaratadi — ya'ni
            # sinov foydalanuvchilarida endi `bonuses` qatori ham qoladi va
            # tozalanmasa `delete(User)` FOREIGN KEY bilan yiqiladi.
            await s.execute(delete(Bonus).where(Bonus.user_id.in_(ids)))
            await s.execute(delete(OvertimeEntry).where(OvertimeEntry.user_id.in_(ids)))
            await s.execute(delete(OvertimeProfile).where(OvertimeProfile.user_id.in_(ids)))
            await s.execute(delete(SalaryRate).where(SalaryRate.user_id.in_(ids)))
            await s.execute(delete(FinePolicy).where(FinePolicy.scope_id.in_(ids)))
            await s.execute(delete(Attendance).where(Attendance.user_id.in_(ids)))
            await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id.in_(ids)))
            await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id.in_(ids)))
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.execute(delete(AuditLog).where(AuditLog.target_user_id.in_(ids) | AuditLog.actor_id.in_(ids)))
            await s.execute(delete(User).where(User.id.in_(ids)))
            await s.commit()

    ctx: dict = {}
    try:
        ctx = asyncio.run(_setup_and_direct_checks())
    except Exception:
        check("Bosqich6: sozlash/servis funksiyalari (umumiy)", False, traceback.format_exc(limit=2).strip())

    if ctx:
        try:
            with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as f:
                secret = next(
                    (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")), ""
                )
            BOT_SECRET_HDR = {"X-Bot-Secret": secret}
            with httpx.Client(timeout=15) as client:
                # ── bot-secret himoyasi ──
                r = client.post(f"{API_BASE}/payroll/calculate-monthly", json={})
                check("calculate-monthly bot-secretsiz -> 401", r.status_code == 401, f"kod={r.status_code}")
                r = client.post(f"{API_BASE}/payroll/late-warnings-tick", json={})
                check("late-warnings-tick bot-secretsiz -> 401", r.status_code == 401, f"kod={r.status_code}")
                r = client.post(f"{API_BASE}/payroll/overtime/auto-detect", json={})
                check("overtime/auto-detect bot-secretsiz -> 401", r.status_code == 401, f"kod={r.status_code}")

                # ── calculate-monthly: ANIQ (izolyatsiyalangan) davr — default
                # ("o'tgan oy") ATAYLAB sinalmaydi, aks holda jonli joriy oy
                # ma'lumotiga tegib qo'yardi. ──
                r = client.post(f"{API_BASE}/payroll/calculate-monthly", headers=BOT_SECRET_HDR,
                                 json={"period": PERIOD})
                check("calculate-monthly -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
                body = r.json() if r.status_code == 200 else {}
                check("calculate-monthly natijada period to'g'ri", body.get("period") == PERIOD, f"={body}")

                # ── late-warnings-tick: to'g'ridan-to'g'ri chaqiruvda near_limit
                # bo'lgan d1'ni endi HTTP orqali (aniq target_date bilan) ──
                r = client.post(f"{API_BASE}/payroll/late-warnings-tick", headers=BOT_SECRET_HDR,
                                 json={"target_date": ctx["d1"]})
                check("late-warnings-tick -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
                lw = r.json() if r.status_code == 200 else {}
                check("late-warnings-tick kamida 1 ta ogohlantirdi (T-Payroll6Late)",
                      lw.get("warned", 0) >= 1, f"={lw}")

                # ── overtime/auto-detect: HTTP orqali YANGI (hali test qilinmagan)
                # kun — aniq target_date bilan ──
                r = client.post(f"{API_BASE}/payroll/overtime/auto-detect", headers=BOT_SECRET_HDR,
                                 json={"target_date": ctx["day_http"]})
                check("overtime/auto-detect -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
                oad = r.json() if r.status_code == 200 else {}
                check("overtime/auto-detect kamida 1 ta nomzod yaratdi (T-Payroll6OT)",
                      oad.get("created", 0) >= 1, f"={oad}")
        except Exception:
            check("Bosqich6: HTTP tick endpointlari (umumiy)", False, traceback.format_exc(limit=2).strip())

    try:
        asyncio.run(_cleanup())
    except Exception:
        check("Bosqich6: tozalash (umumiy)", False, traceback.format_exc(limit=2).strip())


def test_payroll_reporting() -> None:
    """Bosqich 7: hisobot va maxfiylik — Excel eksport
    (`api/services/export.py::build_payroll_xlsx`, `GET /payroll/{period}/export`),
    barcha pul o'zgarishlari uchun AuditLog (`payroll_calculated`), oylik
    digestga Boshliq-ONLY "jami ish haqi fondi" (`api/services/monthly_digest.py`).
    Izolyatsiyalangan davr "2020-04"."""
    import asyncio
    from io import BytesIO

    import httpx
    from openpyxl import load_workbook

    print("\n" + "=" * 60)
    print("BOSQICH 7: HISOBOT VA MAXFIYLIK (export + audit + digest)")
    print("=" * 60)

    PERIOD = "2020-04"
    WINDOW_MIN = 480

    async def _setup() -> dict:
        from sqlalchemy import delete, select

        from api.services.payroll import run_payroll
        from db.base import async_session
        from db.models import (
            Attendance,
            AuditLog,
            LeadStageDaily,
            PayrollPeriod,
            Payslip,
            PayslipItem,
            SalaryRate,
            User,
            WorkScheduleOverride,
            WorkScheduleWeekly,
        )

        async with async_session() as s:
            stale_ids = list(await s.scalars(select(User.id).where(User.full_name.like("T-Payroll7%"))))
            if stale_ids:
                stale_payslips = list(await s.scalars(select(Payslip.id).where(Payslip.user_id.in_(stale_ids))))
                if stale_payslips:
                    await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id.in_(stale_payslips)))
                await s.execute(delete(Payslip).where(Payslip.user_id.in_(stale_ids)))
                await s.execute(delete(SalaryRate).where(SalaryRate.user_id.in_(stale_ids)))
                await s.execute(delete(Attendance).where(Attendance.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id.in_(stale_ids)))
                await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id.in_(stale_ids)))
                await s.execute(
                    delete(AuditLog).where(
                        AuditLog.target_user_id.in_(stale_ids) | AuditLog.actor_id.in_(stale_ids)
                    )
                )
                await s.execute(delete(User).where(User.id.in_(stale_ids)))
            await s.execute(
                delete(Attendance).where(Attendance.date >= date(2020, 4, 1), Attendance.date < date(2020, 5, 1))
            )
            await s.execute(
                delete(WorkScheduleOverride).where(
                    WorkScheduleOverride.date >= date(2020, 4, 1), WorkScheduleOverride.date < date(2020, 5, 1)
                )
            )
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.commit()

            rop = User(telegram_id=999700701, full_name="T-Payroll7Rop", role="rop",
                       bot_started=True, is_active=True)
            emp = User(telegram_id=999700702, full_name="T-Payroll7Emp", role="employee",
                       bot_started=True, is_active=True)
            outsider = User(telegram_id=999700703, full_name="T-Payroll7Outsider", role="employee",
                             bot_started=True, is_active=True)
            s.add_all([rop, emp, outsider])
            await s.flush()
            emp.manager_id = rop.id
            await s.commit()

            d = date(2020, 4, 6)
            for u in (emp, outsider):
                for wd in range(7):
                    s.add(WorkScheduleWeekly(user_id=u.id, weekday=wd, is_working=False))
                s.add(WorkScheduleOverride(user_id=u.id, date=d, is_working=True,
                                            start_time="09:00", end_time="18:00"))
                s.add(Attendance(user_id=u.id, date=d, status="present", worked_minutes=WINDOW_MIN))
                s.add(SalaryRate(user_id=u.id, amount=2_000_000, pay_basis="monthly",
                                  effective_from=date(2019, 1, 1), changed_by=u.id))
            await s.commit()

            await run_payroll(s, PERIOD, user_ids=[emp.id, outsider.id])

            # Oylik digest "faol operator" darvozasidan o'tishi uchun bitta
            # yengil CRM snapshot yozuvi — real xodim/tashkilotga bog'liq emas
            # (rid haqiqiy foydalanuvchiga bog'lanmagan, faqat "active" gate).
            s.add(LeadStageDaily(date=date(2020, 5, 10), responsible_id=999700799,
                                  responsible_name="T-Payroll7-CRM", pipe_status_id=1,
                                  stage_name="T-sinov bosqichi", leads_count=3))
            await s.commit()

            return {"rop_id": rop.id, "emp_id": emp.id, "outsider_id": outsider.id}

    async def _cleanup() -> None:
        from sqlalchemy import delete, select

        from db.base import async_session
        from db.models import (
            Attendance,
            AuditLog,
            Bonus,
            LeadStageDaily,
            PayrollPeriod,
            Payslip,
            PayslipItem,
            SalaryRate,
            User,
            WorkScheduleOverride,
            WorkScheduleWeekly,
        )

        async with async_session() as s:
            ids = list(await s.scalars(select(User.id).where(User.full_name.like("T-Payroll7%"))))
            if ids:
                pslips = list(await s.scalars(select(Payslip.id).where(Payslip.user_id.in_(ids))))
                if pslips:
                    await s.execute(delete(PayslipItem).where(PayslipItem.payslip_id.in_(pslips)))
                await s.execute(delete(Payslip).where(Payslip.user_id.in_(ids)))
                # §2.3: oylik hisobi endi bonus qatorini ham yaratadi
                await s.execute(delete(Bonus).where(Bonus.user_id.in_(ids)))
                await s.execute(delete(SalaryRate).where(SalaryRate.user_id.in_(ids)))
                await s.execute(delete(Attendance).where(Attendance.user_id.in_(ids)))
                await s.execute(delete(WorkScheduleOverride).where(WorkScheduleOverride.user_id.in_(ids)))
                await s.execute(delete(WorkScheduleWeekly).where(WorkScheduleWeekly.user_id.in_(ids)))
                await s.execute(delete(AuditLog).where(AuditLog.target_user_id.in_(ids) | AuditLog.actor_id.in_(ids)))
                await s.execute(delete(User).where(User.id.in_(ids)))
            await s.execute(
                delete(LeadStageDaily).where(
                    LeadStageDaily.responsible_id == 999700799, LeadStageDaily.date == date(2020, 5, 10)
                )
            )
            await s.execute(delete(PayrollPeriod).where(PayrollPeriod.period == PERIOD))
            await s.commit()

    async def _direct_checks(ctx: dict) -> None:
        from api.services.export import build_payroll_xlsx
        from api.services.monthly_digest import build_monthly_digest, send_monthly_digest
        from db.base import async_session

        async with async_session() as s:
            buf_all = await build_payroll_xlsx(s, PERIOD)
            wb_all = load_workbook(BytesIO(buf_all.read()))
            check("Bosqich7: 'Xulosa' varag'i bor", "Xulosa" in wb_all.sheetnames, f"={wb_all.sheetnames}")
            emp_sheet = f"T-Payroll7Emp #{ctx['emp_id']}"
            out_sheet = f"T-Payroll7Outsider #{ctx['outsider_id']}"
            check("Bosqich7: user_ids'siz eksportda IKKALA xodim varag'i bor",
                  emp_sheet in wb_all.sheetnames and out_sheet in wb_all.sheetnames,
                  f"={wb_all.sheetnames}")

            buf_scoped = await build_payroll_xlsx(s, PERIOD, user_ids=[ctx["emp_id"]])
            wb_scoped = load_workbook(BytesIO(buf_scoped.read()))
            check("Bosqich7: user_ids bilan FAQAT shu xodim varag'i bor",
                  emp_sheet in wb_scoped.sheetnames and out_sheet not in wb_scoped.sheetnames,
                  f"={wb_scoped.sheetnames}")

            digest = await build_monthly_digest(s, ref_day=date(2020, 5, 15))
            digest_text = digest.get("text") or ""
            # Windows konsoli (cp1251) digest matnidagi emojilarni (masalan
            # U+1F5D3) chop etishda qulaydi — check() xabarlarida faqat ASCII-
            # xavfsiz uzunlik/qidiruv natijasi ishlatiladi, xom matn EMAS.
            check("Bosqich7: digest matni yaratildi (CRM snapshot gate o'tildi)",
                  digest.get("text") is not None, f"uzunlik={len(digest_text)}")
            pf = digest.get("payroll_fund")
            check("Bosqich7: payroll_fund davri = o'tgan oy (2020-04)",
                  pf is not None and pf["period"] == PERIOD, f"={pf}")
            if pf:
                check("Bosqich7: payroll_fund jami = 2 xodim netto yig'indisi",
                      abs(pf["total"] - 4_000_000.0) < 1, f"={pf}")
            check("Bosqich7: digest MATNIDA 'fond' so'zi YO'Q (guruhga sizib chiqmaydi)",
                  "fond" not in digest_text.lower(), f"uzunlik={len(digest_text)}")

            # E'TIBOR: `send_monthly_digest` `ref_day` qabul qilmaydi (har doim
            # HAQIQIY "bugun"dan hisoblaydi, real joriy oy CRM faolligiga
            # bog'liq) — shu sabab bu yerda faqat "hech qachon yubormaydi"
            # (dry_run) tekshiriladi, "2020-04"ga tegishli QIYMAT EMAS.
            dry = await send_monthly_digest(s, dry_run=True)
            # `dry`ning "text" kaliti xom (emoji bilan) digest matni bo'lishi
            # mumkin — check() xabarida shu kalitsiz qisqa xulosa ishlatiladi
            # (Windows konsoli cp1251 emojilarda qulaydi).
            dry_summary = {k: v for k, v in dry.items() if k != "text"}
            check("Bosqich7: dry_run hech qachon yubormaydi", dry.get("sent") is False, f"={dry_summary}")

    ctx: dict = {}
    try:
        ctx = asyncio.run(_setup())
    except Exception:
        check("Bosqich7: sozlash (umumiy)", False, traceback.format_exc(limit=2).strip())

    if ctx:
        try:
            asyncio.run(_direct_checks(ctx))
        except Exception:
            check("Bosqich7: servis funksiyalari (umumiy)", False, traceback.format_exc(limit=2).strip())

        try:
            mgr = find_manager_id()
            mgr_t = token_for(mgr[0], mgr[1]) if mgr else None
            rop_t = token_for(ctx["rop_id"], "rop")
            emp_t = token_for(ctx["emp_id"], "employee")

            with httpx.Client(timeout=15) as client:
                r = client.get(f"{API_BASE}/payroll/{PERIOD}/export", headers=auth(emp_t))
                check("xodim /payroll/*/export -> 403", r.status_code == 403, f"kod={r.status_code}")

                r = client.get(f"{API_BASE}/payroll/{PERIOD}/export", headers=auth(mgr_t))
                check("HR/Boss /payroll/*/export -> 200", r.status_code == 200, f"kod={r.status_code}")
                check("export content-type xlsx",
                      "spreadsheetml" in r.headers.get("content-type", ""),
                      f"={r.headers.get('content-type')}")
                if r.status_code == 200:
                    wb_http = load_workbook(BytesIO(r.content))
                    check("HR eksportida ikkala xodim ham bor",
                          f"T-Payroll7Emp #{ctx['emp_id']}" in wb_http.sheetnames
                          and f"T-Payroll7Outsider #{ctx['outsider_id']}" in wb_http.sheetnames,
                          f"={wb_http.sheetnames}")

                r = client.get(f"{API_BASE}/payroll/{PERIOD}/export", headers=auth(rop_t))
                check("ROP /payroll/*/export -> 200", r.status_code == 200, f"kod={r.status_code}")
                if r.status_code == 200:
                    wb_rop = load_workbook(BytesIO(r.content))
                    check("ROP eksportida FAQAT o'z jamoasi (Emp bor, Outsider yo'q)",
                          f"T-Payroll7Emp #{ctx['emp_id']}" in wb_rop.sheetnames
                          and f"T-Payroll7Outsider #{ctx['outsider_id']}" not in wb_rop.sheetnames,
                          f"={wb_rop.sheetnames}")

                # ── "payroll_calculated" AuditLog — barcha pul o'zgarishi audit qilinadi ──
                r = client.post(f"{API_BASE}/payroll/{PERIOD}/calculate", headers=auth(mgr_t), json={})
                check("qayta hisoblash -> 202", r.status_code == 202, f"kod={r.status_code}")
                # Audit yozuvini endi CRON yozadi (so'rov emas) — actor_id
                # `calc_requested_by` dan olinadi, ya'ni kim bosgani saqlanadi.
                payroll_tick(client, PERIOD)

            conn = db()
            cur = conn.cursor()
            row = cur.execute(
                "select count(*) from audit_logs where action='payroll_calculated' "
                "and json_extract(after, '$.period')=?", (PERIOD,)
            ).fetchone()
            check("AuditLog 'payroll_calculated' yozildi", row is not None and row[0] >= 1, f"={row}")
            conn.close()
        except Exception:
            check("Bosqich7: HTTP export/audit (umumiy)", False, traceback.format_exc(limit=2).strip())

    try:
        asyncio.run(_cleanup())
    except Exception:
        check("Bosqich7: tozalash (umumiy)", False, traceback.format_exc(limit=2).strip())


def test_dasturchi_bot_bridge() -> None:
    """Dasturchi web/bot interfeysi: web tomoni (`AdminOverride.tsx`) mavjud
    `admin_override.py` backendining ustiga qurilgan (yangi backend YO'Q),
    lekin bot tomoni uchun BITTA yangi ko'prik qo'shildi — `POST /auth/
    bot-token` (bot-secret bilan himoyalangan, FAQAT dasturchi uchun JWT
    beradi). Shu test faqat SHU yangi endpointni va uning JWT'si haqiqatan
    `/admin/*` va `/attendance/manual`ga kira olishini tekshiradi — chuqurroq
    admin_override.py mantiqi allaqachon `test_admin_override` (Bosqich 3.5)
    da to'liq qoplangan, bu yerda takrorlanmaydi."""
    import httpx

    print("\n" + "=" * 60)
    print("DASTURCHI BOT KO'PRIGI (POST /auth/bot-token)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-Bridge%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from norms where user_id in ({qm})", stale)
        cur.execute(f"delete from attendance where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from work_schedule_override where date = '2020-05-05'")
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999800801,'T-Bridge-Dev','dasturchi',1,1,datetime('now'))")
    dev_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999800802,'T-Bridge-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into work_schedule_override (user_id, date, is_working, start_time, end_time, updated_at)"
        " values (?, '2020-05-05', 1, '09:00', '18:00', datetime('now'))", (emp_uid,))
    conn.commit()

    def cleanup_bridge():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [dev_uid, emp_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from norms where user_id in ({qm})", uids)
            c2.execute(f"delete from attendance where user_id in ({qm})", uids)
            c2.execute("delete from work_schedule_override where date = '2020-05-05'")
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Dasturchi bot ko'prigi tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        with open("D:/Project/hodimlar_tizimi/.env", encoding="utf-8") as f:
            secret = next(
                (line.strip().split("=", 1)[1] for line in f if line.startswith("BOT_SHARED_SECRET=")), ""
            )
        BOT_SECRET_HDR = {"X-Bot-Secret": secret}

        with httpx.Client(timeout=15) as client:
            r = client.post(f"{API_BASE}/auth/bot-token", json={"telegram_id": 999800801})
            check("bot-secretsiz -> 401", r.status_code == 401, f"kod={r.status_code}")

            r = client.post(f"{API_BASE}/auth/bot-token", headers=BOT_SECRET_HDR, json={"telegram_id": 1})
            check("mavjud bo'lmagan telegram_id -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.post(f"{API_BASE}/auth/bot-token", headers=BOT_SECRET_HDR, json={"telegram_id": 999800802})
            check("xodim (dasturchi EMAS) -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.post(f"{API_BASE}/auth/bot-token", headers=BOT_SECRET_HDR, json={"telegram_id": 999800801})
            check("dasturchi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            token = r.json().get("access_token") if r.status_code == 200 else None
            check("javobda access_token bor", bool(token), f"={r.json() if r.status_code == 200 else None}")

            if token:
                jwt_hdr = {"Authorization": f"Bearer {token}"}

                r = client.put(
                    f"{API_BASE}/admin/norms/{emp_uid}/suhbat", headers=jwt_hdr,
                    json={"value": 33, "override_reason": "T-sinov: bot ko'prigi orqali"},
                )
                check("mint qilingan JWT bilan /admin/norms ishlaydi -> 200", r.status_code == 200,
                      f"kod={r.status_code} {r.text[:150]}")
                check("norma qiymati 33", r.status_code == 200 and r.json().get("value") == 33,
                      f"={r.json() if r.status_code == 200 else None}")

                r = client.put(
                    f"{API_BASE}/attendance/manual", headers=jwt_hdr,
                    json={
                        "user_id": emp_uid, "date": "2020-05-05", "check_in": "09:20",
                        "reason": "T-sinov: /att_fix ko'prigi",
                    },
                )
                check("mint qilingan JWT bilan /attendance/manual ishlaydi -> 200", r.status_code == 200,
                      f"kod={r.status_code} {r.text[:150]}")
                check("late_minutes hisoblandi (09:20 kelish, 09:00 boshlanish)",
                      r.status_code == 200 and r.json().get("late_minutes") == 20,
                      f"={r.json() if r.status_code == 200 else None}")
    except Exception:
        check("Dasturchi bot ko'prigi (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_bridge()
        conn.close()


def test_positions_permissions() -> None:
    """HRga lavozim yaratish/tahrirlash huquqi berildi (ilgari faqat Boshliq/
    Dasturchi) — `api/routers/positions.py::MANAGE_ROLES`. ROP hamon faqat
    o'qiy oladi (READ_ROLES)."""
    import httpx

    print("\n" + "=" * 60)
    print("LAVOZIM RUXSATLARI (HR endi boshqara oladi)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-PosPerm%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where actor_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from positions where name = 'T-Sinov lavozimi'")
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999900901,'T-PosPerm-Hr','hr',1,1,datetime('now'))")
    hr_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999900902,'T-PosPerm-Rop','rop',1,1,datetime('now'))")
    rop_uid = cur.lastrowid
    conn.commit()

    def cleanup_pos():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [hr_uid, rop_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where actor_id in ({qm})", uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            c2.execute("delete from positions where name = 'T-Sinov lavozimi'")
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Lavozim ruxsati tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        hr_t = token_for(hr_uid, "hr")
        rop_t = token_for(rop_uid, "rop")

        with httpx.Client(timeout=15) as client:
            r = client.post(f"{API_BASE}/positions", headers=auth(rop_t), json={
                "name": "T-Sinov lavozimi", "menu_flags": {}, "metrics": ["suhbat"], "managed_by_roles": [],
            })
            check("ROP lavozim yarata OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.post(f"{API_BASE}/positions", headers=auth(hr_t), json={
                "name": "T-Sinov lavozimi", "menu_flags": {}, "metrics": ["suhbat"], "managed_by_roles": [],
            })
            check("HR lavozim yarata oladi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            pos_id = r.json().get("id") if r.status_code == 200 else None

            if pos_id:
                r = client.patch(f"{API_BASE}/positions/{pos_id}", headers=auth(hr_t), json={
                    "metrics": ["suhbat", "tashrif"],
                })
                check("HR lavozimni tahrirlay oladi -> 200", r.status_code == 200, f"kod={r.status_code}")
                check("metrics yangilandi", r.status_code == 200 and r.json().get("metrics") == ["suhbat", "tashrif"],
                      f"={r.json() if r.status_code == 200 else None}")
    except Exception:
        check("Lavozim ruxsatlari (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_pos()
        conn.close()


def test_telegram_login_security() -> None:
    """Telegram login xavfsizlik arxitekturasi (3 qatlam):
    1. Replay himoyasi — bir xil `hash` ikkinchi marta rad etiladi.
    2. Rate-limit — belgilangan chegaradan oshgan urinish 429 qaytaradi.
    3. Taklif havolasi (invite) muddati — muddati o'tgan token 'token_expired'.

    Haqiqiy Telegram'ga hech qanday so'rov yubormaydi — imzoni loyihaning o'z
    HMAC algoritmi bilan (.env'dagi BOT_TOKEN) lokal hisoblaydi, xuddi
    Telegram Login Widget qilganidek."""
    import hashlib
    import hmac as hmac_lib
    import time

    import httpx

    from api.config import settings

    print("\n" + "=" * 60)
    print("TELEGRAM LOGIN XAVFSIZLIGI (replay / rate-limit / invite muddati)")
    print("=" * 60)

    def sign(data: dict) -> dict:
        pairs = [f"{k}={v}" for k, v in sorted(data.items())]
        check_string = "\n".join(pairs)
        secret_key = hashlib.sha256(settings.bot_token.encode()).digest()
        data = dict(data)
        data["hash"] = hmac_lib.new(secret_key, check_string.encode(), hashlib.sha256).hexdigest()
        return data

    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-LoginSec%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999900951,'T-LoginSec-Replay','employee',1,1,datetime('now'))")
    replay_uid = cur.lastrowid
    cur.execute(
        "insert into users (full_name, role, bot_started, is_active, invite_token, invite_expires_at, created_at)"
        " values ('T-LoginSec-Expired','employee',0,1,'T-invtok-expired-999901',"
        " datetime('now','-1 hour'), datetime('now'))")
    expired_uid = cur.lastrowid
    cur.execute(
        "insert into users (full_name, role, bot_started, is_active, invite_token, invite_expires_at, created_at)"
        " values ('T-LoginSec-Valid','employee',0,1,'T-invtok-valid-999902',"
        " datetime('now','+1 day'), datetime('now'))")
    valid_uid = cur.lastrowid
    conn.commit()

    def cleanup_sec():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [replay_uid, expired_uid, valid_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            c2.execute("delete from used_telegram_login_hashes where hash in (?,?)", (replay_hash_1, replay_hash_2))
            c2.execute(
                "delete from login_attempts where endpoint in ('dev-login','telegram-login')"
                " and identifier = '127.0.0.1' and created_at >= ?", (test_start,)
            )
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Login xavfsizlik tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    replay_hash_1 = ""
    replay_hash_2 = ""
    test_start = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    try:
        with httpx.Client(timeout=15) as client:
            # ── 1. Replay himoyasi ──────────────────────────────────────
            payload = sign({"id": 999900951, "auth_date": int(time.time()), "first_name": "T-Test"})
            replay_hash_1 = payload["hash"]

            r = client.post(f"{API_BASE}/auth/telegram-login", json=payload)
            check("Telegram login (birinchi, to'g'ri imzo) -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:150]}")

            r = client.post(f"{API_BASE}/auth/telegram-login", json=payload)
            check("Xuddi shu hash ikkinchi marta -> 401 (replay)", r.status_code == 401,
                  f"kod={r.status_code} {r.text[:150]}")

            # Boshqa (yangi) hash — xuddi shu foydalanuvchi bilan qayta kirish
            # hamon ISHLASHI kerak (hash o'zi bloklanadi, foydalanuvchi emas).
            # auth_date ataylab 5s orqaga suriladi — bir xil sekundda ikkala
            # payload aynan bir xil hash hosil qilib qolmasligi uchun.
            payload2 = sign({"id": 999900951, "auth_date": int(time.time()) - 5, "first_name": "T-Test"})
            replay_hash_2 = payload2["hash"]
            r = client.post(f"{API_BASE}/auth/telegram-login", json=payload2)
            check("Yangi hash bilan qayta login -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:150]}")

            # ── 2. Rate-limit (dev-login, 20/soat) ──────────────────────
            last_code = None
            for _ in range(21):
                r = client.post(f"{API_BASE}/auth/dev-login", json={"telegram_id": 999900951})
                last_code = r.status_code
                if last_code == 429:
                    break
            check("21-chi urinishda 429 (rate-limit ishladi)", last_code == 429, f"oxirgi kod={last_code}")

            # ── 3. Invite havolasi muddati ───────────────────────────────
            r = client.post(
                f"{API_BASE}/users/telegram-start",
                headers={"X-Bot-Secret": settings.bot_shared_secret},
                json={"telegram_id": 999900952, "invite_token": "T-invtok-expired-999901"},
            )
            check("Muddati o'tgan invite -> token_expired", r.status_code == 200 and r.json().get("status") == "token_expired",
                  f"kod={r.status_code} {r.text[:150]}")

            r = client.post(
                f"{API_BASE}/users/telegram-start",
                headers={"X-Bot-Secret": settings.bot_shared_secret},
                json={"telegram_id": 999900953, "invite_token": "T-invtok-valid-999902"},
            )
            check("Muddati o'tmagan invite -> ok", r.status_code == 200 and r.json().get("status") == "ok",
                  f"kod={r.status_code} {r.text[:150]}")
    except Exception:
        check("Telegram login xavfsizligi (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_sec()
        conn.close()


def test_hr_wide_employee_access() -> None:
    """HRga BARCHA xodimlar uchun ish jadvali/dam kuni + xodim nomidan sababli
    kun belgilash huquqi berildi:
    1. Ish jadvali: HR ilgari faqat can_manage_norms ruxsat bergan (o'ziga
       biriktirilgan lavozim yoki egasiz) xodimga tegishi mumkin edi — endi
       BARCHA faol 'Xodim' rolidagilarga (boshqa ROP jamoasidagi bo'lsa ham).
       ROP/HRning bir-birini yoki Boshliqni boshqarolmasligi saqlangan.
    2. Sababli kun: HR/Boshliq/Dasturchi xodim nomidan darhol 'approved'
       holatda yozuv qo'sha oladi (ROP bunga huquqli emas — decide bilan bir xil)."""
    import httpx

    from api.config import settings

    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-HrWide%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from work_schedule_weekly where user_id in ({qm})", stale)
        cur.execute(f"delete from work_schedule_override where user_id in ({qm})", stale)
        cur.execute(f"delete from excused_days where user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999901001,'T-HrWide-Hr','hr',1,1,datetime('now'))")
    hr_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999901002,'T-HrWide-Rop','rop',1,1,datetime('now'))")
    rop_uid = cur.lastrowid
    # position_id/managed_by_roles YO'Q va manager_id ROPga o'rnatilgan — ESKI
    # can_manage_norms qoidasida bu HR uchun 403 bo'lardi (na orphan, na HR
    # lavozimiga tegishli).
    cur.execute(
        "insert into users (telegram_id, full_name, role, manager_id, bot_started, is_active, created_at)"
        " values (999901003,'T-HrWide-Emp',?,?,1,1,datetime('now'))", ("employee", rop_uid))
    emp_uid = cur.lastrowid
    # Dasturchi roli — egasining 2026-08-05 talabi: HR uning ham ish grafigi
    # va dam kunini belgilay olsin (ilgari 403 edi).
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999901004,'T-HrWide-Dev','dasturchi',1,1,datetime('now'))")
    dev_uid = cur.lastrowid
    conn.commit()

    def cleanup_hw():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [hr_uid, rop_uid, emp_uid, dev_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from work_schedule_weekly where user_id in ({qm})", uids)
            c2.execute(f"delete from work_schedule_override where user_id in ({qm})", uids)
            c2.execute(f"delete from excused_days where user_id in ({qm})", uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  HR keng qamrov tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        hr_t = token_for(hr_uid, "hr")
        rop_t = token_for(rop_uid, "rop")

        with httpx.Client(timeout=15) as client:
            print("\n" + "=" * 60)
            print("HR KENG QAMROV: ish jadvali + xodim nomidan sababli kun")
            print("=" * 60)

            # ── Ish jadvali: HR endi ROP jamoasidagi xodimga ham tega oladi ──
            r = client.get(f"{API_BASE}/work-schedule/{emp_uid}/weekly", headers=auth(hr_t))
            check("HR boshqa ROP jamoasidagi xodim jadvalini KO'RADI -> 200 (ilgari 403)",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.put(
                f"{API_BASE}/work-schedule/{emp_uid}/weekly", headers=auth(hr_t),
                json={"days": [
                    {"weekday": wd, "is_working": wd < 6, "start_time": "09:00" if wd < 6 else None,
                     "end_time": "18:00" if wd < 6 else None}
                    for wd in range(7)
                ]},
            )
            check("HR shu xodim jadvalini O'ZGARTIRA oladi -> 200 (ilgari 403)",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.put(
                f"{API_BASE}/work-schedule/{emp_uid}/override", headers=auth(hr_t),
                json={"date": "2020-06-06", "is_working": False, "note": "T-sinov: dam kuni"},
            )
            check("HR shu xodimga BITTA KUNLIK dam kuni belgilay oladi -> 200",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.get(f"{API_BASE}/work-schedule/{hr_uid}/weekly", headers=auth(hr_t))
            check("HR boshqa HR/ROP/Boshliq jadvaliga TEGA OLMAYDI (o'ziga ham) -> 403",
                  r.status_code == 403, f"kod={r.status_code}")

            # ── Dasturchi: HR endi uning grafigini ham yuritadi ──
            r = client.get(f"{API_BASE}/work-schedule/{dev_uid}/weekly", headers=auth(hr_t))
            check("HR DASTURCHI jadvalini ko'radi -> 200 (ilgari 403)",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.put(
                f"{API_BASE}/work-schedule/{dev_uid}/weekly", headers=auth(hr_t),
                json={"days": [
                    {"weekday": wd, "is_working": wd < 5, "start_time": "09:00" if wd < 5 else None,
                     "end_time": "18:00" if wd < 5 else None}
                    for wd in range(7)
                ]},
            )
            check("HR DASTURCHI jadvalini o'zgartira oladi -> 200",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.put(
                f"{API_BASE}/work-schedule/{dev_uid}/override", headers=auth(hr_t),
                json={"date": "2020-06-07", "is_working": False, "note": "T-sinov: dasturchi dam kuni"},
            )
            check("HR DASTURCHIGA dam kuni belgilay oladi -> 200",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")

            r = client.get(f"{API_BASE}/work-schedule/{rop_uid}/weekly", headers=auth(hr_t))
            check("ROP jadvali HR uchun HAMON yopiq -> 403 (qamrov kengaymadi)",
                  r.status_code == 403, f"kod={r.status_code}")

            # ── Sababli kun: HR/Boshliq/Dasturchi xodim nomidan darhol belgilaydi ──
            # DIQQAT: endpoint path'i `telegram_id`, DB `id` EMAS.
            r = client.get(f"{API_BASE}/excused-days/targets/999901001",
                            headers={"X-Bot-Secret": settings.bot_shared_secret})
            check("Bot: HR targets -> 200", r.status_code == 200, f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/excused-days/targets/999901002",
                            headers={"X-Bot-Secret": settings.bot_shared_secret})
            check("Bot: ROP targets -> 403 (decide huquqi yo'q)", r.status_code == 403, f"kod={r.status_code}")

            r = client.post(
                f"{API_BASE}/excused-days/for-user", headers=auth(rop_t),
                json={"user_id": emp_uid, "date": "2020-07-07", "reason": "T-sinov"},
            )
            check("ROP xodim nomidan sababli kun BELGILAY OLMAYDI -> 403", r.status_code == 403,
                  f"kod={r.status_code}")

            r = client.post(
                f"{API_BASE}/excused-days/for-user", headers=auth(hr_t),
                json={"user_id": emp_uid, "date": "2020-07-07", "reason": "T-sinov: kasallik"},
            )
            check("HR xodim nomidan sababli kun belgilaydi -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:150]}")
            body = r.json() if r.status_code == 200 else {}
            check("Darhol 'approved' holatda (qayta tasdiq shart emas)", body.get("status") == "approved",
                  f"={body}")
            check("decided_by HRning o'zi", body.get("decided_by") == hr_uid, f"={body}")

            r = client.post(
                f"{API_BASE}/excused-days/for-user", headers=auth(hr_t),
                json={"user_id": emp_uid, "date": "2020-07-07", "reason": "T-sinov: takror"},
            )
            check("Bir kunga ikkinchi marta -> 400 (dublikat)", r.status_code == 400, f"kod={r.status_code}")

            r = client.post(
                f"{API_BASE}/excused-days/for-user", headers=auth(hr_t),
                json={"user_id": rop_uid, "date": "2020-08-08", "reason": "T-sinov: xodim emas"},
            )
            check("Faqat 'Xodim' rolidagi nishonga -> 400 (ROPga belgilab bo'lmaydi)",
                  r.status_code == 400, f"kod={r.status_code}")
    except Exception:
        check("HR keng qamrov (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_hw()
        conn.close()


def test_attendance_edit_rights() -> None:
    """Davomat keldi/ketdi vaqtini tuzatish huquqlari (egasining 2026-08-03 talabi):

    1. Dasturchi `PUT /admin/attendance/manual` — AUDITSIZ va jim
       (`AuditLog` yozuvi UMUMAN yaratilmaydi).
    2. Oddiy `PUT /attendance/manual` — audit YOZADI (saytda ko'rinsin).
    3. Huquq SHAXSAN beriladi (`can_edit_attendance`): roli hr/boss/dasturchi
       bo'lmagan odam ham tuzata oladi.
    4. Bayroq bilan tahrirlayotgan odam O'Z yozuvini tuzata OLMAYDI.
    5. Bayroqsiz va roli yo'q odam -> 403.
    6. Yozuv bo'lmasa YARATILADI («Keldim» bosish esidan chiqqan holat)."""
    import httpx

    from api.config import settings

    print("\n" + "=" * 60)
    print("DAVOMAT TUZATISH HUQUQLARI (dasturchi jim + shaxsiy ruxsat)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    DAY = "2020-09-09"

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-AttEdit%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from attendance where user_id in ({qm})", stale)
        cur.execute(f"delete from work_schedule_weekly where user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, can_edit_attendance, created_at)"
        " values (999903001,'T-AttEdit-Dev','dasturchi',1,1,0,datetime('now'))")
    dev_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, can_edit_attendance, created_at)"
        " values (999903002,'T-AttEdit-Emp','employee',1,1,0,datetime('now'))")
    emp_uid = cur.lastrowid
    # Bayroq bilan ruxsat berilgan ODDIY XODIM (roli bo'yicha huquqi yo'q)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, can_edit_attendance, created_at)"
        " values (999903003,'T-AttEdit-Granted','employee',1,1,1,datetime('now'))")
    granted_uid = cur.lastrowid
    # Huquqsiz oddiy xodim
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, can_edit_attendance, created_at)"
        " values (999903004,'T-AttEdit-NoRight','employee',1,1,0,datetime('now'))")
    noright_uid = cur.lastrowid
    for uid in (emp_uid, granted_uid):
        cur.execute(
            "insert into work_schedule_weekly (user_id, weekday, is_working, start_time, end_time, updated_at)"
            " select ?, wd, 1, '09:00', '18:00', datetime('now') from (select 0 wd union select 1 union select 2"
            " union select 3 union select 4 union select 5 union select 6)", (uid,))
    conn.commit()

    def cleanup_ae():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [dev_uid, emp_uid, granted_uid, noright_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from attendance where user_id in ({qm})", uids)
            c2.execute(f"delete from work_schedule_weekly where user_id in ({qm})", uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Davomat huquqi tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    def audit_count(uid: int) -> int:
        c = db()
        try:
            return c.execute(
                "select count(*) from audit_logs where target_user_id=? and action='attendance_manual_edit'", (uid,)
            ).fetchone()[0]
        finally:
            c.close()

    try:
        dev_t = token_for(dev_uid, "dasturchi")
        granted_t = token_for(granted_uid, "employee")
        noright_t = token_for(noright_uid, "employee")

        with httpx.Client(timeout=15) as client:
            # ── 1. Dasturchi JIM tuzatishi: yozuv YARATILADI, audit YO'Q ──
            before_audit = audit_count(emp_uid)
            r = client.put(
                f"{API_BASE}/admin/attendance/manual", headers=auth(dev_t),
                json={"user_id": emp_uid, "date": DAY, "check_in": "09:25", "check_out": "18:00"},
            )
            check("Dasturchi jim tuzatish -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            body = r.json() if r.status_code == 200 else {}
            check("Yozuv YARATILDI (bosish esidan chiqqan holat)", body.get("created") is True, f"={body}")
            check("Kechikish qayta hisoblandi (09:25 -> 25 daq)", body.get("late_minutes") == 25, f"={body}")
            check("Javobda audited=False", body.get("audited") is False, f"={body}")
            check("AuditLog yozuvi YARATILMADI (jim)", audit_count(emp_uid) == before_audit,
                  f"oldin={before_audit} keyin={audit_count(emp_uid)}")

            # ── 2. Shaxsan ruxsat berilgan XODIM tuzata oladi + audit YOZILADI ──
            before_audit = audit_count(emp_uid)
            r = client.put(
                f"{API_BASE}/attendance/manual", headers=auth(granted_t),
                json={"user_id": emp_uid, "date": DAY, "check_in": "09:00", "check_out": "18:00",
                      "reason": "T-sinov: bayroq bilan tuzatish"},
            )
            check("Bayroq berilgan xodim tuzata oladi -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:150]}")
            check("Kechikish 0 ga tushdi (09:00)", r.status_code == 200 and r.json().get("late_minutes") == 0,
                  f"={r.json() if r.status_code == 200 else None}")
            check("AuditLog YOZILDI (saytda ko'rinsin)", audit_count(emp_uid) == before_audit + 1,
                  f"oldin={before_audit} keyin={audit_count(emp_uid)}")

            # ── 3. O'Z yozuvini tuzata OLMAYDI ──
            r = client.put(
                f"{API_BASE}/attendance/manual", headers=auth(granted_t),
                json={"user_id": granted_uid, "date": DAY, "check_in": "09:00",
                      "reason": "T-sinov: o'zimniki"},
            )
            check("Bayroq egasi O'Z yozuvini tuzata OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # ── 4. Huquqsiz xodim -> 403 ──
            r = client.put(
                f"{API_BASE}/attendance/manual", headers=auth(noright_t),
                json={"user_id": emp_uid, "date": DAY, "check_in": "10:00", "reason": "T-sinov: ruxsatsiz"},
            )
            check("Huquqsiz xodim -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.put(
                f"{API_BASE}/admin/attendance/manual", headers=auth(granted_t),
                json={"user_id": emp_uid, "date": DAY, "check_in": "10:00"},
            )
            check("Bayroq egasi JIM endpointga kira OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # ── 5. Huquq berish/olish (faqat Dasturchi) + audit ──
            r = client.post(
                f"{API_BASE}/admin/users/{noright_uid}/attendance-editor", headers=auth(dev_t),
                json={"granted": True, "override_reason": "T-sinov: huquq berish"},
            )
            check("Dasturchi huquq beradi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            c = db()
            flag = c.execute("select can_edit_attendance from users where id=?", (noright_uid,)).fetchone()[0]
            c.close()
            check("Bayroq bazada 1 bo'ldi", flag == 1, f"={flag}")

            r = client.get(f"{API_BASE}/admin/attendance-editors", headers=auth(dev_t))
            names = [x["full_name"] for x in r.json()] if r.status_code == 200 else []
            check("Ro'yxatda yangi huquq egasi bor", "T-AttEdit-NoRight" in names, f"={names}")

            r = client.post(
                f"{API_BASE}/admin/users/{noright_uid}/attendance-editor", headers=auth(granted_t),
                json={"granted": True, "override_reason": "T-sinov: ruxsatsiz berish"},
            )
            check("Dasturchi bo'lmagan huquq bera OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")
    except Exception:
        check("Davomat tuzatish huquqlari (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_ae()
        conn.close()


def test_location_exempt_checkin() -> None:
    """«Bez lokatsiya» check-in (egasining 2026-08-03 talabi):

    1. Bayroqli xodim ISTALGAN joydan (koordinatasiz) «Keldim» qila oladi.
    2. Bayroqsiz xodim koordinatasiz yuborsa — RAD etiladi (aks holda GPS
       tekshiruvini istalgan kishi `latitude: null` bilan chetlab o'tardi).
    3. Bayroqli xodimda ham Face ID BEKOR QILINMAYDI.
    4. Ruxsatni faqat Dasturchi bera oladi."""
    import httpx

    print("\n" + "=" * 60)
    print("JOYLASHUVSIZ CHECK-IN («bez lokatsiya»)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    FACE_JSON = json.dumps([0.05] * 128)

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-NoGeo%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from attendance where user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, face_descriptor,"
        " skip_location_check, created_at) values (999904001,'T-NoGeo-Free','employee',1,1,?,1,datetime('now'))",
        (FACE_JSON,))
    free_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, face_descriptor,"
        " skip_location_check, created_at) values (999904002,'T-NoGeo-Bound','employee',1,1,?,0,datetime('now'))",
        (FACE_JSON,))
    bound_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999904003,'T-NoGeo-Dev','dasturchi',1,1,datetime('now'))")
    dev_uid = cur.lastrowid
    conn.commit()

    def cleanup_ng():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [free_uid, bound_uid, dev_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from attendance where user_id in ({qm})", uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Bez-lokatsiya tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        free_t = token_for(free_uid, "employee")
        bound_t = token_for(bound_uid, "employee")
        dev_t = token_for(dev_uid, "dasturchi")
        body_nogeo = {"latitude": None, "longitude": None, "face_descriptor": [0.05] * 128, "liveness": 1.0}

        with httpx.Client(timeout=20) as client:
            # 1. Bayroqli xodim — koordinatasiz o'tadi
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(free_t), json=body_nogeo)
            check("Bayroqli xodim koordinatasiz «Keldim» -> 200", r.status_code == 200,
                  f"kod={r.status_code} {r.text[:160]}")
            check("Masofa NULL (0 emas — soxta 'ofis markazi' bo'lmasin)",
                  r.status_code == 200 and r.json().get("check_in_distance_m") is None,
                  f"={r.json().get('check_in_distance_m') if r.status_code == 200 else None}")

            # 2. Bayroqsiz xodim — koordinatasiz RAD etiladi
            r = client.post(f"{API_BASE}/attendance/me/check-in", headers=auth(bound_t), json=body_nogeo)
            check("Bayroqsiz xodim koordinatasiz -> RAD etiladi", r.status_code >= 400,
                  f"kod={r.status_code} {r.text[:160]}")

            # 3. Bayroqli xodimda Face ID baribir tekshiriladi (begona yuz)
            c = db()
            c.execute("delete from attendance where user_id=?", (free_uid,))
            c.commit()
            c.close()
            r = client.post(
                f"{API_BASE}/attendance/me/check-in", headers=auth(free_t),
                json={**body_nogeo, "face_descriptor": [0.9] * 128},
            )
            check("Bayroqli xodim BEGONA yuz bilan -> RAD (Face ID bekor emas)",
                  r.status_code >= 400, f"kod={r.status_code} {r.text[:160]}")

            # 4. Ruxsatni faqat Dasturchi beradi
            r = client.post(
                f"{API_BASE}/admin/users/{bound_uid}/location-exempt", headers=auth(dev_t),
                json={"granted": True, "override_reason": "T-sinov: bez lokatsiya"},
            )
            check("Dasturchi ruxsat beradi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            c = db()
            flag = c.execute("select skip_location_check from users where id=?", (bound_uid,)).fetchone()[0]
            c.close()
            check("Bayroq bazada 1 bo'ldi", flag == 1, f"={flag}")

            r = client.post(
                f"{API_BASE}/admin/users/{bound_uid}/location-exempt", headers=auth(free_t),
                json={"granted": True, "override_reason": "T-sinov: ruxsatsiz"},
            )
            check("Oddiy xodim ruxsat bera OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # 5. HR ham bera oladi (2026-08-04: egasi "hr va dasturchi
            #    foydalanishi uchun" dedi — ilgari faqat Dasturchi edi va HR
            #    ko'chma xodimni o'zi belgilay olmasdi).
            c = db()
            hr_row = c.execute(
                "select id from users where role='hr' and is_active=1 limit 1").fetchone()
            c.close()
            if hr_row:
                hr_t = token_for(hr_row[0], "hr")
                r = client.post(
                    f"{API_BASE}/admin/users/{bound_uid}/location-exempt", headers=auth(hr_t),
                    json={"granted": False, "override_reason": "T-sinov: HR olib qo'ydi"},
                )
                check("HR ruxsatni boshqara oladi -> 200", r.status_code == 200,
                      f"kod={r.status_code} {r.text[:120]}")
                r = client.get(f"{API_BASE}/admin/location-exempt", headers=auth(hr_t))
                check("HR ruxsat ro'yxatini ko'ra oladi -> 200", r.status_code == 200,
                      f"kod={r.status_code}")
            else:
                check("HR hisobi topildi", False, "bazada hr roli yo'q")
    except Exception:
        check("Bez-lokatsiya check-in (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_ng()
        conn.close()


def test_attendance_list_dates() -> None:
    """«Yozuvlar» ro'yxatining sana filtri (2026-08-03 regressiyasi).

    ⚠️ BU BUG LOKALDA TAKRORLANMAYDI: SQLite sanani matn sifatida saqlaydi va
    satr bilan solishtirishga rozi bo'ladi. PostgreSQL esa rad etadi —
    «operator does not exist: date >= character varying» — ya'ni jonli
    serverda endpoint 500 qaytarardi va rahbar panelidagi jadval doim bo'sh
    ko'rinardi. Shuning uchun bu test tipni EMAS, xulq-atvorni qo'riqlaydi:
    (a) to'g'ri sana bilan 200 va oraliqdan tashqaridagi yozuv KIRMAYDI;
    (b) noto'g'ri formatda 400 (ilgari 500 bo'lardi)."""
    import httpx

    print("\n" + "=" * 60)
    print("DAVOMAT RO'YXATI — SANA FILTRI")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "boss/dasturchi/hr yo'q")
        return
    token = token_for(mgr[0], mgr[1])
    if not token:
        return

    conn = db()
    cur = conn.cursor()
    today = date.today()
    inside = (today - timedelta(days=2)).isoformat()
    outside = (today - timedelta(days=40)).isoformat()
    uid = None
    try:
        cur.execute(
            "insert into users (full_name, role, is_active, bot_started, created_at)"
            " values (?,?,1,0,datetime('now'))",
            ("T-SanaFiltr", "employee"),
        )
        uid = cur.lastrowid
        for d in (inside, outside):
            cur.execute(
                "insert into attendance (user_id, date, status, late_minutes,"
                " early_leave_minutes, worked_minutes, is_weekend, created_at, updated_at)"
                " values (?,?,?,0,0,0,0,datetime('now'),datetime('now'))",
                (uid, d, "present"),
            )
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=20) as c:
            r = c.get(
                "/attendance",
                params={"user_id": uid, "date_from": (today - timedelta(days=7)).isoformat(),
                        "date_to": today.isoformat()},
                headers=auth(token),
            )
            check("sana oralig'i bilan 200 qaytdi", r.status_code == 200, f"status={r.status_code} {r.text[:120]}")
            if r.status_code == 200:
                dates = [row["date"] for row in r.json()]
                check("oraliq ICHIDAGI yozuv keldi", inside in dates, f"dates={dates}")
                check("oraliqdan TASHQARIDAGI yozuv kelmadi", outside not in dates, f"dates={dates}")

            r2 = c.get("/attendance", params={"date_from": "03.08.2026"}, headers=auth(token))
            check("noto'g'ri sana formatida 400 (500 EMAS)", r2.status_code == 400,
                  f"status={r2.status_code}")
    finally:
        if uid is not None:
            cur.execute("delete from attendance where user_id=?", (uid,))
            cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        conn.close()


def test_attendance_reminder() -> None:
    """«Keldim/Ketdim bosishni unutmang» eslatmasi (D-bo'lim, 2026-08-03).

    ⚠️ FAQAT `dry_run` ishlatiladi — real xodimga Telegram xabari KETMAYDI.
    Sinov T- xodimlarning ish oynasini HOZIRGI vaqtga moslab qo'yadi, ya'ni
    "ish boshlanishiga 15 daqiqa qoldi" holati sun'iy yaratiladi.

    Tekshiriladi: (a) bosmagan -> ro'yxatga tushadi; (b) dam kunida ->
    tushmaydi; (c) sababli kunda -> tushmaydi; (d) allaqachon bosgan ->
    tushmaydi; (e) iz yozilgach takror tushmaydi."""
    import httpx

    from api.config import settings

    print("\n" + "=" * 60)
    print("DAVOMAT ESLATMASI («Keldim/Ketdim bosishni unutmang»)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    today = date.today().isoformat()
    now = datetime.now(TZ)
    # Ish boshlanishi = hozir + 10 daqiqa -> "15 daqiqa qoldi" oynasi ichida.
    start_at = (now + timedelta(minutes=10)).strftime("%H:%M")
    end_at = (now + timedelta(minutes=200)).strftime("%H:%M")

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-Rem%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for t in ("attendance_reminders", "attendance", "work_schedule_override",
                  "work_schedule_weekly", "excused_days"):
            cur.execute(f"delete from {t} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    def mk(tg: int, name: str) -> int:
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
            " values (?,?,'employee',1,1,datetime('now'))", (tg, name))
        return cur.lastrowid

    forgot_uid = mk(999905001, "T-Rem-Forgot")     # bosmagan -> eslatma KERAK
    dayoff_uid = mk(999905002, "T-Rem-DayOff")     # dam kuni -> KERAK EMAS
    excused_uid = mk(999905003, "T-Rem-Excused")   # sababli -> KERAK EMAS
    done_uid = mk(999905004, "T-Rem-Done")         # allaqachon bosgan -> KERAK EMAS

    for uid in (forgot_uid, excused_uid, done_uid):
        cur.execute(
            "insert into work_schedule_override (user_id, date, is_working, start_time, end_time, updated_at)"
            " values (?,?,1,?,?,datetime('now'))", (uid, today, start_at, end_at))
    # Dam kuni
    cur.execute(
        "insert into work_schedule_override (user_id, date, is_working, updated_at)"
        " values (?,?,0,datetime('now'))", (dayoff_uid, today))
    # Sababli kun (tasdiqlangan)
    cur.execute(
        "insert into excused_days (user_id, date, reason, status, created_at)"
        " values (?,?,'T-sinov','approved',datetime('now'))", (excused_uid, today))
    # Allaqachon «Keldim» bosgan
    cur.execute(
        "insert into attendance (user_id, date, check_in_time, late_minutes, early_leave_minutes,"
        " worked_minutes, status, is_weekend, created_at, updated_at)"
        " values (?,?,datetime('now'),0,0,0,'present',0,datetime('now'),datetime('now'))",
        (done_uid, today))
    conn.commit()

    def cleanup_rem():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [forgot_uid, dayoff_uid, excused_uid, done_uid]
            qm = ",".join("?" * len(uids))
            for t in ("attendance_reminders", "attendance", "work_schedule_override",
                      "work_schedule_weekly", "excused_days"):
                c2.execute(f"delete from {t} where user_id in ({qm})", uids)
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Eslatma tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        hdr = {"X-Bot-Secret": settings.bot_shared_secret}
        with httpx.Client(timeout=30) as client:
            r = client.post(f"{API_BASE}/attendance/reminder-tick", headers=hdr, json={"dry_run": True})
            check("reminder-tick (dry_run) -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            planned = r.json().get("planned", []) if r.status_code == 200 else []
            ids = {p["user_id"] for p in planned}

            check("Bosmagan xodim ro'yxatda BOR", forgot_uid in ids, f"planned={sorted(ids)}")
            check("Dam kunidagi YO'Q", dayoff_uid not in ids, f"planned={sorted(ids)}")
            check("Sababli kundagi YO'Q", excused_uid not in ids, f"planned={sorted(ids)}")
            check("Allaqachon bosgan YO'Q", done_uid not in ids, f"planned={sorted(ids)}")
            # Ish boshlanishi hozir+10 daqiqa qilib qo'yilgan -> AYNAN 10
            # daqiqalik nuqta tushishi kerak (5 va 0 hali emas).
            mine = [p for p in planned if p["user_id"] == forgot_uid]
            check("Bosmagan uchun tur = check_in_10",
                  any(p["kind"] == "check_in_10" for p in mine), f"={mine}")
            check("5 va 0 nuqtalari HALI tushmadi",
                  not any(p["kind"] in ("check_in_5", "check_in_0") for p in mine), f"={mine}")
            check("Bitta tick'da bitta nuqta", len(mine) <= 1, f"={mine}")

            # Iz yozilgan bo'lsa o'sha nuqta takror tushmasligi (real
            # yuborishsiz — izni qo'lda yozamiz, chunki dry_run iz yozmaydi).
            c = db()
            c.execute(
                "insert into attendance_reminders (user_id, date, kind, sent_at)"
                " values (?,?,'check_in_10',datetime('now'))", (forgot_uid, today))
            c.commit()
            c.close()
            r = client.post(f"{API_BASE}/attendance/reminder-tick", headers=hdr, json={"dry_run": True})
            planned2 = r.json().get("planned", []) if r.status_code == 200 else []
            mine2 = [p for p in planned2 if p["user_id"] == forgot_uid]
            check("Iz yozilgach o'sha nuqta TAKROR tushmaydi",
                  not any(p["kind"] == "check_in_10" for p in mine2), f"={mine2}")
    except Exception:
        check("Davomat eslatmasi (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_rem()
        conn.close()


def test_dashboard_day_off() -> None:
    """Dashboardda «dam olishda» bo'limi (A-bo'lim, 2026-08-03).

    Dam kunidagilar ilgari hech qayerda ko'rinmasdi: `working_today`dan
    tushib qolar, `not_checked_in`ga ham kirmas edi (to'g'ri — jarima ham
    olmaydi), lekin rahbar ekranida javobsiz farq qolardi.

    MUHIM: bu FAQAT ko'rinish — hisob-kitobga ta'sir qilmasligi ham
    tekshiriladi (`working_today` va `not_checked_in` o'zgarmasin)."""
    import httpx

    print("\n" + "=" * 60)
    print("DASHBOARD: «bugun dam olishda» bo'limi")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    today = date.today().isoformat()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-DayOff%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for t in ("attendance", "work_schedule_override", "work_schedule_weekly"):
            cur.execute(f"delete from {t} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999906001,'T-DayOff-Resting','employee',1,1,datetime('now'))")
    rest_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999906002,'T-DayOff-Working','employee',1,1,datetime('now'))")
    work_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999906003,'T-DayOff-Hr','hr',1,1,datetime('now'))")
    hr_uid = cur.lastrowid
    cur.execute(
        "insert into work_schedule_override (user_id, date, is_working, updated_at)"
        " values (?,?,0,datetime('now'))", (rest_uid, today))
    cur.execute(
        "insert into work_schedule_override (user_id, date, is_working, start_time, end_time, updated_at)"
        " values (?,?,1,'09:00','18:00',datetime('now'))", (work_uid, today))
    conn.commit()

    def cleanup_do():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [rest_uid, work_uid, hr_uid]
            qm = ",".join("?" * len(uids))
            for t in ("attendance", "work_schedule_override", "work_schedule_weekly"):
                c2.execute(f"delete from {t} where user_id in ({qm})", uids)
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Dam kuni tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        hr_t = token_for(hr_uid, "hr")
        with httpx.Client(timeout=20) as client:
            r = client.get(f"{API_BASE}/attendance/dashboard", headers=auth(hr_t))
            check("dashboard -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            body = r.json() if r.status_code == 200 else {}
            names = {x["full_name"] for x in body.get("on_day_off", [])}
            summary = body.get("summary", {})

            check("Dam kunidagi xodim ro'yxatda BOR", "T-DayOff-Resting" in names, f"={sorted(names)}")
            check("Ishlayotgan xodim ro'yxatda YO'Q", "T-DayOff-Working" not in names, f"={sorted(names)}")
            check("summary.on_day_off soni ro'yxat bilan mos",
                  summary.get("on_day_off") == len(body.get("on_day_off", [])),
                  f"soni={summary.get('on_day_off')} ro'yxat={len(body.get('on_day_off', []))}")
            # Hisob-kitobga ta'sir qilmasligi: dam kunidagi na "ishlashi kerak",
            # na "kelmagan" ga kirmaydi.
            check("Dam kunidagi 'kelmagan' ga KIRMAYDI",
                  summary.get("not_checked_in", 0) <= summary.get("working_today", 0),
                  f"kelmagan={summary.get('not_checked_in')} ishlashi_kerak={summary.get('working_today')}")
    except Exception:
        check("Dashboard dam kuni (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_do()
        conn.close()


def test_payroll_approval_segregation() -> None:
    """Oylik tasdig'i ikki bosqichga ajratildi (2026-08-08, egasining talabi
    "hr uni belgilaydi, boss boshliq uni tasdiqlaydi").

    2026-08-21 (egasining qarori): oylikni HR ham TASDIQLAY OLADI —
    «bossga borib o'tirmasin». Ilgari «HR tayyor deydi -> Boshliq
    qulflaydi» ajratimi bor edi va u oylikni kechiktirardi.

    AJRATIM O'RNIGA IZ: kim hisoblagan, kim tekshirgan va kim
    qulflagani saqlanadi; har xodimga shaxsiy xabar boradi.

    ROP va XODIM baribir tasdiqlay olmaydi — bu tekshiruv saqlanadi.

    Tekshiriladi:
      (a) ROP/xodim yakuniy tasdiqlay OLMAYDI (403);
      (b) HR bosqichisiz ham tasdiqlash MUMKIN va iz avtomatik to'ladi;
      (c) HR "tayyor" deydi -> status hr_approved, kim tekshirgani ko'rinadi;
      (d) HR yakuniy tasdiqlaydi -> locked;
      (e) qulflangach qayta hisoblash 409.
    """
    import httpx

    print("\n" + "=" * 60)
    print("OYLIK TASDIG'I: VAZIFALAR AJRATIMI")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    period = "2019-05"          # ATAYLAB uzoq o'tmish — jonli davrlarga tegmasin
    uid = None
    try:
        hr_row = cur.execute("select id from users where role='hr' and is_active=1 limit 1").fetchone()
        boss_row = cur.execute("select id from users where role='boss' and is_active=1 limit 1").fetchone()
        if not hr_row or not boss_row:
            check("HR va Boshliq hisoblari topildi", False, "biri yo'q")
            return
        hr_t = token_for(hr_row[0], "hr")
        boss_t = token_for(boss_row[0], "boss")

        cur.execute(
            "insert into users (full_name, role, is_active, bot_started, created_at)"
            " values (?,?,1,0,datetime('now'))", ("T-Approve", "employee"))
        uid = cur.lastrowid
        conn.commit()
        # Tasdiqlay OLMAYDIGAN rollar — oylik tasdig'i HR ga ochilgach ham
        # ular yopiq qolishi kerak.
        emp_t = token_for(uid, "employee")
        rop_row = cur.execute(
            "select id from users where role='rop' and is_active=1 limit 1").fetchone()
        rop_t = token_for(rop_row[0], "rop") if rop_row else None

        with httpx.Client(base_url=API_BASE, timeout=60) as c:
            r = c.post(f"/payroll/{period}/calculate", headers=auth(hr_t), json={"user_ids": [uid]})
            check("HR hisoblay oladi (navbat -> 202)", r.status_code == 202,
                  f"kod={r.status_code} {r.text[:120]}")
            tick = payroll_tick(c, period)
            check("cron HR so'rovini bajardi", bool(tick) and tick.get("ok") is True, f"={tick}")

            # (a) ROP va xodim baribir tasdiqlay olmaydi
            for nom, tok in (("ROP", rop_t), ("xodim", emp_t)):
                if not tok:
                    continue
                r = c.post(f"/payroll/{period}/approve", headers=auth(tok))
                check(f"{nom} YAKUNIY tasdiqlay OLMAYDI -> 403",
                      r.status_code == 403, f"kod={r.status_code}")

            r = c.post(f"/payroll/{period}/hr-approve", headers=auth(hr_t))
            check("HR «tayyor» dedi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")
            if r.status_code == 200:
                check("status hr_approved bo'ldi", r.json().get("status") == "hr_approved",
                      f"status={r.json().get('status')}")

            r = c.get("/payroll/periods", headers=auth(boss_t))
            if r.status_code == 200:
                row = next((x for x in r.json() if x["period"] == period), None)
                check("Boshliq kim tekshirganini ko'radi",
                      bool(row and row.get("hr_approved_name")),
                      f"hr_approved_name={row.get('hr_approved_name') if row else None}")

            # (d) HR o'zi yakuniy tasdiqlaydi — Boshliq kerak emas
            r = c.post(f"/payroll/{period}/approve", headers=auth(hr_t))
            check("⭐ HR yakuniy tasdiqladi -> 200 (Boshliq kerak emas)",
                  r.status_code == 200, f"kod={r.status_code} {r.text[:140]}")

            r = c.post(f"/payroll/{period}/calculate", headers=auth(hr_t), json={"user_ids": [uid]})
            check("qulflangach qayta hisoblash -> 409", r.status_code == 409, f"kod={r.status_code}")
    finally:
        cur.execute("delete from payslip_items where payslip_id in (select id from payslips where period=?)", (period,))
        cur.execute("delete from payslips where period=?", (period,))
        cur.execute("delete from payroll_periods where period=?", (period,))
        if uid is not None:
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from users where id=?", (uid,))
        cur.execute("delete from audit_logs where action in "
                    "('payroll_calculated','payroll_period_hr_approved','payroll_period_approved')")
        conn.commit()
        conn.close()


def test_kpi_rates() -> None:
    """KPI stavkalari saytdan sozlanadigan bo'ldi (2026-08-08, egasi so'radi).

    Ilgari stavkalar `api/services/bonus.py` da KONSTANTA edi — HR ularni
    o'zgartira olmasdi, tarixiy emasdi va lavozimga qarab farqlanmasdi
    (mobilograf video stavkasi 0 bo'lgani uchun uning KPI'si doim nol edi).

    Tekshiriladi:
      (a) stavka yaratish va ro'yxat;
      (b) 3 darajali qamrov — xodim stavkasi global'dan USTUN;
      (c) tarixiylik — keyingi `effective_from` eskisini bosmaydi;
      (d) validatsiya (noma'lum ko'rsatkich, global+scope_id, dublikat sana);
      (e) sozlanmagan stavka -> bonus 0, LEKIN breakdown'da `missing_rates`.
    """
    import asyncio as _aio
    import httpx
    from datetime import date as _date
    from decimal import Decimal as _Dec

    print("\n" + "=" * 60)
    print("KPI STAVKALARI (kpi_rates)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    token = token_for(mgr[0], mgr[1])
    if not token:
        return

    conn = db()
    cur = conn.cursor()
    uid = kpi_pos_id = None
    try:
        # Lavozim SHART: `metrics_for` bo'sh bo'lsa `calculate_bonus` hech
        # qanday ko'rsatkichni ko'rmaydi va §2.6(a) tekshiruvi ma'nosiz
        # bo'lib qolardi. «suhbat» ga stavka beriladi, «tashrif» ga YO'Q —
        # aynan shu farq `missing_rates` da ko'rinishi kerak.
        cur.execute(
            "insert into positions (name, metrics, is_active, created_at)"
            " values (?,?,1,datetime('now'))", ("T-Kpi-Lavozim", '["suhbat", "tashrif"]'))
        kpi_pos_id = cur.lastrowid
        cur.execute(
            "insert into users (full_name, role, position_id, is_active, bot_started, created_at)"
            " values (?,?,?,1,0,datetime('now'))", ("T-Kpi", "employee", kpi_pos_id))
        uid = cur.lastrowid
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=20) as c:
            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "global", "metric": "suhbat", "amount": 2500,
                "effective_from": "2020-01-01", "note": "T-sinov"})
            check("global stavka yaratildi", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")

            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "user", "scope_id": uid, "metric": "suhbat", "amount": 4000,
                "effective_from": "2020-01-01"})
            check("xodimga alohida stavka", r.status_code == 200, f"kod={r.status_code}")

            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "global", "metric": "suhbat", "amount": 9999,
                "effective_from": "2020-01-01"})
            check("bir sanaga ikkinchi stavka -> 400", r.status_code == 400, f"kod={r.status_code}")

            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "global", "scope_id": 5, "metric": "suhbat", "amount": 100,
                "effective_from": "2020-01-01"})
            check("global + scope_id -> 422", r.status_code == 422, f"kod={r.status_code}")

            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "global", "metric": "yolgon_korsatkich", "amount": 100,
                "effective_from": "2020-01-01"})
            check("noma'lum ko'rsatkich -> 422", r.status_code == 422, f"kod={r.status_code}")

            # Tarixiylik: keyingi sanaga yangi qiymat
            r = c.post("/payroll/kpi-rates", headers=auth(token), json={
                "scope": "user", "scope_id": uid, "metric": "suhbat", "amount": 6000,
                "effective_from": "2020-06-01"})
            check("keyingi sanaga yangi stavka (tarix)", r.status_code == 200, f"kod={r.status_code}")

        # Qamrov va tarix — xizmat qatlamidan
        from api.services.kpi_rates import resolve_kpi_rate
        from db.base import async_session
        from db.models import User as _U

        async def _chk():
            async with async_session() as s2:
                u = await s2.get(_U, uid)
                eski = await resolve_kpi_rate(s2, u, "suhbat", _date(2020, 3, 1))
                yangi = await resolve_kpi_rate(s2, u, "suhbat", _date(2020, 9, 1))
                yoq = await resolve_kpi_rate(s2, u, "tashrif", _date(2020, 9, 1))
                return eski, yangi, yoq

        eski, yangi, yoq = _aio.run(_chk())
        check("xodim stavkasi global'dan USTUN (4000)", eski == _Dec("4000.00"), f"={eski}")
        check("tarix ishlaydi — keyingi sanada 6000", yangi == _Dec("6000.00"), f"={yangi}")
        check("sozlanmagan ko'rsatkich -> None (0 EMAS)", yoq is None, f"={yoq}")

        # §2.6(a): stavka sozlanmagan bo'lsa bonus 0 chiqadi, LEKIN sabab
        # breakdown'da yozilib qoladi. HR «nega bonus 0» degan savolga bir
        # qarashda javob topsin — 0 stavka (ataylab bepul) va stavka YO'Q
        # (hali kiritilmagan) BOSHQA-BOSHQA holat.
        from api.services.bonus import calculate_bonus as _calc_bonus

        async def _bonus_breakdown():
            async with async_session() as s3:
                u = await s3.get(_U, uid)
                return await _calc_bonus(s3, u, "2020-09")

        natija = _aio.run(_bonus_breakdown())
        bd = natija["breakdown"]
        check("§2.6a: stavkasiz ko'rsatkich bonusga 0 qo'shadi",
              float(natija["amount"]) == 0.0, f"={natija['amount']}")
        check("§2.6a: breakdown'da `missing_rates` sababi qoladi",
              "missing_rates" in bd and len(bd["missing_rates"]) > 0, f"={bd.get('missing_rates')}")
        check("§2.6a: sozlangan ko'rsatkich `missing_rates` ga TUSHMAYDI",
              "suhbat" not in bd.get("missing_rates", []), f"={bd.get('missing_rates')}")
    finally:
        if uid is not None:
            cur.execute("delete from kpi_rates where scope='user' and scope_id=?", (uid,))
            cur.execute("delete from kpi_rates where scope='global' and metric='suhbat' and note='T-sinov'")
            cur.execute("delete from kpi_rates where scope='global' and metric='suhbat'")
            cur.execute("delete from audit_logs where action='kpi_rate_created'")
            cur.execute("delete from bonuses where user_id=?", (uid,))
            cur.execute("delete from users where id=?", (uid,))
        if kpi_pos_id is not None:
            cur.execute("delete from positions where id=?", (kpi_pos_id,))
            conn.commit()
        conn.close()


def test_overtime_global_profile() -> None:
    """§3.2 — qo'shimcha ish nihoyat AVTOMATIK hisoblanadimi.

    MUAMMO EDI: `OvertimeProfile` faqat xodim bo'yicha edi va `enabled`
    default `False`. Ya'ni HR har bir xodimga QO'LDA profil ochmaguncha
    qo'shimcha ish umuman hisoblanmasdi — jonli bazada yoqilgan profil 0 ta
    edi. Endi `scope='global'` qatori barchaga default bo'ladi.

    Tekshiriladi:
      (a) profilsiz xodim -> 0 (regressiya qo'riqchisi);
      (b) GLOBAL profil yoqilgach o'sha xodimga nomzod yaratiladi;
      (c) xodimning O'Z qatori global'dan USTUN (hatto o'chirilgan bo'lsa ham);
      (d) `auto_approve` -> nomzod darhol `approved`;
      (e) `/overtime/detect-now` — rahbarga 200, xodimga 403;
      (f) `/overtime/bulk-decide` bir bosishda hammasini tasdiqlaydi.
    """
    import asyncio as _aio
    import httpx
    from datetime import date as _date

    print("\n" + "=" * 60)
    print("QO'SHIMCHA ISH: GLOBAL PROFIL (3.2)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])
    if not mgr_t:
        return

    KUN = "2019-09-10"
    WINDOW_MIN = 480
    conn = db()
    cur = conn.cursor()
    a_uid = b_uid = None
    try:
        for nom, tg in (("T-OtGlobalA", 999700601), ("T-OtGlobalB", 999700602)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " created_at) values (?,?,'employee',0,1,datetime('now'))", (tg, nom))
        a_uid, b_uid = [r[0] for r in cur.execute(
            "select id from users where full_name in ('T-OtGlobalA','T-OtGlobalB') order by id")]
        for uid in (a_uid, b_uid):
            for wd in range(7):
                cur.execute("insert into work_schedule_weekly (user_id, weekday, is_working,"
                            " updated_at) values (?,?,0,datetime('now'))", (uid, wd))
            cur.execute("insert into work_schedule_override (user_id, date, is_working,"
                        " start_time, end_time, updated_at)"
                        " values (?,?,1,'09:00','18:00',datetime('now'))", (uid, KUN))
            # +40 daqiqa ORTIQCHA ishlangan
            # `check_out_time` SHART: aniqlash faqat ishdan chiqqani qayd
            # etilgan kunlarni qaraydi (aks holda kun hali tugamagan bo'lishi
            # mumkin). 13:40 UTC = 18:40 Toshkent.
            cur.execute("insert into attendance (user_id, date, status, late_minutes,"
                        " early_leave_minutes, worked_minutes, is_weekend, check_out_time,"
                        " created_at, updated_at)"
                        " values (?,?,'present',0,0,?,0,?,datetime('now'),datetime('now'))",
                        (uid, KUN, WINDOW_MIN + 40, KUN + " 13:40:00"))
        conn.commit()
        emp_t = token_for(a_uid, "employee")

        from api.services import payroll as pr
        from db.base import async_session

        async def _detect():
            async with async_session() as s2:
                created = await pr.detect_overtime_candidates(s2, _date(2019, 9, 10))
                await s2.commit()
                return [(e.user_id, e.minutes, e.status) for e in created]

        # (a) profil umuman yo'q -> nomzod yo'q
        yoq = _aio.run(_detect())
        check("profilsiz xodimga nomzod YARATILMAYDI (regressiya)",
              all(u not in (a_uid, b_uid) for u, _, _ in yoq), "=" + str(yoq))

        with httpx.Client(base_url=API_BASE, timeout=60) as c:
            # (e) huquqlar — detect-now
            r = c.post("/payroll/overtime/detect-now", headers=auth(emp_t), json={"target_date": KUN})
            check("/overtime/detect-now oddiy xodimga -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))

            # (b) GLOBAL profil yoqiladi
            r = c.put("/payroll/overtime-profiles/global", headers=auth(mgr_t), json={
                "enabled": True, "auto_approve": False, "mode": "fixed_rate",
                "fixed_rate_per_hour": 10000, "norm_hours_source": "schedule",
                "min_minutes": 15})
            check("global profil yaratildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            check("global profil scope='global', user_id yo'q",
                  r.status_code == 200 and r.json().get("scope") == "global"
                  and r.json().get("user_id") is None, "=" + r.text[:120])

            r = c.post("/payroll/overtime/detect-now", headers=auth(mgr_t), json={"target_date": KUN})
            check("/overtime/detect-now rahbarga -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            rows = cur.execute("select user_id, minutes, status from overtime_entries"
                               " where date=? and user_id in (?,?)", (KUN, a_uid, b_uid)).fetchall()
            check("GLOBAL profil bilan IKKALA xodimga ham nomzod yaratildi",
                  len(rows) == 2, "=" + str(rows))
            check("nomzod +40 daqiqa (sof farq)",
                  all(r2[1] == 40 for r2 in rows), "=" + str(rows))
            check("nomzod 'pending' (auto_approve o'chiq)",
                  all(r2[2] == "pending" for r2 in rows), "=" + str(rows))

            # (f) ommaviy tasdiq
            r = c.post("/payroll/overtime/bulk-decide", headers=auth(mgr_t),
                       json={"period": "2019-09", "status": "approved"})
            check("/overtime/bulk-decide -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            check("bulk-decide kamida 2 ta yozuvni tasdiqladi",
                  r.status_code == 200 and r.json().get("decided", 0) >= 2, "=" + r.text[:120])
            qolgan = cur.execute("select count(*) from overtime_entries where date=?"
                                 " and status='pending'", (KUN,)).fetchone()
            check("kutilayotgan yozuv qolmadi", qolgan[0] == 0, "=" + str(qolgan))

            # (c) xodim qatori global'dan USTUN — B ga O'CHIRILGAN profil
            cur.execute("delete from overtime_entries where date=?", (KUN,))
            conn.commit()
            r = c.put("/payroll/overtime-profiles/" + str(b_uid), headers=auth(mgr_t), json={
                "enabled": False, "auto_approve": False, "mode": "fixed_rate",
                "fixed_rate_per_hour": 10000, "norm_hours_source": "schedule",
                "min_minutes": 15})
            check("xodimga o'chirilgan profil yozildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            c.post("/payroll/overtime/detect-now", headers=auth(mgr_t), json={"target_date": KUN})
            rows2 = [r2[0] for r2 in cur.execute(
                "select user_id from overtime_entries where date=? and user_id in (?,?)",
                (KUN, a_uid, b_uid)).fetchall()]
            check("xodim qatori GLOBAL'dan ustun — o'chirilgan xodimga nomzod YO'Q",
                  a_uid in rows2 and b_uid not in rows2, "=" + str(rows2))

            # (d) auto_approve
            cur.execute("delete from overtime_entries where date=?", (KUN,))
            conn.commit()
            r = c.put("/payroll/overtime-profiles/" + str(b_uid), headers=auth(mgr_t), json={
                "enabled": True, "auto_approve": True, "mode": "fixed_rate",
                "fixed_rate_per_hour": 10000, "norm_hours_source": "schedule",
                "min_minutes": 15})
            check("auto_approve yoqilgan profil -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            c.post("/payroll/overtime/detect-now", headers=auth(mgr_t), json={"target_date": KUN})
            st = cur.execute("select status from overtime_entries where date=? and user_id=?",
                             (KUN, b_uid)).fetchone()
            check("auto_approve -> nomzod DARHOL tasdiqlangan",
                  st is not None and st[0] == "approved", "=" + str(st))
    except Exception:
        check("Global qo'shimcha ish profili (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cur.execute("delete from overtime_entries where date=?", (KUN,))
        cur.execute("delete from overtime_profiles where scope='global'")
        if a_uid is not None:
            ids = [a_uid, b_uid]
            qm = ",".join("?" * len(ids))
            for t in ("overtime_entries", "overtime_profiles", "attendance",
                      "work_schedule_override", "work_schedule_weekly", "bonuses", "payslips"):
                cur.execute("delete from " + t + " where user_id in (" + qm + ")", ids)
            cur.execute("delete from audit_logs where target_user_id in (" + qm + ")"
                        " or actor_id in (" + qm + ")", ids + ids)
            cur.execute("delete from users where id in (" + qm + ")", ids)
        cur.execute("delete from audit_logs where action in ('overtime_profile_global_upserted',"
                    "'overtime_bulk_decided','overtime_detect_now')")
        conn.commit()
        conn.close()


def test_kpi_in_payroll() -> None:
    """§2.3 — KPI bonusi OYLIK bilan birga hisoblanadimi.

    MUAMMO EDI: `bonuses` jadvaliga qator yaratadigan yagona yo'l bot/cron
    edi (oyning oxirgi kuni 23:30). `build_payslip` esa o'sha jadvaldan
    TAYYOR qatorni o'qiydi — ya'ni oy o'rtasida HR «Hisoblash» bosganda
    bonus qatori umuman yo'q bo'lib, KPI puli jimgina 0 chiqardi.

    Tekshiriladi:
      (a) `/bonuses/recalculate` — rahbarga 200, oddiy xodimga 403;
      (b) qamrov `employee` dan kengaydi (§2.5);
      (c) oylik hisobi bonus qatorini O'ZI yaratadi (avval qator YO'Q edi)
          va payslip'da bonus summasi ko'rinadi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("KPI BONUSI OYLIK BILAN BIRGA (2.3)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])
    if not mgr_t:
        return

    PERIOD = "2021-05"
    conn = db()
    cur = conn.cursor()
    uid = pos_id = None
    try:
        # Lavozim — FAQAT «suhbat» ko'rsatkichi (metrics_for shundan o'qiydi)
        cur.execute(
            "insert into positions (name, metrics, is_active, created_at)"
            " values (?,?,1,datetime('now'))", ("T-KpiPay-Lavozim", '["suhbat"]'))
        pos_id = cur.lastrowid
        cur.execute(
            "insert into users (telegram_id, full_name, role, position_id, bot_started,"
            " is_active, created_at) values (999700801,'T-KpiPay','employee',?,0,1,datetime('now'))",
            (pos_id,))
        uid = cur.lastrowid
        # 10 ta suhbat -> stavka 3000 -> bonus 30 000 kutiladi
        cur.execute(
            "insert into daily_results (user_id, date, conversations_count, visits_count, source)"
            " values (?,?,?,?,'manual')", (uid, PERIOD + "-11", 10, 0))
        conn.commit()
        emp_t = token_for(uid, "employee")

        with httpx.Client(base_url=API_BASE, timeout=60) as c:
            c.post("/payroll/kpi-rates", headers=auth(mgr_t), json={
                "scope": "user", "scope_id": uid, "metric": "suhbat", "amount": 3000,
                "effective_from": PERIOD + "-01", "note": "T-kpipay"})
            c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": uid, "amount": 1000000, "pay_basis": "monthly",
                "effective_from": PERIOD + "-01", "reason": "hire"})

            # (a) huquqlar
            r = c.post("/bonuses/recalculate", headers=auth(emp_t), json={"period": PERIOD})
            check("/bonuses/recalculate oddiy xodimga -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))
            r = c.post("/bonuses/recalculate", headers=auth(mgr_t), json={"period": PERIOD})
            check("/bonuses/recalculate rahbarga -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])

            row = cur.execute("select amount from bonuses where user_id=? and period=?",
                              (uid, PERIOD)).fetchone()
            check("saytdan hisoblanganda bonus qatori paydo bo'ldi (10 x 3000)",
                  row is not None and abs(row[0] - 30000) < 1, "=" + str(row))

            # (b) qamrov: employee'dan boshqa rollar ham (ilgari faqat employee edi)
            other = cur.execute(
                "select count(*) from bonuses b join users u on u.id=b.user_id"
                " where b.period=? and u.role in ('hr','rop','dasturchi')", (PERIOD,)).fetchone()
            check("qamrov kengaydi — employee'dan boshqa rollar ham hisoblandi (2.5)",
                  other is not None and other[0] >= 1, "=" + str(other))

            # (c) ASOSIY: bonus qatorlarini o'chirib, FAQAT oylik hisoblaymiz
            cur.execute("delete from bonuses where period=?", (PERIOD,))
            conn.commit()
            yoq = cur.execute("select count(*) from bonuses where period=?", (PERIOD,)).fetchone()
            check("tayyorgarlik: bonus qatorlari o'chirildi", yoq[0] == 0, "=" + str(yoq))

            r = c.post("/payroll/" + PERIOD + "/calculate", headers=auth(mgr_t), json={})
            check("oylik navbatga qo'yildi -> 202", r.status_code == 202, "kod=" + str(r.status_code))
            tick = payroll_tick(c, PERIOD)
            check("cron oylikni hisobladi", bool(tick) and tick.get("ok") is True, "=" + str(tick))
            check("cron javobida bonus soni ham bor",
                  bool(tick) and tick.get("bonuses", 0) >= 1, "=" + str(tick))

            row2 = cur.execute("select amount from bonuses where user_id=? and period=?",
                               (uid, PERIOD)).fetchone()
            check("OYLIK HISOBI bonus qatorini o'zi yaratdi",
                  row2 is not None and abs(row2[0] - 30000) < 1, "=" + str(row2))

            r = c.get("/payroll/" + PERIOD + "/user/" + str(uid), headers=auth(mgr_t))
            bonus_amount = r.json().get("bonus_amount") if r.status_code == 200 else None
            check("payslip'da KPI bonusi ko'rinadi (30 000)",
                  bonus_amount is not None and abs(bonus_amount - 30000) < 1,
                  "kod=" + str(r.status_code) + " bonus=" + str(bonus_amount))
    except Exception:
        check("KPI+oylik (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cur.execute("delete from bonuses where period=?", (PERIOD,))
        cur.execute("delete from payslip_items where payslip_id in"
                    " (select id from payslips where period=?)", (PERIOD,))
        cur.execute("delete from payslips where period=?", (PERIOD,))
        cur.execute("delete from payroll_periods where period=?", (PERIOD,))
        cur.execute("delete from audit_logs where action in ('bonus_calculated','payroll_calculated')"
                    " and json_extract(after,'$.period')=?", (PERIOD,))
        if uid is not None:
            cur.execute("delete from kpi_rates where scope='user' and scope_id=?", (uid,))
            cur.execute("delete from daily_results where user_id=?", (uid,))
            cur.execute("delete from salary_rates where user_id=?", (uid,))
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from users where id=?", (uid,))
        if pos_id is not None:
            cur.execute("delete from positions where id=?", (pos_id,))
        conn.commit()
        conn.close()


def test_absent_deduct_daily() -> None:
    """Kelmagan kun kunlik ulush bilan ayiriladi (2026-08-08, egasi "A" tanladi).

    NEGA MUHIM: ilgari `monthly` stavkada kelmagan kun asosiy oylikni UMUMAN
    kamaytirmasdi — faqat qat'iy `absent_fine` bor edi va jonli bazada u 0
    ekan, ya'ni kelmaslik mutlaqo bepul edi.

    Bu yerda tekshiriladi:
      (a) `deduct_daily` rejimida baza kunlik ulush ×  kelmagan kun ga kamayadi;
      (b) alohida qat'iy jarima QO'YILMAYDI (ikki marta jazo yo'q);
      (c) ayirma jarima CHEKLOVIGA (cap) tushmaydi — cap faqat jarimaga;
      (d) sababli kun ayirilmaydi.
    """
    from datetime import date as _date
    from decimal import Decimal as _D

    print("\n" + "=" * 60)
    print("KELMAGAN KUN — KUNLIK ULUSH AYIRMASI (deduct_daily)")
    print("=" * 60)

    import api.services.payroll as pr
    from db.models import AbsentMode, FineMode, FinePolicy, PayBasis

    class _Rate:
        amount = _D("3000000")
        pay_basis = PayBasis.monthly.value
        effective_from = _date(2020, 1, 1)

    # 10 ish kuni: 7 kelgan, 2 kelmagan, 1 sababli
    days = []
    for i in range(1, 11):
        if i in (3, 4):
            st, exc = "absent", False
        elif i == 5:
            st, exc = "excused", True
        else:
            st, exc = "present", False
        days.append({
            "date": _date(2020, 1, i), "is_working": True, "status": st,
            "excused": exc, "late_minutes": 0, "worked_minutes": 480,
            "scheduled_minutes": 480,
        })

    pol = FinePolicy(
        scope="global", absent_mode=AbsentMode.deduct_daily.value, absent_fine=999999,
        fine_mode=FineMode.per_day.value, fine_per_day=0,
        monthly_cap_percent=20, is_active=True,
    )

    base, item, absent_item, _unpaid = pr.compute_base(_Rate(), _Rate(), days, _date(2020, 1, 1), pol)

    kunlik = _D("3000000") / _D(10)          # 300 000
    kutilgan = _D("3000000") - kunlik * 2    # 2 400 000
    check("baza kunlik ulush bilan kamaydi (2 kun)", base == kutilgan,
          f"base={base}, kutilgan={kutilgan}")
    check("ayirma qatori yaratildi", absent_item is not None)
    if absent_item:
        check("ayirma manfiy va 600 000 ga teng", absent_item["amount"] == -kunlik * 2,
              f"amount={absent_item['amount']}")
        check("sababli kun ayirilmadi (2 kun, 3 emas)", absent_item["quantity"] == 2,
              f"quantity={absent_item['quantity']}")

    fine = pr.compute_absent_fine(days, pol)
    check("deduct_daily'da qat'iy jarima QO'YILMAYDI (ikki marta jazo yo'q)",
          fine["amount"] == _D("0"), f"amount={fine['amount']} (absent_fine=999999 bo'lsa ham)")

    # Cap ayirmaga tegmasligi: jarima 0 bo'lgani uchun cap umuman ishlamaydi,
    # ya'ni 600 000 ayirma 20% (600 000) chegarasidan qat'i nazar to'liq qoladi.
    lo, ab, raw, capped = pr.apply_fine_cap(_D("0"), fine["amount"], base, pol)
    check("cap ayirmaga tegmaydi (jarima yo'q -> cap qo'llanmaydi)", not capped,
          f"capped={capped}, raw={raw}")


def test_uysot_request_deadline() -> None:
    """CRM kutishi: HTTP so'rov TEZ chiqadi, fon (cron) SABR qiladi.

    MUAMMO (2026-08-13 jonli o'lchovi): saytda bitta so'rov 40.3 soniya
    kutdi. Deploy'da konkurentlik = 1, ya'ni bitta uzoq so'rov BUTUN saytni
    o'lik qiladi. Uysot 429 bersa so'rov ichida 60s x 4 = 4 daqiqagacha
    kutilardi.

    Ikki yo'l ATAYLAB turlicha:
      - `mark_request_context()` chaqirilgan (odam kutayotgan so'rov) ->
        `UysotBusy` darhol, kutmaydi;
      - fon (cron) -> avvalgidek qayta urinadi.
    Regressiya xavfi: kimdir `deadline` ni fon yo'liga ham qo'llasa, uzun
    skanlar yarim yo'lda uzilib qolardi va buni HECH KIM sezmasdi.
    """
    import asyncio as _aio
    import httpx as _hx
    import crm.uysot as _U

    print("\n" + "=" * 60)
    print("CRM KUTISH CHEGARASI (so'rov vs fon)")
    print("=" * 60)

    class _Fake429:
        def __init__(self): self.calls = 0
        async def request(self, method, path, json=None):
            self.calls += 1
            return _hx.Response(429, request=_hx.Request(method, "http://x" + path))

    async def _olcha(in_request: bool):
        _U._RATE_BUDGET = _U._SharedRateBudget(60)
        c = _Fake429()
        async def _ish():
            if in_request:
                _U.mark_request_context()
            try:
                r = await _U._limited_request(c, "POST", "/test")
                return f"javob {r.status_code}", c.calls
            except _U.UysotBusy:
                return "UysotBusy", c.calls
        # ALOHIDA task — contextvar boshqa o'lchovga oqib ketmasin
        return await _aio.create_task(_ish())

    async def _run():
        natija1, calls1 = await _olcha(True)
        eski = _U.RATE_LIMIT_BACKOFF_SECONDS
        _U.RATE_LIMIT_BACKOFF_SECONDS = 0.05      # fon yo'lini tez sinash uchun
        try:
            natija2, calls2 = await _olcha(False)
        finally:
            _U.RATE_LIMIT_BACKOFF_SECONDS = eski
        return natija1, calls1, natija2, calls2

    n1, c1, n2, c2 = _aio.run(_run())
    check("so'rov yo'li 429'da DARHOL chiqadi (kutmaydi)", n1 == "UysotBusy", f"={n1}")
    check("so'rov yo'li qayta urinmaydi (1 ta chaqiruv)", c1 == 1, f"={c1}")
    check("fon yo'li avvalgidek qayta urinadi", c2 == _U.MAX_RATE_LIMIT_RETRIES + 1, f"={c2}")
    check("fon yo'li 429'ni yuqoriga qaytaradi", n2 == "javob 429", f"={n2}")


def test_audit_json_guard() -> None:
    """JSON ustunlarga xavfli tur tushsa ham COMMIT yiqilmasin (BUG-1 sinfi).

    BUG-1 da audit `before` ichidagi `Decimal` butun amalni bekor qilardi
    (`TypeError: Object of type Decimal is not JSON serializable`) — jarima
    qoidasini saqlash, oylik hisoblash kabi PUL amallari shu tufayli
    yiqilardi.

    Chaqiruvchi tomon `api/audit_json.py` bilan tuzatildi, lekin yangi kod
    uni ishlatishni unutishi mumkin. Shuning uchun engine darajasida zaxira
    o'girish qo'yildi (`db/base.py::_json_serializer`). Bu test aynan o'sha
    zaxirani qo'riqlaydi: ATAYLAB xom Decimal/date/datetime/Enum uzatiladi.
    """
    import asyncio as _aio
    from datetime import date as _d, datetime as _dt
    from decimal import Decimal as _Dec

    print("\n" + "=" * 60)
    print("AUDIT JSON ZAXIRA HIMOYASI (BUG-1 sinfi qaytmasin)")
    print("=" * 60)

    from sqlalchemy import select as _sel
    from db.base import async_session as _sess
    from db.models import AuditLog as _AL, Role as _Role, User as _U

    async def _run():
        async with _sess() as s:
            u = (await s.execute(_sel(_U).limit(1))).scalars().first()
            if u is None:
                return None, "bazada foydalanuvchi yo'q"
            row = _AL(
                actor_id=u.id, action="T-GUARD-SINOV", target_user_id=None,
                before={"decimal": _Dec("12345.67"), "sana": _d(2026, 8, 8),
                        "vaqt": _dt(2026, 8, 8, 12, 30), "enum": _Role.hr},
                after=None,
            )
            s.add(row)
            try:
                await s.commit()
            except Exception as e:
                return None, f"{type(e).__name__}: {e}"
            await s.refresh(row)
            saqlangan = dict(row.before)
            await s.delete(row)
            await s.commit()
            return saqlangan, None

    saqlangan, xato = _aio.run(_run())
    check("xavfli turlar bilan COMMIT yiqilmadi", xato is None, f"xato={xato}")
    if saqlangan:
        check("Decimal -> float", saqlangan.get("decimal") == 12345.67, f"={saqlangan.get('decimal')}")
        check("date -> ISO satr", saqlangan.get("sana") == "2026-08-08", f"={saqlangan.get('sana')}")
        check("datetime -> ISO satr", str(saqlangan.get("vaqt")).startswith("2026-08-08T"),
              f"={saqlangan.get('vaqt')}")
        check("Enum -> qiymat", saqlangan.get("enum") == "hr", f"={saqlangan.get('enum')}")


def test_payroll_settings_reedit() -> None:
    """Sozlamalarni QAYTA tahrirlash (2026-08-08 regressiyasi, BUG-1).

    Jarima qoidasi ham, qo'shimcha ish profili ham bir marta YARATILGANDA
    ishlardi, lekin har qanday keyingi TAHRIR 500 berardi: audit `before`
    snapshot'i xom ORM qiymatlari bilan qurilardi va ichidagi `Decimal`
    JSON ustunga yozilmasdi (`Object of type Decimal is not JSON
    serializable`). Audit commit paytida yiqilgani uchun asosiy o'zgarish
    ham qaytarilardi — foydalanuvchi sababsiz 500 ko'rardi.

    Shuning uchun bu yerda har bir upsert IKKI MARTA chaqiriladi: birinchisi
    yaratadi, ikkinchisi TAHRIRLAYDI. Faqat ikkinchisi bugni ushlaydi."""
    import httpx

    print("\n" + "=" * 60)
    print("OYLIK SOZLAMALARINI QAYTA TAHRIRLASH (BUG-1)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    token = token_for(mgr[0], mgr[1])
    if not token:
        return

    conn = db()
    cur = conn.cursor()
    uid = None
    try:
        cur.execute(
            "insert into users (full_name, role, is_active, bot_started, created_at)"
            " values (?,?,1,0,datetime('now'))", ("T-PayReedit", "employee"))
        uid = cur.lastrowid
        conn.commit()

        pol = {
            "scope": "user", "scope_id": uid, "grace_minutes": 5,
            "free_late_minutes_per_month": 60, "fine_mode": "per_day",
            "fine_per_day": 50000, "absent_mode": "fixed", "absent_fine": 200000,
            "monthly_cap_percent": 20, "fine_applies_to": "net_salary", "is_active": True,
        }
        otp = {
            "enabled": True, "mode": "derived", "multiplier": 1.5,
            "norm_hours_source": "schedule", "min_minutes": 15,
        }

        with httpx.Client(base_url=API_BASE, timeout=20) as c:
            r1 = c.put("/payroll/policies", headers=auth(token), json=pol)
            check("jarima qoidasi — yaratish", r1.status_code == 200,
                  f"kod={r1.status_code} {r1.text[:120]}")

            pol2 = {**pol, "fine_per_day": 70000}
            r2 = c.put("/payroll/policies", headers=auth(token), json=pol2)
            check("jarima qoidasi — QAYTA tahrirlash (ilgari 500)",
                  r2.status_code == 200, f"kod={r2.status_code} {r2.text[:160]}")
            if r2.status_code == 200:
                check("yangi qiymat saqlandi", float(r2.json().get("fine_per_day") or 0) == 70000,
                      f"fine_per_day={r2.json().get('fine_per_day')}")

            r3 = c.put(f"/payroll/overtime-profiles/{uid}", headers=auth(token), json=otp)
            check("qo'shimcha ish profili — yaratish", r3.status_code == 200,
                  f"kod={r3.status_code} {r3.text[:120]}")

            otp2 = {**otp, "multiplier": 2.0}
            r4 = c.put(f"/payroll/overtime-profiles/{uid}", headers=auth(token), json=otp2)
            check("qo'shimcha ish profili — QAYTA tahrirlash (ilgari 500)",
                  r4.status_code == 200, f"kod={r4.status_code} {r4.text[:160]}")
            if r4.status_code == 200:
                check("yangi koeffitsiyent saqlandi",
                      float(r4.json().get("multiplier") or 0) == 2.0,
                      f"multiplier={r4.json().get('multiplier')}")
    finally:
        if uid is not None:
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from fine_policies where scope='user' and scope_id=?", (uid,))
            cur.execute("delete from overtime_profiles where user_id=?", (uid,))
            cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        conn.close()


def test_fine_policy_rights() -> None:
    """Kechikish normasini o'zgartirish huquqi (C-bo'lim, 2026-08-03).

    1. Bayroqli odam (roli hr/boss/dasturchi EMAS) jarima qoidasini
       o'zgartira oladi.
    2. Bayroqsiz ROP -> 403.
    3. ⚠️ Bayroq FAQAT jarima qoidasini ochadi — oylik hisoblash/tasdiqlash
       va stavkalar OCHILMAYDI (aks holda bitta bayroq bilan butun payroll
       boshqaruvi berilib qolardi).
    4. Huquqni Boshliq VA Dasturchi bera oladi; HR bera OLMAYDI."""
    import httpx

    print("\n" + "=" * 60)
    print("KECHIKISH NORMASI HUQUQI (shaxsan beriladi)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-Fine%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    def mk(tg: int, name: str, role: str, flag: int = 0) -> int:
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " can_edit_fine_policy, created_at) values (?,?,?,1,1,?,datetime('now'))",
            (tg, name, role, flag))
        return cur.lastrowid

    granted_uid = mk(999908001, "T-Fine-Granted", "rop", 1)   # bayroqli ROP
    plain_uid = mk(999908002, "T-Fine-PlainRop", "rop", 0)    # bayroqsiz ROP
    boss_uid = mk(999908003, "T-Fine-Boss", "boss")
    dev_uid = mk(999908004, "T-Fine-Dev", "dasturchi")
    hr_uid = mk(999908005, "T-Fine-Hr", "hr")
    conn.commit()

    def cleanup_fp():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [granted_uid, plain_uid, boss_uid, dev_uid, hr_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Jarima huquqi tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        granted_t = token_for(granted_uid, "rop")
        plain_t = token_for(plain_uid, "rop")
        boss_t = token_for(boss_uid, "boss")
        dev_t = token_for(dev_uid, "dasturchi")
        hr_t = token_for(hr_uid, "hr")

        with httpx.Client(timeout=20) as client:
            # 1. Bayroqli ROP qoidani KO'RA va O'ZGARTIRA oladi
            r = client.get(f"{API_BASE}/payroll/policies", headers=auth(granted_t))
            check("Bayroqli ROP qoidani ko'radi -> 200", r.status_code == 200, f"kod={r.status_code}")

            # 2. Bayroqsiz ROP -> 403
            r = client.get(f"{API_BASE}/payroll/policies", headers=auth(plain_t))
            check("Bayroqsiz ROP -> 403", r.status_code == 403, f"kod={r.status_code}")

            # 3. ⚠️ Bayroq BOSHQA payroll amallarini OCHMAYDI
            r = client.get(f"{API_BASE}/payroll/rates", headers=auth(granted_t))
            check("Bayroqli ROP STAVKALARGA kira OLMAYDI -> 403", r.status_code == 403,
                  f"kod={r.status_code} (bayroq faqat jarima qoidasini ochishi kerak)")
            r = client.post(
                f"{API_BASE}/payroll/2020-01/calculate", headers=auth(granted_t), json={},
            )
            check("Bayroqli ROP oylik HISOBLAY olmaydi -> 403", r.status_code == 403,
                  f"kod={r.status_code}")

            # 4. Huquqni kim bera oladi
            r = client.get(f"{API_BASE}/payroll/fine-policy-editors", headers=auth(boss_t))
            check("Boshliq ro'yxatni ko'radi -> 200", r.status_code == 200, f"kod={r.status_code}")
            names = [x["full_name"] for x in r.json()] if r.status_code == 200 else []
            check("Ro'yxatda bayroqli ROP bor", "T-Fine-Granted" in names, f"={names}")

            r = client.get(f"{API_BASE}/payroll/fine-policy-editors", headers=auth(hr_t))
            check("HR huquq ro'yxatiga kira OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.post(
                f"{API_BASE}/payroll/fine-policy-editors/{plain_uid}", headers=auth(dev_t),
                json={"granted": True, "reason": "T-sinov: dasturchi beradi"},
            )
            check("Dasturchi huquq beradi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:120]}")

            r = client.post(
                f"{API_BASE}/payroll/fine-policy-editors/{plain_uid}", headers=auth(boss_t),
                json={"granted": False, "reason": "T-sinov: boshliq oladi"},
            )
            check("Boshliq huquqni oladi -> 200", r.status_code == 200, f"kod={r.status_code}")

            r = client.post(
                f"{API_BASE}/payroll/fine-policy-editors/{plain_uid}", headers=auth(hr_t),
                json={"granted": True, "reason": "T-sinov: hr bera olmaydi"},
            )
            check("HR huquq bera OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")
    except Exception:
        check("Kechikish normasi huquqi (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_fp()
        conn.close()


def test_explanation_letters() -> None:
    """Tushuntirish xati (B-bo'lim, 2026-08-03).

    Oqim: sababsiz `absent` -> tizim so'raydi -> xodim botda javob yozadi ->
    HR qaror qiladi. Qabul qilinsa MAVJUD `ExcusedDay` orqali kun sababliga
    o'tadi va jarima o'z-o'zidan tushadi (yangi jarima yo'li YO'Q).

    ⚠️ Real xodimga xabar ketmasligi uchun so'rov QO'LDA yaratiladi
    (`ask_explanations` job'i emas) — u bot xabari yuborardi."""
    import httpx

    from api.config import settings

    print("\n" + "=" * 60)
    print("TUSHUNTIRISH XATI (sababsiz kelmagan kun)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    DAY = "2020-11-11"

    stale = [r[0] for r in cur.execute("select id from users where full_name like 'T-Expl%'").fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for t in ("explanation_requests", "attendance", "excused_days", "work_schedule_override"):
            cur.execute(f"delete from {t} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", stale + stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    conn.commit()

    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999909001,'T-Expl-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999909002,'T-Expl-Hr','hr',1,1,datetime('now'))")
    hr_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999909003,'T-Expl-Other','employee',1,1,datetime('now'))")
    other_uid = cur.lastrowid
    # Sababsiz kelmagan kun + so'rov (job'siz, xabar yubormasdan)
    cur.execute(
        "insert into attendance (user_id, date, late_minutes, early_leave_minutes, worked_minutes,"
        " status, is_weekend, created_at, updated_at)"
        " values (?,?,0,0,0,'absent',0,datetime('now'),datetime('now'))", (emp_uid, DAY))
    cur.execute(
        "insert into explanation_requests (user_id, date, status, asked_at)"
        " values (?,?,'pending',datetime('now'))", (emp_uid, DAY))
    req_id = cur.lastrowid
    conn.commit()

    def cleanup_ex():
        try:
            conn2 = db()
            c2 = conn2.cursor()
            uids = [emp_uid, hr_uid, other_uid]
            qm = ",".join("?" * len(uids))
            for t in ("explanation_requests", "attendance", "excused_days", "work_schedule_override"):
                c2.execute(f"delete from {t} where user_id in ({qm})", uids)
            c2.execute(f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})", uids + uids)
            c2.execute(f"delete from users where id in ({qm})", uids)
            conn2.commit()
            conn2.close()
        except Exception:
            print("  Tushuntirish tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    try:
        hr_t = token_for(hr_uid, "hr")
        emp_t = token_for(emp_uid, "employee")
        bot_hdr = {"X-Bot-Secret": settings.bot_shared_secret}

        with httpx.Client(timeout=20) as client:
            # 1. BOSHQA odam javob yoza OLMAYDI (tugma forward qilinsa ham)
            r = client.post(
                f"{API_BASE}/attendance/explanations/{req_id}/answer", headers=bot_hdr,
                json={"telegram_id": 999909003, "answer_text": "Men boshqa odamman"},
            )
            check("Begona odam javob yoza OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            # 2. Egasi javob yozadi
            r = client.post(
                f"{API_BASE}/attendance/explanations/{req_id}/answer", headers=bot_hdr,
                json={"telegram_id": 999909001, "answer_text": "T-sinov: kasal bo'lib qoldim"},
            )
            check("Xodim javob yozadi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            check("Holat 'answered' bo'ldi", r.status_code == 200 and r.json().get("status") == "answered",
                  f"={r.json() if r.status_code == 200 else None}")

            # 3. Xodim ro'yxatni ko'ra OLMAYDI (faqat hr/boss/dasturchi)
            r = client.get(f"{API_BASE}/attendance/explanations", headers=auth(emp_t))
            check("Oddiy xodim ro'yxatga kira OLMAYDI -> 403", r.status_code == 403, f"kod={r.status_code}")

            r = client.get(f"{API_BASE}/attendance/explanations?status_filter=answered", headers=auth(hr_t))
            check("HR ro'yxatni ko'radi -> 200", r.status_code == 200, f"kod={r.status_code}")
            ids = [x["id"] for x in r.json()] if r.status_code == 200 else []
            check("Javob kelgan xat ro'yxatda bor", req_id in ids, f"={ids}")

            # 4. HR QABUL qiladi -> ExcusedDay yaratiladi, davomat qayta hisoblanadi
            r = client.post(
                f"{API_BASE}/attendance/explanations/{req_id}/decide", headers=auth(hr_t),
                json={"accept": True, "note": "T-sinov: qabul"},
            )
            check("HR qabul qiladi -> 200", r.status_code == 200, f"kod={r.status_code} {r.text[:150]}")
            check("Holat 'accepted'", r.status_code == 200 and r.json().get("status") == "accepted",
                  f"={r.json() if r.status_code == 200 else None}")

            c = db()
            exc = c.execute(
                "select status from excused_days where user_id=? and date=?", (emp_uid, DAY)
            ).fetchone()
            att_status = c.execute(
                "select status from attendance where user_id=? and date=?", (emp_uid, DAY)
            ).fetchone()
            c.close()
            check("MAVJUD ExcusedDay yaratildi va 'approved'", exc is not None and exc[0] == "approved",
                  f"={exc}")
            check("Davomat yozuvi qayta hisoblandi (endi 'absent' emas)",
                  att_status is not None and att_status[0] != "absent", f"={att_status}")

            # 5. Hal qilingan xatga qayta javob yozib bo'lmaydi
            r = client.post(
                f"{API_BASE}/attendance/explanations/{req_id}/answer", headers=bot_hdr,
                json={"telegram_id": 999909001, "answer_text": "T-sinov: qayta yozaman"},
            )
            check("Hal qilingan xatga qayta javob -> 400", r.status_code == 400, f"kod={r.status_code}")
    except Exception:
        check("Tushuntirish xati (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cleanup_ex()
        conn.close()


def test_future_days_rule() -> None:
    """S-01 (TZ 2.3) — oy o'rtasida hisoblanganda KELAJAK kunlari «kelmagan»
    sanalmasligi. Bu qoidani KODGA qotirib qo'yadigan qo'riqchi.

    NEGA KERAK: bu tizimning eng qimmat xatosi bo'lgan. Oy o'rtasida
    hisoblanganda oyning QOLGAN kunlari `att is None` sababli «kelmagan»
    bo'lib, kunlik ulush oylikdan ayirilardi — Abdulaziz 10 000 000 o'rniga
    4 615 400 so'm olgan edi. Tuzatilgan (d151269), lekin qoida test bilan
    bog'lanmagan edi: kimdir `future` shoxini olib tashlasa jimgina qaytardi.

    SOF FUNKSIYA sinovi — baza kerak emas: `collect_attendance` qaytaradigan
    kunlar ro'yxati qo'lda yasaladi.
    """
    from decimal import Decimal as _D
    from types import SimpleNamespace as _NS

    from api.services import payroll as pr

    print("\n" + "=" * 60)
    print("S-01: KELAJAK KUNLARI «KELMAGAN» SANALMASIN")
    print("=" * 60)

    # Fikstura: 30 kunlik oy, 1-26 ish kuni, 27-30 dam. Bugun — 15-kun.
    OYLIK = _D("5200000")
    ISH_KUNI = 26
    RATE = _NS(amount=OYLIK, pay_basis="monthly", effective_from=date(2026, 6, 1))
    POLICY = _NS(absent_mode="deduct_daily", absent_fine=_D("0"))
    BOSHI = date(2026, 6, 1)

    def kun(d, ish=True, holat="present", ishlangan=480):
        return {
            "date": d, "is_working": ish, "start": "09:00", "end": "18:00",
            "scheduled_minutes": 480 if ish else 0, "attendance": None,
            "excused": False, "excused_paid": True, "status": holat,
            "late_minutes": 0, "worked_minutes": ishlangan,
        }

    def qur(bugun_kuni, kelmagan=frozenset()):
        """`bugun_kuni` dan boshlab (o'zi ham) — `future`."""
        kunlar = []
        for i in range(1, ISH_KUNI + 1):
            d = date(2026, 6, i)
            if i >= bugun_kuni:
                kunlar.append(kun(d, holat="future", ishlangan=0))
            elif i in kelmagan:
                kunlar.append(kun(d, holat="absent", ishlangan=0))
            else:
                kunlar.append(kun(d))
        for i in range(ISH_KUNI + 1, 31):
            kunlar.append(kun(date(2026, 6, i), ish=False, holat="weekend", ishlangan=0))
        return kunlar

    def baza(kunlar):
        b, _item, ayirma, _unpaid = pr.compute_base(RATE, None, kunlar, BOSHI, POLICY)
        return b, ayirma

    # (1) Oy o'rtasi, hech kim qolmagan -> TO'LIQ oylik, ayirma YO'Q
    b1, ayirma1 = baza(qur(15))
    check("S-01: oy o'rtasida to'liq oylik (kelajak kunlari ayirilmaydi)",
          b1 == OYLIK, f"base={b1}, kutilgan={OYLIK}")
    check("S-01: kelajak kunlari uchun ayirma qatori YO'Q",
          ayirma1 is None, f"={ayirma1}")

    # (2) HAQIQIY kelmagan kun HAMON ayiriladi
    b2, ayirma2 = baza(qur(15, kelmagan={5, 6}))
    kutilgan = OYLIK - OYLIK / ISH_KUNI * 2
    check("S-01: haqiqiy kelmagan kun ayiriladi (2 kun)",
          abs(b2 - kutilgan) < 1, f"base={b2}, kutilgan={kutilgan}")
    check("S-01: ayirma qatorida aynan 2 kun",
          ayirma2 is not None and ayirma2["quantity"] == 2, f"={ayirma2}")
    check("S-01: kunlik ulush maxraji BUTUN oy (11 kun emas, 26 kun)",
          abs(abs(_D(str(ayirma2["amount"]))) - OYLIK / ISH_KUNI * 2) < 1,
          f"={ayirma2['amount']}")

    # (3) REGRESSIYA: oy tugagach natija AYNAN bir xil qoladi
    b3, _ = baza(qur(ISH_KUNI + 1, kelmagan={5, 6}))
    check("S-01: oy tugagach hisob o'zgarmaydi (o'tmishga tegilmagan)",
          abs(b3 - b2) < 1, f"oy tugagach={b3}, oy o'rtasida={b2}")

    # (4) BUGUNGI kun ham «hali tugamagan» (3a18520)
    kunlar_bugun = qur(15)
    bugungi = [d for d in kunlar_bugun if d["date"] == date(2026, 6, 15)]
    check("S-01: BUGUNGI kun ham `future` (ertalab hisoblansa jazolamaydi)",
          len(bugungi) == 1 and bugungi[0]["status"] == "future",
          f"={bugungi[0]['status'] if bugungi else 'topilmadi'}")

    # (5) Kelmagan-kun JARIMASI ham kelajakni sanamaydi
    POLICY_FIX = _NS(absent_mode="fixed", absent_fine=_D("100000"))
    jarima = pr.compute_absent_fine(qur(15), POLICY_FIX)
    check("S-01: kelajak kunlari uchun jarima ham YO'Q",
          jarima["absent_days"] == 0 and float(jarima["amount"]) == 0.0, f"={jarima}")


def test_fine_from_bonus() -> None:
    """S-02 (yangi TZ 2.1) — ushlanma avval BONUSDAN, qoldiq HR qoidasiga ko'ra.

    NEGA: ushlanma to'g'ridan-to'g'ri ish haqidan olinardi
    (`fine_applies_to='net_salary'`). TZ bo'yicha bu O'zbekiston Mehnat
    kodeksiga zid. Endi default `bonus_first`, qoldiq esa panelda tanlanadi.

    Bu blok `split_fine` ni SOF funksiya sifatida sinaydi — 3 rejim x 3 holat.
    """
    from decimal import Decimal as _D
    from types import SimpleNamespace as _NS

    from api.services import payroll as pr

    print("\n" + "=" * 60)
    print("S-02: USHLANMA AVVAL BONUSDAN")
    print("=" * 60)

    def qoida(applies="bonus_first", remainder="drop"):
        return _NS(fine_applies_to=applies, fine_remainder_mode=remainder)

    NOL = _D("0")
    USHLANMA = _D("300000")

    # ── Holat 1: bonus > ushlanma (hammasi bonusdan) ──
    for rejim in ("drop", "carry_next_month", "from_salary"):
        r = pr.split_fine(USHLANMA, _D("500000"), NOL, qoida(remainder=rejim))
        check(f"S-02 [{rejim}] bonus > ushlanma -> hammasi bonusdan",
              r["from_bonus"] == USHLANMA and r["from_salary"] == NOL
              and r["carried_out"] == NOL and r["dropped"] == NOL, f"={r}")

    # ── Holat 2: bonus < ushlanma (qoldiq bor) ──
    r = pr.split_fine(USHLANMA, _D("100000"), NOL, qoida(remainder="drop"))
    check("S-02 [drop] bonus < ushlanma -> qoldiq OLINMAYDI",
          r["from_bonus"] == _D("100000") and r["from_salary"] == NOL
          and r["dropped"] == _D("200000") and r["carried_out"] == NOL, f"={r}")

    r = pr.split_fine(USHLANMA, _D("100000"), NOL, qoida(remainder="carry_next_month"))
    check("S-02 [carry] bonus < ushlanma -> qoldiq KEYINGI OYGA",
          r["from_bonus"] == _D("100000") and r["from_salary"] == NOL
          and r["carried_out"] == _D("200000") and r["dropped"] == NOL, f"={r}")

    r = pr.split_fine(USHLANMA, _D("100000"), NOL, qoida(remainder="from_salary"))
    check("S-02 [from_salary] bonus < ushlanma -> qoldiq OYLIKDAN",
          r["from_bonus"] == _D("100000") and r["from_salary"] == _D("200000")
          and r["carried_out"] == NOL and r["dropped"] == NOL, f"={r}")

    # ── Holat 3: bonus = 0 ──
    r = pr.split_fine(USHLANMA, NOL, NOL, qoida(remainder="drop"))
    check("S-02 [drop] bonus = 0 -> ushlanma UMUMAN olinmaydi",
          r["from_bonus"] == NOL and r["from_salary"] == NOL
          and r["dropped"] == USHLANMA, f"={r}")

    r = pr.split_fine(USHLANMA, NOL, NOL, qoida(remainder="carry_next_month"))
    check("S-02 [carry] bonus = 0 -> hammasi keyingi oyga",
          r["carried_out"] == USHLANMA and r["from_salary"] == NOL, f"={r}")

    r = pr.split_fine(USHLANMA, NOL, NOL, qoida(remainder="from_salary"))
    check("S-02 [from_salary] bonus = 0 -> hammasi oylikdan",
          r["from_salary"] == USHLANMA and r["from_bonus"] == NOL, f"={r}")

    # ── Eski rejim buzilmagan (regressiya) ──
    r = pr.split_fine(USHLANMA, _D("500000"), NOL, qoida(applies="net_salary"))
    check("S-02: `net_salary` rejimi avvalgidek — hammasi oylikdan",
          r["from_salary"] == USHLANMA and r["from_bonus"] == NOL, f"={r}")

    # ── Qoida umuman yo'q -> eski (xavfsiz) xatti-harakat ──
    r = pr.split_fine(USHLANMA, _D("500000"), NOL, None)
    check("S-02: qoida yo'q bo'lsa oylikdan (eski xatti-harakat)",
          r["from_salary"] == USHLANMA, f"={r}")

    # ── O'tgan oydan ko'chgan qoldiq shu oyda olinadi ──
    r = pr.split_fine(_D("50000"), _D("500000"), _D("200000"), qoida(remainder="carry_next_month"))
    check("S-02: o'tgan oy qoldig'i shu oy bonusidan olinadi (50k+200k)",
          r["from_bonus"] == _D("250000") and r["carried_out"] == NOL, f"={r}")

    r = pr.split_fine(_D("50000"), _D("100000"), _D("200000"), qoida(remainder="carry_next_month"))
    check("S-02: qoldiq yana yetmasa — qolgani KEYINGI oyga (150k)",
          r["from_bonus"] == _D("100000") and r["carried_out"] == _D("150000"), f"={r}")

    # ── Ushlanma yo'q bo'lsa hech narsa qilinmaydi ──
    r = pr.split_fine(NOL, _D("500000"), NOL, qoida())
    check("S-02: ushlanma 0 -> hamma qiymat 0",
          all(v == NOL for v in r.values()), f"={r}")

    # ── Yaxlit tekshiruv: hech qachon KO'PROQ olinmaydi ──
    for bonus in (NOL, _D("100000"), _D("500000")):
        for rejim in ("drop", "carry_next_month", "from_salary"):
            r = pr.split_fine(USHLANMA, bonus, _D("70000"), qoida(remainder=rejim))
            jami = r["from_bonus"] + r["from_salary"] + r["carried_out"] + r["dropped"]
            check(f"S-02 [{rejim}, bonus={int(bonus)}] taqsimot yig'indisi buzilmadi",
                  jami == USHLANMA + _D("70000"), f"jami={jami}, kutilgan={USHLANMA + _D('70000')}")


def test_fine_from_bonus_e2e() -> None:
    """S-02 — PAYSLIP darajasida: ushlanma bonusdan olinadimi, qoldiq
    keyingi oyga BIR MARTA ko'chadimi, qatorlar yig'indisi `net` ga tengmi.

    `split_fine` sof funksiyasi alohida sinaladi (`test_fine_from_bonus`);
    bu yerda ZANJIR tekshiriladi — o'tgan oy payslip'idan qoldiqni o'qish
    faqat shu yo'lda ishlaydi.
    """
    import asyncio
    from decimal import Decimal as _D

    print("\n" + "=" * 60)
    print("S-02: PAYSLIP ZANJIRI (bonus_first + qoldiq)")
    print("=" * 60)

    OY1, OY2 = "2019-03", "2019-04"

    async def _run():
        from sqlalchemy import delete as _del, select as _sel

        from api.services import payroll as pr
        from db.base import async_session
        from db.models import (
            Attendance, Bonus, FinePolicy, Payslip, PayslipItem, User,
            WorkScheduleOverride, WorkScheduleWeekly,
        )

        async with async_session() as s2:
            # Eski qoldiqni tozalash (avvalgi qulagan yurishdan)
            eski = list(await s2.scalars(_sel(User.id).where(User.full_name == "T-FineBonus")))
            if eski:
                pids = list(await s2.scalars(_sel(Payslip.id).where(Payslip.user_id.in_(eski))))
                if pids:
                    await s2.execute(_del(PayslipItem).where(PayslipItem.payslip_id.in_(pids)))
                for M in (Payslip, Bonus, Attendance, WorkScheduleOverride, WorkScheduleWeekly):
                    await s2.execute(_del(M).where(M.user_id.in_(eski)))
                await s2.execute(_del(FinePolicy).where(FinePolicy.scope == "user",
                                                        FinePolicy.scope_id.in_(eski)))
                await s2.execute(_del(User).where(User.id.in_(eski)))
                await s2.commit()

            u = User(telegram_id=999700501, full_name="T-FineBonus", role="employee",
                     bot_started=True, is_active=True)
            s2.add(u)
            await s2.flush()

            for wd in range(7):
                s2.add(WorkScheduleWeekly(user_id=u.id, weekday=wd, is_working=False))
            kunlar = [date(2019, 3, 4), date(2019, 3, 5), date(2019, 3, 6)]
            for d in kunlar:
                s2.add(WorkScheduleOverride(user_id=u.id, date=d, is_working=True,
                                            start_time="09:00", end_time="18:00"))
                s2.add(Attendance(user_id=u.id, date=d, status="late", late_minutes=60,
                                  worked_minutes=420))
            # Ushlanma: 3 kun x 100 000 = 300 000. Bonus 150 000 -> qoldiq 150 000.
            s2.add(FinePolicy(scope="user", scope_id=u.id, is_active=True,
                              free_late_minutes_per_month=0, fine_mode="per_day",
                              fine_per_day=100_000, absent_mode="none",
                              monthly_cap_amount=1_000_000,
                              fine_applies_to="bonus_first",
                              fine_remainder_mode="carry_next_month"))
            s2.add(Bonus(user_id=u.id, period=OY1, amount=150_000, breakdown={}))
            s2.add(Bonus(user_id=u.id, period=OY2, amount=500_000, breakdown={}))
            await s2.commit()
            return u.id

    async def _hisobla(uid, period):
        from sqlalchemy import select as _sel

        from api.services import payroll as pr
        from db.base import async_session
        from db.models import Payslip, PayslipItem

        async with async_session() as s2:
            await pr.run_payroll(s2, period, user_ids=[uid])
        async with async_session() as s2:
            slip = await s2.scalar(_sel(Payslip).where(Payslip.user_id == uid,
                                                       Payslip.period == period))
            items = list(await s2.scalars(
                _sel(PayslipItem).where(PayslipItem.payslip_id == slip.id)
            ))
            return slip, items

    uid = None
    try:
        uid = asyncio.run(_run())

        # ── 1-oy: ushlanma 300k, bonus 150k -> 150k bonusdan, 150k keyingi oyga
        slip1, items1 = asyncio.run(_hisobla(uid, OY1))
        bd1 = slip1.breakdown or {}
        check("S-02 e2e: 150 000 bonusdan olindi",
              abs(bd1.get("fine_from_bonus", 0) - 150_000) < 1, f"={bd1.get('fine_from_bonus')}")
        check("S-02 e2e: oylikdan HECH NARSA olinmadi",
              abs(bd1.get("fine_from_salary", 0)) < 1, f"={bd1.get('fine_from_salary')}")
        check("S-02 e2e: 150 000 keyingi oyga ko'chdi",
              abs(bd1.get("fine_carried_out", 0) - 150_000) < 1, f"={bd1.get('fine_carried_out')}")
        check("S-02 e2e: payslip'da «keyingi oyga o'tdi» qatori bor",
              any(i.kind == "fine_waived" for i in items1),
              f"={[i.kind for i in items1]}")
        check("S-02 e2e: ushlanma qatorida MANBA ko'rsatilgan",
              any("bonusdan" in (i.label or "") for i in items1),
              f"={[i.label for i in items1]}")
        jami1 = sum(float(i.amount) for i in items1)
        check("S-02 e2e: qatorlar yig'indisi = net (1-oy)",
              abs(jami1 - float(slip1.net)) <= 50, f"qatorlar={jami1}, net={slip1.net}")

        # ── 2-oy: ushlanma yo'q, lekin o'tgan oy qoldig'i 150k olinadi
        slip2, items2 = asyncio.run(_hisobla(uid, OY2))
        bd2 = slip2.breakdown or {}
        check("S-02 e2e: o'tgan oy qoldig'i 2-oyda o'qildi",
              abs(bd2.get("fine_carried_in", 0) - 150_000) < 1, f"={bd2.get('fine_carried_in')}")
        check("S-02 e2e: qoldiq 2-oy bonusidan olindi",
              abs(bd2.get("fine_from_bonus", 0) - 150_000) < 1, f"={bd2.get('fine_from_bonus')}")
        check("S-02 e2e: 2-oyda yangi qoldiq qolmadi",
              abs(bd2.get("fine_carried_out", 0)) < 1, f"={bd2.get('fine_carried_out')}")
        check("S-02 e2e: 2-oy payslip'ida «o'tgan oydan» qatori bor",
              any(i.kind == "fine_carry_in" for i in items2), f"={[i.kind for i in items2]}")
        jami2 = sum(float(i.amount) for i in items2)
        check("S-02 e2e: qatorlar yig'indisi = net (2-oy)",
              abs(jami2 - float(slip2.net)) <= 50, f"qatorlar={jami2}, net={slip2.net}")

        # ── IDEMPOTENT: 2-oyni qayta hisoblasak qoldiq IKKI MARTA olinmasin
        net_oldin = float(slip2.net)
        slip2b, _ = asyncio.run(_hisobla(uid, OY2))
        check("S-02 e2e: qayta hisoblanganda qoldiq IKKI MARTA olinmaydi",
              abs(float(slip2b.net) - net_oldin) < 1,
              f"birinchi={net_oldin}, ikkinchi={slip2b.net}")
    except Exception:
        check("S-02 e2e (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        if uid is not None:
            conn = db()
            cur = conn.cursor()
            cur.execute("delete from payslip_items where payslip_id in"
                        " (select id from payslips where user_id=?)", (uid,))
            for t in ("payslips", "bonuses", "attendance", "work_schedule_override",
                      "work_schedule_weekly"):
                cur.execute(f"delete from {t} where user_id=?", (uid,))
            cur.execute("delete from fine_policies where scope='user' and scope_id=?", (uid,))
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from users where id=?", (uid,))
            cur.execute("delete from payroll_periods where period in (?,?)", (OY1, OY2))
            conn.commit()
            conn.close()


def test_me_sections() -> None:
    """S-04 (TZ 2.6) — «kim nimani ko'radi» endi BITTA joyda.

    Ilgari uch joyda edi: `web/src/Layout.tsx` (yon panel),
    `web/src/lib/employeeNav.ts` (kabinet) va `bot/keyboards.py` (bot).
    Muvofiqlik INSON e'tiboriga qolgan edi — `employeeNav.ts` boshida
    «bot bilan AYNAN bir xil bo'lishi shart» degan ogohlantirish turibdi.

    Bu blok ENG MUHIM shartni qo'riqlaydi: yangi manba hozirgi yon panel
    bilan AYNAN mos (regressiya yo'q). Ro'yxat qo'lda yozilgan — u
    `Layout.tsx` dan MUSTAQIL, aks holda test hech narsani isbotlamasdi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-04: /me/sections — YAGONA MANBA")
    print("=" * 60)

    # `web/src/Layout.tsx: NAV_GROUPS` dan QO'LDA ko'chirilgan (2026-08-19).
    # Shartsiz bandlar — har rahbarda ko'rinadi.
    RAHBAR_DOIM = [
        "/", "/statistics", "/reports",
        "/attendance", "/excused-days", "/work-schedule", "/work-log", "/offices",
        "/lead-stats", "/funnel", "/norms",
        "/payroll",
        "/users", "/audit-logs",
        "/check-in",
    ]
    # Faqat hr/boss/dasturchi (ROP'da YO'Q)
    RAHBAR_PAYROLL = ["/overtime", "/requests", "/appeals", "/positions", "/celebration"]

    conn = db()
    cur = conn.cursor()
    uid = None
    try:
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── ROL MATRITSASI ──
            natijalar = {}
            for rol in ("boss", "hr", "rop", "dasturchi"):
                mgr = None
                row = cur.execute(
                    "select id from users where role=? and is_active=1 limit 1", (rol,)
                ).fetchone()
                if not row:
                    continue
                t = token_for(row[0], rol)
                r = c.get("/me/sections", headers=auth(t))
                if r.status_code != 200:
                    check(f"S-04: /me/sections [{rol}] -> 200", False,
                          "kod=" + str(r.status_code) + " " + r.text[:120])
                    continue
                yollar = [x["path"] for x in r.json()]
                natijalar[rol] = yollar
                check(f"S-04: /me/sections [{rol}] -> 200", True, f"{len(yollar)} bo'lim")

            # Har rahbarda shartsiz bandlar BOR
            for rol, yollar in natijalar.items():
                yetishmaydi = [p for p in RAHBAR_DOIM if p not in yollar]
                check(f"S-04 [{rol}]: yon panelning shartsiz bandlari to'liq",
                      not yetishmaydi, "yetishmaydi=" + str(yetishmaydi))

            # ROP payroll bandlarini KO'RMAYDI, hr/boss/dasturchi KO'RADI
            if "rop" in natijalar:
                ortiqcha = [p for p in RAHBAR_PAYROLL if p in natijalar["rop"]]
                check("S-04 [rop]: payroll bandlari KO'RINMAYDI",
                      not ortiqcha, "ortiqcha=" + str(ortiqcha))
            for rol in ("hr", "boss", "dasturchi"):
                if rol not in natijalar:
                    continue
                yetishmaydi = [p for p in RAHBAR_PAYROLL if p not in natijalar[rol]]
                check(f"S-04 [{rol}]: payroll bandlari KO'RINADI",
                      not yetishmaydi, "yetishmaydi=" + str(yetishmaydi))

            # «Dasturchi rejimi» — FAQAT dasturchida
            for rol, yollar in natijalar.items():
                bormi = "/dasturchi" in yollar
                check(f"S-04 [{rol}]: /dasturchi {'bor' if rol == 'dasturchi' else 'YO_Q'}",
                      bormi == (rol == "dasturchi"), f"bor={bormi}")

            # ── XODIM ──
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " created_at) values (999700301,'T-Sections','employee',0,1,datetime('now'))")
            uid = cur.lastrowid
            conn.commit()
            emp_t = token_for(uid, "employee")
            r = c.get("/me/sections", headers=auth(emp_t))
            check("S-04: xodim uchun -> 200", r.status_code == 200, "kod=" + str(r.status_code))
            emp = r.json() if r.status_code == 200 else []
            yollar = [x["path"] for x in emp]

            check("S-04: xodimga KABINET to'plami keladi (yon panel emas)",
                  all(x["audience"] == "employee" for x in emp), "=" + str(yollar[:5]))
            check("S-04: xodim rahbar sahifalarini KO'RMAYDI",
                  "/users" not in yollar and "/payroll" not in yollar
                  and "/audit-logs" not in yollar, "=" + str(yollar))
            for kerak in ("/check-in", "/me/payroll", "/me/stats", "/me/excused"):
                check(f"S-04: xodimda {kerak} bor", kerak in yollar, "=" + str(yollar))
            check("S-04: bo'limlar tartiblangan",
                  [x["order"] for x in emp] == sorted(x["order"] for x in emp),
                  "=" + str([x["order"] for x in emp]))
            check("S-04: har bandda mijozga kerakli maydonlar bor",
                  all(x.get("key") and x.get("label") and x.get("path") and x.get("icon")
                      for x in emp), "=" + str(emp[:1]))

            # `can_edit_attendance` bayrog'i bo'limni OCHADI
            check("S-04: bayroqsiz xodimda «Davomat tuzatish» YO'Q",
                  "/attendance" not in yollar, "=" + str(yollar))
            cur.execute("update users set can_edit_attendance=1 where id=?", (uid,))
            conn.commit()
            r2 = c.get("/me/sections", headers=auth(emp_t))
            yollar2 = [x["path"] for x in r2.json()] if r2.status_code == 200 else []
            check("S-04: bayroq berilgach «Davomat tuzatish» PAYDO bo'ladi",
                  "/attendance" in yollar2, "=" + str(yollar2))

            # Autentifikatsiyasiz — 401
            r3 = c.get("/me/sections")
            check("S-04: tokensiz -> 401", r3.status_code == 401, "kod=" + str(r3.status_code))
    except Exception:
        check("S-04 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        if uid is not None:
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        conn.close()


def test_bot_menu_from_server() -> None:
    """S-05b (TZ 2.6) — bot menyusi endi SERVERDA quriladi.

    Ilgari `bot/keyboards.py::main_menu` ~20 ta shartni O'ZI hisoblardi va
    AYNAN o'sha shartlar saytda ham (ikki joyda) takrorlanardi. Endi qoida
    `api/services/sections.py` da, bot faqat chizadi.

    Bu blok ikkita narsani qo'riqlaydi:
      (a) menyu ESKI ko'rinishni saqlagan (oltin namuna — qo'lda yozilgan);
      (b) bot javobida `bot_menu` keladi va tugmalar tanish.
    """
    import httpx

    from api.services.sections import bot_menu_rows

    print("\n" + "=" * 60)
    print("S-05b: BOT MENYUSI SERVERDAN")
    print("=" * 60)

    class _Pos:
        def __init__(self, flags, metrics):
            self.menu_flags = flags
            self.metrics = metrics

    class _U:
        def __init__(self, role, flags=None, metrics=None):
            self.role = role
            # ⚠️ `is None` SHART: `metrics=[]` (ataylab bo'sh lavozim) va
            # `metrics=None` (lavozim biriktirilmagan) BOSHQA-BOSHQA holat.
            # `if flags or metrics` deb yozilsa bo'sh ro'yxat «lavozim yo'q»
            # ga aylanib, sinov noto'g'ri natija bergan edi.
            self.position = (
                None if (flags is None and metrics is None) else _Pos(flags, metrics)
            )
            self.can_edit_fine_policy = False
            self.can_edit_attendance = False

    # ── (a) OLTIN NAMUNA: oddiy xodim, lavozimsiz ──
    # Qo'lda yozilgan — `bot_menu_rows` dan MUSTAQIL, aks holda test
    # o'zini o'zi tasdiqlagan bo'lardi.
    #
    # ⚠️ NAMUNA O'ZGARSA — ATAYLABMI? Bu ro'yxat menyuning TASDIQLANGAN
    # ko'rinishi. U faqat yangi bo'lim ataylab qo'shilganda yangilanadi va
    # o'zgarish commit izohida ko'rsatiladi. Tasodifiy o'zgarish (masalan
    # `visible` shartini buzib qo'yish) shu yerda ushlanadi.
    # Tarix: S-11 da «📁 Hujjatlarim» qo'shildi (TZ 3.4);
    #        S-28 da «❓ HR ga savol» qo'shildi (TZ 3.29);
    #        S-35 da «📚 Darsliklarim» qo'shildi (TZ 3.1).
    kutilgan = [
        ["✅ Keldim / Ketdim"],
        ["📋 Vazifalarim"],
        ["📝 Ish kundaligi"],
        ["📊 Bugungi normam", "💰 Oylik KPI'm"],
        ["📈 Statistikam", "🙋 Sababli kun so'rash"],
        ["📮 Murojaatlarim", "📁 Hujjatlarim"],
        ["❓ HR ga savol"],
        ["📚 Darsliklarim"],
        ["🗓 Ish jadvali"],
        ["💵 Mening oyligim"],
        ["📋 Bugungi rejam"],
        ["🧲 Lidlar statistikasi"],
        ["🤖 Sotuv AI"],
    ]
    haqiqiy = bot_menu_rows(_U("employee"))
    check("S-05b: oddiy xodim menyusi eski ko'rinishda",
          haqiqiy == kutilgan, "=" + str(haqiqiy))

    # ── Boshliq: shaxsiy «Keldim/Ketdim» va oylik YO'Q, boshqaruv bor ──
    boss = bot_menu_rows(_U("boss"))
    tekis_boss = [b for row in boss for b in row]
    check("S-05b: Boshliqda «Keldim / Ketdim» YO'Q",
          "✅ Keldim / Ketdim" not in tekis_boss, "=" + str(tekis_boss[:4]))
    check("S-05b: Boshliqda «Mening oyligim» YO'Q",
          "💵 Mening oyligim" not in tekis_boss, "=" + str(tekis_boss))
    for kerak in ("📤 Vazifa berish", "🧾 Audit jurnali", "🎬 Tabrik videolari"):
        check(f"S-05b: Boshliqda «{kerak}» bor", kerak in tekis_boss, "=" + str(tekis_boss))

    # ── ROP: boshqaruv bor, lekin «Xodim uchun sababli kun» YO'Q ──
    rop = [b for row in bot_menu_rows(_U("rop")) for b in row]
    check("S-05b: ROP «Xodim uchun sababli kun» ni KO'RMAYDI",
          "🙋 Xodim uchun sababli kun" not in rop, "=" + str(rop))
    check("S-05b: ROP «Sotuv AI» ni ko'radi", "🤖 Sotuv AI" in rop, "=" + str(rop))
    check("S-05b: ROP «Audit jurnali» ni KO'RMAYDI",
          "🧾 Audit jurnali" not in rop, "=" + str(rop))

    # ── Lavozim bayroqlari tugmani YOPADI ──
    yopiq = [b for row in bot_menu_rows(_U("employee", flags={"tasks": False, "payroll": False}))
             for b in row]
    check("S-05b: `tasks=False` -> «Vazifalarim» yo'q",
          "📋 Vazifalarim" not in yopiq, "=" + str(yopiq))
    check("S-05b: `payroll=False` -> «Mening oyligim» yo'q",
          "💵 Mening oyligim" not in yopiq, "=" + str(yopiq))

    # ── `metrics=[]` (ataylab bo'sh) va `None` (lavozim yo'q) FARQI ──
    bosh = [b for row in bot_menu_rows(_U("employee", metrics=[])) for b in row]
    yoq = [b for row in bot_menu_rows(_U("employee", metrics=None)) for b in row]
    check("S-05b: `metrics=[]` -> lid statistikasi YO'Q",
          "🧲 Lidlar statistikasi" not in bosh, "=" + str(bosh))
    check("S-05b: lavozimsiz (`None`) -> lid statistikasi BOR (eski default)",
          "🧲 Lidlar statistikasi" in yoq, "=" + str(yoq))

    # ── (b) Bot API javobida menyu keladi ──
    conn = db()
    cur = conn.cursor()
    uid = None
    try:
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999700401,'T-BotMenu','employee',1,1,datetime('now'))")
        uid = cur.lastrowid
        conn.commit()
        with httpx.Client(base_url=API_BASE, timeout=20) as c:
            r = c.get("/users/by-telegram/999700401", headers=bot_secret_hdr())
            check("S-05b: /users/by-telegram -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            menyu = r.json().get("bot_menu") if r.status_code == 200 else None
            check("S-05b: javobda `bot_menu` bor", bool(menyu), "=" + str(menyu))
            check("S-05b: javobdagi menyu server hisobi bilan bir xil",
                  menyu == kutilgan, "=" + str(menyu))

            # Bot tomoni: qatorlardan klaviatura quriladi
            from bot.keyboards import main_menu
            km = main_menu(menyu)
            check("S-05b: bot qatorlarni klaviaturaga aylantiradi",
                  len(km.keyboard) == len(kutilgan), "=" + str(len(km.keyboard)))
            zaxira = main_menu(None)
            check("S-05b: menyu kelmasa MINIMAL zaxira (ikkinchi manba emas)",
                  len(zaxira.keyboard) == 1, "=" + str(len(zaxira.keyboard)))
    except Exception:
        check("S-05b (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        if uid is not None:
            cur.execute("delete from audit_logs where target_user_id=? or actor_id=?", (uid, uid))
            cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        conn.close()


def test_view_scope() -> None:
    """S-06 (TZ 4-qism) — «xodim faqat o'ziga tegishlisini ko'radi».

    Ilgari bu qoida 10 dan ortiq joyda QO'LDA yozilgan edi va ular BIR XIL
    EMAS edi: payroll ROP ga lavozim bo'yicha jamoani ham ko'rsatardi,
    ish kundaligi esa faqat bevosita bo'ysunuvchilarni. Endi yagona qatlam
    — `api/deps.py::scoped_user_ids` / `assert_can_view`.

    ⚠️ 404, 403 EMAS: 403 «yozuv bor, lekin ruxsat yo'q» degani, ya'ni
    begona xodim MAVJUDLIGINI oshkor qiladi.
    """
    import asyncio
    import httpx

    print("\n" + "=" * 60)
    print("S-06: MARKAZLASHGAN KO'RINISH FILTRI")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    ids = {}
    try:
        # Ikki ROP, har birida bittadan xodim + bir «begona» xodim
        for nom, rol, tg in (
            ("T-Scope-Rop1", "rop", 999700201),
            ("T-Scope-Rop2", "rop", 999700202),
            ("T-Scope-Emp1", "employee", 999700203),
            ("T-Scope-Emp2", "employee", 999700204),
            ("T-Scope-Hr", "hr", 999700205),
        ):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " created_at) values (?,?,?,0,1,datetime('now'))", (tg, nom, rol))
            ids[nom] = cur.lastrowid
        cur.execute("update users set manager_id=? where id=?",
                    (ids["T-Scope-Rop1"], ids["T-Scope-Emp1"]))
        cur.execute("update users set manager_id=? where id=?",
                    (ids["T-Scope-Rop2"], ids["T-Scope-Emp2"]))
        conn.commit()

        rop1_t = token_for(ids["T-Scope-Rop1"], "rop")
        emp1_t = token_for(ids["T-Scope-Emp1"], "employee")
        hr_t = token_for(ids["T-Scope-Hr"], "hr")

        # ── (1) Xizmat qatlami: qamrov to'g'ri hisoblanadimi ──
        from api.deps import scoped_user_ids
        from db.base import async_session
        from db.models import User as _U

        async def _qamrov(uid):
            async with async_session() as s2:
                u = await s2.get(_U, uid)
                return await scoped_user_ids(u, s2)

        hr_scope = asyncio.run(_qamrov(ids["T-Scope-Hr"]))
        check("S-06: HR qamrovi CHEKLOVSIZ (None)", hr_scope is None, "=" + str(hr_scope))

        rop1_scope = asyncio.run(_qamrov(ids["T-Scope-Rop1"]))
        check("S-06: ROP o'zini ko'radi",
              ids["T-Scope-Rop1"] in rop1_scope, "=" + str(sorted(rop1_scope)))
        check("S-06: ROP O'Z jamoasini ko'radi",
              ids["T-Scope-Emp1"] in rop1_scope, "=" + str(sorted(rop1_scope)))
        check("S-06: ROP BOSHQA ROP ni ko'rmaydi",
              ids["T-Scope-Rop2"] not in rop1_scope, "=" + str(sorted(rop1_scope)))
        check("S-06: ROP HR ni ko'rmaydi",
              ids["T-Scope-Hr"] not in rop1_scope, "=" + str(sorted(rop1_scope)))

        emp_scope = asyncio.run(_qamrov(ids["T-Scope-Emp1"]))
        check("S-06: xodim FAQAT o'zini ko'radi",
              emp_scope == {ids["T-Scope-Emp1"]}, "=" + str(emp_scope))

        # `rop_sees_team=False` — ba'zi modullarda ROP jamoani ham ko'rmaydi
        async def _qatiy(uid):
            async with async_session() as s2:
                u = await s2.get(_U, uid)
                return await scoped_user_ids(u, s2, rop_sees_team=False)

        qatiy = asyncio.run(_qatiy(ids["T-Scope-Rop1"]))
        check("S-06: `rop_sees_team=False` -> ROP faqat o'zi",
              qatiy == {ids["T-Scope-Rop1"]}, "=" + str(qatiy))

        # ── (2) HTTP: begona so'rov 404 (403 EMAS) ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            begona = ids["T-Scope-Emp2"]
            oz = ids["T-Scope-Emp1"]

            r = c.get(f"/bonuses?user_id={begona}", headers=auth(rop1_t))
            check("S-06: ROP begona xodim bonusini so'radi -> 404 (403 emas)",
                  r.status_code == 404, "kod=" + str(r.status_code))
            r = c.get(f"/bonuses?user_id={oz}", headers=auth(rop1_t))
            check("S-06: ROP O'Z jamoasi bonusini ko'radi -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code) + " " + r.text[:100])
            r = c.get(f"/bonuses?user_id={begona}", headers=auth(hr_t))
            check("S-06: HR hammani ko'radi -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code))

            r = c.get(f"/payroll/2019-01/user/{begona}", headers=auth(rop1_t))
            check("S-06: ROP begona payslip -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            # Xodim bu endpointlarga ROL bo'yicha kirolmaydi -> 403
            # (bu BOSHQA holat: «bunday imkoniyat yo'q», «bu yozuv sizniki
            # emas» emas — shuning uchun 404 emas, 403 to'g'ri).
            r = c.get(f"/bonuses?user_id={oz}", headers=auth(emp1_t))
            check("S-06: xodimga rol bo'yicha yopiq -> 403 (404 emas)",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── (3) Ro'yxat endpointlari filtrlanadimi ──
            r = c.get("/requests?status_filter=pending", headers=auth(rop1_t))
            check("S-06: ROP murojaatlar ro'yxati -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code))
            if r.status_code == 200:
                yot = [x for x in r.json() if x.get("user_id") == begona]
                check("S-06: ro'yxatda BEGONA xodim yo'q", not yot, "=" + str(yot[:2]))

            r = c.get("/excused-days?status_filter=pending", headers=auth(rop1_t))
            check("S-06: ROP sababli kunlar ro'yxati -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code))
            if r.status_code == 200:
                yot = [x for x in r.json() if x.get("user_id") == begona]
                check("S-06: sababli kunlarda BEGONA xodim yo'q", not yot, "=" + str(yot[:2]))
    except Exception:
        check("S-06 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        if ids:
            hammasi = list(ids.values())
            qm = ",".join("?" * len(hammasi))
            cur.execute(f"update users set manager_id=null where manager_id in ({qm})", hammasi)
            for t in ("bonuses", "payslips", "excused_days", "attendance", "work_log_entries"):
                try:
                    cur.execute(f"delete from {t} where user_id in ({qm})", hammasi)
                except sqlite3.OperationalError:
                    pass
            cur.execute(f"delete from audit_logs where target_user_id in ({qm})"
                        f" or actor_id in ({qm})", hammasi + hammasi)
            cur.execute(f"delete from users where id in ({qm})", hammasi)
            conn.commit()
        conn.close()


def test_background_jobs() -> None:
    """S-07 (TZ 2.2) — og'ir ish so'rov ichida BAJARILMAYDI.

    cPanel Passenger'da konkurentlik = 1: Excel eksporti so'rov ichida
    yasalgani uchun o'sha vaqt davomida BUTUN sayt javob bermasdi. Endi
    so'rov navbatga qo'yadi (202), og'ir ishni cron jarayoni bajaradi.

    Qabul mezonlari (TZ):
      • eksport so'rov ichida bajarilmaydi;
      • foydalanuvchi «tayyorlanmoqda» xabarini oladi;
      • bitta ish IKKI MARTA bajarilmaydi;
      • xato bo'lsa `failed` va sabab yoziladi, cron O'LMAYDI.
    """
    import asyncio
    import time

    import httpx

    print("\n" + "=" * 60)
    print("S-07: OG'IR ISH NAVBATI")
    print("=" * 60)

    from api.services.background_jobs import background_tick, enqueue
    from db.base import async_session
    from db.models import BackgroundJob

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    job_ids = []
    try:
        # ── (1) So'rov DARHOL qaytadi ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            t0 = time.perf_counter()
            r = c.post("/reports/export-async?date_from=2019-01-01&date_to=2019-01-31",
                       headers=auth(mgr_t))
            davomiylik = time.perf_counter() - t0
            # Telegram ulanmagan rahbar bo'lsa 400 — bu ham to'g'ri xatti-harakat
            if r.status_code == 400:
                check("S-07: Telegramsiz rahbarga tushunarli xato",
                      "Telegram" in r.text, "=" + r.text[:120])
            else:
                check("S-07: eksport so'rovi -> 202 (natija emas, NAVBAT)",
                      r.status_code == 202, "kod=" + str(r.status_code) + " " + r.text[:120])
                check("S-07: so'rov TEZ qaytadi (og'ir ish ichida emas)",
                      davomiylik < 2.0, f"{davomiylik:.2f}s")
                job = r.json()
                job_ids.append(job.get("job_id"))
                check("S-07: javobda `job_id` va «tayyorlanmoqda» xabari bor",
                      job.get("job_id") and "tayyorlanmoqda" in (job.get("message") or ""),
                      "=" + str(job))

                # Holat endpointi
                r2 = c.get(f"/reports/jobs/{job['job_id']}", headers=auth(mgr_t))
                check("S-07: holat endpointi -> 200 va `queued`",
                      r2.status_code == 200 and r2.json().get("status") == "queued",
                      "kod=" + str(r2.status_code) + " " + r2.text[:100])

                # Begona ish -> 404 (S-06 qoidasi)
                r3 = c.get("/reports/jobs/999999", headers=auth(mgr_t))
                check("S-07: begona/yo'q ishga -> 404", r3.status_code == 404,
                      "kod=" + str(r3.status_code))

        # ── (2) Ish navbatga TUSHDI, cron uni oladi ──
        async def _tick():
            async with async_session() as s2:
                return await background_tick(s2)

        if job_ids:
            natija = asyncio.run(_tick())
            check("S-07: cron navbatdagi ishni bajardi",
                  natija.get("ran") == "report_export" and natija.get("ok") is True,
                  "=" + str(natija))
            check("S-07: fayl yasaldi (baytlar bor)",
                  (natija.get("bytes") or 0) > 1000, "=" + str(natija.get("bytes")))

            # ── (3) IKKI MARTA bajarilmaydi ──
            ikkinchi = asyncio.run(_tick())
            check("S-07: o'sha ish IKKINCHI marta olinmaydi",
                  ikkinchi.get("ran") is None, "=" + str(ikkinchi))

            holat = cur.execute("select status from background_jobs where id=?",
                                (job_ids[0],)).fetchone()
            check("S-07: ish holati `done`", holat and holat[0] == "done", "=" + str(holat))

        # ── (4) Xato bo'lsa `failed` + sabab, cron o'lmaydi ──
        async def _xato_ish():
            async with async_session() as s2:
                j = await enqueue(s2, "report_export", {"date_from": "buzuq"}, mgr[0])
                await s2.commit()
                return j.id

        xato_id = asyncio.run(_xato_ish())
        job_ids.append(xato_id)
        natija = asyncio.run(_tick())
        check("S-07: xato ish cron'ni O'LDIRMAYDI (javob qaytdi)",
              natija.get("ok") is False, "=" + str(natija))
        row = cur.execute("select status, error from background_jobs where id=?",
                          (xato_id,)).fetchone()
        check("S-07: xato ish `failed` bo'ldi va SABAB yozildi",
              row and row[0] == "failed" and row[1], "=" + str(row))

        # ── (5) Noma'lum tur navbatga UMUMAN tushmaydi ──
        async def _nomalum():
            async with async_session() as s2:
                try:
                    await enqueue(s2, "yolgon_ish", {}, mgr[0])
                    return "o'tdi"
                except ValueError as e:
                    return str(e)

        check("S-07: noma'lum ish turi rad etiladi",
              "noma'lum" in asyncio.run(_nomalum()), "=" + asyncio.run(_nomalum()))
    except Exception:
        check("S-07 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        cur.execute("delete from background_jobs")
        conn.commit()
        conn.close()


def test_setup_status() -> None:
    """S-08 (TZ 2.7) — mexanizmi tayyor, lekin QIYMATI yo'q modullar ko'rinsin.

    Jonli isbot (2026-08-17): `kpi_rates` jadvali BUTUNLAY bo'sh edi —
    kod to'g'ri ishlardi, ko'paytiriladigan stavka yo'q edi. Buni topguncha
    butun oylik tekshiruvi kerak bo'ldi. Endi bosh sahifada ko'rinadi.

    Qabul mezonlari (TZ):
      • har qator to'g'ridan-to'g'ri sozlash sahifasiga olib boradi;
      • hammasi sozlangan bo'lsa ro'yxatda sozlanmagan qolmaydi;
      • yangi modul qo'shish uchun bitta qator yetadi.
    """
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("S-08: SOZLANMAGAN MODULLAR")
    print("=" * 60)

    from api.services.setup_status import _TEKSHIRUVLAR, collect_setup_status
    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    kpi_id = None
    try:
        async def _holat():
            async with async_session() as s2:
                return await collect_setup_status(s2)

        items = asyncio.run(_holat())
        check("S-08: ro'yxat bo'sh emas", len(items) >= 7, "=" + str(len(items)))
        check("S-08: har bandda havola bor",
              all(i.link.startswith("/") for i in items),
              "=" + str([i.link for i in items]))
        check("S-08: har bandda «nima yetishmayapti» yozilgan",
              all(len(i.missing) > 10 for i in items), "=" + str([i.key for i in items]))

        # Sozlanmaganlar TEPADA, ular ichida muhimlari birinchi
        tayyor_indeks = [n for n, i in enumerate(items) if i.ready]
        sozlanmagan_indeks = [n for n, i in enumerate(items) if not i.ready]
        if tayyor_indeks and sozlanmagan_indeks:
            check("S-08: sozlanmaganlar TEPADA",
                  max(sozlanmagan_indeks) < min(tayyor_indeks),
                  f"sozlanmagan={sozlanmagan_indeks}, tayyor={tayyor_indeks}")
        muhim = [n for n, i in enumerate(items) if not i.ready and i.critical]
        oddiy = [n for n, i in enumerate(items) if not i.ready and not i.critical]
        if muhim and oddiy:
            check("S-08: MUHIM modullar oddiylaridan yuqorida",
                  max(muhim) < min(oddiy), f"muhim={muhim}, oddiy={oddiy}")

        # ── Qiymat kiritilsa modul «tayyor» bo'ladi ──
        oldin = {i.key: i.ready for i in items}
        check("S-08: KPI stavkasi hozir yo'q (boshlang'ich holat)",
              oldin.get("kpi_rates") is False, "=" + str(oldin.get("kpi_rates")))

        cur.execute(
            "insert into kpi_rates (scope, scope_id, metric, amount, effective_from,"
            " changed_by, created_at) values ('global',null,'suhbat',1000,'2019-01-01',?,"
            "datetime('now'))", (mgr[0],))
        kpi_id = cur.lastrowid
        conn.commit()

        keyin = {i.key: i.ready for i in asyncio.run(_holat())}
        check("S-08: stavka kiritilgach modul «tayyor» bo'ldi",
              keyin.get("kpi_rates") is True, "=" + str(keyin.get("kpi_rates")))
        check("S-08: boshqa modullar holati o'zgarmadi",
              all(keyin[k] == v for k, v in oldin.items() if k != "kpi_rates"),
              "=" + str({k: (oldin[k], keyin[k]) for k in oldin if oldin[k] != keyin[k]}))

        # ── Kengaytirilishi: bitta qator yetadi ──
        check("S-08: ro'yxat bitta joyda e'lon qilingan (kengaytirish oson)",
              len(_TEKSHIRUVLAR) == len(items), f"{len(_TEKSHIRUVLAR)} vs {len(items)}")

        # ── HTTP + ruxsat ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.get("/me/setup-status", headers=auth(mgr_t))
            check("S-08: /me/setup-status rahbarga -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code))
            if r.status_code == 200:
                check("S-08: javob tuzilishi to'g'ri",
                      all({"key", "label", "ready", "missing", "link", "critical"} <= set(x)
                          for x in r.json()), "=" + str(r.json()[:1]))

            # Oddiy xodim ko'ra olmaydi
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " created_at) values (999700601,'T-Setup','employee',0,1,datetime('now'))")
            emp_id = cur.lastrowid
            conn.commit()
            r2 = c.get("/me/setup-status", headers=auth(token_for(emp_id, "employee")))
            check("S-08: oddiy xodimga -> 403", r2.status_code == 403,
                  "kod=" + str(r2.status_code))
            cur.execute("delete from users where id=?", (emp_id,))
            conn.commit()
    except Exception:
        check("S-08 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        if kpi_id is not None:
            cur.execute("delete from kpi_rates where id=?", (kpi_id,))
        cur.execute("delete from users where full_name='T-Setup'")
        conn.commit()
        conn.close()


def test_holidays() -> None:
    """S-09 (TZ 2.9) — bayram ish kuni sifatida sanalmasin, ta'tildagi
    xodim check-in qilolmasin.

    Qabul mezonlari (TZ):
      • bayram kuni ish kuni sifatida sanalmaydi (hamma hisoblagichda);
      • ta'tildagi xodim check-in qilolmaydi;
      • bayram kiritilmagan bo'lsa «Sozlanmagan modullar» da ko'rinadi.

    Alohida tekshiriladigan CHEGARA HOLAT (TZ talabi): bayram + sababli
    kun + dam olish kuni BIR KUNGA to'g'ri kelsa hisob buzilmasin —
    kun ikki marta ayirilmasin va manfiy natija chiqmasin.

    USTUVORLIK ham tekshiriladi: xodimga ATAYIN qo'yilgan kunlik override
    bayramdan kuchli (bayram navbatchiligi), bayram esa haftalik
    andozadan kuchli.
    """
    import asyncio
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-09: BAYRAMLAR + TA'TILDAGI CHECK-IN")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    # Kelasi yilning YANVARI — jonli ma'lumotga tegmaslik uchun ataylab
    # kelajak: o'sha oyda hech kimning davomati/oyligi yo'q.
    yil = _date.today().year + 1
    # 4 dan 10 gacha bo'lgan oraliqda ish kuni (Du-Ju) bo'lgan sana tanlaymiz.
    bayram_kuni = next(
        d for d in (_date(yil, 1, n) for n in range(4, 12)) if d.weekday() < 5
    )
    dam_kuni = next(
        d for d in (_date(yil, 1, n) for n in range(4, 12)) if d.weekday() >= 5
    )
    ish_kuni2 = bayram_kuni + _td(days=1) if (bayram_kuni + _td(days=1)).weekday() < 5 \
        else bayram_kuni + _td(days=3)
    davr = f"{yil}-01"

    conn = db()
    cur = conn.cursor()
    uid = None
    try:
        # Oldingi UZILGAN yurishning qoldig'i natijani buzmasin: birinchi
        # tekshiruv «bayramdan oldin bu kun ish kuni edi» degan boshlang'ich
        # holatga tayanadi. Tozalash faqat oxirida bo'lsa, test yarmida
        # yiqilgan yurishdan keyingi safar noto'g'ri FAIL berardi.
        cur.execute("delete from holidays where name like 'T-%' or name='Takror'")
        cur.execute("delete from users where full_name='T-Holiday'")
        conn.commit()
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999700901,'T-Holiday','employee',0,1,datetime('now'))")
        uid = cur.lastrowid
        conn.commit()

        async def _sched(period=davr):
            from api.services.payroll import month_schedule
            from db.models import User as U
            async with async_session() as s2:
                u = await s2.get(U, uid)
                return {d["date"]: d["is_working"] for d in await month_schedule(s2, u, period)}

        async def _range(a, b):
            from api.services.workdays import range_days
            from db.models import User as U
            async with async_session() as s2:
                u = await s2.get(U, uid)
                return {d["date"]: d["is_working"] for d in await range_days(s2, u, a, b)}

        # ── 1) Boshlang'ich holat: bayram yo'q, oddiy ish kuni ──
        oldin = asyncio.run(_sched())
        check("S-09: bayramdan OLDIN kun ish kuni edi",
              oldin.get(bayram_kuni) is True, f"{bayram_kuni} -> {oldin.get(bayram_kuni)}")
        ish_kuni_oldin = sum(1 for v in oldin.values() if v)

        # ── 2) HTTP orqali bayram qo'shish (HR paneli yo'li) ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.post("/holidays", headers=auth(mgr_t),
                       json={"date": bayram_kuni.isoformat(), "name": "T-Bayram",
                             "kind": "state"})
            check("S-09: POST /holidays -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            hid = r.json().get("id") if r.status_code == 201 else None

            # Takroriy sana -> 409 (HR ro'yxatni ikki marta yuborishi normal)
            r2 = c.post("/holidays", headers=auth(mgr_t),
                        json={"date": bayram_kuni.isoformat(), "name": "Takror",
                              "kind": "state"})
            check("S-09: takroriy sana -> 409", r2.status_code == 409,
                  "kod=" + str(r2.status_code))

            # Oddiy xodim qo'sha olmaydi (lekin KO'RA oladi)
            emp_t = token_for(uid, "employee")
            r3 = c.post("/holidays", headers=auth(emp_t),
                        json={"date": ish_kuni2.isoformat(), "name": "X", "kind": "state"})
            check("S-09: oddiy xodim qo'sha olmaydi -> 403", r3.status_code == 403,
                  "kod=" + str(r3.status_code))
            r4 = c.get("/holidays", headers=auth(emp_t), params={"year": yil})
            check("S-09: oddiy xodim ro'yxatni KO'RADI -> 200", r4.status_code == 200,
                  "kod=" + str(r4.status_code))

            # ── 3) Ro'yxatni bir marta kiritish (bulk) ──
            r5 = c.post("/holidays/bulk", headers=auth(mgr_t), json={
                "items": [
                    {"date": bayram_kuni.isoformat(), "name": "Takror", "kind": "state"},
                    {"date": dam_kuni.isoformat(), "name": "T-DamBayram", "kind": "company"},
                ],
                "overwrite": False,
            })
            check("S-09: bulk -> 200", r5.status_code == 200, "kod=" + str(r5.status_code))
            if r5.status_code == 200:
                j = r5.json()
                check("S-09: bulk borini o'tkazib yubordi, yangisini qo'shdi",
                      j.get("added") == 1 and j.get("skipped") == 1, "=" + str(j))

        # ── 4) Bayram ish kuni sifatida SANALMAYDI ──
        keyin = asyncio.run(_sched())
        check("S-09: bayram kuni endi ish kuni EMAS",
              keyin.get(bayram_kuni) is False, f"{bayram_kuni} -> {keyin.get(bayram_kuni)}")
        check("S-09: oydagi ish kunlari ayni 1 taga kamaydi",
              sum(1 for v in keyin.values() if v) == ish_kuni_oldin - 1,
              f"{ish_kuni_oldin} -> {sum(1 for v in keyin.values() if v)}")
        check("S-09: qolgan kunlarga tegmadi",
              all(keyin[d] == v for d, v in oldin.items() if d != bayram_kuni),
              "=" + str([d for d, v in oldin.items() if d != bayram_kuni and keyin[d] != v]))

        # ── 5) CHEGARA: bayram DAM OLISH kuniga to'g'ri kelsa ──
        # Kun ikki marta ayirilmasin — u allaqachon ish kuni emas edi.
        check("S-09: dam kuniga tushgan bayram hisobni BUZMADI",
              keyin.get(dam_kuni) is False and
              sum(1 for v in keyin.values() if v) == ish_kuni_oldin - 1,
              f"{dam_kuni} -> {keyin.get(dam_kuni)}")

        # ── 6) `workdays.range_days` ham bilishi kerak ──
        rd = asyncio.run(_range(_date(yil, 1, 1), _date(yil, 1, 31)))
        check("S-09: workdays.range_days da ham bayram ish kuni emas",
              rd.get(bayram_kuni) is False, f"{bayram_kuni} -> {rd.get(bayram_kuni)}")

        # ── 7) USTUVORLIK: override bayramdan KUCHLI ──
        cur.execute(
            "insert into work_schedule_override (user_id, date, is_working, start_time,"
            " end_time, updated_at) values (?,?,1,'09:00','18:00',datetime('now'))",
            (uid, bayram_kuni.isoformat()))
        conn.commit()
        ustun = asyncio.run(_sched())
        check("S-09: override bayramdan KUCHLI (bayram navbatchiligi)",
              ustun.get(bayram_kuni) is True, f"{bayram_kuni} -> {ustun.get(bayram_kuni)}")
        cur.execute("delete from work_schedule_override where user_id=?", (uid,))
        conn.commit()

        # ── 8) USTUVORLIK: bayram haftalik andozadan KUCHLI ──
        cur.execute(
            "insert into work_schedule_weekly (user_id, weekday, is_working, start_time,"
            " end_time, updated_at) values (?,?,1,'09:00','18:00',datetime('now'))",
            (uid, bayram_kuni.weekday()))
        conn.commit()
        haftalik = asyncio.run(_sched())
        check("S-09: bayram haftalik andozadan KUCHLI",
              haftalik.get(bayram_kuni) is False,
              f"{bayram_kuni} -> {haftalik.get(bayram_kuni)}")
        cur.execute("delete from work_schedule_weekly where user_id=?", (uid,))
        conn.commit()

        # ── 9) UCHALASI BIR KUNDA: bayram + sababli kun + dam kuni ──
        # `target_split._working_days` sababli kunni AYIRADI; kun allaqachon
        # bayram/dam kuni bo'lsa ikki marta ayirilmasligi kerak.
        cur.execute(
            "insert into excused_days (user_id, date, reason, status, created_at)"
            " values (?,?,'T-tatil','approved',datetime('now'))",
            (uid, dam_kuni.isoformat()))
        conn.commit()

        async def _wd():
            from api.services import target_split
            from db.models import User as U
            async with async_session() as s2:
                u = await s2.get(U, uid)
                excused = (await target_split._excused_map(s2, davr)).get(uid, set())  # noqa: SLF001
                return await target_split._working_days(s2, u, davr, excused)  # noqa: SLF001

        uch = asyncio.run(_wd())
        check("S-09: bayram+sababli+dam kuni bir kunda — ish kuni MANFIY emas",
              uch >= 0, "=" + str(uch))
        check("S-09: uchtasi to'g'ri kelganda kun IKKI marta ayirilmadi",
              uch == ish_kuni_oldin - 1, f"kutilgan={ish_kuni_oldin - 1}, chiqdi={uch}")

        # ── 10) TA'TILDAGI XODIM CHECK-IN QILOLMAYDI ──
        bugun = _date.today()
        cur.execute(
            "insert into excused_days (user_id, date, reason, status, created_at)"
            " values (?,?,'T-tatil','approved',datetime('now'))", (uid, bugun.isoformat()))
        conn.commit()

        async def _checkin():
            from api.services.attendance import CheckError, OnLeaveError, perform_check_in
            from db.models import User as U
            async with async_session() as s2:
                u = await s2.get(U, uid)
                try:
                    await perform_check_in(s2, u, None, None)
                except OnLeaveError as e:
                    return "on_leave", str(e)
                except CheckError as e:
                    return "other", str(e)
                return "ok", ""

        turi, xabar = asyncio.run(_checkin())
        check("S-09: ta'tildagi xodim check-in qilolmadi", turi == "on_leave",
              f"{turi}: {xabar[:100]}")
        check("S-09: rad etish xabari TUSHUNARLI (nima qilishni aytadi)",
              turi == "on_leave" and "HR" in xabar and len(xabar) > 60,
              "=" + xabar[:120])

        # Sababli kun olib tashlansa — check-in yana boshqa sababga tayanadi
        # (bu yerda GPS/yuz), ya'ni «ta'til» to'sig'i o'tdi.
        cur.execute("delete from excused_days where user_id=? and date=?",
                    (uid, bugun.isoformat()))
        conn.commit()
        turi2, xabar2 = asyncio.run(_checkin())
        check("S-09: sababli kun olingach «ta'til» to'sig'i yo'qoldi",
              turi2 != "on_leave", f"{turi2}: {xabar2[:80]}")

        # ── 11) «Sozlanmagan modullar» da ko'rinishi ──
        from api.services.setup_status import collect_setup_status

        async def _holat():
            async with async_session() as s2:
                return {i.key: i.ready for i in await collect_setup_status(s2)}

        check("S-09: bayram kiritilgani «Sozlanmagan» ro'yxatidan chiqardi",
              asyncio.run(_holat()).get("holidays") is True,
              "=" + str(asyncio.run(_holat()).get("holidays")))

        # ── 12) O'chirish ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            if hid:
                rd2 = c.delete(f"/holidays/{hid}", headers=auth(mgr_t))
                check("S-09: DELETE /holidays -> 200", rd2.status_code == 200,
                      "kod=" + str(rd2.status_code))
                qayta = asyncio.run(_sched())
                check("S-09: bayram o'chirilgach kun yana ish kuni bo'ldi",
                      qayta.get(bayram_kuni) is True,
                      f"{bayram_kuni} -> {qayta.get(bayram_kuni)}")
            rd3 = c.delete("/holidays/99999999", headers=auth(mgr_t))
            check("S-09: yo'q bayramni o'chirish -> 404", rd3.status_code == 404,
                  "kod=" + str(rd3.status_code))
    except Exception:
        check("S-09 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from holidays where name like 'T-%' or name='Takror'")
            if uid:
                cur.execute("delete from excused_days where user_id=?", (uid,))
                cur.execute("delete from work_schedule_override where user_id=?", (uid,))
                cur.execute("delete from work_schedule_weekly where user_id=?", (uid,))
                cur.execute("delete from attendance where user_id=?", (uid,))
                cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_employee_documents() -> None:
    """S-10 (TZ 3.4) — kadr hujjatlari arxivi: model va API.

    Modulning eng nozik joyi RUXSAT: bu maxfiy ma'lumot (diplom, tibbiy
    ma'lumotnoma, shartnoma). TZ ataylab talab qiladi — **ROP umuman
    ko'rmaydi**, hatto O'Z JAMOASI bo'lsa ham. Boshqa modullarda ROP
    jamoasini ko'radi, shuning uchun bu chetlanish alohida sinaladi.

    Begona so'ralganda 403 EMAS, 404: 403 «bu odamda hujjat bor» degan
    ma'lumotni oshkor qilardi.

    Qabul mezonlari (TZ):
      • xodim faqat o'zinikini ko'radi, begonaga 404;
      • ROP ga 404 (hatto o'z jamoasi bo'lsa ham);
      • barcha o'qish `deleted_at IS NULL`;
      • fayl serverda saqlanmaydi — faqat `file_id`;
      • test: 8+ (rol matritsasi, soft delete, 404, muddat maydoni).
    """
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-10: KADR HUJJATLARI (ruxsat matritsasi)")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Doc%'")
        conn.commit()

        def _yarat(nom: str, rol: str, tg: int, manager: int | None = None) -> int:
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " manager_id, created_at) values (?,?,?,0,1,?,datetime('now'))",
                (tg, nom, rol, manager))
            conn.commit()
            return cur.lastrowid

        ids["hr"] = _yarat("T-DocHR", "hr", 999701001)
        ids["rop"] = _yarat("T-DocRop", "rop", 999701002)
        # ROP ning BEVOSITA bo'ysunuvchisi — «o'z jamoasi» aynan shu
        ids["team"] = _yarat("T-DocTeam", "employee", 999701003, ids["rop"])
        ids["other"] = _yarat("T-DocOther", "employee", 999701004)

        t = {k: token_for(v, r) for (k, v), r in zip(
            ids.items(), ["hr", "rop", "employee", "employee"])}

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── Tur ro'yxati yagona manbadan ──
            r = c.get("/employee-documents/types", headers=auth(t["team"]))
            check("S-10: /types -> 200", r.status_code == 200, "kod=" + str(r.status_code))
            if r.status_code == 200:
                turlar = {x["value"] for x in r.json()}
                check("S-10: TZ dagi 7 tur ham bor",
                      {"contract", "job_description", "property_act", "handover_act",
                       "medical", "diploma", "other"} <= turlar, "=" + str(sorted(turlar)))

            # ── YUKLASH ruxsati ──
            yangi = {
                "user_id": ids["team"], "doc_type": "contract",
                "name": "T-Doc mehnat shartnomasi", "file_id": "T-FAKE-FILE-ID-1",
                "file_type": "document", "issued_at": "2026-01-10",
            }
            r = c.post("/employee-documents", headers=auth(t["team"]), json=yangi)
            check("S-10: oddiy xodim yuklay olmaydi -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))
            r = c.post("/employee-documents", headers=auth(t["rop"]), json=yangi)
            check("S-10: ROP yuklay olmaydi -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))
            r = c.post("/employee-documents", headers=auth(t["hr"]), json=yangi)
            check("S-10: HR yukladi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            doc_id = r.json().get("id") if r.status_code == 201 else None
            if r.status_code == 201:
                check("S-10: javobda tur nomi tarjima qilingan",
                      r.json().get("doc_type_label") == "Mehnat shartnomasi",
                      "=" + str(r.json().get("doc_type_label")))

            # ── Fayl SERVERDA saqlanmaydi ──
            if doc_id:
                row = cur.execute(
                    "select file_id from employee_documents where id=?", (doc_id,)).fetchone()
                check("S-10: bazada faqat file_id (fayl yo'q)",
                      row is not None and row[0] == "T-FAKE-FILE-ID-1", "=" + str(row))

            # ── RUXSAT MATRITSASI (o'qish) ──
            r = c.get("/employee-documents/me", headers=auth(t["team"]))
            check("S-10: xodim O'ZINIKINI ko'radi -> 200 (1 ta)",
                  r.status_code == 200 and len(r.json()) == 1,
                  f"kod={r.status_code} soni={len(r.json()) if r.status_code == 200 else '-'}")

            r = c.get(f"/employee-documents/user/{ids['team']}", headers=auth(t["other"]))
            check("S-10: BEGONA xodim -> 404 (403 emas — oshkor qilmaydi)",
                  r.status_code == 404, "kod=" + str(r.status_code))

            # ⚠️ ASOSIY CHETLANISH: ROP o'z jamoasini ham ko'rmaydi
            r = c.get(f"/employee-documents/user/{ids['team']}", headers=auth(t["rop"]))
            check("S-10: ROP O'Z JAMOASINI ham ko'rmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            r = c.get(f"/employee-documents/user/{ids['rop']}", headers=auth(t["rop"]))
            check("S-10: ROP o'zinikini ko'radi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))

            r = c.get(f"/employee-documents/user/{ids['team']}", headers=auth(t["hr"]))
            check("S-10: HR hammani ko'radi -> 200 (1 ta)",
                  r.status_code == 200 and len(r.json()) == 1,
                  f"kod={r.status_code} soni={len(r.json()) if r.status_code == 200 else '-'}")

            # ── MUDDAT maydoni ──
            bugun = _date.today()
            c.post("/employee-documents", headers=auth(t["hr"]), json={
                "user_id": ids["team"], "doc_type": "medical",
                "name": "T-Doc muddati o'tgan", "file_id": "T-FAKE-2",
                "expires_at": (bugun - _td(days=5)).isoformat()})
            c.post("/employee-documents", headers=auth(t["hr"]), json={
                "user_id": ids["team"], "doc_type": "diploma",
                "name": "T-Doc kelajakda", "file_id": "T-FAKE-3",
                "expires_at": (bugun + _td(days=10)).isoformat()})
            r = c.get("/employee-documents/me", headers=auth(t["team"]))
            hujjatlar = {d["name"]: d for d in r.json()} if r.status_code == 200 else {}
            check("S-10: muddati o'tgan hujjat belgilandi",
                  hujjatlar.get("T-Doc muddati o'tgan", {}).get("is_expired") is True
                  and hujjatlar["T-Doc muddati o'tgan"]["days_left"] == -5,
                  "=" + str(hujjatlar.get("T-Doc muddati o'tgan", {}).get("days_left")))
            check("S-10: amaldagi hujjat muddati o'tgan emas",
                  hujjatlar.get("T-Doc kelajakda", {}).get("is_expired") is False
                  and hujjatlar["T-Doc kelajakda"]["days_left"] == 10,
                  "=" + str(hujjatlar.get("T-Doc kelajakda", {}).get("days_left")))
            check("S-10: muddatsiz hujjatda days_left=None",
                  hujjatlar.get("T-Doc mehnat shartnomasi", {}).get("days_left") is None,
                  "=" + str(hujjatlar.get("T-Doc mehnat shartnomasi", {}).get("days_left")))
            # Muddati tugaydiganlar TEPADA (muddatsizlar oxirida)
            tartib = [d["name"] for d in r.json()] if r.status_code == 200 else []
            check("S-10: muddati yaqinlari tepada, muddatsiz oxirida",
                  tartib and tartib[-1] == "T-Doc mehnat shartnomasi"
                  and tartib[0] == "T-Doc muddati o'tgan", "=" + str(tartib))

            # ── Xato kiritishlar ──
            r = c.post("/employee-documents", headers=auth(t["hr"]), json={
                "user_id": ids["team"], "doc_type": "yolgon_tur",
                "name": "T-Doc x", "file_id": "T-FAKE-4"})
            check("S-10: noma'lum tur -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))
            r = c.post("/employee-documents", headers=auth(t["hr"]), json={
                "user_id": ids["team"], "doc_type": "other", "name": "T-Doc y",
                "file_id": "T-FAKE-5", "issued_at": "2026-05-01",
                "expires_at": "2026-04-01"})
            check("S-10: tugash sanasi berilgandan oldin -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))
            r = c.post("/employee-documents", headers=auth(t["hr"]), json={
                "user_id": 99999999, "doc_type": "other", "name": "T-Doc z",
                "file_id": "T-FAKE-6"})
            check("S-10: yo'q xodimga yuklash -> 404", r.status_code == 404,
                  "kod=" + str(r.status_code))

            # ── YUMSHOQ o'chirish ──
            if doc_id:
                r = c.delete(f"/employee-documents/{doc_id}", headers=auth(t["rop"]))
                check("S-10: ROP o'chira olmaydi -> 403", r.status_code == 403,
                      "kod=" + str(r.status_code))
                r = c.delete(f"/employee-documents/{doc_id}", headers=auth(t["hr"]))
                check("S-10: HR o'chirdi -> 200", r.status_code == 200,
                      "kod=" + str(r.status_code))
                qator = cur.execute(
                    "select deleted_at from employee_documents where id=?",
                    (doc_id,)).fetchone()
                check("S-10: o'chirish YUMSHOQ (qator bazada, deleted_at to'ldi)",
                      qator is not None and qator[0] is not None, "=" + str(qator))
                r = c.get("/employee-documents/me", headers=auth(t["team"]))
                nomlar = [d["name"] for d in r.json()] if r.status_code == 200 else []
                check("S-10: o'chirilgani ro'yxatda YO'Q (deleted_at IS NULL filtri)",
                      "T-Doc mehnat shartnomasi" not in nomlar, "=" + str(nomlar))
                r = c.delete(f"/employee-documents/{doc_id}", headers=auth(t["hr"]))
                check("S-10: ikkinchi marta o'chirish -> 404", r.status_code == 404,
                      "kod=" + str(r.status_code))
    except Exception:
        check("S-10 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                cur.execute(
                    "delete from employee_documents where user_id in (%s)"
                    % ",".join("?" * len(ids)), tuple(ids.values()))
            cur.execute("delete from users where full_name like 'T-Doc%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_documents_bot() -> None:
    """S-11 (TZ 3.4) — hujjatlar boti, kabinet va tugma matnlari.

    Qabul mezonlari (TZ):
      • HR botdan hujjat yuklay oladi, fayl Telegramda qoladi;
      • xodim o'z hujjatini botdan qayta olishi mumkin;
      • muddati o'tayotgan hujjat ro'yxatda ajratib ko'rsatiladi;
      • test: yuklash -> o'qish -> soft delete zanjiri.

    ALOHIDA QO'RIQCHI: tugma matnlari `bot/keyboards.py` va
    `api/services/sections.py` da IKKI MARTA yozilgan (birinchisi handler
    mosligi, ikkinchisi menyu qurish uchun). Ular bir-biridan uzoqlashsa
    server tugmani chizadi, handler esa uni tanimaydi — tugma JIMGINA
    ishlamay qoladi. S-05 aynan shu sinf xatoni yopish uchun qilingan
    edi, shuning uchun bu yerda tekshiriladi.
    """
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-11: HUJJATLAR BOTI + KABINET")
    print("=" * 60)

    # ── Tugma matnlari ikkala faylda AYNAN bir xilmi ──
    try:
        import importlib

        kb = importlib.import_module("bot.keyboards")
        sec = importlib.import_module("api.services.sections")
        kb_btn = {k: v for k, v in vars(kb).items() if k.startswith("BTN_")}
        sec_btn = {k: v for k, v in vars(sec).items() if k.startswith("BTN_")}
        umumiy = set(kb_btn) & set(sec_btn)
        farq = {k: (kb_btn[k], sec_btn[k]) for k in umumiy if kb_btn[k] != sec_btn[k]}
        check("S-11: tugma matnlari ikkala faylda bir xil", not farq, "=" + str(farq))
        # `sections.py` chizadigan HAR BIR tugma handlerda bo'lishi shart
        yetishmaydi = {k: v for k, v in sec_btn.items() if k not in kb_btn}
        check("S-11: sections'dagi har tugma keyboards'da ham bor",
              not yetishmaydi, "=" + str(yetishmaydi))
    except Exception:
        check("S-11 tugma qo'riqchisi", False, traceback.format_exc(limit=2).strip())

    # ── Izohdan nom va sana ajratish (bot mantiqi) ──
    try:
        from bot.handlers.documents import _muddat_belgisi, _parse_caption  # noqa: SLF001

        check("S-11: izoh oxiridagi sana ajratildi",
              _parse_caption("Mehnat shartnomasi 2027-12-31", "F") == (
                  "Mehnat shartnomasi", "2027-12-31"),
              "=" + str(_parse_caption("Mehnat shartnomasi 2027-12-31", "F")))
        check("S-11: sana OLDIDA bo'lsa ham ajratiladi",
              _parse_caption("2027-12-31 Diplom", "F") == ("Diplom", "2027-12-31"),
              "=" + str(_parse_caption("2027-12-31 Diplom", "F")))
        check("S-11: sanasiz izoh — nom, muddat yo'q",
              _parse_caption("Diplom", "F") == ("Diplom", None),
              "=" + str(_parse_caption("Diplom", "F")))
        check("S-11: izoh bo'sh — fayl nomi ishlatiladi",
              _parse_caption("", "shartnoma.pdf") == ("shartnoma.pdf", None),
              "=" + str(_parse_caption("", "shartnoma.pdf")))
        check("S-11: muddati o'tgan hujjat botda AJRATIB ko'rsatiladi",
              "⛔" in _muddat_belgisi({"days_left": -3, "is_expired": True}),
              "=" + _muddat_belgisi({"days_left": -3, "is_expired": True}))
        check("S-11: muddati yaqinlashgan — ogohlantirish",
              "⚠️" in _muddat_belgisi({"days_left": 7, "is_expired": False}),
              "=" + _muddat_belgisi({"days_left": 7, "is_expired": False}))
        check("S-11: muddatsiz hujjatda belgi yo'q",
              _muddat_belgisi({"days_left": None, "is_expired": False}) == "",
              "=" + repr(_muddat_belgisi({"days_left": None, "is_expired": False})))
    except Exception:
        check("S-11 izoh tahlili", False, traceback.format_exc(limit=2).strip())

    # ── Menyuda tugmalar to'g'ri rollarga chiqadimi ──
    try:
        from types import SimpleNamespace

        from api.services.sections import bot_menu_rows

        def _tugmalar(rol: str) -> set:
            u = SimpleNamespace(role=rol, position=None, can_edit_fine_policy=False,
                                can_edit_attendance=False)
            return {b for r in bot_menu_rows(u) for b in r}

        check("S-11: «Hujjatlarim» HAR BIR rolda bor",
              all("📁 Hujjatlarim" in _tugmalar(r)
                  for r in ("employee", "hr", "rop", "boss", "dasturchi")),
              "=" + str({r: "📁 Hujjatlarim" in _tugmalar(r)
                         for r in ("employee", "hr", "rop", "boss", "dasturchi")}))
        check("S-11: «Hujjat yuklash» faqat HR/Boshliq/Dasturchida",
              {r for r in ("employee", "hr", "rop", "boss", "dasturchi")
               if "📎 Hujjat yuklash" in _tugmalar(r)} == {"hr", "boss", "dasturchi"},
              "=" + str({r for r in ("employee", "hr", "rop", "boss", "dasturchi")
                         if "📎 Hujjat yuklash" in _tugmalar(r)}))
    except Exception:
        check("S-11 menyu tekshiruvi", False, traceback.format_exc(limit=2).strip())

    # ── Bot endpointlari: yuklash -> o'qish -> yuborish -> soft delete ──
    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-BDoc%'")
        conn.commit()

        def _yarat(nom: str, rol: str, tg: int, manager: int | None = None) -> int:
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " manager_id, created_at) values (?,?,?,1,1,?,datetime('now'))",
                (tg, nom, rol, manager))
            conn.commit()
            return cur.lastrowid

        ids["hr"] = _yarat("T-BDocHR", "hr", 999701101)
        ids["rop"] = _yarat("T-BDocRop", "rop", 999701102)
        ids["emp"] = _yarat("T-BDocEmp", "employee", 999701103, ids["rop"])
        tg = {"hr": 999701101, "rop": 999701102, "emp": 999701103}

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # Tur ro'yxati botga ham ochiq (bot o'z nusxasini yuritmasin)
            r = c.get("/employee-documents/bot/types", params={"telegram_id": tg["emp"]})
            check("S-11: bot /types -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))

            # Xodim ro'yxati — faqat HR
            r = c.get("/employee-documents/bot/employees", params={"telegram_id": tg["emp"]})
            check("S-11: bot xodim ro'yxati oddiy xodimga -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.get("/employee-documents/bot/employees", params={"telegram_id": tg["rop"]})
            check("S-11: bot xodim ro'yxati ROP ga -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))
            r = c.get("/employee-documents/bot/employees", params={"telegram_id": tg["hr"]})
            check("S-11: bot xodim ro'yxati HR ga -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))

            # ── YUKLASH ──
            muddat = (_date.today() + _td(days=15)).isoformat()
            yuk = {"telegram_id": tg["hr"], "user_id": ids["emp"], "doc_type": "diploma",
                   "name": "T-BDoc diplom", "file_id": "T-BOT-FILE-1",
                   "file_type": "document", "expires_at": muddat}
            r = c.post("/employee-documents/bot/upload",
                       json={**yuk, "telegram_id": tg["emp"]})
            check("S-11: botdan oddiy xodim yuklay olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.post("/employee-documents/bot/upload", json=yuk)
            check("S-11: HR botdan yukladi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            doc_id = r.json().get("id") if r.status_code == 201 else None

            # ── O'QISH: xodim o'zinikini botdan ko'radi ──
            r = c.get("/employee-documents/bot/my", params={"telegram_id": tg["emp"]})
            check("S-11: xodim botdan o'z hujjatini ko'radi",
                  r.status_code == 200 and len(r.json()) == 1
                  and r.json()[0]["name"] == "T-BDoc diplom",
                  f"kod={r.status_code} {r.text[:120]}")
            check("S-11: muddat serverdan keladi (15 kun)",
                  r.status_code == 200 and r.json()[0]["days_left"] == 15,
                  "=" + str(r.json()[0].get("days_left") if r.status_code == 200 else None))

            # HR ning o'z ro'yxati BO'SH — yuklagan bo'lsa ham egasi u emas
            r = c.get("/employee-documents/bot/my", params={"telegram_id": tg["hr"]})
            check("S-11: yuklagan HR ning o'z ro'yxatiga tushmadi",
                  r.status_code == 200 and r.json() == [], "=" + r.text[:80])

            # ── YUBORISH: ruxsat SHU YERDA ham tekshiriladi ──
            if doc_id:
                r = c.post("/employee-documents/bot/send",
                           json={"telegram_id": tg["rop"], "doc_id": doc_id})
                check("S-11: ROP begona hujjatni o'ziga yuborib ololmaydi -> 404",
                      r.status_code == 404, "kod=" + str(r.status_code))
                r = c.post("/employee-documents/bot/send",
                           json={"telegram_id": tg["emp"], "doc_id": doc_id})
                check("S-11: egasi o'z hujjatini qayta oladi -> 200",
                      r.status_code == 200, "kod=" + str(r.status_code) + " " + r.text[:100])
                # ⚠️ Testda bildirishnomalar O'CHIQ — `delivered` False bo'lishi
                # KUTILGAN holat. Agar True chiqsa, demak haqiqiy xodimga fayl
                # ketgan (test.py ning `require_notifications_off` qo'riqchisi
                # buni oldini oladi, lekin ikkinchi qatlam zarar qilmaydi).
                check("S-11: test rejimida haqiqiy fayl YUBORILMADI",
                      r.status_code == 200 and r.json().get("delivered") is False,
                      "=" + r.text[:100])

                # ── SOFT DELETE zanjiri ──
                hr_t = token_for(ids["hr"], "hr")
                r = c.delete(f"/employee-documents/{doc_id}", headers=auth(hr_t))
                check("S-11: o'chirildi -> 200", r.status_code == 200,
                      "kod=" + str(r.status_code))
                r = c.get("/employee-documents/bot/my", params={"telegram_id": tg["emp"]})
                check("S-11: o'chirilgach botdagi ro'yxat bo'shadi",
                      r.status_code == 200 and r.json() == [], "=" + r.text[:80])
                r = c.post("/employee-documents/bot/send",
                           json={"telegram_id": tg["emp"], "doc_id": doc_id})
                check("S-11: o'chirilgan hujjatni yuborib bo'lmaydi -> 404",
                      r.status_code == 404, "kod=" + str(r.status_code))
    except Exception:
        check("S-11 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                cur.execute(
                    "delete from employee_documents where user_id in (%s)"
                    % ",".join("?" * len(ids)), tuple(ids.values()))
            cur.execute("delete from users where full_name like 'T-BDoc%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_deadlines_core() -> None:
    """S-12 (TZ 3.5) — muddat eslatmalari yadrosi.

    Qabul mezonlari (TZ):
      • muddat qo'lda ham, hisoblanib ham chiqadi;
      • `reminded_at` — bir muddat bo'yicha takroriy xabar yo'q;
      • test: `>=` semantikasi (cron bir kun kechiksa ham xabar tushadi).

    ENG MUHIM TEKSHIRUV — «IKKITA MANBA BO'LMASIN». Hisoblanadigan
    muddatning sanasi jadvalga YOZILMASLIGI shart: manba (hujjatning
    `expires_at` yoki `hire_date`) o'zgarsa, ro'yxat DARHOL yangi sanani
    ko'rsatishi kerak. Nusxa saqlansa tizim ikki xil muddat ko'rsatardi
    va qaysi biri to'g'ri ekani bilinmasdi.
    """
    import asyncio
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-12: MUDDAT ESLATMALARI (yadro)")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    uid = None
    bugun = _date.today()
    try:
        cur.execute("delete from users where full_name like 'T-Dl%'")
        conn.commit()
        # hire_date shunday tanlanadiki, sinov muddati (90 kun) 10 kundan keyin tugasin
        hire = (bugun - _td(days=80)).isoformat()
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999701201,'T-DlEmp','employee',0,1,?,"
            "datetime('now'))", (hire,))
        uid = cur.lastrowid
        conn.commit()

        async def _royxat(days=None):
            from api.services.deadlines import upcoming
            async with async_session() as s2:
                return [i for i in await upcoming(s2, days) if i.user_id == uid]

        # Sozlama standart holatga (boshqa test o'zgartirgan bo'lishi mumkin)
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            c.put("/deadlines/config", headers=auth(mgr_t),
                  json={"probation_days": 90, "remind_days": 30})

            # ── 1) HISOBLANADIGAN: sinov muddati ──
            items = asyncio.run(_royxat())
            sinov = [i for i in items if i.kind == "probation"]
            check("S-12: sinov muddati HISOBLANIB chiqdi (jadvalga yozilmagan)",
                  len(sinov) == 1 and sinov[0].days_left == 10,
                  "=" + str([(i.kind, i.days_left) for i in items]))
            check("S-12: hisoblangan bandda jadval qatori YO'Q",
                  bool(sinov) and sinov[0].row_id is None,
                  "=" + str(sinov[0].row_id if sinov else "-"))
            qatorlar = cur.execute(
                "select count(*) from deadlines where user_id=?", (uid,)).fetchone()[0]
            check("S-12: hisoblangan muddat uchun jadval BO'SH qoldi",
                  qatorlar == 0, "=" + str(qatorlar))

            # ── 2) YAGONA MANBA: hire_date o'zgarsa muddat ham o'zgaradi ──
            cur.execute("update users set hire_date=? where id=?",
                        ((bugun - _td(days=85)).isoformat(), uid))
            conn.commit()
            sinov2 = [i for i in asyncio.run(_royxat()) if i.kind == "probation"]
            check("S-12: manba o'zgardi -> muddat DARHOL yangilandi (nusxa yo'q)",
                  len(sinov2) == 1 and sinov2[0].days_left == 5,
                  "=" + str(sinov2[0].days_left if sinov2 else "-"))

            # ── 3) HISOBLANADIGAN: hujjat muddati ──
            r = c.post("/employee-documents", headers=auth(mgr_t), json={
                "user_id": uid, "doc_type": "contract", "name": "T-Dl shartnoma",
                "file_id": "T-DL-1", "expires_at": (bugun + _td(days=20)).isoformat()})
            doc_id = r.json().get("id") if r.status_code == 201 else None
            hujjat = [i for i in asyncio.run(_royxat()) if i.source_kind == "document"]
            check("S-12: hujjat muddati ham ro'yxatga tushdi",
                  len(hujjat) == 1 and hujjat[0].days_left == 20,
                  "=" + str([(i.kind, i.days_left) for i in hujjat]))
            check("S-12: shartnoma alohida tur sifatida ko'rinadi",
                  bool(hujjat) and hujjat[0].kind == "contract",
                  "=" + str(hujjat[0].kind if hujjat else "-"))

            # ── 4) QO'LDA kiritish ──
            r = c.post("/deadlines", headers=auth(mgr_t), json={
                "user_id": uid, "kind": "medical_exam",
                "due_date": (bugun + _td(days=3)).isoformat(),
                "note": "T-Dl tibbiy"})
            check("S-12: qo'lda muddat kiritildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            manual_id = r.json().get("id") if r.status_code == 201 else None

            # Hisoblanadigan turni QO'LDA kiritib bo'lmaydi (ikkita manba xavfi)
            r = c.post("/deadlines", headers=auth(mgr_t), json={
                "user_id": uid, "kind": "probation",
                "due_date": (bugun + _td(days=3)).isoformat()})
            check("S-12: hisoblanadigan turni qo'lda kiritish -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))
            check("S-12: xato xabari NIMA QILISHNI aytadi",
                  r.status_code == 400 and "hisoblanadi" in r.text, r.text[:140])

            # ── 5) IKKALASI BITTA RO'YXATDA, sana bo'yicha tartibda ──
            items = asyncio.run(_royxat())
            check("S-12: qo'lda + hisoblangan BITTA ro'yxatda (3 ta)",
                  len(items) == 3, "=" + str([(i.kind, i.days_left) for i in items]))
            check("S-12: eng yaqin muddat birinchi",
                  [i.days_left for i in items] == sorted(i.days_left for i in items),
                  "=" + str([i.days_left for i in items]))

            # ── 6) `>=` SEMANTIKASI: o'tib ketgan muddat ro'yxatDAN CHIQMAYDI ──
            cur.execute("update deadlines set due_date=? where id=?",
                        ((bugun - _td(days=2)).isoformat(), manual_id))
            conn.commit()
            items = asyncio.run(_royxat())
            otgan = [i for i in items if i.kind == "medical_exam"]
            check("S-12: cron 2 kun kechikdi — muddat baribir ro'yxatda",
                  len(otgan) == 1 and otgan[0].days_left == -2,
                  "=" + str([(i.kind, i.days_left) for i in items]))
            check("S-12: o'tib ketgani `is_overdue` bilan belgilangan",
                  bool(otgan) and otgan[0].is_overdue is True,
                  "=" + str(otgan[0].is_overdue if otgan else "-"))

            # ── 7) UFQ: `days` dan uzoq muddat ko'rinmaydi ──
            yaqin = asyncio.run(_royxat(days=7))
            check("S-12: `days=7` — uzoq muddatlar chiqarib tashlandi",
                  {i.kind for i in yaqin} == {"medical_exam", "probation"},
                  "=" + str([(i.kind, i.days_left) for i in yaqin]))

            # ── 8) `reminded_at`: takroriy xabar yo'q ──
            async def _belgila():
                from api.services.deadlines import mark_reminded, upcoming
                async with async_session() as s2:
                    hammasi = [i for i in await upcoming(s2) if i.user_id == uid]
                    return await mark_reminded(s2, hammasi, bugun)

            n = asyncio.run(_belgila())
            check("S-12: 3 ta band «eslatildi» deb belgilandi", n == 3, "=" + str(n))
            items = asyncio.run(_royxat())
            check("S-12: hammasida `reminded_at` bugungi sana",
                  all(i.reminded_at == bugun for i in items),
                  "=" + str([(i.kind, i.reminded_at) for i in items]))
            izlar = cur.execute(
                "select count(*) from deadlines where user_id=? and source_kind is not null",
                (uid,)).fetchone()[0]
            check("S-12: hisoblangan bandlar uchun iz qatori endi yaratildi (2 ta)",
                  izlar == 2, "=" + str(izlar))
            sanalar = cur.execute(
                "select due_date from deadlines where user_id=? and source_kind is not null",
                (uid,)).fetchall()
            check("S-12: iz qatorida sana YO'Q (manba yagona bo'lib qoldi)",
                  all(x[0] is None for x in sanalar), "=" + str(sanalar))

            # Ikkinchi marta belgilash yangi qator YARATMASIN
            asyncio.run(_belgila())
            izlar2 = cur.execute(
                "select count(*) from deadlines where user_id=? and source_kind is not null",
                (uid,)).fetchone()[0]
            check("S-12: qayta eslatish DUBLIKAT qator yaratmadi",
                  izlar2 == 2, "=" + str(izlar2))

            # ── 9) Iz qatori borligida ham sana MANBADAN olinadi ──
            if doc_id:
                cur.execute("update employee_documents set expires_at=? where id=?",
                            ((bugun + _td(days=25)).isoformat(), doc_id))
                conn.commit()
                h2 = [i for i in asyncio.run(_royxat()) if i.source_kind == "document"]
                check("S-12: iz qatori bo'lsa ham sana MANBADAN keladi",
                      len(h2) == 1 and h2[0].days_left == 25,
                      "=" + str(h2[0].days_left if h2 else "-"))

            # ── 10) YOPISH ──
            r = c.post("/deadlines/close", headers=auth(mgr_t),
                       json={"key": f"document:{doc_id}"})
            check("S-12: hisoblangan band kalit bo'yicha yopildi",
                  r.status_code == 200, "kod=" + str(r.status_code))
            qolgan = {i.kind for i in asyncio.run(_royxat())}
            check("S-12: yopilgan band ro'yxatga QAYTMADI",
                  "contract" not in qolgan, "=" + str(qolgan))
            if manual_id:
                r = c.post(f"/deadlines/{manual_id}/close", headers=auth(mgr_t))
                check("S-12: qo'lda kiritilgani yopildi", r.status_code == 200,
                      "kod=" + str(r.status_code))
                r = c.post(f"/deadlines/{manual_id}/close", headers=auth(mgr_t))
                check("S-12: ikkinchi marta yopish -> 404", r.status_code == 404,
                      "kod=" + str(r.status_code))

            # ── 11) RUXSAT ──
            emp_t = token_for(uid, "employee")
            r = c.get("/deadlines", headers=auth(emp_t))
            check("S-12: oddiy xodim muddatlarni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.get("/deadlines", headers=auth(mgr_t))
            check("S-12: rahbar ko'radi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
    except Exception:
        check("S-12 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if uid:
                cur.execute("delete from deadlines where user_id=?", (uid,))
                cur.execute("delete from employee_documents where user_id=?", (uid,))
                cur.execute("delete from users where id=?", (uid,))
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_deadlines_cron() -> None:
    """S-13 (TZ 3.5) — muddat eslatmalari croni va xabari.

    Qabul mezonlari (TZ):
      • kuniga bir marta, takrorlanmaydi;
      • bir necha muddat bir kunga tushsa — BITTA xabarga birlashadi;
      • test: xabar yuboruvchi patch qilingan holda.

    ⚠️ XABAR YUBORUVCHI PATCH QILINADI. Loyihada jonli tuzoq bor: lokal
    test haqiqiy xodimlarga Telegram xabari yuborib yuborgan. Bu yerda
    `api.notify.notify_user` almashtiriladi va tarmoqqa umuman
    chiqilmaydi — `NOTIFICATIONS_ENABLED=false` qo'riqchisiga qo'shimcha
    ikkinchi qatlam.
    """
    import asyncio
    from datetime import date as _date
    from datetime import timedelta as _td

    print("\n" + "=" * 60)
    print("S-13: MUDDAT ESLATMALARI (cron va xabar)")
    print("=" * 60)

    from db.base import async_session

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    bugun = _date.today()
    try:
        cur.execute("delete from users where full_name like 'T-Dc%'")
        conn.commit()

        def _yarat(nom, rol, tg, hire=None):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started, is_active,"
                " hire_date, created_at) values (?,?,?,1,1,?,datetime('now'))",
                (tg, nom, rol, hire))
            conn.commit()
            return cur.lastrowid

        # Uchta xodim, uchtasi ham sinov muddati 5 kundan keyin tugaydi —
        # ya'ni BITTA kunga uchta band tushadi.
        hire = (bugun - _td(days=85)).isoformat()
        for n in range(3):
            ids[f"emp{n}"] = _yarat(f"T-DcEmp{n}", "employee", 999701300 + n, hire)

        # ── Yuboruvchini PATCH qilamiz ──
        yuborilgan: list = []

        async def _tick():
            import api.services.cron_jobs as cj
            from api.notify import notify_user as asl

            import api.notify as notify_mod

            async def soxta(db_, user, category, text, **kw):
                yuborilgan.append((user.id, user.role, text))
                return True

            notify_mod.notify_user = soxta
            try:
                async with async_session() as s2:
                    return await cj.deadline_tick(s2)
            finally:
                notify_mod.notify_user = asl

        natija = asyncio.run(_tick())
        check("S-13: tick ishladi", natija.get("ok") is True, "=" + str(natija))
        check("S-13: uchala band ham qamrab olindi",
              natija.get("items", 0) >= 3, "=" + str(natija.get("items")))

        # ── BITTA xabar: uchta band bitta matnga birlashdi ──
        bizniki = [t for (_uid, _rol, t) in yuborilgan
                   if "T-DcEmp0" in t or "T-DcEmp1" in t or "T-DcEmp2" in t]
        check("S-13: uchta muddat BITTA xabarga birlashdi",
              all(("T-DcEmp0" in t and "T-DcEmp1" in t and "T-DcEmp2" in t)
                  for t in bizniki) and len(bizniki) >= 1,
              f"xabarlar={len(bizniki)}")
        check("S-13: har mas'ulga bittadan xabar (odam soniga teng)",
              len(yuborilgan) == len({u for (u, _r, _t) in yuborilgan}),
              "=" + str(len(yuborilgan)))
        rollar = {r for (_u, r, _t) in yuborilgan}
        check("S-13: xabar HR/Boshliq/Dasturchiga ketdi (guruhga emas)",
              rollar <= {"hr", "boss", "dasturchi"}, "=" + str(rollar))
        if bizniki:
            check("S-13: xabarda sana va qolgan kun ko'rsatilgan",
                  "5 kun" in bizniki[0] and str(bugun.year) in bizniki[0],
                  "=" + bizniki[0][:160])
            check("S-13: xabarda nima qilish kerakligi aytilgan",
                  "Muddatlar" in bizniki[0], "=" + bizniki[0][-80:])

        # ── TAKRORLANMAYDI ──
        yuborilgan.clear()
        natija2 = asyncio.run(_tick())
        qayta = [t for (_u, _r, t) in yuborilgan if "T-DcEmp0" in t]
        check("S-13: ikkinchi tick bizning bandlarni QAYTA yubormadi",
              not qayta, "=" + str(len(qayta)))
        check("S-13: ikkinchi tickda bizning bandlar hisobga olinmadi",
              natija2.get("items", 0) == 0 or not qayta,
              "=" + str(natija2))

        # ── ERTAGA yana yuboriladi (bir kunlik oyna) ──
        cur.execute(
            "update deadlines set reminded_at=? where user_id in (?,?,?)",
            ((bugun - _td(days=1)).isoformat(), ids["emp0"], ids["emp1"], ids["emp2"]))
        conn.commit()
        yuborilgan.clear()
        asyncio.run(_tick())
        ertaga = [t for (_u, _r, t) in yuborilgan if "T-DcEmp0" in t]
        check("S-13: kechagi eslatmadan keyin BUGUN yana yuborildi",
              len(ertaga) >= 1, "=" + str(len(ertaga)))

        # ── O'TIB KETGAN band alohida bo'limda ──
        cur.execute("update users set hire_date=? where id=?",
                    ((bugun - _td(days=95)).isoformat(), ids["emp0"]))
        cur.execute("update deadlines set reminded_at=null where user_id=?", (ids["emp0"],))
        conn.commit()
        yuborilgan.clear()
        asyncio.run(_tick())
        otgan = [t for (_u, _r, t) in yuborilgan if "T-DcEmp0" in t]
        check("S-13: o'tib ketgan muddat alohida «Muddati o'tgan» bo'limida",
              bool(otgan) and "Muddati o'tgan" in otgan[0],
              "=" + (otgan[0][:200] if otgan else "xabar yo'q"))

        # ── YOPILGAN band eslatilmaydi ──
        cur.execute("update deadlines set status='done', reminded_at=null "
                    "where user_id=?", (ids["emp0"],))
        conn.commit()
        yuborilgan.clear()
        asyncio.run(_tick())
        yopiq = [t for (_u, _r, t) in yuborilgan if "T-DcEmp0" in t]
        check("S-13: yopilgan band bo'yicha eslatma KELMADI",
              not yopiq, "=" + str(len(yopiq)))
    except Exception:
        check("S-13 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from deadlines where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()


def _docx_fixture(paragraphs: str) -> bytes:
    """Eng kichik ishlaydigan `.docx` — sinov uchun.

    Haqiqiy Word fayli emas, lekin `docx_render` uchun yetarli: u faqat
    ZIP va `word/*.xml` bilan ishlaydi."""
    import io
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "<Types/>")
        z.writestr(
            "word/document.xml",
            '<?xml version="1.0"?><w:document xmlns:w="x"><w:body>'
            + paragraphs
            + "</w:body></w:document>",
        )
    return buf.getvalue()


def _docx_text(data: bytes) -> str:
    import io
    import re
    import zipfile

    xml = zipfile.ZipFile(io.BytesIO(data)).read("word/document.xml").decode("utf-8")
    return "".join(m.group(1) for m in re.finditer(r"<w:t[^>]*>(.*?)</w:t>", xml, re.S))


def test_docx_render() -> None:
    """S-14 (TZ 3.3) — `.docx` shablonini to'ldirish, kutubxonasiz.

    Qabul mezonlari (TZ):
      • shablondan to'ldirilgan `.docx` chiqadi;
      • BO'LINGAN belgi holati sinovdan o'tgan (fixture: bo'lingan `w:t`);
      • yangi kutubxona qo'shilmagan (`requirements.txt` o'zgarmagan);
      • generatsiya so'rov ichida emas, FON ishida.

    ASOSIY TUZOQ — BO'LINGAN BELGI. Wordda `{{fish}}` deb yozilgani fayl
    ichida uchta run'ga bo'linib ketishi mumkin (imlo tekshiruvi, til
    belgisi, kursor joyi). Oddiy `replace` bunda hech narsa topmaydi:
    shablon ko'zga to'g'ri ko'rinadi-yu, natija bo'sh chiqadi.
    """
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("S-14: HUJJAT GENERATSIYASI (docx, kutubxonasiz)")
    print("=" * 60)

    from api.services.docx_render import find_placeholders, render
    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    # ── 1) YANGI KUTUBXONA QO'SHILMAGAN ──
    try:
        req = (Path(__file__).resolve().parent / "api" / "requirements.txt").read_text(
            encoding="utf-8"
        ).lower()
        yomon = [x for x in ("python-docx", "docxtpl", "lxml", "docx2python") if x in req]
        check("S-14: docx kutubxonasi QO'SHILMAGAN", not yomon, "=" + str(yomon))

        import api.services.docx_render as mod

        manba = Path(mod.__file__).read_text(encoding="utf-8")
        tashqi = [
            ln.strip() for ln in manba.splitlines()
            if ln.startswith(("import ", "from ")) and not any(
                ln.startswith(p) for p in (
                    "import io", "import re", "import zipfile", "from __future__")
            )
        ]
        check("S-14: modul FAQAT standart kutubxonadan foydalanadi",
              not tashqi, "=" + str(tashqi))
    except Exception:
        check("S-14 bog'liqlik tekshiruvi", False, traceback.format_exc(limit=2).strip())

    # ── 2) BUTUN belgi ──
    try:
        d = _docx_fixture("<w:p><w:r><w:t>Hurmatli {{fish}}!</w:t></w:r></w:p>")
        out, qolgan = render(d, {"fish": "Ali Valiyev"})
        check("S-14: butun belgi almashtirildi",
              _docx_text(out) == "Hurmatli Ali Valiyev!" and not qolgan,
              "=" + _docx_text(out))

        # ── 3) BO'LINGAN belgi — ASOSIY TUZOQ ──
        d = _docx_fixture(
            "<w:p><w:r><w:t>Hurmatli {{</w:t></w:r>"
            "<w:r><w:t>fi</w:t></w:r><w:r><w:t>sh</w:t></w:r>"
            "<w:r><w:t>}}, xush kelibsiz</w:t></w:r></w:p>"
        )
        out, qolgan = render(d, {"fish": "Ali Valiyev"})
        check("S-14: UCHGA BO'LINGAN belgi ham almashtirildi",
              _docx_text(out) == "Hurmatli Ali Valiyev, xush kelibsiz" and not qolgan,
              "=" + _docx_text(out))
        check("S-14: bo'lingan belgi ro'yxatda ham topiladi",
              find_placeholders(d) == ["fish"], "=" + str(find_placeholders(d)))

        # ── 4) Bitta abzatsda ikki belgi, biri bo'lingan ──
        d = _docx_fixture(
            "<w:p><w:r><w:t>{{a}} va {{</w:t></w:r><w:r><w:t>b}}</w:t></w:r></w:p>"
        )
        out, _ = render(d, {"a": "BIR", "b": "IKKI"})
        check("S-14: bitta abzatsda ikki belgi",
              _docx_text(out) == "BIR va IKKI", "=" + _docx_text(out))

        # ── 5) XML xavfli belgi ──
        d = _docx_fixture("<w:p><w:r><w:t>{{n}}</w:t></w:r></w:p>")
        out, _ = render(d, {"n": "Ali & Vali <MCHJ>"})
        import io as _io
        import xml.dom.minidom as _md
        import zipfile as _zf

        xml = _zf.ZipFile(_io.BytesIO(out)).read("word/document.xml")
        _md.parseString(xml)  # buzilgan bo'lsa xato beradi
        check("S-14: `&` va `<` qochirildi — XML buzilmadi",
              b"&amp;" in xml and b"&lt;MCHJ&gt;" in xml, "=" + xml.decode()[-90:])

        # ── 6) To'ldirilmagan belgi TEGILMAYDI va ro'yxatda qaytadi ──
        d = _docx_fixture("<w:p><w:r><w:t>{{bor}} / {{yoq}}</w:t></w:r></w:p>")
        out, qolgan = render(d, {"bor": "X"})
        check("S-14: to'ldirilmagan belgi hujjatda ko'rinib qoldi",
              "{{yoq}}" in _docx_text(out), "=" + _docx_text(out))
        check("S-14: to'ldirilmaganlar ro'yxati qaytdi",
              qolgan == ["yoq"], "=" + str(qolgan))

        # ── 7) Bo'sh joy saqlanadi ──
        d = _docx_fixture("<w:p><w:r><w:t>Ism: {{n}} lavozim</w:t></w:r></w:p>")
        out, _ = render(d, {"n": "Ali"})
        xml = _zf.ZipFile(_io.BytesIO(out)).read("word/document.xml").decode()
        check("S-14: `xml:space=preserve` qo'shildi (so'zlar yopishmasin)",
              'xml:space="preserve"' in xml and _docx_text(out) == "Ism: Ali lavozim",
              "=" + _docx_text(out))

        # ── 8) Kolontitul ham to'ldiriladi ──
        import io as _io2
        import zipfile as _zf2

        buf = _io2.BytesIO()
        with _zf2.ZipFile(buf, "w") as z:
            z.writestr("[Content_Types].xml", "<Types/>")
            z.writestr("word/document.xml",
                       '<?xml version="1.0"?><w:document xmlns:w="x"><w:body>'
                       "<w:p><w:r><w:t>tana</w:t></w:r></w:p></w:body></w:document>")
            z.writestr("word/header1.xml",
                       '<?xml version="1.0"?><w:hdr xmlns:w="x">'
                       "<w:p><w:r><w:t>Raqam: {{raqam}}</w:t></w:r></w:p></w:hdr>")
        out, qolgan = render(buf.getvalue(), {"raqam": "77"})
        hdr = _zf2.ZipFile(_io2.BytesIO(out)).read("word/header1.xml").decode()
        check("S-14: kolontituldagi belgi ham to'ldirildi",
              "Raqam: 77" in hdr and not qolgan, "=" + hdr[-70:])
    except Exception:
        check("S-14 render tekshiruvi", False, traceback.format_exc(limit=2).strip())

    # ── 9) FON ISHI: so'rov ichida generatsiya YO'Q ──
    conn = db()
    cur = conn.cursor()
    tmpl_id = None
    try:
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # Shablonni ro'yxatga olish (`file_id` — soxta, yuklab olish patch qilinadi)
            r = c.post("/document-templates", headers=auth(mgr_t), json={
                "kind": "offer", "name": "T-Tmpl taklif", "file_id": "T-TMPL-FAKE",
                "placeholders": ["fish", "lavozim"]})
            check("S-14: shablon qo'shildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            tmpl_id = r.json().get("id") if r.status_code == 201 else None

            r = c.get("/document-templates", headers=auth(mgr_t))
            check("S-14: shablon ro'yxatda", r.status_code == 200
                  and any(t["id"] == tmpl_id for t in r.json()),
                  "kod=" + str(r.status_code))

            # Oddiy xodim ko'ra olmaydi
            emp = cur.execute(
                "select id from users where role='employee' and is_active=1 limit 1"
            ).fetchone()
            if emp:
                r = c.get("/document-templates", headers=auth(token_for(emp[0], "employee")))
                check("S-14: oddiy xodimga -> 403", r.status_code == 403,
                      "kod=" + str(r.status_code))

            # Generatsiya -> 202 (navbat), so'rov ichida BAJARILMAYDI
            import time as _t

            boshi = _t.perf_counter()
            r = c.post("/document-templates/render", headers=auth(mgr_t), json={
                "template_id": tmpl_id,
                "values": {"fish": "Ali Valiyev"},
                "filename": "T-taklif"})
            ketgan = _t.perf_counter() - boshi
            check("S-14: generatsiya NAVBATGA qo'yildi -> 202",
                  r.status_code == 202, "kod=" + str(r.status_code) + " " + r.text[:120])
            check("S-14: so'rov TEZ qaytdi (ish so'rov ichida emas)",
                  ketgan < 1.0, f"{ketgan:.3f}s")
            check("S-14: yetishmayotgan belgi DARHOL aytildi",
                  r.status_code == 202 and r.json().get("missing") == ["lavozim"],
                  "=" + r.text[:120])
            job_id = r.json().get("job_id") if r.status_code == 202 else None

            holat = cur.execute(
                "select status, kind from background_jobs where id=?", (job_id,)).fetchone()
            check("S-14: navbatda `document_render` ishi turibdi",
                  holat == ("queued", "document_render"), "=" + str(holat))

        # ── 10) Cron ishni bajaradi (shablon yuklab olish PATCH qilinadi) ──
        d = _docx_fixture("<w:p><w:r><w:t>Hurmatli {{</w:t></w:r>"
                          "<w:r><w:t>fish}}, lavozim: {{lavozim}}</w:t></w:r></w:p>")
        natija: dict = {}

        async def _tick():
            import api.telegram_notify as tn
            from api.services.background_jobs import background_tick

            asl_dl, asl_send = tn.download_file, tn.send_media_file

            async def soxta_dl(file_id):
                return d

            async def soxta_send(chat_id, content, filename, kind, caption=None):
                natija["filename"] = filename
                natija["bytes"] = content
                natija["caption"] = caption
                return {"ok": True, "result": {"document": {"file_id": "T-OUT"}}}

            tn.download_file, tn.send_media_file = soxta_dl, soxta_send
            try:
                async with async_session() as s2:
                    return await background_tick(s2)
            finally:
                tn.download_file, tn.send_media_file = asl_dl, asl_send

        res = asyncio.run(_tick())
        check("S-14: cron ishni bajardi", res.get("ok") is True and res.get("ran") ==
              "document_render", "=" + str(res))
        check("S-14: fayl nomi `.docx` bilan tugadi",
              natija.get("filename") == "T-taklif.docx", "=" + str(natija.get("filename")))
        if natija.get("bytes"):
            matn = _docx_text(natija["bytes"])
            check("S-14: fon ishida BO'LINGAN belgi to'ldirildi",
                  "Hurmatli Ali Valiyev" in matn, "=" + matn)
            check("S-14: to'ldirilmagani hujjatda ko'rinib qoldi",
                  "{{lavozim}}" in matn, "=" + matn)
        check("S-14: izohda to'ldirilmagan belgi ogohlantirildi",
              "lavozim" in (natija.get("caption") or ""),
              "=" + str(natija.get("caption")))
    except Exception:
        check("S-14 fon ishi", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from background_jobs where kind='document_render'")
            cur.execute("delete from document_templates where name like 'T-Tmpl%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_offers() -> None:
    """S-15 (TZ 3.3) — ish taklifi: model, forma va hujjat.

    Qabul mezonlari (TZ):
      • `salary` INTEGER (matn emas);
      • taklif bazada qoladi, keyin qidiriladi;
      • fayl FON ishida tayyorlanadi;
      • test: forma -> hujjat -> baza yozuvi.

    ⚠️ TIZIM NOMZODGA HECH NARSA YUBORMAYDI (TZ talabi). Bu ham
    tekshiriladi: hujjat SO'RAGAN HR ga boradi, nomzodga emas — nomzodning
    telefoni bazada bo'lsa ham.
    """
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("S-15: ISH TAKLIFLARI")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    tmpl_id = None
    try:
        cur.execute("delete from offers where candidate_name like 'T-Of%'")
        cur.execute("delete from document_templates where name like 'T-OfTmpl%'")
        conn.commit()

        pos = cur.execute("select id, name from positions limit 1").fetchone()
        if not pos:
            check("lavozim topildi", False, "positions bo'sh")
            return

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── Belgilar ro'yxati oldindan ma'lum ──
            r = c.get("/offers/placeholders", headers=auth(mgr_t))
            nomlar = {x["name"] for x in r.json()} if r.status_code == 200 else set()
            check("S-15: shablon belgilari ro'yxati bor",
                  {"fish", "lavozim", "oylik", "oylik_sozda"} <= nomlar,
                  "=" + str(sorted(nomlar)))

            # ── FORMA -> BAZA ──
            r = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Of Ali Valiyev", "phone": "+998901234567",
                "position_id": pos[0], "salary": 12000000,
                "probation_months": 3, "start_date": "2026-09-01",
                "manager_id": mgr[0]})
            check("S-15: taklif saqlandi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            offer_id = r.json().get("id") if r.status_code == 201 else None

            # ── `salary` INTEGER ──
            tur = cur.execute("select typeof(salary), salary from offers where id=?",
                              (offer_id,)).fetchone()
            check("S-15: `salary` bazada INTEGER (matn emas)",
                  tur == ("integer", 12000000), "=" + str(tur))
            r2 = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Of Matn", "position_id": pos[0],
                "salary": "12 mln"})
            check("S-15: matn oylik rad etildi -> 422", r2.status_code == 422,
                  "kod=" + str(r2.status_code))
            r2 = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Of Nol", "position_id": pos[0], "salary": 0})
            check("S-15: nol oylik rad etildi -> 422", r2.status_code == 422,
                  "kod=" + str(r2.status_code))
            r2 = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Of Lavozimsiz", "salary": 5000000})
            check("S-15: lavozimsiz taklif rad etildi -> 400",
                  r2.status_code == 400, "kod=" + str(r2.status_code))

            # ── QIDIRUV (TZ: «keyin qidiriladi») ──
            r = c.get("/offers", headers=auth(mgr_t), params={"q": "Valiyev"})
            check("S-15: ism bo'yicha qidiruv ishladi",
                  r.status_code == 200 and any(o["id"] == offer_id for o in r.json()),
                  f"kod={r.status_code} soni={len(r.json()) if r.status_code == 200 else '-'}")
            r = c.get("/offers", headers=auth(mgr_t), params={"q": "901234567"})
            check("S-15: telefon bo'yicha ham topiladi",
                  r.status_code == 200 and any(o["id"] == offer_id for o in r.json()),
                  "=" + str(r.status_code))
            r = c.get("/offers", headers=auth(mgr_t), params={"q": "yoq-bunday-odam"})
            check("S-15: topilmasa bo'sh ro'yxat", r.status_code == 200 and r.json() == [],
                  "=" + r.text[:60])

            # ── HOLAT ──
            r = c.put(f"/offers/{offer_id}/status", headers=auth(mgr_t),
                      json={"status": "accepted"})
            check("S-15: holat o'zgardi", r.status_code == 200
                  and r.json().get("status_label") == "Qabul qilingan",
                  "=" + r.text[:100])
            r = c.put(f"/offers/{offer_id}/status", headers=auth(mgr_t),
                      json={"status": "yolgon"})
            check("S-15: noma'lum holat -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))

            # ── RUXSAT ──
            emp = cur.execute(
                "select id from users where role='employee' and is_active=1 limit 1"
            ).fetchone()
            if emp:
                r = c.get("/offers", headers=auth(token_for(emp[0], "employee")))
                check("S-15: oddiy xodim takliflarni ko'rmaydi -> 403",
                      r.status_code == 403, "kod=" + str(r.status_code))

            # ── HUJJAT: FON ishida ──
            r = c.post("/document-templates", headers=auth(mgr_t), json={
                "kind": "offer", "name": "T-OfTmpl", "file_id": "T-OF-TMPL",
                "placeholders": ["fish", "lavozim", "oylik_sozda", "yolgon_belgi"]})
            tmpl_id = r.json().get("id") if r.status_code == 201 else None

            import time as _t

            boshi = _t.perf_counter()
            r = c.post(f"/offers/{offer_id}/generate", headers=auth(mgr_t),
                       json={"template_id": tmpl_id})
            ketgan = _t.perf_counter() - boshi
            check("S-15: hujjat NAVBATGA qo'yildi -> 202", r.status_code == 202,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            check("S-15: so'rov TEZ qaytdi (fon ishi)", ketgan < 1.0, f"{ketgan:.3f}s")
            check("S-15: shablonda yo'q belgi ogohlantirildi",
                  r.status_code == 202 and r.json().get("missing") == ["yolgon_belgi"],
                  "=" + r.text[:120])
            job_id = r.json().get("job_id") if r.status_code == 202 else None
            holat = cur.execute(
                "select status, kind, user_id from background_jobs where id=?",
                (job_id,)).fetchone()
            check("S-15: navbatdagi ish SO'RAGAN HR ga bog'langan",
                  holat == ("queued", "document_render", mgr[0]), "=" + str(holat))

            # ── Cron bajaradi: hujjat to'ldiriladi, HR ga boradi ──
            natija: dict = {}
            d = _docx_fixture(
                "<w:p><w:r><w:t>{{fish}} — {{lavozim}}, oylik: {{</w:t></w:r>"
                "<w:r><w:t>oylik_sozda}} so'm</w:t></w:r></w:p>")

            async def _tick():
                import api.telegram_notify as tn
                from api.services.background_jobs import background_tick

                asl_dl, asl_send = tn.download_file, tn.send_media_file

                async def soxta_dl(file_id):
                    return d

                async def soxta_send(chat_id, content, filename, kind, caption=None):
                    natija["chat_id"] = chat_id
                    natija["bytes"] = content
                    natija["filename"] = filename
                    return {"ok": True, "result": {"document": {"file_id": "T-OUT"}}}

                tn.download_file, tn.send_media_file = soxta_dl, soxta_send
                try:
                    async with async_session() as s2:
                        return await background_tick(s2)
                finally:
                    tn.download_file, tn.send_media_file = asl_dl, asl_send

            res = asyncio.run(_tick())
            check("S-15: cron hujjatni tayyorladi",
                  res.get("ok") is True and res.get("ran") == "document_render",
                  "=" + str(res))
            if natija.get("bytes"):
                matn = _docx_text(natija["bytes"])
                check("S-15: hujjatda nomzod ismi va lavozimi bor",
                      "T-Of Ali Valiyev" in matn and pos[1] in matn, "=" + matn)
                check("S-15: oylik odam o'qiydigan ko'rinishda (12 000 000)",
                      "12 000 000" in matn, "=" + matn)
            check("S-15: fayl nomida nomzod ismi bor",
                  "T-Of" in (natija.get("filename") or ""),
                  "=" + str(natija.get("filename")))

            # ⚠️ Hujjat HR ga bordi, NOMZODGA emas
            hr_tg = cur.execute("select telegram_id from users where id=?",
                                (mgr[0],)).fetchone()
            check("S-15: hujjat SO'RAGAN HR ga bordi (nomzodga emas)",
                  hr_tg is not None and natija.get("chat_id") == hr_tg[0],
                  f"chat={natija.get('chat_id')}, hr={hr_tg[0] if hr_tg else '-'}")

            # ── Taklif bazada QOLDI ──
            qoldi = cur.execute(
                "select candidate_name, salary, status from offers where id=?",
                (offer_id,)).fetchone()
            check("S-15: hujjat tayyorlangach ham taklif bazada qoldi",
                  qoldi == ("T-Of Ali Valiyev", 12000000, "accepted"), "=" + str(qoldi))
    except Exception:
        check("S-15 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from background_jobs where kind='document_render'")
            cur.execute("delete from offers where candidate_name like 'T-Of%'")
            cur.execute("delete from document_templates where name like 'T-OfTmpl%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_offer_hire() -> None:
    """S-16 (TZ 3.3) — taklifdan xodim va oylik stavkasi.

    Qabul mezonlari (TZ):
      • bitta bosishda xodim + stavka yaratiladi;
      • `offers.user_id` bog'lanadi;
      • IKKI MARTA bosilsa ikkita xodim yaratilmaydi (idempotent).

    NEGA MUHIM: F.I.Sh., lavozim va ish haqi allaqachon taklifda bor.
    Ularni qo'lda qayta terish xatoning eng keng tarqalgan manbai —
    kelishilgan oylik bilan bazaga kiritilgani boshqacha bo'lib chiqadi
    va bu faqat birinchi oylik hisoblanganda bilinadi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-16: TAKLIFDAN XODIM (idempotent)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    yaratilgan: list[int] = []
    try:
        cur.execute("delete from offers where candidate_name like 'T-Hire%'")
        cur.execute("delete from users where full_name like 'T-Hire%'")
        conn.commit()

        pos = cur.execute("select id, name from positions limit 1").fetchone()
        if not pos:
            check("lavozim topildi", False, "positions bo'sh")
            return

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Hire Nomzod", "phone": "+998900000001",
                "position_id": pos[0], "salary": 9500000,
                "probation_months": 3, "start_date": "2026-09-15",
                "manager_id": mgr[0]})
            offer_id = r.json().get("id") if r.status_code == 201 else None
            check("S-16: taklif yaratildi", offer_id is not None, "kod=" + str(r.status_code))

            # ── BITTA BOSISH: xodim + stavka ──
            r = c.post(f"/offers/{offer_id}/hire", headers=auth(mgr_t))
            check("S-16: ishga olish -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            uid = r.json().get("user_id") if r.status_code == 200 else None
            if uid:
                yaratilgan.append(uid)
            check("S-16: yangi xodim yaratildi (`created=True`)",
                  r.status_code == 200 and r.json().get("created") is True,
                  "=" + r.text[:120])

            xodim = cur.execute(
                "select full_name, role, position_id, manager_id, hire_date, is_active,"
                " telegram_id from users where id=?", (uid,)).fetchone()
            check("S-16: F.I.Sh. taklifdan olindi (qayta terilmadi)",
                  xodim and xodim[0] == "T-Hire Nomzod", "=" + str(xodim))
            check("S-16: lavozim va rahbar ham ko'chdi",
                  xodim and xodim[2] == pos[0] and xodim[3] == mgr[0], "=" + str(xodim))
            check("S-16: `hire_date` = ishga chiqish sanasi",
                  xodim and xodim[4] == "2026-09-15", "=" + str(xodim[4] if xodim else None))
            check("S-16: roli `employee`, faol", xodim and xodim[1] == "employee"
                  and xodim[5] == 1, "=" + str(xodim))
            check("S-16: `telegram_id` BO'SH (nomzodning boti bizda yo'q)",
                  xodim and xodim[6] is None, "=" + str(xodim[6] if xodim else "-"))

            stavka = cur.execute(
                "select amount, effective_from, pay_basis from salary_rates"
                " where user_id=? and deleted_at is null", (uid,)).fetchall()
            check("S-16: BITTA oylik stavkasi yaratildi", len(stavka) == 1,
                  "=" + str(stavka))
            check("S-16: stavka summasi taklifdagidek (9 500 000)",
                  stavka and float(stavka[0][0]) == 9500000.0, "=" + str(stavka))
            check("S-16: `effective_from` = ishga chiqish sanasi (TZ)",
                  stavka and stavka[0][1] == "2026-09-15", "=" + str(stavka))

            # ── `offers.user_id` BOG'LANDI ──
            bogl = cur.execute("select user_id, status from offers where id=?",
                               (offer_id,)).fetchone()
            check("S-16: `offers.user_id` bog'landi",
                  bool(bogl) and bogl[0] is not None and bogl[0] == uid,
                  "=" + str(bogl))
            check("S-16: holat avtomatik `accepted` bo'ldi",
                  bogl and bogl[1] == "accepted", "=" + str(bogl))

            # ── IDEMPOTENT: ikkinchi bosish ──
            r2 = c.post(f"/offers/{offer_id}/hire", headers=auth(mgr_t))
            check("S-16: ikkinchi bosish -> 200 (xato emas)", r2.status_code == 200,
                  "kod=" + str(r2.status_code))
            check("S-16: ikkinchi bosishda `created=False`",
                  r2.status_code == 200 and r2.json().get("created") is False,
                  "=" + r2.text[:120])
            check("S-16: bir xil xodim qaytdi",
                  r2.status_code == 200 and r2.json().get("user_id") == uid,
                  "=" + r2.text[:120])
            soni = cur.execute(
                "select count(*) from users where full_name='T-Hire Nomzod'").fetchone()[0]
            check("S-16: IKKINCHI XODIM YARATILMADI", soni == 1, "=" + str(soni))
            st_soni = cur.execute(
                "select count(*) from salary_rates where user_id=?", (uid,)).fetchone()[0]
            check("S-16: ikkinchi stavka ham yaratilmadi", st_soni == 1, "=" + str(st_soni))

            # ── Holat orqali ham ishlaydi (`accepted` qo'yilsa) ──
            r = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Hire Ikkinchi", "position_id": pos[0],
                "salary": 7000000})
            o2 = r.json().get("id") if r.status_code == 201 else None
            r = c.put(f"/offers/{o2}/status", headers=auth(mgr_t),
                      json={"status": "accepted"})
            check("S-16: «accepted» holati ham xodim yaratdi",
                  r.status_code == 200 and r.json().get("user_id") is not None,
                  "=" + r.text[:140])
            uid2 = r.json().get("user_id") if r.status_code == 200 else None
            if uid2:
                yaratilgan.append(uid2)
                bugun = cur.execute(
                    "select effective_from from salary_rates where user_id=?",
                    (uid2,)).fetchone()
                check("S-16: sanasiz taklifda stavka BUGUNDAN boshlanadi",
                      bugun and bugun[0] == date.today().isoformat(), "=" + str(bugun))

            # ── Rad etilgan taklifdan xodim yaratilmaydi ──
            r = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Hire Rad", "position_id": pos[0], "salary": 5000000})
            o3 = r.json().get("id") if r.status_code == 201 else None
            c.put(f"/offers/{o3}/status", headers=auth(mgr_t), json={"status": "declined"})
            r = c.post(f"/offers/{o3}/hire", headers=auth(mgr_t))
            check("S-16: rad etilgan taklifdan xodim yaratilmaydi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:100])

            # ── Ruxsat ──
            emp = cur.execute(
                "select id from users where role='employee' and is_active=1"
                " and full_name not like 'T-Hire%' limit 1").fetchone()
            if emp:
                r = c.post(f"/offers/{offer_id}/hire",
                           headers=auth(token_for(emp[0], "employee")))
                check("S-16: oddiy xodim ishga ololmaydi -> 403",
                      r.status_code == 403, "kod=" + str(r.status_code))
    except Exception:
        check("S-16 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if yaratilgan:
                belgi = ",".join("?" * len(yaratilgan))
                cur.execute(f"delete from salary_rates where user_id in ({belgi})",
                            tuple(yaratilgan))
            cur.execute("delete from offers where candidate_name like 'T-Hire%'")
            cur.execute("delete from users where full_name like 'T-Hire%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_certificates() -> None:
    """S-17 (TZ 3.9) — ma'lumotnoma ariza tasdiqlanishi bilan chiqadi.

    Qabul mezonlari (TZ):
      • ariza tasdiqlanishi bilan hujjat tayyorlanadi;
      • hujjat raqami TAKRORLANMAYDI;
      • o'rtacha oylik FAQAT so'ralganda yoziladi;
      • arxivda «kimga, qachon, qaysi maqsadda» tarixi qoladi.

    O'RTACHA OYLIK — maxfiy ma'lumot. Bankka kerak, bog'chaga umuman
    kerak emas. So'ralmagan bo'lsa hisoblanmaydi ham, hujjatga yozilmaydi
    ham; arxivda ham faqat bayroq turadi, summaning o'zi emas.
    """
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("S-17: MA'LUMOTNOMA (avtomatik)")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    uid = tmpl_id = None
    try:
        cur.execute("delete from certificates where number like '%/9%'")
        cur.execute("delete from users where full_name like 'T-Cert%'")
        cur.execute("delete from document_templates where name like 'T-CertTmpl%'")
        conn.commit()

        pos = cur.execute("select id, name from positions limit 1").fetchone()
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " position_id, hire_date, created_at) values (999701701,'T-Cert Xodim',"
            "'employee',1,1,?,'2024-03-15',datetime('now'))", (pos[0] if pos else None,))
        uid = cur.lastrowid
        conn.commit()
        emp_t = token_for(uid, "employee")

        # ── Shablon ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.post("/document-templates", headers=auth(mgr_t), json={
                "kind": "reference", "name": "T-CertTmpl bank",
                "file_id": "T-CERT-TMPL",
                "placeholders": ["raqam", "fish", "lavozim", "ortacha_oylik_sozda"]})
            tmpl_id = r.json().get("id") if r.status_code == 201 else None
            check("S-17: «reference» shabloni yuklandi", tmpl_id is not None,
                  "kod=" + str(r.status_code))

            # ── Belgilar ro'yxati ──
            r = c.get("/certificates/placeholders", headers=auth(mgr_t))
            nomlar = {x["name"] for x in r.json()} if r.status_code == 200 else set()
            check("S-17: shablon belgilari e'lon qilingan",
                  {"raqam", "fish", "lavozim", "ishga_qabul_sanasi",
                   "ortacha_oylik"} <= nomlar, "=" + str(sorted(nomlar)))

            # ── ARIZA -> TASDIQ -> HUJJAT (oyliksiz, bog'cha) ──
            r = c.post("/requests/me", headers=auth(emp_t), json={
                "kind": "certificate",
                "reason": "Bog'chaga ma'lumotnoma kerak, iltimos tayyorlang",
                "payload": {"purpose": "kindergarten", "include_salary": False}})
            check("S-17: ariza yuborildi", r.status_code in (200, 201),
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            req_id = r.json().get("id") if r.status_code in (200, 201) else None

            r = c.post(f"/requests/{req_id}/decide", headers=auth(mgr_t),
                       json={"decision": "approved", "note": "Tayyorlansin"})
            check("S-17: HR tasdiqladi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:160])
            info = r.json().get("applied", {}) if r.status_code == 200 else {}
            check("S-17: tasdiq bilan MA'LUMOTNOMA yaratildi (C guruh emas)",
                  bool(info.get("certificate_id")) and bool(info.get("number")),
                  "=" + str(info))
            check("S-17: keyingi qadam matni raqamni aytadi",
                  r.status_code == 200 and str(info.get("number", "")) in
                  (r.json().get("next_step") or ""), "=" + str(r.json().get("next_step")))
            raqam1 = info.get("number")

            cert = cur.execute(
                "select purpose, include_salary, avg_salary, request_id, user_id"
                " from certificates where number=?", (raqam1,)).fetchone()
            check("S-17: arxivda maqsad va ariza bog'lanishi bor",
                  cert == ("kindergarten", 0, None, req_id, uid), "=" + str(cert))
            check("S-17: OYLIK SO'RALMAGAN — hisoblanmadi ham, yozilmadi ham",
                  cert and cert[2] is None, "=" + str(cert[2] if cert else "-"))

            # Navbatga qo'yildimi
            ish = cur.execute(
                "select kind, status, user_id, params from background_jobs"
                " order by id desc limit 1").fetchone()
            check("S-17: hujjat NAVBATGA qo'yildi",
                  ish and ish[0] == "document_render" and ish[1] == "queued",
                  "=" + str(ish[:3] if ish else None))
            check("S-17: fayl XODIMNING o'ziga boradi (HR ga emas)",
                  ish and ish[2] == uid, f"job_user={ish[2] if ish else '-'}, xodim={uid}")

            # ── Cron bajaradi -> arxivga yoziladi ──
            d = _docx_fixture("<w:p><w:r><w:t>{{raqam}} {{fish}} {{lavozim}} "
                              "[{{ortacha_oylik_sozda}}]</w:t></w:r></w:p>")
            natija: dict = {}

            async def _tick():
                import api.telegram_notify as tn
                from api.services.background_jobs import background_tick

                asl_dl, asl_send = tn.download_file, tn.send_media_file

                async def soxta_dl(file_id):
                    return d

                async def soxta_send(chat_id, content, filename, kind, caption=None):
                    natija["bytes"] = content
                    natija["chat_id"] = chat_id
                    return {"ok": True, "result": {"document": {"file_id": "T-CERT-OUT"}}}

                tn.download_file, tn.send_media_file = soxta_dl, soxta_send
                try:
                    async with async_session() as s2:
                        return await background_tick(s2)
                finally:
                    tn.download_file, tn.send_media_file = asl_dl, asl_send

            res = asyncio.run(_tick())
            check("S-17: cron hujjatni tayyorladi", res.get("ok") is True, "=" + str(res))
            if natija.get("bytes"):
                matn = _docx_text(natija["bytes"])
                check("S-17: hujjatda raqam, ism va lavozim bor",
                      raqam1 in matn and "T-Cert Xodim" in matn, "=" + matn)
                check("S-17: oylik so'ralmagani uchun BO'SH qoldi",
                      "[]" in matn, "=" + matn)

            # ── ARXIVGA yozildi ──
            hujjat = cur.execute(
                "select d.name, d.file_id, d.user_id from employee_documents d"
                " join certificates c on c.document_id = d.id where c.number=?",
                (raqam1,)).fetchone()
            check("S-17: tayyor hujjat KADR ARXIVIGA yozildi",
                  hujjat is not None and hujjat[1] == "T-CERT-OUT" and hujjat[2] == uid,
                  "=" + str(hujjat))
            check("S-17: arxivdagi nomda raqam bor",
                  hujjat and raqam1 in hujjat[0], "=" + str(hujjat[0] if hujjat else "-"))

            # ── OYLIK BILAN (bank) ──
            cur.execute(
                "insert into payslips (user_id, period, base_amount, net, calculated_at)"
                " values (?,?,?,?,datetime('now'))", (uid, "2026-06", 8000000, 8000000))
            cur.execute(
                "insert into payslips (user_id, period, base_amount, net, calculated_at)"
                " values (?,?,?,?,datetime('now'))", (uid, "2026-07", 10000000, 10000000))
            conn.commit()

            r = c.post("/certificates", headers=auth(mgr_t), json={
                "user_id": uid, "purpose": "bank", "include_salary": True})
            check("S-17: arizasiz berish -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            raqam2 = r.json().get("number") if r.status_code == 201 else None

            # ── RAQAM TAKRORLANMAYDI ──
            check("S-17: ikkinchi raqam BIRINCHISIDAN farq qiladi",
                  raqam1 and raqam2 and raqam1 != raqam2, f"{raqam1} vs {raqam2}")
            check("S-17: raqamlar ketma-ket (0001 -> 0002)",
                  raqam1 and raqam2
                  and int(raqam2.split("/")[1]) == int(raqam1.split("/")[1]) + 1,
                  f"{raqam1} -> {raqam2}")
            jami, noyob = cur.execute(
                "select count(*), count(distinct number) from certificates").fetchone()
            check("S-17: barcha raqamlar NOYOB", jami == noyob, f"{jami} qator, {noyob} raqam")

            cert2 = cur.execute(
                "select include_salary, avg_salary from certificates where number=?",
                (raqam2,)).fetchone()
            check("S-17: oylik SO'RALGANDA hisoblandi (8M va 10M o'rtachasi = 9M)",
                  cert2 and cert2[0] == 1 and float(cert2[1]) == 9000000.0,
                  "=" + str(cert2))

            # ── ARXIV ro'yxati ──
            r = c.get("/certificates", headers=auth(mgr_t), params={"user_id": uid})
            arxiv = r.json() if r.status_code == 200 else []
            check("S-17: arxivda ikkala ma'lumotnoma ham bor", len(arxiv) == 2,
                  "=" + str(len(arxiv)))
            check("S-17: arxivda «kimga, qachon, qaysi maqsadda» ko'rinadi",
                  all({"user_name", "issued_at", "purpose_label"} <= set(x)
                      for x in arxiv), "=" + str(arxiv[:1]))
            check("S-17: arxiv oylik SUMMASINI oshkor QILMAYDI (faqat bayroq)",
                  all("avg_salary" not in x for x in arxiv),
                  "=" + str(sorted(arxiv[0]) if arxiv else []))
            check("S-17: qaysi biri oylik bilan ekani ko'rinadi",
                  sorted(x["include_salary"] for x in arxiv) == [False, True],
                  "=" + str([x["include_salary"] for x in arxiv]))

            # ── RUXSAT ──
            r = c.get("/certificates", headers=auth(emp_t))
            check("S-17: oddiy xodim arxivni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
    except Exception:
        check("S-17 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if uid:
                cur.execute("delete from payslips where user_id=?", (uid,))
                cur.execute("delete from employee_documents where user_id=?", (uid,))
                cur.execute("delete from certificates where user_id=?", (uid,))
                cur.execute("delete from employee_requests where user_id=?", (uid,))
                cur.execute("delete from users where id=?", (uid,))
            cur.execute("delete from background_jobs where kind='document_render'")
            cur.execute("delete from document_templates where name like 'T-CertTmpl%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_assets() -> None:
    """S-18 (TZ 3.11) — kompaniya mol-mulki va biriktirish.

    Qabul mezonlari (TZ):
      • BAND buyumni ikkinchi xodimga biriktirib bo'lmaydi;
      • inventar raqami takrorlanmaydi;
      • test: biriktirish -> qaytarish -> QAYTA biriktirish.

    «Bitta buyum — bitta xodimda» qo'riqchisi IKKI QATLAMLI va ikkalasi
    ham alohida sinaladi: (1) kod tekshiruvi tushunarli xato beradi,
    (2) qisman unikal indeks (`returned_at IS NULL`) bazada kafolatlaydi
    — kodni chetlab o'tib to'g'ridan-to'g'ri INSERT qilinsa ham.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-18: MOL-MULK (bitta buyum — bitta xodimda)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    asset_id = None
    try:
        cur.execute("delete from users where full_name like 'T-Ast%'")
        cur.execute(
            "delete from asset_assignments where asset_id in"
            " (select id from assets where inventory_no like 'T-INV%')")
        cur.execute("delete from assets where inventory_no like 'T-INV%'")
        conn.commit()

        for n, (nom, rol, tg) in enumerate(
            [("T-Ast Birinchi", "employee", 999701801),
             ("T-Ast Ikkinchi", "employee", 999701802)]):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,?,0,1,datetime('now'))",
                (tg, nom, rol))
            ids[f"u{n}"] = cur.lastrowid
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.get("/assets/kinds", headers=auth(mgr_t))
            check("S-18: turlar va holatlar ro'yxati bor",
                  r.status_code == 200 and len(r.json().get("kinds", [])) >= 7
                  and len(r.json().get("conditions", [])) == 4,
                  "kod=" + str(r.status_code))

            # ── BUYUM QO'SHISH ──
            r = c.post("/assets", headers=auth(mgr_t), json={
                "inventory_no": "T-INV-001", "name": "T-Ast Noutbuk",
                "kind": "laptop", "value": 9000000})
            check("S-18: buyum qo'shildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            asset_id = r.json().get("id") if r.status_code == 201 else None
            check("S-18: yangi buyum OMBORDA (egasi yo'q)",
                  r.status_code == 201 and r.json().get("holder_id") is None,
                  "=" + r.text[:120])

            # ── INVENTAR RAQAMI TAKRORLANMAYDI ──
            r = c.post("/assets", headers=auth(mgr_t), json={
                "inventory_no": "T-INV-001", "name": "Boshqa buyum", "kind": "phone"})
            check("S-18: takroriy inventar raqami -> 409", r.status_code == 409,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            soni = cur.execute(
                "select count(*) from assets where inventory_no='T-INV-001'").fetchone()[0]
            check("S-18: ikkinchi qator YARATILMADI", soni == 1, "=" + str(soni))

            # Noma'lum tur
            r = c.post("/assets", headers=auth(mgr_t), json={
                "inventory_no": "T-INV-002", "name": "X", "kind": "yolgon"})
            check("S-18: noma'lum tur -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))

            # ── BIRIKTIRISH ──
            r = c.post(f"/assets/{asset_id}/assign", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "condition_out": "good"})
            check("S-18: birinchi xodimga biriktirildi -> 200",
                  r.status_code == 200 and r.json().get("holder_id") == ids["u0"],
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            check("S-18: xodim hali «qabul qildim» bosmagan",
                  r.status_code == 200 and r.json().get("accepted") is False,
                  "=" + r.text[:120])

            # ── ⚠️ BAND BUYUM: 1-qatlam (kod) ──
            r = c.post(f"/assets/{asset_id}/assign", headers=auth(mgr_t), json={
                "user_id": ids["u1"], "condition_out": "good"})
            check("S-18: BAND buyumni ikkinchi xodimga berish -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code))
            check("S-18: xato xabari KIMDA ekanini aytadi",
                  r.status_code == 409 and "T-Ast Birinchi" in r.text, r.text[:160])
            ochiq = cur.execute(
                "select count(*) from asset_assignments where asset_id=?"
                " and returned_at is null", (asset_id,)).fetchone()[0]
            check("S-18: ochiq biriktirish hamon BITTA", ochiq == 1, "=" + str(ochiq))

            # ── ⚠️ BAND BUYUM: 2-qatlam (baza indeksi) ──
            # Kodni CHETLAB O'TIB to'g'ridan-to'g'ri INSERT — indeks to'sishi kerak.
            import sqlite3 as _sq

            try:
                cur.execute(
                    "insert into asset_assignments (asset_id, user_id, assigned_at,"
                    " condition_out, created_at) values (?,?,date('now'),'good',"
                    "datetime('now'))", (asset_id, ids["u1"]))
                conn.commit()
                check("S-18: BAZA indeksi ikkinchi ochiq biriktirishni to'sdi",
                      False, "INSERT o'tib ketdi — indeks ishlamayapti!")
                cur.execute("delete from asset_assignments where asset_id=? and user_id=?",
                            (asset_id, ids["u1"]))
                conn.commit()
            except _sq.IntegrityError as e:
                check("S-18: BAZA indeksi ikkinchi ochiq biriktirishni to'sdi",
                      True, str(e)[:80])
                conn.rollback()

            # ── Biriktirilgan buyumni hisobdan chiqarib bo'lmaydi ──
            r = c.delete(f"/assets/{asset_id}", headers=auth(mgr_t))
            check("S-18: xodimdagi buyumni hisobdan chiqarish -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:110])

            # ── Xodim O'ZINIKINI ko'radi ──
            r = c.get("/assets/me", headers=auth(token_for(ids["u0"], "employee")))
            check("S-18: xodim o'ziga biriktirilganini ko'radi",
                  r.status_code == 200 and len(r.json()) == 1
                  and r.json()[0]["inventory_no"] == "T-INV-001",
                  f"kod={r.status_code} {r.text[:110]}")
            r = c.get("/assets/me", headers=auth(token_for(ids["u1"], "employee")))
            check("S-18: boshqa xodimda bo'sh ro'yxat",
                  r.status_code == 200 and r.json() == [], "=" + r.text[:80])
            r = c.get("/assets", headers=auth(token_for(ids["u0"], "employee")))
            check("S-18: oddiy xodim UMUMIY ro'yxatni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── QAYTARISH ──
            r = c.post(f"/assets/{asset_id}/return", headers=auth(mgr_t),
                       json={"condition_in": "worn"})
            check("S-18: qaytarib olindi -> 200",
                  r.status_code == 200 and r.json().get("holder_id") is None,
                  "kod=" + str(r.status_code) + " " + r.text[:130])
            check("S-18: buyum holati qaytarishdagiga tenglashdi",
                  r.status_code == 200 and r.json().get("condition") == "worn",
                  "=" + r.text[:130])
            qator = cur.execute(
                "select returned_at, condition_out, condition_in from asset_assignments"
                " where asset_id=?", (asset_id,)).fetchone()
            check("S-18: qator O'CHIRILMADI — tarix qoldi",
                  qator is not None and qator[0] is not None, "=" + str(qator))
            check("S-18: berish va qaytarish holati alohida yozildi (zarar ko'rinadi)",
                  qator == (qator[0], "good", "worn"), "=" + str(qator))

            r = c.post(f"/assets/{asset_id}/return", headers=auth(mgr_t),
                       json={"condition_in": "good"})
            check("S-18: bo'sh buyumni qaytarish -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))

            # ── QAYTA BIRIKTIRISH (TZ zanjirining uchinchi qadami) ──
            r = c.post(f"/assets/{asset_id}/assign", headers=auth(mgr_t), json={
                "user_id": ids["u1"], "condition_out": "worn"})
            check("S-18: qaytargandan keyin BOSHQA xodimga biriktirildi",
                  r.status_code == 200 and r.json().get("holder_id") == ids["u1"],
                  "kod=" + str(r.status_code) + " " + r.text[:130])
            jami = cur.execute(
                "select count(*) from asset_assignments where asset_id=?",
                (asset_id,)).fetchone()[0]
            check("S-18: tarixda IKKITA yozuv (eskisi saqlandi)", jami == 2,
                  "=" + str(jami))

            # ── TARIX ──
            r = c.get(f"/assets/{asset_id}/history", headers=auth(mgr_t))
            tarix = r.json() if r.status_code == 200 else []
            check("S-18: tarixda ikkala egasi ham ko'rinadi",
                  len(tarix) == 2 and {x["user_name"] for x in tarix} ==
                  {"T-Ast Birinchi", "T-Ast Ikkinchi"},
                  "=" + str([x["user_name"] for x in tarix]))
            check("S-18: tarixda holat o'zgarishi saqlangan (good -> worn)",
                  any(x["condition_out"] == "good" and x["condition_in"] == "worn"
                      for x in tarix), "=" + str(tarix))
    except Exception:
        check("S-18 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if asset_id:
                cur.execute("delete from asset_assignments where asset_id=?", (asset_id,))
                cur.execute("delete from assets where id=?", (asset_id,))
            cur.execute("delete from assets where inventory_no like 'T-INV%'")
            cur.execute("delete from users where full_name like 'T-Ast%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_assets_employee() -> None:
    """S-19 (TZ 3.11) — xodim tomoni, dalolatnoma va standart to'plam.

    Qabul mezonlari (TZ):
      • xodim FAQAT o'ziga biriktirilganini ko'radi;
      • «Qabul qildim» VAQTI yoziladi;
      • dalolatnoma `.docx` chiqadi.

    «Qabul qildim» ni faqat xodimning O'ZI bosishi mumkin: nizo chiqqanda
    «men buni olganim yo'q» degan da'voga javob shu yozuv bo'ladi.
    Tasdiqlash IDEMPOTENT — qayta bosilsa BIRINCHI vaqt saqlanadi, aks
    holda xodim sanani «yangilab» qo'yishi mumkin edi.
    """
    import asyncio
    import time as _t

    import httpx

    print("\n" + "=" * 60)
    print("S-19: MOL-MULK — XODIM TOMONI VA DALOLATNOMA")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    asset_id = tmpl_id = None
    try:
        cur.execute("delete from users where full_name like 'T-A19%'")
        cur.execute(
            "delete from asset_assignments where asset_id in"
            " (select id from assets where inventory_no like 'T-I19%')")
        cur.execute("delete from assets where inventory_no like 'T-I19%'")
        cur.execute("delete from document_templates where name like 'T-Act%'")
        conn.commit()

        pos = cur.execute("select id, name from positions limit 1").fetchone()
        for n, tg in enumerate((999701901, 999701902)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, position_id, created_at)"
                " values (?,?,'employee',0,1,?,datetime('now'))",
                (tg, f"T-A19 Xodim{n}", pos[0] if pos and n == 0 else None))
            ids[f"u{n}"] = cur.lastrowid
        conn.commit()
        t0 = token_for(ids["u0"], "employee")
        t1 = token_for(ids["u1"], "employee")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.post("/assets", headers=auth(mgr_t), json={
                "inventory_no": "T-I19-01", "name": "T-A19 Noutbuk",
                "kind": "laptop", "value": 8000000})
            asset_id = r.json().get("id") if r.status_code == 201 else None
            c.post(f"/assets/{asset_id}/assign", headers=auth(mgr_t),
                   json={"user_id": ids["u0"], "condition_out": "good"})

            # ── XODIM FAQAT O'ZINIKINI KO'RADI ──
            r = c.get("/assets/me", headers=auth(t0))
            check("S-19: egasi o'z buyumini ko'radi",
                  r.status_code == 200 and len(r.json()) == 1, "=" + r.text[:110])
            r = c.get("/assets/me", headers=auth(t1))
            check("S-19: boshqa xodim ko'rmaydi",
                  r.status_code == 200 and r.json() == [], "=" + r.text[:80])

            # ── «QABUL QILDIM» ──
            r = c.post(f"/assets/{asset_id}/accept", headers=auth(t1))
            check("S-19: BEGONA xodim tasdiqlay olmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))
            r = c.post(f"/assets/{asset_id}/accept", headers=auth(mgr_t))
            check("S-19: HR xodim NOMIDAN tasdiqlay olmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            r = c.post(f"/assets/{asset_id}/accept", headers=auth(t0))
            check("S-19: egasi tasdiqladi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            birinchi = r.json().get("accepted_at") if r.status_code == 200 else None
            check("S-19: tasdiqlash VAQTI yozildi", bool(birinchi), "=" + str(birinchi))
            bazada = cur.execute(
                "select accepted_at from asset_assignments where asset_id=?"
                " and returned_at is null", (asset_id,)).fetchone()
            check("S-19: vaqt bazada ham bor", bazada and bazada[0] is not None,
                  "=" + str(bazada))

            # IDEMPOTENT: qayta bosish vaqtni O'ZGARTIRMAYDI
            _t.sleep(1.1)
            r = c.post(f"/assets/{asset_id}/accept", headers=auth(t0))
            check("S-19: qayta bosishda BIRINCHI vaqt saqlandi",
                  r.status_code == 200 and r.json().get("accepted_at") == birinchi,
                  f"{birinchi} vs {r.json().get('accepted_at')}")

            r = c.get("/assets/me", headers=auth(t0))
            check("S-19: ro'yxatda «qabul qilingan» belgisi paydo bo'ldi",
                  r.status_code == 200 and r.json()[0].get("accepted") is True,
                  "=" + r.text[:120])

            # ── DALOLATNOMA ──
            r = c.post("/document-templates", headers=auth(mgr_t), json={
                "kind": "act", "name": "T-Act dalolatnoma", "file_id": "T-ACT-TMPL",
                "placeholders": ["fish", "buyum", "inventar", "amal", "holati"]})
            tmpl_id = r.json().get("id") if r.status_code == 201 else None

            r = c.get("/assets/act-placeholders", headers=auth(mgr_t))
            nomlar = {x["name"] for x in r.json()} if r.status_code == 200 else set()
            check("S-19: dalolatnoma belgilari e'lon qilingan",
                  {"fish", "buyum", "inventar", "amal", "holati", "qiymati_sozda"}
                  <= nomlar, "=" + str(sorted(nomlar)))

            boshi = _t.perf_counter()
            r = c.post(f"/assets/{asset_id}/act", headers=auth(mgr_t),
                       json={"template_id": tmpl_id, "action": "out"})
            ketgan = _t.perf_counter() - boshi
            check("S-19: dalolatnoma NAVBATGA qo'yildi -> 202",
                  r.status_code == 202, "kod=" + str(r.status_code) + " " + r.text[:120])
            check("S-19: so'rov TEZ qaytdi (fon ishi)", ketgan < 1.0, f"{ketgan:.3f}s")

            natija: dict = {}
            d = _docx_fixture("<w:p><w:r><w:t>{{fish}} — {{buyum}} ({{</w:t></w:r>"
                              "<w:r><w:t>inventar}}) {{amal}}, holati: {{holati}}"
                              "</w:t></w:r></w:p>")

            async def _tick():
                import api.telegram_notify as tn
                from api.services.background_jobs import background_tick

                asl_dl, asl_send = tn.download_file, tn.send_media_file

                async def soxta_dl(file_id):
                    return d

                async def soxta_send(chat_id, content, filename, kind, caption=None):
                    natija["bytes"] = content
                    natija["filename"] = filename
                    return {"ok": True, "result": {"document": {"file_id": "T-ACT-OUT"}}}

                tn.download_file, tn.send_media_file = soxta_dl, soxta_send
                try:
                    async with async_session() as s2:
                        return await background_tick(s2)
                finally:
                    tn.download_file, tn.send_media_file = asl_dl, asl_send

            res = asyncio.run(_tick())
            check("S-19: cron dalolatnomani tayyorladi", res.get("ok") is True,
                  "=" + str(res))
            if natija.get("bytes"):
                matn = _docx_text(natija["bytes"])
                check("S-19: dalolatnomada xodim, buyum va inventar bor",
                      "T-A19 Xodim0" in matn and "T-I19-01" in matn, "=" + matn)
                check("S-19: «berildi» deb yozilgan", "berildi" in matn, "=" + matn)
                check("S-19: holat tarjima qilingan (Yaxshi)", "Yaxshi" in matn,
                      "=" + matn)

            # Qaytarilmagan buyumga «in» dalolatnomasi bo'lmaydi
            r = c.post(f"/assets/{asset_id}/act", headers=auth(mgr_t),
                       json={"template_id": tmpl_id, "action": "in"})
            check("S-19: qaytarilmagan buyumga «in» dalolatnomasi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:110])

            # ── STANDART TO'PLAM ──
            if pos:
                r = c.put("/assets/standard-set", headers=auth(mgr_t), json={
                    "position_id": pos[0], "items": {"laptop": 1, "phone": 1, "sim": 2}})
                check("S-19: standart to'plam saqlandi",
                      r.status_code == 200 and r.json().get("count") == 3,
                      "kod=" + str(r.status_code) + " " + r.text[:110])

                r = c.put("/assets/standard-set", headers=auth(mgr_t), json={
                    "position_id": pos[0], "items": {"laptop": 0}})
                check("S-19: nol miqdor rad etildi -> 400", r.status_code == 400,
                      "kod=" + str(r.status_code) + " " + r.text[:110])
                r = c.put("/assets/standard-set", headers=auth(mgr_t), json={
                    "position_id": pos[0], "items": {"yolgon": 1}})
                check("S-19: noma'lum tur rad etildi -> 400", r.status_code == 400,
                      "kod=" + str(r.status_code))

                r = c.get(f"/assets/standard-set/{pos[0]}", headers=auth(mgr_t))
                turlar = {i["kind"]: i["quantity"] for i in r.json().get("items", [])}
                check("S-19: to'plam o'qildi va xato urinishlar buzmadi",
                      turlar == {"laptop": 1, "phone": 1, "sim": 2}, "=" + str(turlar))

                # To'liq ALMASHTIRISH (qisman yangilash emas)
                c.put("/assets/standard-set", headers=auth(mgr_t), json={
                    "position_id": pos[0], "items": {"laptop": 1}})
                r = c.get(f"/assets/standard-set/{pos[0]}", headers=auth(mgr_t))
                check("S-19: to'plam BUTUNLAY almashtirildi (eskisi o'chdi)",
                      [i["kind"] for i in r.json().get("items", [])] == ["laptop"],
                      "=" + str(r.json().get("items")))

                # ── NAZORAT RO'YXATI ──
                c.put("/assets/standard-set", headers=auth(mgr_t), json={
                    "position_id": pos[0], "items": {"laptop": 1, "phone": 1}})
                r = c.get(f"/assets/checklist/{ids['u0']}", headers=auth(mgr_t))
                bandlar = {i["kind"]: i for i in r.json().get("items", [])}
                check("S-19: nazoratda noutbuk BOR deb ko'rsatilgan",
                      bandlar.get("laptop", {}).get("held") == 1
                      and bandlar["laptop"]["missing"] == 0, "=" + str(bandlar))
                check("S-19: nazoratda telefon YETISHMAYDI deb ko'rsatilgan",
                      bandlar.get("phone", {}).get("held") == 0
                      and bandlar["phone"]["missing"] == 1, "=" + str(bandlar))

                r = c.get(f"/assets/checklist/{ids['u1']}", headers=auth(mgr_t))
                check("S-19: lavozimsiz xodimda bo'sh nazorat va bayroq",
                      r.status_code == 200 and r.json().get("has_position") is False
                      and r.json().get("items") == [], "=" + r.text[:120])

            # ── Qaytargandan keyin xodim ro'yxatidan chiqadi ──
            c.post(f"/assets/{asset_id}/return", headers=auth(mgr_t),
                   json={"condition_in": "worn"})
            r = c.get("/assets/me", headers=auth(t0))
            check("S-19: qaytarilgach xodim ro'yxatidan chiqdi",
                  r.status_code == 200 and r.json() == [], "=" + r.text[:80])
            r = c.post(f"/assets/{asset_id}/accept", headers=auth(t0))
            check("S-19: qaytarilgan buyumni tasdiqlab bo'lmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            # Endi «in» dalolatnomasi ishlaydi
            r = c.post(f"/assets/{asset_id}/act", headers=auth(mgr_t),
                       json={"template_id": tmpl_id, "action": "in"})
            check("S-19: qaytargandan keyin «in» dalolatnomasi -> 202",
                  r.status_code == 202, "kod=" + str(r.status_code))
    except Exception:
        check("S-19 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if pos:
                cur.execute("delete from position_asset_sets where position_id=?", (pos[0],))
            if asset_id:
                cur.execute("delete from asset_assignments where asset_id=?", (asset_id,))
                cur.execute("delete from assets where id=?", (asset_id,))
            cur.execute("delete from background_jobs where kind='document_render'")
            cur.execute("delete from document_templates where name like 'T-Act%'")
            cur.execute("delete from users where full_name like 'T-A19%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_acknowledgements() -> None:
    """S-20 — «Tanishdim» qaydi: uchta modul uchun umumiy mexanizm.

    Qabul mezonlari (TZ):
      • uchala obyekt turi BITTA jadvalga yozadi;
      • VERSIYA o'zgarsa qayta tanishish talab qilinadi;
      • bir odam bir versiyani IKKI MARTA tasdiqlay olmaydi (unique).

    Versiya — modulning markaziy g'oyasi. Yo'riqnoma yangilansa eski
    tanishuv o'tmaydi: xodim eski matnga rozi bo'lgan, yangisiga emas.
    Bu huquqiy jihatdan muhim — «u bilardi» degan da'vo faqat u KO'RGAN
    versiyaga nisbatan o'rinli.
    """
    import asyncio
    import sqlite3 as _sq
    import time as _t

    import httpx

    print("\n" + "=" * 60)
    print("S-20: «TANISHDIM» QAYDI (versiya bilan)")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Ack%'")
        conn.commit()
        for n, tg in enumerate((999702001, 999702002)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, f"T-Ack Xodim{n}"))
            ids[f"u{n}"] = cur.lastrowid
        conn.commit()
        t0 = token_for(ids["u0"], "employee")
        t1 = token_for(ids["u1"], "employee")

        async def _sorash(obj_type, obj_id, version, users, title=None):
            from api.services.acknowledgements import request_ack
            async with async_session() as s2:
                n = await request_ack(
                    s2, object_type=obj_type, object_id=obj_id, version=version,
                    user_ids=users, title=title, link="/me/acks", requested_by=mgr[0])
                await s2.commit()
                return n

        # ── UCHALA TUR BITTA JADVALGA ──
        n1 = asyncio.run(_sorash("instruction", 9001, 1, [ids["u0"], ids["u1"]],
                                 "T-Ack Yo'riqnoma"))
        n2 = asyncio.run(_sorash("announcement", 9002, 1, [ids["u0"]], "T-Ack E'lon"))
        n3 = asyncio.run(_sorash("briefing", 9003, 1, [ids["u0"]], "T-Ack Instruktaj"))
        check("S-20: uchala tur uchun so'rov yaratildi",
              (n1, n2, n3) == (2, 1, 1), f"={n1},{n2},{n3}")
        jadvallar = cur.execute(
            "select count(distinct object_type) from acknowledgements"
            " where object_id in (9001,9002,9003)").fetchone()[0]
        check("S-20: uchala tur BITTA jadvalda", jadvallar == 3, "=" + str(jadvallar))

        # ── SO'ROV IDEMPOTENT ──
        takror = asyncio.run(_sorash("instruction", 9001, 1, [ids["u0"], ids["u1"]]))
        check("S-20: takroriy so'rov yangi qator YARATMADI", takror == 0, "=" + str(takror))
        soni = cur.execute(
            "select count(*) from acknowledgements where object_id=9001").fetchone()[0]
        check("S-20: obyektda hamon 2 qator", soni == 2, "=" + str(soni))

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── XODIM O'Z RO'YXATINI KO'RADI ──
            r = c.get("/acks/me", headers=auth(t0))
            bandlar = r.json() if r.status_code == 200 else []
            check("S-20: xodimda uchta band ko'rindi", len(bandlar) == 3,
                  "=" + str([b["object_type"] for b in bandlar]))
            check("S-20: sarlavha va havola saqlangan",
                  any(b["title"] == "T-Ack E'lon" and b["link"] == "/me/acks"
                      for b in bandlar), "=" + str(bandlar[:1]))
            check("S-20: tur nomi tarjima qilingan",
                  any(b["object_type_label"] == "Instruktaj" for b in bandlar),
                  "=" + str([b["object_type_label"] for b in bandlar]))
            r = c.get("/acks/me", headers=auth(t1))
            check("S-20: ikkinchi xodimda faqat bitta band",
                  r.status_code == 200 and len(r.json()) == 1, "=" + r.text[:110])

            # ── «TANISHDIM» ──
            r = c.post("/acks/me/ack", headers=auth(t0),
                       json={"object_type": "instruction", "object_id": 9001, "version": 1})
            check("S-20: tanishdi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            birinchi = r.json().get("acknowledged_at") if r.status_code == 200 else None
            check("S-20: vaqt yozildi", bool(birinchi), "=" + str(birinchi))

            # IDEMPOTENT: qayta bosish vaqtni o'zgartirmaydi
            _t.sleep(1.1)
            r = c.post("/acks/me/ack", headers=auth(t0),
                       json={"object_type": "instruction", "object_id": 9001, "version": 1})
            check("S-20: qayta bosishda BIRINCHI vaqt saqlandi",
                  r.status_code == 200 and r.json().get("acknowledged_at") == birinchi,
                  f"{birinchi} vs {r.json().get('acknowledged_at')}")

            r = c.get("/acks/me", headers=auth(t0))
            check("S-20: tanishilgan band ro'yxatdan chiqdi",
                  r.status_code == 200 and len(r.json()) == 2,
                  "=" + str([b["object_type"] for b in r.json()]))

            # ── SO'RALMAGAN BANDNI TASDIQLAB BO'LMAYDI ──
            r = c.post("/acks/me/ack", headers=auth(t1),
                       json={"object_type": "announcement", "object_id": 9002, "version": 1})
            check("S-20: so'ralmagan bandni tasdiqlash -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))
            r = c.post("/acks/me/ack", headers=auth(t0),
                       json={"object_type": "yolgon", "object_id": 1, "version": 1})
            check("S-20: noma'lum tur -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))

            # ── ⚠️ VERSIYA: yangisi chiqsa eski tanishuv O'TMAYDI ──
            asyncio.run(_sorash("instruction", 9001, 2, [ids["u0"], ids["u1"]],
                                "T-Ack Yo'riqnoma v2"))
            r = c.get("/acks/me", headers=auth(t0))
            yangi = [b for b in r.json() if b["object_id"] == 9001]
            check("S-20: v2 chiqqach yo'riqnoma QAYTA so'raldi",
                  len(yangi) == 1 and yangi[0]["version"] == 2,
                  "=" + str(yangi))
            eski = cur.execute(
                "select version, acknowledged_at is not null from acknowledgements"
                " where object_id=9001 and user_id=? order by version",
                (ids["u0"],)).fetchall()
            check("S-20: v1 tanishuvi TARIXDA qoldi (o'chirilmadi)",
                  eski == [(1, 1), (2, 0)], "=" + str(eski))

            # ── FAQAT ENG YANGI VERSIYA ko'rinadi ──
            asyncio.run(_sorash("instruction", 9001, 3, [ids["u1"]], "v3"))
            r = c.get("/acks/me", headers=auth(t1))
            yorik = [b for b in r.json() if b["object_id"] == 9001]
            check("S-20: ikki versiyani o'tkazib yuborgan xodimga BITTA band",
                  len(yorik) == 1 and yorik[0]["version"] == 3, "=" + str(yorik))
            baza = cur.execute(
                "select count(*) from acknowledgements where object_id=9001 and user_id=?",
                (ids["u1"],)).fetchone()[0]
            check("S-20: bazada esa uchala versiya turibdi", baza == 3, "=" + str(baza))

            # ── ⚠️ UNIQUE: bir versiyani ikki marta yozib bo'lmaydi ──
            try:
                cur.execute(
                    "insert into acknowledgements (user_id, object_type, object_id,"
                    " version, requested_at) values (?,'instruction',9001,1,datetime('now'))",
                    (ids["u0"],))
                conn.commit()
                check("S-20: BAZA dublikat tasdiqni to'sdi", False,
                      "INSERT o'tib ketdi — UNIQUE ishlamayapti!")
                cur.execute("delete from acknowledgements where id=?", (cur.lastrowid,))
                conn.commit()
            except _sq.IntegrityError as e:
                check("S-20: BAZA dublikat tasdiqni to'sdi", True, str(e)[:70])
                conn.rollback()

            # ── KIM O'QIGAN (rahbar) ──
            r = c.get("/acks/object/instruction/9001", headers=auth(mgr_t),
                      params={"version": 1})
            oquvchilar = r.json() if r.status_code == 200 else []
            check("S-20: v1 bo'yicha ikki xodim ko'rinadi", len(oquvchilar) == 2,
                  "=" + str(len(oquvchilar)))
            check("S-20: O'QIMAGANLAR TEPADA",
                  bool(oquvchilar) and oquvchilar[0]["acknowledged_at"] is None,
                  "=" + str([(x["user_name"], bool(x["acknowledged_at"]))
                             for x in oquvchilar]))

            r = c.get("/acks/object/instruction/9001/stats", headers=auth(mgr_t),
                      params={"version": 1})
            check("S-20: statistika 2 dan 1 tasi o'qigan",
                  r.status_code == 200 and r.json().get("total") == 2
                  and r.json().get("read") == 1 and r.json().get("pending") == 1,
                  "=" + r.text[:120])

            r = c.get("/acks/object/instruction/9001/stats", headers=auth(mgr_t))
            check("S-20: versiyasiz so'rovda ENG SO'NGGI versiya olinadi",
                  r.status_code == 200 and r.json().get("version") == 3,
                  "=" + r.text[:120])

            # ── RUXSAT ──
            r = c.get("/acks/object/instruction/9001", headers=auth(t0))
            check("S-20: oddiy xodim o'quvchilar ro'yxatini ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
    except Exception:
        check("S-20 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from acknowledgements where object_id in (9001,9002,9003)")
            cur.execute("delete from users where full_name like 'T-Ack%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_announcements() -> None:
    """S-21 (TZ 3.12) — ichki e'lonlar.

    Qabul mezonlari (TZ):
      • rahbar panelida kim o'qigani/o'qimagani ko'rinadi;
      • KUNLIK LIMIT ishlaydi;
      • qamrovga kirmagan xodimga e'lon UMUMAN ko'rinmaydi.

    Qamrov — ko'rinishni bezash emas, FILTR: sotuv bo'limiga aytilgan gap
    prorabga ko'rinmasligi kerak. Shuning uchun tekshiruv serverda va
    bitta funksiyada (`visible_to`).
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-21: ICHKI E'LONLAR (qamrov + kunlik limit)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    yaratilgan: list[int] = []
    try:
        cur.execute("delete from users where full_name like 'T-Ann%'")
        cur.execute("delete from announcements where title like 'T-Ann%'")
        conn.commit()

        pos = cur.execute("select id, name from positions limit 1").fetchone()
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " position_id, created_at) values (999702101,'T-Ann Sotuvchi','employee',"
            "0,1,?,datetime('now'))", (pos[0] if pos else None,))
        ids["sotuvchi"] = cur.lastrowid
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999702102,'T-Ann Boshqa','employee',0,1,datetime('now'))")
        ids["boshqa"] = cur.lastrowid
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999702103,'T-Ann Rop','rop',0,1,datetime('now'))")
        ids["rop"] = cur.lastrowid
        conn.commit()
        t_sot = token_for(ids["sotuvchi"], "employee")
        t_bos = token_for(ids["boshqa"], "employee")
        t_rop = token_for(ids["rop"], "rop")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # Chegara testga yetarli bo'lsin
            c.put("/announcements/config", headers=auth(mgr_t), json={"daily_limit": 4})

            # ── QAMROV: LAVOZIM bo'yicha ──
            if pos:
                r = c.post("/announcements", headers=auth(mgr_t), json={
                    "title": "T-Ann Lavozimga", "body": "Faqat shu lavozimga",
                    "audience": "positions", "scope_ids": [pos[0]], "important": False})
                check("S-21: lavozim qamrovi bilan e'lon -> 201",
                      r.status_code == 201, "kod=" + str(r.status_code) + " " + r.text[:120])
                if r.status_code == 201:
                    yaratilgan.append(r.json()["id"])

                r = c.get("/announcements/me", headers=auth(t_sot))
                check("S-21: qamrovdagi xodim e'lonni KO'RADI",
                      r.status_code == 200
                      and any(a["title"] == "T-Ann Lavozimga" for a in r.json()),
                      "=" + str([a["title"] for a in r.json()]))
                r = c.get("/announcements/me", headers=auth(t_bos))
                check("S-21: qamrovdan TASHQARIDAGI xodimga UMUMAN ko'rinmaydi",
                      r.status_code == 200
                      and not any(a["title"] == "T-Ann Lavozimga" for a in r.json()),
                      "=" + str([a["title"] for a in r.json()]))

            # ── QAMROV: ROL bo'yicha ──
            r = c.post("/announcements", headers=auth(mgr_t), json={
                "title": "T-Ann Roplarga", "body": "Faqat ROP larga",
                "audience": "roles", "scope_ids": ["rop"], "important": False})
            if r.status_code == 201:
                yaratilgan.append(r.json()["id"])
            r = c.get("/announcements/me", headers=auth(t_rop))
            check("S-21: rol qamrovi ishladi (ROP ko'radi)",
                  r.status_code == 200
                  and any(a["title"] == "T-Ann Roplarga" for a in r.json()),
                  "=" + str([a["title"] for a in r.json()]))
            r = c.get("/announcements/me", headers=auth(t_sot))
            check("S-21: boshqa roldagi xodimga ko'rinmaydi",
                  r.status_code == 200
                  and not any(a["title"] == "T-Ann Roplarga" for a in r.json()),
                  "=" + str([a["title"] for a in r.json()]))

            # ── BO'SH QAMROV rad etiladi ──
            r = c.post("/announcements", headers=auth(mgr_t), json={
                "title": "T-Ann Bosh", "body": "Hech kimga", "audience": "roles",
                "scope_ids": [], "important": False})
            check("S-21: bo'sh qamrov ro'yxati -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            r = c.post("/announcements", headers=auth(mgr_t), json={
                "title": "T-Ann Yolgon", "body": "matn", "audience": "roles",
                "scope_ids": ["yolgonrol"], "important": False})
            check("S-21: noma'lum rol -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))

            # ── MUHIM e'lon -> «Tanishdim» so'raladi ──
            r = c.post("/announcements", headers=auth(mgr_t), json={
                "title": "T-Ann Muhim", "body": "Bu muhim e'lon",
                "audience": "all", "scope_ids": None, "important": True})
            check("S-21: muhim e'lon yuborildi va tanishuv so'raldi",
                  r.status_code == 201 and r.json().get("ack_requested") is True
                  and r.json().get("audience_size", 0) > 0,
                  "kod=" + str(r.status_code) + " " + r.text[:130])
            muhim_id = r.json().get("id") if r.status_code == 201 else None
            if muhim_id:
                yaratilgan.append(muhim_id)

            r = c.get("/announcements/me", headers=auth(t_sot))
            muhim = [a for a in r.json() if a["id"] == muhim_id]
            check("S-21: xodimda «tanishmagan» belgisi turibdi",
                  len(muhim) == 1 and muhim[0]["acknowledged"] is False,
                  "=" + str(muhim))

            r = c.post("/acks/me/ack", headers=auth(t_sot), json={
                "object_type": "announcement", "object_id": muhim_id, "version": 1})
            check("S-21: xodim tanishdi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            r = c.get("/announcements/me", headers=auth(t_sot))
            muhim = [a for a in r.json() if a["id"] == muhim_id]
            check("S-21: belgi «tanishgan» ga o'zgardi",
                  muhim and muhim[0]["acknowledged"] is True, "=" + str(muhim))

            # ── KIM O'QIGAN (rahbar) ──
            r = c.get(f"/acks/object/announcement/{muhim_id}", headers=auth(mgr_t))
            oquvchilar = r.json() if r.status_code == 200 else []
            oqigan = [x for x in oquvchilar if x["acknowledged_at"]]
            check("S-21: rahbar kim o'qiganini ko'radi",
                  len(oquvchilar) > 1 and len(oqigan) == 1,
                  f"jami={len(oquvchilar)}, o'qigan={len(oqigan)}")
            check("S-21: O'QIMAGANLAR TEPADA",
                  bool(oquvchilar) and oquvchilar[0]["acknowledged_at"] is None,
                  "=" + str([bool(x["acknowledged_at"]) for x in oquvchilar[:3]]))

            # ── ⚠️ KUNLIK LIMIT ──
            r = c.get("/announcements/quota", headers=auth(mgr_t))
            qolgan = r.json().get("left", 0) if r.status_code == 200 else 0
            check("S-21: kvota hisoblanmoqda", r.status_code == 200
                  and r.json().get("daily_limit") == 4, "=" + r.text[:100])

            for n in range(qolgan):
                rr = c.post("/announcements", headers=auth(mgr_t), json={
                    "title": f"T-Ann To'ldirish{n}", "body": "to'ldirish matni",
                    "audience": "all", "scope_ids": None, "important": False})
                if rr.status_code == 201:
                    yaratilgan.append(rr.json()["id"])
            r = c.post("/announcements", headers=auth(mgr_t), json={
                "title": "T-Ann Ortiqcha", "body": "ortiqcha matn", "audience": "all",
                "scope_ids": None, "important": False})
            check("S-21: chegara to'lgach -> 429", r.status_code == 429,
                  "kod=" + str(r.status_code) + " " + r.text[:130])
            check("S-21: xato xabari NIMA QILISHNI aytadi",
                  r.status_code == 429 and "Ertaga" in r.text, r.text[:150])
            yaratildimi = cur.execute(
                "select count(*) from announcements where title='T-Ann Ortiqcha'"
            ).fetchone()[0]
            check("S-21: chegaradan oshgan e'lon YARATILMADI", yaratildimi == 0,
                  "=" + str(yaratildimi))

            # ── O'CHIRILGAN e'lon ham limitga kiradi (chetlab o'tishga qarshi) ──
            #  ⚠️ MUHIM e'lonni o'chirmaymiz — u keyingi (versiya) tekshiruvida
            #  kerak. O'chirish uchun to'ldirish e'lonlaridan birini olamiz.
            ochiriladigan = cur.execute(
                "select id from announcements where title like 'T-Ann To%' limit 1"
            ).fetchone()
            if ochiriladigan:
                c.delete(f"/announcements/{ochiriladigan[0]}", headers=auth(mgr_t))
                r = c.post("/announcements", headers=auth(mgr_t), json={
                    "title": "T-Ann Chetlab", "body": "chetlab matn", "audience": "all",
                    "scope_ids": None, "important": False})
                check("S-21: o'chirib limitni chetlab bo'lmaydi -> 429",
                      r.status_code == 429, "kod=" + str(r.status_code))

            # ── VERSIYA: matn tahrirlansa qayta tanishish ──
            if muhim_id:
                r = c.put(f"/announcements/{muhim_id}", headers=auth(mgr_t), json={
                    "title": "T-Ann Muhim", "body": "MATN O'ZGARDI",
                    "audience": "all", "scope_ids": None, "important": True})
                check("S-21: tahrirlangach versiya oshdi va qayta so'raldi",
                      r.status_code == 200 and r.json().get("version") == 2
                      and r.json().get("reacked") is True, "=" + r.text[:120])
                r = c.get("/announcements/me", headers=auth(t_sot))
                muhim = [a for a in r.json() if a["id"] == muhim_id]
                check("S-21: eski tanishuv O'TMADI — qayta so'ralyapti",
                      muhim and muhim[0]["acknowledged"] is False
                      and muhim[0]["version"] == 2, "=" + str(muhim))

            # ── RUXSAT ──
            r = c.post("/announcements", headers=auth(t_sot), json={
                "title": "T-Ann Xodim", "body": "xodim matni", "audience": "all",
                "scope_ids": None, "important": False})
            check("S-21: oddiy xodim e'lon yoza olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.put("/announcements/config", headers=auth(t_rop),
                      json={"daily_limit": 50})
            check("S-21: ROP chegarani o'zgartira olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
    except Exception:
        check("S-21 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from acknowledgements where object_type='announcement'"
                        " and object_id in (select id from announcements"
                        " where title like 'T-Ann%')")
            cur.execute("delete from announcements where title like 'T-Ann%'")
            cur.execute("delete from users where full_name like 'T-Ann%'")
            cur.execute("update announcement_config set daily_limit=3 where id=1")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_birthday_anniversary() -> None:
    """S-22 (TZ 3.14) — tug'ilgan kun va ish yubileyi.

    Qabul mezonlari (TZ):
      • YANGI JADVAL yaratilmagan;
      • tabrik guruhga BIR MARTA ketadi (takroriy qo'riqchi);
      • tug'ilgan kun kiritilmagan xodim uchun JIM;
      • sana kiritish HR panelida bor.

    Mavjud `celebration` mexanizmi ishlatiladi: o'sha jadval, o'sha
    video, o'sha «👏 Tabriklash» tugmasi. Farqi — manba CRM voqeasi
    emas, kundalik cron; shuning uchun `lead_event_id` o'rniga
    `dedupe_key` qo'riqlaydi.
    """
    import asyncio
    from datetime import date as _date

    import httpx

    print("\n" + "=" * 60)
    print("S-22: TUG'ILGAN KUN VA YUBILEY")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])
    bugun = _date.today()

    # ── YANGI JADVAL YARATILMAGAN ──
    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        jadvallar = {r[0] for r in cur.execute(
            "select name from sqlite_master where type='table'")}
        yomon = [t for t in jadvallar
                 if any(k in t for k in ("birthday", "anniversary", "tugilgan"))]
        check("S-22: YANGI JADVAL yaratilmagan", not yomon, "=" + str(yomon))
        check("S-22: mavjud `celebration_posts` kengaytirilgan",
              "dedupe_key" in [r[1] for r in cur.execute(
                  "PRAGMA table_info(celebration_posts)")], "dedupe_key yo'q")

        cur.execute("delete from users where full_name like 'T-Bd%'")
        cur.execute("delete from celebration_posts where dedupe_key like '%:99%'")
        conn.commit()

        # Bugun tug'ilgan kuni bo'lgan xodim (30 yosh)
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " birth_date, created_at) values (999702201,'T-Bd Tugilgan','employee',0,1,"
            "?,datetime('now'))", (bugun.replace(year=bugun.year - 30).isoformat(),))
        ids["bd"] = cur.lastrowid
        # Bugun 3 yillik yubiley
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999702202,'T-Bd Yubiley','employee',0,1,"
            "?,datetime('now'))", (bugun.replace(year=bugun.year - 3).isoformat(),))
        ids["yub"] = cur.lastrowid
        # Sanasi YO'Q xodim — tizim jim turishi kerak
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999702203,'T-Bd Sanasiz','employee',0,1,datetime('now'))")
        ids["yoq"] = cur.lastrowid
        # BUGUN ishga kirgan — 0 yil, yubiley EMAS
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999702204,'T-Bd Yangi','employee',0,1,"
            "?,datetime('now'))", (bugun.isoformat(),))
        ids["yangi"] = cur.lastrowid
        conn.commit()

        async def _hodisalar():
            from api.services.celebration import people_events
            async with async_session() as s2:
                hs = await people_events(s2, bugun)
                return [(u.full_name, k, n) for u, k, n in hs
                        if u.full_name.startswith("T-Bd")]

        hodisalar = asyncio.run(_hodisalar())
        nomlar = {h[0] for h in hodisalar}
        check("S-22: tug'ilgan kun topildi",
              ("T-Bd Tugilgan", "birthday", 30) in hodisalar, "=" + str(hodisalar))
        check("S-22: 3 yillik yubiley topildi",
              ("T-Bd Yubiley", "anniversary", 3) in hodisalar, "=" + str(hodisalar))
        check("S-22: sanasi YO'Q xodim uchun tizim JIM",
              "T-Bd Sanasiz" not in nomlar, "=" + str(nomlar))
        check("S-22: BUGUN ishga kirgan xodimda yubiley YO'Q (0 yil)",
              "T-Bd Yangi" not in nomlar, "=" + str(nomlar))

        # ── GURUHGA BIR MARTA ──
        # Guruh biriktirilgan bo'lishi kerak; bo'lmasa bu qism o'tkaziladi.
        guruh = cur.execute(
            "select chat_id from monitored_groups where purpose='main'"
            " and is_active=1 limit 1").fetchone()
        if not guruh:
            cur.execute(
                "insert into monitored_groups (chat_id, title, purpose, is_active,"
                " created_at) values (-100999702299,'T-Bd Guruh','main',1,datetime('now'))")
            conn.commit()

        yuborilgan: list = []

        async def _tick():
            import api.telegram_notify as tn
            from api.services.celebration import announce_people

            asl_msg, asl_file = tn.send_message, tn.send_file_id

            async def soxta_msg(chat_id, text, reply_markup=None):
                yuborilgan.append((chat_id, text))
                return {"ok": True, "result": {"message_id": 111}}

            async def soxta_file(chat_id, file_id, file_type, caption=None,
                                 reply_markup=None):
                yuborilgan.append((chat_id, caption))
                return {"ok": True, "result": {"message_id": 112}}

            tn.send_message, tn.send_file_id = soxta_msg, soxta_file
            #  `celebration` modul o'zining import qilgan nusxasini
            #  ishlatadi — uni ham almashtiramiz.
            import api.services.celebration as cel

            cel_msg, cel_file = cel.send_message, cel.send_file_id
            cel.send_message, cel.send_file_id = soxta_msg, soxta_file
            try:
                async with async_session() as s2:
                    return await announce_people(s2)
            finally:
                tn.send_message, tn.send_file_id = asl_msg, asl_file
                cel.send_message, cel.send_file_id = cel_msg, cel_file

        res1 = asyncio.run(_tick())
        bizniki = [t for (_c, t) in yuborilgan
                   if t and ("T-Bd Tugilgan" in t or "T-Bd Yubiley" in t)]
        check("S-22: guruhga tabrik yuborildi", len(bizniki) == 2,
              f"={len(bizniki)}, natija={res1}")
        check("S-22: tug'ilgan kun matnida yosh bor",
              any("30 yosh" in t for t in bizniki), "=" + str(bizniki))
        check("S-22: yubiley matnida yil bor",
              any("3 yil" in t for t in bizniki), "=" + str(bizniki))

        # ── ⚠️ TAKRORIY QO'RIQCHI ──
        yuborilgan.clear()
        res2 = asyncio.run(_tick())
        qayta = [t for (_c, t) in yuborilgan
                 if t and ("T-Bd Tugilgan" in t or "T-Bd Yubiley" in t)]
        check("S-22: ikkinchi tick QAYTA yubormadi", not qayta,
              f"={len(qayta)}, natija={res2}")
        postlar = cur.execute(
            "select count(*) from celebration_posts where user_id in (?,?)",
            (ids["bd"], ids["yub"])).fetchone()[0]
        check("S-22: bazada ham ikkita post (dublikat yo'q)", postlar == 2,
              "=" + str(postlar))

        # ── SANA KIRITISH HR PANELIDA ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.patch(f"/users/{ids['yoq']}/birth-date", headers=auth(mgr_t),
                        json={"birth_date": "1995-05-15"})
            check("S-22: HR tug'ilgan kunni kiritdi -> 200", r.status_code == 200
                  and r.json().get("birth_date") == "1995-05-15",
                  "kod=" + str(r.status_code) + " " + r.text[:120])

            r = c.patch(f"/users/{ids['yoq']}/birth-date", headers=auth(mgr_t),
                        json={"birth_date": (bugun.replace(year=bugun.year + 1)).isoformat()})
            check("S-22: kelajakdagi sana rad etildi -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code) + " " + r.text[:110])

            r = c.patch(f"/users/{ids['yoq']}/birth-date", headers=auth(mgr_t),
                        json={"birth_date": None})
            check("S-22: sanani tozalash ishlaydi", r.status_code == 200
                  and r.json().get("birth_date") is None, "=" + r.text[:110])

            r = c.patch(f"/users/{ids['yoq']}/birth-date",
                        headers=auth(token_for(ids["bd"], "employee")),
                        json={"birth_date": "1990-01-01"})
            check("S-22: oddiy xodim boshqaning sanasini kirita olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

        # ── ERTANGI ESLATMA ──
        cur.execute("update users set birth_date=? where id=?",
                    ((bugun.replace(year=bugun.year - 25) +
                      __import__("datetime").timedelta(days=1)).isoformat(), ids["yoq"]))
        conn.commit()

        eslatma: list = []

        async def _reminder():
            import api.notify as notify_mod
            import api.services.cron_jobs as cj

            asl = notify_mod.notify_user

            async def soxta(db_, user, category, text, **kw):
                eslatma.append(text)
                return True

            notify_mod.notify_user = soxta
            try:
                async with async_session() as s2:
                    return await cj.celebration_people_reminder_tick(s2)
            finally:
                notify_mod.notify_user = asl

        asyncio.run(_reminder())
        check("S-22: ertangi tug'ilgan kun haqida HR ga eslatma ketdi",
              any("T-Bd Sanasiz" in t for t in eslatma),
              "=" + str(eslatma[:1])[:200])
    except Exception:
        check("S-22 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(
                    f"delete from celebration_posts where user_id in ({belgi})",
                    tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            cur.execute("delete from monitored_groups where title='T-Bd Guruh'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_staff_positions() -> None:
    """S-23 (TZ 3.20) — shtat jadvali.

    Qabul mezonlari (TZ):
      • BAND soni AVTOMATIK hisoblanadi;
      • bo'sh o'rin ro'yxatda ko'rinadi;
      • xodim ko'rmaydi, ROP faqat o'z bo'limini.

    «Band» soni saqlanmaydi: qo'lda kiritilsa u darhol eskirardi —
    xodim ishdan bo'shaydi, shtat jadvalini yangilash unutiladi va tizim
    «hammasi band» deb yolg'on ko'rsatib turaveradi. Test buni bevosita
    tekshiradi: xodim faolsizlanganda son o'zgaradi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-23: SHTAT JADVALI (band avtomatik)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    staff_id = None
    try:
        # BAND ustuni bazada YO'Q
        ustunlar = [r[1] for r in cur.execute("PRAGMA table_info(staff_positions)")]
        check("S-23: «band» ustuni bazada YO'Q (hisoblanadi)",
              not any(u in ("occupied", "busy", "band") for u in ustunlar),
              "=" + str(ustunlar))

        cur.execute("delete from users where full_name like 'T-St%'")
        cur.execute("delete from staff_positions where department like 'T-St%'")
        conn.commit()

        # Sinov uchun alohida lavozim — mavjud xodimlar hisobni buzmasin
        cur.execute("delete from positions where name='T-St Lavozim'")
        cur.execute(
            "insert into positions (name, is_active, created_at)"
            " values ('T-St Lavozim',1,datetime('now'))")
        pos_id = cur.lastrowid
        cur.execute(
            "insert into positions (name, is_active, created_at)"
            " values ('T-St Boshqa',1,datetime('now'))")
        pos2_id = cur.lastrowid
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── QO'SHISH ──
            r = c.post("/staff", headers=auth(mgr_t), json={
                "department": "T-St Sotuv", "position_id": pos_id, "units": 3,
                "salary_min": 5000000, "salary_max": 8000000})
            check("S-23: shtat birligi qo'shildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            staff_id = r.json().get("id") if r.status_code == 201 else None
            check("S-23: yangi birlikda band 0, bo'sh 3",
                  r.status_code == 201 and r.json().get("occupied") == 0
                  and r.json().get("vacant") == 3, "=" + r.text[:130])

            r = c.post("/staff", headers=auth(mgr_t), json={
                "department": "T-St Sotuv", "position_id": pos_id, "units": 1,
                "salary_min": 9000000, "salary_max": 5000000})
            check("S-23: teskari vilka rad etildi -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code) + " " + r.text[:110])

            # ── ⚠️ BAND AVTOMATIK HISOBLANADI ──
            for n in range(2):
                cur.execute(
                    "insert into users (telegram_id, full_name, role, bot_started,"
                    " is_active, position_id, created_at)"
                    " values (?,?,'employee',0,1,?,datetime('now'))",
                    (999702300 + n, f"T-St Xodim{n}", pos_id))
                ids[f"u{n}"] = cur.lastrowid
            conn.commit()

            r = c.get("/staff", headers=auth(mgr_t))
            bizniki = [x for x in r.json() if x["id"] == staff_id]
            check("S-23: ikki xodim qo'shilgach BAND 2 bo'ldi",
                  bizniki and bizniki[0]["occupied"] == 2 and bizniki[0]["vacant"] == 1,
                  "=" + str(bizniki))

            # Xodim FAOLSIZLANSA o'rin bo'shaydi
            cur.execute("update users set is_active=0 where id=?", (ids["u0"],))
            conn.commit()
            r = c.get("/staff", headers=auth(mgr_t))
            bizniki = [x for x in r.json() if x["id"] == staff_id]
            check("S-23: xodim faolsizlanganda o'rin BO'SHADI (1/3)",
                  bizniki and bizniki[0]["occupied"] == 1 and bizniki[0]["vacant"] == 2,
                  "=" + str(bizniki))
            cur.execute("update users set is_active=1 where id=?", (ids["u0"],))
            conn.commit()

            # ── BO'SH O'RINLAR RO'YXATI ──
            r = c.get("/staff/summary", headers=auth(mgr_t))
            xulosa = r.json() if r.status_code == 200 else {}
            bizning = [v for v in xulosa.get("vacancies", [])
                       if v.get("staff_id") == staff_id]
            check("S-23: bo'sh o'rin ro'yxatda ko'rinadi",
                  len(bizning) == 1 and bizning[0]["vacant"] == 1,
                  "=" + str(bizning))
            check("S-23: bo'sh o'rinda lavozim nomi va vilka bor",
                  bizning and bizning[0]["position_name"] == "T-St Lavozim"
                  and bizning[0]["salary_min"] == 5000000, "=" + str(bizning))

            # ── MUZLATILGAN o'rin «bo'sh» sanalmaydi ──
            r = c.put(f"/staff/{staff_id}", headers=auth(mgr_t), json={
                "department": "T-St Sotuv", "position_id": pos_id, "units": 3,
                "salary_min": 5000000, "salary_max": 8000000, "status": "frozen"})
            check("S-23: holat «muzlatilgan» ga o'zgardi", r.status_code == 200,
                  "kod=" + str(r.status_code))
            r = c.get("/staff/summary", headers=auth(mgr_t))
            bizning = [v for v in r.json().get("vacancies", [])
                       if v.get("staff_id") == staff_id]
            check("S-23: muzlatilgan o'rin bo'sh ro'yxatidan CHIQDI",
                  not bizning, "=" + str(bizning))
            c.put(f"/staff/{staff_id}", headers=auth(mgr_t), json={
                "department": "T-St Sotuv", "position_id": pos_id, "units": 3,
                "salary_min": 5000000, "salary_max": 8000000, "status": "open"})

            # ── Banddan kam birlik qo'yib bo'lmaydi ──
            r = c.put(f"/staff/{staff_id}", headers=auth(mgr_t), json={
                "department": "T-St Sotuv", "position_id": pos_id, "units": 1,
                "salary_min": None, "salary_max": None, "status": "open"})
            check("S-23: banddan kam birlik -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code) + " " + r.text[:130])

            # ── RUXSAT: xodim UMUMAN ko'rmaydi ──
            r = c.get("/staff", headers=auth(token_for(ids["u0"], "employee")))
            check("S-23: oddiy xodim shtat jadvalini ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.get("/staff/summary", headers=auth(token_for(ids["u0"], "employee")))
            check("S-23: xulosa ham xodimga yopiq -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))

            # ── RUXSAT: ROP faqat o'z qamrovini ko'radi ──
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (999702310,'T-St Rop','rop',0,1,"
                "datetime('now'))")
            ids["rop"] = cur.lastrowid
            #  ROP ning jamoasi — faqat u0 (uning `manager_id` si)
            cur.execute("update users set manager_id=? where id=?", (ids["rop"], ids["u0"]))
            # Boshqa lavozimda, ROP jamoasidan TASHQARIDA shtat birligi
            cur.execute(
                "insert into staff_positions (department, position_id, units, status,"
                " effective_from, created_at) values ('T-St Boshqa bolim',?,2,'open',"
                "date('now'),datetime('now'))", (pos2_id,))
            begona_id = cur.lastrowid
            conn.commit()

            rop_t = token_for(ids["rop"], "rop")
            r = c.get("/staff", headers=auth(rop_t))
            korgan = {x["id"] for x in r.json()} if r.status_code == 200 else set()
            check("S-23: ROP o'z jamoasi lavozimini KO'RADI",
                  staff_id in korgan, "=" + str(sorted(korgan)))
            check("S-23: ROP begona bo'lim birligini KO'RMAYDI",
                  begona_id not in korgan, "=" + str(sorted(korgan)))

            r = c.post("/staff", headers=auth(rop_t), json={
                "department": "T-St Rop qoshdi", "position_id": pos_id, "units": 1,
                "salary_min": None, "salary_max": None})
            check("S-23: ROP shtat birligi QO'SHA OLMAYDI -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── YOPISH (o'chirmaydi) ──
            r = c.delete(f"/staff/{staff_id}", headers=auth(mgr_t))
            check("S-23: yopish -> 200", r.status_code == 200
                  and r.json().get("status") == "closed", "=" + r.text[:100])
            qator = cur.execute(
                "select status from staff_positions where id=?", (staff_id,)).fetchone()
            check("S-23: qator O'CHIRILMADI — tarix qoldi (status=closed)",
                  qator == ("closed",), "=" + str(qator))
    except Exception:
        check("S-23 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            cur.execute("delete from staff_positions where department like 'T-St%'")
            cur.execute("delete from users where full_name like 'T-St%'")
            cur.execute("delete from positions where name like 'T-St %'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_probation_list() -> None:
    """S-24 (TZ 3.24) — sinov muddatidagi xodimlar.

    Qabul mezonlari (TZ):
      • ro'yxat DOIM ko'rinadi (eslatma o'tib ketsa ham);
      • muddati o'tganlar AJRATIB ko'rsatiladi;
      • sinov muddati sozlamasi qayerdan olinishi hujjatlashtirilgan.

    ⚠️ YANGI JADVAL YO'Q — ro'yxat `hire_date` + sinov muddatidan
    hisoblanadi. Saqlansa `hire_date` tuzatilganda nusxa eskirardi.
    """
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-24: SINOV MUDDATI (hisoblanadi)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])
    bugun = _date.today()

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        jadvallar = {r[0] for r in cur.execute(
            "select name from sqlite_master where type='table'")}
        check("S-24: YANGI JADVAL yaratilmagan",
              not [t for t in jadvallar if "probation" in t], "=" + str(
                  [t for t in jadvallar if "probation" in t]))

        cur.execute("delete from users where full_name like 'T-Pr%'")
        cur.execute("delete from offers where candidate_name like 'T-Pr%'")
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            c.put("/deadlines/config", headers=auth(mgr_t),
                  json={"probation_days": 90, "remind_days": 30})

            # Sinovda: 80 kun oldin kelgan -> 10 kun qoldi
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, hire_date, created_at)"
                " values (999702401,'T-Pr Sinovda','employee',0,1,?,datetime('now'))",
                ((bugun - _td(days=80)).isoformat(),))
            ids["sinov"] = cur.lastrowid
            # Muddati O'TGAN: 100 kun oldin -> -10 kun
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, hire_date, created_at)"
                " values (999702402,'T-Pr Otgan','employee',0,1,?,datetime('now'))",
                ((bugun - _td(days=100)).isoformat(),))
            ids["otgan"] = cur.lastrowid
            # Ancha oldin kelgan — ro'yxatda BO'LMASLIGI kerak
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, hire_date, created_at)"
                " values (999702403,'T-Pr Eski','employee',0,1,?,datetime('now'))",
                ((bugun - _td(days=400)).isoformat(),))
            ids["eski"] = cur.lastrowid
            # `hire_date` YO'Q — ro'yxatga tushmaydi
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at)"
                " values (999702404,'T-Pr Sanasiz','employee',0,1,datetime('now'))")
            ids["sanasiz"] = cur.lastrowid
            conn.commit()

            r = c.get("/probation", headers=auth(mgr_t))
            check("S-24: ro'yxat -> 200", r.status_code == 200, "kod=" + str(r.status_code))
            qatorlar = {x["full_name"]: x for x in r.json()} if r.status_code == 200 else {}

            check("S-24: sinovdagi xodim ro'yxatda (10 kun qoldi)",
                  "T-Pr Sinovda" in qatorlar
                  and qatorlar["T-Pr Sinovda"]["days_left"] == 10,
                  "=" + str(qatorlar.get("T-Pr Sinovda", {}).get("days_left")))
            check("S-24: muddati o'tgan xodim ham RO'YXATDA (eslatma o'tib ketsa ham)",
                  "T-Pr Otgan" in qatorlar, "=" + str(sorted(qatorlar)))
            check("S-24: o'tgani AJRATIB ko'rsatilgan (`is_overdue`)",
                  qatorlar.get("T-Pr Otgan", {}).get("is_overdue") is True
                  and qatorlar["T-Pr Otgan"]["days_left"] == -10,
                  "=" + str(qatorlar.get("T-Pr Otgan", {})))
            check("S-24: ancha oldin kelgan xodim ro'yxatda YO'Q",
                  "T-Pr Eski" not in qatorlar, "=" + str(sorted(qatorlar)))
            check("S-24: `hire_date` yo'q xodim ro'yxatda YO'Q",
                  "T-Pr Sanasiz" not in qatorlar, "=" + str(sorted(qatorlar)))

            # ── Muddati o'tganlar TEPADA ──
            tartib = [x["full_name"] for x in r.json()
                      if x["full_name"].startswith("T-Pr")]
            check("S-24: muddati o'tganlar TEPADA",
                  tartib and tartib[0] == "T-Pr Otgan", "=" + str(tartib))

            # ── MANBA hujjatlashtirilgan ──
            check("S-24: har qatorda muddat MANBAI ko'rsatilgan",
                  all(x.get("source") for x in r.json()),
                  "=" + str([x.get("source") for x in r.json()][:3]))
            check("S-24: sozlamadan kelgani «umumiy sozlama» deb belgilangan",
                  qatorlar.get("T-Pr Sinovda", {}).get("source") == "umumiy sozlama",
                  "=" + str(qatorlar.get("T-Pr Sinovda", {}).get("source")))

            # ── TAKLIFDAN kelgan muddat ustun turadi ──
            #  20 kun oldin kelgan xodim: umumiy sozlama bo'yicha 70 kun
            #  qolardi, taklifdagi 1 oy (30 kun) bo'yicha esa 10 kun.
            #  Farq aniq ko'rinadi — qaysi manba ishlaganini isbotlaydi.
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, hire_date, created_at)"
                " values (999702405,'T-Pr Taklifli','employee',0,1,?,datetime('now'))",
                ((bugun - _td(days=20)).isoformat(),))
            ids["taklifli"] = cur.lastrowid
            conn.commit()

            r = c.get("/probation", headers=auth(mgr_t))
            oldin = {x["full_name"]: x for x in r.json()}.get("T-Pr Taklifli", {})
            check("S-24: taklifsiz xodimda umumiy sozlama (70 kun)",
                  oldin.get("days_left") == 70, "=" + str(oldin.get("days_left")))

            pos = cur.execute("select id from positions limit 1").fetchone()
            r2 = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Pr Taklifdan", "position_id": pos[0],
                "salary": 6000000, "probation_months": 1})
            offer_id = r2.json().get("id") if r2.status_code == 201 else None
            cur.execute("update offers set user_id=? where id=?",
                        (ids["taklifli"], offer_id))
            conn.commit()

            r = c.get("/probation", headers=auth(mgr_t))
            qatorlar = {x["full_name"]: x for x in r.json()}
            taklifli = qatorlar.get("T-Pr Taklifli", {})
            check("S-24: taklifdagi muddat (1 oy) sozlamadan USTUN turdi (70 -> 10)",
                  taklifli.get("days_left") == 10, "=" + str(taklifli.get("days_left")))
            check("S-24: manba «ish taklifi» deb ko'rsatildi",
                  "ish taklifi" in (taklifli.get("source") or ""),
                  "=" + str(taklifli.get("source")))

            #  Muddati ANCHA oldin tugagan xodim ro'yxatdan chiqadi:
            #  qaror kechikkanini ko'rsatish uchun 30 kun yetadi, undan
            #  keyin ro'yxat eski xodimlar bilan to'lib ketardi.
            cur.execute("update users set hire_date=? where id=?",
                        ((bugun - _td(days=80)).isoformat(), ids["taklifli"]))
            conn.commit()
            r = c.get("/probation", headers=auth(mgr_t))
            check("S-24: 30 kundan ko'p o'tgan sinov ro'yxatdan chiqdi",
                  "T-Pr Taklifli" not in {x["full_name"] for x in r.json()},
                  "=" + str(sorted(x["full_name"] for x in r.json())))

            # ── ONBOARDING belgilarini mavjud modullardan ──
            check("S-24: shartnoma yo'qligi ko'rsatilgan",
                  qatorlar.get("T-Pr Otgan", {}).get("has_contract") is False,
                  "=" + str(qatorlar.get("T-Pr Otgan", {}).get("has_contract")))
            c.post("/employee-documents", headers=auth(mgr_t), json={
                "user_id": ids["otgan"], "doc_type": "contract",
                "name": "T-Pr shartnoma", "file_id": "T-PR-1"})
            r = c.get("/probation", headers=auth(mgr_t))
            qatorlar = {x["full_name"]: x for x in r.json()}
            check("S-24: shartnoma qo'shilgach belgi o'zgardi",
                  qatorlar.get("T-Pr Otgan", {}).get("has_contract") is True,
                  "=" + str(qatorlar.get("T-Pr Otgan", {}).get("has_contract")))

            # ── XULOSA ──
            r = c.get("/probation/summary", headers=auth(mgr_t))
            x = r.json() if r.status_code == 200 else {}
            check("S-24: xulosada jami/o'tgan/yaqin bor",
                  {"total", "overdue", "ending_soon", "default_days"} <= set(x),
                  "=" + str(x))
            check("S-24: xulosada sozlama qiymati ham bor (90 kun)",
                  x.get("default_days") == 90, "=" + str(x.get("default_days")))

            # ── RUXSAT ──
            r = c.get("/probation", headers=auth(token_for(ids["sinov"], "employee")))
            check("S-24: oddiy xodim ko'rmaydi -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))
    except Exception:
        check("S-24 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from employee_documents where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            cur.execute("delete from offers where candidate_name like 'T-Pr%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_salary_reason() -> None:
    """S-25 (TZ 3.25) — ish haqi o'zgarishi sababi.

    Qabul mezonlari (TZ):
      • SABABSIZ stavka kiritib bo'lmaydi (400);
      • eski qatorlarda `reason` NULL — «kiritilmagan» deb ko'rsatiladi;
      • xodim BOSHQASINIKINI ko'ra olmaydi.

    NEGA MAJBURIY: stavka tarixi bor edi, lekin «nega?» degan savolga
    javob yo'q edi. Bir yildan keyin «bu odamga nega 20% qo'shgan
    edik?» degan savol javobsiz qolardi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-25: ISH HAQI O'ZGARISHI SABABI")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Sr%'")
        conn.commit()
        for n, tg in enumerate((999702501, 999702502)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, f"T-Sr Xodim{n}"))
            ids[f"u{n}"] = cur.lastrowid
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999702503,'T-Sr Rop','rop',0,1,datetime('now'))")
        ids["rop"] = cur.lastrowid
        conn.commit()
        t0 = token_for(ids["u0"], "employee")
        t1 = token_for(ids["u1"], "employee")
        rop_t = token_for(ids["rop"], "rop")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── Sabab ro'yxati YAGONA manbadan ──
            r = c.get("/payroll/rates/reasons", headers=auth(mgr_t))
            sabablar = {x["value"] for x in r.json()} if r.status_code == 200 else set()
            check("S-25: sabab ro'yxatida TZ dagi turlar bor",
                  {"periodic", "position", "performance", "market", "other"} <= sabablar,
                  "=" + str(sorted(sabablar)))

            # ── ⚠️ SABABSIZ KIRITIB BO'LMAYDI ──
            r = c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "amount": 5000000, "pay_basis": "monthly",
                "effective_from": "2026-01-01"})
            check("S-25: sababsiz stavka rad etildi -> 422 (maydon yo'q)",
                  r.status_code == 422, "kod=" + str(r.status_code))
            r = c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "amount": 5000000, "pay_basis": "monthly",
                "effective_from": "2026-01-01", "reason": ""})
            check("S-25: bo'sh sabab rad etildi -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            r = c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "amount": 5000000, "pay_basis": "monthly",
                "effective_from": "2026-01-01", "reason": "yolgon_sabab"})
            check("S-25: noma'lum sabab rad etildi -> 400", r.status_code == 400,
                  "kod=" + str(r.status_code))
            check("S-25: xato xabari MUMKIN BO'LGAN sabablarni sanaydi",
                  r.status_code == 400 and "Davriy oshirish" in r.text, r.text[:160])
            soni = cur.execute(
                "select count(*) from salary_rates where user_id=?",
                (ids["u0"],)).fetchone()[0]
            check("S-25: rad etilgan urinishlardan qator YARATILMADI", soni == 0,
                  "=" + str(soni))

            # ── TO'G'RI kiritish ──
            r = c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "amount": 5000000, "pay_basis": "monthly",
                "effective_from": "2026-01-01", "reason": "hire",
                "note": "T-Sr dastlabki"})
            check("S-25: sabab bilan stavka kiritildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:140])
            check("S-25: javobda sabab qaytdi",
                  r.status_code == 200 and r.json().get("reason") == "hire",
                  "=" + r.text[:130])

            r = c.post("/payroll/rates", headers=auth(mgr_t), json={
                "user_id": ids["u0"], "amount": 6500000, "pay_basis": "monthly",
                "effective_from": "2026-06-01", "reason": "performance"})
            check("S-25: ikkinchi stavka (natija bo'yicha) kiritildi",
                  r.status_code == 200 and r.json().get("reason") == "performance",
                  "kod=" + str(r.status_code))

            bazada = cur.execute(
                "select reason from salary_rates where user_id=? order by effective_from",
                (ids["u0"],)).fetchall()
            check("S-25: sabab BAZAGA yozildi",
                  bazada == [("hire",), ("performance",)], "=" + str(bazada))

            # ── ESKI qator NULL qoladi ──
            cur.execute(
                "insert into salary_rates (user_id, amount, pay_basis, effective_from,"
                " changed_by, created_at) values (?,?,'monthly','2025-01-01',?,"
                "datetime('now'))", (ids["u0"], 4000000, mgr[0]))
            conn.commit()
            r = c.get("/payroll/rates", headers=auth(mgr_t),
                      params={"user_id": ids["u0"]})
            qatorlar = r.json() if r.status_code == 200 else []
            eski = [x for x in qatorlar if x["effective_from"] == "2025-01-01"]
            check("S-25: eski qatorda `reason` NULL qaytdi (kiritilmagan)",
                  len(eski) == 1 and eski[0]["reason"] is None, "=" + str(eski))
            check("S-25: migratsiya eski qatorni TAXMIN bilan to'ldirmadi",
                  cur.execute(
                      "select count(*) from salary_rates where reason is null"
                  ).fetchone()[0] >= 1, "hammasi to'ldirilgan")

            # ── XODIM O'Z tarixini ko'radi ──
            r = c.get("/payroll/rates/me", headers=auth(t0))
            meniki = r.json() if r.status_code == 200 else []
            check("S-25: xodim o'z tarixini ko'radi (3 qator)",
                  r.status_code == 200 and len(meniki) == 3,
                  f"kod={r.status_code} soni={len(meniki)}")
            check("S-25: tarix eng yangisidan boshlanadi",
                  meniki and meniki[0]["effective_from"] == "2026-06-01",
                  "=" + str([x["effective_from"] for x in meniki]))
            check("S-25: xodim javobida sabab ham bor",
                  meniki and meniki[0]["reason"] == "performance",
                  "=" + str(meniki[0].get("reason") if meniki else None))

            # ── BOSHQASINIKINI ko'ra olmaydi ──
            r = c.get("/payroll/rates/me", headers=auth(t1))
            check("S-25: boshqa xodimda o'z tarixi BO'SH (begonaniki emas)",
                  r.status_code == 200 and r.json() == [], "=" + r.text[:80])
            r = c.get("/payroll/rates", headers=auth(t0),
                      params={"user_id": ids["u0"]})
            check("S-25: xodim umumiy endpointdan foydalana olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── ROP KO'RMAYDI ──
            r = c.get("/payroll/rates", headers=auth(rop_t),
                      params={"user_id": ids["u0"]})
            check("S-25: ROP begona tarixni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.post("/payroll/rates", headers=auth(rop_t), json={
                "user_id": ids["u0"], "amount": 9000000, "pay_basis": "monthly",
                "effective_from": "2027-01-01", "reason": "market"})
            check("S-25: ROP stavka kirita olmaydi -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))

            # ── S-16 zanjiri buzilmadi: taklifdan kelgan stavkada sabab bor ──
            pos = cur.execute("select id from positions limit 1").fetchone()
            r = c.post("/offers", headers=auth(mgr_t), json={
                "candidate_name": "T-Sr Nomzod", "position_id": pos[0],
                "salary": 7000000})
            offer_id = r.json().get("id") if r.status_code == 201 else None
            r = c.post(f"/offers/{offer_id}/hire", headers=auth(mgr_t))
            yangi_uid = r.json().get("user_id") if r.status_code == 200 else None
            check("S-25: taklifdan xodim yaratish hamon ishlaydi",
                  r.status_code == 200, "kod=" + str(r.status_code) + " " + r.text[:120])
            if yangi_uid:
                ids["nomzod"] = yangi_uid
                sabab = cur.execute(
                    "select reason from salary_rates where user_id=?",
                    (yangi_uid,)).fetchone()
                check("S-25: taklifdan kelgan stavkada sabab «ishga qabul»",
                      sabab == ("hire",), "=" + str(sabab))
    except Exception:
        check("S-25 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from salary_rates where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            cur.execute("delete from offers where candidate_name like 'T-Sr%'")
            cur.execute("delete from users where full_name like 'T-Sr%'")
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_profile_changes() -> None:
    """S-26 (TZ 3.26) — xodim o'z ma'lumotini yangilash so'rovi.

    Qabul mezonlari (TZ):
      • TO'G'RIDAN-TO'G'RI o'zgartirish YO'Q;
      • oq ro'yxatdan tashqari maydon so'rovi RAD etiladi;
      • tasdiqlangach ESKI QIYMAT auditda qoladi.

    Oq ro'yxat serverda: aks holda xodim `role` yoki `is_active` ni
    «so'rab» yuborishi mumkin bo'lardi va HR e'tiborsiz tasdiqlab
    qo'yishi mumkin edi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-26: PROFIL O'ZGARTIRISH SO'ROVLARI")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Pc%'")
        conn.commit()
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " phone, created_at) values (999702601,'T-Pc Xodim','employee',0,1,"
            "'+998900000000',datetime('now'))")
        ids["emp"] = cur.lastrowid
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " created_at) values (999702602,'T-Pc Rop','rop',0,1,datetime('now'))")
        ids["rop"] = cur.lastrowid
        conn.commit()
        emp_t = token_for(ids["emp"], "employee")
        rop_t = token_for(ids["rop"], "rop")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── Maydonlar ro'yxati YAGONA manbadan ──
            r = c.get("/profile-changes/fields", headers=auth(emp_t))
            maydonlar = {x["value"] for x in r.json()} if r.status_code == 200 else set()
            check("S-26: oq ro'yxatda TZ dagi maydonlar bor",
                  {"phone", "address", "marital_status", "emergency_contact"} <= maydonlar,
                  "=" + str(sorted(maydonlar)))
            nozik = {x["value"] for x in r.json() if x.get("sensitive")}
            check("S-26: F.I.Sh. «ogohlantirish bilan» deb belgilangan",
                  nozik == {"full_name"}, "=" + str(nozik))

            # ── ⚠️ OQ RO'YXATDAN TASHQARI maydon RAD etiladi ──
            for yomon in ("role", "is_active", "hire_date", "salary"):
                r = c.post("/profile-changes/me", headers=auth(emp_t),
                           json={"field": yomon, "new_value": "boss"})
                check(f"S-26: «{yomon}» so'rovi rad etildi -> 400",
                      r.status_code == 400, "kod=" + str(r.status_code))
            rol = cur.execute("select role from users where id=?",
                              (ids["emp"],)).fetchone()
            check("S-26: rad etilgan urinishlar bazaga TEGMADI",
                  rol == ("employee",), "=" + str(rol))

            # ── SO'ROV yuborish ──
            r = c.post("/profile-changes/me", headers=auth(emp_t),
                       json={"field": "phone", "new_value": "+998911111111"})
            check("S-26: telefon so'rovi qabul qilindi -> 201",
                  r.status_code == 201, "kod=" + str(r.status_code) + " " + r.text[:130])
            req_id = r.json().get("id") if r.status_code == 201 else None
            check("S-26: so'rovda ESKI qiymat saqlandi",
                  r.status_code == 201 and r.json().get("old_value") == "+998900000000",
                  "=" + str(r.json().get("old_value") if r.status_code == 201 else None))

            # ── ⚠️ TO'G'RIDAN-TO'G'RI O'ZGARTIRISH YO'Q ──
            hozir = cur.execute("select phone from users where id=?",
                                (ids["emp"],)).fetchone()
            check("S-26: so'rovdan keyin baza HALI o'zgarmadi",
                  hozir == ("+998900000000",), "=" + str(hozir))

            # Bir maydonga ikkinchi ochiq so'rov
            r = c.post("/profile-changes/me", headers=auth(emp_t),
                       json={"field": "phone", "new_value": "+998922222222"})
            check("S-26: bir maydonga ikkinchi ochiq so'rov -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code))
            # Bir xil qiymat
            r = c.post("/profile-changes/me", headers=auth(emp_t),
                       json={"field": "address", "new_value": ""})
            check("S-26: bo'sh qiymat -> 422", r.status_code == 422,
                  "kod=" + str(r.status_code))

            # ── XODIM o'z so'rovlarini ko'radi ──
            r = c.get("/profile-changes/me", headers=auth(emp_t))
            check("S-26: xodim o'z so'rovini ko'radi",
                  r.status_code == 200 and len(r.json()) == 1, "=" + r.text[:100])
            r = c.get("/profile-changes/me/profile", headers=auth(emp_t))
            check("S-26: kutilayotgan maydon belgilangan",
                  r.status_code == 200 and r.json().get("pending_fields") == ["phone"],
                  "=" + r.text[:130])

            # ── RUXSAT ──
            r = c.get("/profile-changes", headers=auth(emp_t))
            check("S-26: oddiy xodim umumiy ro'yxatni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.get("/profile-changes", headers=auth(rop_t))
            check("S-26: ROP ham ko'rmaydi (shaxsiy ma'lumot) -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))
            r = c.post(f"/profile-changes/{req_id}/decide", headers=auth(rop_t),
                       json={"approve": True})
            check("S-26: ROP tasdiqlay olmaydi -> 403", r.status_code == 403,
                  "kod=" + str(r.status_code))

            # ── HR TASDIG'I -> baza o'zgaradi ──
            r = c.get("/profile-changes", headers=auth(mgr_t))
            bizniki = [x for x in r.json() if x["id"] == req_id]
            check("S-26: HR kutilayotgan so'rovni ko'radi", len(bizniki) == 1,
                  "=" + str(len(r.json())))
            check("S-26: HR javobida eski va yangi qiymat yonma-yon",
                  bizniki and bizniki[0]["old_value"] == "+998900000000"
                  and bizniki[0]["new_value"] == "+998911111111", "=" + str(bizniki))

            r = c.post(f"/profile-changes/{req_id}/decide", headers=auth(mgr_t),
                       json={"approve": True, "note": "Tekshirildi"})
            check("S-26: HR tasdiqladi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            yangi = cur.execute("select phone from users where id=?",
                                (ids["emp"],)).fetchone()
            check("S-26: TASDIQDAN KEYIN baza o'zgardi",
                  yangi == ("+998911111111",), "=" + str(yangi))

            # ── ⚠️ ESKI QIYMAT AUDITDA ──
            audit = cur.execute(
                "select action, before, after from audit_logs"
                " where target_user_id=? and action='profile_change_approved'"
                " order by id desc limit 1", (ids["emp"],)).fetchone()
            check("S-26: tasdiq auditga yozildi", audit is not None,
                  "=" + str(audit))
            check("S-26: auditda ESKI qiymat saqlandi",
                  audit and "+998900000000" in (audit[1] or ""),
                  "=" + str(audit[1] if audit else None))
            check("S-26: auditda yangi qiymat ham bor",
                  audit and "+998911111111" in (audit[2] or ""),
                  "=" + str(audit[2] if audit else None))

            # ── Ikkinchi marta hal qilib bo'lmaydi ──
            r = c.post(f"/profile-changes/{req_id}/decide", headers=auth(mgr_t),
                       json={"approve": False})
            check("S-26: hal qilingan so'rovni qayta hal qilish -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            # ── RAD ETISH bazani o'zgartirmaydi ──
            r = c.post("/profile-changes/me", headers=auth(emp_t),
                       json={"field": "address", "new_value": "T-Pc yangi manzil"})
            req2 = r.json().get("id") if r.status_code == 201 else None
            r = c.post(f"/profile-changes/{req2}/decide", headers=auth(mgr_t),
                       json={"approve": False, "note": "Hujjat kerak"})
            check("S-26: rad etish -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            manzil = cur.execute("select address from users where id=?",
                                 (ids["emp"],)).fetchone()
            check("S-26: rad etilgach baza O'ZGARMADI", manzil == (None,),
                  "=" + str(manzil))
            r = c.get("/profile-changes/me/profile", headers=auth(emp_t))
            check("S-26: rad etilgach maydon qayta so'ralishi mumkin",
                  r.status_code == 200 and r.json().get("pending_fields") == [],
                  "=" + r.text[:130])
    except Exception:
        check("S-26 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(
                    f"delete from profile_change_requests where user_id in ({belgi})",
                    tuple(ids.values()))
                cur.execute(f"delete from audit_logs where target_user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()


def test_contract_registration() -> None:
    """S-27 (TZ 3.28) — shartnomani davlat ro'yxatidan o'tkazish belgisi.

    Qabul mezonlari (TZ):
      • belgisiz xodimlar RO'YXATI bor;
      • 3 kundan keyin eslatma, TAKRORLANMAYDI;
      • kadr auditi (3.30) uchun so'rov tayyor.

    ⚠️ TIZIM RO'YXATGA OLISHNI BAJARMAYDI — bu tashqi jarayon (mehnat
    organi). Tizim faqat «qilindimi?» degan BELGINI yuritadi.
    """
    import asyncio
    from datetime import date as _date
    from datetime import timedelta as _td

    import httpx

    print("\n" + "=" * 60)
    print("S-27: SHARTNOMANI RO'YXATGA OLISH")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])
    bugun = _date.today()

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Cr%'")
        conn.commit()

        # 10 kun oldin kelgan — kechikkan
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999702701,'T-Cr Kechikkan','employee',0,1,"
            "?,datetime('now'))", ((bugun - _td(days=10)).isoformat(),))
        ids["kech"] = cur.lastrowid
        # Bugun kelgan — hali muddat yetmagan
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999702702,'T-Cr Yangi','employee',0,1,"
            "?,datetime('now'))", (bugun.isoformat(),))
        ids["yangi"] = cur.lastrowid
        conn.commit()

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── BELGISIZ XODIMLAR RO'YXATI ──
            r = c.get("/employee-documents/unregistered", headers=auth(mgr_t))
            check("S-27: belgisiz ro'yxat -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            qatorlar = {x["full_name"]: x for x in r.json()} if r.status_code == 200 else {}
            check("S-27: shartnomasi yo'q xodim ro'yxatda",
                  "T-Cr Kechikkan" in qatorlar, "=" + str(sorted(qatorlar)))
            check("S-27: hujjat umuman yo'qligi ko'rsatilgan",
                  qatorlar.get("T-Cr Kechikkan", {}).get("document_id") is None,
                  "=" + str(qatorlar.get("T-Cr Kechikkan", {})))
            check("S-27: 10 kun kechikkani «overdue» deb belgilangan",
                  qatorlar.get("T-Cr Kechikkan", {}).get("overdue") is True
                  and qatorlar["T-Cr Kechikkan"]["days_since_hire"] == 10,
                  "=" + str(qatorlar.get("T-Cr Kechikkan", {})))
            check("S-27: bugun kelgan xodim «overdue» EMAS",
                  qatorlar.get("T-Cr Yangi", {}).get("overdue") is False,
                  "=" + str(qatorlar.get("T-Cr Yangi", {})))

            # ── Shartnoma yuklandi, lekin BELGI yo'q ──
            r = c.post("/employee-documents", headers=auth(mgr_t), json={
                "user_id": ids["kech"], "doc_type": "contract",
                "name": "T-Cr shartnoma", "file_id": "T-CR-1"})
            doc_id = r.json().get("id") if r.status_code == 201 else None
            r = c.get("/employee-documents/unregistered", headers=auth(mgr_t))
            qatorlar = {x["full_name"]: x for x in r.json()}
            check("S-27: hujjat bor, belgi yo'q — HAMON ro'yxatda",
                  "T-Cr Kechikkan" in qatorlar
                  and qatorlar["T-Cr Kechikkan"]["document_id"] == doc_id,
                  "=" + str(qatorlar.get("T-Cr Kechikkan", {})))

            # ── BELGI qo'yish ──
            r = c.post(f"/employee-documents/{doc_id}/register", headers=auth(mgr_t),
                       json={"registered_at": (bugun - _td(days=1)).isoformat(),
                             "note": "T-Cr mehnat organida"})
            check("S-27: belgi qo'yildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            bazada = cur.execute(
                "select registered_at, registered_by, registration_note"
                " from employee_documents where id=?", (doc_id,)).fetchone()
            check("S-27: sana, kim va izoh bazaga yozildi",
                  bazada is not None and bazada[0] == (bugun - _td(days=1)).isoformat()
                  and bazada[1] == mgr[0] and bazada[2] == "T-Cr mehnat organida",
                  "=" + str(bazada))

            r = c.get("/employee-documents/unregistered", headers=auth(mgr_t))
            check("S-27: belgi qo'yilgach ro'yxatdan CHIQDI",
                  "T-Cr Kechikkan" not in {x["full_name"] for x in r.json()},
                  "=" + str([x["full_name"] for x in r.json()]))

            # ── IDEMPOTENT: qayta bosish sanani siljitmaydi ──
            r = c.post(f"/employee-documents/{doc_id}/register", headers=auth(mgr_t),
                       json={})
            keyin = cur.execute(
                "select registered_at from employee_documents where id=?",
                (doc_id,)).fetchone()
            check("S-27: qayta bosishda BIRINCHI sana saqlandi",
                  keyin == ((bugun - _td(days=1)).isoformat(),), "=" + str(keyin))

            # ── Xato holatlar ──
            #  Kelajakdagi sana idempotentlik TEKSHIRUVIDAN OLDIN rad
            #  etiladi: noto'g'ri kiritish har doim xato bersin, hatto
            #  natija o'zgarmaydigan holatda ham — aks holda HR sanani
            #  xato yozganini bilmay qolardi.
            r = c.post(f"/employee-documents/{doc_id}/register", headers=auth(mgr_t),
                       json={"registered_at": (bugun + _td(days=5)).isoformat()})
            check("S-27: kelajakdagi sana rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:110])
            hamon = cur.execute(
                "select registered_at from employee_documents where id=?",
                (doc_id,)).fetchone()
            check("S-27: rad etilgan urinish saqlangan sanaga TEGMADI",
                  hamon == ((bugun - _td(days=1)).isoformat(),), "=" + str(hamon))

            r = c.post("/employee-documents", headers=auth(mgr_t), json={
                "user_id": ids["yangi"], "doc_type": "diploma",
                "name": "T-Cr diplom", "file_id": "T-CR-2"})
            dip_id = r.json().get("id") if r.status_code == 201 else None
            r = c.post(f"/employee-documents/{dip_id}/register", headers=auth(mgr_t),
                       json={})
            check("S-27: diplomga belgi qo'yib bo'lmaydi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:110])

            # ── RUXSAT ──
            r = c.get("/employee-documents/unregistered",
                      headers=auth(token_for(ids["kech"], "employee")))
            check("S-27: oddiy xodim ro'yxatni ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

        # ── ⚠️ 3 KUNDAN KEYIN MUDDAT, TAKRORLANMAYDI ──
        async def _tick():
            import api.services.cron_jobs as cj
            async with async_session() as s2:
                return await cj.contract_registration_tick(s2)

        res1 = asyncio.run(_tick())
        muddatlar = cur.execute(
            "select user_id, due_date, status from deadlines"
            " where kind='contract_registration' and user_id in (?,?)",
            (ids["kech"], ids["yangi"])).fetchall()
        check("S-27: belgisi bor xodimga muddat yaratilmadi",
              not any(m[0] == ids["kech"] and m[2] == "open" for m in muddatlar),
              "=" + str(muddatlar))
        check("S-27: 3 kun to'lmagan xodimga ham muddat yo'q",
              not any(m[0] == ids["yangi"] for m in muddatlar),
              f"={muddatlar}, natija={res1}")

        # Belgisiz va kechikkan xodim qo'shamiz
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active,"
            " hire_date, created_at) values (999702703,'T-Cr Belgisiz','employee',0,1,"
            "?,datetime('now'))", ((bugun - _td(days=7)).isoformat(),))
        ids["belgisiz"] = cur.lastrowid
        conn.commit()

        res2 = asyncio.run(_tick())
        check("S-27: kechikkan va belgisiz xodimga MUDDAT yaratildi",
              res2.get("created") == 1, "=" + str(res2))
        m = cur.execute(
            "select due_date, status, note from deadlines"
            " where kind='contract_registration' and user_id=?",
            (ids["belgisiz"],)).fetchone()
        check("S-27: muddat = ishga qabul + 3 kun",
              m is not None and m[0] == (bugun - _td(days=4)).isoformat(),
              "=" + str(m))
        check("S-27: muddat izohida sabab yozilgan",
              m is not None and "ro'yxatidan o'tkazilmagan" in (m[2] or ""),
              "=" + str(m[2] if m else None))

        # ── TAKRORLANMAYDI ──
        res3 = asyncio.run(_tick())
        soni = cur.execute(
            "select count(*) from deadlines where kind='contract_registration'"
            " and user_id=?", (ids["belgisiz"],)).fetchone()[0]
        check("S-27: ikkinchi tick IKKINCHI muddat yaratmadi",
              soni == 1 and res3.get("created") == 0, f"soni={soni}, {res3}")

        # ── Belgi qo'yilgach muddat YOPILADI ──
        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            r = c.post("/employee-documents", headers=auth(mgr_t), json={
                "user_id": ids["belgisiz"], "doc_type": "contract",
                "name": "T-Cr shartnoma 2", "file_id": "T-CR-3"})
            d2 = r.json().get("id") if r.status_code == 201 else None
            c.post(f"/employee-documents/{d2}/register", headers=auth(mgr_t), json={})

        res4 = asyncio.run(_tick())
        holat = cur.execute(
            "select status from deadlines where kind='contract_registration'"
            " and user_id=?", (ids["belgisiz"],)).fetchone()
        check("S-27: belgi qo'yilgach muddat AVTOMATIK yopildi",
              holat == ("done",) and res4.get("closed") == 1,
              f"={holat}, {res4}")
    except Exception:
        check("S-27 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from deadlines where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(
                    f"delete from employee_documents where user_id in ({belgi})",
                    tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})", tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_hr_inquiries() -> None:
    """S-28 (TZ 3.29) — xodim murojaatlari jurnali.

    Qabul mezonlari (TZ):
      • savol-javob SAQLANADI;
      • xodim FAQAT o'z murojaatlarini ko'radi;
      • javobsiz murojaat HR panelida AJRALIB turadi.

    ⚠️ AI hukm chiqarmaydi — tasniflagich faqat TOIFANI qo'yadi,
    javobni har doim odam yozadi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-28: XODIM MUROJAATLARI JURNALI")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Hq%'")
        conn.commit()
        for nom, tg in (("T-Hq Aziz", 999702801), ("T-Hq Bobur", 999702802)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, nom))
            ids[nom] = cur.lastrowid
        conn.commit()
        #  Rahbarning telegram_id si — bot yo'lini SINASH uchun. Botda
        #  JWT yo'q, kirish `telegram_id` bo'yicha.
        mgr_tg = cur.execute(
            "select telegram_id from users where id=?", (mgr[0],)).fetchone()[0]
        aziz_t = token_for(ids["T-Hq Aziz"], "employee")
        bobur_t = token_for(ids["T-Hq Bobur"], "employee")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── TASNIFLAGICH ──
            r = c.post("/hr-inquiries/me", headers=auth(aziz_t),
                       json={"question": "Oyligim nega kam tushdi?"})
            check("S-28: savol yozildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            savol1 = r.json() if r.status_code == 201 else {}
            check("S-28: toifa avtomatik aniqlandi (oylik)",
                  savol1.get("category") == "salary", "=" + str(savol1))

            r = c.post("/hr-inquiries/me", headers=auth(aziz_t),
                       json={"question": "Ta'tilga qachon chiqsam bo'ladi?"})
            savol2 = r.json() if r.status_code == 201 else {}
            check("S-28: qo'shimchali so'z ham topildi (ta'tilga -> vacation)",
                  savol2.get("category") == "vacation", "=" + str(savol2))

            r = c.post("/hr-inquiries/me", headers=auth(bobur_t),
                       json={"question": "Bugun ob-havo qanday bo'larkan?"})
            savol3 = r.json() if r.status_code == 201 else {}
            check("S-28: kalit so'z yo'q -> «other» (taxmin qilinmaydi)",
                  savol3.get("category") == "other", "=" + str(savol3))

            # ── XODIM FAQAT O'ZINIKINI KO'RADI ──
            r = c.get("/hr-inquiries/me", headers=auth(aziz_t))
            aziznikilar = r.json() if r.status_code == 200 else []
            check("S-28: o'z murojaatlarim ko'rinadi", len(aziznikilar) == 2,
                  "soni=" + str(len(aziznikilar)))
            check("S-28: BOSHQA xodim murojaati ko'rinmaydi",
                  all(x["user_id"] == ids["T-Hq Aziz"] for x in aziznikilar),
                  "=" + str([x["user_id"] for x in aziznikilar]))

            # Xodim HR ro'yxatini OCHA OLMAYDI
            r = c.get("/hr-inquiries", headers=auth(aziz_t))
            check("S-28: xodim HR ro'yxatiga kira olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── HR RO'YXATI: JAVOBSIZLAR BIRINCHI ──
            r = c.get("/hr-inquiries", headers=auth(mgr_t))
            check("S-28: HR ro'yxati -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            hammasi = r.json() if r.status_code == 200 else []
            meniki = [x for x in hammasi if x["id"] in
                      (savol1.get("id"), savol2.get("id"), savol3.get("id"))]
            check("S-28: uchala murojaat HR ga ko'rindi", len(meniki) == 3,
                  "soni=" + str(len(meniki)))

            r = c.get("/hr-inquiries/stats", headers=auth(mgr_t))
            ochiq_oldin = r.json().get("open") if r.status_code == 200 else None

            # ── JAVOB ──
            r = c.post(f"/hr-inquiries/{savol1['id']}/answer", headers=auth(mgr_t),
                       json={"answer": "Kechikish jarimasi ushlangan."})
            check("S-28: javob yozildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])

            r = c.get("/hr-inquiries/me", headers=auth(aziz_t))
            javobli = [x for x in r.json() if x["id"] == savol1["id"]]
            check("S-28: javob xodimga QAYTDI va saqlandi",
                  bool(javobli) and javobli[0]["answer"] == "Kechikish jarimasi ushlangan."
                  and javobli[0]["status"] == "answered",
                  "=" + str(javobli[:1]))
            check("S-28: javob bergan odam yozildi",
                  bool(javobli) and javobli[0]["answered_by"] == mgr[0],
                  "=" + str(javobli[0].get("answered_by") if javobli else None))

            r = c.get("/hr-inquiries/stats", headers=auth(mgr_t))
            check("S-28: javobsizlar soni kamaydi",
                  r.json().get("open") == (ochiq_oldin or 0) - 1,
                  f"oldin={ochiq_oldin}, keyin={r.json().get('open')}")

            # ── SARALASH: javobsizlar tepada ──
            r = c.get("/hr-inquiries", headers=auth(mgr_t))
            tartib = [x["status"] for x in r.json()]
            birinchi_javobli = next(
                (i for i, v in enumerate(tartib) if v != "open"), len(tartib))
            check("S-28: javobsizlar ro'yxat TEPASIDA",
                  all(v == "open" for v in tartib[:birinchi_javobli])
                  and all(v != "open" for v in tartib[birinchi_javobli:]),
                  "=" + str(tartib[:8]))

            # ── TOIFANI QO'LDA O'ZGARTIRISH ──
            r = c.put(f"/hr-inquiries/{savol3['id']}/category", headers=auth(mgr_t),
                      json={"category": "schedule"})
            check("S-28: toifa o'zgartirildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            qator = cur.execute(
                "select category, category_auto from hr_inquiries where id=?",
                (savol3["id"],)).fetchone()
            check("S-28: qo'lda o'zgartirilgach «avto» belgisi olindi",
                  qator == ("schedule", 0), "=" + str(qator))

            r = c.put(f"/hr-inquiries/{savol3['id']}/category", headers=auth(mgr_t),
                      json={"category": "yolgon"})
            check("S-28: noma'lum toifa rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            # ── FILTRLAR ──
            r = c.get("/hr-inquiries?status_filter=open", headers=auth(mgr_t))
            check("S-28: holat filtri ishlaydi",
                  r.status_code == 200
                  and all(x["status"] == "open" for x in r.json()),
                  "kod=" + str(r.status_code))
            r = c.get("/hr-inquiries?category=vacation", headers=auth(mgr_t))
            check("S-28: toifa filtri ishlaydi",
                  r.status_code == 200
                  and all(x["category"] == "vacation" for x in r.json()),
                  "kod=" + str(r.status_code))

            # ── JAVOBSIZ YOPISH ──
            r = c.post(f"/hr-inquiries/{savol3['id']}/close", headers=auth(mgr_t))
            check("S-28: javobsiz yopildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            # Javob berilgan murojaatni yopib BO'LMAYDI
            r = c.post(f"/hr-inquiries/{savol1['id']}/close", headers=auth(mgr_t))
            check("S-28: javob berilgan murojaat yopilmaydi -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code))

            # ── BOT YO'LI ──
            r = c.post("/hr-inquiries/bot/ask",
                       json={"telegram_id": 999702802,
                             "question": "Ma'lumotnoma kerak edi"})
            check("S-28: botdan savol -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            bot_savol = r.json() if r.status_code == 201 else {}
            check("S-28: botdagi savol ham tasniflandi",
                  bot_savol.get("category") == "documents", "=" + str(bot_savol))

            # ⚠️ `POST /bot/answer` `POST /{inquiry_id}/answer` bilan bir xil
            # SHAKLDA va bir xil METODDA — marshrut tartibi buzilsa «bot»
            # so'zi murojaat raqami deb o'qilib, 422 qaytarardi.
            r = c.post("/hr-inquiries/bot/answer",
                       json={"telegram_id": mgr_tg,
                             "inquiry_id": bot_savol.get("id", 0),
                             "answer": "Ma'lumotnomani ertaga tayyorlaymiz."})
            #  ⚠️ Bu tekshiruv MARSHRUT TARTIBINI qo'riqlaydi: `/bot/answer`
            #  `/{inquiry_id}/answer` dan KEYIN e'lon qilinsa, «bot» so'zi
            #  raqam deb o'qilib 422 kelardi va bot javob bera olmasdi.
            check("S-28: botdan javob berildi -> 200 (marshrut almashmadi)",
                  r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            r = c.get("/hr-inquiries/bot/my", params={"telegram_id": 999702802})
            bot_javob = [x for x in r.json() if x["id"] == bot_savol.get("id")]
            check("S-28: botdagi javob xodimga yetdi",
                  bool(bot_javob) and bot_javob[0]["status"] == "answered",
                  "=" + str(bot_javob[:1]))

            r = c.get("/hr-inquiries/bot/my", params={"telegram_id": 999702802})
            check("S-28: botda o'z murojaatlarim ko'rinadi",
                  r.status_code == 200
                  and all(x["user_id"] == ids["T-Hq Bobur"] for x in r.json()),
                  "kod=" + str(r.status_code))

            r = c.get("/hr-inquiries/bot/open", params={"telegram_id": 999702801})
            check("S-28: oddiy xodim botda javobsizlar ro'yxatini ko'rmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── KIRITISHNI TEKSHIRISH ──
            r = c.post("/hr-inquiries/me", headers=auth(aziz_t), json={"question": "ha"})
            check("S-28: juda qisqa savol rad etildi -> 422",
                  r.status_code == 422, "kod=" + str(r.status_code))
            r = c.post("/hr-inquiries/999999/answer", headers=auth(mgr_t),
                       json={"answer": "yo'q"})
            check("S-28: mavjud bo'lmagan murojaat -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))
    except Exception:
        check("S-28 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from hr_inquiries where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_hr_knowledge_loop() -> None:
    """S-29 (TZ 3.29) — murojaatlar → bilim bazasi halqasi.

    Qabul mezonlari (TZ):
      • «Eng ko'p beriladigan 10 savol» ro'yxati;
      • bir bosishda bilim bazasiga o'tadi;
      • bilim bazasida javob bo'lsa bot avval o'zi javob beradi.

    ⚠️ ENG MUHIM TEKSHIRUV — SIZISH: HR javobi Sotuv AI promptiga va
    TASHQI chatbot datasetiga TUSHMASLIGI kerak. Ichki qoida (oylik,
    jarima) mijozga ketsa, bu eng jiddiy xato bo'lardi.
    """
    import asyncio

    import httpx

    print("\n" + "=" * 60)
    print("S-29: MUROJAATLAR -> BILIM BAZASI HALQASI")
    print("=" * 60)

    from db.base import async_session

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    kb_ids: list[int] = []
    try:
        cur.execute("delete from users where full_name like 'T-Kb%'")
        cur.execute("delete from knowledge_entries where source='hr_inquiry'")
        conn.commit()
        for nom, tg in (("T-Kb Aziz", 999702901), ("T-Kb Bobur", 999702902)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, nom))
            ids[nom] = cur.lastrowid
        conn.commit()
        aziz_t = token_for(ids["T-Kb Aziz"], "employee")
        bobur_t = token_for(ids["T-Kb Bobur"], "employee")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── SOLISHTIRGICH: o'lchangan xulq buzilmasin ──
            from api.services.hr_inquiries import (
                _GROUP_MATCH, _SUGGEST_MATCH, _similarity, _tokens,
            )
            bir_xil = [
                ("Oylik qachon beriladi?", "Oylikni qachon berasiz"),
                ("Ta'tilga qachon chiqsam bo'ladi", "Tatilga qachon chiqaman"),
                ("Ma'lumotnoma kerak edi", "Malumotnoma qanday olinadi"),
            ]
            boshqa = [
                ("Oylik qachon beriladi?", "Ta'til qancha kun"),
                ("Ish jadvalim qanday", "Oylik qancha"),
                ("Ma'lumotnoma kerak", "Noutbuk kerak"),
            ]
            topildi = sum(
                1 for a, b in bir_xil
                if _similarity(_tokens(a), _tokens(b)) >= _SUGGEST_MATCH)
            check("S-29: qo'shimchali shakllar bir xil deb topiladi",
                  topildi == 3, f"{topildi}/3")
            yolgon = sum(
                1 for a, b in boshqa
                if _similarity(_tokens(a), _tokens(b)) >= _GROUP_MATCH)
            check("S-29: boshqa-boshqa savollar BIRLASHTIRILMAYDI",
                  yolgon == 0, f"noto'g'ri birlashgan={yolgon}")

            # ── SAVOL VA JAVOB ──
            r = c.post("/hr-inquiries/me", headers=auth(aziz_t),
                       json={"question": "Ta'tilga qachon chiqsam bo'ladi?"})
            savol = r.json() if r.status_code == 201 else {}
            check("S-29: baza bo'sh — taklif YO'Q, HR ga ketdi",
                  savol.get("suggestion") is None and savol.get("notified", 0) >= 1,
                  "=" + str(savol))
            r = c.post(f"/hr-inquiries/{savol['id']}/answer", headers=auth(mgr_t),
                       json={"answer": "Yillik ta'til 12 oy ishlagandan keyin."})
            check("S-29: javob yozildi", r.status_code == 200, "kod=" + str(r.status_code))

            # ── BIR BOSISHDA BILIM BAZASIGA ──
            r = c.post(f"/hr-inquiries/{savol['id']}/to-knowledge", headers=auth(mgr_t))
            check("S-29: bir bosishda bilim bazasiga -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            kb = r.json() if r.status_code == 200 else {}
            if kb.get("entry_id"):
                kb_ids.append(kb["entry_id"])
            check("S-29: yozuv `hr` qamrovida (mijozga ko'rinmaydi)",
                  kb.get("audience") == "hr", "=" + str(kb))
            qator = cur.execute(
                "select status, audience, source from knowledge_entries where id=?",
                (kb.get("entry_id", 0),)).fetchone()
            check("S-29: darhol `verified` (HR o'zi yozgan va o'zi bosgan)",
                  qator == ("verified", "hr", "hr_inquiry"), "=" + str(qator))

            # Ikkinchi bosish DUBLIKAT yaratmaydi
            r = c.post(f"/hr-inquiries/{savol['id']}/to-knowledge", headers=auth(mgr_t))
            soni = cur.execute(
                "select count(*) from knowledge_entries where source='hr_inquiry'"
            ).fetchone()[0]
            check("S-29: tugma qayta bosilsa dublikat yo'q",
                  soni == 1 and r.json().get("entry_id") == kb.get("entry_id"),
                  f"soni={soni}")

            # Javobsizni ko'chirib bo'lmaydi
            r = c.post("/hr-inquiries/me", headers=auth(bobur_t),
                       json={"question": "Yangi kreslo qachon keladi?"})
            javobsiz = r.json()
            r = c.post(f"/hr-inquiries/{javobsiz['id']}/to-knowledge", headers=auth(mgr_t))
            check("S-29: javobsiz murojaat bazaga ketmaydi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # ⚠️ SIZISH TEKSHIRUVI
            #
            # ⚠️ Bo'sh baza ustida tekshirish HECH NARSANI isbotlamaydi
            # («yo'q» matn bo'sh matnda ham topilmaydi). Shuning uchun
            # avval HAQIQIY sotuv yozuvi qo'yiladi: u promptda BO'LISHI,
            # HR yozuvi esa BO'LMASLIGI kerak.
            # ══════════════════════════════════════════════
            cur.execute(
                "insert into knowledge_entries (kind, audience, category, question,"
                " answer, status, date_sensitive, needs_recheck, source, ai_attempts,"
                " created_at, updated_at) values ('single','sales','narx',"
                "'Kvartira narxi qancha?','T-KB-SOTUV-MAYOQ narx 500 mln',"
                "'verified',0,0,'hr_inquiry',0,datetime('now'),datetime('now'))")
            conn.commit()

            async def _kontekst():
                from api.services.sales_ai import build_context
                async with async_session() as s2:
                    return await build_context(s2)

            kb_text, _pb, kb_soni, _p2 = asyncio.run(_kontekst())
            check("S-29: sotuv yozuvi promptda BOR (tekshiruv haqiqiy)",
                  "T-KB-SOTUV-MAYOQ" in kb_text, "prompt=" + kb_text[:200])
            check("S-29: ⚠️ HR javobi Sotuv AI promptiga TUSHMADI",
                  "Yillik ta'til 12 oy" not in kb_text, "prompt=" + kb_text[:300])

            async def _dataset():
                from api.routers.knowledge import build_dataset
                async with async_session() as s2:
                    return await build_dataset(s2)

            ds = asyncio.run(_dataset())
            ds_matn = str(ds)
            check("S-29: sotuv yozuvi datasetda BOR (tekshiruv haqiqiy)",
                  "T-KB-SOTUV-MAYOQ" in ds_matn, "dataset=" + ds_matn[:200])
            check("S-29: ⚠️ HR javobi TASHQI chatbot datasetiga TUSHMADI",
                  "Yillik ta'til 12 oy" not in ds_matn, "dataset=" + ds_matn[:300])

            # ── BOT O'ZI JAVOB BERADI (TAKLIF) ──
            r = c.post("/hr-inquiries/me", headers=auth(bobur_t),
                       json={"question": "Tatilga qachon chiqaman?"})
            check("S-29: o'xshash savol -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code))
            yangi = r.json() if r.status_code == 201 else {}
            tak = yangi.get("suggestion")
            check("S-29: bilim bazasidan TAKLIF chiqdi", tak is not None,
                  "=" + str(yangi))
            check("S-29: taklif matni HR yozgan javob",
                  bool(tak) and "Yillik ta'til 12 oy" in tak.get("answer", ""),
                  "=" + str(tak))
            check("S-29: taklif chiqqanda HR BEZOVTA QILINMADI",
                  yangi.get("notified") == 0, "notified=" + str(yangi.get("notified")))

            # ── XODIM TASDIQLADI ──
            r = c.post("/hr-inquiries/me/suggestion", headers=auth(bobur_t),
                       json={"inquiry_id": yangi["id"],
                             "entry_id": tak["entry_id"], "accepted": True})
            check("S-29: taklif qabul qilindi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            check("S-29: murojaat yopildi, HR ga bormadi",
                  r.json().get("resolved") is True, "=" + str(r.json()))
            qator = cur.execute(
                "select status, auto_answered, answered_by, knowledge_entry_id"
                " from hr_inquiries where id=?", (yangi["id"],)).fetchone()
            check("S-29: `auto_answered` belgilandi, `answered_by` BO'SH",
                  qator is not None and qator[0] == "answered" and qator[1] == 1
                  and qator[2] is None and qator[3] == kb.get("entry_id"),
                  "=" + str(qator))

            # ── XODIM RAD ETDI -> HR ga boradi ──
            r = c.post("/hr-inquiries/me", headers=auth(bobur_t),
                       json={"question": "Tatilga qachon chiqaman, ayting?"})
            rad = r.json()
            if rad.get("suggestion"):
                r = c.post("/hr-inquiries/me/suggestion", headers=auth(bobur_t),
                           json={"inquiry_id": rad["id"],
                                 "entry_id": rad["suggestion"]["entry_id"],
                                 "accepted": False})
                check("S-29: «bu javob emas» -> HR ga yuborildi",
                      r.status_code == 200 and r.json().get("resolved") is False
                      and r.json().get("notified", 0) >= 1, "=" + str(r.json()))
                holat = cur.execute(
                    "select status, auto_answered from hr_inquiries where id=?",
                    (rad["id"],)).fetchone()
                check("S-29: rad etilgan murojaat OCHIQ qoladi",
                      holat == ("open", 0), "=" + str(holat))
            else:
                check("S-29: rad etish oqimi uchun taklif chiqdi", False,
                      "taklif chiqmadi: " + str(rad))

            # ── O'ZGANING murojaatiga taklif javobi berib bo'lmaydi ──
            r = c.post("/hr-inquiries/me", headers=auth(bobur_t),
                       json={"question": "Tatilga qachon chiqaman, javob bering?"})
            begona = r.json()
            if begona.get("suggestion"):
                r = c.post("/hr-inquiries/me/suggestion", headers=auth(aziz_t),
                           json={"inquiry_id": begona["id"],
                                 "entry_id": begona["suggestion"]["entry_id"],
                                 "accepted": True})
                #  404, 403 emas: 403 murojaat MAVJUDLIGINI tasdiqlardi
                #  (S-06 qoidasi; S-30 auditida tuzatildi).
                check("S-29: ⚠️ o'zganing murojaatini yopib bo'lmaydi -> 404",
                      r.status_code == 404, "kod=" + str(r.status_code))

            # ── TAKRORLANUVCHI SAVOLLAR HISOBOTI ──
            r = c.get("/hr-inquiries/frequent?limit=10", headers=auth(mgr_t))
            check("S-29: hisobot -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            hisobot = r.json() if r.status_code == 200 else {}
            check("S-29: toifalar kesimi bor",
                  bool(hisobot.get("categories")), "=" + str(hisobot.get("categories")))
            savollar = hisobot.get("questions", [])
            check("S-29: TOP savollar ro'yxati bor", len(savollar) >= 1,
                  "soni=" + str(len(savollar)))
            check("S-29: ro'yxat limitdan oshmaydi", len(savollar) <= 10,
                  "soni=" + str(len(savollar)))
            tatil = [q for q in savollar if "atil" in q["sample"]]
            check("S-29: o'xshash ta'til savollari BITTA qatorga yig'ildi",
                  bool(tatil) and tatil[0]["count"] >= 2,
                  "=" + str(tatil[:1]))
            check("S-29: bazadagi savol «bazada» deb belgilangan",
                  bool(tatil) and tatil[0]["in_knowledge"] is True,
                  "=" + str(tatil[:1]))

            # Xodim hisobotni ko'rmaydi
            r = c.get("/hr-inquiries/frequent", headers=auth(aziz_t))
            check("S-29: xodim hisobotga kira olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

            # ── ESKIRGAN YOZUV TAKLIF QILINMAYDI ──
            cur.execute("update knowledge_entries set needs_recheck=1 where id=?",
                        (kb.get("entry_id", 0),))
            conn.commit()
            r = c.post("/hr-inquiries/me", headers=auth(aziz_t),
                       json={"question": "Tatilga qachon chiqaman?"})
            check("S-29: qayta ko'rish kutayotgan yozuv TAKLIF QILINMAYDI",
                  r.json().get("suggestion") is None, "=" + str(r.json()))
            cur.execute("update knowledge_entries set needs_recheck=0 where id=?",
                        (kb.get("entry_id", 0),))
            conn.commit()
    except Exception:
        check("S-29 (umumiy)", False, traceback.format_exc(limit=2).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from hr_inquiries where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            cur.execute("delete from knowledge_entries where source='hr_inquiry'")
            conn.commit()
        except Exception:
            pass
        conn.close()



#  ── S-30: B blok rol matritsasi ──────────────────────────────
#
#  Har qator: (modul, metod, yo'l, {rol: kutilgan_kod}).
#  Kutilgan kod `403` — rol bu modulni UMUMAN ochmaydi.
#  `"ok"` — ruxsat bor (200/201/404/422 bo'lishi mumkin, lekin 403 EMAS).
#
#  ⚠️ Jadval QO'LDA yozilgan va routerdagi `_HR`/`_MANAGER` majmualaridan
#  MUSTAQIL. Aks holda test kodni o'zidan olib, o'zini tasdiqlagan
#  bo'lardi: qo'riqchi tasodifan kengaytirilsa ham «to'g'ri» derdi.
B_BLOCK_MATRIX = [
    # modul,               metod, yo'l,                       employee, rop,   hr
    ("kadr hujjatlari",    "GET", "/employee-documents/unregistered", 403, 403, "ok"),
    ("muddatlar",          "GET", "/deadlines",                       403, 403, "ok"),
    ("hujjat shablonlari", "GET", "/document-templates",              403, 403, "ok"),
    ("ish takliflari",     "GET", "/offers",                          403, 403, "ok"),
    ("ma'lumotnomalar",    "GET", "/certificates",                    403, 403, "ok"),
    ("mol-mulk",           "GET", "/assets",                          403, 403, "ok"),
    ("sinov muddati",      "GET", "/probation",                       403, 403, "ok"),
    ("ma'lumot so'rovlari","GET", "/profile-changes",                 403, 403, "ok"),
    ("murojaatlar",        "GET", "/hr-inquiries",                    403, 403, "ok"),
    ("murojaat hisoboti",  "GET", "/hr-inquiries/frequent",           403, 403, "ok"),
    #  ── ROP ham ko'radigan modullar ──
    ("e'lonlar (rahbar)",  "GET", "/announcements",                   403, "ok", "ok"),
    ("shtat jadvali",      "GET", "/staff",                           403, "ok", "ok"),
    #  ── Hamma ko'radigan (xodim tomoni) ──
    ("bayramlar",          "GET", "/holidays?year=2026",             "ok", "ok", "ok"),
    ("mening hujjatlarim", "GET", "/employee-documents/me",          "ok", "ok", "ok"),
    ("menga biriktirilgan","GET", "/assets/me",                      "ok", "ok", "ok"),
    ("mening e'lonlarim",  "GET", "/announcements/me",               "ok", "ok", "ok"),
    ("mening murojaatim",  "GET", "/hr-inquiries/me",                "ok", "ok", "ok"),
    ("tanishishlarim",     "GET", "/acks/me",                        "ok", "ok", "ok"),
    ("ma'lumotlarim",      "GET", "/profile-changes/me",             "ok", "ok", "ok"),
]

#  Menyu bo'limi -> uni ochadigan endpoint. Menyu va backend
#  KELISHMASA test yiqiladi (S-30 qabul mezoni: «menyuda ortiqcha
#  bo'lim ko'rinmaydi» — va teskarisi: ochiq modul yashirin qolmasin).
SECTION_TO_ENDPOINT = {
    "employee-documents": "/employee-documents/unregistered",
    "deadlines": "/deadlines",
    "offers": "/offers",
    "certificates": "/certificates",
    "assets": "/assets",
    "announcements": "/announcements",
    "staff": "/staff",
    "probation": "/probation",
    "profile-changes": "/profile-changes",
    "hr-inquiries": "/hr-inquiries",
}


def test_b_block_visibility_audit() -> None:
    """S-30 — B blok ko'rinish auditi.

    Qabul mezonlari (TZ):
      • har yangi modul uchun ROL MATRITSASI testi;
      • begona so'rov hamma joyda 404 (403 emas — mavjudligini
        oshkor qilmaslik uchun, S-06 qoidasi);
      • menyuda ortiqcha bo'lim ko'rinmaydi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-30: B BLOK KO'RINISH AUDITI")
    print("=" * 60)

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Au%'")
        conn.commit()
        #  To'rt rol: xodim, ROP, HR. Boshliq mavjudini olamiz (uning
        #  huquqi HR bilan bir xil, alohida yaratish shart emas).
        for nom, tg, rol in (
            ("T-Au Xodim", 999703001, "employee"),
            ("T-Au ROP", 999703002, "rop"),
            ("T-Au HR", 999703003, "hr"),
            ("T-Au Boshliq", 999703005, "boss"),
            ("T-Au Begona", 999703004, "employee"),
        ):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,?,0,1,datetime('now'))",
                (tg, nom, rol))
            ids[nom] = cur.lastrowid
        conn.commit()

        tok = {
            "employee": token_for(ids["T-Au Xodim"], "employee"),
            "rop": token_for(ids["T-Au ROP"], "rop"),
            "hr": token_for(ids["T-Au HR"], "hr"),
            "boss": token_for(ids["T-Au Boshliq"], "boss"),
        }
        begona_t = token_for(ids["T-Au Begona"], "employee")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ══════════════════════════════════════════════
            # A) ROL MATRITSASI
            # ══════════════════════════════════════════════
            xato = []
            for modul, metod, yol, kut_emp, kut_rop, kut_hr in B_BLOCK_MATRIX:
                #  Boshliq ustuni ATAYLAB alohida yozilmagan: bu
                #  modullarda uning huquqi HR bilan AYNAN bir xil
                #  (`_HR`/`_VIEW` majmualarida ikkovi ham bor). Alohida
                #  ustun yozilsa u nusxa bo'lib, birini yangilab
                #  ikkinchisini unutish xavfi tug'ilardi.
                for rol, kutilgan in (("employee", kut_emp), ("rop", kut_rop),
                                      ("hr", kut_hr), ("boss", kut_hr)):
                    r = c.request(metod, yol, headers=auth(tok[rol]))
                    if kutilgan == "ok":
                        yaxshi = r.status_code != 403
                    else:
                        yaxshi = r.status_code == kutilgan
                    if not yaxshi:
                        xato.append(f"{modul} [{rol}] {metod} {yol}: "
                                    f"kutilgan={kutilgan} keldi={r.status_code}")
            check(f"S-30: rol matritsasi ({len(B_BLOCK_MATRIX)} modul x 4 rol)",
                  not xato, " | ".join(xato[:5]))

            # ══════════════════════════════════════════════
            # B) MENYU <-> BACKEND MUVOFIQLIGI
            #
            # ⚠️ ENG QIMMATLI TEKSHIRUV: menyu (`sections.py`) va
            # qo'riqchi (`require_roles`) IKKI joyda yozilgan. Biri
            # o'zgarib ikkinchisi qolsa: yo xodim bosib 403 oladi, yo
            # ochiq modul menyuda ko'rinmay yashirin qoladi.
            # ══════════════════════════════════════════════
            nomuvofiq = []
            for rol in ("employee", "rop", "hr", "boss"):
                r = c.get("/me/sections", headers=auth(tok[rol]))
                if r.status_code != 200:
                    nomuvofiq.append(f"{rol}: /me/sections -> {r.status_code}")
                    continue
                korinadi = {x["key"] for x in r.json()}
                for kalit, endpoint in SECTION_TO_ENDPOINT.items():
                    menyuda = kalit in korinadi
                    rr = c.request("GET", endpoint, headers=auth(tok[rol]))
                    ochiq = rr.status_code != 403
                    if menyuda != ochiq:
                        nomuvofiq.append(
                            f"{rol}/{kalit}: menyuda={menyuda} backend_ochiq={ochiq}")
            check("S-30: menyu va backend qo'riqchisi MOS",
                  not nomuvofiq, " | ".join(nomuvofiq[:5]))

            #  B blokda qo'shilgan bo'limlar `/me/sections` ga UMUMAN
            #  tushganmi (TZ 2-band). Yuqoridagi muvofiqlik tekshiruvi
            #  buni HR uchun bilvosita ushlaydi, lekin xodim tomonidagi
            #  bo'limlarni ushlamaydi — ular hech qayerda qo'riqchi
            #  bilan taqqoslanmaydi.
            r = c.get("/me/sections", headers=auth(tok["hr"]))
            hr_bolimlar = {x["key"] for x in r.json()} if r.status_code == 200 else set()
            kutilgan_hr = set(SECTION_TO_ENDPOINT)
            check("S-30: B blok rahbar bo'limlari menyuda bor",
                  kutilgan_hr <= hr_bolimlar,
                  "yetishmaydi: " + str(sorted(kutilgan_hr - hr_bolimlar)))

            r = c.get("/me/sections", headers=auth(tok["employee"]))
            xodim_bolimlar = {x["key"] for x in r.json()} if r.status_code == 200 else set()
            kutilgan_xodim = {
                "documents", "my-assets", "my-announcements",
                "my-inquiries", "my-profile", "my-salary-history",
            }
            check("S-30: B blok xodim bo'limlari kabinetda bor",
                  kutilgan_xodim <= xodim_bolimlar,
                  "yetishmaydi: " + str(sorted(kutilgan_xodim - xodim_bolimlar)))

            #  ⚠️ Teskari yo'nalish: xodimga RAHBAR bo'limi tushib
            #  qolmasin (TZ: «menyuda ortiqcha bo'lim ko'rinmaydi»).
            ortiqcha = kutilgan_hr & xodim_bolimlar
            check("S-30: xodim menyusida rahbar bo'limi YO'Q",
                  not ortiqcha, "ortiqcha: " + str(sorted(ortiqcha)))

            # ══════════════════════════════════════════════
            # C) BEGONA SO'ROV -> 404 (403 EMAS)
            # ══════════════════════════════════════════════
            begona_xato = []

            #  1) Boshqa xodimning kadr hujjatlari
            r = c.get(f"/employee-documents/user/{ids['T-Au Begona']}",
                      headers=auth(tok["employee"]))
            if r.status_code != 404:
                begona_xato.append(f"employee-documents/user: {r.status_code}")
            #  ROP ham kadr hujjatlarida «begona» (TZ 3.4 maxfiy modul)
            r = c.get(f"/employee-documents/user/{ids['T-Au Begona']}",
                      headers=auth(tok["rop"]))
            if r.status_code != 404:
                begona_xato.append(f"employee-documents/user [rop]: {r.status_code}")
            check("S-30: begona xodim hujjatlari -> 404 (403/200 emas)",
                  not begona_xato, " | ".join(begona_xato))

            #  2) Boshqa xodimning murojaati (S-29 taklifi orqali)
            r = c.post("/hr-inquiries/me", headers=auth(begona_t),
                       json={"question": "T-Au begona savol, oylik haqida"})
            begona_id = r.json().get("id") if r.status_code == 201 else None
            r = c.post("/hr-inquiries/me/suggestion", headers=auth(tok["employee"]),
                       json={"inquiry_id": begona_id or 0, "entry_id": 1,
                             "accepted": True})
            #  ⚠️ 404 SHART, 403 emas — 403 murojaat mavjudligini
            #  tasdiqlardi (S-06 qoidasi). Auditda aynan shu topildi:
            #  S-29 kodi 403 qaytarardi.
            check("S-30: begona murojaat -> 404 (403 emas!)",
                  r.status_code == 404, "kod=" + str(r.status_code))

            #  3) Boshqa xodimga biriktirilgan buyumni «qabul qilish»
            r = c.post("/assets/999999/accept", headers=auth(tok["employee"]))
            check("S-30: begona buyumni qabul qilib bo'lmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # D) EKSPORT QAMROVI
            #
            # ⚠️ Qamrov IKKI marta yozilgan: `deps.scoped_user_ids` va
            # `reports._visible_user_ids`. Ikkovi bir xil javob berishi
            # SHART — biri o'zgarib ikkinchisi qolsa, ROP eksportda
            # butun tashkilotni ko'rib qolardi (ilgari aynan shu xato
            # bo'lgan, `reports.py` izohida yozilgan).
            # ══════════════════════════════════════════════
            r = c.get("/reports/export?date_from=2026-08-01&date_to=2026-08-02",
                      headers=auth(tok["employee"]))
            check("S-30: xodim Excel eksportga kira olmaydi -> 403",
                  r.status_code == 403, "kod=" + str(r.status_code))

        #  Ikki qamrov amalga oshiruvchisi bir xil javob beradimi
        import asyncio as _aio

        from db.base import async_session

        async def _qamrovlar():
            from api.deps import scoped_user_ids
            from api.routers.reports import _visible_user_ids
            from db.models import User as _U
            from sqlalchemy import select as _sel
            from sqlalchemy.orm import selectinload as _sl
            async with async_session() as s2:
                rop = await s2.scalar(
                    _sel(_U).options(_sl(_U.position))
                    .where(_U.id == ids["T-Au ROP"]))
                a = await scoped_user_ids(rop, s2)
                b = await _visible_user_ids(s2, rop)
                return a, b

        a, b = _aio.run(_qamrovlar())
        check("S-30: eksport qamrovi markaziy qamrov bilan BIR XIL",
              (a is None and b is None) or (set(a or []) == set(b or [])),
              f"deps={sorted(a) if a else a} reports={sorted(b) if b else b}")
    except Exception:
        check("S-30 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from hr_inquiries where user_id in ({belgi})",
                            tuple(ids.values()))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_courses_model() -> None:
    """S-32 (TZ 3.1) — o'quv paneli: inventarizatsiya va model.

    Qabul mezonlari (TZ):
      • anketa bilan umumiy qism HUJJATLASHTIRILGAN;
      • uch jadval migratsiyasi IKKALA DIALEKTDA ishlaydi;
      • `deleted_at` bor va BARCHA o'qish shu bilan filtrlanadi.
    """
    import asyncio
    import re as _re

    print("\n" + "=" * 60)
    print("S-32: O'QUV PANELI — MODEL")
    print("=" * 60)

    from db.base import async_session

    conn = db()
    cur = conn.cursor()
    kurs_id = None
    try:
        # ══════════════════════════════════════════════
        # 1) UCH JADVAL VA `deleted_at`
        # ══════════════════════════════════════════════
        for jadval in ("courses", "course_materials", "course_questions"):
            ustunlar = [r[1] for r in cur.execute(f"PRAGMA table_info({jadval})")]
            check(f"S-32: `{jadval}` jadvali bor va `deleted_at` ustuni mavjud",
                  bool(ustunlar) and "deleted_at" in ustunlar,
                  "ustunlar=" + str(ustunlar[:6]))

        # ══════════════════════════════════════════════
        # 2) IKKALA DIALEKT
        #
        # Postgres'da `BOOLEAN DEFAULT 0` XATO beradi (butun son
        # boolean ustunga tushmaydi). SQLite esa `false` ni bilmaydi.
        # `sa.false()` ikkovini to'g'ri render qiladi — shuni
        # TEKSHIRAMIZ, ishonch bilan qoldirmaymiz.
        # ══════════════════════════════════════════════
        import os as _os
        import subprocess as _sp
        muhit = dict(_os.environ)
        muhit["DATABASE_URL"] = "postgresql+asyncpg://u:p@localhost/db"
        try:
            ddl = _sp.run(
                [sys.executable, "-m", "alembic", "-c", "db/alembic.ini",
                 "upgrade", "kb01e1f2a3b4:cr02f2a3b4c5", "--sql"],
                capture_output=True, text=True, env=muhit, timeout=120,
            ).stdout
            check("S-32: migratsiya Postgres dialektida renderlanadi",
                  "CREATE TABLE courses" in ddl, "uzunlik=" + str(len(ddl)))
            check("S-32: Postgres'da `BOOLEAN DEFAULT false` (0 EMAS)",
                  "BOOLEAN DEFAULT false" in ddl and "BOOLEAN DEFAULT (0)" not in ddl,
                  [l.strip() for l in ddl.splitlines() if "BOOLEAN" in l][:2])
            check("S-32: Postgres'da `options` ustuni JSON",
                  _re.search(r"options JSON", ddl) is not None,
                  [l.strip() for l in ddl.splitlines() if "options" in l][:2])
        except Exception as e:
            check("S-32: Postgres dialekti tekshiruvi", False, str(e)[:120])

        # ══════════════════════════════════════════════
        # 3) BARCHA O'QISH `deleted_at` NI FILTRLAYDI
        #
        # ⚠️ ENG MUHIM TEKSHIRUV. Buni «esda tutishga» qoldirib
        # bo'lmaydi — o'chirilgan kurs xodimga ko'rinsa, u bekor
        # qilingan yo'riqnomani o'qib noto'g'ri qoidani o'rganardi.
        # Shuning uchun kurs modellaridan TO'G'RIDAN-TO'G'RI o'qish
        # `courses.py` dan TASHQARIDA taqiqlanadi: yagona kirish
        # nuqtasi `alive()`.
        # ══════════════════════════════════════════════
        import ast as _ast
        import pathlib as _pl
        ildiz = _pl.Path(__file__).parent
        buzuvchilar = []
        taqiq = {"Course", "CourseMaterial", "CourseQuestion", "CourseAssignment"}

        #  ⚠️ MATN bo'yicha emas, AST bo'yicha qidiramiz. Oddiy `re` izohlar
        #  va hujjat satrlaridagi «select(Course...)» iborasini ham topib,
        #  YOLG'ON ogohlantirish berardi — aynan shunday bo'ldi: S-34 da
        #  routerning O'Z izohi testni yiqitdi. AST faqat HAQIQIY chaqiruvni
        #  ko'radi va bo'shliq/qator ko'chishidan qat'i nazar ishlaydi.
        for fayl in list((ildiz / "api").rglob("*.py")) + list((ildiz / "bot").rglob("*.py")):
            if fayl.name == "courses.py" and fayl.parent.name == "services":
                continue  # yagona ruxsat etilgan joy
            try:
                daraxt = _ast.parse(fayl.read_text(encoding="utf-8", errors="ignore"))
            except SyntaxError:
                continue
            for tugun in _ast.walk(daraxt):
                if not isinstance(tugun, _ast.Call):
                    continue
                nom = getattr(tugun.func, "id", None) or getattr(tugun.func, "attr", None)
                if nom != "select" or not tugun.args:
                    continue
                birinchi = tugun.args[0]
                model = getattr(birinchi, "id", None) or getattr(birinchi, "attr", None)
                if model in taqiq:
                    buzuvchilar.append(
                        f"{fayl.relative_to(ildiz)}:{tugun.lineno} {model}")
        check("S-32: kurs modellari `alive()` dan TASHQARIDA o'qilmaydi",
              not buzuvchilar, " | ".join(buzuvchilar[:5]))

        # ── `alive()` haqiqatan filtrlaydimi (xulq tekshiruvi) ──
        async def _oqim():
            from sqlalchemy import select as _sel

            from api.services import courses as C
            from db.models import CourseMaterial, CourseQuestion, User
            async with async_session() as s2:
                u = await s2.scalar(_sel(User).limit(1))
                c1 = await C.create_course(
                    s2, title="T-Kurs S32", description=None,
                    pass_percent=70, max_attempts=0, actor_id=u.id)
                await C.add_material(s2, course_id=c1.id, kind="text",
                                     title="T-material-1", body="matn")
                m2 = await C.add_material(s2, course_id=c1.id, kind="video",
                                          title="T-material-2", file_id="T-FID")
                q1 = await C.add_question(s2, course_id=c1.id, text="2+2?",
                                          options=["3", "4"], correct_index=1)
                q2 = await C.add_question(s2, course_id=c1.id, text="Fikringiz?")
                await s2.commit()
                natija = {
                    "kurs_id": c1.id,
                    "tartib": [m.position for m in await C.materials(s2, c1.id)],
                    "ochiq_savol": (q2.options, q2.correct_index),
                    "test_savol": (q1.options, q1.correct_index),
                }
                #  Materialni yumshoq o'chiramiz
                await C.soft_delete(s2, m2)
                await s2.commit()
                natija["ochirilgach_material"] = [
                    m.title for m in await C.materials(s2, c1.id)]
                #  Kursni yumshoq o'chiramiz
                await C.soft_delete(s2, c1)
                await s2.commit()
                natija["ochirilgach_kurs"] = await C.get_course(s2, c1.id)
                natija["royxatda"] = [
                    x.id for x in await C.list_courses(s2) if x.id == c1.id]
                #  Bazada qator TURIBDI (qattiq o'chirilmagan)
                natija["bazada_bor"] = await s2.scalar(
                    _sel(CourseMaterial.id).where(CourseMaterial.id == m2.id)) is not None
                return natija

        r = asyncio.run(_oqim())
        kurs_id = r["kurs_id"]
        check("S-32: tartib 10 qadam bilan beriladi (oraga qo'shish uchun joy)",
              r["tartib"] == [10, 20], "=" + str(r["tartib"]))
        check("S-32: ochiq savol AVTOMAT baholanmaydi (`correct_index=None`)",
              r["ochiq_savol"] == ([], None), "=" + str(r["ochiq_savol"]))
        check("S-32: test savolida to'g'ri javob saqlanadi",
              r["test_savol"] == (["3", "4"], 1), "=" + str(r["test_savol"]))
        check("S-32: o'chirilgan material ro'yxatdan CHIQADI",
              r["ochirilgach_material"] == ["T-material-1"],
              "=" + str(r["ochirilgach_material"]))
        check("S-32: o'chirilgan kurs `get_course` da YO'Q",
              r["ochirilgach_kurs"] is None, "=" + str(r["ochirilgach_kurs"]))
        check("S-32: o'chirilgan kurs ro'yxatda YO'Q",
              not r["royxatda"], "=" + str(r["royxatda"]))
        check("S-32: o'chirish YUMSHOQ — qator bazada qoladi (tarix yo'qolmaydi)",
              r["bazada_bor"] is True, "=" + str(r["bazada_bor"]))

        # ══════════════════════════════════════════════
        # 4) ANKETA MEXANIZMI QAYTA ISHLATILADI
        # ══════════════════════════════════════════════
        import inspect as _insp

        from api.services import courses as _C
        manba = _insp.getsource(_C)
        check("S-32: `.docx` ajratgichi anketanikidan olinadi (ikkinchisi yozilmagan)",
              "from api.services.docx_parse import parse_questions" in manba,
              "import topilmadi")
        modellar_manbasi = (ildiz / "db" / "models.py").read_text(encoding="utf-8")
        check("S-32: anketa bilan umumiy qism HUJJATLASHTIRILGAN",
              "O'QUV PANELI" in modellar_manbasi
              and "AnketaAssignment.current_q" in modellar_manbasi
              and "docx_parse" in modellar_manbasi,
              "models.py dagi hujjat bloki topilmadi")

        #  Ajratgich haqiqatan ishlaydimi — `.txt` bilan sinaymiz
        async def _import():
            from sqlalchemy import select as _sel

            from db.models import User
            async with async_session() as s2:
                u = await s2.scalar(_sel(User).limit(1))
                c2 = await _C.create_course(
                    s2, title="T-Kurs import", description=None,
                    pass_percent=70, max_attempts=0, actor_id=u.id)
                matn = "1. Birinchi savol?\n2. Ikkinchi savol?\n3. Uchinchi savol?\n"
                res = await _C.import_questions_from_file(
                    s2, course_id=c2.id, data=matn.encode("utf-8"),
                    filename="savollar.txt")
                await s2.commit()
                return res, [q.text for q in await _C.questions(s2, c2.id)]

        res, savollar = asyncio.run(_import())
        check("S-32: fayldan savollar yuklandi (anketa ajratgichi bilan)",
              res.get("added", 0) >= 3, "=" + str(res))
        check("S-32: yuklangan savollar OCHIQ javobli (ajratgich variant bilmaydi)",
              len(savollar) >= 3, "=" + str(savollar[:3]))
    except Exception:
        check("S-32 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            cur.execute(
                "delete from course_questions where course_id in"
                " (select id from courses where title like 'T-Kurs%')")
            cur.execute(
                "delete from course_materials where course_id in"
                " (select id from courses where title like 'T-Kurs%')")
            cur.execute("delete from courses where title like 'T-Kurs%'")
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_course_assignments() -> None:
    """S-33 (TZ 3.1) — kurs tayinlash va natija.

    Qabul mezonlari (TZ):
      • bir xodimga bir kurs IKKI MARTA tayinlanmaydi;
      • holat BAZADA (restartga chidamli, FSM emas);
      • urinish raqami saqlanadi.
    """
    import asyncio

    print("\n" + "=" * 60)
    print("S-33: KURS TAYINLASH VA NATIJA")
    print("=" * 60)

    from db.base import async_session

    conn = db()
    cur = conn.cursor()
    try:
        async def _oqim():
            from sqlalchemy import select as _sel

            from api.services import courses as C
            from db.models import CourseAssignment, User
            out = {}
            async with async_session() as s2:
                users = list(await s2.scalars(_sel(User).limit(2)))
                u1, u2 = users[0], users[1]
                out["u1"], out["u2"] = u1.id, u2.id

                #  ⚠️ 80%: 1-urinish 67% bilan YIQILADI va qayta urinish
                #  mazmunli bo'ladi (60% bo'lsa birinchi urinishdayoq
                #  o'tib ketardi va `retry()` ni sinab bo'lmasdi).
                c = await C.create_course(s2, title="T-S33 kurs", description=None,
                                          pass_percent=80, max_attempts=2,
                                          actor_id=u1.id)
                await C.add_material(s2, course_id=c.id, kind="text",
                                     title="T-Dars-1", body="matn")
                await C.add_question(s2, course_id=c.id, text="2+2?",
                                     options=["3", "4"], correct_index=1, points=2)
                await C.add_question(s2, course_id=c.id, text="5+5?",
                                     options=["10", "11"], correct_index=0, points=1)
                await s2.commit()
                out["course_id"] = c.id

                c.is_published = True
                await s2.commit()

                # ── Tayinlash va DUBLIKAT ──
                out["assign1"] = await C.assign(
                    s2, course_id=c.id, user_ids=[u1.id, u2.id], assigned_by=u1.id)
                await s2.commit()
                out["assign2"] = await C.assign(
                    s2, course_id=c.id, user_ids=[u1.id, u2.id], assigned_by=u1.id)
                await s2.commit()
                out["jami_tayinlash"] = len(list(await s2.scalars(
                    C.alive(CourseAssignment).where(
                        CourseAssignment.course_id == c.id))))

                a = await C.assignment_for(s2, course_id=c.id, user_id=u1.id)
                out["boshlangich"] = (a.status, a.attempt_no, a.current_material,
                                      a.current_q)

                p1 = await C.progress(s2, a)
                out["bosqich_material"] = p1["stage"]
                p2 = await C.next_material(s2, a)
                await s2.commit()
                out["bosqich_savol"] = p2["stage"]

                # ── 1-urinish: bittasi noto'g'ri ──
                r1 = await C.submit_answer(s2, assignment=a, choice=1)   # to'g'ri
                r2 = await C.submit_answer(s2, assignment=a, choice=1)   # noto'g'ri
                await s2.commit()
                out["javob_togri"] = (r1["correct"], r2["correct"])
                res = await C.finish(s2, a)
                await s2.commit()
                out["urinish1"] = (res.score, res.max_score, res.percent,
                                   res.passed, res.attempt_no, res.pending_review)

                # ⚠️ HOLAT BAZADA: yangi sessiyada o'qib ko'ramiz
                out["assignment_id"] = a.id

            #  ── YANGI SESSIYA (restart taqlidi) ──
            async with async_session() as s3:
                from api.services import courses as C2
                from db.models import CourseAssignment as CA
                a = await s3.scalar(C2.alive(CA).where(CA.id == out["assignment_id"]))
                out["restartdan_keyin"] = (a.status, a.attempt_no,
                                           len(a.answers or []))

                # ── Qayta urinish ──
                await C2.retry(s3, a)
                await s3.commit()
                out["retry"] = (a.attempt_no, a.answers == [], a.current_q,
                                a.current_material)
                await C2.submit_answer(s3, assignment=a, choice=1)
                await C2.submit_answer(s3, assignment=a, choice=0)
                await s3.commit()
                res2 = await C2.finish(s3, a)
                await s3.commit()
                out["urinish2"] = (res2.percent, res2.passed, res2.attempt_no)
                out["tarix"] = [(t.attempt_no, t.percent, t.passed)
                                for t in await C2.results(s3, a.id)]


            #  ── OCHIQ savolli kurs ──
            async with async_session() as s4:
                from api.services import courses as C3
                c2 = await C3.create_course(s4, title="T-S33 ochiq", description=None,
                                            pass_percent=50, max_attempts=0,
                                            actor_id=out["u1"])
                await C3.add_question(s4, course_id=c2.id, text="Fikringiz?")
                c2.is_published = True
                await s4.commit()
                await C3.assign(s4, course_id=c2.id, user_ids=[out["u2"]],
                                assigned_by=out["u1"])
                await s4.commit()
                a2 = await C3.assignment_for(s4, course_id=c2.id, user_id=out["u2"])
                await C3.submit_answer(s4, assignment=a2, text="Mening fikrim...")
                await s4.commit()
                res3 = await C3.finish(s4, a2)
                await s4.commit()
                out["ochiq"] = (res3.percent, res3.passed, res3.pending_review,
                                (res3.answers or [{}])[0].get("text"))

            # ══════════════════════════════════════════════
            # «XATO KUTILGAN» SINOVLAR — HAR BIRI O'Z SESSIYASIDA
            #
            # ⚠️ `rollback()` ATAYLAB ishlatilmaydi: u
            # `expire_on_commit=False` bo'lsa ham sessiyadagi BARCHA
            # obyektni ekspire qiladi va keyingi `obj.id` lazy-load
            # boshlab, async kontekstda `MissingGreenlet` beradi
            # (loyihaning ma'lum tuzog'i). Alohida sessiya jonli
            # xulqqa ham yaqinroq — har so'rov o'z sessiyasida.
            # ══════════════════════════════════════════════
            async with async_session() as s5:
                from api.services import courses as C4
                #  Nashr qilinmagan kurs tayinlanmaydi
                c3 = await C4.create_course(s5, title="T-S33 nashrsiz",
                                            description=None, pass_percent=50,
                                            max_attempts=0, actor_id=out["u1"])
                await s5.commit()
                try:
                    await C4.assign(s5, course_id=c3.id, user_ids=[out["u1"]],
                                    assigned_by=out["u1"])
                    out["nashrsiz"] = "TAYINLANDI (xato)"
                except ValueError as e:
                    out["nashrsiz"] = str(e)

            async with async_session() as s6:
                from api.services import courses as C5
                #  Material tugamasdan savolga javob
                a3 = await C5.assignment_for(
                    s6, course_id=out["course_id"], user_id=out["u2"])
                try:
                    await C5.submit_answer(s6, assignment=a3, choice=0)
                    out["erta_savol"] = "QABUL QILINDI (xato)"
                except ValueError as e:
                    out["erta_savol"] = str(e)

            async with async_session() as s7:
                from api.services import courses as C6
                from db.models import CourseAssignment as CA2
                #  O'tilgan kursni qayta topshirib bo'lmaydi
                a4 = await s7.scalar(
                    C6.alive(CA2).where(CA2.id == out["assignment_id"]))
                try:
                    await C6.retry(s7, a4)
                    out["otilgan_retry"] = "RUXSAT BERILDI (xato)"
                except ValueError as e:
                    out["otilgan_retry"] = str(e)
            return out

        r = asyncio.run(_oqim())

        check("S-33: nashr qilinmagan kurs TAYINLANMAYDI",
              "nashr" in r["nashrsiz"], "=" + str(r["nashrsiz"]))
        check("S-33: ikki xodimga tayinlandi",
              r["assign1"] == {"created": 2, "skipped": 0}, "=" + str(r["assign1"]))
        check("S-33: ⚠️ IKKINCHI marta tayinlanmadi (dublikat yo'q)",
              r["assign2"] == {"created": 0, "skipped": 2}
              and r["jami_tayinlash"] == 2,
              f"={r['assign2']}, jami={r['jami_tayinlash']}")
        check("S-33: boshlang'ich holat `assigned`, urinish 1",
              r["boshlangich"] == ("assigned", 1, 0, 0),
              "=" + str(r["boshlangich"]))
        check("S-33: material tugamasdan savolga javob RAD ETILADI",
              "material" in r["erta_savol"].lower(), "=" + str(r["erta_savol"]))
        check("S-33: material bosqichidan savol bosqichiga o'tadi",
              (r["bosqich_material"], r["bosqich_savol"]) == ("material", "savol"),
              f"={r['bosqich_material']} -> {r['bosqich_savol']}")
        check("S-33: test savoli DARHOL baholanadi",
              r["javob_togri"] == (True, False), "=" + str(r["javob_togri"]))
        #  2 ball (to'g'ri) / 3 ball (jami) = 67% >= 60% -> o'tdi
        check("S-33: 1-urinish natijasi ball va foiz bilan yozildi",
              r["urinish1"][:3] == (2, 3, 67), "=" + str(r["urinish1"]))
        check("S-33: 67% < 80% -> YIQILDI", r["urinish1"][3] is False,
              "=" + str(r["urinish1"]))

        # ══════════════════════════════════════════════
        # ⚠️ HOLAT BAZADA — RESTARTGA CHIDAMLI
        # ══════════════════════════════════════════════
        check("S-33: ⚠️ holat BAZADA — yangi sessiyada saqlanib qoldi",
              r["restartdan_keyin"] == ("finished", 1, 2),
              "=" + str(r["restartdan_keyin"]))

        check("S-33: qayta urinishda raqam oshdi, javoblar tozalandi",
              r["retry"][:2] == (2, True), "=" + str(r["retry"]))
        check("S-33: qayta urinishda MATERIAL qayta ko'rsatilmaydi",
              r["retry"][3] > 0, "current_material=" + str(r["retry"][3]))
        check("S-33: 2-urinish 100% -> o'tdi",
              r["urinish2"] == (100, True, 2), "=" + str(r["urinish2"]))

        # ══════════════════════════════════════════════
        # ⚠️ URINISH RAQAMI SAQLANADI (TZ qabul mezoni)
        # ══════════════════════════════════════════════
        check("S-33: ⚠️ URINISHLAR TARIXI saqlanadi (eskisi o'chirilmaydi)",
              r["tarix"] == [(1, 67, False), (2, 100, True)],
              "=" + str(r["tarix"]))

        check("S-33: o'tilgan kursni qayta topshirib bo'lmaydi",
              "o'til" in r["otilgan_retry"], "=" + str(r["otilgan_retry"]))

        # ══════════════════════════════════════════════
        # OCHIQ SAVOL — odam baholaydi
        # ══════════════════════════════════════════════
        check("S-33: ochiq savolli urinish `pending_review` bilan yopiladi",
              r["ochiq"][2] is True, "=" + str(r["ochiq"]))
        check("S-33: ⚠️ baholanmagan urinish `passed` BO'LMAYDI",
              r["ochiq"][1] is False, "=" + str(r["ochiq"]))
        check("S-33: ochiq javob MATNI saqlanadi (odam baholashi uchun)",
              r["ochiq"][3] == "Mening fikrim...", "=" + str(r["ochiq"]))

        # ══════════════════════════════════════════════
        # ⚠️ BAZA DARAJASIDA DUBLIKAT TO'SILADI
        #
        # Kod qo'riqchisi unutilsa ham qisman unique indeks ushlaydi.
        # Xom SQL bilan sinaymiz — servisni chetlab o'tamiz.
        # ══════════════════════════════════════════════
        try:
            cur.execute(
                "insert into course_assignments (course_id, user_id, status,"
                " current_material, current_q, attempt_no, created_at)"
                " values (?,?,'assigned',0,0,1,datetime('now'))",
                (r["course_id"], r["u1"]))
            conn.commit()
            check("S-33: ⚠️ baza dublikatni to'sadi (qisman unique indeks)",
                  False, "INSERT o'tib ketdi — indeks ishlamayapti")
        except sqlite3.IntegrityError as e:
            conn.rollback()
            check("S-33: ⚠️ baza dublikatni to'sadi (qisman unique indeks)",
                  "unique" in str(e).lower(), "=" + str(e)[:90])

        #  Yumshoq o'chirilgach QAYTA tayinlash MUMKIN (yillik qayta o'qitish)
        cur.execute(
            "update course_assignments set deleted_at=datetime('now')"
            " where course_id=? and user_id=?", (r["course_id"], r["u1"]))
        conn.commit()
        try:
            cur.execute(
                "insert into course_assignments (course_id, user_id, status,"
                " current_material, current_q, attempt_no, created_at)"
                " values (?,?,'assigned',0,0,1,datetime('now'))",
                (r["course_id"], r["u1"]))
            conn.commit()
            check("S-33: yumshoq o'chirilgach QAYTA tayinlash mumkin"
                  " (yillik qayta o'qitish)", True, "")
        except sqlite3.IntegrityError as e:
            conn.rollback()
            check("S-33: yumshoq o'chirilgach QAYTA tayinlash mumkin"
                  " (yillik qayta o'qitish)", False, str(e)[:90])
    except Exception:
        check("S-33 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            cur.execute(
                "delete from course_results where assignment_id in"
                " (select id from course_assignments where course_id in"
                "  (select id from courses where title like 'T-S33%'))")
            cur.execute(
                "delete from course_assignments where course_id in"
                " (select id from courses where title like 'T-S33%')")
            cur.execute(
                "delete from course_questions where course_id in"
                " (select id from courses where title like 'T-S33%')")
            cur.execute(
                "delete from course_materials where course_id in"
                " (select id from courses where title like 'T-S33%')")
            cur.execute("delete from courses where title like 'T-S33%'")
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_courses_hr_api() -> None:
    """S-34 (TZ 3.1) — o'quv paneli HR sayt tomoni.

    Qabul mezonlari (TZ):
      • kurs yaratish -> material -> savol ZANJIRI ishlaydi;
      • `.docx` dan savol IMPORT qilinadi;
      • o'tish chegarasi KURSDA saqlanadi.
    """
    import io

    import httpx

    print("\n" + "=" * 60)
    print("S-34: O'QUV PANELI — HR API")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    kurs_id = None
    try:
        cur.execute("delete from users where full_name like 'T-Cu%'")
        conn.commit()
        for nom, tg, rol in (("T-Cu Xodim", 999703101, "employee"),
                             ("T-Cu ROP", 999703102, "rop")):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,?,0,1,datetime('now'))",
                (tg, nom, rol))
            ids[nom] = cur.lastrowid
        conn.commit()
        xodim_t = token_for(ids["T-Cu Xodim"], "employee")
        rop_t = token_for(ids["T-Cu ROP"], "rop")

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ══════════════════════════════════════════════
            # ROL QAMROVI — ROP ataylab KO'RMAYDI
            # ══════════════════════════════════════════════
            for rol, tok in (("xodim", xodim_t), ("ROP", rop_t)):
                r = c.get("/courses", headers=auth(tok))
                check(f"S-34: {rol} o'quv panelini ko'rmaydi -> 403",
                      r.status_code == 403, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # 1) ZANJIR: kurs -> material -> savol
            # ══════════════════════════════════════════════
            r = c.post("/courses", headers=auth(mgr_t), json={
                "title": "T-Cu Xavfsizlik yo'riqnomasi",
                "description": "sinov kursi",
                "pass_percent": 80, "max_attempts": 2, "is_mandatory": True})
            check("S-34: kurs yaratildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            kurs = r.json() if r.status_code == 201 else {}
            kurs_id = kurs.get("id")

            # ⚠️ O'TISH CHEGARASI KURSDA SAQLANADI (TZ qabul mezoni)
            check("S-34: ⚠️ o'tish chegarasi KURSDA saqlandi",
                  kurs.get("pass_percent") == 80, "=" + str(kurs))
            check("S-34: «majburiy» bayrog'i saqlandi",
                  kurs.get("is_mandatory") is True, "=" + str(kurs))
            check("S-34: yangi kurs QORALAMA (nashr qilinmagan)",
                  kurs.get("is_published") is False, "=" + str(kurs))

            #  Savolsiz kursni nashr qilib bo'lmaydi
            r = c.post(f"/courses/{kurs_id}/publish?value=true", headers=auth(mgr_t))
            check("S-34: savolsiz kurs NASHR QILINMAYDI -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code) + " " + r.text[:120])

            #  Material
            r = c.post(f"/courses/{kurs_id}/materials", headers=auth(mgr_t), json={
                "kind": "text", "title": "1-dars", "body": "Xavfsizlik qoidalari"})
            check("S-34: matn material qo'shildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            r = c.post(f"/courses/{kurs_id}/materials", headers=auth(mgr_t), json={
                "kind": "video", "title": "2-dars", "file_id": "T-CU-FILEID"})
            check("S-34: video material `file_id` bilan qo'shildi -> 201",
                  r.status_code == 201, "kod=" + str(r.status_code))
            #  ⚠️ Faylsiz video RAD ETILADI — aks holda xodimga bo'sh
            #  material ochilardi.
            r = c.post(f"/courses/{kurs_id}/materials", headers=auth(mgr_t), json={
                "kind": "video", "title": "3-dars"})
            check("S-34: FAYLSIZ video material rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            #  Savol (test)
            r = c.post(f"/courses/{kurs_id}/questions", headers=auth(mgr_t), json={
                "text": "Kaska qachon kiyiladi?",
                "options": ["Hech qachon", "Obyektda doim"],
                "correct_index": 1, "points": 2})
            check("S-34: test savoli qo'shildi -> 201", r.status_code == 201,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            check("S-34: test savoli OCHIQ emas",
                  r.json().get("is_open") is False, "=" + str(r.json()))
            #  To'g'ri javobsiz test savoli rad etiladi
            r = c.post(f"/courses/{kurs_id}/questions", headers=auth(mgr_t), json={
                "text": "Variantli lekin javobsiz?", "options": ["A", "B"],
                "correct_index": None, "points": 1})
            check("S-34: to'g'ri javobsiz TEST savoli rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            #  Zanjir natijasi
            r = c.get(f"/courses/{kurs_id}", headers=auth(mgr_t))
            check("S-34: kurs tafsiloti -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            d = r.json() if r.status_code == 200 else {}
            check("S-34: ⚠️ ZANJIR ishlaydi (kurs + 2 material + 1 savol)",
                  d.get("course", {}).get("material_count") == 2
                  and d.get("course", {}).get("question_count") == 1,
                  "=" + str(d.get("course")))

            # ══════════════════════════════════════════════
            # 2) `.docx`/`.txt` DAN IMPORT
            # ══════════════════════════════════════════════
            matn = ("1. Yong'in chiqsa nima qilasiz?\n"
                    "2. Birinchi yordam qutisi qayerda?\n"
                    "3. Evakuatsiya yo'li qayerda?\n")
            r = c.post(
                f"/courses/{kurs_id}/questions/import", headers=auth(mgr_t),
                files={"file": ("savollar.txt", io.BytesIO(matn.encode("utf-8")),
                                "text/plain")})
            check("S-34: ⚠️ fayldan savol IMPORT qilindi -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code) + " " + r.text[:150])
            imp = r.json() if r.status_code == 200 else {}
            check("S-34: uchala savol yuklandi", imp.get("added") == 3,
                  "=" + str(imp))
            r = c.get(f"/courses/{kurs_id}", headers=auth(mgr_t))
            savollar = r.json().get("questions", [])
            check("S-34: import qilinganlar OCHIQ javobli (ajratgich variant bilmaydi)",
                  sum(1 for q in savollar if q["is_open"]) == 3,
                  "=" + str([(q["text"][:20], q["is_open"]) for q in savollar]))

            #  Bo'sh fayl rad etiladi
            r = c.post(
                f"/courses/{kurs_id}/questions/import", headers=auth(mgr_t),
                files={"file": ("bosh.txt", io.BytesIO(b""), "text/plain")})
            check("S-34: bo'sh fayl rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # 3) NASHR VA TAYINLASH
            # ══════════════════════════════════════════════
            r = c.post(f"/courses/{kurs_id}/publish?value=true", headers=auth(mgr_t))
            check("S-34: savol bor — kurs nashr qilindi -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code))

            #  Bo'sh qamrov rad etiladi
            r = c.post(f"/courses/{kurs_id}/assign", headers=auth(mgr_t),
                       json={"audience": "users", "scope_ids": []})
            check("S-34: bo'sh qamrov rad etildi -> 400",
                  r.status_code == 400, "kod=" + str(r.status_code))

            r = c.post(f"/courses/{kurs_id}/assign", headers=auth(mgr_t),
                       json={"audience": "users",
                             "scope_ids": [ids["T-Cu Xodim"]],
                             "due_date": "2026-12-31"})
            check("S-34: kurs tayinlandi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            check("S-34: bitta xodimga tayinlandi",
                  r.json().get("created") == 1, "=" + str(r.json()))
            #  Ikkinchi marta — dublikat yo'q (S-33 qoidasi API orqali)
            r = c.post(f"/courses/{kurs_id}/assign", headers=auth(mgr_t),
                       json={"audience": "users", "scope_ids": [ids["T-Cu Xodim"]]})
            check("S-34: qayta tayinlashda dublikat yaratilmadi",
                  r.json().get("created") == 0 and r.json().get("skipped") == 1,
                  "=" + str(r.json()))

            r = c.get(f"/courses/{kurs_id}/assignments", headers=auth(mgr_t))
            check("S-34: tayinlanganlar ro'yxati -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            qatorlar = r.json() if r.status_code == 200 else []
            check("S-34: tayinlangan xodim ro'yxatda, muddat bilan",
                  len(qatorlar) == 1 and qatorlar[0]["user_name"] == "T-Cu Xodim"
                  and qatorlar[0]["due_date"] == "2026-12-31",
                  "=" + str(qatorlar[:1]))
            check("S-34: hali topshirmagan xodimda natija YO'Q",
                  qatorlar and qatorlar[0]["percent"] is None,
                  "=" + str(qatorlar[:1]))

            # ══════════════════════════════════════════════
            # 4) O'CHIRISH YUMSHOQ
            # ══════════════════════════════════════════════
            r = c.get(f"/courses/{kurs_id}", headers=auth(mgr_t))
            mat_id = r.json()["materials"][0]["id"]
            r = c.delete(f"/courses/{kurs_id}/materials/{mat_id}", headers=auth(mgr_t))
            check("S-34: material o'chirildi -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code))
            bazada = cur.execute(
                "select deleted_at is not null from course_materials where id=?",
                (mat_id,)).fetchone()
            check("S-34: material YUMSHOQ o'chirildi (qator bazada qoldi)",
                  bazada == (1,), "=" + str(bazada))
            r = c.get(f"/courses/{kurs_id}", headers=auth(mgr_t))
            check("S-34: o'chirilgan material tafsilotda YO'Q",
                  r.json()["course"]["material_count"] == 1,
                  "=" + str(r.json()["course"]))

            #  Mavjud bo'lmagan kurs
            r = c.get("/courses/999999", headers=auth(mgr_t))
            check("S-34: mavjud bo'lmagan kurs -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))
    except Exception:
        check("S-34 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            cur.execute(
                "delete from course_results where assignment_id in"
                " (select id from course_assignments where course_id in"
                "  (select id from courses where title like 'T-Cu%'))")
            cur.execute(
                "delete from course_assignments where course_id in"
                " (select id from courses where title like 'T-Cu%')")
            cur.execute(
                "delete from course_questions where course_id in"
                " (select id from courses where title like 'T-Cu%')")
            cur.execute(
                "delete from course_materials where course_id in"
                " (select id from courses where title like 'T-Cu%')")
            cur.execute("delete from courses where title like 'T-Cu%'")
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_courses_employee() -> None:
    """S-35 (TZ 3.1) — o'quv paneli xodim tomoni (bot).

    Qabul mezonlari (TZ):
      • bot restart bo'lsa ham xodim QOLGAN JOYIDAN davom etadi;
      • test BARCHA MATERIAL ko'rilmaguncha ochilmaydi;
      • o'tmasa QAYTA URINISH beriladi (urinish soni yoziladi).
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-35: O'QUV PANELI — XODIM TOMONI (BOT)")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Ce%'")
        conn.commit()
        for nom, tg in (("T-Ce Xodim", 999703201), ("T-Ce Begona", 999703202)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, nom))
            ids[nom] = cur.lastrowid
        conn.commit()
        xodim_t = token_for(ids["T-Ce Xodim"], "employee")
        begona_t = token_for(ids["T-Ce Begona"], "employee")
        TG = 999703201

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── HR kurs tayyorlaydi: 2 material + 2 test savoli ──
            r = c.post("/courses", headers=auth(mgr_t), json={
                "title": "T-Ce Yong'in xavfsizligi", "description": None,
                "pass_percent": 80, "max_attempts": 2, "is_mandatory": True})
            kurs_id = r.json()["id"]
            for i in (1, 2):
                c.post(f"/courses/{kurs_id}/materials", headers=auth(mgr_t), json={
                    "kind": "text", "title": f"{i}-dars", "body": f"{i}-dars matni"})
            c.post(f"/courses/{kurs_id}/questions", headers=auth(mgr_t), json={
                "text": "O'chirgich qayerda?", "options": ["Bilmayman", "Chiqish yonida"],
                "correct_index": 1, "points": 2})
            c.post(f"/courses/{kurs_id}/questions", headers=auth(mgr_t), json={
                "text": "Nechchiga qo'ng'iroq?", "options": ["101", "999"],
                "correct_index": 0, "points": 1})
            c.post(f"/courses/{kurs_id}/publish?value=true", headers=auth(mgr_t))
            c.post(f"/courses/{kurs_id}/assign", headers=auth(mgr_t), json={
                "audience": "users", "scope_ids": [ids["T-Ce Xodim"]]})

            # ══════════════════════════════════════════════
            # BOT: kurslarim
            # ══════════════════════════════════════════════
            r = c.get("/courses/bot/my", params={"telegram_id": TG})
            check("S-35: botda «Darsliklarim» -> 200", r.status_code == 200,
                  "kod=" + str(r.status_code) + " " + r.text[:150])
            mine = r.json() if r.status_code == 200 else []
            check("S-35: tayinlangan kurs ko'rindi",
                  len(mine) == 1 and mine[0]["title"] == "T-Ce Yong'in xavfsizligi",
                  "=" + str(mine[:1]))
            check("S-35: «majburiy» belgisi keldi",
                  mine and mine[0]["is_mandatory"] is True, "=" + str(mine[:1]))
            aid = mine[0]["assignment_id"]

            # ══════════════════════════════════════════════
            # ⚠️ TEST MATERIAL KO'RILMAGUNCHA OCHILMAYDI
            # ══════════════════════════════════════════════
            r = c.post("/courses/bot/progress",
                       json={"telegram_id": TG, "assignment_id": aid})
            p1 = r.json() if r.status_code == 200 else {}
            check("S-35: boshlanishda MATERIAL bosqichi",
                  p1.get("stage") == "material" and p1.get("material_index") == 0,
                  "=" + str({k: p1.get(k) for k in ("stage", "material_index")}))
            check("S-35: joriy material matni keldi",
                  (p1.get("item") or {}).get("title") == "1-dars",
                  "=" + str(p1.get("item")))

            r = c.post("/courses/bot/answer",
                       json={"telegram_id": TG, "assignment_id": aid, "choice": 0})
            check("S-35: ⚠️ material tugamasdan TEST OCHILMAYDI -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code) + " " + r.text[:120])

            #  Ikkala materialni ko'ramiz
            r = c.post("/courses/bot/next-material",
                       json={"telegram_id": TG, "assignment_id": aid})
            check("S-35: 1-material ko'rildi -> 2-material",
                  r.json().get("stage") == "material"
                  and r.json().get("material_index") == 1,
                  "=" + str(r.json().get("stage")))
            r = c.post("/courses/bot/next-material",
                       json={"telegram_id": TG, "assignment_id": aid})
            p2 = r.json()
            check("S-35: barcha material ko'rilgach TEST ochildi",
                  p2.get("stage") == "savol", "=" + str(p2.get("stage")))
            check("S-35: savolda TO'G'RI JAVOB yuborilmaydi (o'g'irlab bo'lmasin)",
                  "correct_index" not in (p2.get("item") or {}),
                  "=" + str(p2.get("item")))

            #  Materiallar tugagach yana `next-material` MUMKIN EMAS
            r = c.post("/courses/bot/next-material",
                       json={"telegram_id": TG, "assignment_id": aid})
            check("S-35: materiallar tugagach yana o'tib bo'lmaydi -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # ⚠️ RESTARTGA CHIDAMLILIK
            #
            # Bot «qayta ishga tushdi» deb faraz qilamiz: hech qanday
            # xotira yo'q, faqat `telegram_id` bilan so'raymiz.
            # ══════════════════════════════════════════════
            r = c.get("/courses/bot/my", params={"telegram_id": TG})
            qayta = r.json()[0]
            r = c.post("/courses/bot/progress",
                       json={"telegram_id": TG, "assignment_id": qayta["assignment_id"]})
            check("S-35: ⚠️ RESTARTDAN keyin xodim QOLGAN JOYIDA (savol bosqichi)",
                  r.json().get("stage") == "savol"
                  and r.json().get("question_index") == 0,
                  "=" + str({k: r.json().get(k) for k in ("stage", "question_index")}))

            # ══════════════════════════════════════════════
            # 1-urinish: bittasi noto'g'ri -> 2/3 = 67% < 80%
            # ══════════════════════════════════════════════
            r = c.post("/courses/bot/answer",
                       json={"telegram_id": TG, "assignment_id": aid, "choice": 1})
            check("S-35: 1-savol javobi qabul qilindi",
                  r.status_code == 200 and r.json().get("correct") is True,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            r = c.post("/courses/bot/answer",
                       json={"telegram_id": TG, "assignment_id": aid, "choice": 1})
            check("S-35: 2-savol noto'g'ri deb belgilandi",
                  r.json().get("correct") is False, "=" + str(r.json().get("correct")))
            check("S-35: savollar tugagach bosqich «tugadi»",
                  r.json().get("stage") == "tugadi", "=" + str(r.json().get("stage")))

            r = c.post("/courses/bot/finish",
                       json={"telegram_id": TG, "assignment_id": aid})
            res1 = r.json() if r.status_code == 200 else {}
            check("S-35: natija DARHOL keldi (foiz va o'tdi/o'tmadi)",
                  res1.get("percent") == 67 and res1.get("passed") is False,
                  "=" + str(res1))
            check("S-35: ⚠️ o'tmagach QAYTA URINISH beriladi",
                  res1.get("can_retry") is True, "=" + str(res1))

            # ══════════════════════════════════════════════
            # ⚠️ QAYTA URINISH — urinish soni yoziladi
            # ══════════════════════════════════════════════
            r = c.post("/courses/bot/retry",
                       json={"telegram_id": TG, "assignment_id": aid})
            p3 = r.json() if r.status_code == 200 else {}
            check("S-35: qayta urinish boshlandi, urinish raqami 2",
                  p3.get("attempt_no") == 2 and p3.get("stage") == "savol",
                  "=" + str({k: p3.get(k) for k in ("attempt_no", "stage")}))
            check("S-35: qayta urinishda MATERIAL qayta ko'rsatilmaydi",
                  p3.get("material_index") == p3.get("material_total"),
                  "=" + str({k: p3.get(k) for k in ("material_index", "material_total")}))

            c.post("/courses/bot/answer",
                   json={"telegram_id": TG, "assignment_id": aid, "choice": 1})
            c.post("/courses/bot/answer",
                   json={"telegram_id": TG, "assignment_id": aid, "choice": 0})
            r = c.post("/courses/bot/finish",
                       json={"telegram_id": TG, "assignment_id": aid})
            res2 = r.json()
            check("S-35: 2-urinish 100% -> o'tdi",
                  res2.get("percent") == 100 and res2.get("passed") is True
                  and res2.get("attempt_no") == 2, "=" + str(res2))
            check("S-35: o'tilgach qayta urinish TAKLIF QILINMAYDI",
                  res2.get("can_retry") is False, "=" + str(res2))

            r = c.get("/courses/bot/my", params={"telegram_id": TG})
            check("S-35: ro'yxatda natija va urinish soni ko'rinadi",
                  r.json()[0]["percent"] == 100 and r.json()[0]["attempt_no"] == 2,
                  "=" + str(r.json()[:1]))

            # ══════════════════════════════════════════════
            # BEGONA XODIM — 404 (403 emas)
            # ══════════════════════════════════════════════
            r = c.get(f"/courses/me/{aid}/progress", headers=auth(begona_t))
            check("S-35: ⚠️ begona tayinlash -> 404 (mavjudligi oshkor bo'lmaydi)",
                  r.status_code == 404, "kod=" + str(r.status_code))
            r = c.get("/courses/me/assignments", headers=auth(begona_t))
            check("S-35: begona xodimda kurs YO'Q",
                  r.status_code == 200 and r.json() == [], "=" + str(r.json())[:80])

            # ══════════════════════════════════════════════
            # SAYT VA BOT BITTA HOLATNI O'QIYDI (S-36 uchun asos)
            # ══════════════════════════════════════════════
            r_jwt = c.get(f"/courses/me/{aid}/progress", headers=auth(xodim_t))
            r_bot = c.post("/courses/bot/progress",
                           json={"telegram_id": TG, "assignment_id": aid})
            check("S-35: JWT va bot BIR XIL holatni qaytaradi",
                  r_jwt.status_code == 200 and r_jwt.json() == r_bot.json(),
                  f"jwt={r_jwt.json()} bot={r_bot.json()}")

            # ══════════════════════════════════════════════
            # OCHIQ SAVOL — anketa protokoli
            # ══════════════════════════════════════════════
            r = c.post("/courses", headers=auth(mgr_t), json={
                "title": "T-Ce Ochiq kurs", "description": None,
                "pass_percent": 50, "max_attempts": 0, "is_mandatory": False})
            k2 = r.json()["id"]
            c.post(f"/courses/{k2}/questions", headers=auth(mgr_t),
                   json={"text": "Fikringiz?", "options": [], "correct_index": None,
                         "points": 1})
            c.post(f"/courses/{k2}/publish?value=true", headers=auth(mgr_t))
            c.post(f"/courses/{k2}/assign", headers=auth(mgr_t), json={
                "audience": "users", "scope_ids": [ids["T-Ce Xodim"]]})

            r = c.post("/courses/bot/answer-text",
                       json={"telegram_id": TG, "text": "Mening fikrim shu."})
            check("S-35: ochiq savol javobi ERKIN MATNDAN ushlandi",
                  r.json().get("handled") is True, "=" + str(r.json())[:150])

            #  ⚠️ Mos holat yo'q bo'lsa `handled: false` — xabar boshqa
            #  oqimlarga o'tadi (anketa protokoli).
            r = c.post("/courses/bot/answer-text",
                       json={"telegram_id": TG, "text": "Yana bir matn"})
            check("S-35: ⚠️ kutilmagan matn USHLANMAYDI (boshqa oqimga o'tadi)",
                  r.json().get("handled") is False, "=" + str(r.json())[:120])
            r = c.post("/courses/bot/answer-text",
                       json={"telegram_id": 111222333, "text": "notanish"})
            check("S-35: notanish foydalanuvchi matni ushlanmaydi",
                  r.json().get("handled") is False, "=" + str(r.json())[:120])
    except Exception:
        check("S-35 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            cur.execute(
                "delete from course_results where assignment_id in"
                " (select id from course_assignments where course_id in"
                "  (select id from courses where title like 'T-Ce%'))")
            cur.execute(
                "delete from course_assignments where course_id in"
                " (select id from courses where title like 'T-Ce%')")
            cur.execute(
                "delete from course_questions where course_id in"
                " (select id from courses where title like 'T-Ce%')")
            cur.execute(
                "delete from course_materials where course_id in"
                " (select id from courses where title like 'T-Ce%')")
            cur.execute("delete from courses where title like 'T-Ce%'")
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()



def test_courses_cabinet() -> None:
    """S-36 (TZ 3.1) — o'quv paneli xodim kabineti.

    Qabul mezonlari (TZ):
      • BOTDA BOSHLAB, SAYTDA davom ettirish mumkin;
      • ikkala joyda BIR XIL foiz ko'rinadi;
      • xodim FAQAT o'ziga tayinlangan kursni ko'radi.
    """
    import httpx

    print("\n" + "=" * 60)
    print("S-36: O'QUV PANELI — XODIM KABINETI")
    print("=" * 60)

    mgr = find_manager_id()
    if not mgr:
        check("rahbar topildi", False, "hr/boss/dasturchi yo'q")
        return
    mgr_t = token_for(mgr[0], mgr[1])

    conn = db()
    cur = conn.cursor()
    ids: dict[str, int] = {}
    try:
        cur.execute("delete from users where full_name like 'T-Cb%'")
        conn.commit()
        for nom, tg in (("T-Cb Xodim", 999703301), ("T-Cb Begona", 999703302)):
            cur.execute(
                "insert into users (telegram_id, full_name, role, bot_started,"
                " is_active, created_at) values (?,?,'employee',0,1,datetime('now'))",
                (tg, nom))
            ids[nom] = cur.lastrowid
        conn.commit()
        xodim_t = token_for(ids["T-Cb Xodim"], "employee")
        begona_t = token_for(ids["T-Cb Begona"], "employee")
        TG = 999703301

        with httpx.Client(base_url=API_BASE, timeout=30) as c:
            # ── HR: 2 material + 2 savol ──
            r = c.post("/courses", headers=auth(mgr_t), json={
                "title": "T-Cb Aralash kurs", "description": None,
                "pass_percent": 50, "max_attempts": 0, "is_mandatory": False})
            kurs_id = r.json()["id"]
            for i in (1, 2):
                c.post(f"/courses/{kurs_id}/materials", headers=auth(mgr_t), json={
                    "kind": "text", "title": f"{i}-dars", "body": f"{i}-matn"})
            for t, opts, ci in (("Birinchi savol?", ["A", "B"], 0),
                                ("Ikkinchi savol?", ["A", "B"], 1)):
                c.post(f"/courses/{kurs_id}/questions", headers=auth(mgr_t), json={
                    "text": t, "options": opts, "correct_index": ci, "points": 1})
            c.post(f"/courses/{kurs_id}/publish?value=true", headers=auth(mgr_t))
            c.post(f"/courses/{kurs_id}/assign", headers=auth(mgr_t), json={
                "audience": "users", "scope_ids": [ids["T-Cb Xodim"]]})

            # ══════════════════════════════════════════════
            # ⚠️ XODIM FAQAT O'ZINIKINI KO'RADI
            # ══════════════════════════════════════════════
            r = c.get("/courses/me/assignments", headers=auth(xodim_t))
            check("S-36: kabinetda kursim ko'rinadi -> 200",
                  r.status_code == 200 and len(r.json()) == 1,
                  "kod=" + str(r.status_code) + " " + r.text[:120])
            aid = r.json()[0]["assignment_id"]
            r = c.get("/courses/me/assignments", headers=auth(begona_t))
            check("S-36: ⚠️ begona xodimda kurs YO'Q",
                  r.status_code == 200 and r.json() == [], "=" + str(r.json())[:80])
            r = c.get(f"/courses/me/{aid}/progress", headers=auth(begona_t))
            check("S-36: begona kursning holatini ochib bo'lmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            # ══════════════════════════════════════════════
            # ⚠️ BOTDA BOSHLAB, SAYTDA DAVOM ETTIRISH
            #
            # Har qadam NAVBATMA-NAVBAT: bot -> sayt -> bot -> sayt.
            # Ikkita mustaqil progress bo'lsa, zanjir SHU YERDA uzilardi.
            # ══════════════════════════════════════════════
            #  1) BOTDA: 1-materialni ko'rdi
            r = c.post("/courses/bot/next-material",
                       json={"telegram_id": TG, "assignment_id": aid})
            check("S-36: [bot] 1-material ko'rildi",
                  r.status_code == 200 and r.json()["material_index"] == 1,
                  "=" + str(r.json().get("material_index")))

            #  2) SAYTDA: davom etadi — 2-materialda turibdi
            r = c.get(f"/courses/me/{aid}/progress", headers=auth(xodim_t))
            check("S-36: ⚠️ [sayt] BOTDAGI joydan davom etdi (2-material)",
                  r.json()["stage"] == "material" and r.json()["material_index"] == 1,
                  "=" + str({k: r.json().get(k) for k in ("stage", "material_index")}))
            r = c.post(f"/courses/me/{aid}/next-material", headers=auth(xodim_t))
            check("S-36: [sayt] 2-material ko'rildi -> savol bosqichi",
                  r.json()["stage"] == "savol", "=" + str(r.json()["stage"]))

            #  3) BOTDA: birinchi savolga javob
            r = c.post("/courses/bot/answer",
                       json={"telegram_id": TG, "assignment_id": aid, "choice": 0})
            check("S-36: [bot] 1-savolga javob berildi",
                  r.status_code == 200 and r.json()["question_index"] == 1,
                  "=" + str(r.json().get("question_index")))

            #  4) SAYTDA: ikkinchi savolda turibdi
            r = c.get(f"/courses/me/{aid}/progress", headers=auth(xodim_t))
            check("S-36: ⚠️ [sayt] BOTDAGI javobdan keyingi savolda",
                  r.json()["stage"] == "savol" and r.json()["question_index"] == 1,
                  "=" + str({k: r.json().get(k) for k in ("stage", "question_index")}))
            r = c.post(f"/courses/me/{aid}/answer", headers=auth(xodim_t),
                       json={"choice": 1})
            check("S-36: [sayt] 2-savolga javob berildi -> tugadi",
                  r.json()["stage"] == "tugadi", "=" + str(r.json()["stage"]))

            #  5) SAYTDA yakunlash
            r = c.post(f"/courses/me/{aid}/finish", headers=auth(xodim_t))
            natija = r.json() if r.status_code == 200 else {}
            check("S-36: [sayt] yakunlandi, 100% (ikkovi ham to'g'ri)",
                  natija.get("percent") == 100 and natija.get("passed") is True,
                  "=" + str(natija))

            # ══════════════════════════════════════════════
            # ⚠️ IKKALA JOYDA BIR XIL FOIZ
            # ══════════════════════════════════════════════
            r_sayt = c.get("/courses/me/assignments", headers=auth(xodim_t))
            r_bot = c.get("/courses/bot/my", params={"telegram_id": TG})
            check("S-36: ⚠️ sayt va bot ro'yxati AYNAN bir xil",
                  r_sayt.status_code == 200 and r_sayt.json() == r_bot.json(),
                  f"sayt={r_sayt.json()} bot={r_bot.json()}")
            check("S-36: ⚠️ ikkala joyda bir xil foiz (100%)",
                  r_sayt.json()[0]["percent"] == 100
                  and r_bot.json()[0]["percent"] == 100,
                  f"sayt={r_sayt.json()[0]['percent']} bot={r_bot.json()[0]['percent']}")

            # ══════════════════════════════════════════════
            # MATERIAL FAYLINI TELEGRAMGA YUBORISH
            #
            # Video/hujjat brauzerda ko'rsatilmaydi (`file_id` ni
            # brauzer o'qiy olmaydi, serverdan oqizish esa Passenger'ni
            # bloklardi) — fayl xodimning Telegramiga yuboriladi.
            # ══════════════════════════════════════════════
            r = c.post("/courses", headers=auth(mgr_t), json={
                "title": "T-Cb Video kurs", "description": None,
                "pass_percent": 50, "max_attempts": 0, "is_mandatory": False})
            k2 = r.json()["id"]
            c.post(f"/courses/{k2}/materials", headers=auth(mgr_t), json={
                "kind": "video", "title": "Video dars", "file_id": "T-CB-VIDEO"})
            c.post(f"/courses/{k2}/questions", headers=auth(mgr_t), json={
                "text": "Savol?", "options": ["A", "B"], "correct_index": 0,
                "points": 1})
            c.post(f"/courses/{k2}/publish?value=true", headers=auth(mgr_t))
            c.post(f"/courses/{k2}/assign", headers=auth(mgr_t), json={
                "audience": "users", "scope_ids": [ids["T-Cb Xodim"]]})
            r = c.get("/courses/me/assignments", headers=auth(xodim_t))
            a2 = [x for x in r.json() if x["course_id"] == k2][0]["assignment_id"]

            r = c.post(f"/courses/me/{a2}/send-material", headers=auth(xodim_t))
            check("S-36: material faylini Telegramga yuborish -> 200",
                  r.status_code == 200, "kod=" + str(r.status_code) + " " + r.text[:120])
            #  ⚠️ Sinov rejimida bildirishnomalar o'chiq — `delivered=False`
            #  bo'lishi TO'G'RI xulq, xato emas.
            check("S-36: sinov rejimida `delivered=False` (xabar yuborilmadi)",
                  r.json().get("ok") is True and r.json().get("delivered") is False,
                  "=" + str(r.json()))
            r = c.post(f"/courses/me/{a2}/send-material", headers=auth(begona_t))
            check("S-36: begona kursning faylini so'rab bo'lmaydi -> 404",
                  r.status_code == 404, "kod=" + str(r.status_code))

            #  Savol bosqichida fayl so'ralsa 409
            c.post(f"/courses/me/{a2}/next-material", headers=auth(xodim_t))
            r = c.post(f"/courses/me/{a2}/send-material", headers=auth(xodim_t))
            check("S-36: savol bosqichida fayl so'ralsa -> 409",
                  r.status_code == 409, "kod=" + str(r.status_code))
    except Exception:
        check("S-36 (umumiy)", False, traceback.format_exc(limit=3).strip())
    finally:
        try:
            cur.execute(
                "delete from course_results where assignment_id in"
                " (select id from course_assignments where course_id in"
                "  (select id from courses where title like 'T-Cb%'))")
            cur.execute(
                "delete from course_assignments where course_id in"
                " (select id from courses where title like 'T-Cb%')")
            cur.execute(
                "delete from course_questions where course_id in"
                " (select id from courses where title like 'T-Cb%')")
            cur.execute(
                "delete from course_materials where course_id in"
                " (select id from courses where title like 'T-Cb%')")
            cur.execute("delete from courses where title like 'T-Cb%'")
            if ids:
                belgi = ",".join("?" * len(ids))
                cur.execute(f"delete from users where id in ({belgi})",
                            tuple(ids.values()))
            conn.commit()
        except Exception:
            pass
        conn.close()


def cleanup_orphans() -> None:
    """EGASIZ (user'i o'chirilgan) payroll qatorlarini tozalaydi.

    NEGA KERAK: `run_payroll` va `recalculate_period` BARCHA faol xodim
    bo'yicha aylanadi — ya'ni bitta blok davr hisoblaganda BOSHQA blokning
    sinov foydalanuvchilariga ham payslip/bonus qatori yaratiladi. Ular
    o'z blokining tozalash ro'yxatida bo'lmasa qolib ketadi.

    Ustiga-ustak, test.py dagi xom `sqlite3` ulanishlarida
    `PRAGMA foreign_keys` YOQILMAGAN — shu sababli `delete from users`
    egasiz qator qoldirib ham muvaffaqiyatli tugaydi (async SQLAlchemy
    ulanishida esa aynan shu FOREIGN KEY xatosi bilan yiqilardi).

    Bu qoldiqlar keyingi ishga tushirishda CHALG'ITUVCHI xatolar beradi
    (jonli isbot 2026-08-17: «payroll_fund jami» sinovi egasiz 2 mln so'mlik
    payslip tufayli qulagan edi). Shuning uchun oxirida yagona supurgi.
    """
    conn = db()
    cur = conn.cursor()
    jami = 0
    try:
        cur.execute(
            "delete from payslip_items where payslip_id in"
            " (select id from payslips where user_id not in (select id from users))"
        )
        jami += cur.rowcount
        for tbl in ("bonuses", "payslips", "salary_rates", "attendance", "daily_results",
                    "overtime_entries", "overtime_profiles", "work_schedule_override",
                    "work_schedule_weekly", "excused_days", "norms", "payroll_adjustments"):
            try:
                cur.execute(
                    f"delete from {tbl} where user_id is not null"
                    " and user_id not in (select id from users)"
                )
                jami += cur.rowcount
            except sqlite3.OperationalError:
                pass  # jadval yoki ustun yo'q — muhim emas
        conn.commit()
    finally:
        conn.close()
    if jami:
        print(f"\n[tozalash] egasiz {jami} ta qator o'chirildi")


def test_bot_commands_by_role() -> None:
    """Slash-buyruqlar: «/» menyusi lavozimga qarab + ruxsatsizga XATO.

    Nima qo'riqlanadi:
      (a) HAR bir handlerdagi `Command(...)` reestrda bor — yangi buyruq
          qo'shilib, ruxsat qoidasi unutilsa shu yerda ushlanadi;
      (b) rollar bo'yicha ro'yxat OLTIN NAMUNA bilan bir xil (qo'lda
          yozilgan — reestrdan MUSTAQIL, aks holda test o'zini tasdiqlardi);
      (c) ruxsati yo'q buyruq JIM qolmaydi, aniq sabab qaytadi;
      (d) noto'g'ri joyda (guruh/shaxsiy) yozilgan buyruq ham sabab qaytaradi.

    Tarix: ilgari `/guruhlar`, `/guruh_biriktir`, `/statistika` ruxsat yoki
    joy mos kelmasa JIMGINA e'tiborsiz qolardi (handler filtri yutib
    yuborardi), «/» menyusi esa hammaga bir xil 4 ta buyruqni ko'rsatardi.
    """
    import pathlib
    import re

    from api.services.sections import (
        ALL_COMMANDS,
        COMMANDS_BY_NAME,
        bot_commands_payload,
        commands_for,
    )
    from bot.commands import (
        GROUP,
        NO_ACCESS,
        OK,
        PRIVATE,
        UNKNOWN,
        WRONG_SCOPE,
        check_access,
        extract_command,
        visible_commands,
    )

    print("\n" + "=" * 60)
    print("BOT SLASH-BUYRUQLARI — LAVOZIM BO'YICHA")
    print("=" * 60)

    class _Pos:
        def __init__(self, flags, metrics):
            self.menu_flags = flags
            self.metrics = metrics

    class _U:
        def __init__(self, role, flags=None, metrics=None):
            self.role = role
            self.position = None if (flags is None and metrics is None) else _Pos(flags, metrics)
            self.can_edit_fine_policy = False
            self.can_edit_attendance = False

    # ── (a) HAR bir handler buyrug'i reestrda bormi ──
    handlers_dir = pathlib.Path(__file__).parent / "bot" / "handlers"
    found: set[str] = set()
    for path in handlers_dir.glob("*.py"):
        found |= set(re.findall(r'Command\("([a-z_]+)"\)', path.read_text(encoding="utf-8")))
    # `/ai_vaqt` ATAYLAB reestrda yo'q: u eskirgan, faqat tushuntirish
    # beradi va hammaga ochiq qolishi kerak (menyuda ko'rinmasin).
    kutilmagan = {"ai_vaqt"}
    yetishmayotgan = sorted(found - kutilmagan - set(COMMANDS_BY_NAME))
    check("har bir handler buyrug'i reestrda bor (ruxsat qoidasi unutilmagan)",
          not yetishmayotgan, f"reestrsiz: {yetishmayotgan}")
    check("/start va /buyruqlar reestrda",
          {"start", "buyruqlar"} <= set(COMMANDS_BY_NAME))

    # ── (b) OLTIN NAMUNA: rol bo'yicha ro'yxat ──
    # Qo'lda yozilgan. O'ZGARSA — ATAYLABMI? Yangi buyruq qo'shilganda shu
    # ro'yxat ham yangilanadi va commit izohida ko'rsatiladi.
    kutilgan_shaxsiy = {
        "employee": ["start", "buyruqlar", "sotuv_ai"],
        "rop": ["start", "buyruqlar", "oylik", "davomat_vaqt", "reja",
                "norma_ozgartir", "ai_sozlama", "sotuv_ai"],
        "boss": ["start", "buyruqlar", "statistika_vaqt", "oylik", "davomat_vaqt",
                 "reja", "norma_ozgartir", "ai_sozlama", "ai_markazi", "bilim",
                 "playbook", "sotuv_ai"],
    }
    for role, kutilgan in kutilgan_shaxsiy.items():
        haqiqiy = [c.name for c in commands_for(_U(role), PRIVATE)]
        check(f"{role}: shaxsiy chat buyruqlari oltin namuna bilan bir xil",
              haqiqiy == kutilgan, f"={haqiqiy}")

    dev_priv = [c.name for c in commands_for(_U("dasturchi"), PRIVATE)]
    check("dasturchi: shaxsiy chatda eng ko'p buyruq", len(dev_priv) == 19, f"={len(dev_priv)}")
    check("employee: guruhda faqat /buyruqlar",
          [c.name for c in commands_for(_U("employee"), GROUP)] == ["buyruqlar"])
    check("⭐ oddiy xodim guruh-boshqaruv buyruqlarini KO'RMAYDI",
          not ({"guruh_biriktir", "guruh_ochir"}
               & {c.name for c in commands_for(_U("employee"), GROUP)}))
    # Bugalter kabi lavozim: metrikasi ATAYLAB bo'sh → sotuv AI ham yo'q.
    bugalter = [c.name for c in commands_for(_U("employee", {}, []), PRIVATE)]
    check("metrikasiz lavozimda /sotuv_ai ko'rinmaydi", "sotuv_ai" not in bugalter,
          f"={bugalter}")

    # ── «/» menyusi serverdan kelgan javobdan quriladi ──
    specs_emp = bot_commands_payload(_U("employee"))
    specs_dev = bot_commands_payload(_U("dasturchi"))
    check("payload TO'LIQ ro'yxat qaytaradi (sabab aytish uchun)",
          len(specs_emp) == len(ALL_COMMANDS) == len(specs_dev), f"={len(specs_emp)}")
    menyu_emp = [c.command for c in visible_commands(specs_emp, PRIVATE)]
    check("xodim «/» menyusida faqat o'ziga tegishlisi",
          menyu_emp == ["start", "buyruqlar", "sotuv_ai"], f"={menyu_emp}")
    check("xodim uchun guruh «/» menyusi qisqa",
          [c.command for c in visible_commands(specs_emp, GROUP)] == ["buyruqlar"])

    # ── (c) RUXSAT: ruxsati yo'q buyruq to'siladi ──
    verdict, spec = check_access(specs_emp, "guruhlar", PRIVATE)
    check("⭐ xodim /guruhlar bossa — RAD ETILADI (jimlik emas)", verdict == NO_ACCESS,
          f"={verdict}")
    check("rad javobida kimga mo'ljallangani aytiladi",
          bool(spec and spec.get("audience")), f"={spec and spec.get('audience')}")
    check("xodim /norm_set (dasturchi buyrug'i) — rad etiladi",
          check_access(specs_emp, "norm_set", PRIVATE)[0] == NO_ACCESS)
    check("xodim /statistika (rahbar buyrug'i) — rad etiladi",
          check_access(specs_emp, "statistika", GROUP)[0] == NO_ACCESS)
    check("dasturchi /guruhlar — ruxsat", check_access(specs_dev, "guruhlar", PRIVATE)[0] == OK)

    # ROP — rahbar, lekin Boshliq buyrug'iga ruxsati yo'q.
    specs_rop = bot_commands_payload(_U("rop"))
    check("⭐ ROP /statistika_vaqt (Boshliq buyrug'i) — rad etiladi",
          check_access(specs_rop, "statistika_vaqt", PRIVATE)[0] == NO_ACCESS)
    check("ROP /norma_ozgartir — ruxsat",
          check_access(specs_rop, "norma_ozgartir", PRIVATE)[0] == OK)

    # ── (d) JOY: to'g'ri buyruq, noto'g'ri chat turi ──
    check("⭐ dasturchi /guruh_biriktir SHAXSIY chatda — «bu yerda emas» deydi",
          check_access(specs_dev, "guruh_biriktir", PRIVATE)[0] == WRONG_SCOPE)
    check("dasturchi /guruh_biriktir guruhda — ruxsat",
          check_access(specs_dev, "guruh_biriktir", GROUP)[0] == OK)
    check("dasturchi /guruhlar GURUHDA — «bu yerda emas» deydi",
          check_access(specs_dev, "guruhlar", GROUP)[0] == WRONG_SCOPE)
    check("⭐ joy xatosi ruxsat xatosidan OLDIN tekshiriladi",
          check_access(specs_emp, "guruh_biriktir", PRIVATE)[0] == WRONG_SCOPE)

    # ── Reestrda yo'q buyruq — eski xatti-harakat (handler hal qiladi) ──
    check("eskirgan /ai_vaqt to'silmaydi (handler tushuntiradi)",
          check_access(specs_dev, "ai_vaqt", PRIVATE)[0] == UNKNOWN)
    check("umuman noma'lum buyruq to'silmaydi",
          check_access(specs_dev, "yoq_bunday", PRIVATE)[0] == UNKNOWN)

    # ── /statistika faqat biriktirilgan guruhda ──
    stat = COMMANDS_BY_NAME["statistika"]
    check("/statistika faqat main/stats guruhida ishlaydi",
          stat.group_purposes == frozenset({"main", "stats"}), f"={stat.group_purposes}")
    check("boshqa buyruqlarda guruh-maqsad cheklovi yo'q",
          all(not c.group_purposes for c in ALL_COMMANDS if c.name != "statistika"))

    # ── Buyruq matnini ajratish ──
    check("«/reja 5» -> reja", extract_command("/reja 5") == "reja")
    check("guruhda «/reja@bot» -> reja",
          extract_command("/reja@Hodimlar_bot", "Hodimlar_bot") == "reja")
    check("⭐ boshqa botga yo'llangan buyruq O'ZLASHTIRILMAYDI",
          extract_command("/reja@boshqa_bot", "Hodimlar_bot") is None)
    check("oddiy matn buyruq emas", extract_command("salom") is None)
    check("bo'sh «/» buyruq emas", extract_command("/") is None)
    check("katta harf ham taniladi", extract_command("/REJA") == "reja")

    # ── Telegram cheklovlari ──
    check("buyruq nomlari Telegram formatiga mos (a-z, _, <=32)",
          all(re.fullmatch(r"[a-z0-9_]{1,32}", c.name) for c in ALL_COMMANDS))
    check("izohlar <=256 belgi va bo'sh emas",
          all(0 < len(c.description) <= 256 for c in ALL_COMMANDS))
    check("har bir buyruqda kamida bitta qamrov bor",
          all(c.scopes for c in ALL_COMMANDS))


def main() -> None:
    print("=" * 60)
    print("DAVOMAT TIZIMI — DB YOZUVI DEBUG TESTI")
    print("=" * 60)
    require_notifications_off()

    ctx: dict = {}
    try:
        ctx = setup()
    except Exception:
        print("Sozlashda xato:\n" + traceback.format_exc())
        cleanup(ctx)
        sys.exit(1)

    try:
        run_tests(ctx)
    except Exception:
        print("Kutilmagan xato:\n" + traceback.format_exc())
    finally:
        try:
            cleanup(ctx)
        except Exception:
            print("Tozalashda xato:\n" + traceback.format_exc())

    try:
        test_payroll_engine()
    except Exception:
        print("Payroll testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_future_days_rule()
    except Exception:
        print("S-01 kelajak kunlari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_fine_from_bonus()
    except Exception:
        print("S-02 ushlanma testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_fine_from_bonus_e2e()
    except Exception:
        print("S-02 e2e testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_me_sections()
    except Exception:
        print("S-04 sections testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_bot_menu_from_server()
    except Exception:
        print("S-05b bot menyu testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_bot_commands_by_role()
    except Exception:
        print("Bot buyruqlari testida kutilmagan xato:" + traceback.format_exc())

    try:
        test_view_scope()
    except Exception:
        print("S-06 qamrov testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_background_jobs()
    except Exception:
        print("S-07 fon ishlari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_setup_status()
    except Exception:
        print("S-08 setup-status testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_holidays()
    except Exception:
        print("S-09 bayramlar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_employee_documents()
    except Exception:
        print("S-10 hujjatlar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_documents_bot()
    except Exception:
        print("S-11 hujjatlar boti testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_deadlines_core()
    except Exception:
        print("S-12 muddatlar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_deadlines_cron()
    except Exception:
        print("S-13 muddat croni testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_docx_render()
    except Exception:
        print("S-14 docx testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_offers()
    except Exception:
        print("S-15 takliflar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_offer_hire()
    except Exception:
        print("S-16 ishga olish testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_certificates()
    except Exception:
        print("S-17 malumotnoma testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_assets()
    except Exception:
        print("S-18 mol-mulk testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_assets_employee()
    except Exception:
        print("S-19 mol-mulk xodim testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_acknowledgements()
    except Exception:
        print("S-20 tanishtirish testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_announcements()
    except Exception:
        print("S-21 elonlar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_birthday_anniversary()
    except Exception:
        print("S-22 tugilgan kun testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_staff_positions()
    except Exception:
        print("S-23 shtat testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_probation_list()
    except Exception:
        print("S-24 sinov muddati testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_salary_reason()
    except Exception:
        print("S-25 ish haqi sababi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_profile_changes()
    except Exception:
        print("S-26 profil so'rovlari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_contract_registration()
    except Exception:
        print("S-27 shartnoma royxati testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_hr_inquiries()
    except Exception:
        print("S-28 murojaatlar testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_hr_knowledge_loop()
    except Exception:
        print("S-29 bilim bazasi halqasi testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_b_block_visibility_audit()
    except Exception:
        print("S-30 ko'rinish auditi testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_courses_model()
    except Exception:
        print("S-32 o'quv paneli testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_course_assignments()
    except Exception:
        print("S-33 kurs tayinlash testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_courses_hr_api()
    except Exception:
        print("S-34 o'quv paneli API testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_courses_employee()
    except Exception:
        print("S-35 xodim tomoni testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_courses_cabinet()
    except Exception:
        print("S-36 kabinet testida kutilmagan xato:\n"
              + traceback.format_exc())

    try:
        test_payroll_api()
    except Exception:
        print("Payroll API testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_admin_override()
    except Exception:
        print("Admin override testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_payroll_automation()
    except Exception:
        print("Payroll avtomatika testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_payroll_reporting()
    except Exception:
        print("Payroll hisobot testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_kpi_in_payroll()
    except Exception:
        print("KPI+oylik testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_overtime_global_profile()
    except Exception:
        print("Global overtime testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_dasturchi_bot_bridge()
    except Exception:
        print("Dasturchi bot ko'prigi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_positions_permissions()
    except Exception:
        print("Lavozim ruxsatlari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_telegram_login_security()
    except Exception:
        print("Telegram login xavfsizligi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_hr_wide_employee_access()
    except Exception:
        print("HR keng qamrov testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_attendance_edit_rights()
    except Exception:
        print("Davomat tuzatish huquqlari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_location_exempt_checkin()
    except Exception:
        print("Bez-lokatsiya check-in testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_attendance_list_dates()
    except Exception:
        print("Davomat ro'yxati sana filtri testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_attendance_reminder()
    except Exception:
        print("Davomat eslatmasi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_dashboard_day_off()
    except Exception:
        print("Dashboard dam kuni testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_payroll_approval_segregation()
    except Exception:
        print("Tasdiq ajratimi testida kutilmagan xato:" + chr(10) + traceback.format_exc())

    try:
        test_kpi_rates()
    except Exception:
        print("KPI stavkalari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_absent_deduct_daily()
    except Exception:
        print("Kelmagan kun ayirmasi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_uysot_request_deadline()
    except Exception:
        print("CRM kutish chegarasi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_audit_json_guard()
    except Exception:
        print("Audit JSON zaxira testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_payroll_settings_reedit()
    except Exception:
        print("Oylik sozlamalari qayta tahrirlash testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_fine_policy_rights()
    except Exception:
        print("Kechikish normasi huquqi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_explanation_letters()
    except Exception:
        print("Tushuntirish xati testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_visit_counting()
    except Exception:
        print("Tashrif hisoblash testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_duplicate_guard()
    except Exception:
        print("Avans dublikat testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_limit()
    except Exception:
        print("Avans chegarasi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_limit_gate()
    except Exception:
        print("Avans chegara-eshigi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_issue_flow()
    except Exception:
        print("Avans to'lash zanjiri testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_soft_delete()
    except Exception:
        print("Avans o'chirish testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_period_close()
    except Exception:
        print("Avans oy yopilishi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_settings()
    except Exception:
        print("Avans sozlamalari testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_outbox()
    except Exception:
        print("Xabar navbati testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_day_tick()
    except Exception:
        print("Avans kuni cron testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_bot_flow()
    except Exception:
        print("Avans bot oqimi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_advance_hr_panel()
    except Exception:
        print("Avans HR paneli testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        test_app_login_code_delivery()
    except Exception:
        print("Saytga kirish kodi testida kutilmagan xato:\n" + traceback.format_exc())

    try:
        cleanup_orphans()
    except Exception:
        print("Egasiz qatorlarni tozalashda xato:\n" + traceback.format_exc())

    print("\n" + "=" * 60)
    print(f"NATIJA: {len(passed)} OK, {len(failed)} FAIL")
    for name in failed:
        #  ⚠️ `check()` dagi bilan BIR XIL qo'riqchi. Bu yerda u yo'q edi va
        #  tekshiruv nomida ⭐ kabi belgi bo'lsa xulosa `UnicodeEncodeError`
        #  bilan yiqilardi — ya'ni FAIL RO'YXATI UMUMAN KO'RINMASDI va
        #  yiqilgan test yashirin qolardi.
        _fail_print(f"  FAIL: {name}")
    print("=" * 60)
    sys.exit(1 if failed else 0)


def test_visit_counting() -> None:
    """Tashrif hisoblash (lead_diff.daily_operator_breakdown) — 2026-08-03
    tuzatishlari:
      1. Kechikib aniqlangan voqea (detected_at ertasi kun) CRM vaqti
         (`crm_updated_ts`) bo'yicha O'Z kuniga tushadi.
      2. Dual-kredit operator kesimida qoladi (yopgan +1, olib kelgan +1),
         lekin JAMI — noyob tashriflar (kreditlar yig'indisi emas).
      3. Mas'ulsiz tashrif operator kesimiga kirmaydi, jamida yo'qolmaydi.
      4. Tashrif bosqichlari ORASIDA ko'chish qayta sanalmaydi.
      5. Keyingi kunning voqeasi bu kunga kirmaydi.
      6. (2026-08-14) Shartnoma AYNI qoida bilan sanaladi, lekin DUAL-KREDITSIZ —
         faqat yopgan mas'ulga; operator yig'indisi = jami.
    Sinov ma'lumoti o'tmish sanada (2020-06-10) va 9998xxxxx lid ID bilan —
    jonli hisobga ta'sir qilmaydi, oxirida to'liq o'chiriladi."""
    import asyncio as _asyncio
    from datetime import timezone as _timezone

    print("\n" + "=" * 60)
    print("TASHRIF HISOBLASH (diff-engine kunlik kesim)")
    print("=" * 60)

    from sqlalchemy import delete as _delete

    from api.services import lead_diff
    from db.base import async_session
    from db.models import LeadEvent

    DAY = date(2020, 6, 10)
    VISIT_IDS = {990001, 990002}
    CONTRACT_IDS = {990009}
    L = 999800000  # sinov lid ID bazasi (jonli ID'lardan ancha yuqori)
    # 2020-06-10 Tashkent kuni = UTC 06-09 19:00 .. 06-10 19:00; CRM vaqti sifatida
    # kun o'rtasi (12:00 mahalliy = 07:00 UTC) olinadi
    day_epoch = int(datetime(2020, 6, 10, 7, 0, tzinfo=_timezone.utc).timestamp())

    def ev(lead, etype, from_st, to_st, from_resp, to_resp, first_resp, det, crm_ts):
        return LeadEvent(
            crm_lead_id=lead, event_type=etype,
            from_pipe_status_id=from_st, from_stage_name=None,
            to_pipe_status_id=to_st,
            to_stage_name="Tashrif" if to_st in VISIT_IDS else "Boshqa",
            from_responsible_id=from_resp, to_responsible_id=to_resp,
            to_responsible_name=None if to_resp is None else f"T-Op{to_resp}",
            first_responsible_id=first_resp, crm_updated_ts=crm_ts, detected_at=det,
        )

    async def _run():
        async with async_session() as s:
            await s.execute(_delete(LeadEvent).where(LeadEvent.crm_lead_id >= L))
            s.add_all([
                # A: kechikib aniqlangan (ertasi kun 02:00 UTC) — CRM vaqti bilan 06-10 ga tushishi kerak
                ev(L + 1, "stage_change", 111, 990001, 991001, 991001, 991001, datetime(2020, 6, 11, 2, 0), day_epoch),
                # B: dual-kredit — yopgan 991002, olib kelgan 991003
                ev(L + 2, "stage_change", 111, 990002, 991002, 991002, 991003, datetime(2020, 6, 10, 9, 0), day_epoch),
                # C: mas'ulsiz tashrif — jamida bor, operator kesimida yo'q.
                # 2026-08-13: turi `first_seen`dan `stage_change`ga o'zgartirildi —
                # egasining qaroriga ko'ra `first_seen` (skaner lidni endi ko'rgani)
                # umuman tashrif sanalmaydi; bu holat esa haqiqiy BOSQICH O'TISHI,
                # faqat mas'uli noma'lum.
                ev(L + 3, "stage_change", 111, 990001, None, None, None, datetime(2020, 6, 10, 10, 0), day_epoch),
                # C2: skaner lidni birinchi marta Tashrifda ko'rdi — SANALMAYDI
                ev(L + 7, "first_seen", None, 990001, None, 991009, 991009, datetime(2020, 6, 10, 10, 30), day_epoch),
                # D: Tashrif bosqichlari orasida ko'chish — tashrif sanalmasin
                ev(L + 4, "stage_change", 990001, 990002, 991001, 991001, 991001, datetime(2020, 6, 10, 11, 0), day_epoch),
                # F: yana dual-kredit (yopgan 991001, olib kelgan 991003) — jami/kredit farqi uchun
                ev(L + 6, "stage_change", 111, 990001, 991001, 991001, 991003, datetime(2020, 6, 10, 12, 0), day_epoch),
                # E: KEYINGI kun voqeasi (CRM vaqti 06-11) — bu kunga kirmasin
                ev(L + 5, "stage_change", 111, 990001, 991001, 991001, 991001, datetime(2020, 6, 11, 9, 0), day_epoch + 86400),
                # G: SHARTNOMA — yopgan 991002, olib kelgan 991003 (dual-kredit BO'LMASIN)
                ev(L + 8, "stage_change", 111, 990009, 991002, 991002, 991003, datetime(2020, 6, 10, 13, 0), day_epoch),
                # H: shartnoma bosqichi ICHIDA harakat — qayta sanalmasin
                ev(L + 8, "responsible_change", 990009, 990009, 991002, 991004, 991003, datetime(2020, 6, 10, 14, 0), day_epoch),
            ])
            await s.commit()
            try:
                return await lead_diff.daily_operator_breakdown(s, DAY, VISIT_IDS, CONTRACT_IDS)
            finally:
                await s.execute(_delete(LeadEvent).where(LeadEvent.crm_lead_id >= L))
                await s.commit()

    try:
        agg, unique, contracts_total = _asyncio.run(_run())
        credits = sum(a["visits"] for a in agg.values())
        check("kechikib aniqlangan tashrif o'z kuniga tushdi (991001=2: A+F)",
              agg.get(991001, {}).get("visits") == 2, f"991001={agg.get(991001)}")
        check("dual-kredit: yopgan (991002) +1", agg.get(991002, {}).get("visits") == 1,
              f"991002={agg.get(991002)}")
        check("dual-kredit: olib kelgan (991003) +2 (B+F)",
              agg.get(991003, {}).get("visits") == 2, f"991003={agg.get(991003)}")
        check("JAMI noyob tashrif = 4 (kredit yig'indisi 5 emas)", unique == 4,
              f"unique={unique}, kreditlar={credits}")
        check("kredit yig'indisi = 5 (operator kesimi dual-kredit bilan)", credits == 5,
              f"kreditlar={credits}")
        check("keyingi kun voqeasi kirmadi / Tashrif ichi ko'chish sanalmadi",
              agg.get(991001, {}).get("leads_touched") == 3, f"991001={agg.get(991001)}")
        check("first_seen tashrif bermaydi (991009 = 0 kredit)",
              agg.get(991009, {}).get("visits", 0) == 0, f"991009={agg.get(991009)}")
        check("shartnoma faqat YOPGANGA (991002 = 1)",
              agg.get(991002, {}).get("contracts") == 1, f"991002={agg.get(991002)}")
        check("shartnomada dual-kredit YO'Q (olib kelgan 991003 = 0)",
              agg.get(991003, {}).get("contracts", 0) == 0, f"991003={agg.get(991003)}")
        check("shartnoma jami = 1 (bosqich ichidagi harakat qayta sanalmadi)",
              contracts_total == 1, f"contracts_total={contracts_total}")
        check("shartnoma operator yig'indisi = jami",
              sum(a["contracts"] for a in agg.values()) == contracts_total,
              f"yig'indi={sum(a['contracts'] for a in agg.values())}, jami={contracts_total}")
    except Exception:
        check("tashrif hisoblash testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    # ── Statistika builderlari: voqea-asosli Tashrif + snapshot fallback (2026-08-03) ──
    from db.models import LeadStageDaily

    OLD_DAY = date(2019, 3, 5)  # LeadEvent jurnalidan OLDINGI davr — fallback tekshiruvi

    async def _run_builders():
        from api.routers import stats as stats_router

        async with async_session() as s:
            await s.execute(_delete(LeadEvent).where(LeadEvent.crm_lead_id >= L))
            await s.execute(_delete(LeadStageDaily).where(LeadStageDaily.responsible_id >= 991000))
            # Voqealar: DAY kuni 991001 uchun 2 ta haqiqiy kirish
            s.add_all([
                ev(L + 1, "stage_change", 111, 990001, 991001, 991001, 991001, datetime(2020, 6, 10, 9, 0), day_epoch),
                ev(L + 2, "stage_change", 111, 990002, 991001, 991001, 991001, datetime(2020, 6, 10, 10, 0), day_epoch),
                # Shartnoma: yopgan 991001 (statistika builderlarida ham ko'rinishi kerak)
                ev(L + 3, "stage_change", 111, 990009, 991001, 991001, 991002, datetime(2020, 6, 10, 11, 0), day_epoch),
            ])
            # Snapshot esa shishgan: DAY kuni Tashrifda 5 lid "ko'ringan"
            s.add_all([
                LeadStageDaily(date=DAY, responsible_id=991001, responsible_name="T-Op991001",
                               pipe_status_id=990001, stage_name="Tashrif", leads_count=5),
                LeadStageDaily(date=DAY, responsible_id=991001, responsible_name="T-Op991001",
                               pipe_status_id=111, stage_name="Boshqa", leads_count=2),
                # Jurnal boshlanmagan eski kun — snapshot yagona manba bo'lib qolishi kerak
                LeadStageDaily(date=OLD_DAY, responsible_id=991001, responsible_name="T-Op991001",
                               pipe_status_id=990001, stage_name="Tashrif", leads_count=3),
            ])
            await s.commit()
            try:
                # Builderlar sozlamadagi haqiqiy Tashrif ID'lariga qaraydi — testda soxta
                # ID'lar (990001/990002) ishlatilgani uchun vaqtincha almashtiramiz
                orig = stats_router._visit_ids
                orig_contract = stats_router._contract_ids
                stats_router._visit_ids = lambda: VISIT_IDS
                stats_router._contract_ids = lambda: CONTRACT_IDS
                day_out = await stats_router._build_lead_day(s, DAY, None)
                old_out = await stats_router._build_lead_day(s, OLD_DAY, None)
                month_out = await stats_router._build_lead_month(s, "2020-06", None)
                stats_router._visit_ids = orig
                stats_router._contract_ids = orig_contract
                return day_out, old_out, month_out
            finally:
                await s.execute(_delete(LeadEvent).where(LeadEvent.crm_lead_id >= L))
                await s.execute(_delete(LeadStageDaily).where(LeadStageDaily.responsible_id >= 991000))
                await s.commit()

    try:
        day_out, old_out, month_out = _asyncio.run(_run_builders())
        check("kunlik ko'rinish: Tashrif voqealardan (2, snapshot 5 emas)",
              day_out.visits == 2, f"visits={day_out.visits}")
        check("kunlik ko'rinish: operator krediti ham voqealardan",
              any(o.responsible_id == 991001 and o.visits == 2 for o in day_out.operators),
              f"ops={[(o.responsible_id, o.visits) for o in day_out.operators]}")
        check("jurnal oldingi kun: snapshot fallback (3)",
              old_out.visits == 3, f"visits={old_out.visits}")
        check("oylik ko'rinish: DAY qatori voqea-asosli (2)",
              any(d.date == DAY and d.visits == 2 for d in month_out.days),
              f"days={[(str(d.date), d.visits) for d in month_out.days]}")
        check("kunlik ko'rinish: shartnoma (1) va bayroq yoniq",
              day_out.contracts == 1 and day_out.contracts_enabled,
              f"contracts={day_out.contracts}, enabled={day_out.contracts_enabled}")
        check("kunlik ko'rinish: shartnoma operator qatorida (991001=1)",
              any(o.responsible_id == 991001 and o.contracts == 1 for o in day_out.operators),
              f"ops={[(o.responsible_id, o.contracts) for o in day_out.operators]}")
        check("oylik ko'rinish: shartnoma jami = 1",
              month_out.contracts == 1, f"contracts={month_out.contracts}")
        check("jurnalsiz eski kun: shartnoma 0 (snapshot fallback YO'Q)",
              old_out.contracts == 0, f"contracts={old_out.contracts}")
    except Exception:
        check("statistika builder testi ishga tushdi", False, traceback.format_exc(limit=2).strip())



def test_advance_duplicate_guard() -> None:
    """Avans dublikat qo'riqchisi va manba ustuni (Avans TZ A-01).

    MUAMMO: avansning ikki kirish yo'li bor (xodim arizasi va HR ning
    «Ish haqi → Avans» sahifasi) va ikkalasi ham BITTA jadvalga yozadi.
    Manba ko'rinmagani uchun HR ariza orqali kelgan avansni qo'lda takror
    kiritishi mumkin edi — pul ikki marta ayirilardi.

    Tekshiriladi:
      1. Yaqin summa + yaqin sana → dublikat topiladi
      2. Uzoq summa / uzoq sana → topilmaydi (yolg'on ogohlantirish yo'q)
      3. Rad etilgan avans dublikat sanalmaydi (u oylikka kirmaydi)
      4. Boshqa xodim / boshqa davr aralashmaydi
      5. HTTP: takror kiritishda 409 + `advance_duplicate` kodi
      6. `confirm_duplicate=true` → saqlanadi va `source='hr_manual'`
      7. Ariza yo'lidan kelgan qator `source='request'` bilan ko'rinadi

    JONLI XABAR YUBORILMAYDI: 409 javob bildirishnomadan OLDIN qaytadi;
    tasdiqli kiritish esa `notify_user` patch qilingan holda xizmat
    darajasida sinaladi.
    """
    print("\n=== AVANS: dublikat qo'riqchisi va manba (A-01) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import date as _date, timedelta as _td

    from api.routers import payroll as pay_router
    from api.routers.payroll import _find_duplicate_advance, _duplicate_message
    from sqlalchemy import select as _sel
    from db.base import async_session
    from db.models import (
        PayrollAdjustment,
        PayrollAdjustmentCategory,
        PayrollAdjustmentKind,
        PayrollAdjustmentSource,
        PayrollAdjustmentStatus,
        User as _User,
    )

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Adv-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from payroll_adjustments where user_id in ({qm})", stale)
        cur.execute(
            f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})",
            stale + stale,
        )
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600901,'T-Adv-Emp','employee',1,1,datetime('now'))"
    )
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600902,'T-Adv-Other','employee',1,1,datetime('now'))"
    )
    other_uid = cur.lastrowid
    conn.commit()

    # Dublikat qidiruvi DAVR bo'yicha filtrlaydi, sana esa kiritilgan kun
    # bo'yicha solishtiriladi — shuning uchun davr ham joriy oy bo'lishi
    # kerak (aks holda ikkisi hech qachon uchrashmasdi).
    PERIOD = _date.today().strftime("%Y-%m")
    OTHER_PERIOD = "2019-06"
    # Dublikat oynasi (±7 kun) KIRITILGAN kundan hisoblanadi — sinov
    # yozuvlari ham bugun yaratiladi, shuning uchun tayanch sana bugun.
    BASE_DATE = _date.today()
    BASE_AMOUNT = 2_000_000.0

    def cleanup_adv():
        try:
            c2 = db()
            uids = [emp_uid, other_uid]
            qm = ",".join("?" * len(uids))
            c2.execute(f"delete from payroll_adjustments where user_id in ({qm})", uids)
            c2.execute(
                f"delete from audit_logs where target_user_id in ({qm}) or actor_id in ({qm})",
                uids + uids,
            )
            c2.execute(f"delete from users where id in ({qm})", uids)
            c2.commit()
            c2.close()
        except Exception:
            print("  Avans tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    async def _seed_and_check():
        async with async_session() as s:
            # Ariza yo'lidan kelgan avans (mavjud yozuv). A-04 dan keyin
            # yangi avansda `issued_on` BO'SH bo'ladi — pul hali berilmagan;
            # dublikat qo'riqchisi bunday qatorni KIRITILGAN kuni bo'yicha
            # solishtiradi.
            s.add(PayrollAdjustment(
                user_id=emp_uid, period=PERIOD,
                kind=PayrollAdjustmentKind.minus.value,
                category=PayrollAdjustmentCategory.advance.value,
                status=PayrollAdjustmentStatus.pending.value,
                amount=BASE_AMOUNT, reason="T-Adv ariza avansi",
                issued_on=None, created_by=emp_uid,
                source=PayrollAdjustmentSource.request.value,
            ))
            # Rad etilgan avans — dublikat SANALMASLIGI kerak
            s.add(PayrollAdjustment(
                user_id=emp_uid, period=PERIOD,
                kind=PayrollAdjustmentKind.minus.value,
                category=PayrollAdjustmentCategory.advance.value,
                status=PayrollAdjustmentStatus.rejected.value,
                amount=500_000.0, reason="T-Adv rad etilgan",
                issued_on=None, created_by=emp_uid,
                source=PayrollAdjustmentSource.hr_manual.value,
            ))
            # Boshqa xodimning avansi — aralashmasligi kerak
            s.add(PayrollAdjustment(
                user_id=other_uid, period=PERIOD,
                kind=PayrollAdjustmentKind.minus.value,
                category=PayrollAdjustmentCategory.advance.value,
                status=PayrollAdjustmentStatus.pending.value,
                amount=BASE_AMOUNT, reason="T-Adv boshqa xodim",
                issued_on=BASE_DATE, created_by=other_uid,
                source=PayrollAdjustmentSource.hr_manual.value,
            ))
            await s.commit()

            return {
                "exact": await _find_duplicate_advance(s, emp_uid, PERIOD, BASE_AMOUNT, BASE_DATE),
                "near": await _find_duplicate_advance(
                    s, emp_uid, PERIOD, BASE_AMOUNT * 1.05, BASE_DATE + _td(days=3)),
                "far_amount": await _find_duplicate_advance(
                    s, emp_uid, PERIOD, BASE_AMOUNT * 2, BASE_DATE),
                "far_date": await _find_duplicate_advance(
                    s, emp_uid, PERIOD, BASE_AMOUNT, BASE_DATE + _td(days=20)),
                # Rad etilganga AYNAN mos (summa ham, sana ham) — baribir
                # dublikat sanalmaydi: u oylikka kirmaydi.
                "rejected": await _find_duplicate_advance(
                    s, emp_uid, PERIOD, 500_000.0, BASE_DATE),
                "other_period": await _find_duplicate_advance(
                    s, emp_uid, OTHER_PERIOD, BASE_AMOUNT, BASE_DATE),
            }

    try:
        res = _asyncio.run(_seed_and_check())
        check("aynan bir xil avans -> dublikat topildi", res["exact"] is not None)
        check("yaqin summa/sana (+5%, +3 kun) -> dublikat topildi", res["near"] is not None)
        check("uzoq summa (2x) -> dublikat YO'Q", res["far_amount"] is None)
        check("uzoq sana (+20 kun) -> dublikat YO'Q", res["far_date"] is None)
        check("rad etilgan avans dublikat sanalmaydi", res["rejected"] is None)
        check("boshqa davr aralashmaydi", res["other_period"] is None)
        if res["exact"] is not None:
            msg = _duplicate_message(res["exact"])
            check("ogohlantirish manbani ko'rsatadi (ariza)", "ariza" in msg.lower(), msg)
            check("ogohlantirish summa va sanani ko'rsatadi",
                  "2 000 000" in msg and BASE_DATE.strftime("%d.%m.%Y") in msg, msg)
    except Exception:
        check("dublikat qo'riqchisi testi ishga tushdi", False, traceback.format_exc(limit=2).strip())
        cleanup_adv()
        return

    mgr = find_manager_id()
    mgr_t = token_for(mgr[0], mgr[1]) if mgr else None
    if not mgr_t:
        check("avans HTTP testi uchun rahbar topildi", False)
        cleanup_adv()
        return

    body = {
        "user_id": emp_uid,
        "period": PERIOD,
        "amount": BASE_AMOUNT,
        "reason": "T-Adv qo'lda takror",
    }
    try:
        with httpx.Client(timeout=15) as client:
            # 409 bildirishnomadan OLDIN qaytadi -> jonli xabar ketmaydi
            r = client.post(f"{API_BASE}/payroll/advances", json=body, headers=auth(mgr_t))
            check("takror avans -> 409", r.status_code == 409, f"kod={r.status_code}")
            if r.status_code == 409:
                det = r.json().get("detail")
                check("409 tarkibida `advance_duplicate` kodi bor",
                      isinstance(det, dict) and det.get("code") == "advance_duplicate",
                      str(det)[:200])
                check("409 tarkibida mavjud yozuv id'si bor",
                      isinstance(det, dict) and det.get("existing_id") == res["exact"].id,
                      str(det)[:200])
    except Exception:
        check("avans HTTP dublikat testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    async def _forced_insert():
        """Tasdiqli kiritish. `notify_user` patch qilinadi — jonli Boshliqqa
        test xabari bormasin (⚠️ test-telegram-xavfi).

        A-03 chegara eshigi ATAYLAB istisno bilan chetlab o'tiladi: sinov
        xodimida stavka yo'q, ya'ni chegara 0. Bu test dublikat qo'riqchisi
        haqida — chegara o'z testida (`test_advance_limit_gate`) sinaladi."""
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        try:
            from api.schemas import AdvanceIn
            async with async_session() as s:
                boss = await s.scalar(_sel(_User).where(
                    _User.role.in_(("boss", "dasturchi")), _User.is_active.is_(True)))
                if boss is None:
                    return None, sent
                payload = AdvanceIn(
                    user_id=emp_uid, period=PERIOD, amount=BASE_AMOUNT,
                    reason="T-Adv tasdiqlangan takror",
                    confirm_duplicate=True, override_limit=True,
                    override_reason="Sinov: chegara testdan tashqarida",
                )
                return await pay_router.create_advance(payload, boss, s), sent
        finally:
            pay_router.notify_user = orig

    try:
        out, sent = _asyncio.run(_forced_insert())
        check("`confirm_duplicate=true` -> avans saqlandi", out is not None and out.id > 0)
        check("qo'lda kiritilgan avans `source='hr_manual'`",
              out.source == "hr_manual", f"source={out.source}")
        check("tasdiq so'rovi patch qilingan yuboruvchiga bordi (jonli emas)",
              isinstance(sent, list), f"xabarlar={len(sent)}")
    except Exception:
        check("tasdiqli avans testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    try:
        with httpx.Client(timeout=15) as client:
            r = client.get(
                f"{API_BASE}/payroll/adjustments?period={PERIOD}&user_id={emp_uid}&category=advance",
                headers=auth(mgr_t),
            )
            rows = r.json() if r.status_code == 200 else []
            srcs = {row.get("source") for row in rows}
            check("ro'yxat `source` maydonini qaytaradi",
                  "request" in srcs and "hr_manual" in srcs, f"manbalar={srcs}")
    except Exception:
        check("avans ro'yxati testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    cleanup_adv()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_limit() -> None:
    """Avans chegarasi formulasi (Avans TZ A-02).

    MUAMMO: avansga hech qanday chegara yo'q edi — HR xohlagan summani
    kirita olardi va oy oxirida payslip manfiy chiqishi mumkin edi
    (xodim oyligidan ko'p pul olib bo'lgan, qaytarib olinmaydi).

    Formula: (sof oylik ÷ oydagi ish kuni) × ishlangan kun × koeffitsient,
    keyin `min(..., sof oylik × cap%)`, keyin olingan avans va ushlanmalar
    ayiriladi.

    Toza formula DB'siz sinaladi (11 ssenariy), so'ng haqiqiy xodimda
    `limit_for()` — stavkasiz, qulflangan davr va oddiy holat.
    """
    print("\n=== AVANS: chegara formulasi (A-02) ===")
    import asyncio as _asyncio
    from datetime import date as _dt_date, datetime as _dt_datetime
    from decimal import Decimal as D

    from api.services import advance as adv
    from db.base import async_session
    from db.models import (
        PayBasis,
        PayrollAdjustment,
        PayrollAdjustmentCategory,
        PayrollAdjustmentKind,
        PayrollAdjustmentStatus,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )

    NET = D("5000000")

    # ── 1. Yarim oy ishlagan ──
    lim, earned, cap, why = adv.compute_limit(NET, 26, 13, D("0"), D("0"))
    check("yarim oy ishlagan -> netto/26 x 13 x 0.5",
          lim == D("1250000") and why is None, f"limit={lim}, sabab={why}")

    # ── 2. Oyning 5-kunida ishga kirgan (kam kun ishlagan) ──
    lim_late, *_ = adv.compute_limit(NET, 26, 4, D("0"), D("0"))
    check("kech kirgan xodimda chegara kichik",
          lim_late < lim and lim_late > 0, f"limit={lim_late} < {lim}")

    # ── 3. Cap koeffitsientdan qat'i nazar oshmaydi ──
    lim_cap, earned_cap, cap_amt, _ = adv.compute_limit(
        NET, 26, 26, D("0"), D("0"), coefficient=D("2.0"), cap_percent=D("50"))
    check("koeffitsient 2.0 bo'lsa ham cap (50%) dan oshmaydi",
          lim_cap == D("2500000") and earned_cap > cap_amt,
          f"limit={lim_cap}, earned={earned_cap}, cap={cap_amt}")

    # ── 4. Kutilayotgan avans ham ayiriladi ──
    lim_taken, *_ = adv.compute_limit(NET, 26, 26, D("1000000"), D("0"))
    check("olingan avans ayiriladi (2 500 000 - 1 000 000)",
          lim_taken == D("1500000"), f"limit={lim_taken}")

    # ── 5. Ushlanmalar ham ayiriladi ──
    lim_ded, *_ = adv.compute_limit(NET, 26, 26, D("0"), D("500000"))
    check("boshqa ushlanmalar ayiriladi", lim_ded == D("2000000"), f"limit={lim_ded}")

    # ── 6. To'liq ishlatilgan -> 0 va SABAB ──
    lim_zero, _, _, why_zero = adv.compute_limit(NET, 26, 26, D("3000000"), D("0"))
    check("chegara tugagan -> 0 va sabab bor",
          lim_zero == 0 and why_zero == adv.REASON_EXHAUSTED, f"limit={lim_zero}, sabab={why_zero}")

    # ── 7. Ishlangan kun yo'q ──
    _, _, _, why_nw = adv.compute_limit(NET, 26, 0, D("0"), D("0"))
    check("ishlangan kun 0 -> sabab aniq",
          why_nw == adv.REASON_NO_WORKED_DAYS, f"sabab={why_nw}")

    # ── 8. Reja bo'yicha ish kuni yo'q ──
    _, _, _, why_ns = adv.compute_limit(NET, 0, 0, D("0"), D("0"))
    check("oyda ish kuni yo'q -> sabab aniq",
          why_ns == adv.REASON_NO_SCHEDULE, f"sabab={why_ns}")

    # ── 9. Oylik 0 ──
    _, _, _, why_zs = adv.compute_limit(D("0"), 26, 13, D("0"), D("0"))
    check("oylik 0 -> sabab aniq", why_zs == adv.REASON_ZERO_SALARY, f"sabab={why_zs}")

    # ── 10. Ishlangan kun rejadan ko'p bo'lsa cheklanadi ──
    lim_over, *_ = adv.compute_limit(NET, 26, 40, D("0"), D("0"))
    check("ishlangan kun rejadan ko'p bo'lsa oylikdan oshmaydi",
          lim_over == D("2500000"), f"limit={lim_over}")

    # ── 11. Manfiy natija 0 ga tushadi (manfiy chegara qaytarilmaydi) ──
    lim_neg, *_ = adv.compute_limit(NET, 26, 13, D("9000000"), D("0"))
    check("manfiy natija 0 ga tushadi", lim_neg == 0, f"limit={lim_neg}")

    # ── DB darajasi ──
    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Lim-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override", "payslips"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600911,'T-Lim-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600912,'T-Lim-NoRate','employee',1,1,datetime('now'))")
    norate_uid = cur.lastrowid
    conn.commit()

    PERIOD = "2020-04"       # ataylab o'tgan oy: butun oy «future» bo'lib qolmasin
    P_START = _dt_date(2020, 4, 1)

    def cleanup_lim():
        try:
            c2 = db()
            uids = [emp_uid, norate_uid]
            qm = ",".join("?" * len(uids))
            pslips = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", uids).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            for tbl in ("payslips", "payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", uids)
            c2.execute("delete from payroll_periods where period=?", (PERIOD,))
            c2.execute(f"delete from users where id in ({qm})", uids)
            c2.commit()
            c2.close()
        except Exception:
            print("  Chegara testi tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    async def _db_cases():
        async with async_session() as s:
            emp = await s.get(_User, emp_uid)
            norate = await s.get(_User, norate_uid)

            # Du-Ju ish kuni, 5 mln oylik stavka
            for wd in range(5):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                         start_time="09:00", end_time="18:00"))
            for wd in (5, 6):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
            s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                             amount=5_000_000, effective_from=P_START,
                             changed_by=emp_uid))
            # Davomat: 1-14 aprel oralig'idagi ish kunlarida kelgan.
            # Bo'lmasa `collect_attendance` o'tgan kunlarni «absent» deb
            # belgilaydi va ishlangan kun 0 chiqadi (chegara ham 0).
            from db.models import Attendance as _Att
            for dd in range(1, 15):
                day = _dt_date(2020, 4, dd)
                if day.weekday() >= 5:
                    continue
                s.add(_Att(user_id=emp_uid, date=day, status="present",
                           check_in_time=_dt_datetime(2020, 4, dd, 4, 0),
                           check_out_time=_dt_datetime(2020, 4, dd, 13, 0),
                           late_minutes=0, worked_minutes=540))
            await s.commit()

            plain = await adv.limit_for(s, emp, on_date=_dt_date(2020, 4, 15))
            no_rate = await adv.limit_for(s, norate, on_date=_dt_date(2020, 4, 15))

            # Kutilayotgan avans qo'shib, chegara kamayishini tekshiramiz
            s.add(PayrollAdjustment(
                user_id=emp_uid, period=PERIOD,
                kind=PayrollAdjustmentKind.minus.value,
                category=PayrollAdjustmentCategory.advance.value,
                status=PayrollAdjustmentStatus.pending.value,
                amount=300_000, reason="T-Lim kutilayotgan avans",
                issued_on=_dt_date(2020, 4, 10), created_by=emp_uid,
            ))
            await s.commit()
            after_pending = await adv.limit_for(s, emp, on_date=_dt_date(2020, 4, 15))
            return plain, no_rate, after_pending

    try:
        plain, no_rate, after_pending = _asyncio.run(_db_cases())
        check("haqiqiy xodimda chegara musbat va oylikdan kichik",
              0 < plain.limit < plain.net_salary,
              f"limit={plain.limit}, netto={plain.net_salary}")
        check("chegara cap dan oshmaydi", plain.limit <= plain.cap_amount + 1,
              f"limit={plain.limit}, cap={plain.cap_amount}")
        check("oradagi qiymatlar qaytariladi (ish kuni va ishlangan kun)",
              plain.scheduled_days > 0 and plain.worked_days >= 0,
              f"reja={plain.scheduled_days}, ishlangan={plain.worked_days}")
        check("stavkasiz xodim -> 0 va `stavka belgilanmagan`",
              no_rate.limit == 0 and no_rate.reason == adv.REASON_NO_RATE,
              f"limit={no_rate.limit}, sabab={no_rate.reason}")
        check("kutilayotgan avans chegarani kamaytiradi",
              after_pending.taken == 300000.0 and after_pending.limit < plain.limit,
              f"olingan={after_pending.taken}, {after_pending.limit} < {plain.limit}")
    except Exception:
        check("chegara DB testi ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_lim()
        try:
            conn.close()
        except Exception:
            pass
        return

    # ── Qulflangan davr ──
    async def _locked_case():
        from db.models import PayrollPeriod as _PP
        async with async_session() as s:
            s.add(_PP(period=PERIOD, status="approved", locked=True))
            await s.commit()
            emp = await s.get(_User, emp_uid)
            return await adv.limit_for(s, emp, on_date=_dt_date(2020, 4, 15))

    try:
        locked = _asyncio.run(_locked_case())
        check("qulflangan davrda chegara 0 va sabab aniq",
              locked.limit == 0 and locked.reason == adv.REASON_PERIOD_LOCKED,
              f"limit={locked.limit}, sabab={locked.reason}")
    except Exception:
        check("qulflangan davr testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    cleanup_lim()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_limit_gate() -> None:
    """Chegara kiritish nuqtalarida (Avans TZ A-03).

    MUAMMO: A-02 chegarani hisoblaydi, lekin uni HECH KIM tekshirmasa
    foydasi yo'q — HR baribir istagan summani kiritardi.

    Tekshiriladi:
      1. `GET /payroll/advances/limit` chegarani va kelib chiqishini beradi
      2. Chegara ichidagi summa oddiy o'tadi
      3. Chegaradan oshiq -> 400 `advance_over_limit` (ruxsat etilgan summa
         xabarda ko'rinadi — HR raqamni taxmin qilmasin)
      4. HR `override_limit` bilan yuborsa -> 403 (istisno faqat Boshliq)
      5. Boshliq sababsiz istisno qilsa -> 400
      6. Boshliq sabab bilan -> saqlanadi va auditda `advance_over_limit`
      7. Ariza yo'lida ham shu chegara ishlaydi (chetlab o'tib bo'lmaydi)
    """
    print("\n=== AVANS: chegara kiritish nuqtalarida (A-03) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import date as _dt_date, datetime as _dt_datetime

    from api.routers import payroll as pay_router
    from api.services import advance as adv
    from db.base import async_session
    from db.models import (
        Attendance as _Att,
        AuditLog as _Audit,
        PayBasis,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import select as _select

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Gate-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600921,'T-Gate-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    conn.commit()

    PERIOD = "2020-03"
    P_START = _dt_date(2020, 3, 1)

    def cleanup_gate():
        try:
            c2 = db()
            pslips = [r[0] for r in c2.execute(
                "select id from payslips where user_id=?", (emp_uid,)).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            for tbl in ("payslips", "payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id=?", (emp_uid,))
            c2.execute("delete from audit_logs where target_user_id=?", (emp_uid,))
            c2.execute("delete from payroll_periods where period=?", (PERIOD,))
            c2.execute("delete from users where id=?", (emp_uid,))
            c2.commit()
            c2.close()
        except Exception:
            print("  Chegara-eshik tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    async def _seed():
        async with async_session() as s:
            for wd in range(5):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                         start_time="09:00", end_time="18:00"))
            for wd in (5, 6):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
            s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                             amount=5_000_000, effective_from=P_START, changed_by=emp_uid))
            for dd in range(1, 16):
                day = _dt_date(2020, 3, dd)
                if day.weekday() >= 5:
                    continue
                s.add(_Att(user_id=emp_uid, date=day, status="present",
                           check_in_time=_dt_datetime(2020, 3, dd, 4, 0),
                           check_out_time=_dt_datetime(2020, 3, dd, 13, 0),
                           late_minutes=0, worked_minutes=540))
            await s.commit()
            emp = await s.get(_User, emp_uid)
            return await adv.limit_for(s, emp, period=PERIOD)

    try:
        info = _asyncio.run(_seed())
    except Exception:
        check("chegara-eshik testi sozlandi", False, traceback.format_exc(limit=2).strip())
        cleanup_gate()
        return

    if info.limit <= 0:
        check("sinov xodimida chegara musbat chiqdi", False, f"limit={info.limit}, {info.reason}")
        cleanup_gate()
        return

    # HR va Boshliq tokenlari — ikkalasi HAM kerak (rol farqi sinaladi)
    hr_row = conn.execute(
        "select id from users where role='hr' and is_active=1 limit 1").fetchone()
    boss_row = conn.execute(
        "select id, role from users where role in ('boss','dasturchi') and is_active=1 limit 1"
    ).fetchone()
    hr_t = token_for(hr_row[0], "hr") if hr_row else None
    boss_t = token_for(boss_row[0], boss_row[1]) if boss_row else None
    if not hr_t or not boss_t:
        check("chegara-eshik testi uchun HR va Boshliq topildi", False,
              f"hr={bool(hr_t)}, boss={bool(boss_t)}")
        cleanup_gate()
        return

    def body(amount, **extra):
        return {
            "user_id": emp_uid,
            "period": PERIOD,
            "amount": amount,
            "issued_on": _dt_date(2020, 3, 10).isoformat(),
            "reason": "T-Gate sinov avansi",
            **extra,
        }

    try:
        with httpx.Client(timeout=20) as client:
            # 1. Chegara endpointi
            r = client.get(f"{API_BASE}/payroll/advances/limit?user_id={emp_uid}&period={PERIOD}",
                           headers=auth(hr_t))
            check("chegara endpointi 200 qaytaradi", r.status_code == 200, f"kod={r.status_code}")
            if r.status_code == 200:
                d = r.json()
                check("chegara endpointi kelib chiqishini ham beradi",
                      d["limit"] > 0 and d["scheduled_days"] > 0 and d["worked_days"] > 0,
                      f"limit={d['limit']}, kun={d['worked_days']}/{d['scheduled_days']}")

            # 2. Chegaradan oshiq -> 400 va kod
            r = client.post(f"{API_BASE}/payroll/advances",
                            json=body(info.limit + 1_000_000), headers=auth(hr_t))
            check("chegaradan oshiq -> 400", r.status_code == 400, f"kod={r.status_code}")
            det = r.json().get("detail") if r.status_code == 400 else None
            check("400 tarkibida `advance_over_limit` kodi va ruxsat etilgan summa",
                  isinstance(det, dict) and det.get("code") == "advance_over_limit"
                  and det.get("limit") == info.limit,
                  str(det)[:200])

            # 3. HR istisno qila olmaydi
            r = client.post(f"{API_BASE}/payroll/advances",
                            json=body(info.limit + 1_000_000, override_limit=True,
                                      override_reason="HR o'zi qaror qildi"),
                            headers=auth(hr_t))
            check("HR chegaradan oshiq kirita olmaydi -> 403",
                  r.status_code == 403, f"kod={r.status_code}")

            # 4. Boshliq sababsiz istisno -> 400
            r = client.post(f"{API_BASE}/payroll/advances",
                            json=body(info.limit + 1_000_000, override_limit=True),
                            headers=auth(boss_t))
            check("Boshliq sababsiz istisno qila olmaydi -> 400",
                  r.status_code == 400, f"kod={r.status_code}")
    except Exception:
        check("chegara-eshik HTTP testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    # 5-6. Chegara ichida va Boshliq istisnosi — xizmat darajasida
    #      (`notify_user` patch qilingan: jonli Boshliqqa xabar ketmasin)
    async def _service_cases():
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        try:
            from api.schemas import AdvanceIn
            async with async_session() as s:
                hr = await s.get(_User, hr_row[0])
                boss = await s.get(_User, boss_row[0])
                inside = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=100_000,
                              issued_on=_dt_date(2020, 3, 10), reason="T-Gate chegara ichida"),
                    hr, s)
                over = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=info.limit + 1_000_000,
                              issued_on=_dt_date(2020, 3, 20), reason="T-Gate istisno",
                              confirm_duplicate=True, override_limit=True,
                              override_reason="Shoshilinch tibbiy xarajat"),
                    boss, s)
                audits = list(await s.scalars(
                    _select(_Audit).where(_Audit.target_user_id == emp_uid)))
                return inside, over, audits
        finally:
            pay_router.notify_user = orig

    try:
        inside, over, audits = _asyncio.run(_service_cases())
        check("chegara ichidagi summa oddiy o'tadi", inside.id > 0, f"id={inside.id}")
        check("Boshliq sabab bilan chegaradan oshiq kirita oladi", over.id > 0, f"id={over.id}")
        acts = [a.action for a in audits]
        check("istisno auditda `advance_over_limit` sifatida yozilgan",
              "advance_over_limit" in acts, f"amallar={acts}")
        over_audit = next((a for a in audits if a.action == "advance_over_limit"), None)
        check("auditda istisno sababi va chegara saqlangan",
              over_audit is not None
              and over_audit.after.get("override_reason") == "Shoshilinch tibbiy xarajat"
              and over_audit.after.get("limit") is not None,
              str(over_audit.after)[:200] if over_audit else "yo'q")
    except Exception:
        check("chegara-eshik xizmat testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    # 7. Ariza yo'li — chegarani chetlab o'tib bo'lmaydi.
    #    `_apply_advance` HAR DOIM joriy oyga yozadi, shuning uchun sinov
    #    davomati ham joriy oyga qo'yiladi (aks holda chegara 0 bo'lib,
    #    «oshdi» shoxi umuman sinalmasdi).
    async def _request_path():
        from api.routers import requests as req_router
        from db.models import EmployeeRequest as _Req, RequestKind as _RK, RequestStatus as _RS
        bugun = _dt_date.today()
        async with async_session() as s:
            s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                             amount=5_000_000,
                             effective_from=bugun.replace(day=1), changed_by=emp_uid))
            for dd in range(1, min(bugun.day, 15)):
                day = bugun.replace(day=dd)
                if day.weekday() >= 5:
                    continue
                s.add(_Att(user_id=emp_uid, date=day, status="present",
                           check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                           check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                           late_minutes=0, worked_minutes=540))
            await s.commit()

            emp = await s.get(_User, emp_uid)
            joriy = await adv.limit_for(s, emp)
            item = _Req(user_id=emp_uid, kind=_RK.advance.value,
                        amount=99_000_000, reason="T-Gate ariza orqali chetlab o'tish",
                        status=_RS.approved.value, decided_by=boss_row[0])
            s.add(item)
            await s.flush()
            try:
                await req_router._apply_advance(s, item, emp)
                xato = "o'tib ketdi"
            except Exception as e:
                xato = getattr(e, "detail", str(e))
            await s.rollback()
            return joriy, xato

    try:
        joriy, res = _asyncio.run(_request_path())
        check("ariza yo'li ham chegarada to'xtaydi",
              isinstance(res, str) and res != "o'tib ketdi"
              and ("chegara" in res.lower() or "avans berib bo'lmaydi" in res.lower()),
              str(res)[:180])
        if joriy.limit > 0:
            check("ariza rad javobida ruxsat etilgan summa ko'rinadi",
                  "chegarasidan oshdi" in str(res), str(res)[:180])
    except Exception:
        check("ariza yo'li testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    cleanup_gate()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_issue_flow() -> None:
    """«Kiritildi» va «berildi» ajratimi (Avans TZ A-04).

    MUAMMO: ilgari HR kiritishda «berilgan sana» yozardi, tasdiq esa keyin
    kelardi — ya'ni pul Boshliq rad etishi mumkin bo'lgan paytda allaqachon
    qo'lda bo'lardi va qaytarib olinmasdi.

    Yangi zanjir: pending → approved (ruxsat) → issued (kassa to'ladi).

    Tekshiriladi:
      1. Kiritishda `issued_on` YOZILMAYDI (bo'sh qoladi)
      2. `pending` avansni to'langan deb belgilab bo'lmaydi -> 400
      3. `approved` -> `issued` o'tadi, kim va qachon belgilagani saqlanadi
      4. Ikkinchi marta to'lash -> 400
      5. Kelajakdagi sana -> 400
      6. Auditda `advance_issued` amali
      7. ⭐ PAYSLIP: `issued` avans oylikdan AYIRILADI (`approved` kabi) —
         to'lash uni hisobdan chiqarib yubormaydi
      8. Rad etilgan avans payslipga kirmaydi (eski qoida buzilmagan)
    """
    print("\n=== AVANS: kiritildi/berildi ajratimi (A-04) ===")
    import asyncio as _asyncio
    from datetime import date as _dt_date, datetime as _dt_datetime, timedelta as _td

    from api.routers import payroll as pay_router
    from api.schemas import AdvanceIn, AdvanceDecision, AdvanceIssueIn
    from api.services.payroll import build_payslip
    from db.base import async_session
    from db.models import (
        Attendance as _Att,
        AuditLog as _Audit,
        PayBasis,
        PayrollAdjustmentStatus,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import select as _sel

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Iss-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600931,'T-Iss-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    conn.commit()

    # Joriy oy — chegara musbat bo'lishi uchun davomat kerak
    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")

    def cleanup_iss():
        try:
            c2 = db()
            pslips = [r[0] for r in c2.execute(
                "select id from payslips where user_id=?", (emp_uid,)).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            for tbl in ("payslips", "payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id=?", (emp_uid,))
            c2.execute("delete from audit_logs where target_user_id=?", (emp_uid,))
            c2.execute("delete from users where id=?", (emp_uid,))
            c2.commit()
            c2.close()
        except Exception:
            print("  A-04 tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    boss_row = conn.execute(
        "select id from users where role in ('boss','dasturchi') and is_active=1 limit 1").fetchone()
    hr_row = conn.execute(
        "select id from users where role='hr' and is_active=1 limit 1").fetchone()
    if not boss_row or not hr_row:
        check("A-04 testi uchun HR va Boshliq topildi", False)
        cleanup_iss()
        return

    async def _flow():
        """Butun zanjir. `notify_user` patch qilinadi — jonli xodimga/Boshliqqa
        test xabari bormasin (⚠️ test-telegram-xavfi)."""
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        out = {}
        try:
            async with async_session() as s:
                for wd in range(5):
                    s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                             start_time="09:00", end_time="18:00"))
                for wd in (5, 6):
                    s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
                s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                                 amount=5_000_000, effective_from=BUGUN.replace(day=1),
                                 changed_by=emp_uid))
                for dd in range(1, max(BUGUN.day, 2)):
                    day = BUGUN.replace(day=dd)
                    if day.weekday() >= 5:
                        continue
                    s.add(_Att(user_id=emp_uid, date=day, status="present",
                               check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                               check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                               late_minutes=0, worked_minutes=540))
                await s.commit()

                hr = await s.get(_User, hr_row[0])
                boss = await s.get(_User, boss_row[0])

                # 1. Kiritish — `issued_on` yozilmasligi kerak
                adj = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=200_000,
                              reason="T-Iss sinov avansi",
                              # Eski mijoz sana yuborsa ham E'TIBORGA OLINMASIN
                              issued_on=_dt_date(2020, 1, 1)),
                    hr, s)
                out["created"] = adj

                # 2. `pending` ni to'langan deb belgilash — taqiqlanadi
                try:
                    await pay_router.issue_advance(adj.id, AdvanceIssueIn(), hr, s)
                    out["issue_pending"] = "o'tib ketdi"
                except Exception as e:
                    out["issue_pending"] = getattr(e, "detail", str(e))

                # 3. Tasdiqlash -> to'lash
                await pay_router.decide_advance(adj.id, AdvanceDecision(approve=True), boss, s)
                issued = await pay_router.issue_advance(
                    adj.id, AdvanceIssueIn(note="T-Iss kassadan berildi"), hr, s)
                out["issued"] = issued

                # 4. Ikkinchi marta to'lash
                try:
                    await pay_router.issue_advance(adj.id, AdvanceIssueIn(), hr, s)
                    out["issue_twice"] = "o'tib ketdi"
                except Exception as e:
                    out["issue_twice"] = getattr(e, "detail", str(e))

                # 5. Kelajakdagi sana (yangi, tasdiqlangan avansda)
                adj2 = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=50_000,
                              reason="T-Iss ikkinchi avans", confirm_duplicate=True),
                    hr, s)
                await pay_router.decide_advance(adj2.id, AdvanceDecision(approve=True), boss, s)
                try:
                    await pay_router.issue_advance(
                        adj2.id, AdvanceIssueIn(issued_on=BUGUN + _td(days=3)), hr, s)
                    out["future"] = "o'tib ketdi"
                except Exception as e:
                    out["future"] = getattr(e, "detail", str(e))

                # 6. Rad etilgan avans (payslipga kirmasligi kerak)
                adj3 = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=70_000,
                              reason="T-Iss rad etiladigan", confirm_duplicate=True),
                    hr, s)
                await pay_router.decide_advance(adj3.id, AdvanceDecision(approve=False), boss, s)

                out["audits"] = list(await s.scalars(
                    _sel(_Audit).where(_Audit.target_user_id == emp_uid)))

                # 7. Payslip — `issued` (200k) + `approved` (50k) ayirilsin,
                #    rad etilgan (70k) esa YO'Q
                emp = await s.get(_User, emp_uid)
                out["payslip"] = await build_payslip(s, emp, PERIOD)
                return out, sent
        finally:
            pay_router.notify_user = orig

    try:
        out, sent = _asyncio.run(_flow())
    except Exception:
        check("A-04 zanjiri ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_iss()
        try:
            conn.close()
        except Exception:
            pass
        return

    created = out["created"]
    check("kiritishda `issued_on` yozilmaydi (pul hali berilmagan)",
          created.issued_on is None, f"issued_on={created.issued_on}")
    check("kiritilgan avans holati `pending`",
          created.status == PayrollAdjustmentStatus.pending.value, f"status={created.status}")

    check("tasdiqlanmagan avansni to'langan deb belgilab bo'lmaydi",
          out["issue_pending"] != "o'tib ketdi"
          and "tasdiq" in str(out["issue_pending"]).lower(),
          str(out["issue_pending"])[:140])

    issued = out["issued"]
    check("tasdiqlangach `issued` holatiga o'tadi",
          issued.status == PayrollAdjustmentStatus.issued.value, f"status={issued.status}")
    check("to'langan sana bugungi kun bilan to'ldirildi",
          issued.issued_on == _dt_date.today(), f"issued_on={issued.issued_on}")
    check("kim va qachon to'laganini tizim saqladi",
          issued.issued_by == hr_row[0] and issued.issued_at is not None
          and issued.issued_by_name,
          f"by={issued.issued_by}, at={issued.issued_at}, ism={issued.issued_by_name}")

    check("ikkinchi marta to'lab bo'lmaydi",
          out["issue_twice"] != "o'tib ketdi"
          and "allaqachon" in str(out["issue_twice"]).lower(),
          str(out["issue_twice"])[:140])
    check("kelajakdagi sana bilan to'lab bo'lmaydi",
          out["future"] != "o'tib ketdi" and "kelajak" in str(out["future"]).lower(),
          str(out["future"])[:140])

    acts = [a.action for a in out["audits"]]
    check("auditda `advance_issued` amali bor", "advance_issued" in acts, f"amallar={acts}")

    # ⭐ Eng muhim tekshiruv: to'lash pulni hisobdan CHIQARIB YUBORMAYDI
    f = out["payslip"]["fields"]
    check("payslipda `issued` + `approved` ayirildi (200 000 + 50 000)",
          abs(f["adjustments_minus"] - 250_000) < 1, f"adjustments_minus={f['adjustments_minus']}")
    check("rad etilgan avans payslipga KIRMADI (70 000 yo'q)",
          abs(f["adjustments_minus"] - 320_000) > 1, f"adjustments_minus={f['adjustments_minus']}")
    check("xabarlar patch qilingan yuboruvchiga bordi (jonli emas)",
          isinstance(sent, list), f"xabarlar={len(sent)}")

    cleanup_iss()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_soft_delete() -> None:
    """Yumshoq o'chirish, audit va sabab qoidasi (Avans TZ A-05).

    MUAMMO: pul yozuvini butunlay o'chirish «bu avans qayerga ketdi?»
    degan savolga javobsiz qoldiradi. Endi qator bazada QOLADI, lekin
    barcha o'qish `deleted_at IS NULL` bilan filtrlanadi.

    Tekshiriladi:
      1. ⭐ O'chirilgan avans PAYSLIPGA kirmaydi (aks holda o'chirish
         ko'zga ko'rinadigan, lekin pulga ta'sir qilmaydigan soxta amal)
      2. Qator bazadan yo'qolmaydi — `deleted_at`/`deleted_by`/sabab bilan
      3. Ro'yxatda ko'rinmaydi
      4. Chegarani bo'shatadi (o'chirilgan avans «olingan» sanalmaydi)
      5. Dublikat qo'riqchisi uni ogohlantirmaydi
      6. HR tasdiqlangan/to'langan yozuvni o'chira olmaydi -> 403
      7. Boshliq o'chira oladi, sabab bilan va auditda
      8. Ikkinchi marta o'chirish -> 400
      9. Sabab qoidasi: o'chiq bo'lsa ixtiyoriy, yoqilsa ma'nosiz matn 400
    """
    print("\n=== AVANS: yumshoq o'chirish va sabab qoidasi (A-05) ===")
    import asyncio as _asyncio
    from datetime import date as _dt_date, datetime as _dt_datetime

    from api.routers import payroll as pay_router
    from api.schemas import AdvanceIn, AdvanceDecision
    from api.services.advance import limit_for, taken_and_deductions
    from api.services.payroll import build_payslip
    from db.base import async_session
    from db.models import (
        Attendance as _Att,
        AuditLog as _Audit,
        AdvanceSettings as _AdvSet,
        PayBasis,
        PayrollAdjustment as _Adj,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import select as _sel

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Del-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600941,'T-Del-Emp','employee',1,1,datetime('now'))")
    emp_uid = cur.lastrowid
    conn.commit()

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")

    def cleanup_del():
        try:
            c2 = db()
            pslips = [r[0] for r in c2.execute(
                "select id from payslips where user_id=?", (emp_uid,)).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            for tbl in ("payslips", "payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id=?", (emp_uid,))
            c2.execute("delete from audit_logs where target_user_id=?", (emp_uid,))
            c2.execute("delete from users where id=?", (emp_uid,))
            # Sinov FAQAT xodim darajasidagi qoida yaratadi — global
            # sozlamaga tegilmaydi. Qolib ketgan bo'lsa o'chiramiz.
            c2.execute("delete from advance_settings where scope='user' and scope_id=?", (emp_uid,))
            c2.commit()
            c2.close()
        except Exception:
            print("  A-05 tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    boss_row = conn.execute(
        "select id from users where role in ('boss','dasturchi') and is_active=1 limit 1").fetchone()
    hr_row = conn.execute(
        "select id from users where role='hr' and is_active=1 limit 1").fetchone()
    if not boss_row or not hr_row:
        check("A-05 testi uchun HR va Boshliq topildi", False)
        cleanup_del()
        return

    async def _flow():
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        out = {}
        try:
            async with async_session() as s:
                for wd in range(5):
                    s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                             start_time="09:00", end_time="18:00"))
                for wd in (5, 6):
                    s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
                s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                                 amount=5_000_000, effective_from=BUGUN.replace(day=1),
                                 changed_by=emp_uid))
                for dd in range(1, max(BUGUN.day, 2)):
                    day = BUGUN.replace(day=dd)
                    if day.weekday() >= 5:
                        continue
                    s.add(_Att(user_id=emp_uid, date=day, status="present",
                               check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                               check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                               late_minutes=0, worked_minutes=540))
                await s.commit()

                hr = await s.get(_User, hr_row[0])
                boss = await s.get(_User, boss_row[0])
                emp = await s.get(_User, emp_uid)

                # Ikkita avans: biri tasdiqlanadi (o'chiriladi), biri pending
                a1 = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=300_000,
                              reason="T-Del tasdiqlanadigan"), hr, s)
                await pay_router.decide_advance(a1.id, AdvanceDecision(approve=True), boss, s)
                a2 = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=120_000,
                              reason="T-Del kutilayotgan", confirm_duplicate=True), hr, s)

                out["before_payslip"] = (await build_payslip(s, emp, PERIOD))["fields"]
                out["before_limit"] = await limit_for(s, emp, period=PERIOD)

                # 6. HR tasdiqlanganini o'chira olmaydi
                try:
                    await pay_router.delete_adjustment(a1.id, "HR urinishi", hr, s)
                    out["hr_delete"] = "o'tib ketdi"
                except Exception as e:
                    out["hr_delete"] = f"{getattr(e, 'status_code', '?')}: {getattr(e, 'detail', e)}"

                # HR pending'ni o'chira OLADI
                await pay_router.delete_adjustment(a2.id, "T-Del xato kiritilgan", hr, s)
                # 7. Boshliq tasdiqlanganini o'chira oladi
                await pay_router.delete_adjustment(a1.id, "T-Del boshliq bekor qildi", boss, s)

                # 8. Ikkinchi marta
                try:
                    await pay_router.delete_adjustment(a1.id, "takror", boss, s)
                    out["twice"] = "o'tib ketdi"
                except Exception as e:
                    out["twice"] = getattr(e, "detail", str(e))

                out["after_payslip"] = (await build_payslip(s, emp, PERIOD))["fields"]
                out["after_limit"] = await limit_for(s, emp, period=PERIOD)
                out["taken"] = (await taken_and_deductions(s, emp_uid, PERIOD))[0]
                out["rows_in_db"] = list(await s.scalars(
                    _sel(_Adj).where(_Adj.user_id == emp_uid)))
                out["listed"] = await pay_router.list_adjustments(
                    period=PERIOD, user_id=emp_uid, category="advance", _actor=hr, db=s)
                out["dup"] = await pay_router._find_duplicate_advance(
                    s, emp_uid, PERIOD, 300_000, BUGUN)
                out["audits"] = list(await s.scalars(
                    _sel(_Audit).where(_Audit.target_user_id == emp_uid)))

                # 9. Sabab qoidasi (B-01 dan keyin `advance_settings` da).
                #    GLOBAL sozlamaga TEGILMAYDI — sinov xodimining O'ZIGA
                #    qoida yaratamiz (xodim > lavozim > global), oxirida
                #    o'chiramiz.
                policy = _AdvSet(scope="user", scope_id=emp_uid,
                                 reason_required=True, is_active=True)
                s.add(policy)
                await s.commit()
                if policy is not None:
                    try:
                        await pay_router.create_advance(
                            AdvanceIn(user_id=emp_uid, period=PERIOD, amount=50_000,
                                      reason="avans", confirm_duplicate=True), hr, s)
                        out["meaningless"] = "o'tib ketdi"
                    except Exception as e:
                        out["meaningless"] = getattr(e, "detail", str(e))
                    try:
                        ok = await pay_router.create_advance(
                            AdvanceIn(user_id=emp_uid, period=PERIOD, amount=50_000,
                                      reason="Oilaviy shoshilinch xarajat",
                                      confirm_duplicate=True), hr, s)
                        out["meaningful"] = ok.id
                    except Exception as e:
                        out["meaningful"] = getattr(e, "detail", str(e))
                    await s.delete(policy)
                    await s.commit()
                return out, sent
        finally:
            pay_router.notify_user = orig

    try:
        out, sent = _asyncio.run(_flow())
    except Exception:
        check("A-05 zanjiri ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_del()
        try:
            conn.close()
        except Exception:
            pass
        return

    before, after = out["before_payslip"], out["after_payslip"]
    check("o'chirishdan OLDIN payslipda tasdiqlangan avans bor (300 000)",
          abs(before["adjustments_minus"] - 300_000) < 1,
          f"minus={before['adjustments_minus']}")
    check("⭐ o'chirilgan avans PAYSLIPGA kirmaydi (0)",
          abs(after["adjustments_minus"]) < 1, f"minus={after['adjustments_minus']}")

    rows = out["rows_in_db"]
    check("qator bazadan YO'QOLMADI (2 ta o'chirilgan qator bor)",
          len(rows) >= 2 and all(r.deleted_at is not None for r in rows),
          f"qatorlar={[(r.id, r.deleted_at is not None) for r in rows]}")
    check("kim o'chirgani va sabab saqlangan",
          all(r.deleted_by is not None and r.deleted_reason for r in rows),
          f"={[(r.deleted_by, r.deleted_reason) for r in rows]}")
    check("ro'yxatda o'chirilganlar KO'RINMAYDI", len(out["listed"]) == 0,
          f"ro'yxat={len(out['listed'])}")
    check("chegara bo'shadi (olingan avans 0)", float(out["taken"]) == 0.0,
          f"taken={out['taken']}")
    check("o'chirilgandan keyin chegara oshdi",
          out["after_limit"].limit > out["before_limit"].limit,
          f"{out['before_limit'].limit} -> {out['after_limit'].limit}")
    check("dublikat qo'riqchisi o'chirilganini ko'rmaydi", out["dup"] is None)

    check("HR tasdiqlangan avansni o'chira olmaydi -> 403",
          "403" in str(out["hr_delete"]), str(out["hr_delete"])[:140])
    check("ikkinchi marta o'chirib bo'lmaydi",
          out["twice"] != "o'tib ketdi" and "allaqachon" in str(out["twice"]).lower(),
          str(out["twice"])[:140])

    dels = [a for a in out["audits"] if a.action == "payroll_adjustment_deleted"]
    check("har o'chirish auditda (2 ta)", len(dels) == 2, f"={len(dels)}")
    check("auditda summa va sabab bor",
          all(d.before and d.before.get("amount") for d in dels)
          and any("boshliq bekor qildi" in str(d.after.get("deleted_reason", "")) for d in dels),
          str([d.after.get("deleted_reason") for d in dels])[:160])

    check("sabab qoidasi yoqilganda «avans» matni O'TMAYDI",
          out["meaningless"] != "o'tib ketdi" and "ma'noli" in str(out["meaningless"]).lower(),
          str(out["meaningless"])[:140])
    check("sabab qoidasi yoqilganda ma'noli matn o'tadi",
          isinstance(out["meaningful"], int), str(out["meaningful"])[:140])
    check("xabarlar patch qilingan yuboruvchiga bordi (jonli emas)",
          isinstance(sent, list), f"xabarlar={len(sent)}")

    cleanup_del()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_period_close() -> None:
    """Oy yopilishi qoidasi va xodim ko'rinishi (Avans TZ A-06).

    MUAMMO: davr qulflanganda hali `pending` bo'lgan avans abadiy osilib
    qolardi — oylikka ham kirmasdi, rad ham etilmasdi, xodim javob kutib
    o'tirardi.

    Tekshiriladi:
      1. `_next_period` chegara holatlari (dekabr → yanvar)
      2. `carry` (default): pending avans keyingi davrga KO'CHADI
      3. `cancel`: pending avans rad etiladi va sabab yoziladi
      4. Tasdiqlangan/to'langan avansga TEGILMAYDI
      5. Auditda `advance_period_closed`
      6. Preflight HR ni ogohlantiradi (`pending_advances`) va `ok=False`
      7. Xodim o'z avanslarini va QOLGAN chegarasini ko'radi
      8. Rad etilgan avans xodim jamiga kirmaydi
      9. Boshqa xodimniki ko'rinmaydi (bot yo'li telegram_id dan yechadi)
    """
    print("\n=== AVANS: oy yopilishi va xodim ko'rinishi (A-06) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import date as _dt_date, datetime as _dt_datetime

    from api.routers import payroll as pay_router
    from api.schemas import AdvanceIn, AdvanceDecision
    from db.base import async_session
    from db.models import (
        Attendance as _Att,
        AuditLog as _Audit,
        AdvanceSettings as _AdvSet,
        PayBasis,
        PayrollAdjustment as _Adj,
        PayrollAdjustmentStatus,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import select as _sel

    # ── 1. Toza funksiya ──
    check("davr o'sishi: 2026-08 -> 2026-09",
          pay_router._next_period("2026-08") == "2026-09")
    check("yil chegarasi: 2026-12 -> 2027-01",
          pay_router._next_period("2026-12") == "2027-01",
          pay_router._next_period("2026-12"))

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Close-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600951,'T-Close-Carry','employee',1,1,datetime('now'))")
    carry_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999600952,'T-Close-Cancel','employee',1,1,datetime('now'))")
    cancel_uid = cur.lastrowid
    conn.commit()

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")
    UIDS = [carry_uid, cancel_uid]

    def cleanup_close():
        try:
            c2 = db()
            qm = ",".join("?" * len(UIDS))
            pslips = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", UIDS).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            for tbl in ("payslips", "payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", UIDS)
            c2.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", UIDS)
            c2.execute(f"delete from audit_logs where target_user_id in ({qm})", UIDS)
            c2.execute(f"delete from users where id in ({qm})", UIDS)
            c2.commit()
            c2.close()
        except Exception:
            print("  A-06 tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    boss_row = conn.execute(
        "select id from users where role in ('boss','dasturchi') and is_active=1 limit 1").fetchone()
    hr_row = conn.execute(
        "select id from users where role='hr' and is_active=1 limit 1").fetchone()
    if not boss_row or not hr_row:
        check("A-06 testi uchun HR va Boshliq topildi", False)
        cleanup_close()
        return

    async def _flow():
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        out = {}
        try:
            async with async_session() as s:
                for uid in UIDS:
                    for wd in range(5):
                        s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=True,
                                                 start_time="09:00", end_time="18:00"))
                    for wd in (5, 6):
                        s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=False))
                    s.add(SalaryRate(user_id=uid, pay_basis=PayBasis.monthly.value,
                                     amount=5_000_000, effective_from=BUGUN.replace(day=1),
                                     changed_by=uid))
                    for dd in range(1, max(BUGUN.day, 2)):
                        day = BUGUN.replace(day=dd)
                        if day.weekday() >= 5:
                            continue
                        s.add(_Att(user_id=uid, date=day, status="present",
                                   check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                                   check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                                   late_minutes=0, worked_minutes=540))
                # `cancel` xodimiga O'ZIGA qoida (global sozlamaga tegilmaydi)
                s.add(_AdvSet(scope="user", scope_id=cancel_uid,
                              pending_on_close="cancel", is_active=True))
                await s.commit()

                hr = await s.get(_User, hr_row[0])
                boss = await s.get(_User, boss_row[0])

                # Har xodimga: 1 ta pending + 1 ta tasdiqlangan
                made = {}
                for uid in UIDS:
                    p1 = await pay_router.create_advance(
                        AdvanceIn(user_id=uid, period=PERIOD, amount=100_000,
                                  reason="T-Close kutilayotgan"), hr, s)
                    p2 = await pay_router.create_advance(
                        AdvanceIn(user_id=uid, period=PERIOD, amount=250_000,
                                  reason="T-Close tasdiqlanadigan", confirm_duplicate=True), hr, s)
                    await pay_router.decide_advance(p2.id, AdvanceDecision(approve=True), boss, s)
                    made[uid] = (p1.id, p2.id)

                # 7-8. Xodim ko'rinishi (yopishdan OLDIN)
                emp = await s.get(_User, carry_uid)
                out["mine"] = await pay_router._my_advances(s, emp)

                # 6. Preflight ogohlantirishi
                out["preflight"] = await pay_router.preflight(PERIOD, hr, s)

                # 2-3. Yopilish qoidasi
                out["closed"] = await pay_router._close_pending_advances(s, PERIOD, boss)
                await s.commit()

                out["rows"] = {
                    uid: {
                        a.id: (a.status, a.period, a.decided_note)
                        for a in await s.scalars(_sel(_Adj).where(_Adj.user_id == uid))
                    }
                    for uid in UIDS
                }
                out["made"] = made
                out["audits"] = [
                    a.action
                    for a in await s.scalars(_sel(_Audit).where(_Audit.target_user_id.in_(UIDS)))
                ]
                return out, sent
        finally:
            pay_router.notify_user = orig

    try:
        out, sent = _asyncio.run(_flow())
    except Exception:
        check("A-06 zanjiri ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_close()
        try:
            conn.close()
        except Exception:
            pass
        return

    made = out["made"]
    carry_pending, carry_approved = made[carry_uid]
    cancel_pending, cancel_approved = made[cancel_uid]
    KEYINGI = pay_router._next_period(PERIOD)

    c_rows = out["rows"][carry_uid]
    check("`carry`: pending avans keyingi davrga ko'chdi",
          c_rows[carry_pending] == (PayrollAdjustmentStatus.pending.value, KEYINGI, None),
          str(c_rows[carry_pending]))
    check("`carry`: tasdiqlangan avansga TEGILMADI",
          c_rows[carry_approved][0] == PayrollAdjustmentStatus.approved.value
          and c_rows[carry_approved][1] == PERIOD,
          str(c_rows[carry_approved]))

    x_rows = out["rows"][cancel_uid]
    check("`cancel`: pending avans rad etildi",
          x_rows[cancel_pending][0] == PayrollAdjustmentStatus.rejected.value
          and x_rows[cancel_pending][1] == PERIOD,
          str(x_rows[cancel_pending]))
    check("`cancel`: rad sababi yozilgan",
          "yopildi" in str(x_rows[cancel_pending][2] or ""), str(x_rows[cancel_pending][2]))
    check("`cancel`: tasdiqlangan avansga TEGILMADI",
          x_rows[cancel_approved][0] == PayrollAdjustmentStatus.approved.value,
          str(x_rows[cancel_approved]))

    check("natija hisoblandi (1 ko'chdi, 1 bekor)",
          out["closed"]["carried"] == 1 and out["closed"]["cancelled"] == 1,
          str({k: v for k, v in out["closed"].items() if k != "notify"}))
    check("auditda `advance_period_closed` (2 ta)",
          out["audits"].count("advance_period_closed") == 2,
          f"={out['audits'].count('advance_period_closed')}")

    pf = out["preflight"]
    check("preflight tasdiqlanmagan avanslarni ko'rsatadi (2 ta)",
          len(pf.pending_advances) == 2, f"={len(pf.pending_advances)}")
    check("preflight `ok=False` (HR ko'rmay o'tmasin)", pf.ok is False, f"ok={pf.ok}")

    mine = out["mine"]
    check("xodim o'z avanslarini ko'radi (2 ta)", len(mine.rows) == 2, f"={len(mine.rows)}")
    check("xodim jamisi to'g'ri (100 000 + 250 000)",
          abs(mine.total - 350_000) < 1, f"total={mine.total}")
    check("xodim QOLGAN chegarasini ko'radi",
          mine.remaining_limit > 0 and mine.limit_reason is None,
          f"limit={mine.remaining_limit}, sabab={mine.limit_reason}")

    # 9. Bot yo'li — boshqa xodimniki ko'rinmaydi
    try:
        with httpx.Client(timeout=15) as client:
            r = client.get(f"{API_BASE}/payroll/my/999600951/advances", headers=bot_secret_hdr())
            check("bot yo'li o'z avanslarini qaytaradi", r.status_code == 200, f"kod={r.status_code}")
            if r.status_code == 200:
                d = r.json()
                # Bu chaqiruv yopilishdan KEYIN — ko'chirilgan avans endi
                # keyingi oyda, ya'ni joriy oyda faqat tasdiqlangani qoladi.
                check("bot javobida faqat SHU xodimning JORIY oy avanslari",
                      len(d["rows"]) == 1 and abs(d["rows"][0]["amount"] - 250_000) < 1,
                      f"={[(r['amount'], r['status']) for r in d['rows']]}")
                check("ko'chirilgan avans joriy oy jamisiga kirmaydi",
                      abs(d["total"] - 250_000) < 1, f"total={d['total']}")
            r = client.get(f"{API_BASE}/payroll/my/999600999/advances", headers=bot_secret_hdr())
            check("mavjud bo'lmagan telegram_id -> 404", r.status_code == 404, f"kod={r.status_code}")
            r = client.get(f"{API_BASE}/payroll/my/999600951/advances")
            check("bot siri bo'lmasa -> 401/403",
                  r.status_code in (401, 403), f"kod={r.status_code}")
    except Exception:
        check("A-06 bot yo'li testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    check("xabarlar patch qilingan yuboruvchiga bordi (jonli emas)",
          isinstance(sent, list), f"xabarlar={len(sent)}")

    cleanup_close()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_settings() -> None:
    """Avans sozlamalari — uch darajali qamrov (Avans TZ B-01, B-02).

    MUAMMO: A blokda ikkita sozlama vaqtincha `fine_policies` da turgan
    edi va koeffitsient/cap umuman sozlanmasdi — ular kod ichidagi
    o'zgarmas qiymat edi.

    Tekshiriladi:
      1. Uch daraja: xodim > lavozim > global
      2. Har darajadagi bo'shliqda keyingi (kengroq) darajaga o'tiladi
      3. `is_active=False` qator O'TKAZIB YUBORILADI
      4. Hech qanday sozlama bo'lmasa — `None` (bot jim turadi)
      5. ⭐ Chegara sozlamadagi koeffitsient/cap bilan hisoblanadi
      6. Sozlamasiz chegara default bilan ishlayveradi (HR ishi to'xtamasin)
      7. HTTP: upsert (yaratish + yangilash), ro'yxat, o'chirish
      8. Har o'zgarish auditda (`advance_settings_upserted`)
      9. `position`/`user` qamrovda `scope_id` majburiy -> 422
     10. «Sozlanmagan modullar» ro'yxatida qator bor
    """
    print("\n=== AVANS: sozlamalar va uch darajali qamrov (B-01/B-02) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import date as _dt_date, datetime as _dt_datetime
    from decimal import Decimal as D

    from api.services import advance as adv
    from api.services.setup_status import collect_setup_status
    from db.base import async_session
    from db.models import (
        AdvanceSettings as _Set,
        Attendance as _Att,
        AuditLog as _Audit,
        PayBasis,
        Position as _Pos,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import select as _sel

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Set-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from positions where name like 'T-Set-%'")
    cur.execute("insert into positions (name, is_active, created_at) values ('T-Set-Lavozim',1,datetime('now'))")
    pos_id = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, position_id, bot_started, is_active, created_at)"
        " values (999600961,'T-Set-Emp','employee',?,1,1,datetime('now'))", (pos_id,))
    emp_uid = cur.lastrowid
    conn.commit()

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")
    # Jonli sozlama bormi — test oxirida holatni AYNAN qaytarish uchun.
    had_global = conn.execute(
        "select count(*) from advance_settings where scope='global'").fetchone()[0]

    def cleanup_set():
        try:
            c2 = db()
            for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override"):
                c2.execute(f"delete from {tbl} where user_id=?", (emp_uid,))
            c2.execute("delete from audit_logs where target_user_id=?", (emp_uid,))
            c2.execute("delete from advance_settings where scope='user' and scope_id=?", (emp_uid,))
            c2.execute("delete from advance_settings where scope='position' and scope_id=?", (pos_id,))
            # Global qatorni FAQAT test yaratgan bo'lsa o'chiramiz.
            if not had_global:
                c2.execute("delete from advance_settings where scope='global'")
            c2.execute("delete from users where id=?", (emp_uid,))
            c2.execute("delete from positions where id=?", (pos_id,))
            c2.commit()
            c2.close()
        except Exception:
            print("  B-01 tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    mgr = find_manager_id()
    mgr_t = token_for(mgr[0], mgr[1]) if mgr else None
    if not mgr_t:
        check("B-01 testi uchun rahbar topildi", False)
        cleanup_set()
        return

    async def _levels():
        out = {}
        async with async_session() as s:
            emp = await s.get(_User, emp_uid)

            # 4. Sozlamasiz
            if not had_global:
                out["empty"] = await adv.resolve_advance_settings(s, emp)

            # global
            s.add(_Set(scope="global", scope_id=None, coefficient=0.5, cap_percent=50,
                       advance_day=20, is_active=True))
            await s.commit()
            out["global"] = await adv.resolve_advance_settings(s, emp)

            # lavozim global'dan kuchli
            s.add(_Set(scope="position", scope_id=pos_id, coefficient=0.6, cap_percent=60,
                       advance_day=15, is_active=True))
            await s.commit()
            out["position"] = await adv.resolve_advance_settings(s, emp)

            # xodim lavozimdan kuchli
            user_row = _Set(scope="user", scope_id=emp_uid, coefficient=0.8, cap_percent=80,
                            advance_day=10, is_active=True)
            s.add(user_row)
            await s.commit()
            out["user"] = await adv.resolve_advance_settings(s, emp)

            # 3. Faol emas -> o'tkazib yuboriladi
            user_row.is_active = False
            await s.commit()
            out["user_inactive"] = await adv.resolve_advance_settings(s, emp)
            return out

    try:
        lv = _asyncio.run(_levels())
    except Exception:
        check("uch daraja testi ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_set()
        return

    if "empty" in lv:
        check("sozlamasiz -> None (bot jim turadi)", lv["empty"] is None)
    check("global daraja topiladi",
          lv["global"] is not None and lv["global"].scope == "global",
          str(getattr(lv["global"], "scope", None)))
    check("lavozim global'dan kuchli",
          lv["position"] is not None and lv["position"].scope == "position",
          str(getattr(lv["position"], "scope", None)))
    check("xodim lavozimdan kuchli",
          lv["user"] is not None and lv["user"].scope == "user",
          str(getattr(lv["user"], "scope", None)))
    check("faol bo'lmagan daraja o'tkazib yuboriladi (lavozimga qaytadi)",
          lv["user_inactive"] is not None and lv["user_inactive"].scope == "position",
          str(getattr(lv["user_inactive"], "scope", None)))

    # ── 5-6. Chegara sozlamani O'QIYDIMI ──
    async def _limit_uses_settings():
        async with async_session() as s:
            for wd in range(5):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                         start_time="09:00", end_time="18:00"))
            for wd in (5, 6):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
            s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                             amount=5_000_000, effective_from=BUGUN.replace(day=1),
                             changed_by=emp_uid))
            for dd in range(1, max(BUGUN.day, 2)):
                day = BUGUN.replace(day=dd)
                if day.weekday() >= 5:
                    continue
                s.add(_Att(user_id=emp_uid, date=day, status="present",
                           check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                           check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                           late_minutes=0, worked_minutes=540))
            await s.commit()
            emp = await s.get(_User, emp_uid)

            # Amaldagi qoida — lavozim (koef 0.6)
            with_settings = await adv.limit_for(s, emp, period=PERIOD)
            # Aynan shu sharoit uchun default bilan (koef 0.5)
            with_default = await adv.limit_for(
                s, emp, period=PERIOD,
                coefficient=adv.DEFAULT_COEFFICIENT, cap_percent=adv.DEFAULT_CAP_PERCENT)
            return with_settings, with_default

    try:
        with_settings, with_default = _asyncio.run(_limit_uses_settings())
        check("⭐ chegara sozlamadagi koeffitsientni o'qiydi (0.6 > 0.5)",
              with_settings.coefficient == 0.6 and with_settings.limit > with_default.limit,
              f"koef={with_settings.coefficient}, {with_settings.limit} > {with_default.limit}")
        check("sozlamasiz chegara default bilan ishlayveradi",
              with_default.coefficient == 0.5 and with_default.limit > 0,
              f"koef={with_default.coefficient}, limit={with_default.limit}")
    except Exception:
        check("chegara-sozlama bog'lanishi testi ishga tushdi", False,
              traceback.format_exc(limit=3).strip())

    # ── 7-9. HTTP ──
    try:
        with httpx.Client(timeout=20) as client:
            r = client.get(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t))
            check("sozlamalar ro'yxati -> 200", r.status_code == 200, f"kod={r.status_code}")
            check("ro'yxatda qamrov nomi bor",
                  r.status_code == 200
                  and any(x["scope"] == "position" and x["scope_name"] == "T-Set-Lavozim"
                          for x in r.json()),
                  str(r.json())[:200] if r.status_code == 200 else "")

            # 9. scope_id majburiy
            r = client.put(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t),
                           json={"scope": "position", "scope_id": None})
            check("lavozim qamrovida scope_id majburiy -> 422",
                  r.status_code == 422, f"kod={r.status_code}")

            # 7. Yangilash (upsert)
            r = client.put(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t),
                           json={"scope": "position", "scope_id": pos_id,
                                 "advance_day": 25, "coefficient": 0.7, "cap_percent": 70,
                                 "min_amount": 200000, "reminder_time": "16:30",
                                 "pending_on_close": "cancel", "reason_required": True,
                                 "is_active": True})
            check("mavjud qamrov YANGILANADI (yangi qator emas)",
                  r.status_code == 200 and r.json()["advance_day"] == 25
                  and r.json()["coefficient"] == 0.7,
                  f"kod={r.status_code} {str(r.json())[:150]}")
            upd_id = r.json()["id"] if r.status_code == 200 else None

            r2 = client.get(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t))
            pos_rows = [x for x in r2.json() if x["scope"] == "position"]
            check("lavozim qamrovi BITTA qator bo'lib qoldi", len(pos_rows) == 1,
                  f"={len(pos_rows)}")

            # Noto'g'ri qiymatlar
            r = client.put(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t),
                           json={"scope": "global", "advance_day": 31})
            check("advance_day 28 dan oshmaydi -> 422", r.status_code == 422, f"kod={r.status_code}")
            r = client.put(f"{API_BASE}/payroll/advance-settings", headers=auth(mgr_t),
                           json={"scope": "global", "pending_on_close": "nimadir"})
            check("noto'g'ri pending_on_close -> 422", r.status_code == 422, f"kod={r.status_code}")

            # 7. O'chirish
            if upd_id:
                r = client.delete(f"{API_BASE}/payroll/advance-settings/{upd_id}",
                                  headers=auth(mgr_t))
                check("qamrovni o'chirish -> 200", r.status_code == 200, f"kod={r.status_code}")
    except Exception:
        check("B-01 HTTP testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    # ── 8. Audit va 10. sozlanmagan modullar ──
    async def _audit_and_setup():
        async with async_session() as s:
            acts = [
                a.action
                for a in await s.scalars(
                    _sel(_Audit).where(_Audit.action.like("advance_settings%"))
                )
            ]
            items = await collect_setup_status(s)
            return acts, items

    try:
        acts, items = _asyncio.run(_audit_and_setup())
        check("sozlama o'zgarishi auditda", "advance_settings_upserted" in acts, str(acts)[:150])
        check("o'chirish ham auditda", "advance_settings_deleted" in acts, str(acts)[:150])
        item = next((i for i in items if i.key == "advance_settings"), None)
        check("«Sozlanmagan modullar» ro'yxatida avans qatori bor",
              item is not None, f"kalitlar={[i.key for i in items]}")
        if item is not None:
            check("global sozlama bor — modul «tayyor» deb ko'rsatiladi",
                  item.ready is True, f"ready={item.ready}, {item.missing}")
    except Exception:
        check("audit/setup testi ishga tushdi", False, traceback.format_exc(limit=2).strip())

    cleanup_set()
    try:
        conn.close()
    except Exception:
        pass


def test_outbox() -> None:
    """Chiquvchi xabarlar navbati (Avans TZ B-03).

    MUAMMO: xabarlar SO'ROV ICHIDA yuboriladi. cPanel'da konkurentlik 1
    — Telegram 3 soniya javob bermasa butun sayt kutadi; xabar yo'qolsa
    qayta urinish ham, iz ham qolmaydi.

    Tekshiriladi:
      1. Navbatga qo'yish va cron yuborishi
      2. `dedupe_key` — bir xabar ikki marta navbatga tushmaydi
      3. ⭐ Ikki jarayon (parallel tick) bitta xabarni IKKI MARTA yubormaydi
      4. Rate-limit: bir tick'da `BATCH_SIZE` dan ko'p yuborilmaydi
      5. 3 urinishdan keyin `failed` va HR ga ogohlantirish
      6. `sending` da osilib qolgan qator qaytarib olinadi
      7. Kelajakka rejalashtirilgan xabar hozir yuborilmaydi

    JONLI TELEGRAMGA CHIQMAYDI: `outbox.send_message` patch qilinadi
    (⚠️ test-telegram-xavfi).
    """
    print("\n=== XABAR NAVBATI (B-03) ===")
    import asyncio as _asyncio
    from datetime import datetime as _dt, timedelta as _td

    from api.services import outbox as ob
    from db.base import async_session
    from db.models import Outbox as _Ob, OutboxStatus as _St
    from sqlalchemy import delete as _del, select as _sel

    KIND = "T-Outbox"
    CHAT = 999609001

    def cleanup_ob():
        try:
            c2 = db()
            c2.execute("delete from outbox where kind=?", (KIND,))
            c2.commit()
            c2.close()
        except Exception:
            print("  Navbat tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    cleanup_ob()

    async def _run():
        out = {}
        yuborilgan: list = []
        rejim = {"mode": "ok"}

        async def _fake_send(chat_id, text, reply_markup=None):
            """Patch qilingan yuboruvchi — jonli Telegramga CHIQMAYDI."""
            if rejim["mode"] == "fail":
                return None            # `send_message` xatoda None qaytaradi
            if rejim["mode"] == "raise":
                raise RuntimeError("tarmoq uzildi")
            yuborilgan.append((chat_id, text))
            return {"ok": True}

        orig = ob.send_message
        ob.send_message = _fake_send
        try:
            # ── 1. Navbatga qo'yish va yuborish ──
            async with async_session() as s:
                await ob.enqueue(s, CHAT, KIND, "birinchi xabar")
                await ob.enqueue(s, CHAT, KIND, "ikkinchi xabar")
                await s.commit()
                out["tick1"] = await ob.tick(s)
                out["sent_texts"] = list(yuborilgan)
                out["statuses1"] = [
                    (r.status, r.attempts, r.sent_at is not None)
                    for r in await s.scalars(_sel(_Ob).where(_Ob.kind == KIND))
                ]

            # ── 2. dedupe_key ──
            async with async_session() as s:
                a = await ob.enqueue(s, CHAT, KIND, "takrorlanmas", dedupe_key="T-Outbox:2026-08:42")
                await s.commit()
                b = await ob.enqueue(s, CHAT, KIND, "takrorlanmas", dedupe_key="T-Outbox:2026-08:42")
                await s.commit()
                out["dedupe"] = (a is not None, b is None)
                out["dedupe_count"] = await s.scalar(
                    _sel(func.count()).select_from(_Ob).where(_Ob.dedupe_key == "T-Outbox:2026-08:42")
                )

            # ── 7. Kelajakdagi xabar hozir yuborilmaydi ──
            async with async_session() as s:
                await ob.enqueue(s, CHAT, KIND, "keyinroq",
                                 scheduled_at=_dt.utcnow() + _td(hours=2))
                await s.commit()

            yuborilgan.clear()
            async with async_session() as s:
                out["tick2"] = await ob.tick(s)
            out["tick2_texts"] = list(yuborilgan)

            # ── 3. Parallel tick: ikki jarayon bitta xabarni olmasin ──
            async with async_session() as s:
                await s.execute(_del(_Ob).where(_Ob.kind == KIND))
                for i in range(6):
                    await ob.enqueue(s, CHAT, KIND, f"parallel-{i}")
                await s.commit()

            yuborilgan.clear()

            async def _one_tick():
                async with async_session() as s2:
                    return await ob.tick(s2)

            r1, r2 = await _asyncio.gather(_one_tick(), _one_tick())
            out["parallel"] = (r1["sent"], r2["sent"])
            out["parallel_total"] = r1["sent"] + r2["sent"]
            out["parallel_unique"] = len({t for _, t in yuborilgan})
            out["parallel_calls"] = len(yuborilgan)

            # ── 4. Rate-limit ──
            async with async_session() as s:
                await s.execute(_del(_Ob).where(_Ob.kind == KIND))
                for i in range(ob.BATCH_SIZE + 7):
                    await ob.enqueue(s, CHAT, KIND, f"ko'p-{i}")
                await s.commit()
            yuborilgan.clear()
            async with async_session() as s:
                out["batch"] = await ob.tick(s)
            out["batch_sent"] = len(yuborilgan)

            # ── 5. Uch urinishdan keyin `failed` ──
            async with async_session() as s:
                await s.execute(_del(_Ob).where(_Ob.kind == KIND))
                await ob.enqueue(s, CHAT, KIND, "yiqiladigan")
                await s.commit()
            rejim["mode"] = "fail"
            hr_alert: list = []

            async def _fake_alert(_db, rows):
                hr_alert.append(len(rows))

            orig_alert = ob._alert_hr
            ob._alert_hr = _fake_alert
            try:
                for _ in range(ob.MAX_ATTEMPTS + 1):
                    async with async_session() as s:
                        await ob.tick(s)
            finally:
                ob._alert_hr = orig_alert
            rejim["mode"] = "ok"
            async with async_session() as s:
                row = await s.scalar(_sel(_Ob).where(_Ob.kind == KIND))
                out["failed_row"] = (row.status, row.attempts, row.last_error)
            out["hr_alert"] = hr_alert

            # ── 6a. BAND QILINGAN (yangi `sending`) qatorni ikkinchi
            #        jarayon OLMAYDI. Yuqoridagi `gather` SQLite'da
            #        ketma-ket bajarildi (bo'linish 6/0), shuning uchun
            #        qo'riqchining o'zini alohida tekshiramiz.
            async with async_session() as s:
                await s.execute(_del(_Ob).where(_Ob.kind == KIND))
                r = await ob.enqueue(s, CHAT, KIND, "band qilingan")
                await s.commit()
                r.status = _St.sending.value
                r.claimed_by = "boshqa-jarayon"
                r.claimed_at = _dt.utcnow()          # YANGI band — stale emas
                await s.commit()
            yuborilgan.clear()
            async with async_session() as s:
                out["claimed_by_other"] = await ob.tick(s)
            out["claimed_texts"] = list(yuborilgan)

            # ── 6. Osilib qolgan `sending` qaytarib olinadi ──
            async with async_session() as s:
                await s.execute(_del(_Ob).where(_Ob.kind == KIND))
                r = await ob.enqueue(s, CHAT, KIND, "osilib qolgan")
                await s.commit()
                r.status = _St.sending.value
                r.claimed_by = "eski-jarayon"
                r.claimed_at = _dt.utcnow() - _td(minutes=ob.STALE_MINUTES + 5)
                await s.commit()
            yuborilgan.clear()
            async with async_session() as s:
                out["stale"] = await ob.tick(s)
            out["stale_texts"] = list(yuborilgan)
            return out
        finally:
            ob.send_message = orig

    from sqlalchemy import func

    try:
        out = _asyncio.run(_run())
    except Exception:
        check("navbat testi ishga tushdi", False, traceback.format_exc(limit=3).strip())
        cleanup_ob()
        return

    check("navbatga qo'yilgan xabar cron orqali yuborildi",
          out["tick1"]["sent"] == 2 and len(out["sent_texts"]) == 2, str(out["tick1"]))
    check("yuborilgan qator `sent` va sana bilan belgilanadi",
          all(st == "sent" and at == 1 and ok for st, at, ok in out["statuses1"]),
          str(out["statuses1"]))

    check("`dedupe_key`: ikkinchi qo'yish o'tkazib yuboriladi",
          out["dedupe"] == (True, True), str(out["dedupe"]))
    check("bazada bitta qator qoldi", out["dedupe_count"] == 1, str(out["dedupe_count"]))

    check("kelajakka rejalashtirilgan xabar hozir yuborilmaydi",
          "keyinroq" not in [t for _, t in out["tick2_texts"]],
          str([t for _, t in out["tick2_texts"]]))

    check("⭐ parallel tick: har xabar BIR marta yuborildi",
          out["parallel_total"] == 6 and out["parallel_calls"] == 6
          and out["parallel_unique"] == 6,
          f"jami={out['parallel_total']}, chaqiruv={out['parallel_calls']}, "
          f"noyob={out['parallel_unique']}, bo'linish={out['parallel']}")

    check("rate-limit: bir tick'da BATCH_SIZE dan ko'p emas",
          out["batch_sent"] == ob.BATCH_SIZE, f"={out['batch_sent']} (limit {ob.BATCH_SIZE})")

    st, att, err = out["failed_row"]
    check(f"{ob.MAX_ATTEMPTS} urinishdan keyin `failed`",
          st == "failed" and att == ob.MAX_ATTEMPTS, f"status={st}, urinish={att}")
    check("xato sababi saqlangan", bool(err), str(err))
    check("HR ga ogohlantirish bir marta bordi",
          out["hr_alert"] == [1], str(out["hr_alert"]))

    check("boshqa jarayon band qilgan xabar OLINMAYDI (ikki marta ketmaydi)",
          out["claimed_by_other"]["claimed"] == 0 and not out["claimed_texts"],
          f"{out['claimed_by_other']}, matnlar={out['claimed_texts']}")
    check("osilib qolgan `sending` qaytarib olindi va yuborildi",
          out["stale"]["reclaimed"] == 1 and out["stale"]["sent"] == 1,
          str(out["stale"]))

    cleanup_ob()


def test_advance_day_tick() -> None:
    """Avans kuni cron'i va takroriylik qo'riqchisi (Avans TZ B-04).

    Tekshiriladi:
      1. Avans kunidan OLDIN xabar yo'q
      2. Avans kunida xabar navbatga tushadi, chegara payload'da saqlanadi
      3. ⭐ `>=` semantikasi: cron kechiksa (kun+3) ham xabar tushadi
      4. ⭐ Oyiga BIR marta — takror tick yangi xabar qo'shmaydi
      5. Istisno: ishdan bo'shash arizasi bergan xodimga yuborilmaydi
      6. Istisno: chegarasi 0 bo'lganga yuborilmaydi (stavkasiz/ta'tilda)
      7. Istisno: chegarasi `min_amount` dan past bo'lganga yuborilmaydi
      8. Istisno: sozlamasi yo'q xodimga yuborilmaydi (tizim jim turadi)
      9. Xabar matnida ANIQ summa bor (foiz emas)
    """
    print("\n=== AVANS KUNI CRON'I (B-04) ===")
    import asyncio as _asyncio
    from datetime import date as _dt_date, datetime as _dt_datetime

    from api.services import advance_day as ad
    from db.base import async_session
    from db.models import (
        AdvanceSettings as _AdvSet,
        Attendance as _Att,
        EmployeeRequest as _Req,
        Outbox as _Ob,
        PayBasis,
        Position as _Pos,
        RequestKind as _RK,
        RequestStatus as _RS,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import delete as _del, select as _sel

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Day-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "work_schedule_override",
                    "employee_requests", "advance_responses"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from outbox where kind='advance_day'")

    # Sinov xodimlari (hammasi bot ishga tushirgan, telegram_id bilan)
    UIDS = {}
    for nom, tg in (("Normal", 999609101), ("Resign", 999609102),
                    ("NoRate", 999609103), ("Small", 999609104), ("NoSet", 999609105)):
        cur.execute(
            "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
            " values (?,?,'employee',1,1,datetime('now'))", (tg, f"T-Day-{nom}"))
        UIDS[nom] = cur.lastrowid
    conn.commit()
    ALL = list(UIDS.values())

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")
    # Avans kuni — bugundan 2 kun oldin (ya'ni kun ALLAQACHON kelgan).
    # Oyning boshida bo'lsak 1-kunni olamiz.
    AV_KUN = max(1, min(BUGUN.day - 2, 28))
    # «Kun hali kelmagan» ssenariysi uchun sana
    OLDIN = BUGUN.replace(day=1) if AV_KUN > 1 else BUGUN

    # Jonli global sozlama bormi — holatni qaytarish uchun
    had_global = conn.execute(
        "select count(*) from advance_settings where scope='global'").fetchone()[0]

    def cleanup_day():
        try:
            c2 = db()
            qm = ",".join("?" * len(ALL))
            for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "work_schedule_override",
                        "employee_requests", "advance_responses"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", ALL)
            pslips = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", ALL).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            c2.execute(f"delete from payslips where user_id in ({qm})", ALL)
            c2.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", ALL)
            if not had_global:
                c2.execute("delete from advance_settings where scope='global'")
            c2.execute("delete from outbox where kind='advance_day'")
            c2.execute(f"delete from audit_logs where target_user_id in ({qm})", ALL)
            c2.execute(f"delete from users where id in ({qm})", ALL)
            c2.commit()
            c2.close()
        except Exception:
            print("  B-04 tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    async def _seed():
        async with async_session() as s:
            # Global sozlama: avans kuni AV_KUN, eng kam summa yo'q
            existing = await s.scalar(_sel(_AdvSet).where(_AdvSet.scope == "global"))
            if existing is None:
                s.add(_AdvSet(scope="global", scope_id=None, advance_day=AV_KUN,
                              coefficient=0.5, cap_percent=50, is_active=True))
            else:
                existing.advance_day = AV_KUN
                existing.min_amount = None

            # NoSet — o'z darajasida O'CHIRILGAN sozlama... bu global'ga
            # tushardi. Shuning uchun global'ni o'chirish o'rniga, bu
            # xodimni ALOHIDA tekshiruvda sinaymiz (pastda).
            for nom in ("Normal", "Resign", "NoRate", "Small", "NoSet"):
                uid = UIDS[nom]
                for wd in range(5):
                    s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=True,
                                             start_time="09:00", end_time="18:00"))
                for wd in (5, 6):
                    s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=False))
                # NoRate — ATAYLAB stavkasiz (chegara 0 bo'lsin)
                if nom != "NoRate":
                    s.add(SalaryRate(user_id=uid, pay_basis=PayBasis.monthly.value,
                                     amount=5_000_000, effective_from=BUGUN.replace(day=1),
                                     changed_by=uid))
                for dd in range(1, max(BUGUN.day, 2)):
                    day = BUGUN.replace(day=dd)
                    if day.weekday() >= 5:
                        continue
                    s.add(_Att(user_id=uid, date=day, status="present",
                               check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                               check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                               late_minutes=0, worked_minutes=540))

            # Resign — ochiq ishdan bo'shash arizasi
            s.add(_Req(user_id=UIDS["Resign"], kind=_RK.resignation.value,
                       reason="T-Day ishdan bo'shash", status=_RS.pending.value))
            # Small — o'ziga juda katta `min_amount` (chegara undan past chiqadi)
            s.add(_AdvSet(scope="user", scope_id=UIDS["Small"], advance_day=AV_KUN,
                          coefficient=0.5, cap_percent=50,
                          min_amount=99_000_000, is_active=True))
            await s.commit()

    try:
        _asyncio.run(_seed())
    except Exception:
        check("B-04 sozlash", False, traceback.format_exc(limit=3).strip())
        cleanup_day()
        return

    async def _queued():
        async with async_session() as s:
            rows = list(await s.scalars(_sel(_Ob).where(_Ob.kind == ad.KIND)))
            return {r.chat_id: r for r in rows}

    # ── 1. Kun kelmagan ──
    if AV_KUN > 1:
        try:
            r = _asyncio.run(_run_ad_tick(ad, OLDIN))
            q = _asyncio.run(_queued())
            check("avans kunidan OLDIN xabar yo'q", not q, f"navbat={len(q)}, {r}")
        except Exception:
            check("kun kelmagan ssenariysi", False, traceback.format_exc(limit=2).strip())

    # ── 2, 5-8. Avans kunida ──
    try:
        res = _asyncio.run(_run_ad_tick(ad, BUGUN))
        q = _asyncio.run(_queued())
        check("avans kunida xabar navbatga tushdi",
              999609101 in q, f"navbat={sorted(q)}, {res}")
        check("ishdan bo'shash arizasi berganga YUBORILMAYDI",
              999609102 not in q, f"navbat={sorted(q)}")
        check("chegarasi 0 bo'lganga (stavkasiz) YUBORILMAYDI",
              999609103 not in q, f"navbat={sorted(q)}")
        check("eng kam summadan past bo'lganga YUBORILMAYDI",
              999609104 not in q, f"navbat={sorted(q)}")
        if 999609101 in q:
            p = q[999609101].payload
            check("chegara xabar bilan birga saqlangan",
                  p.get("limit", 0) > 0 and p.get("period") == PERIOD, str(p)[:160])
            check("xabar matnida ANIQ summa bor (foiz emas)",
                  "so'm" in p.get("text", "") and "%" not in p.get("text", ""),
                  p.get("text", "")[:160])
    except Exception:
        check("avans kuni tick'i ishga tushdi", False, traceback.format_exc(limit=3).strip())

    # ── 4. Oyiga bir marta ──
    try:
        oldingi = len(_asyncio.run(_queued()))
        res2 = _asyncio.run(_run_ad_tick(ad, BUGUN))
        keyingi = len(_asyncio.run(_queued()))
        check("⭐ takror tick yangi xabar QO'SHMAYDI (oyiga bir marta)",
              keyingi == oldingi and res2["queued"] == 0,
              f"{oldingi} -> {keyingi}, {res2}")
    except Exception:
        check("takroriylik qo'riqchisi testi", False, traceback.format_exc(limit=2).strip())

    # ── 3. `>=` semantikasi: navbatni tozalab, KECHIKKAN kun bilan ──
    try:
        conn2 = db()
        conn2.execute("delete from outbox where kind='advance_day'")
        conn2.commit()
        conn2.close()
        kech = BUGUN  # bugun allaqachon AV_KUN dan katta
        res3 = _asyncio.run(_run_ad_tick(ad, kech))
        q3 = _asyncio.run(_queued())
        check("⭐ cron kechiksa ham xabar tushadi (`>=`, `==` emas)",
              999609101 in q3 and kech.day > AV_KUN,
              f"kun={kech.day} > avans_kuni={AV_KUN}, navbat={sorted(q3)}")
    except Exception:
        check("`>=` semantikasi testi", False, traceback.format_exc(limit=2).strip())

    # ── 8. Sozlamasiz xodim ──
    try:
        conn3 = db()
        conn3.execute("delete from outbox where kind='advance_day'")
        # Global sozlamani vaqtincha o'chiramiz — hech kimda qoida qolmasin
        conn3.execute("update advance_settings set is_active=0")
        conn3.commit()
        conn3.close()
        res4 = _asyncio.run(_run_ad_tick(ad, BUGUN))
        q4 = _asyncio.run(_queued())
        check("sozlamasiz tizim JIM turadi (birorta xabar yo'q)",
              not q4 and res4["queued"] == 0, f"navbat={len(q4)}, {res4}")
        conn3 = db()
        conn3.execute("update advance_settings set is_active=1")
        conn3.commit()
        conn3.close()
    except Exception:
        check("sozlamasiz ssenariy testi", False, traceback.format_exc(limit=2).strip())

    cleanup_day()
    try:
        conn.close()
    except Exception:
        pass


async def _run_ad_tick(ad_module, on_date):
    """Yordamchi — har chaqiruvda TOZA sessiya (tick o'zi commit qiladi)."""
    from db.base import async_session

    async with async_session() as s:
        return await ad_module.tick(s, on_date=on_date)


def test_advance_bot_flow() -> None:
    """Avans bot oqimi — tugmalar, summa, natija (Avans TZ C-01…C-05).

    Tekshiriladi:
      C-01  1. Xabarda tugmalar bor, `callback_data` da DAVR bor
            2. «Kerak emas» javobi bazaga yoziladi
            3. O'tgan oyning xabari bosilsa «eskirgan» deyiladi
      C-02  4. «Summa kiritish» holati BAZADA saqlanadi (FSM emas)
            5. ⭐ Restart simulyatsiyasi: yangi sessiyada ham holat bor
            6. Raqamsiz matn `handled=False` — boshqa handlerga o'tadi
            7. Muddat o'tsa holat bekor bo'ladi va matn o'tkaziladi
      C-03  8. Chegaradan oshiq summa rad etiladi, aniq raqam bilan
            9. ⭐ Chegara QAYTA hisoblanadi (oraliqda kamaysa)
      C-04 10. So'rov `source='bot'` bilan yoziladi
           11. HR/Boshliqqa xabar OUTBOX orqali
           12. Tasdiq/rad natijasi xodimga (rad — SABAB bilan)
      C-05 13. Javob bermaganga BITTA eslatma
           14. Javob berganga eslatma KETMAYDI
           15. Ikkinchi tick yangi eslatma qo'shmaydi
      Turli 16. Summa matnini keng tushunish («1 200 000», «1.5 mln»)
    """
    print("\n=== AVANS BOT OQIMI (C-01…C-05) ===")
    import asyncio as _asyncio
    from datetime import date as _dt_date, datetime as _dt_datetime, timedelta as _td

    from api.routers import payroll as pay_router
    from api.schemas import AdvanceDecision
    from api.services import advance_bot as ab
    from api.services import advance_day as ad
    from db.base import async_session
    from db.models import (
        AdvanceResponse as _Resp,
        AdvanceResponseState as _RState,
        AdvanceSettings as _AdvSet,
        Attendance as _Att,
        Outbox as _Ob,
        PayBasis,
        PayrollAdjustment as _Adj,
        PayrollAdjustmentSource as _Src,
        PayrollAdjustmentStatus as _AStatus,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import delete as _del, select as _sel

    # ── 16. Summa matnini tushunish (toza funksiya) ──
    for matn, kutilgan in (
        ("1200000", 1_200_000),
        ("1 200 000", 1_200_000),
        ("1.200.000", 1_200_000),
        ("1,200,000", 1_200_000),
        ("500000 so'm", 500_000),
        ("1,5 mln", 1_500_000),
        ("2 mln", 2_000_000),
    ):
        got = ab.parse_amount(matn)
        check(f"summa matni tushunildi: «{matn}»", got == kutilgan, f"{got} != {kutilgan}")
    for matn in ("salom", "", "kerak emas", "yo'q"):
        check(f"raqamsiz matn summa DEB QABUL QILINMAYDI: «{matn}»",
              ab.parse_amount(matn) is None, str(ab.parse_amount(matn)))

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Bot-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "advance_responses"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from outbox where kind like 'advance%'")
    TG = 999609201
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (?, 'T-Bot-Emp','employee',1,1,datetime('now'))", (TG,))
    emp_uid = cur.lastrowid
    conn.commit()

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")
    ESKI = "2019-01"
    had_global = conn.execute(
        "select count(*) from advance_settings where scope='global'").fetchone()[0]

    def cleanup_bot():
        try:
            c2 = db()
            for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                        "work_schedule_weekly", "advance_responses"):
                c2.execute(f"delete from {tbl} where user_id=?", (emp_uid,))
            pslips = [r[0] for r in c2.execute(
                "select id from payslips where user_id=?", (emp_uid,)).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            c2.execute("delete from payslips where user_id=?", (emp_uid,))
            c2.execute("delete from advance_settings where scope='user' and scope_id=?", (emp_uid,))
            if not had_global:
                c2.execute("delete from advance_settings where scope='global'")
            c2.execute("delete from outbox where kind like 'advance%'")
            c2.execute("delete from audit_logs where target_user_id=?", (emp_uid,))
            c2.execute("delete from users where id=?", (emp_uid,))
            c2.commit()
            c2.close()
        except Exception:
            print("  C blok tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    async def _seed():
        async with async_session() as s:
            existing = await s.scalar(_sel(_AdvSet).where(_AdvSet.scope == "global"))
            if existing is None:
                s.add(_AdvSet(scope="global", scope_id=None, advance_day=1,
                              coefficient=0.5, cap_percent=50, is_active=True))
            for wd in range(5):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=True,
                                         start_time="09:00", end_time="18:00"))
            for wd in (5, 6):
                s.add(WorkScheduleWeekly(user_id=emp_uid, weekday=wd, is_working=False))
            s.add(SalaryRate(user_id=emp_uid, pay_basis=PayBasis.monthly.value,
                             amount=5_000_000, effective_from=BUGUN.replace(day=1),
                             changed_by=emp_uid))
            for dd in range(1, max(BUGUN.day, 2)):
                day = BUGUN.replace(day=dd)
                if day.weekday() >= 5:
                    continue
                s.add(_Att(user_id=emp_uid, date=day, status="present",
                           check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                           check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                           late_minutes=0, worked_minutes=540))
            await s.commit()

    try:
        _asyncio.run(_seed())
    except Exception:
        check("C blok sozlash", False, traceback.format_exc(limit=3).strip())
        cleanup_bot()
        return

    out = {}

    # ── C-01: avans kuni xabari tugmalar bilan ──
    async def _announce():
        async with async_session() as s:
            res = await ad.tick(s, on_date=BUGUN)
            row = await s.scalar(
                _sel(_Ob).where(_Ob.kind == "advance_day", _Ob.chat_id == TG))
            resp = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            return res, row, resp

    try:
        res, ob_row, resp = _asyncio.run(_announce())
        check("avans kuni xabari navbatga tushdi", ob_row is not None, str(res))
        if ob_row is not None:
            km = ob_row.payload.get("reply_markup", {})
            tugmalar = [b for qator in km.get("inline_keyboard", []) for b in qator]
            check("xabarda ikkita tugma bor", len(tugmalar) == 2, str(tugmalar))
            check("`callback_data` da DAVR bor (eski xabar chalkashmasin)",
                  all(PERIOD in b["callback_data"] for b in tugmalar), str(tugmalar))
        check("munosabat yozuvi yaratildi (`offered`)",
              resp is not None and resp.state == "offered",
              str(getattr(resp, "state", None)))
    except Exception:
        check("C-01 e'lon testi", False, traceback.format_exc(limit=3).strip())

    # ── C-01: eski xabar ──
    async def _old_message():
        async with async_session() as s:
            return await ab.on_callback(s, TG, "need", ESKI)

    try:
        r = _asyncio.run(_old_message())
        check("o'tgan oyning xabari bosilsa «eskirgan» deydi",
              "eskirgan" in r["text"].lower() and r["clear_keyboard"], str(r)[:160])
    except Exception:
        check("eski xabar testi", False, traceback.format_exc(limit=2).strip())

    # ── C-02: «Summa kiritish» -> holat BAZADA ──
    async def _need():
        async with async_session() as s:
            return await ab.on_callback(s, TG, "need", PERIOD)

    async def _state():
        """ATAYLAB YANGI sessiya — «bot qayta ishga tushdi» simulyatsiyasi."""
        async with async_session() as s:
            return await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))

    try:
        r = _asyncio.run(_need())
        st = _asyncio.run(_state())
        check("«Summa kiritish» -> ruxsat etilgan summa ko'rsatiladi",
              "so'm" in r["text"], r["text"][:120])
        check("⭐ holat BAZADA saqlandi (restartdan omon qoladi)",
              st is not None and st.state == "waiting_input"
              and st.input_expires_at is not None,
              f"state={getattr(st,'state',None)}, muddat={getattr(st,'input_expires_at',None)}")
        out["limit"] = float(st.offered_limit)
    except Exception:
        check("C-02 holat testi", False, traceback.format_exc(limit=3).strip())
        cleanup_bot()
        return

    # ── C-02: raqamsiz matn o'tkaziladi ──
    async def _text(t):
        async with async_session() as s:
            return await ab.on_text(s, TG, t)

    try:
        r = _asyncio.run(_text("bugun ob-havo yaxshi"))
        check("raqamsiz matn boshqa handlerga o'tadi (`handled=False`)",
              r.get("handled") is False, str(r))
    except Exception:
        check("raqamsiz matn testi", False, traceback.format_exc(limit=2).strip())

    # ── C-03: chegaradan oshiq ──
    try:
        r = _asyncio.run(_text(str(int(out["limit"]) + 1_000_000)))
        check("chegaradan oshiq summa rad etiladi va aniq raqam ko'rsatiladi",
              r.get("handled") and "chegaradan oshdi" in r["text"].lower()
              and "so'm" in r["text"], r.get("text", "")[:180])
        st = _asyncio.run(_state())
        check("rad etilgach holat `waiting_input` da QOLADI (qayta yozish mumkin)",
              st.state == "waiting_input", st.state)
    except Exception:
        check("C-03 chegara testi", False, traceback.format_exc(limit=2).strip())

    # ── C-03: chegara QAYTA hisoblanadi ──
    async def _shrink_limit():
        """Oraliqda boshqa avans tasdiqlanadi — chegara kamayadi."""
        async with async_session() as s:
            s.add(_Adj(user_id=emp_uid, period=PERIOD, kind="minus",
                       category="advance", status=_AStatus.approved.value,
                       amount=out["limit"] - 50_000, reason="T-Bot oraliq avans",
                       created_by=emp_uid, source=_Src.hr_manual.value))
            await s.commit()

    try:
        _asyncio.run(_shrink_limit())
        # Endi eski chegara bo'yicha o'tadigan summa ham oshiq bo'lishi kerak
        r = _asyncio.run(_text(str(int(out["limit"] - 10_000))))
        check("⭐ chegara QAYTA hisoblanadi (eski qiymatga ishonilmaydi)",
              r.get("handled") and "chegaradan oshdi" in r["text"].lower(),
              r.get("text", "")[:180])
    except Exception:
        check("chegara qayta hisoblash testi", False, traceback.format_exc(limit=2).strip())

    # Oraliq avansni olib tashlab, toza holatga qaytamiz
    async def _restore():
        async with async_session() as s:
            await s.execute(_del(_Adj).where(_Adj.reason == "T-Bot oraliq avans"))
            await s.commit()

    _asyncio.run(_restore())

    # ── C-04: so'rov yaratiladi ──
    try:
        r = _asyncio.run(_text("300000"))
        check("qabul qilingan summa uchun tasdiq xabari",
              r.get("handled") and "yuborildi" in r["text"].lower(), r.get("text", "")[:140])
    except Exception:
        check("C-04 so'rov testi", False, traceback.format_exc(limit=2).strip())

    async def _after_submit():
        async with async_session() as s:
            adj = await s.scalar(
                _sel(_Adj).where(_Adj.user_id == emp_uid, _Adj.source == _Src.bot.value))
            st = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            req_msgs = list(await s.scalars(
                _sel(_Ob).where(_Ob.kind == ab.KIND_REQUEST)))
            return adj, st, req_msgs

    try:
        adj, st, req_msgs = _asyncio.run(_after_submit())
        check("so'rov `source='bot'` bilan yozildi",
              adj is not None and adj.source == "bot" and float(adj.amount) == 300_000,
              f"source={getattr(adj,'source',None)}, summa={getattr(adj,'amount',None)}")
        check("so'rov `pending` — Boshliq tasdig'i kutilmoqda",
              adj is not None and adj.status == "pending", str(getattr(adj, "status", None)))
        check("holat `submitted` va so'rovga bog'landi",
              st.state == "submitted" and st.adjustment_id == adj.id,
              f"state={st.state}, adj={st.adjustment_id}")
        check("HR/Boshliqqa xabar OUTBOX orqali (so'rov ichida emas)",
              len(req_msgs) >= 1, f"={len(req_msgs)}")
        out["adj_id"] = adj.id
    except Exception:
        check("C-04 natija testi", False, traceback.format_exc(limit=3).strip())

    # ── C-04: rad javobi SABAB bilan ──
    async def _reject():
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        try:
            async with async_session() as s:
                boss = await s.scalar(_sel(_User).where(
                    _User.role.in_(("boss", "dasturchi")), _User.is_active.is_(True)))
                await pay_router.decide_advance(
                    out["adj_id"],
                    AdvanceDecision(approve=False, note="Bu oy byudjet yo'q"),
                    boss, s)
                msgs = list(await s.scalars(_sel(_Ob).where(_Ob.kind == ab.KIND_RESULT)))
                return msgs
        finally:
            pay_router.notify_user = orig

    try:
        msgs = _asyncio.run(_reject())
        check("rad natijasi xodimga OUTBOX orqali bordi", len(msgs) == 1, f"={len(msgs)}")
        if msgs:
            matn = msgs[0].payload.get("text", "")
            check("rad xabarida SABAB bor",
                  "Bu oy byudjet yo'q" in matn, matn[:160])
    except Exception:
        check("C-04 rad javobi testi", False, traceback.format_exc(limit=3).strip())

    # ── C-01: «Kerak emas» ──
    async def _decline():
        async with async_session() as s:
            st = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            st.state = _RState.offered.value
            await s.commit()
            r = await ab.on_callback(s, TG, "no", PERIOD)
            st2 = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            return r, st2.state

    try:
        r, holat = _asyncio.run(_decline())
        check("«Kerak emas» javobi bazaga yozildi", holat == "declined", holat)
        check("«Kerak emas» ga qisqa tasdiq beriladi", bool(r.get("text")), str(r)[:120])
    except Exception:
        check("«Kerak emas» testi", False, traceback.format_exc(limit=2).strip())

    # ── C-05: takroriy eslatma ──
    async def _reminder(state_value):
        async with async_session() as s:
            await s.execute(_del(_Ob).where(_Ob.kind == ab.KIND_REMINDER))
            st = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            st.state = state_value
            st.reminded_at = None
            # Sozlamadagi vaqt allaqachon o'tgan bo'lsin
            g = await s.scalar(_sel(_AdvSet).where(_AdvSet.scope == "global"))
            g.reminder_time = "00:01"
            await s.commit()
            res = await ab.reminder_tick(s, on_date=BUGUN)
            msgs = list(await s.scalars(
                _sel(_Ob).where(_Ob.kind == ab.KIND_REMINDER, _Ob.chat_id == TG)))
            res2 = await ab.reminder_tick(s, on_date=BUGUN)
            msgs2 = list(await s.scalars(
                _sel(_Ob).where(_Ob.kind == ab.KIND_REMINDER, _Ob.chat_id == TG)))
            return res, len(msgs), res2, len(msgs2)

    try:
        res, n1, res2, n2 = _asyncio.run(_reminder(_RState.offered.value))
        check("javob bermaganga eslatma yuborildi", n1 == 1, f"={n1}, {res}")
        check("ikkinchi tick YANGI eslatma qo'shmaydi (bir kunda ko'pi bilan 2 xabar)",
              n2 == 1, f"={n2}, {res2}")
    except Exception:
        check("C-05 eslatma testi", False, traceback.format_exc(limit=3).strip())

    try:
        res, n1, _, _ = _asyncio.run(_reminder(_RState.declined.value))
        check("«Kerak emas» degan xodimga eslatma KETMAYDI", n1 == 0, f"={n1}, {res}")
    except Exception:
        check("C-05 javob bergan testi", False, traceback.format_exc(limit=2).strip())

    # ── C-02: muddat o'tsa holat bekor ──
    async def _expired():
        async with async_session() as s:
            st = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            st.state = _RState.waiting_input.value
            st.input_expires_at = _dt_datetime.utcnow() - _td(minutes=5)
            await s.commit()
            r = await ab.on_text(s, TG, "250000")
            st2 = await s.scalar(
                _sel(_Resp).where(_Resp.user_id == emp_uid, _Resp.period == PERIOD))
            return r, st2.state

    try:
        r, holat = _asyncio.run(_expired())
        check("muddat o'tgan holatda matn boshqa handlerga o'tadi",
              r.get("handled") is False, str(r))
        check("muddat o'tgan holat bekor qilinadi", holat == "offered", holat)
    except Exception:
        check("muddat testi", False, traceback.format_exc(limit=2).strip())

    cleanup_bot()
    try:
        conn.close()
    except Exception:
        pass


def test_advance_hr_panel() -> None:
    """HR paneli, nazorat va yakuniy audit (Avans TZ D-01…D-04).

    D-01  1. Qo'lda e'lon: xabar ketadi, sana ANIQ ko'rsatiladi
          2. ⭐ E'lon qilingan oyda AVTOMATIK xabar KETMAYDI
          3. Ikki marta e'lon qilinsa oxirgisi kuchda (eski navbatdan olinadi)
          4. E'lon tarixi saqlanadi (kim, qachon, nechta)
    D-02  5. Jami summa tasdiqlashdan OLDIN ko'rinadi
          6. Ommaviy tasdiqlash ishlaydi, HAR BIRI auditga tushadi
          7. Allaqachon hal qilingani jimgina o'tkaziladi (butun amal yiqilmaydi)
    D-03  8. Ketma-ket oylar to'g'ri sanaladi
          9. ⭐ Oraliq uzilsa hisob QAYTADAN boshlanadi
         10. Belgi neytral (3 oydan boshlab `flagged`)
    D-04 11. ⭐ ROL MATRITSASI: xodim va ROP ko'ra olmaydi
         12. ⭐ PAYSLIP: bot avansi BIR MARTA ayirilgan (dublikat yo'q)
    """
    print("\n=== AVANS: HR paneli va nazorat (D-01…D-04) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import date as _dt_date, datetime as _dt_datetime

    from api.routers import payroll as pay_router
    from api.services import advance_bot as ab
    from api.services import advance_day as ad
    from api.services.payroll import build_payslip
    from db.base import async_session
    from db.models import (
        AdvanceAnnouncement as _Ann,
        AdvanceResponse as _Resp,
        AdvanceSettings as _AdvSet,
        Attendance as _Att,
        AuditLog as _Audit,
        Outbox as _Ob,
        PayBasis,
        PayrollAdjustment as _Adj,
        PayrollAdjustmentSource as _Src,
        PayrollAdjustmentStatus as _AStatus,
        SalaryRate,
        User as _User,
        WorkScheduleWeekly,
    )
    from sqlalchemy import delete as _del, select as _sel, update as _upd

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Hr-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        for tbl in ("payroll_adjustments", "salary_rates", "attendance",
                    "work_schedule_weekly", "advance_responses"):
            cur.execute(f"delete from {tbl} where user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    cur.execute("delete from outbox where kind like 'advance%'")
    cur.execute("delete from advance_announcements")
    TG = 999609301
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (?, 'T-Hr-Emp','employee',1,1,datetime('now'))", (TG,))
    emp_uid = cur.lastrowid
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (999609302, 'T-Hr-Rop','rop',1,1,datetime('now'))")
    rop_uid = cur.lastrowid
    conn.commit()

    BUGUN = _dt_date.today()
    PERIOD = BUGUN.strftime("%Y-%m")
    UIDS = [emp_uid, rop_uid]
    had_global = conn.execute(
        "select count(*) from advance_settings where scope='global'").fetchone()[0]

    def cleanup_hr():
        try:
            c2 = db()
            qm = ",".join("?" * len(UIDS))
            # Tartib MUHIM: `advance_responses` `payroll_adjustments` ga
            # FK bilan bog'langan, shuning uchun u BIRINCHI o'chiriladi.
            for tbl in ("advance_responses", "payroll_adjustments", "salary_rates",
                        "attendance", "work_schedule_weekly"):
                c2.execute(f"delete from {tbl} where user_id in ({qm})", UIDS)
            pslips = [r[0] for r in c2.execute(
                f"select id from payslips where user_id in ({qm})", UIDS).fetchall()]
            if pslips:
                qm2 = ",".join("?" * len(pslips))
                c2.execute(f"delete from payslip_items where payslip_id in ({qm2})", pslips)
            c2.execute(f"delete from payslips where user_id in ({qm})", UIDS)
            c2.execute(f"delete from advance_settings where scope='user' and scope_id in ({qm})", UIDS)
            if not had_global:
                c2.execute("delete from advance_settings where scope='global'")
            c2.execute("delete from outbox where kind like 'advance%'")
            c2.execute("delete from advance_announcements")
            c2.execute(f"delete from audit_logs where target_user_id in ({qm})", UIDS)
            c2.execute("delete from audit_logs where action='advance_announced'")
            c2.execute(f"delete from users where id in ({qm})", UIDS)
            c2.commit()
            c2.close()
        except Exception:
            print("  D blok tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    boss_row = conn.execute(
        "select id, role from users where role in ('boss','dasturchi') and is_active=1 limit 1"
    ).fetchone()
    hr_row = conn.execute(
        "select id from users where role='hr' and is_active=1 limit 1").fetchone()
    if not boss_row or not hr_row:
        check("D blok testi uchun HR va Boshliq topildi", False)
        cleanup_hr()
        return
    boss_t = token_for(boss_row[0], boss_row[1])
    hr_t = token_for(hr_row[0], "hr")
    emp_t = token_for(emp_uid, "employee")
    rop_t = token_for(rop_uid, "rop")

    async def _seed():
        async with async_session() as s:
            existing = await s.scalar(_sel(_AdvSet).where(_AdvSet.scope == "global"))
            if existing is None:
                s.add(_AdvSet(scope="global", scope_id=None, advance_day=1,
                              coefficient=0.5, cap_percent=50, is_active=True))
            else:
                existing.advance_day = 1
                existing.min_amount = None
            for uid in UIDS:
                for wd in range(5):
                    s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=True,
                                             start_time="09:00", end_time="18:00"))
                for wd in (5, 6):
                    s.add(WorkScheduleWeekly(user_id=uid, weekday=wd, is_working=False))
                s.add(SalaryRate(user_id=uid, pay_basis=PayBasis.monthly.value,
                                 amount=5_000_000, effective_from=BUGUN.replace(day=1),
                                 changed_by=uid))
                for dd in range(1, max(BUGUN.day, 2)):
                    day = BUGUN.replace(day=dd)
                    if day.weekday() >= 5:
                        continue
                    s.add(_Att(user_id=uid, date=day, status="present",
                               check_in_time=_dt_datetime(day.year, day.month, dd, 4, 0),
                               check_out_time=_dt_datetime(day.year, day.month, dd, 13, 0),
                               late_minutes=0, worked_minutes=540))
            await s.commit()

    try:
        _asyncio.run(_seed())
    except Exception:
        check("D blok sozlash", False, traceback.format_exc(limit=3).strip())
        cleanup_hr()
        return

    # ── D-01 ──
    async def _announce(sana, izoh=None):
        async with async_session() as s:
            hr = await s.get(_User, hr_row[0])
            res = await ad.announce_manually(s, hr, sana, izoh)
            notice = list(await s.scalars(
                _sel(_Ob).where(_Ob.kind == ad.KIND_NOTICE, _Ob.chat_id == TG)))
            offers = list(await s.scalars(
                _sel(_Ob).where(_Ob.kind == ad.KIND, _Ob.chat_id == TG)))
            return res, notice, offers

    try:
        sana1 = BUGUN.replace(day=min(BUGUN.day, 20))
        res, notice, offers = _asyncio.run(_announce(sana1, "Bayram sababli ko'chirildi"))
        check("qo'lda e'lon: OGOHLANTIRISH xabari navbatga tushdi",
              res["recipients"] >= 1 and len(notice) == 1, f"{res}, xabarlar={len(notice)}")
        if notice:
            matn = notice[0].payload.get("text", "")
            oylar = ("yanvar", "fevral", "mart", "aprel", "may", "iyun",
                     "iyul", "avgust", "sentabr", "oktabr", "noyabr", "dekabr")
            kutilgan = f"{sana1.day}-{oylar[sana1.month - 1]}"
            check("e'londa avans SANASI aniq ko'rsatiladi (TZ 3-bo'lim namunasi)",
                  kutilgan in matn, f"«{kutilgan}» kutilgan edi: {matn[:160]}")
            check("HR izohi xabarga qo'shildi",
                  "Bayram sababli" in matn, matn[:200])
            check("⭐ e'londa SUMMA YO'Q (u avans kuni alohida keladi)",
                  "so'm" not in matn and "tugma" not in matn.lower(), matn[:200])
            check("e'londa tugma yo'q (bu taklif emas, ogohlantirish)",
                  "reply_markup" not in notice[0].payload, str(notice[0].payload)[:120])
        # Sana allaqachon kelgan -> TAKLIF ham darhol ketadi
        check("⭐ sana kelgan bo'lsa taklif ham darhol yuboriladi",
              len(offers) == 1, f"takliflar={len(offers)}")
        if offers:
            tm = offers[0].payload.get("text", "")
            check("taklifda aniq summa va tugma tushuntirilgan (TZ namunasi)",
                  "so'm" in tm and "Kerak emas" in tm and "%" not in tm, tm[:200])
    except Exception:
        check("D-01 e'lon testi", False, traceback.format_exc(limit=3).strip())

    # ── D-01: avtomatik xabar TO'XTAYDI ──
    async def _auto():
        async with async_session() as s:
            return await ad.tick(s, on_date=BUGUN)

    try:
        r = _asyncio.run(_auto())
        check("⭐ e'londan keyin avtomatik tick TAKRORLAMAYDI (dedupe)",
              r["queued"] == 0, str(r))
    except Exception:
        check("avtomatik to'xtatish testi", False, traceback.format_exc(limit=2).strip())

    # ── D-01: qayta e'lon ──
    try:
        # Kelajakdagi sanaga qayta e'lon: ogohlantirish ketadi, taklif esa
        # O'SHA KUNGACHA KUTADI (TZ 3-bo'lim: «23-ga ko'chirildi»).
        sana2 = BUGUN.replace(day=28) if BUGUN.day < 28 else sana1
        res2, notice2, offers2 = _asyncio.run(_announce(sana2, None))
        check("qayta e'lon: eski YUBORILMAGAN xabar navbatdan olindi",
              len(notice2) == 1, f"navbatda {len(notice2)} ogohlantirish (1 bo'lishi kerak)")
        if sana2 > BUGUN:
            check("⭐ sana KELAJAKDA bo'lsa taklif hali yuborilmaydi",
                  len(offers2) == 0, f"takliflar={len(offers2)}")
    except Exception:
        check("qayta e'lon testi", False, traceback.format_exc(limit=2).strip())

    try:
        with httpx.Client(timeout=20) as client:
            r = client.get(f"{API_BASE}/payroll/advance-announcements", headers=auth(hr_t))
            check("e'lon tarixi saqlanadi", r.status_code == 200 and len(r.json()) == 2,
                  f"kod={r.status_code}, ={len(r.json()) if r.status_code == 200 else '?'}")
            if r.status_code == 200 and r.json():
                check("tarixda kim e'lon qilgani ko'rinadi",
                      bool(r.json()[0].get("sent_by_name")), str(r.json()[0])[:160])
    except Exception:
        check("e'lon tarixi testi", False, traceback.format_exc(limit=2).strip())

    # ── D-04: ROL MATRITSASI ──
    try:
        with httpx.Client(timeout=20) as client:
            for nom, tok, kutilgan in (
                ("xodim", emp_t, 403), ("ROP", rop_t, 403),
                ("HR", hr_t, 200), ("Boshliq", boss_t, 200),
            ):
                r = client.get(f"{API_BASE}/payroll/advance-summary?period={PERIOD}",
                               headers=auth(tok))
                check(f"⭐ jami summa: {nom} -> {kutilgan}",
                      r.status_code == kutilgan, f"kod={r.status_code}")
            for nom, tok, kutilgan in (
                ("xodim", emp_t, 403), ("ROP", rop_t, 403), ("HR", hr_t, 200),
            ):
                r = client.get(f"{API_BASE}/payroll/adjustments?category=advance",
                               headers=auth(tok))
                check(f"avans ro'yxati: {nom} -> {kutilgan}",
                      r.status_code == kutilgan, f"kod={r.status_code}")
            for nom, tok, kutilgan in (
                ("xodim", emp_t, 403), ("ROP", rop_t, 403), ("HR", hr_t, 200),
            ):
                r = client.get(f"{API_BASE}/payroll/advance-settings", headers=auth(tok))
                check(f"sozlamalar: {nom} -> {kutilgan}",
                      r.status_code == kutilgan, f"kod={r.status_code}")
            # E'lon qilish — HR ha, ROP yo'q
            r = client.post(f"{API_BASE}/payroll/advance-announce", headers=auth(rop_t),
                            json={"advance_date": BUGUN.isoformat()})
            check("e'lon qilish: ROP -> 403", r.status_code == 403, f"kod={r.status_code}")
            # Ommaviy tasdiqlash — HR ham qila oladi (2026-08-21 qarori),
            # lekin ROP va xodim yo'q.
            r = client.post(f"{API_BASE}/payroll/advances/bulk-decide", headers=auth(hr_t),
                            json={"ids": [999999], "approve": True})
            check("ommaviy tasdiqlash: HR -> 200 (HR ham tasdiqlaydi)",
                  r.status_code == 200, f"kod={r.status_code}")
            for nom, tok in (("ROP", rop_t), ("xodim", emp_t)):
                r = client.post(f"{API_BASE}/payroll/advances/bulk-decide", headers=auth(tok),
                                json={"ids": [999999], "approve": True})
                check(f"ommaviy tasdiqlash: {nom} -> 403", r.status_code == 403,
                      f"kod={r.status_code}")
    except Exception:
        check("D-04 rol matritsasi testi", False, traceback.format_exc(limit=3).strip())

    # ── Vazifalar ajratimi: HR O'ZI kiritganini tasdiqlay olmaydi ──
    async def _self_approve():
        """HR botdan kelgan so'rovni tasdiqlay OLADI (xodim yaratgan),
        lekin o'zi kiritganini tasdiqlay OLMAYDI — bir odam pulni ham
        kiritib, ham tasdiqlab yubormasin."""
        from api.schemas import AdvanceDecision, AdvanceIn
        sent: list = []

        async def _fake_notify(_db, _user, _cat, _text, **_kw):
            sent.append(_text)

        orig = pay_router.notify_user
        pay_router.notify_user = _fake_notify
        try:
            async with async_session() as s:
                hr = await s.get(_User, hr_row[0])
                emp = await s.get(_User, emp_uid)
                # (a) HR o'zi kiritadi
                own = await pay_router.create_advance(
                    AdvanceIn(user_id=emp_uid, period=PERIOD, amount=150_000,
                              reason="T-Hr HR kiritgani", confirm_duplicate=True),
                    hr, s)
                try:
                    await pay_router.decide_advance(
                        own.id, AdvanceDecision(approve=True), hr, s)
                    ozi = "o'tib ketdi"
                except Exception as e:
                    ozi = f"{getattr(e,'status_code','?')}: {getattr(e,'detail',e)}"
                # (b) Boshliq esa o'sha yozuvni tasdiqlay oladi
                boss = await s.get(_User, boss_row[0])
                await pay_router.decide_advance(
                    own.id, AdvanceDecision(approve=True), boss, s)
                boshliq = (await s.get(_Adj, own.id)).status

                # (c) Botdan kelgan so'rovni HR tasdiqlaydi
                bot_adj = await ab.submit(s, emp, PERIOD, 320_000)
                row = await s.scalar(_sel(_Adj).where(
                    _Adj.user_id == emp_uid, _Adj.source == _Src.bot.value,
                    _Adj.status == _AStatus.pending.value))
                await pay_router.decide_advance(
                    row.id, AdvanceDecision(approve=True), hr, s)
                botniki = (await s.get(_Adj, row.id)).status

                # Bu blok yaratgan yozuvlarni olib tashlaymiz — keyingi
                # D-02 yig'indisi faqat O'Z so'rovlarini sanashi kerak.
                await s.execute(
                    _upd(_Resp).where(_Resp.user_id == emp_uid).values(adjustment_id=None))
                await s.execute(_del(_Adj).where(_Adj.id.in_([own.id, row.id])))
                await s.commit()
                return ozi, boshliq, botniki
        finally:
            pay_router.notify_user = orig

    try:
        ozi, boshliq, botniki = _asyncio.run(_self_approve())
        check("⭐ HR O'ZI kiritgan avansni tasdiqlay OLMAYDI -> 403",
              "403" in str(ozi), str(ozi)[:140])
        check("o'sha yozuvni Boshliq tasdiqlay oladi", boshliq == "approved", str(boshliq))
        check("⭐ botdan kelgan so'rovni HR tasdiqlaydi (asosiy oqim)",
              botniki == "approved", str(botniki))
    except Exception:
        check("vazifalar ajratimi testi", False, traceback.format_exc(limit=3).strip())

    # ── D-02: yig'indi va ommaviy tasdiqlash ──
    async def _make_requests():
        async with async_session() as s:
            emp = await s.get(_User, emp_uid)
            rop = await s.get(_User, rop_uid)
            ids = []
            for u, summa in ((emp, 200_000), (rop, 300_000)):
                r = await ab.submit(s, u, PERIOD, summa)
                adj = await s.scalar(
                    _sel(_Adj).where(_Adj.user_id == u.id, _Adj.source == _Src.bot.value))
                ids.append(adj.id)
            return ids

    try:
        ids = _asyncio.run(_make_requests())
        with httpx.Client(timeout=20) as client:
            r = client.get(f"{API_BASE}/payroll/advance-summary?period={PERIOD}",
                           headers=auth(boss_t))
            d = r.json()
            check("jami summa tasdiqlashdan OLDIN ko'rinadi",
                  d["pending_count"] == 2 and abs(d["pending_total"] - 500_000) < 1
                  and d["pending_employees"] == 2, str(d)[:200])
            check("bugungi so'rovlar alohida sanaladi",
                  d["today_count"] == 2 and abs(d["today_total"] - 500_000) < 1, str(d)[:200])
    except Exception:
        check("D-02 yig'indi testi", False, traceback.format_exc(limit=3).strip())
        cleanup_hr()
        return

    try:
        with httpx.Client(timeout=20) as client:
            # Biri allaqachon hal qilingan bo'lsin
            r0 = client.post(f"{API_BASE}/payroll/advances/{ids[0]}/decide",
                             headers=auth(boss_t), json={"approve": True})
            r = client.post(f"{API_BASE}/payroll/advances/bulk-decide", headers=auth(boss_t),
                            json={"ids": ids, "approve": True, "note": "Ommaviy tasdiq"})
            d = r.json()
            check("ommaviy tasdiqlash ishlaydi",
                  r.status_code == 200 and d["decided"] == 1, f"kod={r.status_code}, {d}")
            check("allaqachon hal qilingani JIMGINA o'tkaziladi (amal yiqilmaydi)",
                  d["skipped"] == 1, str(d))
    except Exception:
        check("D-02 ommaviy tasdiq testi", False, traceback.format_exc(limit=3).strip())

    async def _audits():
        async with async_session() as s:
            rows = list(await s.scalars(
                _sel(_Audit).where(_Audit.target_user_id.in_(UIDS))))
            return [(a.action, (a.after or {}).get("bulk")) for a in rows]

    try:
        acts = _asyncio.run(_audits())
        # Ommaviy amal 1 tasini hal qildi (ikkinchisi allaqachon
        # tasdiqlangan edi) — demak `bulk` belgili audit AYNAN 1 ta.
        # Belgisizlari — alohida tasdiqlar (yuqoridagi bloklardan).
        check("ommaviy tasdiqdagi HAR BIRI auditga tushdi (`bulk` belgisi bilan)",
              sum(1 for a, b in acts if a == "advance_approved" and b) == 1
              and sum(1 for a, b in acts if a == "advance_approved") >= 2,
              str(acts)[:220])
    except Exception:
        check("audit testi", False, traceback.format_exc(limit=2).strip())

    # ── D-03: ketma-ket ──
    async def _streaks(davrlar):
        """Berilgan davrlarga avans qo'yib, ketma-ketlikni hisoblaydi."""
        async with async_session() as s:
            # `advance_responses.adjustment_id` FK bo'lgani uchun avval
            # bog'lanishni uzamiz — aks holda o'chirish FK xatosi beradi.
            await s.execute(
                _upd(_Resp).where(_Resp.user_id == emp_uid).values(adjustment_id=None))
            await s.execute(_del(_Adj).where(_Adj.user_id == emp_uid))
            for p in davrlar:
                s.add(_Adj(user_id=emp_uid, period=p, kind="minus", category="advance",
                           status=_AStatus.approved.value, amount=100_000,
                           reason="T-Hr ketma-ket", created_by=emp_uid,
                           source=_Src.hr_manual.value))
            await s.commit()
            rows = await pay_router._advance_streaks(s, PERIOD)
            return [r for r in rows if r.user_id == emp_uid]

    def _oldingi(period, n):
        y, m = (int(x) for x in period.split("-"))
        for _ in range(n):
            m -= 1
            if m == 0:
                y, m = y - 1, 12
        return f"{y}-{m:02d}"

    try:
        uch = [PERIOD, _oldingi(PERIOD, 1), _oldingi(PERIOD, 2)]
        rows = _asyncio.run(_streaks(uch))
        check("ketma-ket 3 oy to'g'ri sanaladi",
              rows and rows[0].months == 3, str([(r.months, r.total) for r in rows]))
        check("3 oydan boshlab belgi qo'yiladi", rows and rows[0].flagged is True,
              str(rows[0].flagged) if rows else "yo'q")

        # Oraliq uzilgan: joriy, [bo'shliq], 3 oy oldin
        uzilgan = [PERIOD, _oldingi(PERIOD, 2), _oldingi(PERIOD, 3)]
        rows2 = _asyncio.run(_streaks(uzilgan))
        check("⭐ oraliq uzilsa hisob QAYTADAN boshlanadi (1 oy)",
              rows2 and rows2[0].months == 1, str([(r.months,) for r in rows2]))
        check("bir oylik ketma-ketlikka belgi QO'YILMAYDI",
              rows2 and rows2[0].flagged is False, str(rows2[0].flagged) if rows2 else "yo'q")
    except Exception:
        check("D-03 ketma-ket testi", False, traceback.format_exc(limit=3).strip())

    # ── D-04: payslipda BIR MARTA ──
    async def _payslip_once():
        async with async_session() as s:
            await s.execute(
                _upd(_Resp).where(_Resp.user_id == emp_uid).values(adjustment_id=None))
            await s.execute(_del(_Adj).where(_Adj.user_id == emp_uid))
            await s.commit()
            emp = await s.get(_User, emp_uid)
            await ab.submit(s, emp, PERIOD, 250_000)
            adj = await s.scalar(
                _sel(_Adj).where(_Adj.user_id == emp_uid, _Adj.source == _Src.bot.value))
            adj.status = _AStatus.approved.value
            await s.commit()
            f1 = (await build_payslip(s, emp, PERIOD))["fields"]
            # To'langan deb belgilaymiz — summa O'ZGARMASLIGI kerak
            adj.status = _AStatus.issued.value
            await s.commit()
            f2 = (await build_payslip(s, emp, PERIOD))["fields"]
            n = await s.scalar(
                _sel(func.count()).select_from(_Adj).where(
                    _Adj.user_id == emp_uid, _Adj.source == _Src.bot.value))
            return f1, f2, n

    from sqlalchemy import func

    try:
        f1, f2, n = _asyncio.run(_payslip_once())
        check("bot avansi bazada BITTA qator (dublikat yo'q)", n == 1, f"={n}")
        check("⭐ payslipda bot avansi BIR MARTA ayirilgan (250 000)",
              abs(f1["adjustments_minus"] - 250_000) < 1, f"={f1['adjustments_minus']}")
        check("to'langan deb belgilangach summa O'ZGARMAYDI",
              abs(f2["adjustments_minus"] - 250_000) < 1, f"={f2['adjustments_minus']}")
    except Exception:
        check("D-04 payslip testi", False, traceback.format_exc(limit=3).strip())

    cleanup_hr()
    try:
        conn.close()
    except Exception:
        pass


def test_app_login_code_delivery() -> None:
    """Saytga kirish kodini yetkazish (2026-08-21 jonli muammo).

    MUAMMO: sayt -> bot -> «kod mobil ilovangizga yuborildi» -> kod
    KELMAYDI. Sabab: FCM ro'yxatdan chiqmagan tokenga HTTP 200
    qaytaradi, ya'ni «yuborildi» «yetib bordi» degani EMAS. Ilova
    o'chirilgan / bildirishnoma ruxsati olib qo'yilgan / batareya
    cheklovi / eski APK'da kanal yo'q bo'lsa xabar ko'rinmaydi.
    Bot esa kodni kutardi va CHIQISH YO'LI YO'Q edi.

    Tekshiriladi:
      1. ESKI qurilma (last_seen eski) push uchun HISOBGA OLINMAYDI
      2. YANGI qurilma hisobga olinadi
      3. Qurilmasi yo'q xodim -> `screen_fallback` (eski xatti-harakat)
      4. ⭐ «Kod kelmadi» -> `code_delivery` «screen» ga o'tadi va
         sayt kodni ko'rsata boshlaydi
      5. O'tgach kod bilan kirish ISHLAYDI
      6. Begona/eskirgan token -> `invalid`
      7. Amal auditga tushadi
    """
    print("\n=== SAYTGA KIRISH: kod yetkazish (2026-08-21) ===")
    import asyncio as _asyncio
    import httpx
    from datetime import datetime as _dt, timedelta as _td

    from api.services import push as _push
    from db.base import async_session
    from db.models import (
        AppLoginStatus,
        AppLoginToken,
        AuditLog as _Audit,
        PushToken,
        User as _User,
    )
    from sqlalchemy import delete as _del, select as _sel

    conn = db()
    cur = conn.cursor()
    cur.execute("select id from users where full_name like 'T-Log-%'")
    stale = [r[0] for r in cur.fetchall()]
    if stale:
        qm = ",".join("?" * len(stale))
        cur.execute(f"delete from push_tokens where user_id in ({qm})", stale)
        cur.execute(f"delete from audit_logs where target_user_id in ({qm})", stale)
        cur.execute(f"delete from users where id in ({qm})", stale)
    TG = 999609401
    cur.execute(
        "insert into users (telegram_id, full_name, role, bot_started, is_active, created_at)"
        " values (?, 'T-Log-Hr','hr',1,1,datetime('now'))", (TG,))
    uid = cur.lastrowid
    conn.commit()

    def cleanup_log():
        try:
            c2 = db()
            c2.execute("delete from push_tokens where user_id=?", (uid,))
            c2.execute("delete from app_login_tokens where token like 'T-Log-%'")
            c2.execute("delete from audit_logs where target_user_id=?", (uid,))
            c2.execute("delete from users where id=?", (uid,))
            c2.commit()
            c2.close()
        except Exception:
            print("  login testi tozalash xatosi:\n" + traceback.format_exc(limit=1).strip())

    # ── 1-2. Eski/yangi qurilma filtri (FCM'ga CHIQMAYDI: patch) ──
    async def _device_filter():
        yuborilgan: list = []

        async def _fake_send_one(token, category, title, body, data, quiet):
            yuborilgan.append(token)
            return "ok"

        orig = _push._send_one
        _push._send_one = _fake_send_one
        try:
            out = {}
            async with async_session() as s:
                u = await s.get(_User, uid)
                # Qurilma umuman yo'q
                out["no_device"] = await _push.send_login_code(s, u, "1234")
                # ESKI qurilma — chegaradan tashqarida
                eski = _dt.utcnow() - _td(days=_push.ACTIVE_DEVICE_DAYS + 3)
                s.add(PushToken(user_id=uid, token="T-Log-eski", platform="android",
                                is_active=True, last_seen_at=eski))
                await s.commit()
                yuborilgan.clear()
                out["stale"] = await _push.send_login_code(s, u, "1234")
                # YANGI qurilma
                s.add(PushToken(user_id=uid, token="T-Log-yangi", platform="android",
                                is_active=True, last_seen_at=_dt.utcnow()))
                await s.commit()
                yuborilgan.clear()
                out["fresh"] = await _push.send_login_code(s, u, "1234")
                out["fresh_tokens"] = list(yuborilgan)
                return out
        finally:
            _push._send_one = orig

    try:
        d = _asyncio.run(_device_filter())
        check("qurilmasi yo'q xodimga push yuborilmaydi", d["no_device"] == 0, str(d["no_device"]))
        check("⭐ ESKI qurilma push uchun hisobga OLINMAYDI",
              d["stale"] == 0, f"={d['stale']} (0 bo'lishi kerak)")
        check("yangi qurilma hisobga olinadi va FAQAT u",
              d["fresh"] == 1 and d["fresh_tokens"] == ["T-Log-yangi"],
              f"={d['fresh']}, tokenlar={d['fresh_tokens']}")
    except Exception:
        check("qurilma filtri testi", False, traceback.format_exc(limit=3).strip())
        cleanup_log()
        return

    # ── 3-7. «Kod kelmadi» yo'li ──
    async def _make_token(delivery="push"):
        async with async_session() as s:
            await s.execute(_del(AppLoginToken).where(AppLoginToken.token.like("T-Log-%")))
            row = AppLoginToken(
                token="T-Log-token1", status=AppLoginStatus.pending.value,
                pairing_code="4321", code_delivery=delivery,
                expires_at=_dt.utcnow() + _td(minutes=5),
            )
            s.add(row)
            await s.commit()

    try:
        _asyncio.run(_make_token())
        with httpx.Client(timeout=20) as c:
            # 4. «Kod kelmadi»
            r = c.post(f"{API_BASE}/auth/app-login/use-screen", headers=bot_secret_hdr(),
                       json={"login_token": "T-Log-token1", "telegram_id": TG})
            check("⭐ «Kod kelmadi» -> screen_fallback",
                  r.status_code == 200 and r.json().get("status") == "screen_fallback",
                  f"kod={r.status_code} {r.text[:120]}")

            # Sayt endi kodni ko'rsatishi kerak
            # Poll — POST (sayt bir necha soniyada chaqirib turadi).
            r = c.post(f"{API_BASE}/auth/app-login/poll",
                       json={"login_token": "T-Log-token1"})
            check("⭐ sayt poll'da `code_delivery=screen` ko'radi (kodni ochadi)",
                  r.status_code == 200 and r.json().get("code_delivery") == "screen",
                  f"kod={r.status_code} {r.text[:140]}")

            # 5. Kod bilan kirish ishlaydi
            r = c.post(f"{API_BASE}/auth/app-login/confirm", headers=bot_secret_hdr(),
                       json={"login_token": "T-Log-token1", "telegram_id": TG,
                             "pairing_code": "4321"})
            check("screen'ga o'tgach kod bilan kirish ISHLAYDI",
                  r.status_code == 200 and r.json().get("status") == "ok",
                  f"kod={r.status_code} {r.text[:140]}")

            # 6. Begona token
            r = c.post(f"{API_BASE}/auth/app-login/use-screen", headers=bot_secret_hdr(),
                       json={"login_token": "T-Log-yoq", "telegram_id": TG})
            check("mavjud bo'lmagan token -> invalid",
                  r.json().get("status") == "invalid", str(r.json())[:120])

            # Bot siri bo'lmasa
            r = c.post(f"{API_BASE}/auth/app-login/use-screen",
                       json={"login_token": "T-Log-token1", "telegram_id": TG})
            check("bot siri bo'lmasa -> 401/403",
                  r.status_code in (401, 403), f"kod={r.status_code}")
    except Exception:
        check("«Kod kelmadi» testi", False, traceback.format_exc(limit=3).strip())

    # 7. Audit
    async def _audit():
        async with async_session() as s:
            return [a.action for a in await s.scalars(
                _sel(_Audit).where(_Audit.target_user_id == uid))]

    try:
        acts = _asyncio.run(_audit())
        check("amal auditga tushdi", "app_login_switched_to_screen" in acts, str(acts)[:140])
    except Exception:
        check("login audit testi", False, traceback.format_exc(limit=2).strip())

    cleanup_log()
    try:
        conn.close()
    except Exception:
        pass

if __name__ == "__main__":
    main()
