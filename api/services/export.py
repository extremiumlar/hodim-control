from datetime import date, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.routers.norms import METRIC_LABELS, VIDEO_METRIC_TYPES, metrics_for
from api.routers.stats import VISIT_STAGE_NAME, _confirmed_videos_count
from api.timeutil import TASHKENT_TZ, local_range_utc_naive
from db.models import (
    Bonus,
    DailyResult,
    ExcusedDay,
    ExcusedStatus,
    HourlyActual,
    LeadStageDaily,
    OperatorCallsDaily,
    Payslip,
    PayslipItem,
    Role,
    TaskModel,
    TaskStatus,
    User,
)

# Ustunlar tartibi barqaror bo'lishi uchun (faqat faol xodimlar lavozimlarida
# uchraydigan metrikalar ko'rsatiladi, lekin tartib doim shu).
METRIC_ORDER = list(METRIC_LABELS)
METRIC_TOTAL_LABELS = {
    "suhbat": "Suhbatlar (jami)",
    "tashrif": "Tashriflar (jami)",
    "oddiy_video": "Oddiy videolar (jami)",
    "dumaloq_video": "Dumaloq videolar (jami)",
}

# CRM snapshot ustunlari (operator kesimi — crm_visit_external_id orqali bog'lanadi)
CRM_HEADERS = [
    "Qo'ng'iroqlar (jami)",
    "Gaplashgan vaqt",
    "Ishlangan lidlar (jami)",
]

FIXED_HEADERS = [
    "Vazifalar (bajarilgan/jami)",
    "Sababli kunlar (tasdiqlangan)",
    "Bonus (agar bitta oy tanlangan bo'lsa)",
]


def _fmt_talk_cell(sec: int) -> str:
    """5977s → '1:39' (soat:daqiqa) — Excel katakchasi uchun ixcham matn."""
    minutes = sec // 60
    hours, minutes = divmod(minutes, 60)
    return f"{hours}:{minutes:02d}"


async def _crm_daily_maps(db: AsyncSession, date_from: date, date_to: date) -> tuple[dict, dict, dict]:
    """Snapshot jadvallaridan kunlik xaritalar (bitta-bitta grouped so'rov):
    calls[(date, rid)], leads[(date, rid)] = (lid, tashrif), talk[(date, user_id)]."""
    call_rows = await db.execute(
        select(
            OperatorCallsDaily.date,
            OperatorCallsDaily.responsible_id,
            func.sum(OperatorCallsDaily.calls_in + OperatorCallsDaily.calls_out),
        )
        .where(OperatorCallsDaily.date >= date_from, OperatorCallsDaily.date <= date_to)
        .group_by(OperatorCallsDaily.date, OperatorCallsDaily.responsible_id)
    )
    calls = {(d, rid): int(v or 0) for d, rid, v in call_rows.all()}

    is_visit = func.lower(func.trim(LeadStageDaily.stage_name)) == VISIT_STAGE_NAME
    lead_rows = await db.execute(
        select(
            LeadStageDaily.date,
            LeadStageDaily.responsible_id,
            func.sum(LeadStageDaily.leads_count),
            func.sum(case((is_visit, LeadStageDaily.leads_count), else_=0)),
        )
        .where(LeadStageDaily.date >= date_from, LeadStageDaily.date <= date_to)
        .group_by(LeadStageDaily.date, LeadStageDaily.responsible_id)
    )
    leads = {(d, rid): (int(t or 0), int(v or 0)) for d, rid, t, v in lead_rows.all()}

    talk_rows = await db.execute(
        select(HourlyActual.date, HourlyActual.user_id, func.sum(HourlyActual.talk_sec))
        .where(HourlyActual.date >= date_from, HourlyActual.date <= date_to)
        .group_by(HourlyActual.date, HourlyActual.user_id)
    )
    talk = {(d, uid): int(v or 0) for d, uid, v in talk_rows.all()}

    return calls, leads, talk


async def build_report_xlsx(db: AsyncSession, date_from: date, date_to: date) -> BytesIO:
    day_start, day_end = local_range_utc_naive(date_from, date_to)

    # Bonus faqat aniq bitta oy tanlanganda ko'rsatiladi (davr YYYY-MM formatida saqlanadi,
    # ixtiyoriy sana oralig'i uchun bir nechta oyni yig'ish MVP doirasidan tashqarida).
    single_month_period = (
        date_from.strftime("%Y-%m") if date_from.strftime("%Y-%m") == date_to.strftime("%Y-%m") else None
    )

    employees = list(
        await db.scalars(
            select(User).where(User.role == Role.employee.value, User.is_active == True).order_by(User.full_name)  # noqa: E712
        )
    )

    # Ustunlar dinamik: kamida bitta faol xodim lavozimida uchraydigan metrikalar.
    # Xodimda kuzatilmaydigan metrika 0 emas "—" bilan ko'rsatiladi — "0 natija"
    # va "umuman kuzatilmaydi" farqi yo'qolmasligi uchun.
    metrics_by_user = {emp.id: metrics_for(emp) for emp in employees}
    used_metrics = [m for m in METRIC_ORDER if any(m in keys for keys in metrics_by_user.values())]

    # CRM snapshot xaritalari (qo'ng'iroq / lid / gaplashgan vaqt) va rid bog'lanishi
    calls_map, leads_map, talk_map = await _crm_daily_maps(db, date_from, date_to)
    rid_by_user: dict[int, int] = {}
    for emp in employees:
        try:
            if emp.crm_visit_external_id:
                rid_by_user[emp.id] = int(emp.crm_visit_external_id)
        except (TypeError, ValueError):
            continue

    # Vazifalarni mahalliy kun bo'yicha chelaklash — "Kunlik" varaq uchun
    # (created_at bazada naive-UTC, Toshkent kuniga o'girib sanaymiz)
    all_tasks = list(
        await db.scalars(
            select(TaskModel).where(TaskModel.created_at >= day_start, TaskModel.created_at < day_end)
        )
    )
    tasks_by_user_day: dict[tuple[int, date], list[TaskModel]] = {}
    for t in all_tasks:
        local_day = t.created_at.replace(tzinfo=timezone.utc).astimezone(TASHKENT_TZ).date()
        tasks_by_user_day.setdefault((t.assigned_to, local_day), []).append(t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    ws.append([f"Hisobot davri: {date_from.isoformat()} — {date_to.isoformat()}"])
    ws.append(["Xodim", *(METRIC_TOTAL_LABELS[m] for m in used_metrics), *CRM_HEADERS, *FIXED_HEADERS])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for emp in employees:
        results = list(
            await db.scalars(
                select(DailyResult).where(
                    DailyResult.user_id == emp.id, DailyResult.date >= date_from, DailyResult.date <= date_to
                )
            )
        )
        totals = {
            "suhbat": sum(r.conversations_count for r in results),
            "tashrif": sum(r.visits_count for r in results),
        }
        for metric_key, video_type in VIDEO_METRIC_TYPES.items():
            if metric_key in metrics_by_user[emp.id]:
                totals[metric_key] = await _confirmed_videos_count(
                    db, emp.id, date_from, date_to, video_type=video_type
                )

        metric_cells = [
            totals.get(m, 0) if m in metrics_by_user[emp.id] else "—" for m in used_metrics
        ]

        tasks = list(
            await db.scalars(
                select(TaskModel).where(
                    TaskModel.assigned_to == emp.id,
                    TaskModel.created_at >= day_start,
                    TaskModel.created_at < day_end,
                )
            )
        )
        tasks_total = len(tasks)
        tasks_done = sum(1 for t in tasks if t.status == TaskStatus.done.value)

        excused_count = len(
            list(
                await db.scalars(
                    select(ExcusedDay).where(
                        ExcusedDay.user_id == emp.id,
                        ExcusedDay.date >= date_from,
                        ExcusedDay.date <= date_to,
                        ExcusedDay.status == ExcusedStatus.approved.value,
                    )
                )
            )
        )

        bonus_amount = ""
        if single_month_period:
            bonus = await db.scalar(
                select(Bonus).where(Bonus.user_id == emp.id, Bonus.period == single_month_period)
            )
            if bonus:
                bonus_amount = float(bonus.amount)

        # CRM snapshot jami (qo'ng'iroq / gaplashgan / lid) — bog'lanmagan xodimda "—"
        rid = rid_by_user.get(emp.id)
        if rid is not None:
            emp_calls = sum(v for (d, r), v in calls_map.items() if r == rid)
            emp_leads = sum(t for (d, r), (t, _v) in leads_map.items() if r == rid)
            emp_talk = sum(v for (d, uid), v in talk_map.items() if uid == emp.id)
            crm_cells = [emp_calls, _fmt_talk_cell(emp_talk) if emp_talk else "0:00", emp_leads]
        else:
            crm_cells = ["—", "—", "—"]

        ws.append(
            [
                emp.full_name,
                *metric_cells,
                *crm_cells,
                f"{tasks_done}/{tasks_total}",
                excused_count,
                bonus_amount,
            ]
        )

    # ─── "Kunlik" varaq: kun × xodim kesimi ────────────────────────────────────
    ws_daily = wb.create_sheet("Kunlik")
    ws_daily.append(["Sana", "Xodim", "Qo'ng'iroqlar", "Gaplashgan vaqt", "Ishlangan lidlar", "Tashriflar", "Vazifalar"])
    for cell in ws_daily[1]:
        cell.font = Font(bold=True)

    day = date_from
    while day <= date_to:
        for emp in employees:
            rid = rid_by_user.get(emp.id)
            calls = calls_map.get((day, rid), 0) if rid is not None else 0
            leads, visits = leads_map.get((day, rid), (0, 0)) if rid is not None else (0, 0)
            talk = talk_map.get((day, emp.id), 0)
            day_tasks = tasks_by_user_day.get((emp.id, day), [])
            # Faoliyatsiz kun qatori chiqarilmaydi — varaq shishib ketmasin
            if not calls and not leads and not talk and not day_tasks:
                continue
            done = sum(1 for t in day_tasks if t.status == TaskStatus.done.value)
            ws_daily.append(
                [
                    day.isoformat(),
                    emp.full_name,
                    calls,
                    _fmt_talk_cell(talk) if talk else "",
                    leads,
                    visits,
                    f"{done}/{len(day_tasks)}" if day_tasks else "",
                ]
            )
        day += timedelta(days=1)

    for column_cells in ws_daily.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None) if any(
            cell.value is not None for cell in column_cells
        ) else 10
        ws_daily.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None) if any(
            cell.value is not None for cell in column_cells
        ) else 10
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)


_SHEET_NAME_INVALID = set(r"\/:*?[]")


def _safe_sheet_name(full_name: str, user_id: int) -> str:
    """Excel varaq nomi <=31 belgi va `\\/:*?[]` belgilarisiz bo'lishi shart;
    ism qisqartirilgach ikki xodim to'qnashmasligi uchun `#user_id` doim
    qo'shiladi (ism o'zgarishi/takrorlanishidan qat'i nazar barqaror)."""
    cleaned = "".join(c for c in full_name if c not in _SHEET_NAME_INVALID).strip() or f"Xodim {user_id}"
    suffix = f" #{user_id}"
    return cleaned[: 31 - len(suffix)] + suffix


async def build_payroll_xlsx(db: AsyncSession, period: str, user_ids: list[int] | None = None) -> BytesIO:
    """Oylik ish haqi varag'i — Bosqich 7 (OYLIK_JARIMA_REJASI.md). Bitta
    kitobda: "Xulosa" (barcha xodim, bir qator) + har bir xodim uchun
    ALOHIDA varaq (to'liq `PayslipItem` tafsiloti — "nega bu summa" savoliga
    bir qarashda javob, 1-bo'lim ⭐ band bilan bir xil printsip). `user_ids`
    berilsa faqat o'sha xodimlar (ROP faqat o'z jamoasini ko'rishi uchun,
    `payroll.py::_visible_users` bilan bir xil qamrov)."""
    q = select(Payslip).join(User, Payslip.user_id == User.id).where(Payslip.period == period)
    if user_ids is not None:
        q = q.where(Payslip.user_id.in_(user_ids))
    payslips = list(await db.scalars(q))

    names = {}
    if payslips:
        result = await db.execute(
            select(User.id, User.full_name).where(User.id.in_([p.user_id for p in payslips]))
        )
        names = dict(result.all())
    payslips.sort(key=lambda p: names.get(p.user_id, ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Xulosa"
    ws.append([f"Oylik ish haqi — {period}"])
    ws.append(
        [
            "Xodim", "Holat", "Asosiy oylik", "Qo'shimcha ish", "Bonus", "Qo'shimchalar (+)",
            "Kechikish jarimasi", "Kelmagan kun ushlanmasi", "Qo'shimchalar (-)", "Jami (gross)", "Netto",
        ]
    )
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for p in payslips:
        ws.append(
            [
                names.get(p.user_id, f"#{p.user_id}"),
                p.status,
                float(p.base_amount),
                float(p.overtime_amount),
                float(p.bonus_amount),
                float(p.adjustments_plus),
                -float(p.fine_amount),
                -float(p.absent_deduction),
                -float(p.adjustments_minus),
                float(p.gross),
                float(p.net),
            ]
        )
    if payslips:
        ws.append(
            [
                "JAMI", "",
                sum(float(p.base_amount) for p in payslips),
                sum(float(p.overtime_amount) for p in payslips),
                sum(float(p.bonus_amount) for p in payslips),
                sum(float(p.adjustments_plus) for p in payslips),
                -sum(float(p.fine_amount) for p in payslips),
                -sum(float(p.absent_deduction) for p in payslips),
                -sum(float(p.adjustments_minus) for p in payslips),
                sum(float(p.gross) for p in payslips),
                sum(float(p.net) for p in payslips),
            ]
        )
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    _autosize_columns(ws)

    if payslips:
        items_by_payslip: dict[int, list[PayslipItem]] = {}
        item_rows = list(
            await db.scalars(
                select(PayslipItem)
                .where(PayslipItem.payslip_id.in_([p.id for p in payslips]))
                .order_by(PayslipItem.sort_order)
            )
        )
        for it in item_rows:
            items_by_payslip.setdefault(it.payslip_id, []).append(it)

        for p in payslips:
            ws_emp = wb.create_sheet(_safe_sheet_name(names.get(p.user_id, f"Xodim {p.user_id}"), p.user_id))
            ws_emp.append([f"{names.get(p.user_id, f'#{p.user_id}')} — {period} oyi uchun oylik varaqasi"])
            ws_emp.append(["Holat", p.status])
            if p.approved_at is not None:
                ws_emp.append(["Tasdiqlangan", p.approved_at.isoformat()])
            ws_emp.append([])
            ws_emp.append(["Tavsif", "Miqdor", "Stavka", "Summa"])
            for cell in ws_emp[ws_emp.max_row]:
                cell.font = Font(bold=True)
            for it in items_by_payslip.get(p.id, []):
                ws_emp.append(
                    [
                        it.label,
                        float(it.quantity) if it.quantity is not None else "",
                        float(it.rate) if it.rate is not None else "",
                        float(it.amount),
                    ]
                )
            ws_emp.append(["", "", "Netto", float(p.net)])
            for cell in ws_emp[ws_emp.max_row]:
                cell.font = Font(bold=True)
            _autosize_columns(ws_emp)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
