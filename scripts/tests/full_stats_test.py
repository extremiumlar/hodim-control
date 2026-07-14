"""Statistika steki uchun to'liq test batareyasi (V0-V5 + digestlar).

Ishga tushirish (repo ildizidan, jonli app.db ustida ishlaydi):
    .venv/Scripts/python scripts/tests/full_stats_test.py

Tamoyillar:
- Guruhga HECH NARSA yuborilmaydi: dry_run yoki send_message monkeypatch.
- Har raqam MUSTAQIL so'rov bilan qayta hisoblanadi (endpoint o'z-o'zini tasdiqlamaydi).
- DB mutatsiyalari (GroupPostConfig) testdan keyin QAT'IY asl holatga tiklanadi.
- Har test try/except ichida — bitta xato qolganlarini to'xtatmaydi.

Eslatma: ba'zi tekshiruvlar jonli ma'lumot mavjudligiga bog'liq (masalan Firuzabonu
operatori, bugungi snapshot) — bo'sh bazada ular shartli o'tkazib yuboriladi.
"""
import asyncio
import datetime as dt
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
# Repo ildizi: scripts/tests/ dan ikki daraja yuqorida
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

PASS = 0
FAIL = 0
RESULTS: list[str] = []


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    mark = "PASS" if cond else "FAIL"
    if cond:
        PASS += 1
    else:
        FAIL += 1
    line = f"[{mark}] {name}" + (f" — {detail}" if detail else "")
    RESULTS.append(line)
    print(line)


async def db_sums(db, day_from, day_to):
    """Mustaqil hisob: [day_from, day_to] jami calls/leads/visits/talk (aktiv filtrisiz, rid kesimi bilan)."""
    from sqlalchemy import case, func, select

    from db.models import HourlyActual, LeadStageDaily, OperatorCallsDaily

    calls_by_rid = {
        rid: int(v or 0)
        for rid, v in (
            await db.execute(
                select(OperatorCallsDaily.responsible_id, func.sum(OperatorCallsDaily.calls_in + OperatorCallsDaily.calls_out))
                .where(OperatorCallsDaily.date >= day_from, OperatorCallsDaily.date <= day_to)
                .group_by(OperatorCallsDaily.responsible_id)
            )
        ).all()
    }
    is_visit = func.lower(func.trim(LeadStageDaily.stage_name)) == "tashrif"
    leads_by_rid = {}
    visits_by_rid = {}
    for rid, t, v in (
        await db.execute(
            select(
                LeadStageDaily.responsible_id,
                func.sum(LeadStageDaily.leads_count),
                func.sum(case((is_visit, LeadStageDaily.leads_count), else_=0)),
            )
            .where(LeadStageDaily.date >= day_from, LeadStageDaily.date <= day_to)
            .group_by(LeadStageDaily.responsible_id)
        )
    ).all():
        leads_by_rid[rid] = int(t or 0)
        visits_by_rid[rid] = int(v or 0)
    talk_by_user = {
        uid: int(v or 0)
        for uid, v in (
            await db.execute(
                select(HourlyActual.user_id, func.sum(HourlyActual.talk_sec))
                .where(HourlyActual.date >= day_from, HourlyActual.date <= day_to)
                .group_by(HourlyActual.user_id)
            )
        ).all()
    }
    # "Aktiv" (digest bilan bir xil filtr): calls yoki leads bor rid'lar
    rids = set(calls_by_rid) | set(leads_by_rid)
    active = {r for r in rids if calls_by_rid.get(r, 0) or leads_by_rid.get(r, 0)}
    return {
        "calls": sum(calls_by_rid.get(r, 0) for r in active),
        "leads": sum(leads_by_rid.get(r, 0) for r in active),
        "visits": sum(visits_by_rid.get(r, 0) for r in active),
        "calls_by_rid": calls_by_rid,
        "leads_by_rid": leads_by_rid,
        "talk_by_user": talk_by_user,
        "active": active,
    }


async def main():
    from sqlalchemy import select

    from api.config import settings
    from api.routers.stats import group_post_tick, web_operator_summary, web_stats_overview
    from api.services import daily_digest as dd
    from api.services.daily_digest import (
        build_daily_digest,
        digest_group_targets,
        send_daily_digest,
        send_yesterday_correction,
    )
    from api.services.export import build_report_xlsx
    from api.services.monthly_digest import build_monthly_digest
    from api.services.weekly_digest import _pct_change, build_weekly_digest
    from api.timeutil import today_local
    from db.base import async_session
    from db.models import Bonus, GroupPostConfig, ShortfallReason, User

    today = today_local()
    yesterday = today - dt.timedelta(days=1)

    async with async_session() as db:
        # ═══ T1: digest_group_targets (V0) ═══
        t = digest_group_targets(555)
        check("T1a chat_id berilsa faqat o'sha", t == [555], f"{t}")
        t = digest_group_targets(None)
        expected_all = [settings.telegram_group_chat_id, *settings.stats_group_ids]
        check("T1b default: asosiy + stats guruhlar", t == expected_all, f"{t}")
        check("T1c takror/0 yo'q", len(t) == len(set(t)) and 0 not in t)

        # ═══ T2: kunlik digest jami raqamlari mustaqil hisob bilan mos (V5 asosi) ═══
        digest = await build_daily_digest(db)
        ind = await db_sums(db, today, today)
        if digest["text"] is None:
            check("T2 bugun digest (ma'lumot yo'q kuni)", ind["calls"] == 0, "ma'lumot yo'q — mos")
        else:
            check(
                "T2a digest.totals.calls == mustaqil hisob",
                digest["totals"]["calls"] == ind["calls"],
                f"digest={digest['totals']['calls']} mustaqil={ind['calls']}",
            )
            check(
                "T2b digest.totals.leads == mustaqil hisob",
                digest["totals"]["leads"] == ind["leads"],
                f"digest={digest['totals']['leads']} mustaqil={ind['leads']}",
            )
            check(
                "T2c digest.totals.visits == mustaqil hisob",
                digest["totals"]["visits"] == ind["visits"],
                f"digest={digest['totals']['visits']} mustaqil={ind['visits']}",
            )
            check("T2d legend qatori bor", "qo'ng'iroq (kechaga nisbatan)" in digest["text"])
            check(
                "T2e 'Boshqa operatorlar' izohi (rid=0 bo'lsa)",
                (0 in ind["active"]) == ("bog'lanmagan xodimlar" in digest["text"]),
            )

        # ═══ T3: send_daily_digest dry_run hech narsa yubormaydi, totals qaytaradi ═══
        r = await send_daily_digest(db, dry_run=True)
        check("T3a dry_run sent=False + matn bor", r.get("sent") is False and bool(r.get("text")))
        check("T3b dry_run'da ham AI yo'q/bor — matn buzilmagan", "Kun yakuni" in r.get("text", ""))

        # ═══ T4: haftalik digest — jami va % matematikasi ═══
        wk = await build_weekly_digest(db)
        wk_ind = await db_sums(db, today - dt.timedelta(days=6), today)
        if wk["text"]:
            # Jami qatori matnidan sonni sug'urib olish o'rniga mustaqil qayta hisob:
            check(
                "T4a haftalik jami calls matnda bor",
                f"📞 {wk_ind['calls']}" in wk["text"],
                f"kutilgan {wk_ind['calls']}",
            )
            # % formulasi: _pct_change to'g'riligi
            check("T4b pct_change(150,100)=+50", _pct_change(150, 100) == 50)
            check("T4c pct_change kichik bazada None", _pct_change(50, 5) is None)
            check("T4d pct_change prev=None → None", _pct_change(50, None) is None)

        # ═══ T5: oylik digest — jami, bonus yig'indisi ═══
        mo = await build_monthly_digest(db)
        mo_ind = await db_sums(db, today.replace(day=1), today)
        bonuses = list(await db.scalars(select(Bonus).where(Bonus.period == today.strftime("%Y-%m"))))
        if mo["text"]:
            check(
                "T5a oylik jami calls matnda",
                f"📞 {mo_ind['calls']}" in mo["text"],
                f"kutilgan {mo_ind['calls']}",
            )
            total_bonus = sum(float(b.amount) for b in bonuses)
            if total_bonus:
                fmt = f"{total_bonus:,.0f}".replace(",", " ")
                check("T5b bonus jami to'g'ri", fmt in mo["text"], f"kutilgan {fmt}")
            # O'tgan oy (iyun) ma'lumoti yo'q — Jami qatorida oy nomi bo'lmasligi kerak
            check("T5c iyun bo'sh → taqqos yo'q", "iyun" not in mo["text"].split("Jami:")[1].split("\n")[0])

        # ═══ T6: Excel eksport raqamlari ═══
        from openpyxl import load_workbook

        buf = await build_report_xlsx(db, dt.date(2026, 7, 1), today)
        wb = load_workbook(buf)
        ws = wb["Hisobot"]
        headers = [c.value for c in ws[2]]
        check("T6a yangi ustunlar bor", "Qo'ng'iroqlar (jami)" in headers and "Gaplashgan vaqt" in headers)
        # Firuzabonu qatori vs mustaqil hisob
        month_ind = await db_sums(db, dt.date(2026, 7, 1), today)
        users = list(await db.scalars(select(User)))
        firuza = next((u for u in users if "Firuzabonu" in u.full_name), None)
        if firuza and firuza.crm_visit_external_id:
            rid = int(firuza.crm_visit_external_id)
            row = next((r for r in ws.iter_rows(min_row=3, values_only=True) if r[0] == firuza.full_name), None)
            ci = headers.index("Qo'ng'iroqlar (jami)")
            li = headers.index("Ishlangan lidlar (jami)")
            check(
                "T6b Firuzabonu calls (Excel == DB)",
                row and row[ci] == month_ind["calls_by_rid"].get(rid, 0),
                f"excel={row[ci] if row else '?'} db={month_ind['calls_by_rid'].get(rid, 0)}",
            )
            check(
                "T6c Firuzabonu leads (Excel == DB)",
                row and row[li] == month_ind["leads_by_rid"].get(rid, 0),
                f"excel={row[li] if row else '?'} db={month_ind['leads_by_rid'].get(rid, 0)}",
            )
            talk_sec = month_ind["talk_by_user"].get(firuza.id, 0)
            expected_talk = f"{talk_sec // 3600}:{(talk_sec % 3600) // 60:02d}"
            ti = headers.index("Gaplashgan vaqt")
            check("T6d Firuzabonu talk h:mm", row and row[ti] == expected_talk, f"excel={row[ti] if row else '?'} kutilgan={expected_talk}")
        # Bog'lanmagan xodim "—"
        hayot_row = next((r for r in ws.iter_rows(min_row=3, values_only=True) if r[0] and "Hayot" in str(r[0])), None)
        if hayot_row:
            ci = headers.index("Qo'ng'iroqlar (jami)")
            check("T6e bog'lanmagan xodim '—'", hayot_row[ci] == "—")
        # Kunlik varaq bitta kun filtri
        buf1 = await build_report_xlsx(db, dt.date(2026, 7, 13), dt.date(2026, 7, 13))
        wb1 = load_workbook(buf1)
        d_rows = list(wb1["Kunlik"].iter_rows(min_row=2, values_only=True))
        check(
            "T6f Kunlik varaq sana filtri",
            d_rows and all(r[0] == "2026-07-13" for r in d_rows),
            f"{len(d_rows)} qator, sanalar: {set(r[0] for r in d_rows)}",
        )

        # ═══ T7: web overview seriya va sabablar ═══
        ov = await web_stats_overview(days=30, _=None, db=db)
        check("T7a seriya uzunligi 30", len(ov["series"]) == 30)
        last_p = ov["series"][-1]
        check(
            "T7b seriya[bugun].calls == mustaqil (aktivsiz to'liq)",
            last_p["calls"] == sum(ind["calls_by_rid"].values()),
            f"seriya={last_p['calls']} db={sum(ind['calls_by_rid'].values())}",
        )
        # O'rtadagi kun (13.07)
        p13 = next(p for p in ov["series"] if p["date"] == "2026-07-13")
        ind13 = await db_sums(db, dt.date(2026, 7, 13), dt.date(2026, 7, 13))
        check(
            "T7c seriya[13.07] calls/leads",
            p13["calls"] == sum(ind13["calls_by_rid"].values()) and p13["leads"] == sum(ind13["leads_by_rid"].values()),
            f"{p13['calls']}/{p13['leads']}",
        )
        reasons = ov["reasons"]
        week_ago = (today - dt.timedelta(days=6)).isoformat()
        check("T7d sabablar faqat 7 kun ichida", all(r["date"] >= week_ago for r in reasons), f"{len(reasons)} ta")
        db_reason_count = len(
            list(
                await db.scalars(
                    select(ShortfallReason).where(ShortfallReason.date >= today - dt.timedelta(days=6))
                )
            )
        )
        check("T7e sabablar soni DB bilan mos", len(reasons) == db_reason_count, f"api={len(reasons)} db={db_reason_count}")
        check("T7f days chegara 90", (await web_stats_overview(days=500, _=None, db=db))["days"] == 90)

        # ═══ T8: operator-summary raqamlari va % ═══
        sm = await web_operator_summary(period="week", _=None, db=db)
        wk_prev_ind = await db_sums(db, today - dt.timedelta(days=13), today - dt.timedelta(days=7))
        if firuza and firuza.crm_visit_external_id:
            rid = int(firuza.crm_visit_external_id)
            op_row = next((o for o in sm["operators"] if o["responsible_id"] == rid), None)
            cur_calls = wk_ind["calls_by_rid"].get(rid, 0)
            prev_calls = wk_prev_ind["calls_by_rid"].get(rid, 0)
            check("T8a Firuzabonu week calls", op_row and op_row["calls"] == cur_calls, f"api={op_row['calls'] if op_row else '?'} db={cur_calls}")
            expected_pct = _pct_change(cur_calls, prev_calls)
            check("T8b Firuzabonu pct", op_row and op_row["calls_pct"] == expected_pct, f"api={op_row['calls_pct'] if op_row else '?'} kutilgan={expected_pct}")
            wk_talk = wk_ind["talk_by_user"].get(firuza.id, 0)
            check("T8c Firuzabonu talk", op_row and op_row["talk_sec"] == wk_talk)
        # Noto'g'ri period → 400
        try:
            await web_operator_summary(period="yil", _=None, db=db)
            check("T8d noto'g'ri period 400", False, "xato bermadi")
        except Exception as e:  # noqa: BLE001
            check("T8d noto'g'ri period 400", "400" in str(getattr(e, "status_code", "")) or getattr(e, "status_code", 0) == 400)

        # ═══ T9: uch manba bir xil raqam (digest == overview == summary, bugun) ═══
        sm_today = await web_operator_summary(period="today", _=None, db=db)
        if digest["text"] is not None:
            check(
                "T9 digest==summary(today) calls",
                digest["totals"]["calls"] == sm_today["totals"]["calls"],
                f"digest={digest['totals']['calls']} summary={sm_today['totals']['calls']}",
            )

        # ═══ T10: kecha-yakuni tuzatish stsenariylari (V5) ═══
        cfg = await db.get(GroupPostConfig, 1)
        orig = (
            cfg.last_posted_date, cfg.last_posted_calls, cfg.last_posted_leads,
            cfg.last_posted_visits, cfg.correction_last_posted,
        )
        try:
            y_ind = await db_sums(db, yesterday, yesterday)
            y_final = y_ind["calls"]

            # D: yakuniy KAMAYGAN (snapshot anomaliyasi) → jim
            cfg.last_posted_date = yesterday
            cfg.last_posted_calls = y_final + 50
            cfg.last_posted_leads = 0
            cfg.last_posted_visits = 0
            cfg.correction_last_posted = None
            await db.commit()
            r = await send_yesterday_correction(db, dry_run=True)
            check("T10d kamaygan → jim", r.get("sent") is False and not r.get("text"), r.get("reason", ""))

            # F1: posted=0, final kichik (<5) bo'lsa jim / F2: final>=5 bo'lsa xabar
            cfg.last_posted_calls = 0
            await db.commit()
            r = await send_yesterday_correction(db, dry_run=True)
            check(
                "T10f posted=0 chegara (abs 5)",
                (y_final >= 5) == bool(r.get("text")),
                f"final={y_final}, text={'bor' if r.get('text') else 'yoq'}",
            )

            # E: guard — non-dry, farq sezilarli EMAS holatda (yubormaydi, guard yoziladi)
            cfg.last_posted_calls = y_final  # farq 0
            cfg.correction_last_posted = None
            await db.commit()
            r1 = await send_yesterday_correction(db, dry_run=False)
            r2 = await send_yesterday_correction(db, dry_run=False)
            check("T10e1 farqsiz non-dry yubormaydi", r1.get("sent") is False and r1.get("reason") == "Farq sezilarli emas")
            check("T10e2 guard ikkinchi chaqiriqda", r2.get("reason") == "Bugun allaqachon tekshirilgan")
        finally:
            (
                cfg.last_posted_date, cfg.last_posted_calls, cfg.last_posted_leads,
                cfg.last_posted_visits, cfg.correction_last_posted,
            ) = orig
            await db.commit()

        # ═══ T11: group_post_tick to'liq oqimi (send_message patch — yuborilmaydi) ═══
        sent_messages: list[tuple] = []

        async def fake_send(chat_id, text):
            sent_messages.append((chat_id, text[:60]))
            return {"ok": True}

        orig_send = dd.send_message
        orig_ai = dd._ai_summary_text

        async def no_ai(_db):
            return None

        cfg = await db.get(GroupPostConfig, 1)
        orig2 = (
            cfg.post_hour, cfg.post_minute, cfg.last_posted_date, cfg.last_posted_calls,
            cfg.last_posted_leads, cfg.last_posted_visits, cfg.correction_last_posted,
        )
        try:
            dd.send_message = fake_send
            dd._ai_summary_text = no_ai
            now = dt.datetime.now()
            cfg.post_hour, cfg.post_minute = 0, 0  # vaqt allaqachon o'tgan
            cfg.last_posted_date = None
            await db.commit()
            tick = await group_post_tick(db=db)
            await db.refresh(cfg)
            check("T11a tick fired", tick.get("fired") is True)
            check(
                "T11b tick xabarlarni barcha guruhlarga 'yubordi' (patch)",
                len(sent_messages) == len(digest_group_targets(None)),
                f"{len(sent_messages)} ta chat",
            )
            check("T11c last_posted_date bugun", cfg.last_posted_date == today)
            check(
                "T11d totals saqlandi",
                cfg.last_posted_calls == ind["calls"] and cfg.last_posted_leads == ind["leads"],
                f"saqlangan={cfg.last_posted_calls}/{cfg.last_posted_leads} kutilgan={ind['calls']}/{ind['leads']}",
            )
            # Ikkinchi tick — qo'riqchi
            tick2 = await group_post_tick(db=db)
            check("T11e ikkinchi tick jim (bir kunda bir marta)", tick2.get("fired") is False)
        finally:
            dd.send_message = orig_send
            dd._ai_summary_text = orig_ai
            (
                cfg.post_hour, cfg.post_minute, cfg.last_posted_date, cfg.last_posted_calls,
                cfg.last_posted_leads, cfg.last_posted_visits, cfg.correction_last_posted,
            ) = orig2
            await db.commit()

        # ═══ T12: bot handlerlar mavjudligi ═══
        import bot.handlers.group_stats as gs

        check("T12a /statistika chat to'plami to'g'ri", 0 not in gs.STATS_COMMAND_CHATS and len(gs.STATS_COMMAND_CHATS) == 3, f"{gs.STATS_COMMAND_CHATS}")
        check("T12b /oylik handler bor", hasattr(gs, "cmd_oylik"))
        import bot.api_client as ac

        check("T12c api_client'da monthly/daily digest", hasattr(ac, "trigger_monthly_digest") and hasattr(ac, "trigger_daily_digest"))

        # ═══ T13: my_stats week_totals (xodim haftalik kesimi) ═══
        from api.routers.stats import my_stats as my_stats_ep

        emp = next((u for u in users if u.role == "employee" and u.telegram_id and u.is_active), None)
        if emp:
            ms = await my_stats_ep(emp.telegram_id, db=db)
            has_week = getattr(ms, "week_totals", None)
            check("T13 my_stats week_totals bor", has_week is not None and isinstance(has_week, dict), f"{has_week}")

    print()
    print(f"═══ JAMI: {PASS} PASS, {FAIL} FAIL ═══")


asyncio.run(main())
